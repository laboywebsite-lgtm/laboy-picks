"""
LABOY PICKS — MLB
══════════════════════════════════════════════════════
  python3 mlb.py                     → jalaa data, actualiza Excel, muestra líneas
  python3 mlb.py --lines             → muestra líneas (lee Excel, no actualiza)
  python3 mlb.py --picks             → recomendaciones EV+ del modelo (solo terminal)
  python3 mlb.py --export-lines      → genera imagen PNG de las líneas del modelo
  python3 mlb.py --export-picks      → genera imagen PNG de los picks que logueaste hoy
  python3 mlb.py --stats             → estadísticas blended por equipo
  python3 mlb.py --stats-raw         → estadísticas RAW (2026 y 2025 por separado)
  python3 mlb.py --export            → genera .txt para comunidad/dubclub
  python3 mlb.py --refresh           → actualiza MLB Data (wRC+, bullpen xFIP)
  python3 mlb.py --debug-game A H    → diagnóstico de modelo vs mercado para un juego
  python3 mlb.py --log               → registra jugada + genera JPG automático
  python3 mlb.py --export-log [N]   → re-exporta pick #N como JPG (default: último)
  python3 mlb.py --grade IDX W|L|P   → califica jugada por índice
  python3 mlb.py --record            → muestra récord de jugadas
  python3 mlb.py --feedback          → análisis de rendimiento
  python3 mlb.py 2026-04-11          → fecha específica
  python3 mlb.py --export-lines --publish  → genera HTML y lo sube a GitHub Pages
  python3 mlb.py --lines --publish         → lines desde Excel + publica

  Fotos: --export-lines y --export-picks NO se generan automáticamente.
  Solo cuando tú lo pidas con esas flags.

  En iPhone / sin Chrome:
    export FANGRAPHS_COOKIE="your_cookie_here"
    python3 mlb.py

Requiere:
  pip3 install requests pandas openpyxl browser-cookie3 --break-system-packages
  (opcional) pip3 install tabulate --break-system-packages

══════════════════════════════════════════════════════
  NOTA: BetMGM no tiene API pública para usuarios.
  Usa --log / --record / --feedback para trackear
  tus picks localmente con análisis completo.
══════════════════════════════════════════════════════
  MODELO: Para máxima precisión necesitas FanGraphs FIP.
  Si ves error 403: export FANGRAPHS_COOKIE='tu_cookie'
  Usa --debug-game para diagnosticar gaps con el mercado.
══════════════════════════════════════════════════════
"""

import sys, os, re, json, time, warnings, hashlib
import requests
import pandas as pd
import numpy as np
from copy import copy
from openpyxl import load_workbook
from datetime import datetime, date, timedelta

warnings.filterwarnings("ignore")

# ── Cargar .env automáticamente (si existe) ───────────────────────────────────
def _load_dotenv():
    _root = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
    _env  = os.path.join(_root, ".env")
    if not os.path.exists(_env):
        return
    with open(_env) as _f:
        for _line in _f:
            _line = _line.strip()
            if not _line or _line.startswith("#") or "=" not in _line:
                continue
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())
_load_dotenv()

# ── Pillow (PNG exports) ──────────────────────────────
try:
    from PIL import Image, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# Market signals (sharp money / line movement)
try:
    from mlb_market import fetch_market_signals as _fetch_mlb_market_signals
    from mlb_market import sharp_confirm as _mlb_sharp_confirm
    from mlb_market import format_signal as _format_mlb_market_signal
    _HAS_MLB_MARKET = True
except ImportError:
    _HAS_MLB_MARKET = False
    def _fetch_mlb_market_signals(sport="mlb"): return {}
    def _mlb_sharp_confirm(signals, key, bet_type, side):
        return {"lean":"NEUTRAL","strength":0,"signals":[],"confirm":True,"fade":False,"available":False}
    def _format_mlb_market_signal(conf): return "📊 mlb_market.py no disponible"

# ── tabulate (opcional) ───────────────────────────────
try:
    from tabulate import tabulate as _tabulate
    HAS_TABULATE = True
    def tab(rows, headers, fmt="rounded_outline", **kw):
        return _tabulate(rows, headers=headers, tablefmt=fmt, **kw)
except ImportError:
    HAS_TABULATE = False
    def tab(rows, headers, fmt=None, **kw):
        if not rows:
            return "  ".join(str(h) for h in headers)
        w = [max(len(str(h)), max((len(str(r[i])) for r in rows), default=0))
             for i, h in enumerate(headers)]
        sep = "  ".join(str(h).ljust(w[i]) for i, h in enumerate(headers))
        lines = [sep, "─" * len(sep)]
        for row in rows:
            lines.append("  ".join(str(c).ljust(w[i]) for i, c in enumerate(row)))
        return "\n".join(lines)

# ──────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(SCRIPT_DIR, "Laboy Picks - MLB Model.xlsx")
_FONTS_DIR  = os.path.join(SCRIPT_DIR, "..", ".claude", "skills", "canvas-design", "canvas-fonts")

def _fnt(name, size):
    if not HAS_PIL: return None
    try:    return ImageFont.truetype(os.path.join(_FONTS_DIR, name), size)
    except:
        try:    return ImageFont.load_default(size=size)
        except: return ImageFont.load_default()
LOG_FILE         = os.path.join(SCRIPT_DIR, "laboy_picks_log.json")
MODEL_PICKS_FILE = os.path.join(SCRIPT_DIR, "mlb_model_picks.json")  # histórico picks del modelo
PRED_LOG_FILE    = os.path.join(SCRIPT_DIR, "mlb_predictions_log.json")  # log modelo vs mercado vs real
RESULTS_CACHE_FILE = os.path.join(SCRIPT_DIR, "mlb_results_cache.json")  # cache resultados modelo por fecha
DEBUG_STATE_FILE = os.path.join(SCRIPT_DIR, "mlb_debug_state.json")   # snapshot del último --picks run (para --export-debug)
LOG_STATE_FILE   = os.path.join(SCRIPT_DIR, "mlb_log_state.json")     # snapshot del último --export-debug HTML (fuente de verdad para --log)
WEATHER_OVERRIDES_FILE = os.path.join(SCRIPT_DIR, "mlb_weather_overrides.json")  # overrides manuales de clima
LINES_SHEET = "MLB Lines"
DATA_SHEET  = "MLB Data"

# ── Help ──────────────────────────────────────────────
if "--help" in sys.argv or "-h" in sys.argv:
    print("""
╔══════════════════════════════════════════════════════════════════╗
║               ⚾  LABOY PICKS — MLB  |  Comandos                 ║
╚══════════════════════════════════════════════════════════════════╝

  DATOS Y LÍNEAS
  ──────────────────────────────────────────────────────────────
  python3 mlb.py                     Jala data, actualiza Excel, muestra líneas del día
  python3 mlb.py 2026-04-15          Lo mismo pero para una fecha específica
  python3 mlb.py --lines             Muestra líneas leyendo Excel (sin jalar data nueva)
  python3 mlb.py --refresh           Actualiza wRC+, BP xFIP y SP xFIP desde FanGraphs

  EXPORTAR HTML
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --export-lines      Genera HTML con todas las líneas del modelo
  python3 mlb.py --export-picks      Genera HTML con los picks del log de hoy
  python3 mlb.py --export-lines --publish   Exporta HTML y lo sube a GitHub Pages
  python3 mlb.py --lines --publish          Lee Excel, exporta HTML y publica

  PICKS — LOG
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --log               Registra una jugada nueva (te pregunta los datos)
  python3 mlb.py --grade IDX W|L|P   Califica jugada por número (W=Win L=Loss P=Push)
                                     Ejemplo: python3 mlb.py --grade 3 W
  python3 mlb.py --remove IDX        Elimina el pick #IDX del log
  python3 mlb.py --remove IDX1 IDX2  Elimina múltiples picks (ej: --remove 2 4 7)
  python3 mlb.py --export-log        Re-exporta el último pick como HTML
  python3 mlb.py --export-log N      Re-exporta el pick #N como HTML

  ESTADÍSTICAS Y RÉCORD
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --record            Muestra récord completo de jugadas (W/L/P + P&L)
  python3 mlb.py --export-record            Exporta récord personal como HTML
  python3 mlb.py --export-record --publish  Exporta récord y lo publica en GitHub Pages
  python3 mlb.py --feedback          Análisis de rendimiento por tipo de pick (+ AI si ANTHROPIC_API_KEY)
  python3 mlb.py --stats             Estadísticas blended por equipo (wRC+ / xFIP)
  python3 mlb.py --stats-raw         Stats RAW separados por año (2025 y 2026)

  MODELO — CALIBRACIÓN HISTÓRICA
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --grade-picks              Auto-evalúa picks del modelo de hoy vs scores reales (ESPN)
  python3 mlb.py --grade-picks DATE         Evalúa picks de una fecha específica (ej: 2026-04-11)
  python3 mlb.py --grade-picks DATE --publish   Evalúa Y publica Picks HTML + Model Card a GitHub Pages
  python3 mlb.py --grade-picks URL  --publish   Evalúa desde URL pública y publica resultados

  SESIONES Y LOCKING DE PICKS
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --picks --day              Solo picks de juegos DÍA (antes 5 PM ET)
  python3 mlb.py --picks --night            Solo picks de juegos NOCHE (5 PM ET+)
  python3 mlb.py --picks                    Picks completos (todas las sesiones)
  python3 mlb.py --picks --confirmed        Omite juegos sin lineups confirmados
  python3 mlb.py --picks --force-repick     Sobreescribe picks ya guardados (rompe calibración)

  DIAGNÓSTICO
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --debug-game A H    Diagnóstico del modelo para un juego específico
                                     Ejemplo: python3 mlb.py --debug-game YANKEES RED SOX

  OTROS
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --export            Genera .txt para comunidad / Dubclub
  python3 mlb.py --picks             Recomendaciones EV+ del modelo (solo terminal)
  python3 mlb.py --help              Muestra este menú

  SESIONES (lineups 100% confirmados)
  ──────────────────────────────────────────────────────────────
  python3 mlb.py --picks --day       ☀️  Solo juegos con 1er pitch < 5 PM ET
                                         Correr a las 11:30 AM ET para datos completos
  python3 mlb.py --picks --night     🌙  Solo juegos con 1er pitch ≥ 5 PM ET
                                         Correr a las 6:00 PM ET para datos completos
  python3 mlb.py --lines --day       Igual pero muestra líneas (sin filtro EV+)
  python3 mlb.py --lines --night     Igual pero muestra líneas nocturnas

  python3 mlb.py --picks --confirmed ✅  Solo recomienda picks si AMBOS lineups están
                                         publicados en MLB.com (omite juegos sin lineup)
  python3 mlb.py --picks --day --confirmed   Combinación recomendada para sesión de día
  python3 mlb.py --picks --night --confirmed Combinación recomendada para sesión de noche

  VARIABLES DE ENTORNO
  ──────────────────────────────────────────────────────────────
  export FANGRAPHS_COOKIE="tu_cookie"    Cookie de FanGraphs (si hay error 403)
  export MLB_GITHUB_REPO="/tu/path"      Path local del repo mlb-picks para --publish
""")
    sys.exit(0)

# ── Parse flags ───────────────────────────────────────
DEBUG            = "--debug"        in sys.argv
REFRESH_MODE     = "--refresh"        in sys.argv
LINES_MODE       = "--lines"          in sys.argv
PICKS_MODE       = "--picks"          in sys.argv
PICKS_EXPORT     = "--picks-export"   in sys.argv   # legacy alias
EXPORT_LINES     = "--export-lines"      in sys.argv   # genera PNG de lines
EXPORT_PICKS     = "--export-picks"      in sys.argv   # genera PNG de picks del log
EXPORT_LINES_PDF = "--export-lines-pdf"  in sys.argv   # PDF de lines (PNGs → PDF)
EXPORT_PICKS_PDF = "--export-picks-pdf"  in sys.argv   # PDF de picks (PNGs → PDF)
EXPORT_STORY     = "--export-story"      in sys.argv   # alias --export-picks (story 1080×1920)
EXPORT_POST      = "--export-post"       in sys.argv   # pick card recortado 1080×1080
STATS_MODE       = "--stats"          in sys.argv
STATS_RAW_MODE   = "--stats-raw"      in sys.argv
EXPORT_MODE      = "--export"         in sys.argv
RECORD_MODE      = "--record"       in sys.argv
PENDING_MODE     = "--pending"      in sys.argv   # solo picks sin gradear (útil para --grade)
FEEDBACK_MODE    = "--feedback"     in sys.argv
LOG_MODE         = "--log"          in sys.argv
GRADE_MODE       = "--grade"        in sys.argv
DEBUG_GAME_MODE  = "--debug-game"   in sys.argv
EXPORT_LOG_MODE  = "--export-log"   in sys.argv   # exporta pick logueado como HTML
EXPORT_RECORD_MODE = "--export-record" in sys.argv # exporta récord como HTML+JPG
EXPORT_DEBUG_MODE  = "--export-debug"  in sys.argv # exporta debug HTML de picks
PUBLISH_MODE     = "--publish"      in sys.argv   # push HTMLs a GitHub Pages
FORCE_EXPORT     = "--force-export" in sys.argv   # sobreescribir picks HTML aunque ya exista (rompe protección)
REMOVE_MODE      = "--remove"       in sys.argv   # elimina pick(s) del log
GRADE_PICKS_MODE = "--grade-picks"  in sys.argv   # auto-grade picks del modelo
DAY_SESSION      = "--day"          in sys.argv   # solo juegos antes de 5 PM ET
PM_SESSION       = "--pm"           in sys.argv   # solo juegos 1 PM–4:59 PM ET  (tarde, sub-sesión)
NIGHT_SESSION    = "--night"        in sys.argv   # solo juegos 5 PM ET en adelante
REQUIRE_LINEUPS  = "--confirmed"    in sys.argv   # omitir picks sin lineups confirmados de AMBOS equipos
FORCE_REPICK     = "--force-repick" in sys.argv   # fuerza re-guardar picks aunque ya existan (sobreescribe lock)
WEATHER_MODE     = "--weather"     in sys.argv   # muestra clima en vivo para todos los juegos de hoy
SET_WEATHER_MODE = "--set-weather" in sys.argv   # override manual de clima para un juego
CLEAR_WEATHER_MODE = "--clear-weather" in sys.argv  # borra todos los overrides de hoy

# Sesión actual: "day" | "pm" | "night" | "full"
CURRENT_SESSION  = "day" if DAY_SESSION else ("pm" if PM_SESSION else ("night" if NIGHT_SESSION else "full"))

# ── Positional args (excluye argumentos de --grade y --debug-game) ───
_skip_indices = set()
for _flag, _n in [("--grade", 2), ("--debug-game", 2)]:
    if _flag in sys.argv:
        _fi = sys.argv.index(_flag)
        for _x in range(1, _n+1):
            _skip_indices.add(_fi + _x)

args        = [a for i, a in enumerate(sys.argv[1:], 1)
               if not a.startswith("--") and i not in _skip_indices]
TARGET_DATE = args[0] if args else date.today().strftime("%Y-%m-%d")
SEASON      = TARGET_DATE[:4]

# ── The Odds API ──────────────────────────────────────
ODDS_API_KEY = os.environ.get("ODDS_API_KEY", "")  # configura en .env

# ── FanGraphs cookie manual ───────────────────────────
FANGRAPHS_COOKIE_ENV = os.environ.get("FANGRAPHS_COOKIE", "")

# Variable global para saber si FanGraphs FIP fue cargado
FG_FIP_AVAILABLE = False

# ── FanGraphs data cache (JSON, bypasses Excel for wRC+/xFIP) ───
FG_CACHE_FILE = os.path.join(SCRIPT_DIR, "mlb_fg_cache.json")

# ── Umpire tendency cache (JSON) ─────────────────────────────────
UMP_CACHE_FILE = os.path.join(SCRIPT_DIR, "mlb_ump_cache.json")

# ── GitHub Pages publish config ───────────────────────────────
# Pon aquí el path a tu clon local de laboywebsite-lgtm/mlb-picks
# Ejemplo: "/Users/jose/repos/mlb-picks"
GITHUB_PAGES_REPO = os.environ.get(
    "MLB_GITHUB_REPO",
    os.path.join(os.path.expanduser("~"), "repos", "mlb-picks")
)
GITHUB_PAGES_URL  = "https://laboywebsite-lgtm.github.io/mlb-picks"

# ── URL token — seguridad por oscuridad ──────────────
# Salt privado. Cámbialo (o pon env var MLB_TOKEN_SALT) para rotar todos los tokens.
_URL_TOKEN_SALT = os.environ.get("MLB_TOKEN_SALT", "laboyPicksSalt2026")

def _url_token(date_str: str) -> str:
    """
    Token corto (7 chars) determinístico derivado de la fecha + salt.
    Hace que los nombres de archivo sean impredecibles para alguien que
    sólo conoce la fecha (e.g. cambiar 04-13 por 04-14 en la URL no funciona).
    """
    raw = f"{date_str}:{_URL_TOKEN_SALT}"
    return hashlib.sha256(raw.encode()).hexdigest()[:7]

# Dashboard secreto — solo tú tienes la URL.
# Cámbialo o pon env var MLB_DASHBOARD_TOKEN para personalizarlo.
DASHBOARD_TOKEN = os.environ.get("MLB_DASHBOARD_TOKEN", "changeme")  # configura en .env

# ──────────────────────────────────────────────────────
# MAPEOS DE EQUIPOS
# ──────────────────────────────────────────────────────

TEAM_ABB = {
    "ARI":"D-BACKS",   "ATL":"BRAVES",    "BAL":"ORIOLES",  "BOS":"RED SOX",
    "CHC":"CUBS",      "CWS":"WHITE SOX", "CHW":"WHITE SOX","CIN":"REDS",
    "CLE":"GUARDIANS", "COL":"ROCKIES",   "DET":"TIGERS",   "HOU":"ASTROS",
    "KCR":"ROYALS",    "KCA":"ROYALS",    "LAA":"ANGELS",   "LAD":"DODGERS",
    "MIA":"MARLINS",   "MIL":"BREWERS",   "MIN":"TWINS",    "NYM":"METS",
    "NYY":"YANKEES",   "OAK":"ATHLETICS", "ATH":"ATHLETICS","PHI":"PHILLIES",
    "PIT":"PIRATES",   "SDP":"PADRES",    "SDN":"PADRES",   "SEA":"MARINERS",
    "SFG":"GIANTS",    "SLN":"CARDINALS", "STL":"CARDINALS","TBR":"RAYS",
    "TBA":"RAYS",      "TEX":"RANGERS",   "TOR":"BLUE JAYS","WSN":"NATIONALS",
    "WAS":"NATIONALS",
}

TEAM_MAP = {
    "Arizona Diamondbacks":"D-BACKS",   "Atlanta Braves":"BRAVES",
    "Baltimore Orioles":"ORIOLES",      "Boston Red Sox":"RED SOX",
    "Chicago Cubs":"CUBS",              "Chicago White Sox":"WHITE SOX",
    "Cincinnati Reds":"REDS",           "Cleveland Guardians":"GUARDIANS",
    "Colorado Rockies":"ROCKIES",       "Detroit Tigers":"TIGERS",
    "Houston Astros":"ASTROS",          "Kansas City Royals":"ROYALS",
    "Los Angeles Angels":"ANGELS",      "Los Angeles Dodgers":"DODGERS",
    "Miami Marlins":"MARLINS",          "Milwaukee Brewers":"BREWERS",
    "Minnesota Twins":"TWINS",          "New York Mets":"METS",
    "New York Yankees":"YANKEES",       "Oakland Athletics":"ATHLETICS",
    "Philadelphia Phillies":"PHILLIES", "Pittsburgh Pirates":"PIRATES",
    "San Diego Padres":"PADRES",        "San Francisco Giants":"GIANTS",
    "Seattle Mariners":"MARINERS",      "St. Louis Cardinals":"CARDINALS",
    "Tampa Bay Rays":"RAYS",            "Texas Rangers":"RANGERS",
    "Toronto Blue Jays":"BLUE JAYS",    "Washington Nationals":"NATIONALS",
    "Athletics":"ATHLETICS",
}

BOOK_IDS = {
    "Pinnacle":   "pinnacle",
    "FanDuel":    "fanduel",
    "BetMGM":     "betmgm",
    "DraftKings": "draftkings",
    "Circa":      "circasports",
}

STADIUMS = {
    "D-BACKS":"Phoenix,AZ",       "BRAVES":"Cumberland,GA",
    "ORIOLES":"Baltimore,MD",     "RED SOX":"Boston,MA",
    "CUBS":"Chicago,IL",          "WHITE SOX":"Chicago,IL",
    "REDS":"Cincinnati,OH",       "GUARDIANS":"Cleveland,OH",
    "ROCKIES":"Denver,CO",        "TIGERS":"Detroit,MI",
    "ASTROS":"Houston,TX",        "ROYALS":"Kansas City,MO",
    "ANGELS":"Anaheim,CA",        "DODGERS":"Los Angeles,CA",
    "MARLINS":"Miami,FL",         "BREWERS":"Milwaukee,WI",
    "TWINS":"Minneapolis,MN",     "METS":"New York,NY",
    "YANKEES":"New York,NY",      "ATHLETICS":"Sacramento,CA",
    "PHILLIES":"Philadelphia,PA", "PIRATES":"Pittsburgh,PA",
    "PADRES":"San Diego,CA",      "GIANTS":"San Francisco,CA",
    "MARINERS":"Seattle,WA",      "CARDINALS":"St. Louis,MO",
    "RAYS":"St. Petersburg,FL",   "RANGERS":"Arlington,TX",
    "BLUE JAYS":"Toronto,Canada", "NATIONALS":"Washington,DC",
}

# Exact stadium lat/lon for Open-Meteo (primary weather source)
STADIUM_COORDS = {
    "D-BACKS":   (33.4453, -112.0667),  # Chase Field, Phoenix AZ
    "BRAVES":    (33.8901,  -84.4678),  # Truist Park, Cumberland GA
    "ORIOLES":   (39.2838,  -76.6218),  # Camden Yards, Baltimore MD
    "RED SOX":   (42.3467,  -71.0972),  # Fenway Park, Boston MA
    "CUBS":      (41.9484,  -87.6553),  # Wrigley Field, Chicago IL
    "WHITE SOX": (41.8300,  -87.6339),  # Guaranteed Rate Field, Chicago IL
    "REDS":      (39.0975,  -84.5081),  # Great American Ball Park, Cincinnati OH
    "GUARDIANS": (41.4962,  -81.6852),  # Progressive Field, Cleveland OH
    "ROCKIES":   (39.7559, -104.9942),  # Coors Field, Denver CO
    "TIGERS":    (42.3390,  -83.0485),  # Comerica Park, Detroit MI
    "ROYALS":    (39.0517,  -94.4803),  # Kauffman Stadium, Kansas City MO
    "ANGELS":    (33.8003, -117.8827),  # Angel Stadium, Anaheim CA
    "DODGERS":   (34.0739, -118.2400),  # Dodger Stadium, Los Angeles CA
    "TWINS":     (44.9817,  -93.2778),  # Target Field, Minneapolis MN
    "METS":      (40.7571,  -73.8458),  # Citi Field, Queens NY
    "YANKEES":   (40.8296,  -73.9262),  # Yankee Stadium, Bronx NY
    "PHILLIES":  (39.9061,  -75.1665),  # Citizens Bank Park, Philadelphia PA
    "PIRATES":   (40.4469,  -80.0057),  # PNC Park, Pittsburgh PA
    "PADRES":    (32.7073, -117.1566),  # Petco Park, San Diego CA
    "GIANTS":    (37.7786, -122.3893),  # Oracle Park, San Francisco CA
    "MARINERS":  (47.5914, -122.3325),  # T-Mobile Park, Seattle WA
    "CARDINALS": (38.6226,  -90.1928),  # Busch Stadium, St. Louis MO
    "NATIONALS": (38.8730,  -77.0074),  # Nationals Park, Washington DC
    # Athletics moved to Sacramento 2025 → Sutter Health Park
    "ATHLETICS": (38.5804, -121.5001),
}

# ── Estadios con techo / sin impacto de clima ───────────────────────────────
# Fuente autoritativa: mlb_stadiums.json (mismo directorio que mlb.py).
# Si el JSON existe, DOME_TEAMS y STADIUM_ROOF se derivan de él automáticamente.
# Para actualizar: edita mlb_stadiums.json y reinicia. No tocar código.
_STADIUM_FULL_TO_ABB = {
    "Arizona Diamondbacks":"D-BACKS",   "Atlanta Braves":"BRAVES",
    "Baltimore Orioles":"ORIOLES",      "Boston Red Sox":"RED SOX",
    "Chicago Cubs":"CUBS",              "Chicago White Sox":"WHITE SOX",
    "Cincinnati Reds":"REDS",           "Cleveland Guardians":"GUARDIANS",
    "Colorado Rockies":"ROCKIES",       "Detroit Tigers":"TIGERS",
    "Houston Astros":"ASTROS",          "Kansas City Royals":"ROYALS",
    "Los Angeles Angels":"ANGELS",      "Los Angeles Dodgers":"DODGERS",
    "Miami Marlins":"MARLINS",          "Milwaukee Brewers":"BREWERS",
    "Minnesota Twins":"TWINS",          "New York Mets":"METS",
    "New York Yankees":"YANKEES",       "Oakland Athletics":"ATHLETICS",
    "Philadelphia Phillies":"PHILLIES", "Pittsburgh Pirates":"PIRATES",
    "San Diego Padres":"PADRES",        "San Francisco Giants":"GIANTS",
    "Seattle Mariners":"MARINERS",      "St. Louis Cardinals":"CARDINALS",
    "Tampa Bay Rays":"RAYS",            "Texas Rangers":"RANGERS",
    "Toronto Blue Jays":"BLUE JAYS",    "Washington Nationals":"NATIONALS",
}
# STADIUM_ROOF[abbr] = {'roof': str, 'impact': bool, 'name': str, 'lat': float, 'lon': float}
STADIUM_ROOF = {}
try:
    import json as _json_stadiums
    _sj_path = os.path.join(SCRIPT_DIR, "mlb_stadiums.json")
    with open(_sj_path, encoding="utf-8") as _sj:
        _sj_data = _json_stadiums.load(_sj)
    for _s in _sj_data.get("stadiums", []):
        _abb = _STADIUM_FULL_TO_ABB.get(_s.get("team",""))
        if _abb:
            STADIUM_ROOF[_abb] = {
                "roof":   _s.get("roof", "open"),
                "impact": _s.get("weather_impact", True),
                "name":   _s.get("stadium", ""),
                "lat":    _s.get("lat"),
                "lon":    _s.get("lon"),
            }
    # Derive DOME_TEAMS from JSON (weather_impact=false means covered/no-weather)
    DOME_TEAMS = {abb for abb, info in STADIUM_ROOF.items() if not info["impact"]}
except Exception:
    # Fallback hardcoded set (kept in sync manually as backup)
    DOME_TEAMS = {"RAYS","BREWERS","MARLINS","ASTROS","RANGERS","BLUE JAYS",
                  "D-BACKS","MARINERS"}
# ATHLETICS removidos: Sutter Health Park (Sacramento) es estadio abierto

PARK_FACTORS = {
    "D-BACKS":1.01,"BRAVES":1.0,"ORIOLES":0.99,"RED SOX":1.04,"CUBS":0.98,
    "WHITE SOX":0.98,"REDS":1.05,"GUARDIANS":0.99,"ROCKIES":1.13,"TIGERS":1.0,
    "ASTROS":0.99,"ROYALS":1.03,"ANGELS":1.01,"DODGERS":0.99,"MARLINS":1.01,
    "BREWERS":0.99,"TWINS":1.0,"METS":0.96,"YANKEES":1.02,"ATHLETICS":1.03,
    "PHILLIES":1.02,"PIRATES":1.02,"PADRES":0.96,"GIANTS":0.97,"MARINERS":0.94,
    "CARDINALS":0.96,"RAYS":1.01,"RANGERS":1.0,"BLUE JAYS":1.01,"NATIONALS":1.0,
}

# ── Orientación de parques (bearing home plate → CF, grados desde Norte) ─
# Convención: mirando DESDE home plate HACIA center field.
# 0°=Norte, 90°=Este, 180°=Sur, 270°=Oeste.
# CF bearing = dirección desde home plate HACIA center field (la que mira el bateador).
# Convención: ≤60° diff → OUT (viento favorece jonrones), ≥120° → IN, resto → CROSS.
# Verificados/corregidos 2026-04-30 contra Prop Finder + Google Maps satelital.
#
# ✅ PADRES=45   Petco Park: home plate en el sur, CF abre hacia el NE.
#    Verificación 2026-05-09: W wind → toward=90°, diff=|90-45|=45 → OUT ✓ (PF muestra OUT)
#    Verificación 2026-05-07: N wind → toward=180°, diff=|180-45|=135 → IN ✓ (PF verificado)
#    NOTA: valor 315° era INCORRECTO (espejo del correcto) — daba W→IN cuando PF muestra OUT.
# ✅ ROYALS=10   Kauffman Stadium: home plate al sur, CF mira al N/NNE.
#    Verificación 2026-05-07: SW wind → toward=45°, diff=|45-10|=35 → OUT ✓ (vs PF que mostraba OUT)
#    NOTA: valor anterior era 225° (INCORRECTO — daba SW→IN cuando PF muestra OUT)
# ✅ NATIONALS=350 Nationals Park: CF apunta al NNW (el Capitolio es visible desde home plate).
#    SW/SE → OUT ✓, N/NW → IN ✓
#    NOTA: valor anterior era 225° (corrección del 2026-04-30 fue errónea — original 350° era correcto)
# ✅ DODGERS=0   Dodger Stadium: home plate S, CF mira N.
#    Verificación: SW wind → toward=45°, diff=|45-0|=45 → OUT ✓
# ✅ PIRATES=175 PNC Park: ribera N río Allegheny, outfield abre al S/SSW.
#    Verificación: NW wind → toward=135°, diff=|135-175|=40 → OUT ✓ (confirmado vs PF hoy)
#
# 🔧 CORREGIDOS 2026-04-30 (valores anteriores daban IN donde PF muestra CROSS/OUT):
# BRAVES    5→315  Truist Park: home plate SE, outfield abre al NW (Battery Atlanta al NO).
#   N wind → toward=180°, diff=|180-315|=135 → IN ✓  (PropFinder confirmed)
# REDS    345→180  GABP: río Ohio al SUR del estadio, outfield abre al S.
#   NW wind → toward=135°, diff=|135-180|=45 → OUT ✓
# PHILLIES  5→225  Citizens Bank Park: home plate NE (estándar), CF al SW.
#   NW wind → toward=135°, diff=|135-225|=90 → CROSS ✓
# ✅ ORIOLES=45   Camden Yards: home plate al S, CF apunta al NE (~45°).
#   SW wind → toward=45°, diff=|45-45|=0  → OUT   ✓ (verificado vs PropFinder 2026-05-09)
#   NW wind → toward=135°, diff=|135-45|=90 → CROSS ✓
#   NOTA: valor 225° era INCORRECTO (exactamente invertido) — daba SW→IN cuando PF muestra OUT.
# YANKEES 355→225  Yankee Stadium: home plate NE, CF al SW.
# NATIONALS 350→225 Nationals Park: home plate NE, CF al SW.
# RED SOX  95→50   Fenway Park: home plate SW, CF al NE (skyline visible más allá del CF).
#   NW wind → toward=135°, diff=|135-50|=85 → CROSS
# MARINERS 335→200 T-Mobile Park: abre hacia el sur (First Ave S).
#   NW wind → toward=135°, diff=|135-200|=65 → CROSS
# GUARDIANS 320→270 Progressive Field: CF abre al W (downtown CLE).
#   NW wind → toward=135°, diff=|135-270|=135 → IN (viento de cara para jonrones al oeste)
# METS     45→225  Citi Field: home plate NE, CF al SW. (45° daba CROSS por casualidad,
#                  225° es geográficamente correcto y sigue dando CROSS para NW wind)
# BLUE JAYS 5→225  Rogers Centre: orientación estándar NE→SW.
# CF bearings: degrees from HOME PLATE toward CENTER FIELD (0°=N, 90°=E, 180°=S, 270°=W)
# Source: Andrew Clem's Baseball Stadium Statistics (andrewclem.com) — verified from satellite imagery
# Last audited: 2026-05-10 (full audit — previous values had widespread mirror/rotation errors)
# Dome teams: bearing retained for completeness but not used in weather calc (DOME_TEAMS filter applies)
STADIUM_CF_BEARING = {
    # ── Open-air / weather-relevant ────────────────────────────────────────
    "BRAVES":   158,  # SSE  — Truist Park (most southward NL park)
    "ORIOLES":   22,  # NNE  — Camden Yards
    "RED SOX":   45,  # NE   — Fenway Park
    "CUBS":      45,  # NE   — Wrigley Field
    "WHITE SOX": 18,  # NNE  — Guaranteed Rate Field (home plate SW, CF NNE toward downtown)
    "REDS":     113,  # ESE  — Great American Ball Park (CF toward ESE; N wind = CROSS con tendencia RF)
    "GUARDIANS":  0,  # N    — Progressive Field
    "ROCKIES":    0,  # N    — Coors Field
    "TIGERS":   158,  # SSE  — Comerica Park (southernmost AL park)
    "ROYALS":    45,  # NE   — Kauffman Stadium
    "ANGELS":    45,  # NE   — Angel Stadium
    "DODGERS":   22,  # NNE  — Dodger Stadium
    "TWINS":     90,  # E    — Target Field
    "METS":      22,  # NNE  — Citi Field
    "YANKEES":   68,  # ENE  — Yankee Stadium
    "ATHLETICS": 45,  # NE   — Sutter Health Park (Sacramento)
    "PHILLIES":  22,  # NNE  — Citizens Bank Park
    "PIRATES":   30,  # NNE  — PNC Park (home plate SW, CF toward NNE/downtown Pittsburgh)
    "PADRES":     0,  # N    — Petco Park
    "GIANTS":   315,  # NW   — Oracle Park (home plate SE, CF NW; bay/marine wind blows IN from CF)
    "CARDINALS": 45,  # NE   — Busch Stadium III
    "NATIONALS": 22,  # NNE  — Nationals Park
    # ── Retractable / fixed dome (weather_impact=False — bearing unused) ───
    "D-BACKS":    0,  # N    — Chase Field (dome)
    "ASTROS":    68,  # ENE  — Minute Maid Park (dome)
    "MARLINS":  113,  # ESE  — loanDepot Park (dome)
    "BREWERS":  135,  # SE   — American Family Field (dome)
    "MARINERS":  45,  # NE   — T-Mobile Park (dome)
    "RAYS":      45,  # NE   — Tropicana Field (dome)
    "RANGERS":   68,  # ENE  — Globe Life Field (dome)
    "BLUE JAYS":338,  # NNW  — Rogers Centre (dome)
}

STADIUM_TZ_OFFSET = {
    "D-BACKS":-7,"BRAVES":-4,"ORIOLES":-4,"RED SOX":-4,
    "CUBS":-5,"WHITE SOX":-5,"REDS":-4,"GUARDIANS":-4,
    "ROCKIES":-6,"TIGERS":-4,"ASTROS":-5,"ROYALS":-5,
    "ANGELS":-7,"DODGERS":-7,"MARLINS":-4,"BREWERS":-5,
    "TWINS":-5,"METS":-4,"YANKEES":-4,"ATHLETICS":-7,
    "PHILLIES":-4,"PIRATES":-4,"PADRES":-7,"GIANTS":-7,
    "MARINERS":-7,"CARDINALS":-5,"RAYS":-4,"RANGERS":-5,
    "BLUE JAYS":-4,"NATIONALS":-4,
}

TZ_ABBR   = {-4:"ET",-5:"CT",-6:"MT",-7:"PT"}

COMPASS_BEARING = {
    "N":0,"NNE":22.5,"NE":45,"ENE":67.5,
    "E":90,"ESE":112.5,"SE":135,"SSE":157.5,
    "S":180,"SSW":202.5,"SW":225,"WSW":247.5,
    "W":270,"WNW":292.5,"NW":315,"NNW":337.5,
}

# ESPN logos (para HTML export)
ESPN_ABB = {
    "D-BACKS":"ari","BRAVES":"atl","ORIOLES":"bal","RED SOX":"bos",
    "CUBS":"chc","WHITE SOX":"cws","REDS":"cin","GUARDIANS":"cle",
    "ROCKIES":"col","TIGERS":"det","ASTROS":"hou","ROYALS":"kc",
    "ANGELS":"laa","DODGERS":"lad","MARLINS":"mia","BREWERS":"mil",
    "TWINS":"min","METS":"nym","YANKEES":"nyy","ATHLETICS":"oak",
    "PHILLIES":"phi","PIRATES":"pit","PADRES":"sd","GIANTS":"sf",
    "MARINERS":"sea","CARDINALS":"stl","RAYS":"tb","RANGERS":"tex",
    "BLUE JAYS":"tor","NATIONALS":"wsh",
}

def logo_url(team):
    abb = ESPN_ABB.get(team,"")
    if not abb: return ""
    return f"https://a.espncdn.com/combiner/i?img=/i/teamlogos/mlb/500/{abb}.png&w=80&h=80&scale=crop&cquality=40"


# ──────────────────────────────────────────────────────
# FANGRAPHS SESSION
# ──────────────────────────────────────────────────────

def fg_session():
    """
    Crea sesión autenticada con FanGraphs.
    FanGraphs corre en WordPress — la cookie de auth es wordpress_logged_in_*
    (NO es __Secure-next-auth.session-token como antes).
    """
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.fangraphs.com/leaders/major-league",
        "Origin": "https://www.fangraphs.com",
    })

    # ── 1. Variable de entorno ─────────────────────────────────
    # Acepta dos formatos:
    #   A) "wordpress_logged_in_<hash>=laboypicks%7C18..."  (nombre=valor)
    #   B) "laboypicks%7C18..."                             (solo valor — necesita nombre en FANGRAPHS_COOKIE_NAME)
    if FANGRAPHS_COOKIE_ENV:
        if "=" in FANGRAPHS_COOKIE_ENV and "wordpress_logged_in" in FANGRAPHS_COOKIE_ENV:
            # Formato A: "nombre=valor"
            name_part, val_part = FANGRAPHS_COOKIE_ENV.split("=", 1)
            cookie_name  = name_part.strip()
            cookie_value = val_part.strip()
        else:
            # Formato B: solo valor — leer nombre desde FANGRAPHS_COOKIE_NAME
            cookie_value = FANGRAPHS_COOKIE_ENV.strip()
            cookie_name  = os.environ.get("FANGRAPHS_COOKIE_NAME", "").strip()
            if not cookie_name:
                # Sin nombre → no podemos autenticar correctamente
                print("\n" + "="*64)
                print("  ⚠️  FANGRAPHS_COOKIE_NAME no definido.")
                print("  WordPress requiere el nombre EXACTO de la cookie.")
                print()
                print("  1. En DevTools → Application → Cookies → fangraphs.com")
                print("  2. Copia el NOMBRE de la cookie (empieza con wordpress_logged_in_)")
                print("     ejemplo: wordpress_logged_in_0cae6f5cb...")
                print()
                print("  Luego exporta AMBAS variables:")
                print("  export FANGRAPHS_COOKIE_NAME='wordpress_logged_in_0cae6f5cb...'")
                print("  export FANGRAPHS_COOKIE='laboypicks%7C18...'")
                print("="*64 + "\n")
                return s

        s.cookies.set(cookie_name, cookie_value, domain=".fangraphs.com")
        # Mandar también como Cookie header raw (más robusto que solo el jar)
        s.headers["Cookie"] = f"{cookie_name}={cookie_value}"
        print(f"  🍪 FanGraphs: WordPress cookie desde variable de entorno")
        print(f"     Nombre: {cookie_name[:50]}")
        return s

    # ── 2. browser_cookie3 — extracción automática ────────────
    bc3_ok = False
    try:
        import browser_cookie3
        bc3_ok = True
    except ImportError:
        import subprocess
        subprocess.run(["pip3","install","browser-cookie3","--break-system-packages","-q"],
                       capture_output=True)
        try:
            import browser_cookie3
            bc3_ok = True
        except ImportError:
            pass

    if bc3_ok:
        for browser_fn, bname in [
            (lambda: __import__("browser_cookie3").chrome(domain_name=".fangraphs.com"),  "Chrome"),
            (lambda: __import__("browser_cookie3").safari(domain_name=".fangraphs.com"),  "Safari"),
            (lambda: __import__("browser_cookie3").firefox(domain_name=".fangraphs.com"), "Firefox"),
        ]:
            try:
                cookies    = browser_fn()
                clist      = list(cookies)
                fg_cookies = [c for c in clist if "fangraphs" in c.domain.lower()]
                # FanGraphs usa WordPress → cookie auth = wordpress_logged_in_*
                wp_cookie  = next((c for c in fg_cookies
                                   if c.name.startswith("wordpress_logged_in")), None)
                if wp_cookie:
                    s.cookies.update(cookies)
                    print(f"  🍪 FanGraphs: WordPress cookie en {bname} "
                          f"({wp_cookie.name[:40]}...)")
                    return s
                elif fg_cookies:
                    s.cookies.update(cookies)
                    print(f"  🍪 FanGraphs: {bname} ({len(fg_cookies)} cookies, sin sesión WP)")
                    print(f"     → Asegúrate de estar logueado en fangraphs.com en {bname}")
                    return s
            except Exception:
                pass

    # ── 3. Sin cookies — instrucciones paso a paso ────────────
    print("\n" + "="*64)
    print("  ⚠️  FanGraphs FIP no disponible — sin sesión WordPress activa")
    print("="*64)
    print("  SOLUCIÓN (una sola vez):")
    print()
    print("  1. Abre Chrome y entra a fangraphs.com")
    print("  2. Inicia sesión con tu cuenta (si no has hecho esto)")
    print("  3. Presiona Cmd+Option+I (Mac) o F12 (Windows) → DevTools")
    print("  4. Click en la pestaña 'Application'")
    print("  5. En el panel izquierdo: Storage → Cookies → https://www.fangraphs.com")
    print("  6. Busca la cookie que empieza con:  wordpress_logged_in_")
    print("     (ejemplo: wordpress_logged_in_0cae6f5cb...)")
    print("  7. Click en ella → copia el valor del campo 'Value' (empieza con tu usuario)")
    print()
    print("  8. En tu terminal, ANTES de correr mlb.py:")
    print("     export FANGRAPHS_COOKIE='laboypicks|18...'")
    print()
    print("  Para no repetirlo: agrega esa línea a ~/.zshrc o ~/.bash_profile")
    print("="*64 + "\n")
    print("     export FANGRAPHS_COOKIE='tu_cookie'")
    print("  Así no tienes que hacerlo cada vez.")
    print("="*62 + "\n")
    return s


def fg_api(session, params):
    base = "https://www.fangraphs.com/api/leaders/major-league/data"
    defaults = {"month":"0","ind":"0","rost":"0","age":"0",
                "filter":"","players":"0","startdate":"","enddate":"",
                "sort":"4,1","pageitems":"1000","pagenum":"1"}
    defaults.update(params)
    r = session.get(base, params=defaults, timeout=20)
    r.raise_for_status()
    return pd.DataFrame(r.json().get("data", []))


def _strip_accents(s):
    """Normaliza caracteres acentuados → ASCII. Ej: SÁNCHEZ → SANCHEZ."""
    import unicodedata
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

def name_keys(name):
    raw = str(name).strip().upper()
    n   = _strip_accents(raw)        # SÁNCHEZ → SANCHEZ
    parts = n.split()
    keys = [n]
    if len(parts) > 1:
        keys.append(parts[-1])       # apellido solo
    # Si hubo acento, incluir también la versión con acento por si el cache la tiene
    if raw != n:
        keys.append(raw)
        if len(raw.split()) > 1:
            keys.append(raw.split()[-1])
    return keys


# ──────────────────────────────────────────────────────
# DATOS FANGRAPHS
# ──────────────────────────────────────────────────────

def _fg_team_col(df):
    """Detecta la columna de equipo en un DataFrame de FanGraphs."""
    for candidate in ("Team", "Tm", "team", "tm", "TeamName", "team_name"):
        if candidate in df.columns:
            return candidate
    return next((c for c in df.columns if "team" in c.lower()), None)


def _fg_extract_abb(raw_val):
    """
    FanGraphs devuelve la columna Team como HTML anchor:
      '<a href="...team=22...">LAD</a>'
    Esta función extrae la abreviatura del texto del anchor.
    También funciona con abreviaturas directas ('LAD').
    """
    s = str(raw_val).strip()
    m = re.search(r'>([A-Za-z0-9 \-]+)</a>', s)
    if m:
        return m.group(1).strip().upper()
    # Sin HTML → asumir que ya es la abreviatura
    return s.upper()


def _fg_resolve_team(raw_val):
    """Extrae abreviatura de HTML (si aplica) y mapea al nombre interno del modelo."""
    abb = _fg_extract_abb(raw_val)
    # 1. Abreviatura directa (ARI, ATL, LAD, ...)
    t = TEAM_ABB.get(abb)
    if t: return t
    # 2. Nombre completo exacto
    t = TEAM_MAP.get(abb) or TEAM_MAP.get(str(raw_val).strip())
    return t


def get_team_wrc(session, season):
    """Fetch wRC+ por equipo desde FanGraphs (batting, team totals)."""
    print(f"  📊 wRC+ equipos {season}...")
    # Params que corresponden exactamente a la URL del usuario:
    # /leaders/major-league?stats=bat&type=8&team=0,ts&qual=0&season=XXXX
    df = fg_api(session, {"pos":"all","stats":"bat","lg":"all","qual":"0",
                          "type":"8","season":str(season),"season1":str(season),
                          "team":"0,ts","rost":"0","ind":"0"})
    result = {}
    if df.empty:
        print(f"    ⚠️  API retornó 0 filas"); return result

    team_col = _fg_team_col(df)

    # wRC+ es una tasa (league avg=100). wRC es la acumulación cruda (>500 en temporada completa).
    # Buscamos "wRC+" con el signo + primero para no agarrar la columna "wRC" equivocada.
    wrc_col = (
        next((c for c in df.columns if c == "wRC+"), None) or
        next((c for c in df.columns if c.lower() == "wrc+"), None) or
        next((c for c in df.columns if "+" in c and "wrc" in c.lower()), None) or
        next((c for c in df.columns if c.lower() == "wrc"), None)   # último recurso
    )

    if wrc_col is None:
        print(f"    ⚠️  Col wRC+ no encontrada")
        print(f"         Columnas disponibles: {list(df.columns)}")
        return result

    for _, row in df.iterrows():
        raw_team = row.get(team_col, "") if team_col else ""
        team = _fg_resolve_team(raw_team)
        val  = row.get(wrc_col)
        if team and val is not None:
            try: result[team] = round(float(val), 1)
            except: pass

    if not result:
        print(f"    ⚠️  0 equipos mapeados — team_col={team_col!r}  wrc_col={wrc_col!r}")
        print(f"         Todas las columnas: {list(df.columns)}")
        for i in range(min(3, len(df))):
            row = df.iloc[i]
            t_raw = row.get(team_col, "?") if team_col else "?"
            print(f"         fila {i}: raw_team={str(t_raw)[:80]!r}  abb={_fg_extract_abb(t_raw)!r}  {wrc_col}={row.get(wrc_col,'?')!r}")
    else:
        print(f"    ✅ {len(result)} equipos  (col usada: {wrc_col!r})")
    return result


def get_team_wrc_splits(session, season):
    """
    Fetch wRC+ split por equipo vs LHP y vs RHP desde FanGraphs.
    Retorna: {"vs_lhp": {team: wrc+}, "vs_rhp": {team: wrc+}}

    FIX 2026-05-10: El endpoint team=0,ts IGNORA el parámetro split=5/6.
    La forma correcta es usar month=13 (vs LHP) y month=14 (vs RHP),
    que corresponden a los filtros internos de FanGraphs para platoon splits.
    type=1 (Standard) es el que devuelve wRC+ en splits de month.
    """
    def _fetch_split(month_id, label):
        try:
            df = fg_api(session, {
                "pos":"all","stats":"bat","lg":"all","qual":"0",
                "type":"8","season":str(season),"season1":str(season),
                "team":"0,ts","rost":"0","ind":"0","month":str(month_id),
            })
            if df.empty:
                print(f"    ⚠️  wRC+ splits ({label}): API retornó 0 filas")
                return {}
            team_col = _fg_team_col(df)
            wrc_col = (
                next((c for c in df.columns if c == "wRC+"), None) or
                next((c for c in df.columns if c.lower() == "wrc+"), None)
            )
            if not wrc_col or not team_col:
                print(f"    ⚠️  wRC+ splits ({label}): col wRC+ o team no encontrada. Cols: {list(df.columns)}")
                return {}
            result = {}
            for _, row in df.iterrows():
                team = _fg_resolve_team(row.get(team_col, ""))
                val  = row.get(wrc_col)
                if team and val is not None:
                    try: result[team] = round(float(val), 1)
                    except: pass
            if result:
                print(f"    ✅ wRC+ {label}: {len(result)} equipos")
            else:
                print(f"    ⚠️  wRC+ splits ({label}): 0 equipos resueltos")
            return result
        except Exception as e:
            print(f"    ⚠️  wRC+ splits ({label}) no disponible: {e}")
            return {}

    print(f"  📊 wRC+ splits (vs LHP/RHP) {season} — usando month=13/14...")
    vs_lhp = _fetch_split(13, "vs LHP")   # month=13 → Batters vs LHP
    vs_rhp = _fetch_split(14, "vs RHP")   # month=14 → Batters vs RHP
    return {"vs_rhp": vs_rhp, "vs_lhp": vs_lhp}


def get_bullpen_xfip(session, season):
    """Fetch xFIP por equipo desde FanGraphs (relievers, team totals).
    URL base del usuario: stats=rel&type=8&team=0,ts&qual=0&season=XXXX
    """
    print(f"  🔥 Bullpen xFIP {season}...")
    # stats=rel (Reliever leaderboard), type=8 (Dashboard), team=0,ts (team totals)
    # Esto coincide con la URL exacta que el usuario proporcionó
    df = fg_api(session, {"pos":"all","stats":"rel","lg":"all","qual":"0",
                          "type":"8","season":str(season),"season1":str(season),
                          "team":"0,ts","rost":"0","ind":"0"})
    result = {}
    if df.empty:
        print(f"    ⚠️  API retornó 0 filas"); return result

    xfip_col = next((c for c in df.columns if "xfip" in c.lower()), None)
    team_col = _fg_team_col(df)

    if xfip_col is None:
        print(f"    ⚠️  Col xFIP no encontrada")
        print(f"         Columnas disponibles: {list(df.columns)}")
        return result

    for _, row in df.iterrows():
        raw_team = row.get(team_col, "") if team_col else ""
        team = _fg_resolve_team(raw_team)
        val  = row.get(xfip_col)
        if team and val is not None:
            try: result[team] = round(float(val), 3)
            except: pass

    if not result:
        print(f"    ⚠️  0 equipos mapeados — team_col={team_col!r}  xfip_col={xfip_col!r}")
        print(f"         Todas las columnas: {list(df.columns)}")
        for i in range(min(3, len(df))):
            row = df.iloc[i]
            t_raw = row.get(team_col, "?") if team_col else "?"
            print(f"         fila {i}: raw_team={str(t_raw)[:80]!r}  abb={_fg_extract_abb(t_raw)!r}  {xfip_col}={row.get(xfip_col,'?')!r}")
    else:
        print(f"    ✅ {len(result)} equipos  (col usada: {xfip_col!r})")
    return result


def _fg_extract_name(raw_val):
    """
    FanGraphs devuelve el nombre del pitcher también como HTML anchor:
      '<a href="...">Gerrit Cole</a>'
    Esta función extrae el texto del anchor (el nombre).
    """
    s = str(raw_val).strip()
    m = re.search(r'>([^<]+)</a>', s)
    if m:
        return m.group(1).strip().upper()
    return s.upper()


def get_sp_xfip(session, season):
    """
    Fetch xFIP por SP desde FanGraphs (starters individuales).
    URL base: stats=sta & type=8 & team=0 & qual=0 & pageitems=2000000000
    Retorna {name_key: xfip_value} con claves de nombre completo y apellido.
    """
    print(f"  ⚾  SP xFIP {season}...")
    df = fg_api(session, {"pos":"all","stats":"sta","lg":"all","qual":"0",
                          "type":"8","season":str(season),"season1":str(season),
                          "team":"0","rost":"0","ind":"0",
                          "pageitems":"2000000000"})
    result = {}
    if df.empty:
        print(f"    ⚠️  API retornó 0 filas"); return result

    xfip_col = next((c for c in df.columns if "xfip" in c.lower()), None)
    name_col = next((c for c in df.columns
                     if "playername" in c.lower() or c.lower() in ("name","playerid")), None)
    # fallback: primera columna de texto
    if name_col is None:
        name_col = next((c for c in df.columns if "name" in c.lower()), None)

    if xfip_col is None:
        print(f"    ⚠️  Col xFIP no encontrada — cols: {list(df.columns)[:15]}")
        return result
    if name_col is None:
        print(f"    ⚠️  Col nombre no encontrada — cols: {list(df.columns)[:15]}")
        return result

    for _, row in df.iterrows():
        raw_name = row.get(name_col, "")
        name = _fg_extract_name(raw_name)      # strip HTML anchor
        val  = row.get(xfip_col)
        if name and val is not None:
            try:
                v = round(float(val), 2)
                for k in name_keys(name):      # guarda por nombre completo + apellido
                    result[k] = v
            except: pass

    if not result:
        print(f"    ⚠️  0 pitchers — name_col={name_col!r}  xfip_col={xfip_col!r}")
        for i in range(min(3, len(df))):
            row = df.iloc[i]
            print(f"         fila {i}: name={str(row.get(name_col,'?'))[:60]!r}  xFIP={row.get(xfip_col,'?')!r}")
    else:
        print(f"    ✅ {len(df)} starters  (col usada: {xfip_col!r}  nombre: {name_col!r})")
    return result


def get_pitcher_fip(session, season):
    """Legacy — mantener por si acaso. SP xFIP ahora viene de get_sp_xfip."""
    print(f"  ⚾  Pitcher FIP {season} (legacy)...")
    df = fg_api(session, {"pos":"all","stats":"pit","lg":"all","qual":"0",
                          "type":"4","season":str(season),"season1":str(season),"team":"0"})
    fip_col  = next((c for c in df.columns if c.lower()=="fip"), "FIP")
    name_col = next((c for c in df.columns if "playername" in c.lower() or c.lower()=="name"), "PlayerName")
    result = {}
    for _, row in df.iterrows():
        name = _fg_extract_name(str(row.get(name_col,"")))
        val  = row.get(fip_col)
        if name and val is not None:
            try:
                for k in name_keys(name): result[k] = round(float(val), 2)
            except: pass
    print(f"    ✅ {len(df)} pitchers")
    return result


# ──────────────────────────────────────────────────────
# FANGRAPHS CACHE — wRC+ y Bullpen xFIP (bypasa Excel)
# ──────────────────────────────────────────────────────

def load_fg_cache():
    """Lee caché JSON de FanGraphs. Retorna dict o {}."""
    if os.path.exists(FG_CACHE_FILE):
        try:
            with open(FG_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_fg_cache(raw):
    """Guarda datos FanGraphs a caché JSON junto al script."""
    raw["fetched_at"] = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    try:
        with open(FG_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(raw, f, indent=2)
    except Exception as e:
        print(f"  ⚠️  No se pudo guardar caché FG: {e}")


def blend_fg_data(raw):
    """
    Combina datos 2025/2026 de FanGraphs con pesos ADAPTATIVOS por temporada.

    wRC+ equipo — peso actual crece con juegos jugados:
      Abril (~15j): ~12% actual / 88% previo
      Mayo  (~40j): ~27% actual
      Junio (~70j): ~46% actual
      Septiembre (162j): ~75% actual

    BP xFIP — cambia más lento (bullpen rota más, muestra sample pequeño):
      Abril: ~15% actual / 85% previo
      Final: máx 40% actual (bullpens son más variables)

    Retorna {team: {"wrc": X, "bp_xfip": Y}}
    """
    g_cur  = _season_games_estimate(TARGET_DATE)
    w_cur  = _adaptive_wrc_blend(g_cur)       # peso año actual para wRC+
    w_prev = 1.0 - w_cur
    # Bullpen: más conservador (cap 40% actual — más ruido a principio)
    bp_cur  = min(0.40, w_cur * 0.55)
    bp_prev = 1.0 - bp_cur

    data = {}
    all_teams = (set(raw.get("wrc_2025", {})) | set(raw.get("wrc_2026", {})) |
                 set(raw.get("bp_2025",  {})) | set(raw.get("bp_2026",  {})))
    # ── Parámetros para el ajuste de desviación temprana ──────────────────
    # Cuando un equipo muestra una diferencia grande entre su wRC+ de 2025 y
    # el de 2026 (p.ej. Mets: 112 → 81), el blend puro ignora el 90% de esa
    # señal en abril. Este ajuste aplica parcialmente la dirección del cambio
    # para no inflar (o desinflar) las proyecciones de manera irreal.
    #   DEV_MIN_DIFF : diferencia mínima (puntos wRC+) para activar el ajuste
    #   DEV_SIGNAL   : fracción de la desviación que se trata como señal real
    #   DEV_MAX_ADJ  : tope del ajuste para no sobreponderar muestras pequeñas
    # El ajuste escala con w_prev → se desvanece naturalmente cuando la
    # temporada avanza y el año actual ya tiene su propio peso significativo.
    DEV_MIN_DIFF = 20.0   # ≥20 puntos de diferencia para activar
    DEV_SIGNAL   = 0.35   # 35% de la desviación se considera señal
    DEV_MAX_ADJ  = 15.0   # tope: máximo ±15 puntos de corrección

    for team in all_teams:
        w25 = raw.get("wrc_2025", {}).get(team)
        w26 = raw.get("wrc_2026", {}).get(team)
        b25 = raw.get("bp_2025",  {}).get(team)
        b26 = raw.get("bp_2026",  {}).get(team)
        wrc = (w25 * w_prev + w26 * w_cur) if (w25 and w26) else (w25 or w26)
        bp  = (b25 * bp_prev + b26 * bp_cur) if (b25 and b26) else (b25 or b26)

        # Ajuste de señal temprana: si el equipo muestra una desviación grande
        # en wRC+ respecto al año previo, aplicar una corrección parcial.
        if wrc and w25 and w26 and abs(w26 - w25) >= DEV_MIN_DIFF:
            delta      = w26 - w25                        # positivo = mejoró, negativo = empeoró
            correction = delta * DEV_SIGNAL * w_prev      # se desvanece al avanzar la temporada
            correction = max(-DEV_MAX_ADJ, min(DEV_MAX_ADJ, correction))
            wrc        = max(50.0, min(165.0, wrc + correction))

        if wrc:
            data[team] = {"wrc":     round(float(wrc), 1),
                          "bp_xfip": round(float(bp),  3) if bp else None}
    return data


def load_wrc_splits():
    """
    Lee los platoon splits wRC+ vs LHP/RHP desde la caché JSON de FanGraphs.
    Retorna {"vs_rhp": {team: wrc}, "vs_lhp": {team: wrc}}.
    Si no hay datos en caché, o si los splits son idénticos (FanGraphs
    ignoró el split= param y devolvió wRC+ general para ambos), retorna
    dicts vacíos para que el modelo use base wRC+ sin ajuste falso.
    """
    raw    = load_fg_cache()
    vs_rhp = raw.get("wrc_vs_rhp", {})
    vs_lhp = raw.get("wrc_vs_lhp", {})

    # Validación: si >70% de los equipos tienen el mismo valor en ambos
    # splits, el API devolvió datos inválidos (team totals ignora split=).
    if vs_rhp and vs_lhp:
        common = set(vs_rhp) & set(vs_lhp)
        if common:
            n_same = sum(1 for t in common if vs_rhp.get(t) == vs_lhp.get(t))
            if n_same / len(common) > 0.70:
                # Splits rotos — evitar ajuste de platoon basado en datos falsos
                return {"vs_rhp": {}, "vs_lhp": {}, "_splits_broken": True}

    return {"vs_rhp": vs_rhp, "vs_lhp": vs_lhp}


def load_sp_xfip_blended():
    """
    Lee SP xFIP blended desde caché JSON con pesos ADAPTATIVOS.

    Un SP hace ~1 salida cada 5 días, por lo que tiene menos sample
    que los batters en las mismas semanas de temporada.
      ~3 GS (abril):   ~12% actual / 88% previo
      ~10 GS (mayo):   ~35% actual
      ~20 GS (julio):  ~65% actual
      ~30+ GS:         ~80% actual

    Retorna {name_key: xfip_value} — claves: nombre completo y apellido en mayúsculas.
    """
    raw  = load_fg_cache()
    sp25 = raw.get("sp_xfip_2025", {})
    sp26 = raw.get("sp_xfip_2026", {})

    g_cur  = _season_games_estimate(TARGET_DATE)
    gs_cur = max(0, g_cur // 5)                  # aprox GS del pitcher
    w_cur  = _adaptive_sp_blend(gs_cur)
    w_prev = 1.0 - w_cur

    result = {}
    for k in set(sp25) | set(sp26):
        v25 = sp25.get(k); v26 = sp26.get(k)
        if v25 and v26: result[k] = round(w_prev * v25 + w_cur * v26, 2)
        elif v25:       result[k] = v25
        elif v26:       result[k] = v26
    return result


def load_fg_blended(fg=None):
    """
    Carga wRC+, Bullpen xFIP y SP xFIP blended desde caché JSON.
    - Si la caché es del día de hoy → usa caché (no refetch)
    - Si stale y fg disponible → refetch todo + actualiza caché
    - Si fg no disponible → usa caché stale si existe
    - Sin caché → defaults (wRC+=100, xFIP=4.2)
    Retorna {team: {wrc, bp_xfip}} — SP xFIP queda en caché y se
    accede por separado con load_sp_xfip_blended().
    """
    raw = load_fg_cache()
    fetched_at = raw.get("fetched_at", "")
    cache_date = fetched_at[:10] if fetched_at else ""
    today      = date.today().strftime("%Y-%m-%d")

    if cache_date == today:
        n_teams = len(blend_fg_data(raw))
        n_sp    = len(raw.get("sp_xfip_2025", {}))
        print(f"  📋 FG caché del día ({fetched_at[11:16]}) — {n_teams} equipos, {n_sp} SP")
        return blend_fg_data(raw)

    if fg is not None:
        print("  🔄 FG caché desactualizada — fetching wRC+, BP xFIP y SP xFIP...")
        def safe_fetch(fn, *a):
            try:   return fn(*a)
            except requests.exceptions.HTTPError as e:
                if "403" in str(e): print("  ❌ 403 FanGraphs — necesitas cookie")
                else:               print(f"  ⚠️  {e}")
                return {}
            except Exception as e: print(f"  ⚠️  {e}"); return {}

        wrc26  = safe_fetch(get_team_wrc,       fg, SEASON)
        wrc25  = safe_fetch(get_team_wrc,       fg, "2025")
        bp26   = safe_fetch(get_bullpen_xfip,   fg, SEASON)
        bp25   = safe_fetch(get_bullpen_xfip,   fg, "2025")
        sp26   = safe_fetch(get_sp_xfip,        fg, SEASON)
        sp25   = safe_fetch(get_sp_xfip,        fg, "2025")
        splits = safe_fetch(get_team_wrc_splits, fg, SEASON)

        got_something = any([wrc26, wrc25, bp26, bp25, sp26, sp25])
        if got_something:
            if wrc26:  raw["wrc_2026"]       = wrc26
            if wrc25:  raw["wrc_2025"]       = wrc25
            if bp26:   raw["bp_2026"]        = bp26
            if bp25:   raw["bp_2025"]        = bp25
            if sp26:   raw["sp_xfip_2026"]   = sp26
            if sp25:   raw["sp_xfip_2025"]   = sp25
            if splits.get("vs_rhp"):
                raw["wrc_vs_rhp"] = splits["vs_rhp"]
            if splits.get("vs_lhp"):
                raw["wrc_vs_lhp"] = splits["vs_lhp"]
            save_fg_cache(raw)
            n_sp = len(raw.get("sp_xfip_2025", {}))
            print(f"  ✅ FG caché actualizada — {len(blend_fg_data(raw))} equipos, {n_sp} SP")
        else:
            if raw.get("wrc_2025"):
                print(f"  ⚠️  Fetch falló — usando caché stale ({cache_date})")
            else:
                print(f"  ⚠️  Sin datos FG — usando defaults (wRC+=100, xFIP=4.2)")
    else:
        n_sp = len(raw.get("sp_xfip_2025", {}))
        if raw.get("wrc_2025"):
            print(f"  📋 FG caché ({cache_date}) — sin sesión FG, reutilizando ({n_sp} SP en caché)")
        else:
            print(f"  ⚠️  Sin caché FG y sin sesión — usando defaults")

    return blend_fg_data(raw)


# ──────────────────────────────────────────────────────
# LEER DATOS DE MLB DATA (Excel — pitcher xFIP manual)
# ──────────────────────────────────────────────────────

def load_mlb_data(wb):
    """Lee wRC+ y BP xFIP blended desde MLB Data."""
    ws = wb[DATA_SHEET]
    data = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        abb = str(row[2]).strip().upper() if row[2] else None
        if not abb or abb not in TEAM_ABB: continue
        team = TEAM_ABB[abb]
        k = row[10]; l = row[11]; n = row[13]; o = row[14]
        wrc   = (n*0.75 + l*0.25) if (n and l) else (n or l)
        bpxfp = (o*0.9  + k*0.1)  if (o and k) else (o or k)
        if wrc:
            try: data[team] = {"wrc":round(float(wrc),1),
                               "bp_xfip":round(float(bpxfp),3) if bpxfp else None}
            except: pass
    return data


def load_mlb_data_raw(wb):
    """Lee valores RAW 2026 y 2025 separados desde MLB Data."""
    ws = wb[DATA_SHEET]
    data = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        abb = str(row[2]).strip().upper() if row[2] else None
        if not abb or abb not in TEAM_ABB: continue
        team = TEAM_ABB[abb]
        data[team] = {
            "bp_2026":  row[10],  # col K
            "wrc_2026": row[11],  # col L
            "wrc_2025": row[13],  # col N
            "bp_2025":  row[14],  # col O
        }
    return data


def load_batx_table(wb):
    ws = wb[DATA_SHEET]
    batx = {}
    for row in ws.iter_rows(min_row=5, min_col=16, max_col=17, values_only=True):
        name, xfip = row[0], row[1]
        if name and isinstance(xfip, (int, float)):
            for k in name_keys(str(name)):
                batx[k] = round(float(xfip), 2)
    return batx


# ──────────────────────────────────────────────────────
# SCHEDULE MLB
# ──────────────────────────────────────────────────────

def get_mlb_schedule(target_date, silent=False):
    if not silent: print(f"  📅 Schedule MLB {target_date}...")
    r = requests.get(
        f"https://statsapi.mlb.com/api/v1/schedule?"
        f"sportId=1&date={target_date}&hydrate=probablePitcher,team,venue",
        timeout=15)
    r.raise_for_status()
    games = []
    seen_pks = set()   # deduplicar por gamePk — el API a veces devuelve el mismo juego 2x
    for de in r.json().get("dates", []):
        for g in de.get("games", []):
            if g.get("status",{}).get("abstractGameState") == "Final": continue
            pk = g.get("gamePk")
            if pk and pk in seen_pks:
                continue          # duplicado exacto — descartar
            if pk:
                seen_pks.add(pk)
            away = TEAM_MAP.get(g["teams"]["away"]["team"]["name"],
                                 g["teams"]["away"]["team"]["name"].upper())
            home = TEAM_MAP.get(g["teams"]["home"]["team"]["name"],
                                 g["teams"]["home"]["team"]["name"].upper())
            away_pp  = g["teams"]["away"].get("probablePitcher", {})
            home_pp  = g["teams"]["home"].get("probablePitcher", {})
            away_sp  = away_pp.get("fullName", "TBD")
            home_sp  = home_pp.get("fullName", "TBD")
            away_pid = away_pp.get("id")
            home_pid = home_pp.get("id")
            game_num   = g.get("gameNumber", 1)       # 1 = single/G1, 2 = G2
            dh         = g.get("doubleHeader", "N")   # "Y" = DH tradicional, "S" = split
            is_dh      = dh in ("Y", "S")
            games.append({
                "away": away, "home": home,
                "away_sp":     away_sp.upper() if away_sp != "TBD" else "TBD",
                "home_sp":     home_sp.upper() if home_sp != "TBD" else "TBD",
                "away_sp_id":  away_pid,
                "home_sp_id":  home_pid,
                "game_time_utc": g.get("gameDate",""),
                "game_pk":     pk,
                "game_number": game_num,
                "double_header": dh,
                # Label extra para doubleheaders → "ASTROS @ ORIOLES G2"
                "dh_label":    f" G{game_num}" if is_dh else "",
                "away_team_id": g["teams"]["away"]["team"].get("id"),
                "home_team_id": g["teams"]["home"]["team"].get("id"),
            })
    if not silent: print(f"    ✅ {len(games)} juegos")
    return games


# Caché de lateralidad de pitcher (id → "L" | "R" | "S")
_PITCHER_HAND_CACHE = {}

def _get_pitcher_hand(pitcher_id):
    """
    Retorna la lateralidad del pitcher ('L', 'R', 'S') desde la API de MLB.
    Usa caché en memoria para evitar requests repetidos.
    """
    if pitcher_id is None:
        return None
    if pitcher_id in _PITCHER_HAND_CACHE:
        return _PITCHER_HAND_CACHE[pitcher_id]
    try:
        url = f"https://statsapi.mlb.com/api/v1/people/{pitcher_id}?fields=people,pitchHand"
        r   = requests.get(url, timeout=10)
        r.raise_for_status()
        hand = r.json()["people"][0]["pitchHand"]["code"]   # "L", "R" o "S"
        _PITCHER_HAND_CACHE[pitcher_id] = hand
        return hand
    except Exception:
        return None


# ──────────────────────────────────────────────────────
# UMPIRE MODULE — tendencias de HP ump sobre totales
# ──────────────────────────────────────────────────────

# Tendencias históricas de umpires conocidos (backup si scraping falla)
# Actualizado con datos 2023-2025 de umpirescorecards.com. Unidades: runs/game vs avg.
UMP_STATIC_FALLBACK = {
    "CB BUCKNOR":        +0.52, "BUCKNOR":       +0.52,
    "LAZ DIAZ":          +0.44, "DIAZ":          +0.44,
    "ANGEL HERNANDEZ":   +0.38, "HERNANDEZ":     +0.38,
    "JEREMIE REHAK":     +0.35, "REHAK":         +0.35,
    "HUNTER WENDELSTEDT":+0.32, "WENDELSTEDT":  +0.32,
    "BRIAN GORMAN":      +0.28, "GORMAN":        +0.28,
    "QUINN WOLCOTT":     +0.26, "WOLCOTT":       +0.26,
    "MARVIN HUDSON":     +0.24, "HUDSON":        +0.24,
    "RON KULPA":         +0.21, "KULPA":         +0.21,
    "PAUL NAUERT":       +0.18, "NAUERT":        +0.18,
    "BILL MILLER":       +0.15, "MILLER":        +0.15,
    "MARK CARLSON":      +0.09, "CARLSON":       +0.09,
    "JEFF NELSON":       +0.07, "NELSON":        +0.07,
    "DAN BELLINO":       +0.05, "BELLINO":       +0.05,
    "MIKE ESTABROOK":    +0.04, "ESTABROOK":     +0.04,
    "ROBERTO ORTIZ":     -0.04, "ORTIZ":         -0.04,
    "CHAD FAIRCHILD":    -0.07, "FAIRCHILD":     -0.07,
    "ADAM HAMARI":       -0.09, "HAMARI":        -0.09,
    "JANSEN VISCONTI":   -0.12, "VISCONTI":      -0.12,
    "MIKE MUCHLINSKI":   -0.15, "MUCHLINSKI":    -0.15,
    "JOHN TUMPANE":      -0.17, "TUMPANE":       -0.17,
    "MARK WEGNER":       -0.19, "WEGNER":        -0.19,
    "TED BARRETT":       -0.21, "BARRETT":       -0.21,
    "TOM HALLION":       -0.23, "HALLION":       -0.23,
    "WILL LITTLE":       -0.26, "LITTLE":        -0.26,
    "MIKE WINTERS":      -0.29, "WINTERS":       -0.29,
    "CHRIS GUCCIONE":    -0.31, "GUCCIONE":      -0.31,
    "DOUG EDDINGS":      -0.34, "EDDINGS":       -0.34,
    "JAMES HOYE":        -0.41, "HOYE":          -0.41,
    "FIELDIN CULBRETH":  -0.44, "CULBRETH":      -0.44,
    "PHIL CUZZI":        -0.48, "CUZZI":         -0.48,
}


def _parse_ump_names(name, val):
    """Normaliza nombre de umpire y agrega variantes (completo, apellido, F.Apellido)."""
    out = {}
    n = name.strip().upper()
    if not n: return out
    out[n] = val
    parts = n.split()
    if len(parts) >= 2:
        out[parts[-1]] = val
        out[f"{parts[0][0]}.{parts[-1]}"] = val
    return out


def _scrape_ump_tendencies():
    """
    Obtiene tendencias de umpires desde umpscorecards.com.
    Intenta varios endpoints en orden; si todos fallan usa fallback estático.
    """
    import re as _re

    year = TARGET_DATE[:4]
    hdrs = {
        "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36"),
        "Accept": "application/json, */*",
        "Referer": "https://umpscorecards.com/",
        "Origin":  "https://umpscorecards.com",
    }

    # ── Prueba varios endpoints en orden ──────────────────────────────
    _candidates = [
        f"https://umpscorecards.com/data/games?season={year}",
        f"https://umpscorecards.com/data/games?year={year}",
        "https://umpscorecards.com/data/games",
        f"https://umpscorecards.com/api/games?season={year}",
        "https://umpscorecards.com/api/games",
    ]
    games = None
    _last_err = ""
    for _url in _candidates:
        try:
            r = requests.get(_url, headers=hdrs, timeout=15)
            r.raise_for_status()
            if not r.text or not r.text.strip():
                _last_err = "respuesta vacía"
                continue
            if r.text.strip()[0] not in ("[", "{"):
                _last_err = "respuesta no-JSON (HTML?)"
                continue
            games = r.json()
            break   # éxito
        except Exception as e:
            _last_err = str(e)
            continue

    if games is None:
        # No anunciar error verbose — simplemente retorna {} para usar fallback
        return {}

    # ── Normalizar estructura ──────────────────────────────────────────
    try:

        # Normalizar: aceptar lista directa o wrapper {"data": [...]}
        if isinstance(games, dict):
            games = (games.get("data") or games.get("games") or
                     games.get("results") or list(games.values())[0])
        if not isinstance(games, list) or not games:
            raise ValueError("estructura inesperada")

        # Detectar nombres de campos automáticamente desde el primer item
        sample = games[0] if games else {}
        keys   = [k.lower() for k in sample.keys()]

        # Campo del umpire HP
        ump_field = next((k for k in sample.keys()
                         if any(x in k.lower() for x in
                                ("hp_umpire","hp_ump","homeplate","home_plate",
                                 "umpire","ump","official"))), None)

        # Campo de run favor pre-calculado (preferido)
        favor_field = next((k for k in sample.keys()
                           if any(x in k.lower() for x in
                                  ("run_favor","run_diff","runs_above","runs_diff",
                                   "runfavor","run_advantage","favor"))), None)

        # Campos para calcular run favor manualmente si no viene pre-calc
        actual_field   = next((k for k in sample.keys()
                               if any(x in k.lower() for x in
                                      ("total_runs","total_score","runs_total","actual_runs"))), None)
        expected_field = next((k for k in sample.keys()
                               if any(x in k.lower() for x in
                                      ("expected","proj","predict","xruns","x_runs"))), None)

        # Campo de fecha para filtrar por temporada
        date_field = next((k for k in sample.keys()
                          if any(x in k.lower() for x in ("date","game_date","gamedate"))), None)

        if not ump_field:
            print(f"  [ump] no se encontró campo de umpire. Campos: {list(sample.keys())[:10]}")
            return {}

        # Agregar por umpire
        from collections import defaultdict as _dd
        ump_runs = _dd(list)

        for g in games:
            # Filtrar por temporada actual
            if date_field:
                gdate = str(g.get(date_field, ""))
                if not gdate.startswith(year):
                    continue

            ump_name = str(g.get(ump_field, "")).strip()
            if not ump_name or ump_name.lower() in ("", "tbd", "none", "null"):
                continue

            # Obtener run favor
            favor = None
            if favor_field:
                try: favor = float(g[favor_field])
                except (ValueError, TypeError): pass

            if favor is None and actual_field and expected_field:
                try:
                    favor = float(g[actual_field]) - float(g[expected_field])
                except (ValueError, TypeError): pass

            if favor is not None and -15 < favor < 15:  # filtrar outliers
                ump_runs[ump_name].append(favor)

        if not ump_runs:
            print(f"  [ump] sin datos válidos (año {year}). ¿El endpoint filtra por año?")
            # Intentar sin filtro de año (quizás solo devuelve la temporada activa)
            ump_runs = _dd(list)
            for g in games:
                ump_name = str(g.get(ump_field, "")).strip()
                if not ump_name or ump_name.lower() in ("", "tbd", "none", "null"):
                    continue
                favor = None
                if favor_field:
                    try: favor = float(g[favor_field])
                    except (ValueError, TypeError): pass
                if favor is None and actual_field and expected_field:
                    try: favor = float(g[actual_field]) - float(g[expected_field])
                    except (ValueError, TypeError): pass
                if favor is not None and -15 < favor < 15:
                    ump_runs[ump_name].append(favor)

        if not ump_runs:
            print(f"  [ump] sin datos de run favor en el endpoint")
            return {}

        # Calcular promedio por umpire → dict de tendencias
        data = {}
        for ump_name, favors in ump_runs.items():
            if len(favors) >= 1:
                avg = round(sum(favors) / len(favors), 3)
                data.update(_parse_ump_names(ump_name, avg))

        print(f"  ✅ {len(ump_runs)} umpires desde umpscorecards.com "
              f"({sum(len(v) for v in ump_runs.values())} juegos)")
        return data

    except Exception:
        return {}  # silencioso — load_ump_cache() mostrará fallback message


def load_ump_cache():
    """
    Lee caché de tendencias de umpires.
    Orden: caché del día → scrape live → fallback estático.
    """
    # 1. Caché del día
    try:
        if os.path.exists(UMP_CACHE_FILE):
            with open(UMP_CACHE_FILE, "r", encoding="utf-8") as f:
                c = json.load(f)
            if c.get("date") == TARGET_DATE and c.get("tendencies"):
                n = len([k for k in c["tendencies"] if " " in k])  # solo nombres completos
                print(f"  🧑‍⚖️  Umpire caché del día ({n} umps)")
                return c["tendencies"]
    except Exception:
        pass

    # 2. Intentar scrape live
    print("  🧑‍⚖️  Actualizando tendencias de umpires desde umpirescorecards.com...")
    data = _scrape_ump_tendencies()

    if data:
        # Combinar con fallback estático (live gana en caso de conflicto)
        merged = {**UMP_STATIC_FALLBACK, **data}
        try:
            with open(UMP_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump({"date": TARGET_DATE, "tendencies": merged, "source": "live"}, f, indent=2)
            print(f"  ✅ {len([k for k in data if ' ' in k])} umpires live  +  {len(UMP_STATIC_FALLBACK)//2} fallback")
        except Exception as e:
            print(f"  [ump] cache write error: {e}")
        return merged

    # 3. Fallback estático (siempre disponible)
    print(f"  🧑‍⚖️  Usando tabla histórica de umpires ({len(UMP_STATIC_FALLBACK)//2} umps conocidos)")
    try:
        with open(UMP_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({"date": TARGET_DATE, "tendencies": UMP_STATIC_FALLBACK, "source": "static"}, f, indent=2)
    except Exception:
        pass
    return UMP_STATIC_FALLBACK


def get_game_umpires(target_date):
    """
    Obtiene HP umpire por juego desde MLB Stats API (hydrate=officials).
    Retorna {(away, home): ump_name_upper} — empty dict si no disponible.
    Umpires normalmente disponibles desde la mañana del día del juego.
    """
    try:
        r = requests.get(
            f"https://statsapi.mlb.com/api/v1/schedule?"
            f"sportId=1&date={target_date}&hydrate=officials,team",
            timeout=12)
        r.raise_for_status()
        result = {}
        for de in r.json().get("dates", []):
            for g in de.get("games", []):
                officials = g.get("officials", [])
                hp = next((o for o in officials
                           if o.get("officialType","").lower() == "home plate"), None)
                if hp:
                    ump_name = hp["official"]["fullName"].upper()
                    away = TEAM_MAP.get(g["teams"]["away"]["team"]["name"],
                                        g["teams"]["away"]["team"]["name"].upper())
                    home = TEAM_MAP.get(g["teams"]["home"]["team"]["name"],
                                        g["teams"]["home"]["team"]["name"].upper())
                    result[(away, home)] = ump_name
        return result
    except Exception as e:
        print(f"  [ump] assignment fetch error: {e}")
        return {}


def ump_total_factor(ump_name, ump_tendencies, baseline=8.6):
    """
    Convierte runs/juego ajuste del ump → factor multiplicativo para el total.
    Cap: ±3% (máximo ±0.26 runs en un juego de ~8.6) para no sobreponderar.

    Ejemplo:
      CB Bucknor runs_adj = +0.45 → factor = 1 + (0.45/8.6) = 1.052 → capped 1.030
      Doug Eddings runs_adj = -0.38 → factor = 1 - (0.38/8.6) = 0.956 → capped 0.970
    """
    if not ump_name or not ump_tendencies:
        return 1.0
    # Buscar en distintos formatos de nombre
    runs_adj = None
    for key in [ump_name, ump_name.split()[-1],
                f"{ump_name.split()[0][0]}.{ump_name.split()[-1]}" if len(ump_name.split()) >= 2 else ""]:
        if key in ump_tendencies:
            runs_adj = ump_tendencies[key]
            break
    if runs_adj is None:
        return 1.0  # umpire desconocido → sin ajuste
    raw_factor = 1.0 + (runs_adj / baseline)
    # Cap en ±3% para no sobreponderar ningún ump
    return round(max(0.970, min(1.030, raw_factor)), 4)


def format_game_time(game_utc, home_team):
    if not game_utc: return ""
    try:
        clean  = game_utc.replace("Z","").split("+")[0]
        utc_dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
        offset = STADIUM_TZ_OFFSET.get(home_team, -5)
        ldt    = utc_dt + timedelta(hours=offset)
        h, m   = ldt.hour, ldt.minute
        ap     = "PM" if h >= 12 else "AM"
        h12    = h % 12 or 12
        tz     = TZ_ABBR.get(offset, "LT")
        return f"{h12}:{m:02d} {ap} {tz}"
    except: return ""


def _utc_to_local_hour(game_utc, home_team):
    if not game_utc: return None
    try:
        clean  = game_utc.replace("Z","").split("+")[0]
        utc_dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
        offset = STADIUM_TZ_OFFSET.get(home_team, -5)
        return (utc_dt + timedelta(hours=offset)).hour
    except: return None


# ──────────────────────────────────────────────────────
# WEATHER — hora del juego + orientación del parque
# ──────────────────────────────────────────────────────

def wind_direction_for_park(raw_dir, home_team):
    if home_team in DOME_TEAMS or not raw_dir:
        return "DOME" if home_team in DOME_TEAMS else raw_dir
    from_deg = COMPASS_BEARING.get(str(raw_dir).upper())
    cf_deg   = STADIUM_CF_BEARING.get(home_team)
    if from_deg is None or cf_deg is None:
        return _wind_generic(raw_dir)
    toward   = (from_deg + 180) % 360
    diff     = abs(toward - cf_deg)
    if diff > 180: diff = 360 - diff
    if diff <= 60:  return "OUT"
    if diff >= 120: return "IN"
    return "CROSS"


def _wind_generic(d):
    if d in {"N","NNE","NNW"}:           return "OUT"
    if d in {"S","SSE","SSW","SE","SW"}: return "IN"
    if d in {"E","ENE","ESE"}:           return "L-R"
    if d in {"W","WNW","NW"}:            return "R-L"
    return d


def _deg_to_compass(deg):
    """Convert wind direction degrees (meteorological) to 16-point compass string."""
    directions = ["N","NNE","NE","ENE","E","ESE","SE","SSE",
                  "S","SSW","SW","WSW","W","WNW","NW","NNW"]
    idx = round(float(deg) / 22.5) % 16
    return directions[idx]


# ── Pre-computed NWS grid URLs for all MLB stadiums ─────────────────────────
# These are the NWS hourly forecast endpoints per stadium.
# Derived from: GET https://api.weather.gov/points/{lat},{lon}
# Avoids a round-trip lookup on every call. Stable indefinitely (grid rarely changes).
NWS_HOURLY_URLS = {
    "D-BACKS":   "https://api.weather.gov/gridpoints/PSR/162,56/forecast/hourly",
    "BRAVES":    "https://api.weather.gov/gridpoints/FFC/53,87/forecast/hourly",
    "ORIOLES":   "https://api.weather.gov/gridpoints/LWX/100,71/forecast/hourly",
    "RED SOX":   "https://api.weather.gov/gridpoints/BOX/68,49/forecast/hourly",
    "CUBS":      "https://api.weather.gov/gridpoints/LOT/74,73/forecast/hourly",
    "WHITE SOX": "https://api.weather.gov/gridpoints/LOT/71,66/forecast/hourly",
    "REDS":      "https://api.weather.gov/gridpoints/ILN/79,59/forecast/hourly",
    "GUARDIANS": "https://api.weather.gov/gridpoints/CLE/91,65/forecast/hourly",
    "ROCKIES":   "https://api.weather.gov/gridpoints/BOU/60,75/forecast/hourly",
    "TIGERS":    "https://api.weather.gov/gridpoints/DTX/69,65/forecast/hourly",
    "ROYALS":    "https://api.weather.gov/gridpoints/EAX/94,77/forecast/hourly",
    "ANGELS":    "https://api.weather.gov/gridpoints/SGX/55,43/forecast/hourly",
    "DODGERS":   "https://api.weather.gov/gridpoints/LOX/153,46/forecast/hourly",
    "TWINS":     "https://api.weather.gov/gridpoints/MPX/107,65/forecast/hourly",
    "METS":      "https://api.weather.gov/gridpoints/OKX/39,36/forecast/hourly",
    "YANKEES":   "https://api.weather.gov/gridpoints/OKX/34,43/forecast/hourly",
    "PHILLIES":  "https://api.weather.gov/gridpoints/PHI/53,64/forecast/hourly",
    "PIRATES":   "https://api.weather.gov/gridpoints/PBZ/76,57/forecast/hourly",
    "PADRES":    "https://api.weather.gov/gridpoints/SGX/54,14/forecast/hourly",
    "GIANTS":    "https://api.weather.gov/gridpoints/MTR/97,87/forecast/hourly",
    "MARINERS":  "https://api.weather.gov/gridpoints/SEW/124,68/forecast/hourly",
    "CARDINALS": "https://api.weather.gov/gridpoints/LSX/90,72/forecast/hourly",
    "NATIONALS": "https://api.weather.gov/gridpoints/LWX/97,67/forecast/hourly",
    "ATHLETICS": "https://api.weather.gov/gridpoints/STO/71,93/forecast/hourly",
}
# Runtime cache: team → hourly_url (populated lazily from /points lookup if not above)
_NWS_URL_CACHE: dict = {}


def _get_weather_open_meteo(home, game_time_utc=None):
    """
    Primary weather source: Open-Meteo (free, no API key, lat/lon based).
    Returns weather dict or None on failure.
    """
    coords = STADIUM_COORDS.get(home)
    if not coords:
        return None
    lat, lon = coords
    try:
        r = requests.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude":         lat,
                "longitude":        lon,
                "hourly":           "temperature_2m,windspeed_10m,winddirection_10m,relative_humidity_2m",
                "temperature_unit": "fahrenheit",
                "windspeed_unit":   "mph",
                "timezone":         "auto",
                "forecast_days":    2,
            },
            timeout=12,
        )
        r.raise_for_status()
        d     = r.json()
        times = d["hourly"]["time"]           # ["2026-04-16T00:00", ...]
        temps  = d["hourly"]["temperature_2m"]
        winds  = d["hourly"]["windspeed_10m"]
        wdirs  = d["hourly"]["winddirection_10m"]
        humids = d["hourly"].get("relative_humidity_2m", [])

        # Determine target local time for this game
        # game_time_utc is a string like "2026-04-16T23:05:00Z" (same format as rest of codebase)
        # Open-Meteo with timezone=auto returns hourly times in LOCAL timezone of the stadium.
        # We convert UTC game time to local time using STADIUM_TZ_OFFSET, then find the
        # closest hourly slot (rounding to nearest hour, not floor).
        best_idx = 0
        tz_off   = STADIUM_TZ_OFFSET.get(home, -5)
        _debug_target_str = "?"

        def _find_closest_hour(times, target_str):
            """Find index of hourly slot closest to target_str (YYYY-MM-DDTHH:MM)."""
            best_diff_h = 9999.0
            best_i      = 0
            try:
                g_dt = datetime.strptime(target_str, "%Y-%m-%dT%H:%M")
            except ValueError:
                return 0
            for i, t in enumerate(times):
                try:
                    t_dt = datetime.strptime(t, "%Y-%m-%dT%H:%M")
                    diff = abs((t_dt - g_dt).total_seconds() / 3600)
                    if diff < best_diff_h:
                        best_diff_h = diff
                        best_i = i
                except ValueError:
                    continue
            return best_i

        if game_time_utc:
            try:
                if isinstance(game_time_utc, str):
                    clean  = game_time_utc.replace("Z", "").split("+")[0]
                    utc_dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
                else:
                    utc_dt = game_time_utc          # already a datetime
                game_local = utc_dt + timedelta(hours=tz_off)
                # Round to nearest hour (not floor) so 6:40 PM → 7 PM slot not 6 PM
                _mins = game_local.minute
                if _mins >= 30:
                    game_local_rounded = game_local + timedelta(minutes=(60 - _mins))
                else:
                    game_local_rounded = game_local - timedelta(minutes=_mins)
                target_str = game_local_rounded.strftime("%Y-%m-%dT%H:00")
                _debug_target_str = f"{game_local.strftime('%I:%M %p')} local → slot {game_local_rounded.strftime('%I:%M %p')}"
                best_idx = _find_closest_hour(times, target_str)
            except Exception as e:
                print(f"  [weather/open-meteo] time parse error for {home}: {e}")
                best_idx = 0
        else:
            # No game time → use current UTC mapped to local
            now_local = datetime.utcnow() + timedelta(hours=tz_off)
            if now_local.minute >= 30:
                now_local_rounded = now_local + timedelta(minutes=(60 - now_local.minute))
            else:
                now_local_rounded = now_local - timedelta(minutes=now_local.minute)
            now_str   = now_local_rounded.strftime("%Y-%m-%dT%H:00")
            _debug_target_str = f"now → {now_local_rounded.strftime('%I:%M %p')} local"
            best_idx = _find_closest_hour(times, now_str)

        temp_val  = int(round(temps[best_idx]))
        mph_val   = int(round(winds[best_idx]))
        raw_deg   = round(float(wdirs[best_idx]), 1)
        raw_dir   = _deg_to_compass(raw_deg)
        wind_dir  = wind_direction_for_park(raw_dir, home)
        humid_val = int(round(humids[best_idx])) if humids and best_idx < len(humids) else 50
        return {
            "dir":        wind_dir,
            "mph":        mph_val,
            "temp":       temp_val,
            "raw_dir":    raw_dir,
            "raw_deg":    raw_deg,
            "humidity":   humid_val,
            "source":     "Open-Meteo",
            "slot":       _debug_target_str,
        }

    except requests.RequestException as e:
        print(f"  [weather/open-meteo] network error for {home}: {e}")
        return None
    except (KeyError, IndexError, ValueError, TypeError) as e:
        print(f"  [weather/open-meteo] parse error for {home}: {e}")
        return None


def _get_weather_wttr(home, game_time_utc=None):
    """
    Fallback weather source: wttr.in.
    Returns weather dict or None on failure.
    """
    city = STADIUMS.get(home, "")
    if not city:
        return None
    try:
        r = requests.get(
            f"https://wttr.in/{city.replace(' ','+')}?format=j1",
            timeout=10,
        )
        r.raise_for_status()
        d = r.json()
        raw_dir = None; mph = None; temp = None

        target_hour = _utc_to_local_hour(game_time_utc, home)
        if target_hour is not None:
            weather_days = d.get("weather", [])
            cur_local_h  = (datetime.utcnow().hour + STADIUM_TZ_OFFSET.get(home, -5)) % 24
            day_idx = 1 if (target_hour < 4 and cur_local_h >= 20) else 0
            if day_idx < len(weather_days):
                hourly = weather_days[day_idx].get("hourly", [])
                best, best_diff = None, 999
                for entry in hourly:
                    hh   = int(entry.get("time", "0")) // 100
                    diff = abs(hh - target_hour)
                    if diff > 12: diff = 24 - diff
                    if diff < best_diff:
                        best_diff = diff; best = entry
                if best:
                    raw_dir = best.get("winddir16Point", "")
                    mph     = int(best.get("windspeedMiles", 0))
                    temp    = int(best.get("tempF", best.get("temp_F", 70)))

        if raw_dir is None:
            cur     = d["current_condition"][0]
            raw_dir = cur.get("winddir16Point", "")
            mph     = int(cur.get("windspeedMiles", 0))
            temp    = int(cur.get("temp_F", 70))

        wind_dir = wind_direction_for_park(raw_dir, home)
        return {"dir": wind_dir, "mph": mph, "temp": temp, "raw_dir": raw_dir, "humidity": 50}

    except requests.RequestException as e:
        print(f"  [weather/wttr] network error for {home}: {e}")
        return None
    except Exception as e:
        print(f"  [weather/wttr] error for {home}: {e}")
        return None


def _get_weather_nws(home, game_time_utc=None):
    """
    Primary weather source: NWS (National Weather Service, weather.gov).
    No API key required. This is the official US government forecast model
    used by most US sports weather apps (Prop Finder, The Athletic, etc.).

    Two-step call:
      1) GET /points/{lat},{lon}    → resolves NWS grid office + x,y
      2) GET /gridpoints/.../forecast/hourly  → hourly periods

    Wind direction is returned as compass string (NNW, SSW, etc.)
    representing the direction the wind is coming FROM (met convention).
    Returns weather dict or None on failure.
    """
    coords = STADIUM_COORDS.get(home)
    if not coords:
        return None
    lat, lon = coords

    _HEADERS = {
        "User-Agent": "MLB-PicksModel/2.0 (contact: picks-model@local)",
        "Accept": "application/geo+json",
    }

    try:
        # Step 1: resolve NWS hourly URL — use pre-computed table first, then lazy lookup
        hourly_url = (NWS_HOURLY_URLS.get(home)
                      or _NWS_URL_CACHE.get(home))
        if not hourly_url:
            pts_r = requests.get(
                f"https://api.weather.gov/points/{lat},{lon}",
                headers=_HEADERS, timeout=10,
            )
            pts_r.raise_for_status()
            pts = pts_r.json()["properties"]
            hourly_url = pts["forecastHourly"]
            _NWS_URL_CACHE[home] = hourly_url   # cache for this session

        # Step 2: get hourly forecast
        fc_r = requests.get(hourly_url, headers=_HEADERS, timeout=12)

        # If hardcoded URL is stale (404) or server error (500), re-derive via /points/
        if fc_r.status_code in (404, 500) and hourly_url != _NWS_URL_CACHE.get(home):
            print(f"[weather/nws] {fc_r.status_code} on cached URL for {home} → re-deriving via /points/ …")
            pts_r = requests.get(
                f"https://api.weather.gov/points/{lat},{lon}",
                headers=_HEADERS, timeout=10,
            )
            pts_r.raise_for_status()
            hourly_url = pts_r.json()["properties"]["forecastHourly"]
            _NWS_URL_CACHE[home] = hourly_url   # update session cache
            fc_r = requests.get(hourly_url, headers=_HEADERS, timeout=12)

        fc_r.raise_for_status()
        periods = fc_r.json()["properties"]["periods"]
        # Each period: {startTime, endTime, temperature, temperatureUnit,
        #               windSpeed ("13 mph"), windDirection ("SSW"), ...}

        if not periods:
            return None

        # Determine target local time
        tz_off = STADIUM_TZ_OFFSET.get(home, -5)
        _debug_slot = "?"

        if game_time_utc:
            try:
                clean  = game_time_utc.replace("Z","").split("+")[0]
                utc_dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
                game_local = utc_dt + timedelta(hours=tz_off)
                _debug_slot = game_local.strftime("%I:%M %p local")
            except Exception:
                game_local = None
        else:
            game_local = datetime.utcnow() + timedelta(hours=tz_off)
            _debug_slot = f"now {game_local.strftime('%I:%M %p')} local"

        # Find best period: NWS startTime has tz offset, e.g. "2026-04-22T18:00:00-04:00"
        best_period = None
        best_diff   = 9999.0
        for p in periods:
            try:
                st_raw = p["startTime"]  # "2026-04-22T18:00:00-04:00"
                # Parse with offset → naive UTC for comparison
                # Remove colon in tz offset for strptime compat
                st_clean = st_raw[:19]  # "2026-04-22T18:00:00"
                tz_part  = st_raw[19:]  # "-04:00" or "Z"
                if tz_part in ("Z", "+00:00"):
                    tz_h = 0
                elif len(tz_part) >= 6:
                    sign = 1 if tz_part[0] == "+" else -1
                    tz_h = sign * (int(tz_part[1:3]) + int(tz_part[4:6]) / 60)
                else:
                    tz_h = 0
                period_utc = datetime.strptime(st_clean, "%Y-%m-%dT%H:%M:%S") - timedelta(hours=tz_h)
                if game_time_utc:
                    game_utc_dt = utc_dt
                else:
                    game_utc_dt = game_local - timedelta(hours=tz_off) if game_local else datetime.utcnow()
                diff = abs((period_utc - game_utc_dt).total_seconds() / 3600)
                if diff < best_diff:
                    best_diff   = diff
                    best_period = p
            except Exception:
                continue

        if not best_period:
            best_period = periods[0]

        # Parse temperature (always Fahrenheit if ?units=us; default is Fahrenheit for US)
        temp_val  = int(round(float(best_period.get("temperature", 70))))
        temp_unit = best_period.get("temperatureUnit", "F")
        if temp_unit == "C":
            temp_val = int(round(temp_val * 9/5 + 32))

        # Parse wind speed: "13 mph" or "Calm" or "5 to 10 mph" → take average of range
        ws_raw = best_period.get("windSpeed", "0 mph")
        import re as _re
        ws_nums = _re.findall(r"\d+", str(ws_raw))
        if ws_nums:
            mph_val = int(round(sum(int(x) for x in ws_nums) / len(ws_nums)))
        else:
            mph_val = 0

        # Wind direction: compass string, already "FROM" convention (e.g., "SSW")
        raw_dir  = str(best_period.get("windDirection", "")).upper().strip()
        wind_dir = wind_direction_for_park(raw_dir, home)

        return {
            "dir":      wind_dir,
            "mph":      mph_val,
            "temp":     temp_val,
            "raw_dir":  raw_dir,
            "raw_deg":  "",           # NWS gives compass string, not degrees
            "humidity": 50,           # NWS hourly doesn't always have humidity
            "source":   "NWS",
            "slot":     _debug_slot,
        }

    except requests.RequestException as e:
        print(f"  [weather/nws] network error for {home}: {e}")
        return None
    except Exception as e:
        print(f"  [weather/nws] parse error for {home}: {e}")
        return None


# ──────────────────────────────────────────────────────
# WEATHER OVERRIDES — manual input from Prop Finder
# ──────────────────────────────────────────────────────

def _load_weather_overrides():
    """Load manual weather overrides for today. Returns {key: weather_dict}."""
    if not os.path.exists(WEATHER_OVERRIDES_FILE):
        return {}
    try:
        with open(WEATHER_OVERRIDES_FILE, "r") as f:
            all_overrides = json.load(f)
        return all_overrides.get(TARGET_DATE, {})
    except Exception:
        return {}

def _save_weather_override(away, home, temp, wind_dir, mph):
    """Save a manual weather override for a game."""
    # Load all existing overrides
    all_overrides = {}
    if os.path.exists(WEATHER_OVERRIDES_FILE):
        try:
            with open(WEATHER_OVERRIDES_FILE, "r") as f:
                all_overrides = json.load(f)
        except Exception:
            pass
    if TARGET_DATE not in all_overrides:
        all_overrides[TARGET_DATE] = {}

    key = f"{away}@{home}"
    all_overrides[TARGET_DATE][key] = {
        "dir":     wind_dir.upper(),
        "mph":     int(mph),
        "temp":    int(temp),
        "raw_dir": wind_dir.upper(),
        "raw_deg": "",
        "humidity": 50,
        "source":  "MANUAL (Prop Finder)",
        "slot":    "manual override",
    }
    with open(WEATHER_OVERRIDES_FILE, "w") as f:
        json.dump(all_overrides, f, indent=2)
    return all_overrides[TARGET_DATE][key]

def _get_weather_override(home, away=None):
    """Check if a manual override exists for this game. Returns dict or None."""
    overrides = _load_weather_overrides()
    if not overrides:
        return None
    # Try specific away@home key first, then home-only fallback
    if away:
        key = f"{away}@{home}"
        if key in overrides:
            return overrides[key]
    # Fallback: search by home team
    for k, v in overrides.items():
        if k.endswith(f"@{home}"):
            return v
    return None

def cmd_set_weather():
    """
    --set-weather AWAY HOME TEMP IN|OUT|CROSS MPH

    Registra el clima manualmente desde Prop Finder para que el modelo
    use esos valores en vez de la API. El override persiste todo el día.

    Ejemplos:
      python3 mlb.py --set-weather STL PIT 70 IN 13
      python3 mlb.py --set-weather SEA MIN 53 OUT 17
      python3 mlb.py --set-weather NYY TEX 85 DOME 0
    """
    try:
        idx = sys.argv.index("--set-weather")
        away     = sys.argv[idx+1].upper()
        home     = sys.argv[idx+2].upper()
        temp_s   = sys.argv[idx+3]
        dir_s    = sys.argv[idx+4].upper()
        mph_s    = sys.argv[idx+5]
    except (ValueError, IndexError):
        print("\n  Uso: python3 mlb.py --set-weather AWAY HOME TEMP DIR MPH")
        print("  Ejemplo: python3 mlb.py --set-weather STL PIT 70 IN 13")
        print("  DIR puede ser: IN | OUT | CROSS | DOME")
        return

    # Validate dir
    valid_dirs = {"IN","OUT","CROSS","DOME","L-R","R-L"}
    if dir_s not in valid_dirs:
        print(f"  ❌ Dirección inválida: '{dir_s}'. Usa: IN, OUT, CROSS, DOME")
        return

    try:
        temp = int(float(temp_s))
        mph  = int(float(mph_s))
    except ValueError:
        print(f"  ❌ Temperatura o velocidad inválida: {temp_s}, {mph_s}")
        return

    w = _save_weather_override(away, home, temp, dir_s, mph)
    print(f"\n  ✅ Clima guardado para {away} @ {home}  ({TARGET_DATE})")
    print(f"     🌡️  {w['temp']}°F   💨 {w['dir']} {w['mph']}mph   📍 {w['source']}")
    print(f"\n  Ahora corre --picks o --lines para usar estos valores.")

def cmd_clear_weather():
    """--clear-weather: borra todos los overrides manuales de hoy."""
    all_overrides = {}
    if os.path.exists(WEATHER_OVERRIDES_FILE):
        try:
            with open(WEATHER_OVERRIDES_FILE, "r") as f:
                all_overrides = json.load(f)
        except Exception:
            pass
    count = len(all_overrides.get(TARGET_DATE, {}))
    if TARGET_DATE in all_overrides:
        del all_overrides[TARGET_DATE]
        with open(WEATHER_OVERRIDES_FILE, "w") as f:
            json.dump(all_overrides, f, indent=2)
    print(f"\n  🗑️  {count} override(s) de clima borrados para {TARGET_DATE}")

def cmd_list_weather_overrides():
    """Show all active weather overrides for today."""
    overrides = _load_weather_overrides()
    if not overrides:
        print(f"\n  📋 No hay overrides de clima para {TARGET_DATE}")
        return
    print(f"\n  📋 Weather overrides activos — {TARGET_DATE}:")
    print(f"  {'─'*50}")
    for key, w in overrides.items():
        away, home = key.split("@") if "@" in key else ("?", key)
        print(f"  {away} @ {home:<12}  {w['temp']}°F  {w['dir']} {w['mph']}mph  [{w['source']}]")
    print(f"  {'─'*50}")
    print(f"  Para borrar: python3 mlb.py --clear-weather")


def get_weather(home, game_time_utc=None, away=None):
    """
    Get stadium weather for run total model.
    Priority: Manual override (Prop Finder) → NWS (US gov) → Open-Meteo → wttr.in
    Domes always return DOME with no wind/temp adjustments.
    """
    if home in DOME_TEAMS:
        return {"dir": "DOME", "mph": 0, "temp": None, "raw_dir": "DOME",
                "raw_deg": "", "humidity": 50, "source": "DOME", "slot": ""}

    # ── Priority 0: Manual override from Prop Finder ──────────────────────────
    override = _get_weather_override(home, away)
    if override:
        return override

    # ── Primary: NWS (National Weather Service — same as Prop Finder) ────────
    result = _get_weather_nws(home, game_time_utc)
    if result:
        return result

    print(f"  [weather] NWS failed for {home} → trying Open-Meteo fallback …")

    # ── Secondary: Open-Meteo (ECMWF model) ──────────────────────────────────
    result = _get_weather_open_meteo(home, game_time_utc)
    if result:
        return result

    print(f"  [weather] Open-Meteo failed for {home} → trying wttr.in fallback …")

    # ── Tertiary: wttr.in ─────────────────────────────────────────────────────
    result = _get_weather_wttr(home, game_time_utc)
    if result:
        return result

    print(f"  [weather] ⚠️  All sources failed for {home}. Weather unavailable.")
    return {"dir": None, "mph": None, "temp": None, "raw_dir": None,
            "raw_deg": "", "humidity": 50, "source": "N/A", "slot": ""}


# ──────────────────────────────────────────────────────
# MARKET ODDS — The Odds API
# ──────────────────────────────────────────────────────

def get_market_odds():
    if not ODDS_API_KEY: return {}
    try:
        books = ",".join(BOOK_IDS.values())
        r = requests.get(
            "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds/",
            params={"apiKey":ODDS_API_KEY,"regions":"us,eu",
                    "markets":"h2h,totals,spreads",
                    "bookmakers":books,"oddsFormat":"american"},
            timeout=15)
        r.raise_for_status()
        odds_by_game = {}
        skipped = 0
        for game in r.json():
            # ── Filtrar por fecha: solo juegos de TARGET_DATE ─────────────────
            # commence_time viene como "2026-04-17T22:40:00Z" (UTC)
            # Comparar la fecha UTC con TARGET_DATE; para juegos nocturnos
            # también aceptar TARGET_DATE+1 a medianoche UTC (hora PR = UTC-4)
            ct = game.get("commence_time", "")
            if ct:
                ct_date = ct[:10]   # "2026-04-17"
                # Aceptar TARGET_DATE; también la siguiente fecha UTC si es un
                # juego de noche que empieza pasada medianoche UTC (PR = UTC-4)
                if ct_date not in (TARGET_DATE,):
                    # Si el juego es del día siguiente en UTC pero del mismo día
                    # en hora local PR (UTC-4), aceptarlo si commence_time < 04:00 UTC
                    ct_hour = int(ct[11:13]) if len(ct) >= 13 else 12
                    from datetime import datetime as _dt, timedelta as _td
                    day_after = (_dt.strptime(TARGET_DATE, "%Y-%m-%d") + _td(days=1)).strftime("%Y-%m-%d")
                    if ct_date == day_after and ct_hour < 4:
                        pass   # juego nocturno PR — aceptar
                    else:
                        skipped += 1
                        continue  # juego de otro día — ignorar

            away = TEAM_MAP.get(game["away_team"], game["away_team"].upper())
            home = TEAM_MAP.get(game["home_team"], game["home_team"].upper())
            key  = f"{away} vs {home}"
            odds_by_game[key] = {"books":{}}
            for bm in game.get("bookmakers",[]):
                bname = next((k for k,v in BOOK_IDS.items() if v==bm["key"]), bm["key"])
                odds_by_game[key]["books"][bname] = {}
                for mkt in bm.get("markets",[]):
                    if mkt["key"] == "h2h":
                        for o in mkt["outcomes"]:
                            t = TEAM_MAP.get(o["name"], o["name"].upper())
                            odds_by_game[key]["books"][bname][f"ML_{t}"] = o["price"]
                    elif mkt["key"] == "totals":
                        for o in mkt["outcomes"]:
                            odds_by_game[key]["books"][bname][f"Total_{o['name']}"] = {
                                "line": o.get("point"), "odds": o["price"]}
                    elif mkt["key"] == "spreads":
                        for o in mkt["outcomes"]:
                            t = TEAM_MAP.get(o["name"], o["name"].upper())
                            odds_by_game[key]["books"][bname][f"Spread_{t}"] = {
                                "line": o.get("point"), "odds": o["price"]}
                    elif mkt["key"] == "alternate_totals":
                        # Agrupar por línea: Over y Under pueden venir en outcomes separados
                        _alt_by_pt = {}
                        for o in mkt["outcomes"]:
                            _pt = o.get("point")
                            if _pt is None: continue
                            _pt = float(_pt)
                            if _pt not in _alt_by_pt:
                                _alt_by_pt[_pt] = {}
                            _alt_by_pt[_pt][o["name"]] = o["price"]  # "Over" o "Under"
                        _alt_list = []
                        for _pt in sorted(_alt_by_pt.keys()):
                            _entry = {"line": _pt}
                            if "Over"  in _alt_by_pt[_pt]: _entry["over_odds"]  = _alt_by_pt[_pt]["Over"]
                            if "Under" in _alt_by_pt[_pt]: _entry["under_odds"] = _alt_by_pt[_pt]["Under"]
                            _alt_list.append(_entry)
                        if _alt_list:
                            odds_by_game[key]["books"][bname]["Alt_Totals"] = _alt_list
                    elif mkt["key"] == "alternate_spreads":
                        _alt_sp = []
                        for o in mkt["outcomes"]:
                            t = TEAM_MAP.get(o["name"], o["name"].upper())
                            _pt = o.get("point")
                            if _pt is None: continue
                            _alt_sp.append({"team": t, "line": float(_pt), "odds": o["price"]})
                        if _alt_sp:
                            odds_by_game[key]["books"][bname]["Alt_Spreads"] = _alt_sp
        if skipped:
            print(f"  ℹ️  {skipped} juegos de otras fechas ignorados (filtro por {TARGET_DATE})")
        print(f"  ✅ Odds jaladas para {len(odds_by_game)} juegos de {TARGET_DATE}")
        return odds_by_game
    except Exception as e:
        print(f"  ❌ Odds API error: {e}")
        # Si es un error HTTP, mostrar el cuerpo de la respuesta para diagnóstico
        try:
            if hasattr(e, 'response') and e.response is not None:
                print(f"  ❌ HTTP {e.response.status_code}: {e.response.text[:300]}")
        except Exception:
            pass
        return {}


def _get_game_books(odds, away, home):
    """Busca odds del juego — prueba ambos ordenes de equipos."""
    d = odds.get(f"{away} vs {home}", odds.get(f"{home} vs {away}", {}))
    return d.get("books", {})


def _fmt_odds(o):
    try:
        v = int(o)
        return f"+{v}" if v > 0 else str(v)
    except: return str(o)


# ──────────────────────────────────────────────────────
# CALCULAR LÍNEAS DEL MODELO
# ──────────────────────────────────────────────────────

def _season_games_estimate(date_str=None):
    """
    Estima cuántos juegos ha jugado cada equipo en la temporada actual
    basándose en el calendario (no requiere llamada a API).
    MLB 2026: apertura estimada ~26 marzo 2026.
    Retorna int con el estimado de juegos jugados por equipo.
    """
    from datetime import datetime as _dt
    OPENING_DAY = _dt(int(SEASON), 3, 26)   # ajusta si opening day cambia
    if date_str:
        today = _dt.strptime(date_str[:10], "%Y-%m-%d")
    else:
        today = _dt.now()
    days_in = max(0, (today - OPENING_DAY).days)
    # ~0.87 juegos/día por equipo en la primera mitad de temporada
    return min(162, int(days_in * 0.87))


def _adaptive_wrc_blend(g_cur):
    """
    Peso adaptativo para wRC+ de temporada actual.
    - Abril (~15 juegos): ~10% actual, 90% año previo
    - Mayo  (~40 juegos): ~25% actual
    - Junio (~70 juegos): ~45% actual
    - Julio (~100 juegos): ~60% actual
    - Septiembre (162 juegos): ~75% actual
    """
    return round(min(0.75, (g_cur / 162) * 0.75 + 0.02), 3)


def _adaptive_sp_blend(gs_cur):
    """
    Peso adaptativo para SP xFIP de temporada actual.
    Un SP sale cada ~5 días → GS ≈ g_cur / 5.
    - 3 GS (abril): ~10% actual
    - 10 GS (mayo): ~35% actual
    - 20 GS (julio): ~65% actual
    - 30+ GS: ~80% actual
    """
    return round(min(0.80, (gs_cur / 32) * 0.80 + 0.02), 3)


def _tto_adj_xfip(sp_xfip, expected_ip=5.0):
    """
    Ajusta el SP xFIP por el 'Times-Through-Order' (TTO) penalty.

    Investigación muestra que los pitchers son progresivamente peores
    cada vez que enfrentan la alineación completa:
      1ª vuelta (innings 1-3):  baseline
      2ª vuelta (innings 4-6):  +8% efectividad
      3ª vuelta (innings 7+):   +20% efectividad

    Con expected_ip=5.0 (promedio MLB actual):
      Peso TTO1 = 3/5 = 0.60, TTO2 = 2/5 = 0.40
      Factor = 0.60*1.0 + 0.40*1.08 = 1.032
      → un SP con xFIP 3.50 efectivamente rinde como 3.61
    """
    ip = max(1.0, min(9.0, expected_ip))
    if ip <= 3.0:
        return sp_xfip                                            # solo 1ª vuelta
    elif ip <= 6.0:
        w1 = 3.0 / ip; w2 = (ip - 3.0) / ip
        return sp_xfip * (w1 * 1.00 + w2 * 1.11)                # mezcla TTO1+TTO2 (era 1.08)
    else:
        w1 = 3.0 / ip; w2 = 3.0 / ip; w3 = (ip - 6.0) / ip
        return sp_xfip * (w1 * 1.00 + w2 * 1.11 + w3 * 1.23)   # mezcla TTO1+2+3 (era 1.20)


def _rtm_xfip(xfip, lg_avg=4.20, rtm_weight=0.18):
    """
    Regression-to-mean: jala xFIP extremos hacia el promedio de liga (4.20).
    Especialmente importante en abril/inicio de temporada con sample pequeño.
    rtm_weight=0.18 → 18% de la diferencia se regresa a la media.
    Ejemplo: xFIP 3.10 → 3.10 * 0.82 + 4.20 * 0.18 = 3.30 (más conservador)
             xFIP 5.20 → 5.20 * 0.82 + 4.20 * 0.18 = 5.02 (menos extremo)
    """
    return round(xfip * (1.0 - rtm_weight) + lg_avg * rtm_weight, 3)


def calc_xfip_tot(fip, bp_xfip, expected_sp_ip=5.0):
    """
    Calcula xFIP efectivo del juego combinando SP (ajustado por TTO + RTM) y BP.
    Split dinámico basado en innings proyectados del SP:
      SP  = expected_sp_ip / 9
      BP  = (9 - expected_sp_ip) / 9
    Default 5.0 IP → SP 55.6% / BP 44.4%  (antes fijo 60/40)
    RTM aplicado al SP para reducir sobreconfianza en SPs con sample pequeño.
    """
    sp_w    = min(0.70, max(0.40, expected_sp_ip / 9.0))
    bp_w    = 1.0 - sp_w
    tto_fip = _tto_adj_xfip(fip, expected_sp_ip)
    rtm_fip = _rtm_xfip(tto_fip)          # regresa valores extremos a la media
    raw     = rtm_fip * sp_w + bp_xfip * bp_w
    if raw < 2.5: return 2.8
    if raw > 5.5: return 5.2
    return round(raw, 3)


def calc_pf_combined(home, wind_dir, wind_mph, temp, humidity=50):
    pf = PARK_FACTORS.get(home, 1.0)
    # Temperatura: +3.5% por cada 10°F sobre 70°F (aire menos denso = más carreras)
    temp_adj = 1.0
    if temp and home not in DOME_TEAMS:
        temp_adj = 1 + ((temp - 70) / 10 * 0.035)
    # Viento: efecto directo sobre batazos
    wind_adj = 1.0
    if wind_dir and home not in DOME_TEAMS and wind_mph:
        if   wind_dir == "OUT":                 wind_adj = 1 + (wind_mph * 0.010)
        elif wind_dir == "IN":                  wind_adj = 1 - (wind_mph * 0.010)
        elif wind_dir in ("L-R","R-L","CROSS"): wind_adj = 1 + (wind_mph * 0.005)
    # Humedad (ADI): aire húmedo = menos denso = +0.5% por cada 10% sobre 50% HR
    # Efecto real pero sutil (inferior al de temperatura)
    humid_adj = 1.0
    if humidity and home not in DOME_TEAMS:
        humid_adj = 1.0 + ((humidity - 50) / 100 * 0.005)   # ±0.25% max aprox
    return round(pf * temp_adj * wind_adj * humid_adj, 4)


def _daygame_factor(game_hour_local):
    """
    Ajuste por juego de día vs noche.

    Juegos de día (hora local < 17:00) promedian ~0.2-0.3 menos carreras que noche:
    - "Getaway days": managers descansan titulares, bullpen usa pitchers B
    - Calor de mediodía reduce rendimiento físico en innings tardíos
    - Estadios con sol en el cuadro (errores de fielding que cancelan carreras)

    No se aplica a domes (ya filtrados por temp_adj=None).
    Juegos nocturnos (≥17:00) → factor 1.0 (sin ajuste).
    Juegos de día (10-16:59 local) → factor 0.978 (~2.2% menos carreras).
    """
    if game_hour_local is None:
        return 1.0
    h = int(game_hour_local)
    if 10 <= h < 17:   # día: 10am–4:59pm
        return 0.978
    return 1.0         # noche (≥5pm) o madrugada


# ──────────────────────────────────────────────────────
# RECENT FORM MODULE — últimos 10 juegos + días de descanso
# ──────────────────────────────────────────────────────

_RECENT_FORM_CACHE      = {}
_RECENT_FORM_CACHE_TS   = 0.0
_STANDINGS_CACHE        = {}
_STANDINGS_CACHE_TS     = 0.0
_BP_FATIGUE_CACHE       = {}
_BP_FATIGUE_CACHE_TS    = 0.0
_SP_RECENT_CACHE        = {}   # pitcher_id → recent ERA (last 5 starts)
_SP_GS_CACHE            = {}   # pitcher_id → total GS this season (debut detection)
_SP_SEASON_FIP_CACHE    = {}   # pitcher_id → {fip, era, gs, ip, k, bb, hr} por MLB ID
_LINEUP_CACHE           = {}   # game_pk → {"away": [player_id,...], "home": [...]}
_LINEUP_NAMES_CACHE     = {}   # game_pk → {"away": ["Full Name",...], "home": [...]}
_PLAYER_STAT_CACHE      = {}   # player_id → {ops, pa} for wRC+ proxy

# ── Disk cache diario para player OPS y SP FIP ───────────────────────────
# Reduce llamadas API en corridas repetidas del mismo día a 0.
_DISK_CACHE_FILE = os.path.join(SCRIPT_DIR, ".mlb_daily_cache.json")
_DISK_CACHE_DATE = ""   # fecha del cache cargado

def _load_disk_cache():
    """Carga cache de disco si es del día de hoy y rellena los dicts en memoria."""
    global _DISK_CACHE_DATE
    try:
        if not os.path.exists(_DISK_CACHE_FILE):
            return
        with open(_DISK_CACHE_FILE, "r", encoding="utf-8") as _f:
            _dc = json.load(_f)
        if _dc.get("date") != TARGET_DATE:
            return  # caché de otro día — ignorar
        _DISK_CACHE_DATE = TARGET_DATE
        # Restaurar player OPS (keys son "pid|season")
        for _k, _v in _dc.get("player_ops", {}).items():
            _parts = _k.split("|")
            if len(_parts) == 2:
                _PLAYER_STAT_CACHE[(int(_parts[0]), _parts[1])] = _v
        # Restaurar SP FIP (keys son str(pitcher_id))
        for _k, _v in _dc.get("sp_fip", {}).items():
            _SP_SEASON_FIP_CACHE[int(_k)] = _v
    except Exception:
        pass  # cache corrupto → ignorar, se regenera

def _save_disk_cache():
    """Persiste player OPS y SP FIP a disco para el día actual."""
    try:
        _player_ops_serial = {f"{pid}|{season}": v
                              for (pid, season), v in _PLAYER_STAT_CACHE.items()}
        _sp_fip_serial     = {str(k): v for k, v in _SP_SEASON_FIP_CACHE.items()}
        with open(_DISK_CACHE_FILE, "w", encoding="utf-8") as _f:
            json.dump({"date": TARGET_DATE,
                       "player_ops": _player_ops_serial,
                       "sp_fip": _sp_fip_serial}, _f, ensure_ascii=False)
    except Exception:
        pass

# Cargar cache de disco al importar el módulo
_load_disk_cache()
_ROLLING_WRC_CACHE      = {}
_ROLLING_WRC_CACHE_TS   = 0.0
_HOME_AWAY_CACHE        = {}   # team → {"home": wrc_proxy, "away": wrc_proxy}
_HOME_AWAY_CACHE_TS     = 0.0

def _get_mlb_recent_form(target_date=None, window=10, cache_hours=4):
    """
    Fetches the last `window` completed regular-season games for every team
    and computes:
      - wp        : win% over those games  (0.0 – 1.0)
      - run_diff  : avg run differential   (positive = outscoring opponents)
      - games     : actual games found     (may be < window early in season)
      - rest_days : calendar days since last completed game (0 = B2B)

    Results are cached for `cache_hours` hours to avoid repeated API calls.
    Uses MLB Stats API /v1/schedule with linescore hydration.
    """
    global _RECENT_FORM_CACHE, _RECENT_FORM_CACHE_TS
    import time as _time
    now_ts = _time.time()
    if _RECENT_FORM_CACHE and (now_ts - _RECENT_FORM_CACHE_TS) < cache_hours * 3600:
        return _RECENT_FORM_CACHE

    if target_date is None:
        target_date = TARGET_DATE

    try:
        from collections import defaultdict
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").date()
        start_dt  = target_dt - timedelta(days=21)   # 3 semanas atrás
        end_dt    = target_dt - timedelta(days=1)

        url = (f"https://statsapi.mlb.com/api/v1/schedule?"
               f"sportId=1&startDate={start_dt}&endDate={end_dt}"
               f"&hydrate=linescore,teams&gameType=R")
        resp = requests.get(url, timeout=20)
        resp.raise_for_status()

        # Acumular juegos por equipo, más recientes primero
        team_games = defaultdict(list)
        for date_entry in sorted(resp.json().get("dates", []),
                                  key=lambda d: d["date"], reverse=True):
            gdate = date_entry["date"]
            for game in date_entry.get("games", []):
                if game.get("status", {}).get("abstractGameState") != "Final":
                    continue
                ls  = game.get("linescore", {}).get("teams", {})
                rs  = ls.get("away", {}).get("runs")
                rh  = ls.get("home", {}).get("runs")
                if rs is None or rh is None:
                    continue
                away_name = game["teams"]["away"]["team"]["name"]
                home_name = game["teams"]["home"]["team"]["name"]
                aabb = TEAM_MAP.get(away_name, away_name.upper()[:3])
                habb = TEAM_MAP.get(home_name, home_name.upper()[:3])
                aw = 1 if rs > rh else 0
                team_games[aabb].append({"date": gdate, "win": aw,     "diff": rs - rh})
                team_games[habb].append({"date": gdate, "win": 1 - aw, "diff": rh - rs})

        result = {}
        for abb, games in team_games.items():
            recent = games[:window]       # ya ordenado más reciente → más antiguo
            n = len(recent)
            if n < 1:
                continue
            wins    = sum(g["win"]  for g in recent)
            avg_diff = sum(g["diff"] for g in recent) / n
            last_dt  = datetime.strptime(recent[0]["date"], "%Y-%m-%d").date()
            rest     = max(0, (target_dt - last_dt).days)
            # ── Racha activa (streak) ──────────────────────────────────────
            # games[0] = más reciente. Contar cuántos consecutivos W o L
            streak_val = 0
            if recent:
                streak_dir = recent[0]["win"]   # 1=W, 0=L
                for g in recent:
                    if g["win"] == streak_dir:
                        streak_val += 1
                    else:
                        break
                if streak_dir == 0:
                    streak_val = -streak_val   # negativo = racha de derrotas
            result[abb] = {
                "wp":        round(wins / n, 3),
                "run_diff":  round(avg_diff, 2),
                "games":     n,
                "rest_days": rest,
                "streak":    streak_val,   # +N = N wins seguidas, -N = N losses
            }

        _RECENT_FORM_CACHE    = result
        _RECENT_FORM_CACHE_TS = now_ts
        return result

    except Exception as e:
        print(f"  ⚠️  Recent form fetch error: {e}")
        return {}


def _form_factor(team_abb, recent_form):
    """
    Convierte el récord reciente en un multiplicador para el run expectation.

    Win% últimos 10:  cada 10pp de desviación desde .500 = ±2.0%  (cap ±8%)
    Run differential: cada 1.0 carrera avg diff           = ±1.5%  (cap ±4%)
    Streak momentum:  rachas ≥4 aplican ±1.5% adicional  (cap ±1.5%)
    Total cap: ±10%

    Un equipo 7-3 (70 WP) con +1.5 avg run diff → factor ≈ 1.066
    Un equipo 3-7 (30 WP) con -1.0 avg run diff →  factor ≈ 0.940
    Equipo en racha de 5 pérdidas seguidas → -1.5% extra sobre lo anterior
    """
    form = recent_form.get(team_abb)
    if not form or form["games"] < 5:
        return 1.0   # muestra insuficiente → no ajustar
    wp_adj = max(-0.08, min(0.08, (form["wp"] - 0.500) * 0.20))
    rd_adj = max(-0.04, min(0.04,  form["run_diff"]    * 0.015))
    # Streak momentum: rachas significativas (≥4) tienen señal real
    # +0.015 si ganan 4+ seguidas, -0.015 si pierden 4+ seguidas
    streak = form.get("streak", 0)
    if   streak >=  4: str_adj = +0.015
    elif streak <= -4: str_adj = -0.015
    else:              str_adj = 0.0
    return round(max(0.90, min(1.10, 1.0 + wp_adj + rd_adj + str_adj)), 4)


def _rest_factor(team_abb, recent_form):
    """
    Ajuste por días de descanso antes del juego de hoy.
      0 días (B2B)  → -2.5%  (fatiga acumulada)
      1–2 días      →  0%    (ritmo normal)
      3+ días       → +1.0%  (descansados)
    """
    form = recent_form.get(team_abb)
    if not form:
        return 1.0
    rest = form["rest_days"]
    if   rest == 0: return 0.975
    elif rest >= 3: return 1.010
    else:           return 1.000


# ──────────────────────────────────────────────────────
# STANDINGS — win% de temporada
# ──────────────────────────────────────────────────────

def _get_mlb_standings(cache_hours=6):
    """Retorna {team: win_pct} con el récord de temporada actual desde MLB API."""
    global _STANDINGS_CACHE, _STANDINGS_CACHE_TS
    import time as _time
    now_ts = _time.time()
    if _STANDINGS_CACHE and (now_ts - _STANDINGS_CACHE_TS) < cache_hours * 3600:
        return _STANDINGS_CACHE
    try:
        season = TARGET_DATE[:4]
        url = (f"https://statsapi.mlb.com/api/v1/standings?"
               f"leagueId=103,104&season={season}&standingsTypes=regularSeason")
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        result = {}
        for record in r.json().get("records", []):
            for entry in record.get("teamRecords", []):
                name = entry["team"]["name"]
                team = TEAM_MAP.get(name, name.upper())
                wins   = entry.get("wins", 0)
                losses = entry.get("losses", 0)
                total  = wins + losses
                if total > 0:
                    result[team] = round(wins / total, 3)
        _STANDINGS_CACHE    = result
        _STANDINGS_CACHE_TS = now_ts
        if result:
            print(f"  📊 Standings: {len(result)} equipos cargados")
        return result
    except Exception as e:
        print(f"  ⚠️  Standings fetch error: {e}")
        return {}

def _standings_factor(team, standings):
    """
    Convierte win% de temporada en factor multiplicador.
    .600 → 1.04  |  .500 → 1.00  |  .400 → 0.96  (cap ±6%)
    Se combina con form_factor (L10) para señal más robusta.
    """
    wp  = standings.get(team, 0.500)
    adj = max(-0.06, min(0.06, (wp - 0.500) * 0.20))
    return round(1.0 + adj, 4)


# ──────────────────────────────────────────────────────
# SP RECENT FORM — últimas 5 salidas
# ──────────────────────────────────────────────────────

def _get_sp_recent_era(pitcher_id, n_starts=5):
    """
    Fetches ERA de las últimas N salidas del SP desde MLB Stats API.
    Solo salidas de ≥2.0 IP (filtra aperturas falsas / emergencias).

    Retorna dict {"era": float, "ip": float, "n": int} o None si no hay muestra suficiente.

    Requisitos mínimos (sample integrity):
      - Al menos 3 aperturas con ≥2.0 IP cada una.
      - Al menos 12.0 IP acumulados.
    Por qué: 0.75 ERA en 1 salida (6 IP) es ruido estadístico;
             0.75 ERA en 3+ salidas (15 IP) es señal real de forma.
    """
    if pitcher_id is None:
        return None
    if pitcher_id in _SP_RECENT_CACHE:
        return _SP_RECENT_CACHE[pitcher_id]
    try:
        season = TARGET_DATE[:4]
        url = (f"https://statsapi.mlb.com/api/v1/people/{pitcher_id}/stats?"
               f"stats=gameLog&season={season}&gameType=R&group=pitching")
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        splits = r.json().get("stats", [{}])[0].get("splits", [])
        # Track total GS for debut detection (any start, regardless of IP)
        all_starts = [s for s in splits
                      if int(s.get("stat", {}).get("gamesStarted", 0) or 0) >= 1]
        _SP_GS_CACHE[pitcher_id] = len(all_starts)
        starts = [s for s in splits
                  if float(s.get("stat", {}).get("inningsPitched", "0") or "0") >= 2.0]
        recent = starts[-n_starts:]
        # Mínimo: 3 salidas con ≥2 IP cada una
        if len(recent) < 3:
            _SP_RECENT_CACHE[pitcher_id] = None
            return None
        total_er = sum(int(s.get("stat", {}).get("earnedRuns", 0) or 0) for s in recent)
        total_ip = sum(float(s.get("stat", {}).get("inningsPitched", 0) or 0) for s in recent)
        # Mínimo: 12.0 IP acumulados (≈3 salidas normales)
        if total_ip < 12.0:
            _SP_RECENT_CACHE[pitcher_id] = None
            return None
        era = round((total_er / total_ip) * 9, 2)
        result = {"era": era, "ip": round(total_ip, 1), "n": len(recent)}
        _SP_RECENT_CACHE[pitcher_id] = result
        return result
    except Exception:
        _SP_RECENT_CACHE[pitcher_id] = None
        return None

def _get_sp_season_fip_by_id(pitcher_id):
    """
    Fetch FIP de temporada directamente desde statsapi usando el MLB pitcher ID.
    ═══════════════════════════════════════════════════════════════════════════
    FUENTE PRIMARIA DE SP FIP — sin matching por nombre, sin colisiones.
    El ID de MLB es único e inequívoco.

    FIP = ((13*HR + 3*(BB+HBP) - 2*K) / IP) + FIP_constant
    FIP_constant ≈ 3.15 (ajusta FIP a escala ERA; varía ~0.05 por temporada)

    Retorna dict {fip, era, gs, ip, k, bb, hr, hbp, source} o None si falla.
    """
    if pitcher_id is None:
        return None
    if pitcher_id in _SP_SEASON_FIP_CACHE:
        return _SP_SEASON_FIP_CACHE[pitcher_id]
    try:
        season = TARGET_DATE[:4]
        url = (f"https://statsapi.mlb.com/api/v1/people/{pitcher_id}/stats?"
               f"stats=season&season={season}&gameType=R&group=pitching")
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        splits = resp.json().get("stats", [{}])[0].get("splits", [])
        if not splits:
            # Sin datos esta temporada → verdadero debut (nunca ha pitcheado en MLB este año)
            result = {"fip": None, "era": None, "gs": 0, "g": 0, "ip": 0.0,
                      "k": 0, "bb": 0, "hr": 0, "hbp": 0,
                      "source": "debut-no-season-data", "is_opener": False}
            _SP_SEASON_FIP_CACHE[pitcher_id] = result
            return result
        st  = splits[0].get("stat", {})
        ip  = float(st.get("inningsPitched", 0) or 0)
        gs  = int(st.get("gamesStarted",   0) or 0)
        g   = int(st.get("gamesPlayed",    0) or st.get("games", 0) or 0)
        k   = int(st.get("strikeOuts",     0) or 0)
        bb  = int(st.get("baseOnBalls",    0) or 0)
        hr  = int(st.get("homeRuns",       0) or 0)
        hbp = int(st.get("hitByPitch",     0) or 0)
        era_raw = st.get("era")
        era = float(era_raw) if era_raw and era_raw not in ("-.--", "-") else None
        if ip < 1.0:
            # Sin IP suficiente — puede ser debutante O reliever con muy pocas IP
            result = {"fip": None, "era": era, "gs": gs, "g": g, "ip": 0.0,
                      "k": k, "bb": bb, "hr": hr, "hbp": hbp, "source": "statsapi-ID-nodebut",
                      "is_opener": gs == 0 and g >= 3}
            _SP_SEASON_FIP_CACHE[pitcher_id] = result
            return result
        FIP_CONSTANT = 3.15   # 2026 — reajustar si liga sube/baja mucho
        fip = round(((13 * hr + 3 * (bb + hbp) - 2 * k) / ip) + FIP_CONSTANT, 2)
        # Clamp a rango razonable
        fip = round(max(2.50, min(7.50, fip)), 2)
        # is_opener detection:
        #   Case A: 0 GS, 3+ RP appearances → pure reliever opener
        #   Case B: 1-4 GS but IP/game < 3.0 AND mostly RP (g >= gs*3) →
        #           MLB counts opener appearances as GS; this catches pitchers like
        #           Griffin Jax who have 2 "GS" but avg < 2 IP per appearance.
        _ip_per_game = ip / g if g > 0 else 0
        is_opener = (
            (gs == 0 and g >= 3) or
            (0 < gs <= 4 and g >= gs * 3 and _ip_per_game < 3.0)
        )
        result = {"fip": fip, "era": era, "gs": gs, "g": g, "ip": round(ip, 1),
                  "k": k, "bb": bb, "hr": hr, "hbp": hbp, "source": "statsapi-ID",
                  "is_opener": is_opener}
        _SP_SEASON_FIP_CACHE[pitcher_id] = result
        return result
    except Exception:
        _SP_SEASON_FIP_CACHE[pitcher_id] = None
        return None


def _sp_recent_adj_xfip(season_xfip, recent_stats):
    """
    Ajusta SP xFIP de temporada usando ERA reciente con peso proporcional al IP disponible.

    Recibe dict {"era": float, "ip": float, "n": int} o None.
    Si None → retorna season_xfip sin cambios.

    Peso base del ERA reciente escalado por muestra (IP confidence):
      < 12 IP  → filtrado en _get_sp_recent_era (nunca llega aquí)
      12-17 IP → w_recent_max = 0.35  (3 salidas normales, muestra ajustada)
      17-22 IP → w_recent_max = 0.50  (4 salidas, confianza moderada)
      22+ IP   → w_recent_max = 0.60  (5 salidas, muestra sólida)
    Adaptativo: si la divergencia ERA vs xFIP ≥ 1.0, se usa el máximo del tier.
                si la divergencia es normal, se usa 80% del máximo del tier.
    Cap de bajada: el ERA reciente no puede comprimir más de 0.65 xFIP desde
                   la línea de temporada (evita ERA=0.00 en racha de suerte).
    """
    if recent_stats is None:
        return season_xfip
    recent_era = recent_stats["era"]
    recent_ip  = recent_stats["ip"]

    # Escalar confianza por IP disponible
    if recent_ip >= 22.0:
        w_max = 0.60   # muestra sólida (5 salidas completas)
    elif recent_ip >= 17.0:
        w_max = 0.50   # muestra moderada (4 salidas)
    else:
        w_max = 0.35   # muestra ajustada (3 salidas mínimas)

    # Peso adaptativo: más peso si la forma diverge claramente del xFIP de temporada
    divergence = abs(recent_era - season_xfip)
    if divergence >= 1.0:
        w_recent = w_max            # SP claramente fuera/dentro de forma → usar máximo del tier
    else:
        w_recent = w_max * 0.80     # variación normal → usar 80% del máximo

    blended = season_xfip * (1.0 - w_recent) + recent_era * w_recent
    # Cap de bajada: SOLO aplica cuando ERA reciente < 1.50 (territorio de suerte probable).
    # Un ERA como 0.00 ó 0.90 en cualquier muestra puede ser varianza pura; lo limitamos.
    # ERAs genuinos (≥ 1.50) en muestra suficiente (ya validada arriba) actúan libremente.
    # Esto diferencia un pitcher con 0.00 ERA en 3 salidas de uno con 2.62 ERA en 5:
    # el segundo recibe el crédito completo proporcional al IP que tiene.
    if recent_era < 1.50:
        blended = max(season_xfip - 0.65, blended)
    return round(max(2.80, min(5.80, blended)), 3)


# ──────────────────────────────────────────────────────
# SP H2H vs EQUIPO ESPECÍFICO — el factor que faltaba
# ──────────────────────────────────────────────────────
# El mercado considera el historial del SP contra este equipo
# en particular. Un pitcher con 6.80 ERA lifetime vs los Yankees
# es diferente a uno con 1.90 ERA contra ellos, aunque su xFIP
# de temporada sea el mismo. Esto afecta especialmente picks ML
# de underdogs donde el SP local tiene historial favorable.

_SP_VS_TEAM_CACHE = {}

def _get_sp_vs_team(pitcher_id, opp_team_abb):
    """
    Historial del SP contra el equipo oponente específico (últimas 3 temporadas).
    Retorna {"era": float, "ip": float, "gs": int} o None si no hay sample.
    Requiere ≥3 aperturas de ≥2 IP para ser confiable.
    """
    if pitcher_id is None or not opp_team_abb:
        return None
    cache_key = f"{pitcher_id}_{opp_team_abb.upper()}"
    if cache_key in _SP_VS_TEAM_CACHE:
        return _SP_VS_TEAM_CACHE[cache_key]

    cur_year  = int(TARGET_DATE[:4])
    seasons   = [cur_year, cur_year - 1, cur_year - 2]  # últimas 3 temporadas
    # Algunos equipos tienen abreviaciones distintas en el MLB API
    _API_ABB_MAP = {
        "CWS": "CHW", "SD": "SDP", "SF": "SFG", "TB": "TBR",
        "KC": "KCR", "NYY": "NYY", "NYM": "NYM", "LAD": "LAD",
        "LAA": "LAA",
    }
    api_abb = _API_ABB_MAP.get(opp_team_abb.upper(), opp_team_abb.upper())

    all_starts = []
    for season in seasons:
        try:
            url = (f"https://statsapi.mlb.com/api/v1/people/{pitcher_id}/stats?"
                   f"stats=gameLog&season={season}&gameType=R&group=pitching")
            r = requests.get(url, timeout=10)
            if r.status_code != 200:
                continue
            splits = r.json().get("stats", [{}])[0].get("splits", [])
            for s in splits:
                ip = float(s.get("stat", {}).get("inningsPitched", "0") or "0")
                if ip < 2.0:
                    continue
                # El game log incluye info del oponente en s["opponent"]
                opp_info = s.get("opponent", {})
                opp_abb_raw = (opp_info.get("abbreviation") or
                               opp_info.get("teamCode", "")).upper()
                if opp_abb_raw in (opp_team_abb.upper(), api_abb):
                    all_starts.append(s)
        except Exception:
            continue

    if len(all_starts) < 3:   # muestra mínima: 3 aperturas
        _SP_VS_TEAM_CACHE[cache_key] = None
        return None

    total_er = sum(int(s.get("stat", {}).get("earnedRuns", 0) or 0) for s in all_starts)
    total_ip = sum(float(s.get("stat", {}).get("inningsPitched", 0) or 0) for s in all_starts)
    if total_ip < 6.0:
        _SP_VS_TEAM_CACHE[cache_key] = None
        return None

    era = round((total_er / total_ip) * 9, 2)
    result = {"era": era, "ip": round(total_ip, 1), "gs": len(all_starts)}
    _SP_VS_TEAM_CACHE[cache_key] = result
    return result


def _h2h_xfip_adj(base_xfip, h2h_stats):
    """
    Ajusta xFIP de temporada usando el historial específico del SP vs este equipo.
    El mercado pondera esto — especialmente cuando hay un historial claro.

    Blend conservador: más starts → más confianza en el H2H.
      3 GS  → 20% peso H2H / 80% season xFIP
      6 GS  → 27% peso H2H
      10+ GS → 35% peso H2H (cap)
    Ajuste máximo: ±0.80 xFIP (equivale a ~2 carreras en 9 innings)
    """
    if h2h_stats is None:
        return base_xfip
    gs      = h2h_stats["gs"]
    h2h_era = h2h_stats["era"]
    # Confianza crece con sample: 3GS → 20%, 6GS → 27%, 10+GS → 35%
    h2h_w   = min(0.35, 0.15 + gs * 0.02)
    blended = base_xfip * (1.0 - h2h_w) + h2h_era * h2h_w
    # Cap: no mover xFIP más de ±0.80 desde el base de temporada
    blended = max(base_xfip - 0.80, min(base_xfip + 0.80, blended))
    return round(max(2.50, min(5.80, blended)), 3)


# ──────────────────────────────────────────────────────
# BULLPEN FATIGUE — uso de bullpen últimos 3 días
# ──────────────────────────────────────────────────────

def _get_bullpen_fatigue(target_date, cache_hours=4):
    """
    Calcula innings de bullpen (excluyendo abridor) en los últimos 3 días.
    Retorna {team: total_bp_ip}. Bullpenes cansados = peor xFIP efectivo.

    Estrategia: jala schedule (sin hydrate) para obtener gamePks de juegos
    Final → luego fetcha /api/v1/game/{pk}/boxscore individualmente en paralelo.
    (hydrate=boxscore en schedule NO embebe el boxscore — MLB API lo ignora.)
    """
    global _BP_FATIGUE_CACHE, _BP_FATIGUE_CACHE_TS
    import time as _time
    import concurrent.futures
    from collections import defaultdict

    now_ts = _time.time()
    if _BP_FATIGUE_CACHE and (now_ts - _BP_FATIGUE_CACHE_TS) < cache_hours * 3600:
        return _BP_FATIGUE_CACHE
    try:
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").date()
        start_dt  = target_dt - timedelta(days=3)
        end_dt    = target_dt - timedelta(days=1)

        # ── Paso 1: schedule sin hydrate → gamePks de juegos terminados ──
        url = (f"https://statsapi.mlb.com/api/v1/schedule?"
               f"sportId=1&startDate={start_dt}&endDate={end_dt}&gameType=R")
        r = requests.get(url, timeout=15)
        r.raise_for_status()

        games_meta = []   # [(gamePk, away_name, home_name)]
        for date_entry in r.json().get("dates", []):
            for game in date_entry.get("games", []):
                if game.get("status", {}).get("abstractGameState") != "Final":
                    continue
                pk = game.get("gamePk")
                if not pk:
                    continue
                away_name = game["teams"]["away"]["team"]["name"]
                home_name = game["teams"]["home"]["team"]["name"]
                games_meta.append((pk, away_name, home_name))

        if not games_meta:
            _BP_FATIGUE_CACHE    = {}
            _BP_FATIGUE_CACHE_TS = now_ts
            return {}

        # ── Paso 2: boxscore individual por juego (paralelo) ──────────────
        def _fetch_one_bx(meta):
            pk, away_name, home_name = meta
            try:
                bx_url = f"https://statsapi.mlb.com/api/v1/game/{pk}/boxscore"
                br = requests.get(bx_url, timeout=8)
                br.raise_for_status()
                teams_bx = br.json().get("teams", {})
                game_ips = {}
                for side, tname in [("away", away_name), ("home", home_name)]:
                    team     = TEAM_MAP.get(tname, tname.upper())
                    pitchers = teams_bx.get(side, {}).get("pitchers", [])
                    players  = teams_bx.get(side, {}).get("players",  {})
                    total_ip = 0.0
                    for pid in pitchers[1:]:   # pitchers[0] = abridor
                        pstats = (players.get(f"ID{pid}", {})
                                         .get("stats", {})
                                         .get("pitching", {}))
                        ip_raw = str(pstats.get("inningsPitched", "0") or "0")
                        try:
                            parts   = ip_raw.split(".")
                            total_ip += int(parts[0]) + (int(parts[1]) / 3 if len(parts) > 1 else 0)
                        except Exception:
                            pass
                    game_ips[team] = total_ip
                return game_ips
            except Exception:
                return {}

        team_bp_ip = defaultdict(float)
        with concurrent.futures.ThreadPoolExecutor(max_workers=12) as exe:
            for game_result in exe.map(_fetch_one_bx, games_meta):
                for team, ip in game_result.items():
                    team_bp_ip[team] += ip

        result = dict(team_bp_ip)
        _BP_FATIGUE_CACHE    = result
        _BP_FATIGUE_CACHE_TS = now_ts
        n_teams = len([v for v in result.values() if v > 0])
        print(f"  💪 Bullpen fatigue: {len(result)} equipos, {len(games_meta)} juegos — "
              f"{n_teams} con IP > 0")
        return result

    except Exception as e:
        print(f"  ⚠️  Bullpen fatigue fetch error: {e}")
        if DEBUG: raise
        return {}

def _bullpen_fatigue_adj(team, bp_xfip_base, fatigue_data):
    """
    Ajusta bp_xfip según cansancio de los últimos 3 días.
      >13 IP → bullpen muy usado → +0.25 xFIP
      >9  IP → moderado         → +0.12 xFIP
      <3  IP → descansado       → -0.08 xFIP (bono)
    """
    ip3 = fatigue_data.get(team, 6.0)   # default: uso normal
    if   ip3 > 13.0: penalty = +0.25
    elif ip3 >  9.0: penalty = +0.12
    elif ip3 <  3.0: penalty = -0.08
    else:            penalty =  0.0
    return round(min(5.80, max(2.50, bp_xfip_base + penalty)), 3)


# ──────────────────────────────────────────────────────
# CONFIRMED LINEUPS — batting order real del día
# ──────────────────────────────────────────────────────

def _get_confirmed_lineups(game_pk):
    """
    Retorna el batting order confirmado del día desde MLB Stats API.

    Endpoint: /api/v1/game/{pk}/boxscore
    Path:     teams.{away|home}.battingOrder  → [player_id, ...]

    battingOrder se llena en cuanto el manager entrega la alineación oficial
    (típicamente 1-2h antes del primer pitch). Vacío = lineup aún no confirmado.

    Nota: el endpoint /api/v1/game/{pk}/lineups devuelve claves "awayPlayers" /
    "homePlayers" (no "away"/"home"), por eso se usa boxscore que ya confirmamos
    funciona en el resto del código.

    Retorna {"away": [player_id,...], "home": [player_id,...]} o {}.
    """
    if game_pk is None:
        return {}
    if game_pk in _LINEUP_CACHE:
        return _LINEUP_CACHE[game_pk]
    try:
        url = f"https://statsapi.mlb.com/api/v1/game/{game_pk}/boxscore"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        teams = r.json().get("teams", {})
        result = {}
        names  = {}
        for side in ("away", "home"):
            side_data     = teams.get(side, {})
            batting_order = side_data.get("battingOrder", [])
            players_dict  = side_data.get("players", {})
            if batting_order:
                # battingOrder puede venir como ints o strings según versión de API
                result[side] = [int(pid) for pid in batting_order]
                # Capturar nombres del batting order (para debug report)
                names[side] = []
                for pid in batting_order:
                    info = players_dict.get(f"ID{int(pid)}", {})
                    names[side].append(info.get("person", {}).get("fullName", str(pid)))
        _LINEUP_CACHE[game_pk]       = result
        _LINEUP_NAMES_CACHE[game_pk] = names
        if DEBUG:
            for side, ids in result.items():
                print(f"    🔎 lineup {side} ({game_pk}): {len(ids)} bateadores")
        return result
    except Exception as exc:
        if DEBUG:
            print(f"  ⚠️  _get_confirmed_lineups({game_pk}): {exc}")
        _LINEUP_CACHE[game_pk]       = {}
        _LINEUP_NAMES_CACHE[game_pk] = {}
        return {}


def _get_player_ops(player_id, season):
    """
    Retorna OPS del jugador en la temporada actual desde MLB Stats API.
    Usa caché global _PLAYER_STAT_CACHE.
    """
    if player_id is None:
        return None
    key = (player_id, season)
    if key in _PLAYER_STAT_CACHE:
        return _PLAYER_STAT_CACHE[key]
    try:
        url = (f"https://statsapi.mlb.com/api/v1/people/{player_id}/stats?"
               f"stats=season&group=hitting&season={season}")
        r = requests.get(url, timeout=8)
        r.raise_for_status()
        splits = r.json().get("stats", [{}])[0].get("splits", [])
        if not splits:
            _PLAYER_STAT_CACHE[key] = None
            return None
        stat = splits[0].get("stat", {})
        pa  = int(stat.get("plateAppearances", 0) or 0)
        ops_str = stat.get("ops", "0") or "0"
        ops = float(ops_str) if ops_str != "0" else None
        if pa < 30 or ops is None:   # demasiado small sample
            _PLAYER_STAT_CACHE[key] = None
            return None
        _PLAYER_STAT_CACHE[key] = ops
        return ops
    except Exception:
        _PLAYER_STAT_CACHE[key] = None
        return None


# Peso por posición en el batting order (más PAs en los primeros spots)
_BATTING_ORDER_WEIGHTS = {1: 1.15, 2: 1.12, 3: 1.10, 4: 1.08,
                           5: 1.00, 6: 0.95, 7: 0.90, 8: 0.85, 9: 0.80}
_LEAGUE_AVG_OPS = 0.725   # OPS promedio de liga MLB 2024-2026 ≈ wRC+ 100

def _lineup_weighted_wrc(player_ids, base_wrc, season):
    """
    Calcula wRC+ ponderado por el batting order real del día.
    OPS individual → proxy wRC+: (ops / 0.725) * 100
    Peso por posición de bateo (1=más PAs, 9=menos).
    Blend: 65% lineup calculado + 35% team baseline (maneja small samples).
    Si menos de 5 jugadores tienen stats válidos → retorna base_wrc.
    """
    season_str = str(season)
    total_wt = 0.0
    total_wrc = 0.0
    valid = 0
    for pos_idx, pid in enumerate(player_ids[:9]):
        pos = pos_idx + 1
        wt = _BATTING_ORDER_WEIGHTS.get(pos, 0.85)
        ops = _get_player_ops(pid, season_str)
        if ops is not None:
            player_wrc = (ops / _LEAGUE_AVG_OPS) * 100
            player_wrc = max(40, min(200, player_wrc))   # sanity bounds
            total_wrc += player_wrc * wt
            total_wt  += wt
            valid += 1

    if valid < 5 or total_wt == 0:
        return base_wrc   # sin suficientes datos → team average
    lineup_wrc = total_wrc / total_wt
    blended = round(lineup_wrc * 0.65 + base_wrc * 0.35, 1)
    return blended


# ──────────────────────────────────────────────────────
# ROLLING 14-DAY wRC+ — forma ofensiva reciente del equipo
# ──────────────────────────────────────────────────────

def _parse_boxscore_batting(bx_side):
    """
    Extrae stats de bateo acumulables desde un lado del boxscore MLB API.
    Usa la lista `batters` + dict `players` (ambos CONFIRMADOS en hydrate=boxscore).
    teamStats.batting NO está disponible en el schedule+hydrate endpoint.
    Retorna dict con ab, h, bb, hbp, sf, d, t, hr o None si falta data.
    """
    try:
        batters_list = bx_side.get("batters", [])    # [player_id, ...]
        players_dict = bx_side.get("players", {})    # {"ID123": {stats:{batting:{...}}}}
        totals = {"ab": 0, "h": 0, "bb": 0, "hbp": 0,
                  "sf": 0, "d": 0, "t": 0, "hr": 0}
        for pid in batters_list:
            key = f"ID{pid}"
            player = players_dict.get(key, {})
            bat = player.get("stats", {}).get("batting", {})
            totals["ab"]  += int(bat.get("atBats",      0) or 0)
            totals["h"]   += int(bat.get("hits",        0) or 0)
            totals["bb"]  += int(bat.get("baseOnBalls", 0) or 0)
            totals["hbp"] += int(bat.get("hitByPitch",  0) or 0)
            totals["sf"]  += int(bat.get("sacFlies",    0) or 0)
            totals["d"]   += int(bat.get("doubles",     0) or 0)
            totals["t"]   += int(bat.get("triples",     0) or 0)
            totals["hr"]  += int(bat.get("homeRuns",    0) or 0)
        if totals["ab"] == 0:
            return None
        return totals
    except Exception:
        return None


def _ops_from_batting(stats):
    """
    Calcula OPS desde stats acumulados.
    OBP = (H + BB + HBP) / (AB + BB + HBP + SF)
    SLG = TB / AB   donde TB = H + D + 2T + 3HR
    """
    ab  = stats["ab"];  h  = stats["h"]
    bb  = stats["bb"];  hbp = stats["hbp"]
    sf  = stats["sf"]
    d   = stats["d"];   t  = stats["t"];  hr = stats["hr"]
    obp_denom = ab + bb + hbp + sf
    if obp_denom == 0 or ab == 0:
        return None
    obp = (h + bb + hbp) / obp_denom
    tb  = h + d + 2 * t + 3 * hr   # totalBases
    slg = tb / ab
    return round(obp + slg, 3)


def _get_rolling_wrc(target_date, days=14, cache_hours=4):
    """
    Calcula runs/game de los últimos N días usando schedule BÁSICO (sin hydrate).
    game.teams.{side}.score está SIEMPRE presente en juegos Final — garantizado.
    Convierte a wRC+ proxy: (avg_runs / 4.5) × 100
    Requiere mínimo 5 juegos para ser confiable.
    Cache 4 horas.
    """
    global _ROLLING_WRC_CACHE, _ROLLING_WRC_CACHE_TS
    import time as _time
    from collections import defaultdict
    now_ts = _time.time()
    if _ROLLING_WRC_CACHE and (now_ts - _ROLLING_WRC_CACHE_TS) < cache_hours * 3600:
        return _ROLLING_WRC_CACHE
    try:
        target_dt  = datetime.strptime(target_date, "%Y-%m-%d").date()
        start_dt   = target_dt - timedelta(days=days)
        end_dt     = target_dt - timedelta(days=1)   # solo juegos completados
        # Sin hydrate — schedule básico siempre devuelve score en juegos Final
        url = (f"https://statsapi.mlb.com/api/v1/schedule?"
               f"sportId=1&startDate={start_dt}&endDate={end_dt}&gameType=R")
        r = requests.get(url, timeout=15)
        r.raise_for_status()

        team_runs = defaultdict(list)   # team → [runs_game1, runs_game2, ...]
        for date_entry in r.json().get("dates", []):
            for game in date_entry.get("games", []):
                if game.get("status", {}).get("abstractGameState") != "Final":
                    continue
                for side in ("away", "home"):
                    tname = game["teams"][side]["team"]["name"]
                    team  = TEAM_MAP.get(tname, tname.upper())
                    score = game["teams"][side].get("score")
                    if score is not None:
                        team_runs[team].append(int(score))

        LEAGUE_AVG_RUNS = 4.5   # promedio MLB 2024-2026 ≈ 4.5 R/equipo/juego
        MIN_GAMES       = 5

        result = {}
        for team, runs_list in team_runs.items():
            if len(runs_list) >= MIN_GAMES:
                avg = sum(runs_list) / len(runs_list)
                wrc_proxy = round((avg / LEAGUE_AVG_RUNS) * 100, 1)
                # PA equiv para el sistema de trust (36 PA/juego ≈ MLB average)
                pa_equiv  = len(runs_list) * 36
                result[team] = {"wrc": wrc_proxy, "pa": pa_equiv}

        _ROLLING_WRC_CACHE    = result
        _ROLLING_WRC_CACHE_TS = now_ts
        if result:
            print(f"  📈 Rolling {days}-day R/G: {len(result)} equipos (≥{MIN_GAMES} juegos)")
        return result
    except Exception as e:
        print(f"  ⚠️  Rolling wRC+ fetch error: {e}")
        return {}


def _rolling_wrc_factor(team, rolling_wrc_data, base_wrc):
    """
    Compara wRC+ proxy reciente (14 días) con baseline de temporada.
    Equipo caliente (rolling >> base) → factor > 1.0
    Equipo frío  (rolling << base) → factor < 1.0

    El peso del ajuste escala con PAs para proteger small samples:
      <  80 PA → 20% del ajuste   (datos muy limitados)
      < 140 PA → 50% del ajuste   (muestra parcial)
      ≥ 140 PA → 100% del ajuste  (muestra sólida, ≥10 juegos)

    Cap: ±8% sobre el base_wrc.
    """
    entry = rolling_wrc_data.get(team)
    if entry is None or base_wrc <= 0:
        return base_wrc
    rolling = entry["wrc"]
    pa      = entry.get("pa", 0)

    raw_factor = rolling / base_wrc
    capped_factor = max(0.92, min(1.08, raw_factor))
    delta = capped_factor - 1.0   # ej: +0.05 para equipo caliente

    # Escalar ajuste por tamaño de muestra
    if   pa < 80:  trust = 0.20
    elif pa < 140: trust = 0.50
    else:          trust = 1.00

    final_factor = 1.0 + (delta * trust)
    return round(base_wrc * final_factor, 1)


# ──────────────────────────────────────────────────────
# HOME/AWAY wRC+ SPLITS — rendimiento según venue
# ──────────────────────────────────────────────────────

def _get_home_away_splits(season, cache_hours=12):
    """
    Calcula runs/game home vs away por equipo usando schedule básico (sin hydrate).
    game.teams.{side}.score → siempre disponible en juegos Final.
    Retorna {team: {"home": wrc_proxy, "away": wrc_proxy}}.
    Cache 12 horas.
    """
    global _HOME_AWAY_CACHE, _HOME_AWAY_CACHE_TS
    import time as _time
    from collections import defaultdict
    now_ts = _time.time()
    if _HOME_AWAY_CACHE and (now_ts - _HOME_AWAY_CACHE_TS) < cache_hours * 3600:
        return _HOME_AWAY_CACHE
    try:
        today_dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d").date()
        start_dt = max(
            date(int(season), 3, 20),         # inicio típico de temporada MLB
            today_dt - timedelta(days=45)      # máximo 45 días atrás
        )
        end_dt = today_dt - timedelta(days=1)

        url = (f"https://statsapi.mlb.com/api/v1/schedule?"
               f"sportId=1&startDate={start_dt}&endDate={end_dt}&gameType=R")
        r = requests.get(url, timeout=20)
        r.raise_for_status()

        home_runs = defaultdict(list)   # team → [runs when HOME]
        away_runs = defaultdict(list)   # team → [runs when AWAY]

        for date_entry in r.json().get("dates", []):
            for game in date_entry.get("games", []):
                if game.get("status", {}).get("abstractGameState") != "Final":
                    continue
                for side, bucket in [("away", away_runs), ("home", home_runs)]:
                    tname = game["teams"][side]["team"]["name"]
                    team  = TEAM_MAP.get(tname, tname.upper())
                    score = game["teams"][side].get("score")
                    if score is not None:
                        bucket[team].append(int(score))

        LEAGUE_AVG_RUNS = 4.5
        MIN_GAMES       = 4    # mínimo 4 juegos home o away para tener split

        result = {}
        all_teams = set(home_runs) | set(away_runs)
        for team in all_teams:
            entry = {}
            for key, bucket in [("home", home_runs), ("away", away_runs)]:
                lst = bucket.get(team, [])
                if len(lst) >= MIN_GAMES:
                    avg = sum(lst) / len(lst)
                    entry[key] = round((avg / LEAGUE_AVG_RUNS) * 100, 1)
            if entry:
                result[team] = entry

        _HOME_AWAY_CACHE    = result
        _HOME_AWAY_CACHE_TS = now_ts
        if result:
            print(f"  🏟️  Home/Away splits: {len(result)} equipos (season-to-date)")
        return result
    except Exception as e:
        print(f"  ⚠️  Home/Away splits error: {e}")
        return {}


def _apply_home_away_split(team, is_home, home_away_data, base_wrc):
    """
    Aplica split home/away al wRC+ base del equipo.

    CALIBRACIÓN 2026-05-11 — basada en análisis histórico MLB 2015-2024:
      - Equipos locales anotan ~5.5% más en casa que en la carretera
      - Equipos visitantes rinden ~5.5% menos (asimetría real)
      - Blend 65% split / 35% base: el dato empírico YTD pesa más que el base
      - Cap ±15%: permite reflejar equipos con splits extremos (ej: Dodgers en Chavez Ravine)
      - Fallback: away=-5.5%, home=+2% (asimétrico — el local tiene ventaja estructural
        de familiarity + crowd incluso cuando no hay datos YTD suficientes)
    """
    splits = home_away_data.get(team, {})
    side_key = "home" if is_home else "away"
    split_wrc = splits.get(side_key)
    if split_wrc is None:
        # Fallback calibrado: asimetría deliberada (local siempre tiene
        # ventaja estructural aunque sea nueva temporada / pocos juegos)
        if not is_home:
            return round(base_wrc * 0.945, 1)   # −5.5% visitante
        return round(base_wrc * 1.020, 1)        # +2.0% local
    blended = round(split_wrc * 0.65 + base_wrc * 0.35, 1)
    # Cap ±15% sobre base — más margen para equipos con splits extremos reales
    capped = max(base_wrc * 0.85, min(base_wrc * 1.15, blended))
    return round(capped, 1)


# ──────────────────────────────────────────────────────
# FRACTIONAL KELLY — bet sizing proporcional al edge
# ──────────────────────────────────────────────────────

def _kelly_fraction(model_p, odds_american, fraction=0.25):
    """
    Calcula el tamaño óptimo de apuesta usando Fractional Kelly.
    fraction=0.25 = Quarter Kelly (conservador, recomendado para deportes).

    Fórmula Kelly completa: f = (p*b - q) / b
      donde b = payout neto por unidad, q = 1 - p
    Kelly fraccionario: f * fraction
    Cap: máximo 5% del bankroll por apuesta.

    Retorna % del bankroll como string, ej. "2.3%".
    """
    import math as _math
    try:
        if odds_american >= 100:
            b = odds_american / 100.0        # +150 → paga 1.50 por unidad
        else:
            b = 100.0 / abs(odds_american)   # -150 → paga 0.667 por unidad
        q = 1.0 - model_p
        full_kelly = (model_p * b - q) / b
        if full_kelly <= 0:
            return "—"   # apuesta sin valor real
        frac = full_kelly * fraction
        pct  = round(min(5.0, max(0.1, frac * 100)), 1)
        return f"{pct}%"
    except Exception:
        return "—"


def _monte_carlo_totals(tA, tB, market_line, n=20_000):
    """Monte Carlo con distribución Poisson para P(OVER/UNDER).
    tA, tB  : carreras proyectadas por equipo (lambda de Poisson).
    market_line : línea de total del mercado (ej: 8.5).
    n       : simulaciones (20k = precisión ±0.5% con velocidad óptima).
    Retorna dict con p_over, p_under, p_push.
    """
    rng    = np.random.default_rng()
    runs_a = rng.poisson(max(tA, 0.01), n)
    runs_b = rng.poisson(max(tB, 0.01), n)
    total  = runs_a + runs_b
    p_over  = float(np.mean(total >  market_line))
    p_under = float(np.mean(total <  market_line))
    p_push  = float(np.mean(total == market_line))
    return {
        "p_over":  round(p_over,  4),
        "p_under": round(p_under, 4),
        "p_push":  round(p_push,  4),
    }


def _monte_carlo_spreads(tA, tB, n=20_000):
    """Monte Carlo para probabilidades de cobertura por spread (margen de victoria).
    tA : lambda Poisson del equipo away.
    tB : lambda Poisson del equipo home.
    Retorna función cover_prob(spread_point, team_is_away) → float.

    Ejemplo: cover_prob(-1.5, True)  → P(away gana por 2+ carreras)
             cover_prob(+1.5, False) → P(home pierde por 1 o menos = cubre +1.5)
    """
    rng    = np.random.default_rng()
    runs_a = rng.poisson(max(tA, 0.01), n).astype(int)
    runs_b = rng.poisson(max(tB, 0.01), n).astype(int)
    margin = runs_a - runs_b   # positivo = away gana

    def cover_prob(spread_point, team_is_away):
        """P(equipo cubre el spread dado).
        spread_point: línea del equipo (ej: -1.5 para favorito, +1.5 para underdog).
        team_is_away: True si el equipo es visitante.
        """
        if team_is_away:
            # Away cubre si margin > -spread_point
            # Ej: spread -1.5 → margin > 1.5 → away gana por 2+
            return float(np.mean(margin > -spread_point))
        else:
            # Home cubre si -margin > -spread_point, es decir margin < spread_point
            # Ej: spread -1.5 → margin < 1.5 → home gana por 2+
            return float(np.mean(margin < spread_point))

    return cover_prob


def _save_predictions_log(entries):
    """Guarda/actualiza el log de predicciones del modelo en PRED_LOG_FILE.
    entries: lista de dicts con keys date, game, model_total, market_total.
    Hace upsert por (date, game): actualiza si ya existe, agrega si no.
    """
    try:
        if os.path.exists(PRED_LOG_FILE):
            with open(PRED_LOG_FILE, "r", encoding="utf-8") as f:
                log = json.load(f)
        else:
            log = []
        for entry in entries:
            key = (entry["date"], entry["game"])
            found = False
            for row in log:
                if (row.get("date"), row.get("game")) == key:
                    row.update({k: v for k, v in entry.items() if v is not None})
                    found = True
                    break
            if not found:
                log.append(entry)
        with open(PRED_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(log, f, indent=2, ensure_ascii=False)
    except Exception as _e:
        print(f"  ⚠️  predictions log error: {_e}")


# ──────────────────────────────────────────────────────────────────────
# ACES RECONOCIDOS — lista curada para la temporada 2026.
# Fuente: ESPN/TSN 2026 MLB Ace Rankings (Kiley McDaniel, abril 2026).
#
# IMPORTANTE — DOBLE FILTRO:
#   1. Este set bypasea el GS qualifier (aces probados no necesitan 8 GS)
#   2. El xFIP gate SIGUE ACTIVO: si el pitcher anda mal (xFIP > 3.20),
#      _ace_suppressor devuelve 1.0 de todas formas. La lista solo le
#      da el beneficio de la duda en sample size, no en performance.
#
# Ejemplo: Castillo con xFIP 4.5 = sin suppressor aunque esté en lista.
#          Skenes con 3 GS y xFIP 2.3 = suppressor completo.
#
# Matching: substring sobre el nombre completo del SP (case-insensitive).
# Agrega / quita según cambien los status de los pitchers.
# ──────────────────────────────────────────────────────────────────────
KNOWN_ACES_2026 = {
    # Top 2 — consenso absoluto
    "SKUBAL",           # Tarik Skubal (DET) — #1 ESPN, back-to-back Cy Young
    "SKENES",           # Paul Skenes (PIT) — #2 ESPN, generacional

    # Challengers al top tier
    "CROCHET",          # Garrett Crochet (BOS) — #3 ESPN, 255K en 2025
    "YAMAMOTO",         # Yoshinobu Yamamoto (LAD) — #4 ESPN

    # Ace confirmado, surgió en 2025
    "CRISTOPHER SANCHEZ",   # Cristopher Sanchez (PHI) — #5 ESPN, 202IP/2.50ERA 2025
                            # ⚠ Usar nombre completo para no confundir con otros Sanchez

    # Ace tier sólido
    "HUNTER BROWN",     # Hunter Brown (HOU) — #6 ESPN, mejoría constante 3 años
                        # ⚠ Nombre completo para evitar falsos positivos
    "CEASE",            # Dylan Cease (TOR) — #7 ESPN, nuevo contrato $210M
    "WEBB",             # Logan Webb (SF) — #8 ESPN, líder en IP desde 2022
    "SALE",             # Chris Sale (ATL) — #9 ESPN, velo up 1.1mph en 2026
    "WHEELER",          # Zack Wheeler (PHI) — #10 ESPN, líder WAR pitchers desde 2021
                        # (rehabbing TOS, pero el xFIP gate manejará su forma real)
    "LUZARDO",          # Jesus Luzardo (PHI) — #11 ESPN, breakout candidato 2026
}


def _is_known_ace(sp_name):
    """True si el SP está en la lista de aces reconocidos por la industria."""
    if not sp_name:
        return False
    name_up = sp_name.upper()
    for ace in KNOWN_ACES_2026:
        if ace in name_up:
            return True
    return False


def _ace_suppressor(sp_xfip_raw, gs=None, known_ace=False):
    """
    Factor adicional de supresión para aces PROBADOS.

    known_ace=True  → pitcher en KNOWN_ACES_2026. Bypasea el GS qualifier
                      y recibe suppressor completo desde salida #1.
                      (Skenes con 3 GS sigue siendo Skenes.)

    known_ace=False → GS qualifier activo:
      < 8  GS → sin suppressor. 3 buenas salidas no hacen un ace.
      8-12 GS → suppressor parcial (50%).
      ≥ 13 GS → suppressor completo.

    TIERS (xFIP RAW antes de RTM):
      Tier 1 (≤ 2.50): -18% — Skenes, Cole modo dios, Yamamoto en forma
      Tier 2 (2.51–2.75): -13% — Wheeler, Skubal sólido
      Tier 3 (2.76–3.00): -8%  — Burnes, Cease, Gallen
      Tier 4 (3.01–3.20): -4%  — top starter sin tier ace
      > 3.20: sin suppressor extra (RTM/xFIP ya lo maneja)
    """
    if sp_xfip_raw is None:
        return 1.0

    # Determinar factor de xFIP tier
    if sp_xfip_raw <= 2.50:
        raw_factor = 0.82
    elif sp_xfip_raw <= 2.75:
        raw_factor = 0.87
    elif sp_xfip_raw <= 3.00:
        raw_factor = 0.92
    elif sp_xfip_raw <= 3.20:
        raw_factor = 0.96
    else:
        return 1.0  # no ace tier

    # Qualifier de sample size (bypaseado para aces conocidos)
    if known_ace:
        gs_weight = 1.00   # ace probado → suppressor completo, sin importar GS
    elif gs is None:
        gs_weight = 0.50   # sin data GS → precaución, aplicar mitad
    elif gs < 8:
        return 1.0         # muy pocos starts y no es ace conocido — no califica
    elif gs < 13:
        gs_weight = 0.50   # en proceso de validación
    else:
        gs_weight = 1.00   # sample sólido — suppressor completo

    # Interpolar hacia 1.0 según gs_weight
    return round(1.0 - (1.0 - raw_factor) * gs_weight, 4)


def calc_lines(wrc_away, wrc_home, xfip_tot_away, xfip_tot_home, pf,
               game_hour_local=None, ump_factor=1.0,
               form_a=1.0, form_b=1.0, rest_a=1.0, rest_b=1.0,
               standings_a=1.0, standings_b=1.0,
               sp_xfip_raw_away=None, sp_xfip_raw_home=None,
               gs_away=None, gs_home=None,
               sp_name_away=None, sp_name_home=None):
    # Baseline 4.30 calibrado vs 2024-2026 MLB (~4.3 runs/equipo/juego)
    # form_a / form_b      : ajuste L10 récord              (0.90 – 1.10)
    # rest_a / rest_b      : ajuste por días de descanso    (0.975 – 1.01)
    # standings_a/b        : ajuste win% de temporada       (0.94 – 1.06)
    # sp_xfip_raw_away/home: xFIP raw del SP (pre-RTM) para ace suppressor
    # gs_away/gs_home      : GS del SP esta temporada (qualifier de ace)
    # sp_name_away/home    : nombre del SP para lookup en KNOWN_ACES_2026
    dg = _daygame_factor(game_hour_local)

    # Ace suppressor: equipo A enfrenta al SP de HOME, equipo B enfrenta al SP de AWAY
    # known_ace bypasea GS qualifier — Skenes con 3 GS sigue siendo Skenes
    ace_sup_a = _ace_suppressor(sp_xfip_raw_home, gs=gs_home,
                                known_ace=_is_known_ace(sp_name_home))   # away bats vs home SP
    ace_sup_b = _ace_suppressor(sp_xfip_raw_away, gs=gs_away,
                                known_ace=_is_known_ace(sp_name_away))   # home bats vs away SP

    tA = (wrc_away / 100) * 4.30 * (xfip_tot_home / 4.2) * pf * dg * ump_factor * form_a * rest_a * standings_a * ace_sup_a
    tB = (wrc_home / 100) * 4.30 * (xfip_tot_away / 4.2) * pf * dg * ump_factor * form_b * rest_b * standings_b * ace_sup_b

    # ── Total y spread usan runs crudos (sin bonus estructural) ──────────
    total     = tA + tB
    pre_total = round(total * 2) / 2
    mtotals   = f"O {pre_total}" if total > pre_total else f"U {pre_total}"

    diff   = abs(tA - tB)
    diff_r = round(diff * 2) / 2
    if diff_r <= 0.5:
        mspread_team = "AWAY" if tA > tB else "HOME"
        mspread = f"{mspread_team} ML"
    elif diff_r == 1.0:
        mspread_team = "HOME" if tA > tB else "AWAY"
        mspread = f"{mspread_team} +1.5"
    else:
        mspread_team = "AWAY" if tA > tB else "HOME"
        mspread = f"{mspread_team} -1.5"

    # ── HFA Estructural — aplicado SOLO a win probability ────────────────
    # Captura ventajas que no se reflejan en runs: último turno al bate
    # (walk-off opportunity), crowd, familiaridad con el ambiente.
    # Calibrado en 1.025 → equivale a ~+1.5-2% WP para el local en matchup parejo.
    # Fuente: estudios Pythagorean MLB 2015-2024 muestran local gana 53.8% sin
    # diferencias de calidad → ~2.5% sobre el 50/50 puro.
    HFA_STRUCTURAL = 1.025
    tB_wp = tB * HFA_STRUCTURAL   # tB ajustado solo para win prob

    exp   = 1.83
    win_a = (tA**exp) / (tA**exp + tB_wp**exp)
    win_b = 1 - win_a

    def to_am(p):
        if p >= 0.5: return f"-{round(p/(1-p)*100)}"
        else:        return f"+{round((1-p)/p*100)}"

    return {"tA":round(tA,3),"tB":round(tB,3),
            "total":round(total,2),"preTotals":pre_total,
            "mTotals":mtotals,"mSpread":mspread,"mSpreadTeam":mspread_team,
            "ml_pct":f"{to_am(win_a)} / {to_am(win_b)}",
            "winA":round(win_a*100,1),"winB":round(win_b*100,1),
            "form_a":round(form_a,4),"form_b":round(form_b,4),
            "rest_a":round(rest_a,4),"rest_b":round(rest_b,4),
            "standings_a":round(standings_a,4),"standings_b":round(standings_b,4),
            "hfa_structural": HFA_STRUCTURAL}


# ──────────────────────────────────────────────────────
# PICK TRACKER (local)
# ──────────────────────────────────────────────────────

def _load_log():
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return []

def _save_log(log):
    with open(LOG_FILE,"w",encoding="utf-8") as f: json.dump(log, f, indent=2, ensure_ascii=False)

def _entry_stake(e):
    """Retorna el monto apostado. Soporta 'stake' (nuevo, $) y 'units' (legacy)."""
    return e.get("stake") if e.get("stake") is not None else e.get("units", 1)

def _fmt_stake(e):
    """Formatea apuesta como '$15.00' (nuevo) o '1u' (legacy)."""
    if e.get("stake") is not None:
        return f"${e['stake']:.2f}"
    return f"{e.get('units', 1)}u"

def _fmt_pnl(e):
    """Formatea P&L como '+$11.54' / '-$15.00' (nuevo) o '+0.77u' (legacy)."""
    pnl = e.get("pnl")
    if pnl is None: return "—"
    if e.get("stake") is not None:
        return (f"+${pnl:.2f}" if pnl >= 0 else f"-${abs(pnl):.2f}")
    return f"{pnl:+.2f}u"


def _parse_odds_input(s):
    """
    Convierte string de odds a entero. Maneja: '-110', '+105', '110', '-1.5' etc.
    Fix: strip el signo ANTES de int() para evitar doble negación.
    """
    s = s.strip()
    negative = s.startswith("-")
    digits = re.sub(r"[^0-9]", "", s)
    if not digits:
        raise ValueError(f"No se pudo parsear odds: '{s}'")
    val = int(digits)
    return -val if negative else val


def _normalize_pick_str(s):
    """
    Normaliza el formato del pick para display y guardado:
      'O 8.5'   → 'OVER 8.5'
      'U 9.0'   → 'UNDER 9.0'
      'Over 8.5' → 'OVER 8.5'
      'Under 9' → 'UNDER 9'
      'CUBS ML' → 'CUBS ML'   (sin cambios)
    """
    su = s.strip().upper()
    m = re.match(r'^O(?:VER)?\s+([\d.]+)', su)
    if m:
        return f"OVER {m.group(1)}"
    m = re.match(r'^U(?:NDER)?\s+([\d.]+)', su)
    if m:
        return f"UNDER {m.group(1)}"
    return su


def _bk_lookup_odds(bk, pick_upper, away):
    """
    Extrae el odd de un dict de libro para un pick dado.
    Soporta: 'TEAM ML', 'OVER X', 'UNDER X', 'TEAM +/-1.5'
    Retorna int o None.
    """
    # Over / Under
    if pick_upper.startswith("OVER") or pick_upper.startswith("O "):
        d = bk.get("Total_Over", {})
        return int(d["odds"]) if isinstance(d, dict) and "odds" in d else None
    if pick_upper.startswith("UNDER") or pick_upper.startswith("U "):
        d = bk.get("Total_Under", {})
        return int(d["odds"]) if isinstance(d, dict) and "odds" in d else None
    # Detectar equipo
    pick_team = next((t for t in TEAM_ABB.values() if t in pick_upper), away)
    # ML
    if "ML" in pick_upper:
        val = bk.get(f"ML_{pick_team}")
        if val is not None: return int(val)
        for k, v in bk.items():
            if k.startswith("ML_") and pick_team in k: return int(v)
    # Spread / Run Line
    elif re.search(r'[+-]\d+\.?\d*$', pick_upper):
        sp = bk.get(f"Spread_{pick_team}")
        if sp and isinstance(sp, dict) and "odds" in sp:
            return int(sp["odds"])
    return None


def _lookup_log_odds(game_str, pick_str, book="BetMGM", odds_prefetched=None):
    """
    Busca el odd de mercado para ML, Over/Under o Run Line desde la API.
    Si odds_prefetched se pasa, lo usa directamente (sin llamada extra a la API).
    Retorna int (ej. -110) o None si no se encuentra.
    """
    try:
        pick_upper = _normalize_pick_str(pick_str)

        parts    = game_str.replace(" VS ", "@").replace(" VS.", "@").replace(" @ ", "@").split("@")
        away_raw = parts[0].strip() if len(parts) > 0 else ""
        home_raw = parts[1].strip() if len(parts) > 1 else ""
        away = next((v for v in TEAM_ABB.values() if v in away_raw.upper()), away_raw.upper())
        home = next((v for v in TEAM_ABB.values() if v in home_raw.upper()), home_raw.upper())

        if odds_prefetched is not None:
            odds_all = odds_prefetched
        else:
            print(f"  🔍 Buscando odds de {book} para {away} vs {home}…")
            odds_all = get_market_odds()

        bk = _get_game_books(odds_all, away, home).get(book, {})
        if not bk:
            return None

        return _bk_lookup_odds(bk, pick_upper, away)
    except Exception:
        return None


def cmd_log_pick():
    print(f"\n{'═'*60}")
    print(f"  LABOY PICKS — REGISTRAR JUGADA  ({TARGET_DATE})")
    print(f"{'═'*60}\n")

    # ── Cargar picks del modelo ───────────────────────────────────────────
    # Fuente de verdad: mlb_log_state.json — snapshot escrito SOLO cuando se genera el debug HTML.
    # Garantiza que --log muestre EXACTAMENTE los picks del último HTML exportado.
    # Fallback: mlb_debug_state.json (último --picks). Último recurso: API.
    model_picks  = []
    cached_odds  = {}
    _from_cache  = False

    def _enrich_log(p):
        if not p.get("away") and " @ " in p.get("game", ""):
            parts = p["game"].split(" @ ", 1)
            p = dict(p)
            p.setdefault("away", parts[0].strip())
            p.setdefault("home", parts[1].strip())
        return dict(p)

    # Prioridad 1: mlb_log_state.json (picks del último --export-debug HTML)
    _log_state  = _load_log_state()
    _log_picks  = _log_state.get("picks", [])
    _log_date   = _log_state.get("date", "")

    if _log_picks and _log_date == TARGET_DATE and not FORCE_REPICK:
        model_picks = [_enrich_log(p) for p in _log_picks]
        _from_cache = True
        _sess       = _log_state.get("session", "full")
        print(f"  ⚡ {len(model_picks)} pick(s) del último debug HTML (sesión: {_sess})\n")
    else:
        # Prioridad 2: mlb_debug_state.json (último --picks run, si no hay log state de hoy)
        _debug_state = _load_debug_state()
        _debug_picks = _debug_state.get("picks", [])
        _debug_date  = _debug_state.get("date", "")
        if _debug_picks and _debug_date == TARGET_DATE and not FORCE_REPICK:
            model_picks = [_enrich_log(p) for p in _debug_picks]
            _from_cache = True
            _sess       = _debug_state.get("session", "full")
            print(f"  ⚡ {len(model_picks)} pick(s) del debug state (sesión: {_sess}) [sin HTML exportado hoy]\n")
        else:
            # Prioridad 3: API (no hay ningún estado guardado para hoy, o --force-repick)
            print("  ⏳ Jalando picks del modelo...", end="", flush=True)
            try:
                games_today, cached_odds = compute_lines_from_api(silent=True, skip_lineups=False)
                if games_today:
                    model_picks = _compute_picks(games_today, cached_odds)
                print(f" {len(model_picks)} pick(s) encontrado(s).\n")

                if model_picks:
                    _model_picks_save_today(model_picks)
                _save_debug_state(model_picks)
            except Exception as _lpe:
                print(f" ⚠️  {_lpe}\n")

    try:
        # ── Menú de picks del modelo ──────────────────────────────────────
        game = ""
        pick = ""

        if model_picks:
            print("  🎯 Picks recomendados por el modelo:\n")
            for i, p in enumerate(model_picks, 1):
                pick_disp = _normalize_pick_str(p["pick"])
                ev_s      = p.get("ev", "")
                edge_s    = p.get("edge", "")
                print(f"   {i:>2}.  {p['game']:<30}  {pick_disp:<14}  {p['odds']:<7}  "
                      f"EV:{ev_s}  edge:{edge_s}")
            print()
            print("   M.  Ingresar manualmente")
            print()
            sel = input(f"  Escoge [1-{len(model_picks)} / M]: ").strip().upper()
        else:
            sel = "M"

        # chosen_odds: odds del pick seleccionado del modelo (ya guardados, sin API)
        chosen_odds = None

        if sel == "M" or sel == "" or not model_picks:
            # ── Entrada manual ───────────────────────────────────────────
            game = input("  Juego  (ej: CUBS @ PIRATES): ").strip().upper()
            raw_pick = input("  Pick   (ej: CUBS ML / OVER 8.5 / CUBS -1.5): ").strip()
            pick = _normalize_pick_str(raw_pick)
        else:
            try:
                idx = int(sel) - 1
                if not (0 <= idx < len(model_picks)):
                    raise ValueError(f"Selección fuera de rango")
                chosen = model_picks[idx]
                game   = chosen["game"]
                pick   = _normalize_pick_str(chosen["pick"])
                # Si viene del caché, los odds ya están guardados → no necesitamos API
                if _from_cache:
                    try:
                        chosen_odds = int(chosen.get("odds", 0)) or None
                    except (TypeError, ValueError):
                        chosen_odds = None
                print(f"\n  ✓  {game}  →  {pick}\n")
            except ValueError as ve:
                print(f"  ❌ {ve}. Intenta de nuevo.\n"); return

        # ── Sportsbook ───────────────────────────────────────────────────
        book_raw = input("  Sportsbook [BetMGM]: ").strip()
        book     = book_raw if book_raw else "BetMGM"

        # ── Odds (del caché si disponible, si no busca en la API) ────────
        if chosen_odds is not None:
            # Odds directamente del pick guardado → instantáneo
            auto_odds = chosen_odds
        else:
            auto_odds = _lookup_log_odds(game, pick, book,
                                         odds_prefetched=cached_odds if cached_odds else None)
        if auto_odds is not None:
            print(f"  💰 Odds: {_fmt_odds(auto_odds)}  (Enter para aceptar, o escribe otro)")
            override = input("  Odds: ").strip()
            odds_v   = _parse_odds_input(override) if override else auto_odds
        else:
            odds_s = input(f"  Odds (no encontrados en {book}, ej: +150 / -110): ").strip()
            odds_v = _parse_odds_input(odds_s)

        # ── Stake ────────────────────────────────────────────────────────
        stake_s     = input("  Apostado (ej: 15 para $15.00): ").strip()
        stake_clean = re.sub(r"[^\d.]", "", stake_s.split()[0])
        stake       = float(stake_clean)

        # ── Análisis ─────────────────────────────────────────────────────
        analysis = input("  Análisis (opcional — razón del pick, Enter para omitir):\n  > ").strip()

    except (ValueError, EOFError) as _e:
        print(f"\n  ❌ Entrada inválida ({_e}). Intenta de nuevo.\n"); return

    log   = _load_log()
    entry = {"id": len(log), "date": TARGET_DATE, "game": game, "pick": pick,
             "odds": odds_v, "stake": stake, "book": book, "result": None, "pnl": None,
             "analysis": analysis or ""}
    log.append(entry)
    _save_log(log)

    pot = round(stake * (odds_v / 100) if odds_v > 0 else stake * (100 / abs(odds_v)), 2)
    print(f"\n  ✅ Pick #{entry['id']}: {game} │ {pick} │ {_fmt_odds(odds_v)} │ ${stake:.2f} → potencial +${pot:.2f}")
    html_path = None
    try:
        html_path = export_log_pick_html(entry)
        if html_path:
            print(f"     → {os.path.basename(html_path)}")
    except Exception as e:
        print(f"     ⚠️  HTML export falló: {e}")
    if html_path:   # siempre publica al loguear
        cmd_publish([html_path])
    print()


def cmd_export_log():
    """
    --export-log [IDX] [--publish]
    Re-exporta un pick logueado como HTML + JPG.
    Si no se da IDX, exporta el último pick logueado.
    Con --publish sube el HTML y el JPG al dashboard de GitHub Pages.
    Ejemplo: python3 mlb.py --export-log 3 --publish
    """
    log = _load_log()
    if not log:
        print("  ❌ No hay picks logueados. Usa --log primero.\n"); return
    try:
        ei  = sys.argv.index("--export-log")
        idx = int(sys.argv[ei+1]) if ei+1 < len(sys.argv) and sys.argv[ei+1].lstrip('-').isdigit() else len(log)-1
    except (ValueError, IndexError):
        idx = len(log) - 1
    if not (0 <= idx < len(log)):
        print(f"  ❌ Índice {idx} inválido. Hay {len(log)} picks (0–{len(log)-1}).\n"); return
    entry = log[idx]
    print(f"\n  🌐 Exportando Pick #{idx}: {entry['game']} │ {entry['pick']} │ {_fmt_odds(entry['odds'])}")
    html_path = export_log_pick_html(entry)
    if html_path:
        print(f"  ✅ Guardado: {os.path.basename(html_path)}\n")
        if PUBLISH_MODE:
            cmd_publish([html_path])
    else:
        print("  ❌ Error al generar HTML.\n")


def cmd_grade_pick():
    """
    Uso: python3 mlb.py --grade IDX W|L|P
    IDX  = número del pick (sale en --record, empieza en 0)
    W    = Win (ganaste)
    L    = Loss (perdiste)
    P    = Push (devuelven dinero)

    Ejemplo: python3 mlb.py --grade 0 W
    """
    try:
        gi  = sys.argv.index("--grade")
        idx = int(sys.argv[gi+1])
        res = sys.argv[gi+2].upper()
        assert res in ("W","L","P")
    except (ValueError, IndexError, AssertionError):
        print("  ❌ Uso: python3 mlb.py --grade IDX W|L|P")
        print("     Ejemplo: python3 mlb.py --grade 0 W")
        print("\n  IDX  = número del pick (ver python3 mlb.py --record)")
        print("  W    = Win   L = Loss   P = Push")
        return
    log = _load_log()
    if not (0 <= idx < len(log)):
        print(f"  ❌ Índice {idx} no válido. Hay {len(log)} picks (0–{len(log)-1}).\n"
              f"     Corre: python3 mlb.py --record  para ver los índices.")
        return
    e = log[idx]; e["result"] = res
    sv = _entry_stake(e)   # stake en $ o units (legacy)
    if res == "W":
        e["pnl"] = round(sv * (e["odds"] / 100) if e["odds"] > 0
                         else sv * (100 / abs(e["odds"])), 2)
    elif res == "L": e["pnl"] = round(-sv, 2)
    else:            e["pnl"] = 0.0
    _save_log(log)
    emoji = {"W":"✅","L":"❌","P":"🔄"}[res]
    print(f"\n  {emoji} Pick #{idx} → {res} | {e['game']} {e['pick']} {_fmt_odds(e['odds'])} | P&L: {_fmt_pnl(e)}")
    # Regenerar HTML pick card con badge WIN/LOSS actualizado
    try:
        html_path = export_log_pick_html(e)
        if html_path:
            print(f"  🖼️  Pick card actualizado: {os.path.basename(html_path)}")
    except Exception as _ge:
        print(f"  ⚠️  No se pudo regenerar HTML: {_ge}")
    print()


def cmd_remove_pick():
    """
    Elimina uno o varios picks del log por su índice (ID).
    Uso: python3 mlb.py --remove 2
         python3 mlb.py --remove 2 4 7

    Los IDs restantes se renumeran automáticamente para mantener consistencia.
    """
    # Colectar todos los índices que siguen a --remove
    try:
        ri = sys.argv.index("--remove")
        idxs = []
        for a in sys.argv[ri+1:]:
            if a.startswith("--"): break
            idxs.append(int(a))
    except (ValueError, IndexError):
        print("  ❌ Uso: python3 mlb.py --remove IDX [IDX2 ...]")
        print("     Ejemplo: python3 mlb.py --remove 3")
        return

    if not idxs:
        print("  ❌ Especifica al menos un índice. Usa --record para ver los IDs.\n")
        return

    log = _load_log()
    if not log:
        print("  ❌ No hay picks en el log.\n"); return

    # Validar todos los índices primero
    invalid = [i for i in idxs if not (0 <= i < len(log))]
    if invalid:
        print(f"  ❌ Índice(s) inválido(s): {invalid}. El log tiene {len(log)} picks (0–{len(log)-1}).\n")
        return

    # Mostrar qué se va a borrar y pedir confirmación
    print(f"\n{'═'*52}")
    print(f"  ⚠️  ELIMINAR {len(idxs)} PICK(S)")
    print(f"{'═'*52}")
    for i in sorted(idxs):
        e = log[i]
        print(f"  #{i} — {e['date']} | {e['game']} | {e['pick']} {_fmt_odds(e['odds'])}")
    print()
    confirm = input("  ¿Confirmar eliminación? (s/N): ").strip().lower()
    if confirm not in ("s","si","sí","y","yes"):
        print("  ↩️  Cancelado.\n"); return

    # Eliminar y renumerar
    to_remove = set(idxs)
    new_log   = [e for i, e in enumerate(log) if i not in to_remove]
    for new_id, e in enumerate(new_log):
        e["id"] = new_id
    _save_log(new_log)

    print(f"\n  ✅ {len(idxs)} pick(s) eliminado(s). Quedan {len(new_log)} en el log.\n")


def cmd_record():
    """
    --record              → últimos 30 picks (más recientes primero)
    --record all          → todos los picks
    --record 2026-04-18   → solo picks de esa fecha
    --record --pending    → solo picks sin gradear
    """
    log = _load_log()
    print(f"\n{'═'*80}")
    print(f"  LABOY PICKS — MLB · REGISTRO")
    print(f"{'═'*80}")
    if not log:
        print("\n  No hay jugadas. Usa: python3 mlb.py --log\n"); return

    # ── Parse argument after --record ────────────────────────────────────────
    date_filter  = None
    show_all     = False
    pending_only = PENDING_MODE
    try:
        ri  = sys.argv.index("--record")
        arg = sys.argv[ri + 1] if ri + 1 < len(sys.argv) and not sys.argv[ri + 1].startswith("--") else None
        if arg:
            if arg.lower() == "all":
                show_all = True
            elif re.match(r"^\d{4}-\d{2}-\d{2}$", arg):
                date_filter = arg
    except (ValueError, IndexError):
        pass

    # ── Build running balance over ALL picks ─────────────────────────────────
    running_balance = 0.0
    balance_by_id   = {}
    for e in log:
        res   = e.get("result") or "—"
        pnl   = e.get("pnl")
        stake = _entry_stake(e)
        if res == "W":
            running_balance += pnl if pnl is not None else stake
        elif res == "L":
            running_balance -= stake
        balance_by_id[e["id"]] = running_balance

    # ── Filter ───────────────────────────────────────────────────────────────
    if pending_only:
        display_log  = [e for e in log if not e.get("result")]
        filter_label = f"  ⏳ {len(display_log)} picks PENDIENTES de {len(log)} totales"
        if not display_log:
            print("\n  ✅ Todos los picks están calificados.\n"); return
    elif date_filter:
        display_log  = [e for e in log if e.get("date","") == date_filter]
        filter_label = f"  📅 Filtrado: {date_filter}  ({len(display_log)} picks)"
        if not display_log:
            print(f"\n  Sin picks para {date_filter}.\n"); return
    elif show_all:
        display_log  = log
        filter_label = f"  📋 Todos los picks ({len(log)})"
    else:
        display_log  = log[-30:]
        filter_label = f"  📋 Últimos {len(display_log)} picks  (usa --record all para ver todos)"

    # ── Newest first ─────────────────────────────────────────────────────────
    display_log = list(reversed(display_log))

    rows = []
    for e in display_log:
        res     = e.get("result") or "—"
        bal_val = balance_by_id.get(e["id"], 0.0)
        bal_fmt = f"+${bal_val:.2f}" if bal_val >= 0 else f"-${abs(bal_val):.2f}"
        rows.append([e["id"], e["date"], e["game"][:24], e["pick"][:14],
                     _fmt_odds(e["odds"]), _fmt_stake(e), res, _fmt_pnl(e), bal_fmt])

    print(f"\n{filter_label}")
    print("\n" + tab(rows, ["#","Fecha","Juego","Pick","Odds","Apostado","Res","P&L","Ganancia"]))

    # ── Stats always over full log ────────────────────────────────────────────
    graded = [e for e in log if e.get("result") in ("W","L","P")]
    wins   = [e for e in graded if e["result"]=="W"]
    pnl_t  = sum(e["pnl"] for e in graded if e.get("pnl") is not None)
    wag    = sum(_entry_stake(e) for e in graded)
    roi    = (pnl_t / wag * 100) if wag > 0 else 0
    use_dollars = any(e.get("stake") is not None for e in graded)
    pnl_str = f"+${pnl_t:.2f}" if (use_dollars and pnl_t >= 0) else (f"-${abs(pnl_t):.2f}" if use_dollars else f"{pnl_t:+.2f}u")
    wag_str = f"${wag:.2f}" if use_dollars else f"{wag}u"
    bal_final = f"+${running_balance:.2f}" if running_balance >= 0 else f"-${abs(running_balance):.2f}"

    print(f"\n  📊 Récord total: {len(wins)}-{len([e for e in graded if e['result']=='L'])}-"
          f"{len([e for e in graded if e['result']=='P'])}  "
          f"Pending: {len(log)-len(graded)}")
    if graded:
        print(f"     Win%: {len(wins)/len(graded)*100:.1f}%  │  P&L: {pnl_str}  │  Jugado: {wag_str}  │  ROI: {roi:+.1f}%")
    print(f"  💰 Ganancia actual: {bal_final}")

    # Export card
    try:
        card_path = export_record_card(TARGET_DATE)
        if card_path:
            print(f"  💳 Card exportada: {os.path.basename(card_path)}")
    except Exception:
        pass

    print(f"\n  Tips:")
    print(f"    python3 mlb.py --record all              ← ver todos los picks")
    print(f"    python3 mlb.py --record 2026-04-18       ← picks de una fecha")
    print(f"    python3 mlb.py --record --pending        ← solo picks sin gradear")
    print(f"    python3 mlb.py --grade N W|L|P           ← califica pick #N")
    print(f"    python3 mlb.py --remove N                ← elimina pick #N del log")
    print(f"    python3 mlb.py --export-record           ← exporta card como HTML+JPG\n")


def cmd_feedback():
    log    = _load_log()
    graded = [e for e in log if e.get("result") in ("W","L")]
    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS — ANÁLISIS DE RENDIMIENTO")
    print(f"{'═'*72}")
    if len(graded) < 3:
        print(f"\n  Necesitas ≥3 picks calificados. Tienes {len(graded)}.\n"); return
    use_dollars = any(e.get("stake") is not None for e in graded)
    def _pnl_s(v): return (f"+${v:.2f}" if v >= 0 else f"-${abs(v):.2f}") if use_dollars else f"{v:+.2f}u"

    by_type = {}
    for e in graded:
        p = e.get("pick","").upper()
        bt = "ML" if "ML" in p else "Over" if (p.startswith("O ") or "OVER" in p) \
             else "Under" if (p.startswith("U ") or "UNDER" in p) \
             else "Spread" if any(x in p for x in ["-1.5","+1.5","SPREAD"]) else "Other"
        s = by_type.setdefault(bt, {"W":0,"L":0,"pnl":0.0})
        s[e["result"]] += 1; s["pnl"] += e.get("pnl", 0)
    type_rows = []
    for bt, s in sorted(by_type.items()):
        tot = s["W"] + s["L"]
        type_rows.append([bt, f"{s['W']}-{s['L']}", f"{s['W']/tot*100:.0f}%", _pnl_s(s["pnl"])])
    print(f"\n  Por tipo:\n" + tab(type_rows, ["Tipo","W-L","Win%","P&L"], fmt="simple"))

    def bucket(o):
        if o >= 200:  return "+200 y más"
        if o >= 101:  return "+101 a +199"
        if o >= -120: return "EVEN a -120"
        if o >= -150: return "-121 a -150"
        return "-151 y menos"
    by_o = {}
    for e in graded:
        b = bucket(e.get("odds", 0))
        s = by_o.setdefault(b, {"W":0,"L":0,"pnl":0.0})
        s[e["result"]] += 1; s["pnl"] += e.get("pnl", 0)
    odds_rows = []
    for b in ["+200 y más","+101 a +199","EVEN a -120","-121 a -150","-151 y menos"]:
        if b in by_o:
            s = by_o[b]; tot = s["W"] + s["L"]
            odds_rows.append([b, f"{s['W']}-{s['L']}", f"{s['W']/tot*100:.0f}%", _pnl_s(s["pnl"])])
    if odds_rows:
        print(f"\n  Por rango odds:\n" + tab(odds_rows, ["Odds Range","W-L","Win%","P&L"], fmt="simple"))
        # Flag específico para el rango problemático
        _dog_range = by_o.get("+101 a +199")
        if _dog_range:
            _dog_tot = _dog_range["W"] + _dog_range["L"]
            _dog_wp  = _dog_range["W"] / _dog_tot * 100
            if _dog_wp < 25:
                print(f"\n  🚨 +101→+199: {_dog_range['W']}-{_dog_range['L']} ({_dog_wp:.0f}%) — "
                      f"FILTRO ACTIVO en el modelo (bloqueado automáticamente)")
            elif _dog_wp < 40:
                print(f"\n  ⚠️  +101→+199: {_dog_range['W']}-{_dog_range['L']} ({_dog_wp:.0f}%) — "
                      f"por debajo del breakeven (52.4%)")
    wp  = sum(1 for e in graded if e["result"]=="W") / len(graded) * 100
    pnl = sum(e.get("pnl", 0) for e in graded)
    print(f"\n  💡 Win rate: {wp:.0f}%  P&L: {_pnl_s(pnl)}")
    if by_type:
        best  = max(by_type.items(), key=lambda x: x[1]["pnl"])
        worst = min(by_type.items(), key=lambda x: x[1]["pnl"])
        print(f"  ✅ Mejor tipo: {best[0]} ({_pnl_s(best[1]['pnl'])})")
        if worst[0] != best[0]:
            print(f"  ⚠️  Área a mejorar: {worst[0]} ({_pnl_s(worst[1]['pnl'])})")

    # AI-powered feedback
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if api_key:
        print(f"\n  🤖 Análisis AI en progreso...\n")
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)

            # Construir prompt con resumen de picks
            recent_picks_text = ""
            for e in graded[-10:]:  # últimos 10
                recent_picks_text += f"  {e['date']} | {e['game']} | {e['pick']} | {_fmt_odds(e['odds'])} | {e.get('result','?')}\n"

            prompt = f"""Eres un analista de deportes experto en MLB.

Resumen de picks de béisbol:
- Win rate: {wp:.0f}%
- P&L total: {_pnl_s(pnl)}
- Por tipo: {', '.join(f"{bt} {s['W']}-{s['L']}" for bt, s in by_type.items())}

Picks recientes:
{recent_picks_text}

Analiza estos picks y da feedback constructivo en español. Para cada patrón de pérdida, sugiere: ¿fue clima adverso? ¿error del modelo? ¿trampa (favorite para el mercado)? ¿muestra pequeña? Sé específico y usa los datos. Máximo 200 palabras."""

            message = client.messages.create(
                model="claude-opus-4-6",
                max_tokens=1024,
                messages=[{"role":"user","content": prompt}]
            )

            feedback_text = message.content[0].text
            print(f"  💭 {feedback_text}\n")

        except Exception as e:
            print(f"  ⚠️  AI análisis falló: {e}\n")
    else:
        print(f"\n  💡 Para análisis AI, configura: export ANTHROPIC_API_KEY='tu_key'\n")

    print()


# ──────────────────────────────────────────────────────
# STATS DISPLAY
# ──────────────────────────────────────────────────────

def show_stats():
    """Muestra blend de wRC+ y BP xFIP desde caché FanGraphs."""
    data   = load_fg_blended()
    raw    = load_fg_cache()
    splits = load_wrc_splits()
    fetched_at = raw.get("fetched_at", "desconocida")
    splits_ok  = bool(splits.get("vs_rhp") and splits.get("vs_lhp"))
    splits_broken = splits.get("_splits_broken", False)

    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS — TEAM STATS (BLENDED)")
    print(f"  wRC+: 75% 2025 + 25% 2026   │   BP xFIP: 90% 2025 + 10% 2026")
    print(f"  Fuente: FanGraphs caché — {fetched_at}")
    if splits_broken:
        print(f"  ⚠️  PLATOON SPLITS: datos inválidos (FanGraphs devuelve wRC+ general")
        print(f"     para ambos splits — el modelo usa base wRC+ sin ajuste de platoon)")
    elif not splits_ok:
        print(f"  ⚠️  PLATOON SPLITS: no disponibles en caché. Corre --refresh para intentar.")
    else:
        print(f"  ✅ Platoon splits (vs LHP / vs RHP) disponibles")
    print(f"{'═'*72}")

    if splits_ok and not splits_broken:
        vs_rhp = splits["vs_rhp"]; vs_lhp = splits["vs_lhp"]
        rows = []
        for t, d in data.items():
            rhp = vs_rhp.get(t); lhp = vs_lhp.get(t)
            rows.append((
                t,
                f"{d.get('wrc','—'):.1f}" if isinstance(d.get("wrc"), float) else "—",
                f"{rhp:.1f}" if rhp else "—",
                f"{lhp:.1f}" if lhp else "—",
                f"{d.get('bp_xfip','—'):.3f}" if isinstance(d.get("bp_xfip"), float) else "—",
            ))
        rows.sort(key=lambda x: float(x[1]) if x[1] != "—" else 0, reverse=True)
        print("\n" + tab(rows, ["Team","wRC+ Blend","vs RHP","vs LHP","BP xFIP"]))
    else:
        rows = sorted([(t, d.get("wrc","—"), d.get("bp_xfip","—"))
                       for t, d in data.items()],
                      key=lambda x: float(x[1]) if str(x[1]) not in ("—","None") else 0,
                      reverse=True)
        print("\n" + tab(rows, ["Team","wRC+ Blend","BP xFIP Blend"]))
        if splits_broken:
            print(f"\n  ℹ️  Para obtener splits reales necesitas FanGraphs individual leaderboard")
            print(f"     (team=0,ts no soporta split= en el API). Platoon no afecta el modelo ahora.")
    print()


def show_stats_raw():
    """Muestra valores RAW 2026 y 2025 desde caché FanGraphs."""
    raw  = load_fg_cache()
    data = blend_fg_data(raw)
    fetched_at = raw.get("fetched_at", "desconocida")
    print(f"\n{'═'*80}")
    print(f"  LABOY PICKS — STATS RAW (DATOS CRUDOS — FanGraphs caché)")
    print(f"  Fuente: {FG_CACHE_FILE}")
    print(f"  Actualizado: {fetched_at}")
    print(f"{'═'*80}")
    wrc25_d = raw.get("wrc_2025", {}); wrc26_d = raw.get("wrc_2026", {})
    bp25_d  = raw.get("bp_2025",  {}); bp26_d  = raw.get("bp_2026",  {})
    all_teams = sorted(set(wrc25_d) | set(wrc26_d) | set(bp25_d) | set(bp26_d))
    rows = []
    for team in all_teams:
        wrc25 = wrc25_d.get(team); wrc26 = wrc26_d.get(team)
        bp25  = bp25_d.get(team);  bp26  = bp26_d.get(team)
        d = data.get(team, {})
        rows.append([
            team,
            f"{wrc25:.1f}" if wrc25 else "—",
            f"{wrc26:.1f}" if wrc26 else "—",
            f"{d.get('wrc','—'):.1f}" if isinstance(d.get("wrc"), float) else d.get("wrc","—"),
            f"{bp25:.3f}" if bp25 else "—",
            f"{bp26:.3f}" if bp26 else "—",
            f"{d.get('bp_xfip','—'):.3f}" if isinstance(d.get("bp_xfip"), float) else d.get("bp_xfip","—"),
        ])
    rows.sort(key=lambda x: float(x[1]) if x[1] != "—" else 0, reverse=True)
    print("\n" + tab(rows, ["Team","wRC+2025","wRC+2026","wRC+Blend",
                             "BP_xFIP2025","BP_xFIP2026","BP_Blend"]))
    print(f"\n  ⚠️  Si 2026 tiene pocos datos (inicio de temporada), el blend lo pondera menos.")
    print(f"  Para refrescar: python3 mlb.py --refresh\n")


# ──────────────────────────────────────────────────────
# DISPLAY DE LÍNEAS
# ──────────────────────────────────────────────────────

def display_lines(results, odds={}):
    dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    fip_warn = not FG_FIP_AVAILABLE
    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS — MLB   {dt.strftime('%A, %B %d %Y').upper()}")
    if fip_warn:
        print(f"  ⚠️  FIP: FanGraphs no disponible — usando BP xFIP como proxy")
        print(f"      Agrega cookie: export FANGRAPHS_COOKIE='tu_cookie'")
    print(f"{'═'*72}")

    for r in results:
        away, home = r["away"], r["home"]
        gkey       = f"{away} vs {home}"
        lines      = r["lines"]
        weather    = r["weather"]
        away_sp    = r["away_sp"]
        home_sp    = r["home_sp"]
        gtime      = r.get("game_time_local","")
        fip_src    = r.get("fip_source","")

        mspread = lines["mSpread"].replace("AWAY",away).replace("HOME",home)
        mtotals = lines["mTotals"]

        w_str = ""
        if weather["dir"] == "DOME":
            _ri = STADIUM_ROOF.get(home, {})
            w_str = ("🔒 Techo Fijo" if _ri.get("roof") == "fixed_dome" else "🏟️ Retráctil")
            if _ri.get("name"): w_str += f" · {_ri['name']}"
        elif weather["temp"] is not None and weather["dir"]:
            raw_dir = weather.get("raw_dir","")
            raw_deg = weather.get("raw_deg","")
            src     = weather.get("source","Open-Meteo")
            slot    = weather.get("slot","")
            # Show: "70°F  ↓IN 13mph  [SSW(202°) from Open-Meteo · 6:40PM slot]"
            deg_str  = f"({raw_deg}°)" if raw_deg != "" else ""
            eff_icon = {"OUT":"↑","IN":"↓","CROSS":"↔","L-R":"↔","R-L":"↔"}.get(weather["dir"],"")
            w_str = (f"{weather['temp']}°F  {eff_icon}{weather['dir']} {weather['mph']}mph"
                     f"   [desde {raw_dir}{deg_str}  ·  📡{src}]")
            if slot:
                w_str += f"  ⏱{slot}"

        print(f"\n  {'─'*68}")
        hdr = f"  {away:<14} @  {home}"
        if gtime: hdr += f"   │   {gtime}"
        print(hdr)
        _op_a_disp = " [OPENER/RP]" if r.get("_opener_away") else ""
        _op_b_disp = " [OPENER/RP]" if r.get("_opener_home") else ""
        print(f"  SP: {(away_sp[:26]+_op_a_disp):<36}  vs  {home_sp[:26]+_op_b_disp}")
        if fip_src: print(f"  📋 FIP src: {fip_src}")
        print(f"  {'─'*68}")
        print(f"  🎯 MODEL   │  Total: {mtotals:<10}│  Spread: {mspread:<22}│  ML: {lines['ml_pct']}")
        if w_str:
            print(f"  🌤️  WEATHER │  {w_str}")
        ump_hp  = r.get("ump_hp", "")
        ump_f   = r.get("ump_factor", 1.0)
        if ump_hp:
            ump_tend = "avg" if ump_f == 1.0 else (f"+{(ump_f-1)*100:.1f}% runs" if ump_f > 1.0 else f"{(ump_f-1)*100:.1f}% runs")
            print(f"  🧑‍⚖️  UMPIRE   │  HP: {ump_hp}  ({ump_tend})")

        # ── Recent form display ────────────────────────────────────────────
        rf_a = r.get("recent_form_away", {})
        rf_b = r.get("recent_form_home", {})
        if rf_a or rf_b:
            def _form_str(team, rf, form_f, rest_f):
                if not rf:
                    return f"{team}: N/A"
                n      = rf.get("games", 0)
                wp     = rf.get("wp", 0)
                wins   = round(wp * n)
                rd     = rf.get("run_diff", 0)
                rest   = rf.get("rest_days", 1)
                streak = rf.get("streak", 0)
                rd_s   = f"+{rd:.1f}" if rd >= 0 else f"{rd:.1f}"
                rest_s = "B2B" if rest == 0 else (f"3+d rest" if rest >= 3 else f"{rest}d rest")
                form_s = f"{form_f:.3f}" if form_f != 1.0 else "1.000"
                # Mostrar racha si es notable (≥3 seguidas)
                str_s  = ""
                if streak >=  3: str_s = f" 🔥W{streak}"
                elif streak <= -3: str_s = f" ❄️L{abs(streak)}"
                return f"{team}: {wins}-{n-wins} L{n} | AvgRD {rd_s} | {rest_s}{str_s} [{form_s}]"
            fa = r.get("form_a", 1.0); fb = r.get("form_b", 1.0)
            ra = r.get("rest_a", 1.0); rb = r.get("rest_b", 1.0)
            print(f"  📈 FORM     │  {_form_str(away, rf_a, fa, ra)}")
            print(f"  📈 FORM     │  {_form_str(home, rf_b, fb, rb)}")

        # ── Standings display ──────────────────────────────────────────────
        wp_a = r.get("win_pct_away"); wp_b = r.get("win_pct_home")
        std_fa = r.get("standings_a", 1.0); std_fb = r.get("standings_b", 1.0)
        if wp_a is not None and wp_b is not None:
            def _std_str(team, wp, sf):
                arrow = "▲" if sf > 1.005 else ("▼" if sf < 0.995 else "—")
                return f"{team}: {wp*100:.1f}% win [{arrow}{abs(sf-1)*100:.1f}%]"
            print(f"  📊 STANDS   │  {_std_str(away, wp_a, std_fa)}   |   {_std_str(home, wp_b, std_fb)}")

        # ── BP fatigue display ─────────────────────────────────────────────
        bp_fa = r.get("bp_fatigue_away"); bp_fb = r.get("bp_fatigue_home")
        if bp_fa is not None or bp_fb is not None:
            def _bp_str(team, ip):
                if ip is None: return f"{team}: —"
                tag = " ⚠️ CANSADO" if ip > 12 else (" ⚡" if ip > 8 else " ✅")
                return f"{team} BP: {ip:.1f}IP/3d{tag}"
            print(f"  💪 BULLPEN  │  {_bp_str(away, bp_fa)}   |   {_bp_str(home, bp_fb)}")

        # ── Recent SP ERA display ──────────────────────────────────────────
        r_era_a = r.get("recent_era_away"); r_era_b = r.get("recent_era_home")
        if r_era_a is not None or r_era_b is not None:
            def _era_str(sp, era_stats, xfip):
                # era_stats = {"era": float, "ip": float, "n": int} or None
                if era_stats is None: return f"{sp[:15]}: — ERA rec."
                era = era_stats["era"]; ip = era_stats["ip"]; n = era_stats["n"]
                diff = era - (xfip or 4.2)
                tag = f" ▲{diff:+.2f}" if diff > 0.3 else (f" ▼{diff:+.2f}" if diff < -0.3 else "")
                return f"{sp[:15]}: ERA rec. {era:.2f}{tag} ({n}gs/{ip:.0f}IP)"
            print(f"  🔥 SP REC   │  {_era_str(away, r_era_a, r.get('fip_a'))}   |   {_era_str(home, r_era_b, r.get('fip_b'))}")

        # ── SP H2H vs este equipo específico ──────────────────────────────
        h2h_a = r.get("h2h_away_sp"); h2h_b = r.get("h2h_home_sp")
        if h2h_a or h2h_b:
            def _h2h_str(sp, h2h, opp):
                if not h2h: return f"{sp[:12]}: sin H2H vs {opp}"
                diff = h2h["era"] - (r.get("fip_a") or r.get("fip_b") or 4.2)
                tag = f" ▲{diff:+.2f}" if diff > 0.4 else (f" ▼{diff:+.2f}" if diff < -0.4 else "")
                return f"{sp[:12]} vs {opp}: {h2h['era']:.2f} ERA ({h2h['gs']}GS){tag}"
            print(f"  ⚔️  SP H2H   │  {_h2h_str(away_sp, h2h_a, home)}   |   {_h2h_str(home_sp, h2h_b, away)}")

        # ── Platoon split display ──────────────────────────────────────────
        ph_a = r.get("away_sp_hand"); ph_b = r.get("home_sp_hand")
        wrc_a_b = r.get("wrc_a_base"); wrc_b_b = r.get("wrc_b_base")
        wrc_a_v = r.get("wrc_a"); wrc_b_v = r.get("wrc_b")
        if (ph_a or ph_b) and wrc_a_b is not None:
            hand_a = ph_a or "?"
            hand_b = ph_b or "?"
            pa = r.get("platoon_a"); pb = r.get("platoon_b")
            adj_a = f" → {wrc_a_v:.0f} vs {hand_b}HP" if (pa and wrc_a_v != wrc_a_b) else ""
            adj_b = f" → {wrc_b_v:.0f} vs {hand_a}HP" if (pb and wrc_b_v != wrc_b_b) else ""
            print(f"  🏏 PLATOON  │  {away} wRC+ {wrc_a_b:.0f}{adj_a}   |   {home} wRC+ {wrc_b_b:.0f}{adj_b}")

        # ── Model vs Market gap ────────────────────────
        game_books = _get_game_books(odds, away, home)
        if game_books:
            # Best ML available
            best_ml_a = max((bk.get(f"ML_{away}") for bk in game_books.values()
                             if bk.get(f"ML_{away}")), default=None)
            best_ml_h = max((bk.get(f"ML_{home}") for bk in game_books.values()
                             if bk.get(f"ML_{home}")), default=None)
            if best_ml_a and best_ml_h:
                model_ml_a = lines["ml_pct"].split(" / ")[0]
                gap_note   = ""
                try:
                    mkt_a = int(best_ml_a)
                    mod_a = int(model_ml_a.replace("+","").replace("-","-"))
                    gap   = abs(mkt_a - int(model_ml_a.replace("+","")))
                    if model_ml_a.startswith("-"):
                        gap = abs(int(model_ml_a[1:]) - abs(mkt_a)) if mkt_a < 0 else abs(int(model_ml_a[1:]) + mkt_a)
                    if gap > 20:
                        fav_mkt  = away if best_ml_a < best_ml_h else home
                        fav_mod  = away if lines["winA"] > lines["winB"] else home
                        conflict = " ⚠️  FAVORITO DISTINTO" if fav_mkt != fav_mod else f" (gap {gap} pts)"
                        gap_note = f"  🔍 GAP    │  Model {model_ml_a} / Mkt {_fmt_odds(best_ml_a)}{conflict}"
                except: pass
                if gap_note:
                    print(gap_note)

        # ── Market odds table ──────────────────────────
        if game_books:
            mkt_rows = []
            for book, bk in game_books.items():
                ml_a = bk.get(f"ML_{away}"); ml_h = bk.get(f"ML_{home}")
                if ml_a is None and ml_h is None: continue
                ml_s = f"{_fmt_odds(ml_a)} / {_fmt_odds(ml_h)}" if (ml_a and ml_h) else "—"
                tot  = bk.get("Total_Over")
                tot_s = f"O {tot['line']} ({_fmt_odds(tot['odds'])})" if tot else "—"
                spd  = bk.get(f"Spread_{away}")
                if spd:
                    sign = "+" if (spd.get("line") or 0) >= 0 else ""
                    spd_s = f"{away} {sign}{spd['line']} ({_fmt_odds(spd['odds'])})"
                else:
                    spd_s = "—"
                mkt_rows.append([book, ml_s, tot_s, spd_s])
            if mkt_rows:
                print(f"  {'─'*68}")
                tbl = tab(mkt_rows,
                          ["Book", f"ML {away}/{home}", "Total (Over)", f"Spread {away}"],
                          fmt="simple")
                for line in tbl.split("\n"):
                    print(f"  {line}")

    print(f"\n{'═'*72}\n")

    # ── Write mlb_model_lines.json for the web dashboard ──────────────────
    try:
        entries = []
        for r in results:
            away = r["away"]; home = r["home"]
            lines = r["lines"]
            weather = r["weather"]
            w = {}
            if weather.get("dir") == "DOME":
                ri = STADIUM_ROOF.get(home, {})
                w = {"dome": True, "name": ri.get("name","")}
            elif weather.get("temp") is not None:
                w = {
                    "dome":  False,
                    "temp":  weather.get("temp"),
                    "dir":   weather.get("dir",""),
                    "mph":   weather.get("mph",0),
                    "raw_dir": weather.get("raw_dir",""),
                }
            lu_a = bool(r.get("lineup_used_away"))
            lu_h = bool(r.get("lineup_used_home"))
            entries.append({
                "game":     f"{away} @ {home}",
                "away":     away,
                "home":     home,
                "away_sp":  r.get("away_sp","TBD"),
                "home_sp":  r.get("home_sp","TBD"),
                "away_sp_hand": r.get("away_sp_hand",""),
                "home_sp_hand": r.get("home_sp_hand",""),
                "game_time":    r.get("game_time_local",""),
                "lineup_away":  lu_a,
                "lineup_home":  lu_h,
                "lineup_confirmed": lu_a and lu_h,
                "ump_hp":   r.get("ump_hp",""),
                "ump_factor": r.get("ump_factor",1.0),
                "weather":  w,
                "model": {
                    "total":    lines.get("total"),
                    "mTotals":  lines.get("mTotals",""),
                    "mSpread":  lines.get("mSpread","").replace("AWAY", away).replace("HOME", home),
                    "mSpreadTeam": lines.get("mSpreadTeam",""),
                    "ml_pct":   lines.get("ml_pct",""),
                    "wp_away":  lines.get("winA"),
                    "wp_home":  lines.get("winB"),
                    "tA":       lines.get("tA"),
                    "tB":       lines.get("tB"),
                },
                "mkt": {k: v for k, v in (odds.get(f"{away}_{home}") or
                         odds.get(f"{away} vs {home}") or {}).items()}
                       if odds else {},
            })
        out_path = os.path.join(os.path.dirname(__file__), "mlb_model_lines.json")
        existing = {}
        if os.path.exists(out_path):
            with open(out_path) as _f: existing = json.load(_f)
        existing[TARGET_DATE] = entries
        cutoff = (datetime.strptime(TARGET_DATE,"%Y-%m-%d") - timedelta(days=14)).strftime("%Y-%m-%d")
        existing = {k: v for k, v in existing.items() if k >= cutoff}
        with open(out_path, "w") as _f: json.dump(existing, _f, indent=2, ensure_ascii=False)
    except Exception as _e:
        print(f"  ⚠️  mlb_model_lines.json: {_e}")


# ──────────────────────────────────────────────────────
# PUBLISH — push HTMLs a GitHub Pages
# ──────────────────────────────────────────────────────

def cmd_publish(html_paths):
    """
    Publica HTMLs al repo de GitHub Pages (laboywebsite-lgtm/mlb-picks).
    Usa GitHub API directamente — no requiere git CLI ni repo local.
    """
    import base64 as _b64, urllib.request as _ur, urllib.error as _ue
    import shutil, subprocess
    import glob as _glob

    _gh_token = os.environ.get("GITHUB_TOKEN", "") or os.environ.get("LABOY_GITHUB_TOKEN", "")
    _pages_user = os.environ.get("GITHUB_USER", "laboywebsite-lgtm")
    _pages_repo = os.environ.get("GITHUB_REPO", "mlb-picks")

    def _api_push(file_path, repo_filename):
        """Push a single file to GitHub Pages repo via API."""
        from urllib.parse import quote as _quote
        if not _gh_token:
            return False, "GITHUB_TOKEN no configurado"
        with open(file_path, "rb") as _f:
            _content = _b64.b64encode(_f.read()).decode()
        _api_url = f"https://api.github.com/repos/{_pages_user}/{_pages_repo}/contents/{_quote(repo_filename)}"
        _hdrs = {
            "Authorization": f"token {_gh_token}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
            "User-Agent": "laboy-mlb-publish",
        }
        # Get current SHA if file exists
        _sha = ""
        try:
            _req = _ur.Request(_api_url, headers=_hdrs)
            with _ur.urlopen(_req) as _r:
                _sha = json.loads(_r.read()).get("sha", "")
        except Exception:
            pass
        _payload = {"message": f"MLB record {TARGET_DATE}", "content": _content}
        if _sha:
            _payload["sha"] = _sha
        try:
            _req2 = _ur.Request(_api_url, data=json.dumps(_payload).encode(), headers=_hdrs, method="PUT")
            with _ur.urlopen(_req2) as _r2:
                _res = json.loads(_r2.read())
            return True, _res["commit"]["sha"][:8]
        except Exception as _e:
            return False, str(_e)

    # ── Fallback: git clone approach (local dev) ─────────────────────────────
    repo = GITHUB_PAGES_REPO
    _use_api = not repo or not os.path.isdir(os.path.dirname(repo) if repo else "")

    if not _use_api:
        _clone_url = (f"https://{_gh_token}@github.com/{_pages_user}/{_pages_repo}"
                      if _gh_token else f"https://github.com/{_pages_user}/{_pages_repo}")
        if not os.path.isdir(repo):
            print(f"\n  📥 Repo mlb-picks no encontrado. Clonando...")
            os.makedirs(os.path.dirname(repo), exist_ok=True)
            _r = subprocess.run(["git", "clone", _clone_url, repo], capture_output=True, text=True)
            if _r.returncode != 0:
                _use_api = True  # fallback to API
            else:
                print(f"  ✅ Repo clonado.\n")

    def _api_update_dashboard(new_files):
        """
        Actualiza SOLO manifest.json en GitHub — nunca toca el dashboard HTML.
        El dashboard (PWA) lee manifest.json para mostrar picks, lines y records.
        """
        from urllib.parse import quote as _quote
        import re as _re
        if not _gh_token:
            return
        _base  = f"https://api.github.com/repos/{_pages_user}/{_pages_repo}"
        _hdrs2 = {"Authorization": f"token {_gh_token}", "Accept": "application/vnd.github.v3+json",
                  "Content-Type": "application/json", "User-Agent": "laboy-mlb-publish"}

        try:
            # Leer manifest.json actual
            _murl = f"{_base}/contents/manifest.json"
            _req_m = _ur.Request(_murl, headers=_hdrs2)
            try:
                with _ur.urlopen(_req_m, timeout=15) as _r_m:
                    _mdata = json.loads(_r_m.read())
                import base64 as _b64m
                _manifest = json.loads(_b64m.b64decode(_mdata["content"]).decode("utf-8"))
                _manifest_sha = _mdata["sha"]
            except Exception:
                _manifest = {"sport": "MLB", "base_url": GITHUB_PAGES_URL, "files": []}
                _manifest_sha = ""

            # Agregar las nuevas entradas al manifest (evitar duplicados por nombre)
            existing_names = {f["name"] for f in _manifest.get("files", [])}
            added = 0
            for fname in (new_files or []):
                if fname in existing_names:
                    continue
                enc   = _quote(fname)
                dm    = _re.search(r"(\d{4}-\d{2}-\d{2})", fname)
                fdate = dm.group(1) if dm else TARGET_DATE
                ftype = ("record" if ("Record" in fname or "Model Card" in fname)
                         else "picks" if "Picks" in fname
                         else "lines" if "Lines" in fname
                         else "debug")
                fsub  = ("my_record" if "Record Card" in fname
                         else "model_record" if "Model Card" in fname
                         else "")
                _sess_m  = _re.search(r"(?:Debug|Picks) (DAY|NIGHT|PM) ", fname)
                _slate_m = _re.search(r" S(\d+)[- ]", fname)
                _manifest.setdefault("files", []).insert(0, {
                    "name":    fname,
                    "url":     f"{GITHUB_PAGES_URL}/{enc}",
                    "type":    ftype,
                    "subtype": fsub,
                    "date":    fdate,
                    "session": _sess_m.group(1).lower() if _sess_m else "",
                    "slate":   f"S{_slate_m.group(1)}" if _slate_m else "",
                })
                added += 1

            if added == 0:
                print(f"  ℹ️  manifest.json ya tiene estos archivos.")
                return

            # Pushear manifest.json actualizado
            import base64 as _b64m2
            _new_content = _b64m2.b64encode(
                json.dumps(_manifest, indent=2, ensure_ascii=False).encode("utf-8")
            ).decode()
            _payload = {"message": f"auto: manifest update ({', '.join(new_files or [])})",
                        "content": _new_content}
            if _manifest_sha:
                _payload["sha"] = _manifest_sha
            _req_put = _ur.Request(_murl, data=json.dumps(_payload).encode(),
                                   headers=_hdrs2, method="PUT")
            with _ur.urlopen(_req_put, timeout=20) as _r_put:
                _res = json.loads(_r_put.read())
            print(f"  📋 manifest.json actualizado → {_res['commit']['sha'][:8]}")

        except Exception as _e_dash:
            print(f"  ⚠️  manifest.json update falló: {_e_dash}")

    def _api_push_content(repo_filename, raw_bytes):
        from urllib.parse import quote as _quote
        _api_url = f"https://api.github.com/repos/{_pages_user}/{_pages_repo}/contents/{_quote(repo_filename)}"
        _hdrs3 = {"Authorization": f"token {_gh_token}", "Accept": "application/vnd.github.v3+json",
                  "Content-Type": "application/json", "User-Agent": "laboy-mlb-publish"}
        import base64 as _b64_inner
        _content = _b64_inner.b64encode(raw_bytes).decode()
        _sha = ""
        try:
            _req = _ur.Request(_api_url, headers=_hdrs3)
            with _ur.urlopen(_req) as _r: _sha = json.loads(_r.read()).get("sha", "")
        except Exception: pass
        _payload = {"message": f"deploy: {repo_filename}", "content": _content}
        if _sha: _payload["sha"] = _sha
        try:
            _req2 = _ur.Request(_api_url, data=json.dumps(_payload).encode(), headers=_hdrs3, method="PUT")
            with _ur.urlopen(_req2) as _r2: _res = json.loads(_r2.read())
            return True, _res["commit"]["sha"][:8]
        except Exception as _e:
            return False, str(_e)

    # ── API publish ───────────────────────────────────────────────────────────
    if _use_api:
        published = []
        for hp in (html_paths or []):
            if hp and os.path.isfile(hp):
                fname = os.path.basename(hp)
                ok, info = _api_push(hp, fname)
                if ok:
                    print(f"  ✅ {fname} → GitHub Pages ({info})")
                    published.append(fname)
                else:
                    print(f"  ⚠️  No se pudo publicar {fname}: {info}")
        if published:
            print(f"\n  🌐 {GITHUB_PAGES_URL}/{published[0]}")
            _api_update_dashboard(published)
        return

    copied = []
    for hp in (html_paths or []):
        if hp and os.path.isfile(hp):
            dest = os.path.join(repo, os.path.basename(hp))
            shutil.copy2(hp, dest)
            copied.append(os.path.basename(hp))

    # Copiar imágenes de picks personales (JPG/PNG)
    for img_pat in ["Laboy Pick *.jpg", "Laboy Pick *.png"]:
        for img in _glob.glob(os.path.join(SCRIPT_DIR, img_pat)):
            dest = os.path.join(repo, os.path.basename(img))
            shutil.copy2(img, dest)

    if not copied:
        print("\n  ⚠️  No hay HTMLs para publicar. Corre con --export-lines primero.")
        return

    # ── Regenerar index.html ────────────────────────────────────
    _publish_update_index(repo)

    # ── git add / commit / push ─────────────────────────────────
    def _git(args, timeout=60):
        try:
            r = subprocess.run(["git", "-C", repo] + args,
                               capture_output=True, text=True, timeout=timeout)
            return r.returncode, r.stdout.strip(), r.stderr.strip()
        except subprocess.TimeoutExpired:
            return 1, "", f"git {args[0]} timeout ({timeout}s)"

    _git(["add", "--all"])
    msg = f"MLB {TARGET_DATE} — {', '.join(copied)}"
    code, out, err = _git(["commit", "-m", msg])
    if code != 0 and "nothing to commit" in (out + err):
        print("\n  Sin cambios nuevos en el repo (archivos identicos).")
    elif code != 0:
        print(f"\n  git commit fallo: {err or out}")
        return

    # Pull remoto antes de push para evitar "rejected — fetch first"
    print("  git pull --rebase...")
    code, out, err = _git(["pull", "--rebase"], timeout=45)
    if code != 0:
        print(f"\n  git pull fallo: {err or out}")
        print(f"     Resuelve conflictos manualmente y vuelve a intentar.")
        return

    print("  git push...")
    code, out, err = _git(["push"], timeout=60)
    if code != 0:
        print(f"\n  git push fallo: {err or out}")
        print(f"     Verifica que tienes acceso SSH/HTTPS configurado.")
        return

    print(f"\n  ✅ Publicado en GitHub Pages!")
    from urllib.parse import quote as _url_quote
    for fname in copied:
        encoded = _url_quote(fname)
        print(f"  🌐 {GITHUB_PAGES_URL}/{encoded}")
    print(f"\n  📱 Dashboard privado:")
    print(f"     {GITHUB_PAGES_URL}/dashboard-{DASHBOARD_TOKEN}.html")
    print(f"     (Guarda este enlace en Safari / iPhone)")
    print()


def _publish_update_index(repo):
    """
    Regenera index.html (en blanco) y dashboard-{DASHBOARD_TOKEN}.html en el repo.
    """
    import glob as _glob
    from urllib.parse import quote as _url_quote
    import json as _json

    # ── .nojekyll ──────────────────────────────────────────────────────────────
    with open(os.path.join(repo, ".nojekyll"), "w") as f:
        f.write("")

    # ── index.html — completamente en blanco ───────────────────────────────────
    blank = (
        "<!DOCTYPE html>\n"
        '<html lang="es"><head><meta charset="UTF-8">'
        "<title>Laboy Picks</title></head><body></body></html>\n"
    )
    with open(os.path.join(repo, "index.html"), "w", encoding="utf-8") as f:
        f.write(blank)

    all_html = sorted(
        _glob.glob(os.path.join(repo, "Laboy *.html")),
        key=os.path.getmtime, reverse=True,
    )
    all_imgs = sorted(
        _glob.glob(os.path.join(repo, "Laboy Pick *.jpg")) +
        _glob.glob(os.path.join(repo, "Laboy Pick *.png")),
        key=os.path.getmtime, reverse=True,
    )

    # Cargar log de picks para obtener nombres de juego
    _pick_game = {}   # {pick_id: game_name}
    try:
        with open(LOG_FILE, encoding="utf-8") as _lf:
            _log_data = _json.load(_lf)
        _entries = _log_data if isinstance(_log_data, list) else _log_data.get("picks", [])
        for _e in _entries:
            _pick_game[int(_e["id"])] = _e.get("game", "")
    except Exception:
        pass

    # Separar fotos de hoy vs archivo (para manifest.json)
    # Siempre usar la fecha real de hoy, no TARGET_DATE (que puede ser una fecha pasada)
    _real_today  = date.today().strftime("%Y-%m-%d")
    today_imgs   = []
    archive_imgs = []
    for ip in all_imgs:
        base     = os.path.basename(ip)
        enc      = _url_quote(base)
        url      = f"{GITHUB_PAGES_URL}/{enc}"
        dm       = re.search(r"(\d{4}-\d{2}-\d{2})", base)
        img_date = dm.group(1) if dm else ""
        # Extraer número de pick (#N) para buscar juego en el log
        nm       = re.search(r"#(\d+)", base)
        game     = _pick_game.get(int(nm.group(1)), "") if nm else ""
        if img_date == _real_today:
            today_imgs.append({"base": base, "url": url, "game": game})
        else:
            archive_imgs.append({"base": base, "url": url, "date": img_date, "game": game})

    # ── dashboard-{TOKEN}.html — lista privada de links ────────────────────────
    def _fname_icon(fname):
        if "Lines" in fname:       return "📊"
        if "Model Card" in fname:  return "🏆"
        if "Record Card" in fname: return "📈"
        return "🎯"

    def _fname_label(fname):
        base = os.path.basename(fname)
        name = base[:-5]
        if len(name) > 8 and name[-8] == "-":
            name = name[:-8]
        name = name.replace("Laboy ", "", 1)
        return f"{_fname_icon(base)} {name}"

    # Separar picks por sesión: DAY, PM, NIGHT, y full-day (sin sesión)
    picks_day   = [hp for hp in all_html if re.search(r"Picks DAY ",   os.path.basename(hp))]
    picks_pm    = [hp for hp in all_html if re.search(r"Picks PM ",    os.path.basename(hp))]
    picks_night = [hp for hp in all_html if re.search(r"Picks NIGHT ", os.path.basename(hp))]
    picks_full  = [hp for hp in all_html
                   if "Picks" in os.path.basename(hp)
                   and not re.search(r"Picks (DAY|PM|NIGHT) ", os.path.basename(hp))]
    lines_files = [hp for hp in all_html if "Lines" in os.path.basename(hp)]
    rec_files   = [hp for hp in all_html if "Record" in os.path.basename(hp)
                   or "Model Card" in os.path.basename(hp)]
    dbg_day   = [hp for hp in all_html if re.search(r"Debug DAY ",   os.path.basename(hp))]
    dbg_pm    = [hp for hp in all_html if re.search(r"Debug PM ",    os.path.basename(hp))]
    dbg_night = [hp for hp in all_html if re.search(r"Debug NIGHT ", os.path.basename(hp))]
    dbg_full  = [hp for hp in all_html
                 if "Debug" in os.path.basename(hp)
                 and not re.search(r"Debug (DAY|PM|NIGHT) ", os.path.basename(hp))]

    _MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def _file_links(flist, label, emoji=""):
        if not flist:
            return ""
        items = ""
        for hp in flist[:12]:
            name   = os.path.basename(hp)
            enc    = _url_quote(name)
            dm     = re.search(r"(\d{4})-(\d{2})-(\d{2})", name)
            sm     = re.search(r" S(\d+)[- ]", name)
            if dm:
                _mon   = _MONTH_ABBR[int(dm.group(2)) - 1]
                dstr_l = f"{_mon} {dm.group(3)}"
                if sm:
                    dstr_l += f" S{sm.group(1)}"
            else:
                dstr_l = name
            items += (f'<li><a href="{GITHUB_PAGES_URL}/{enc}" target="_blank">'
                      f'{dstr_l}</a></li>\n')
        lbl = f"{emoji} {label}".strip() if emoji else label
        return f"<h3>{lbl}</h3><ul>{items}</ul>"

    dash_html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>⚾ MLB Picks — Dashboard</title>
  <style>
    body{{font-family:system-ui,sans-serif;background:#0f172a;color:#e2e8f0;
          padding:24px;max-width:600px;margin:0 auto}}
    h1{{color:#f8fafc;font-size:1.4rem;border-bottom:1px solid #334155;
        padding-bottom:10px;margin-bottom:16px}}
    h3{{color:#94a3b8;font-size:0.85rem;text-transform:uppercase;
        letter-spacing:.08em;margin-top:20px;margin-bottom:6px}}
    a{{color:#60a5fa;text-decoration:none}}
    a:hover{{text-decoration:underline}}
    ul{{list-style:none;padding:0;margin:0}}
    li{{padding:6px 0;border-bottom:1px solid #1e293b}}
    .ts{{font-size:0.75rem;color:#475569;margin-top:20px}}
    .section-sep{{border-top:1px solid #1e293b;margin:18px 0 0}}
  </style>
</head>
<body>
  <h1>⚾ MLB Picks — Dashboard</h1>
  {_file_links(picks_day,   "Picks Sesión Día",   "☀️")}
  {_file_links(picks_pm,    "Picks Sesión Tarde", "🌤️")}
  {_file_links(picks_night, "Picks Sesión Noche", "🌙")}
  {('<div class="section-sep"></div>' + _file_links(picks_full, "Picks Completos")) if picks_full else ""}
  {_file_links(dbg_day,   "Reports · Día",   "☀️")}
  {_file_links(dbg_pm,    "Reports · Tarde", "🌤️")}
  {_file_links(dbg_night, "Reports · Noche", "🌙")}
  {('<div class="section-sep"></div>' + _file_links(dbg_full, "Reports Completos", "🔬")) if dbg_full else ""}
  {_file_links(lines_files, "Lines del Modelo",   "📊")}
  {_file_links(rec_files,   "Model Record",       "📈")}
  <p class="ts">Actualizado: {datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
</body>
</html>"""

    dash_path = os.path.join(repo, f"dashboard-{DASHBOARD_TOKEN}.html")
    with open(dash_path, "w", encoding="utf-8") as f:
        f.write(dash_html)
    print(f"  📋 Dashboard actualizado: dashboard-{DASHBOARD_TOKEN}.html")

    # ── manifest.json ─────────────────────────────────────────────────────────
    manifest = {"sport": "MLB", "base_url": GITHUB_PAGES_URL, "files": []}
    for hp in all_html[:20]:
        base = os.path.basename(hp)
        if base.startswith("dashboard-") or base.startswith("archive-picks-"):
            continue
        enc   = _url_quote(base)
        ftype = ("debug" if "Debug" in base
                 else "picks" if "Picks" in base
                 else "lines" if "Lines" in base
                 else "record")
        fsubtype = ("model_record" if ("Model Card" in base or "Model Record" in base) else "my_record") if ftype == "record" else ""
        dm        = re.search(r"(\d{4}-\d{2}-\d{2})", base)
        _sess_m   = re.search(r"(?:Debug|Picks) (DAY|NIGHT|PM) ", base)
        _slate_m  = re.search(r" S(\d+)[- ]", base)
        _fsession = _sess_m.group(1).lower() if _sess_m else "full"
        _fslate   = f"S{_slate_m.group(1)}" if _slate_m else ""
        manifest["files"].append({
            "name": base, "url": f"{GITHUB_PAGES_URL}/{enc}",
            "type": ftype, "subtype": fsubtype,
            "date":    dm.group(1) if dm else "",
            "session": _fsession if ftype in ("debug", "picks") else "",
            "slate":   _fslate   if ftype in ("debug", "picks") else "",
        })
    for img in today_imgs:
        manifest["files"].append({
            "name": img["base"], "url": img["url"],
            "type": "mypick_img", "subtype": "today",
            "date": TARGET_DATE, "game": img.get("game", ""),
        })
    for img in archive_imgs[:30]:
        manifest["files"].append({
            "name": img["base"], "url": img["url"],
            "type": "mypick_img", "subtype": "archive",
            "date": img["date"], "game": img.get("game", ""),
        })

    with open(os.path.join(repo, "manifest.json"), "w", encoding="utf-8") as f:
        _json.dump(manifest, f, indent=2, ensure_ascii=False)


# ──────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────
# WEATHER — verificación de clima en vivo para todos los juegos de hoy
# ──────────────────────────────────────────────────────

def cmd_weather(games):
    """
    --weather  →  muestra clima en vivo para todos los juegos de hoy.
    Muestra fuente, temperatura cruda, dirección del viento en grados y
    compass, el efecto IN/OUT para el parque, y la hora exacta usada.
    Útil para contrastar con Prop Finder u otras fuentes externas.
    """
    print(f"\n{'═'*70}")
    print(f"  🌤️  CLIMA EN VIVO — {TARGET_DATE}   (fuente: Open-Meteo / wttr.in)")
    print(f"{'═'*70}")
    tz_labels = {-4:"ET",-5:"CT",-6:"MT",-7:"PT"}

    if not games:
        print("  ❌ No hay juegos para mostrar. Corre con --lines o --picks primero.")
        return

    COL_W = 26
    hdr = (f"  {'JUEGO':<{COL_W}}  {'HORA':>8}  {'TEMP':>5}  "
           f"{'VIENTO':>12}  {'EFECTO':>6}  {'SLOT USADO'}")
    print(hdr)
    print(f"  {'─'*68}")

    for g in games:
        away   = g.get("away","?")
        home   = g.get("home","?")
        game_utc = g.get("game_time_utc","")
        gtime  = g.get("game_time_local","") or format_game_time(game_utc, home)

        # Fetch fresh weather
        w = get_weather(home, game_time_utc=game_utc)

        game_str = f"{away} @ {home}"
        if home in DOME_TEAMS:
            print(f"  {game_str:<{COL_W}}  {gtime:>8}  {'—':>5}  {'DOME':>12}  {'DOME':>6}")
            continue

        temp_str   = f"{w['temp']}°F" if w.get("temp") is not None else "N/A"
        raw_deg    = w.get("raw_deg", "?")
        raw_dir    = w.get("raw_dir","?")
        mph        = w.get("mph","?")
        park_dir   = w.get("dir","?")
        slot       = w.get("slot","?")
        source     = w.get("source","?")

        # Format wind: "SSW(202°) → IN  · 13 mph"
        wind_raw   = f"{raw_dir}({raw_deg}°)"
        wind_str   = f"{wind_raw} {mph}mph"

        # Color-code effect
        eff_icons = {"OUT":"↑OUT","IN":"↓IN","CROSS":"↔CROSS","L-R":"↔L-R","R-L":"↔R-L"}
        eff_str = eff_icons.get(park_dir, park_dir or "?")

        print(f"  {game_str:<{COL_W}}  {gtime:>8}  {temp_str:>5}  {wind_str:>16}  {eff_str:>7}  [{slot}]  📡{source}")

    print(f"\n  {'─'*68}")
    print(f"  ℹ️  Temp: Open-Meteo ECMWF model forecast.")
    print(f"  ℹ️  IN = viento soplando hacia home plate (→ menos carreras).")
    print(f"  ℹ️  OUT = viento soplando hacia CF (→ más carreras, activa el parque).")
    print(f"  ℹ️  Prop Finder usa The Weather Channel/NWS — puede diferir ±5°F y ±1 hora.")
    print(f"  ⚠️  Si ves discrepancias > 5°F o dirección opuesta, contrasta con weather.com\n")


# DEBUG GAME — diagnóstico de inputs del modelo
# ──────────────────────────────────────────────────────

def cmd_debug_game(results, odds):
    """
    Muestra todos los inputs del modelo para un juego específico.
    Uso: python3 mlb.py --debug-game AWAY HOME
    Ejemplo: python3 mlb.py --debug-game PIRATES CUBS
    """
    try:
        di    = sys.argv.index("--debug-game")
        d_a   = sys.argv[di+1].upper()
        d_h   = sys.argv[di+2].upper()
    except (ValueError, IndexError):
        print("  ❌ Uso: python3 mlb.py --debug-game AWAY HOME")
        print("     Ejemplo: python3 mlb.py --debug-game PIRATES CUBS")
        return

    target = next((r for r in results
                   if d_a in r["away"] and d_h in r["home"]
                   or d_h in r["away"] and d_a in r["home"]), None)
    if not target:
        available = ", ".join(f"{r['away']} @ {r['home']}" for r in results)
        print(f"  ❌ No se encontró {d_a} vs {d_h}")
        print(f"     Disponibles: {available}")
        return

    away, home = target["away"], target["home"]
    lines      = target["lines"]
    weather    = target["weather"]
    gkey       = f"{away} vs {home}"

    print(f"\n{'═'*72}")
    print(f"  🔍 DEBUG: {away} @ {home}   [{target.get('game_time_local','')}]")
    print(f"{'═'*72}")

    # SP info
    print(f"\n  STARTING PITCHERS:")
    print(f"  Away: {target['away_sp']}")
    print(f"    FIP usado en modelo: {target.get('fip_a','N/A')}")
    print(f"    Fuente: {target.get('fip_source_a','FIP N/A (FanGraphs no disponible)')}")
    print(f"  Home: {target['home_sp']}")
    print(f"    FIP usado en modelo: {target.get('fip_b','N/A')}")
    print(f"    Fuente: {target.get('fip_source_b','FIP N/A (FanGraphs no disponible)')}")

    if not FG_FIP_AVAILABLE:
        print(f"\n  ⚠️  FanGraphs NO disponible → SP FIP basado en BP como proxy")
        print(f"     Para FIP real de SP: export FANGRAPHS_COOKIE='tu_cookie'")
        print(f"     el modelo puede favorecer al equipo equivocado.")
        print(f"     Solución: export FANGRAPHS_COOKIE='tu_cookie'")

    # Team data
    print(f"\n  TEAM DATA:")
    print(f"  {away}: wRC+ = {target.get('wrc_a','—')}  │  BP xFIP = {target.get('bp_a','—')}")
    print(f"  {home}: wRC+ = {target.get('wrc_b','—')}  │  BP xFIP = {target.get('bp_b','—')}")

    # xFIP total
    print(f"\n  xFIP TOTAL (SP 60% + BP 40%):")
    print(f"  {away} vs {home}'s pitching: {target.get('xfip_a','—')}")
    print(f"  {home} vs {away}'s pitching: {target.get('xfip_b','—')}")

    # Park & weather
    pf = calc_pf_combined(home, weather["dir"], weather["mph"], weather["temp"], weather.get("humidity", 50))
    print(f"\n  PARK & WEATHER:")
    print(f"  Park Factor base ({home}): {PARK_FACTORS.get(home,1.0)}")
    _w_src = weather.get("source","?")
    _w_raw = weather.get("raw_dir","—")
    _w_deg = weather.get("raw_deg","")
    _w_deg_s = f" ({_w_deg}°)" if _w_deg else ""
    print(f"  Weather [{_w_src}]: {weather['temp']}°F  Wind {weather['dir']} {weather['mph']}mph "
          f"(desde: {_w_raw}{_w_deg_s})")
    print(f"  PF combinado (park × temp × wind): {pf}")

    # Expected runs
    print(f"\n  PROYECCIÓN DE CARRERAS:")
    print(f"  {away}: {lines['tA']} runs esperados")
    print(f"  {home}: {lines['tB']} runs esperados")
    print(f"  Total modelo: {lines['total']}  →  Línea: {lines['mTotals']}")

    # Win %
    print(f"\n  PROBABILIDADES:")
    print(f"  {away}: {lines['winA']}%  →  ML modelo: {lines['ml_pct'].split(' / ')[0]}")
    print(f"  {home}: {lines['winB']}%  →  ML modelo: {lines['ml_pct'].split(' / ')[1]}")

    # Market comparison
    game_books = _get_game_books(odds, away, home)
    if game_books:
        print(f"\n  MERCADO vs MODELO:")
        rows = []
        for book, bk in game_books.items():
            ml_a = bk.get(f"ML_{away}"); ml_h = bk.get(f"ML_{home}")
            if ml_a and ml_h:
                rows.append([book, _fmt_odds(ml_a), _fmt_odds(ml_h)])
        rows.append(["── MODELO ──", lines['ml_pct'].split(' / ')[0],
                     lines['ml_pct'].split(' / ')[1]])
        print(tab(rows, ["Book", f"ML {away}", f"ML {home}"], fmt="simple"))

        # Flag large gaps
        best_ml_a = max((bk.get(f"ML_{away}") for bk in game_books.values()
                         if bk.get(f"ML_{away}")), default=None)
        best_ml_h = max((bk.get(f"ML_{home}") for bk in game_books.values()
                         if bk.get(f"ML_{home}")), default=None)
        if best_ml_a and best_ml_h:
            fav_mkt = away if best_ml_a < best_ml_h else home
            fav_mod = away if lines["winA"] > lines["winB"] else home
            if fav_mkt != fav_mod:
                print(f"\n  ⚠️  ALERTA: Modelo favorece {fav_mod} pero mercado favorece {fav_mkt}")
                print(f"  Posibles causas:")
                print(f"  1. FIP del SP — FanGraphs no disponible (usando BP proxy)")
                print(f"  2. Lesiones o cambios de lineup no reflejados")
                print(f"  3. Diferencia de bullpen no capturada")
                if not FG_FIP_AVAILABLE:
                    print(f"  → Resolución: activa FanGraphs cookie para FIP actualizado")
    print()


# ──────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────
# EV PICKS RECOMMENDATIONS
# ──────────────────────────────────────────────────────

def _parse_time_sort(t):
    """Convierte '7:05 PM ET' a int para ordenar cronológicamente."""
    if not t: return 9999
    try:
        parts = t.split()
        h, m  = map(int, parts[0].split(':'))
        ap    = parts[1].upper() if len(parts) > 1 else 'PM'
        if ap == 'PM' and h != 12: h += 12
        if ap == 'AM' and h == 12: h  = 0
        return h * 100 + m
    except: return 9999

def _am_to_prob(o):
    o = int(o)
    return abs(o)/(abs(o)+100) if o < 0 else 100/(o+100)

def _am_to_payout(o):
    o = int(o)
    return o/100 if o > 0 else 100/abs(o)

def _prob_to_american(p):
    """Win probability (0–1) → American odds int."""
    p = max(0.01, min(0.99, float(p)))
    if p >= 0.5:
        return int(round(-p / (1 - p) * 100))
    else:
        return int(round((1 - p) / p * 100))


def show_picks(results, odds):
    """
    Muestra picks EV+ del día usando _compute_picks() como único motor.
    Aplica todos los thresholds calibrados y el daily cap.
    """
    dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS — RECOMENDACIONES EV+   {dt.strftime('%a %b %d %Y').upper()}")
    if not FG_FIP_AVAILABLE:
        print(f"  ⚠️  FIP no disponible — picks basados en BP xFIP como proxy")
    print(f"{'═'*72}")

    # ── Motor único: _compute_picks con thresholds calibrados + daily cap ──
    picks = _compute_picks(results, odds)

    # Leer contadores expuestos por _compute_picks
    _skipped_no_odds = getattr(_compute_picks, "_skipped_no_odds", 0)
    _skipped_started = getattr(_compute_picks, "_skipped_started", 0)
    _skipped_lineup  = getattr(_compute_picks, "_skipped_lineup",  0)
    _games_evaluated = getattr(_compute_picks, "_games_evaluated", 0)
    _near_misses     = getattr(_compute_picks, "_near_misses",     [])

    confirmed_picks = [p for p in picks if p.get("lineup_confirmed")]
    watchlist_picks = [p for p in picks if not p.get("lineup_confirmed")]

    if not picks:
        print("\n  No se encontraron picks con edge suficiente hoy.")
        print(f"  📊 Resumen: {len(results)} juegos en schedule  →  "
              f"{_skipped_no_odds} sin odds  ·  "
              f"{_skipped_started} ya iniciados  ·  "
              f"{_skipped_lineup} sin lineup  ·  "
              f"{_games_evaluated} evaluados")
        if _skipped_no_odds:
            # Mostrar qué juegos no tienen odds para diagnosticar el mismatch
            _missing_odds = [f"{r['away']} @ {r['home']}" for r in results
                             if not _get_game_books(odds, r["away"], r["home"])]
            print(f"  ⚠️  Sin odds: {', '.join(_missing_odds[:8])}")
        if _skipped_started:
            print(f"  ⏱  {_skipped_started} juego(s) ya iniciado(s) — filtrados (odds en vivo no válidos).")
        print(f"  📏 Thresholds: ML ≥5.5% edge / EV ≥5%  ·  Totals diff ≥ Over 1.50 / Under 1.25 / EV ≥8%")
        print(f"  🐶 Dog shrinkage: fav-lig +10% mkt  |  dog+101 +20% mkt / EV≥8%  |  dog+150 +30% mkt / EV≥9%")
        if _near_misses:
            # Mostrar los mejores near-misses (max 5, ordenados por EV)
            _nm_sorted = sorted(_near_misses, key=lambda x: -x["ev"])[:5]
            print(f"\n  📋 Candidatos que no alcanzaron el threshold ({len(_near_misses)} total):")
            for nm in _nm_sorted:
                print(f"     · {nm['game']:<30}  {nm['pick']:<14}  {nm['odds']:<7}  "
                      f"EV:{nm['ev']*100:+.1f}%  edge:{nm['edge']:+.1f}%  ← {nm['gap']}")
        else:
            print("  (Ningún candidato estuvo cerca del threshold — líneas bien calibradas hoy)")
        print("  (Tip: si no ves odds del mercado, verifica ODDS_API_KEY)")
    else:
        def _print_picks_table(pk_list, label):
            rows = [[p["game"], p["time"],
                     ("🔒 " if p.get("candado") else ("⛈️  " if p.get("tormenta") else "")) + p["pick"],
                     p["odds"],
                     p.get("modelo", p.get("modelo%","")),
                     p.get("mercado", p.get("mercado%","")),
                     p["edge"], p["ev"],
                     p.get("kelly", "—")] for p in pk_list]
            print(f"\n  {label}\n")
            print(tab(rows, ["Juego","Hora","Pick","Odds","Modelo","Mercado","Edge","EV","Kelly%"],
                      fmt="rounded_outline"))

        if confirmed_picks:
            _print_picks_table(confirmed_picks, f"🎯 {len(confirmed_picks)} PICK(S) — Lineup Confirmado ✅")
        else:
            print("\n  ⚠️  No hay picks con lineup confirmado aún.")

        if watchlist_picks:
            print(f"\n  {'─'*68}")
            _print_picks_table(watchlist_picks, f"⏳ {len(watchlist_picks)} WATCH LIST — Lineup Pendiente (NO apostar aún)")
            print(f"  ⚠️  Estos picks pueden cambiar cuando salgan los lineups.")
            print(f"     Vuelve a correr --picks cuando estén confirmados.")

        # Triple Lock / Perfect Storm solo para picks confirmados
        all_display = confirmed_picks + watchlist_picks
        tc_picks = [p for p in all_display if p.get("candado")]
        if tc_picks:
            print(f"\n  🔒 TRIPLE LOCK — Under de alta probabilidad:")
            for p in tc_picks:
                fip_ok  = "✅" if p.get("_tc_fip")  else "❌"
                pf_ok   = "✅" if p.get("_tc_pf")   else "❌"
                wrc_ok  = "✅" if p.get("_tc_wrc")  else "❌"
                lu_icon = "📋" if p.get("lineup_confirmed") else "⚠️ "
                print(f"     {lu_icon} {p['game']:<28}  FIP<3.80:{fip_ok}  PF<98:{pf_ok}  wRC+<100:{wrc_ok}")
        tp_picks = [p for p in all_display if p.get("tormenta")]
        if tp_picks:
            print(f"\n  ⛈️  PERFECT STORM — Over de alta probabilidad:")
            for p in tp_picks:
                fip_ok  = "✅" if p.get("_tp_fip")  else "❌"
                pf_ok   = "✅" if p.get("_tp_pf")   else "❌"
                wrc_ok  = "✅" if p.get("_tp_wrc")  else "❌"
                lu_icon = "📋" if p.get("lineup_confirmed") else "⚠️ "
                print(f"     {lu_icon} {p['game']:<28}  FIP>4.20:{fip_ok}  PF>102:{pf_ok}  wRC+>105:{wrc_ok}")

        print(f"\n  Edge   = modelo% − mercado% (sin vig)")
        print(f"  EV     = Expected Value por unidad apostada")
        print(f"  Kelly% = % recomendado del bankroll (Quarter Kelly, conservador)")

    print(f"\n{'═'*72}\n")

    # ── Persistir picks en MODEL_PICKS_FILE ──────────────────────────────────
    # Siempre guardar los picks del run actual (confirmados primero, watchlist
    # como fallback) para que el dashboard pueda leerlos vía mlb_model_picks.json.
    confirmed_picks = [p for p in picks if p.get("lineup_confirmed")]
    watchlist_picks = [p for p in picks if not p.get("lineup_confirmed")]
    if confirmed_picks:
        _model_picks_save_today(confirmed_picks)
    elif watchlist_picks:
        # Sin lineups confirmados aún — guardar watchlist para mostrar en dashboard
        _model_picks_save_today(watchlist_picks)
    # También guardar snapshot de debug para --export-debug
    _save_debug_state(picks)

    # ── Predictions log: modelo vs mercado vs real ────────────────────────
    # Guarda model_total y market_total por juego para calibración futura.
    # actual_runs se llena cuando el usuario hace grade en el dashboard.
    _pred_entries = []
    for _p in picks:
        if _p.get("_type") == "TOT" and _p.get("market_total") is not None:
            _pred_entries.append({
                "date":         TARGET_DATE,
                "game":         _p.get("game",""),
                "model_total":  round(float(_p.get("modelo","0").replace("Proj ","") or 0), 2),
                "market_total": _p.get("market_total"),
                "pick":         _p.get("pick",""),
                "p_over":       _p.get("p_over"),
                "p_under":      _p.get("p_under"),
                "actual_runs":  None,   # se llena al hacer grade
            })
    if _pred_entries:
        _save_predictions_log(_pred_entries)

    return picks


# ──────────────────────────────────────────────────────
# EXPORT TXT para comunidad / dubclub
# ──────────────────────────────────────────────────────

def export_lines(results, odds={}):
    dt      = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr    = dt.strftime("%A, %B %d %Y").upper()
    fname   = f"Laboy MLB {TARGET_DATE}.txt"
    fpath   = os.path.join(SCRIPT_DIR, fname)
    W = 60

    def bl(text=""):  return f"│  {str(text):<{W-4}}│"
    def div(l="├",r="┤"): return f"{l}{'─'*(W-2)}{r}"

    out = ["═"*W, f"  LABOY MLB DATA MODEL — {dstr}"[:W], "═"*W, ""]
    for r in results:
        away, home = r["away"], r["home"]
        gkey   = f"{away} vs {home}"
        ld     = r["lines"]
        wth    = r["weather"]
        gtime  = r.get("game_time_local","")

        mspread = ld["mSpread"].replace("AWAY",away).replace("HOME",home)
        w_s = "DOME" if wth["dir"]=="DOME" \
              else (f"{wth['temp']}°F · Wind {wth['dir']} {wth['mph']}mph"
                    if wth["temp"] is not None and wth["dir"] else "Weather N/A")

        out.append(f"┌{'─'*(W-2)}┐")
        hdr = f"{away} @ {home}" + (f"  ·  {gtime}" if gtime else "")
        out.append(bl(hdr[:W-5]))
        out.append(bl(f"SP: {r['away_sp'][:20]} vs {r['home_sp'][:20]}"))
        out.append(bl(f"⛅ {w_s}"))
        out.append(div())
        out.append(bl("MODEL LINES"))
        out.append(bl(f"  Total:  {ld['mTotals']}"))
        out.append(bl(f"  Spread: {mspread}"))
        out.append(bl(f"  ML:     {ld['ml_pct']}"))

        books = _get_game_books(odds, away, home)
        pin = books.get("Pinnacle",{})
        if pin:
            out.append(div())
            out.append(bl("MARKET — Pinnacle"))
            ml_a = pin.get(f"ML_{away}"); ml_h = pin.get(f"ML_{home}")
            if ml_a and ml_h:
                out.append(bl(f"  ML: {away} {_fmt_odds(ml_a)} / {home} {_fmt_odds(ml_h)}"))
            tot_o = pin.get("Total_Over")
            if tot_o:
                tot_u = pin.get("Total_Under",{})
                out.append(bl(f"  Total: O/U {tot_o['line']} ({_fmt_odds(tot_o['odds'])} / {_fmt_odds(tot_u.get('odds','—'))})"))
            spd = pin.get(f"Spread_{away}")
            if spd:
                sign = "+" if (spd.get("line") or 0)>=0 else ""
                out.append(bl(f"  Spread: {away} {sign}{spd['line']} ({_fmt_odds(spd['odds'])})"))

        out.append(f"└{'─'*(W-2)}┘")
        out.append("")

    out += ["─"*W, "  Data Model by Laboy Picks  |  dubclub.win", "─"*W]
    with open(fpath,"w",encoding="utf-8") as f: f.write("\n".join(out))
    print(f"\n  📄 Exportado: {fname}")
    return fpath


# ──────────────────────────────────────────────────────
# EXPORT HTML — visual con logos (diseño original)
# ──────────────────────────────────────────────────────

def _team_color_hex(team):
    """Convierte color RGB del equipo a hex CSS."""
    c = _TEAM_COLORS.get(team)
    if c: return "#{:02X}{:02X}{:02X}".format(*c)
    return "#e85d04"


def _to_et(time_str):
    """Convierte un string de hora como '6:40 PM CT' → '7:40 PM AST'. Siempre retorna AST (Puerto Rico, UTC-4)."""
    if not time_str: return time_str
    offsets = {"ET": 0, "CT": 1, "MT": 2, "PT": 3}
    import re as _re
    m = _re.match(r"(\d+):(\d+)\s*(AM|PM)\s*([A-Z]+)", time_str.strip())
    if not m: return time_str
    h, mn, ap, tz = int(m.group(1)), int(m.group(2)), m.group(3), m.group(4)
    diff = offsets.get(tz, 0)
    h24 = h % 12 + (12 if ap == "PM" else 0)
    h24_ast = h24 + diff
    if h24_ast >= 24: h24_ast -= 24
    ap_ast = "PM" if h24_ast >= 12 else "AM"
    h12_ast = h24_ast % 12 or 12
    return f"{h12_ast}:{mn:02d} {ap_ast} AST"


def _filter_by_session(games):
    """Filtra lista de juegos por sesión (--day / --night).

    --day   → first pitch ANTES de 5:00 PM ET  (juegos de día)
    --pm    → first pitch 1:00 PM – 4:59 PM ET (sub-sesión tarde, dentro del día)
    --night → first pitch 5:00 PM ET en adelante (juegos de noche)
    Sin flags → retorna todos los juegos sin filtrar.
    """
    if not DAY_SESSION and not PM_SESSION and not NIGHT_SESSION:
        return games

    import re as _re

    def _et_hour(g):
        """Hora ET (0-23) del juego, o None si no puede determinarse."""
        gtime = g.get("game_time_local", "")
        # _to_et convierte cualquier zona a AST (UTC-4) = EDT en verano = ET
        et_str = _to_et(gtime) if gtime else ""
        m = _re.match(r"(\d+):(\d+)\s*(AM|PM)", et_str.strip())
        if m:
            h, ap = int(m.group(1)), m.group(3)
            return h % 12 + (12 if ap == "PM" else 0)
        # Fallback: parsear directamente de UTC
        utc_str = g.get("game_time_utc", "")
        if utc_str:
            try:
                clean  = utc_str.replace("Z", "").split("+")[0]
                utc_dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
                return (utc_dt.hour - 4) % 24   # EDT = UTC-4
            except Exception:
                pass
        return None

    result = []
    for g in games:
        h = _et_hour(g)
        if h is None:                               # hora desconocida → incluir siempre
            result.append(g)
        elif DAY_SESSION   and h < 17:
            result.append(g)
        elif PM_SESSION    and h >= 13 and h < 17:
            result.append(g)
        elif NIGHT_SESSION and h >= 17:
            result.append(g)
    return result


def _logo_b64():
    """
    Carga laboy_logo.png desde SCRIPT_DIR, remueve el fondo negro y retorna
    un data-URI base64 listo para usar en <img src="...">.
    Retorna None si el archivo no existe o hay error.
    """
    logo_path = os.path.join(SCRIPT_DIR, "laboy_logo.png")
    if not os.path.exists(logo_path):
        return None
    try:
        from PIL import Image
        import io, base64
        img = Image.open(logo_path).convert("RGBA")
        data = img.load()
        w, h = img.size
        # Remover píxeles negros/casi-negros → transparentes
        for y in range(h):
            for x in range(w):
                r, g, b, a = data[x, y]
                if r < 35 and g < 35 and b < 35:
                    data[x, y] = (0, 0, 0, 0)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/png;base64,{b64}"
    except Exception:
        return None


def _html_css():
    return """
  :root{
    --bg:#080808;--card:#0f0f0f;--card2:#141414;--inner:#1a1a1a;
    --accent:#f07820;--green:#22c55e;--red:#ef4444;
    --text:#f1f5f9;--muted:#64748b;--muted2:#94a3b8;--border:#202020;
  }
  *{box-sizing:border-box;margin:0;padding:0}
  body{
    background:var(--bg);color:var(--text);
    font-family:-apple-system,BlinkMacSystemFont,'Inter','Segoe UI',sans-serif;
    padding:0 0 56px;
  }
  /* ── Header ── */
  .header{
    background:#000;padding:18px 24px 14px;text-align:center;
    border-bottom:1px solid #1a1a1a;
    position:relative;
  }
  .header::after{
    content:'';position:absolute;bottom:-1px;left:50%;transform:translateX(-50%);
    width:120px;height:2px;background:var(--accent);border-radius:1px;
  }
  .header img.logo{height:110px;width:auto;display:block;margin:0 auto 2px}
  .header h1{font-size:1.8rem;font-weight:900;letter-spacing:4px;color:var(--accent)}
  .header .date{
    color:var(--muted);margin-top:4px;font-size:0.72rem;
    letter-spacing:2px;text-transform:uppercase;
  }
  /* ── Section wrapper ── */
  .section{max-width:800px;margin:0 auto;padding:28px 18px 0}
  .section-title{
    font-size:0.65rem;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;
    color:var(--muted);margin:24px 0 10px;
    display:flex;align-items:center;gap:8px;
  }
  .section-title::after{
    content:'';flex:1;height:1px;background:var(--border);
  }
  /* ── Pick card ── */
  .pick-card{
    background:var(--card);
    border:1px solid var(--border);
    border-radius:16px;
    padding:16px 18px 14px;
    margin-bottom:16px;
    overflow:hidden;
    box-shadow:0 4px 24px rgba(0,0,0,0.6);
  }
  .pick-time{font-size:0.68rem;color:var(--muted);margin-bottom:6px;letter-spacing:.5px}
  .teams-row{display:flex;align-items:center;gap:12px;margin-bottom:14px}
  .teams-row img{width:48px;height:48px;object-fit:contain}
  .pick-main{flex:1;min-width:0}
  .game-label{font-size:0.78rem;color:var(--muted);margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .pick-label{font-size:1.25rem;font-weight:800;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
  .odds-badge{
    background:#f0782018;color:var(--accent);
    border:1px solid #f0782030;
    border-radius:5px;padding:2px 8px;font-size:0.85rem;font-weight:700;
  }
  /* ── Stat chips ── */
  .stats-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:6px}
  .stat{
    background:var(--inner);border:1px solid var(--border);
    border-radius:8px;padding:8px 6px;text-align:center;
  }
  .stat-label{font-size:0.58rem;text-transform:uppercase;letter-spacing:1px;color:var(--muted);margin-bottom:3px}
  .stat-val{font-size:0.9rem;font-weight:700}
  /* ── No picks ── */
  .no-picks{color:var(--muted);text-align:center;padding:32px;font-size:0.85rem}
  /* ── Lines card ── */
  .line-card{
    background:var(--card);border:1px solid var(--border);
    border-radius:12px;padding:14px 16px;margin-bottom:10px;
    box-shadow:0 4px 20px rgba(0,0,0,0.5);
  }
  .line-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px}
  .team-logo{display:flex;align-items:center;gap:6px;font-size:0.8rem;font-weight:700}
  .team-logo img{width:34px;height:34px;object-fit:contain}
  .line-time{font-size:0.72rem;color:var(--muted)}
  .sp-row{font-size:0.7rem;color:var(--muted);margin-bottom:6px}
  .line-stats{display:flex;gap:10px;flex-wrap:wrap;font-size:0.78rem;margin-bottom:4px}
  .weather-row{font-size:0.72rem;color:var(--muted);margin-top:4px}
  .mkt-row{font-size:0.72rem;color:#475569;margin-top:2px}
  /* ── Footer ── */
  .footer{
    text-align:center;padding:36px 16px 0;
    color:#2a2a2a;font-size:0.72rem;letter-spacing:1px;
  }
  .footer a{color:var(--accent);text-decoration:none}"""


def _html_wrap(title, header_sub, dstr, yr, body_html):
    """Envuelve body_html en el shell HTML completo con header y footer — AI style."""
    logo_src = _logo_b64()
    logo_html = (f'<img class="dbg-logo" src="{logo_src}" alt="Laboy Picks">'
                 if logo_src else '<span class="dbg-wordmark">LABOY</span>')
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
<style>
{_html_css()}

/* ══════════════════════════════════════════════
   AI-STYLE — MLB Picks / Lines / Record
   ══════════════════════════════════════════════ */
body {{
  background:#050508;
  background-image:
    linear-gradient(rgba(0,220,255,.018) 1px, transparent 1px),
    linear-gradient(90deg, rgba(0,220,255,.018) 1px, transparent 1px);
  background-size:32px 32px;
}}
.dbg-header {{
  background:linear-gradient(180deg,#000 0%,#06060a 100%);
  padding:22px 24px 18px;
  text-align:center;
  border-bottom:1px solid #0ff2;
  position:relative;
  overflow:hidden;
}}
@keyframes dbg-scan {{
  0%   {{ transform:translateY(-120%) }}
  100% {{ transform:translateY(1200%) }}
}}
.dbg-header::before {{
  content:'';
  position:absolute;
  top:0;left:0;right:0;height:60px;
  background:linear-gradient(180deg,transparent,rgba(0,220,255,.09),transparent);
  animation:dbg-scan 4s linear infinite;
  pointer-events:none;
}}
.dbg-header::after {{
  content:'';
  position:absolute;
  bottom:-1px;left:10%;right:10%;
  height:1px;
  background:linear-gradient(90deg,transparent,#00dcff80,#f0782080,transparent);
}}
.dbg-logo {{
  height:80px;width:auto;display:block;margin:0 auto 10px;
  filter:drop-shadow(0 0 12px rgba(240,120,32,.45));
}}
.dbg-wordmark {{
  font-size:2rem;font-weight:900;letter-spacing:6px;
  background:linear-gradient(90deg,#f07820,#00dcff);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  display:block;margin-bottom:8px;
}}
.dbg-title {{
  font-size:0.6rem;font-weight:800;letter-spacing:4px;
  text-transform:uppercase;margin-bottom:4px;
  background:linear-gradient(90deg,#00dcff,#7c3aed,#00dcff);
  background-size:200% auto;
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  animation:dbg-gradient 4s linear infinite;
}}
@keyframes dbg-gradient {{
  0%   {{ background-position:0% center }}
  100% {{ background-position:200% center }}
}}
.dbg-date {{
  color:#475569;font-size:0.65rem;letter-spacing:2.5px;text-transform:uppercase;
  margin-top:2px;
}}
.dbg-badge {{
  display:inline-flex;align-items:center;gap:5px;
  background:rgba(0,220,255,.07);border:1px solid rgba(0,220,255,.2);
  border-radius:20px;padding:2px 10px;margin-top:8px;
  font-size:0.58rem;font-weight:700;letter-spacing:2px;color:#00dcff99;
  text-transform:uppercase;
}}
.dbg-badge-dot {{
  width:5px;height:5px;border-radius:50%;
  background:#00dcff;box-shadow:0 0 6px #00dcff;
  animation:dbg-pulse 1.8s ease-in-out infinite;
}}
@keyframes dbg-pulse {{
  0%,100%{{ opacity:1;transform:scale(1) }}
  50%    {{ opacity:.4;transform:scale(.7) }}
}}
.pick-card {{
  border:1px solid rgba(0,220,255,.12);
  background:linear-gradient(160deg,#0d0d10 0%,#0a0a0d 100%);
  box-shadow:0 0 0 1px rgba(0,220,255,.04),0 4px 32px rgba(0,0,0,.8),
             inset 0 1px 0 rgba(255,255,255,.03);
}}
.line-card {{
  border:1px solid rgba(0,220,255,.10);
  background:linear-gradient(160deg,#0d0d10 0%,#0a0a0d 100%);
  box-shadow:0 0 0 1px rgba(0,220,255,.03),0 4px 24px rgba(0,0,0,.75);
}}
.section-title {{
  color:transparent;
  background:linear-gradient(90deg,#00dcff,#94a3b8);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  font-size:0.6rem;font-weight:800;letter-spacing:3px;
}}
.section-title::after {{
  background:linear-gradient(90deg,rgba(0,220,255,.25),transparent);
}}
.stat {{
  background:rgba(0,220,255,.025);
  border:1px solid rgba(0,220,255,.08);
}}
.stat-val {{
  font-family:'JetBrains Mono','SF Mono','Fira Code','Courier New',monospace;
  font-size:0.88rem;letter-spacing:-.5px;
}}
.stat-label {{ color:#4a6272 }}
@keyframes dbg-odds-glow {{
  0%,100%{{ box-shadow:0 0 4px rgba(240,120,32,.3) }}
  50%    {{ box-shadow:0 0 10px rgba(240,120,32,.6) }}
}}
.odds-badge {{ animation:dbg-odds-glow 2.5s ease-in-out infinite; }}
.footer {{ color:#1a2530;border-top:1px solid #0ff1;margin-top:20px;padding-top:16px; }}
.footer a {{ color:#00dcff44 }}
::-webkit-scrollbar{{ width:5px }}
::-webkit-scrollbar-track{{ background:#050508 }}
::-webkit-scrollbar-thumb{{
  background:linear-gradient(180deg,#00dcff40,#f0782040);border-radius:3px;
}}
.fa-icon{{font-size:0.85em;opacity:0.75;margin-right:4px}}
.section-title .fa-icon{{font-size:0.9em;opacity:1;margin-right:6px}}
</style>
</head>
<body>
<div class="dbg-header">
  {logo_html}
  <div class="dbg-title">&#9632;&nbsp;Laboy Picks · {header_sub}&nbsp;&#9632;</div>
  <div class="dbg-date">{dstr}</div>
  <div><span class="dbg-badge"><span class="dbg-badge-dot"></span>Model Engine · Active</span></div>
</div>
<div class="section">
{body_html}
</div>
<div class="footer" style="text-align:center;padding:28px 16px 0;font-size:0.68rem;letter-spacing:1px">
  <p>Data Model by <a href="https://instagram.com/laboypicks">Laboy Picks</a> &nbsp;·&nbsp; {yr}</p>
</div>
</body>
</html>"""


def _debug_html_wrap(title, dstr, yr, body_html):
    """HTML wrapper específico para el Debug Report — diseño AI-style."""
    logo_src = _logo_b64()
    logo_html = (f'<img class="dbg-logo" src="{logo_src}" alt="Laboy Picks">'
                 if logo_src else '<span class="dbg-wordmark">LABOY</span>')
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
<style>
{_html_css()}

/* ══════════════════════════════════════════════
   AI-STYLE OVERRIDES — Debug Report only
   ══════════════════════════════════════════════ */

/* Grid background */
body {{
  background:#050508;
  background-image:
    linear-gradient(rgba(0,220,255,.018) 1px, transparent 1px),
    linear-gradient(90deg, rgba(0,220,255,.018) 1px, transparent 1px);
  background-size:32px 32px;
}}

/* ── Header ── */
.dbg-header {{
  background:linear-gradient(180deg,#000 0%,#06060a 100%);
  padding:22px 24px 18px;
  text-align:center;
  border-bottom:1px solid #0ff2;
  position:relative;
  overflow:hidden;
}}
/* Animated horizontal scan line */
@keyframes dbg-scan {{
  0%   {{ transform:translateY(-120%) }}
  100% {{ transform:translateY(1200%) }}
}}
.dbg-header::before {{
  content:'';
  position:absolute;
  top:0;left:0;right:0;height:60px;
  background:linear-gradient(180deg,transparent,rgba(0,220,255,.09),transparent);
  animation:dbg-scan 4s linear infinite;
  pointer-events:none;
}}
/* Bottom glow line */
.dbg-header::after {{
  content:'';
  position:absolute;
  bottom:-1px;left:10%;right:10%;
  height:1px;
  background:linear-gradient(90deg,transparent,#00dcff80,#f0782080,transparent);
}}
.dbg-logo {{
  height:80px;width:auto;display:block;margin:0 auto 10px;
  filter:drop-shadow(0 0 12px rgba(240,120,32,.45));
}}
.dbg-wordmark {{
  font-size:2rem;font-weight:900;letter-spacing:6px;
  background:linear-gradient(90deg,#f07820,#00dcff);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  display:block;margin-bottom:8px;
}}
.dbg-title {{
  font-size:0.6rem;font-weight:800;letter-spacing:4px;
  text-transform:uppercase;margin-bottom:4px;
  background:linear-gradient(90deg,#00dcff,#7c3aed,#00dcff);
  background-size:200% auto;
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  animation:dbg-gradient 4s linear infinite;
}}
@keyframes dbg-gradient {{
  0%   {{ background-position:0% center }}
  100% {{ background-position:200% center }}
}}
.dbg-date {{
  color:#475569;font-size:0.65rem;letter-spacing:2.5px;text-transform:uppercase;
  margin-top:2px;
}}
/* AI badge */
.dbg-badge {{
  display:inline-flex;align-items:center;gap:5px;
  background:rgba(0,220,255,.07);border:1px solid rgba(0,220,255,.2);
  border-radius:20px;padding:2px 10px;margin-top:8px;
  font-size:0.58rem;font-weight:700;letter-spacing:2px;color:#00dcff99;
  text-transform:uppercase;
}}
.dbg-badge-dot {{
  width:5px;height:5px;border-radius:50%;
  background:#00dcff;
  box-shadow:0 0 6px #00dcff;
  animation:dbg-pulse 1.8s ease-in-out infinite;
}}
@keyframes dbg-pulse {{
  0%,100%{{ opacity:1;transform:scale(1) }}
  50%    {{ opacity:.4;transform:scale(.7) }}
}}

/* ── Pick card AI glow ── */
.pick-card {{
  border:1px solid rgba(0,220,255,.12);
  background:linear-gradient(160deg,#0d0d10 0%,#0a0a0d 100%);
  box-shadow:
    0 0 0 1px rgba(0,220,255,.04),
    0 4px 32px rgba(0,0,0,.8),
    inset 0 1px 0 rgba(255,255,255,.03);
}}

/* ── Section titles ── */
.section-title {{
  color:transparent;
  background:linear-gradient(90deg,#00dcff,#94a3b8);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  font-size:0.6rem;font-weight:800;letter-spacing:3px;
}}
.section-title::after {{
  background:linear-gradient(90deg,rgba(0,220,255,.25),transparent);
}}

/* ── Stat chips ── */
.stat {{
  background:rgba(0,220,255,.025);
  border:1px solid rgba(0,220,255,.08);
}}
.stat-val {{
  font-family:'JetBrains Mono','SF Mono','Fira Code','Courier New',monospace;
  font-size:0.88rem;letter-spacing:-.5px;
}}
.stat-label {{ color:#4a6272 }}

/* ── Odds badge pulse ── */
@keyframes dbg-odds-glow {{
  0%,100%{{ box-shadow:0 0 4px rgba(240,120,32,.3) }}
  50%    {{ box-shadow:0 0 10px rgba(240,120,32,.6) }}
}}
.odds-badge {{
  animation:dbg-odds-glow 2.5s ease-in-out infinite;
}}

/* ── Inner stat cells in debug cards ── */
/* Numbers inside debug sections get mono font automatically */
.dbg-mono {{
  font-family:'JetBrains Mono','SF Mono','Fira Code','Courier New',monospace;
  letter-spacing:-.3px;
}}

/* ── Footer ── */
.footer {{
  color:#1a2530;border-top:1px solid #0ff1;
  margin-top:20px;padding-top:16px;
}}
.footer a {{ color:#00dcff44 }}

/* ── Scrollbar AI style ── */
::-webkit-scrollbar{{ width:5px }}
::-webkit-scrollbar-track{{ background:#050508 }}
::-webkit-scrollbar-thumb{{
  background:linear-gradient(180deg,#00dcff40,#f0782040);
  border-radius:3px;
}}

  .fa-icon{{font-size:0.85em;opacity:0.75;margin-right:4px}}
  .section-title .fa-icon{{font-size:0.9em;opacity:1;margin-right:6px}}
</style>
</head>
<body>
<div class="dbg-header">
  {logo_html}
  <div class="dbg-title">&#9632;&nbsp;Model Report&nbsp;&#9632;</div>
  <div class="dbg-date">MLB &nbsp;·&nbsp; {dstr}</div>
  <div><span class="dbg-badge"><span class="dbg-badge-dot"></span>AI Analysis Engine · Active</span></div>
</div>
<div class="section">
{body_html}
</div>
<div class="footer" style="text-align:center;padding:28px 16px 0;font-size:0.68rem;letter-spacing:1px">
  <p>Data Model by <a href="https://instagram.com/laboypicks">Laboy Picks</a> &nbsp;·&nbsp; {yr}</p>
</div>
</body>
</html>"""


def _compute_picks(results, odds):
    """
    Calcula picks EV+ para ML, Run Line y Total con reglas conservadoras:

      ML   : edge ≥ 3%, EV ≥ 3%;  modelo = probabilidad de ganar (%)
      RL   : solo favoritos por ≥2 runs proyectados; odds en [-125, +115];
             NO recomendar si el mismo equipo ya tiene ML con edge ≥ 3%
      Total: P(cover) = norm_cdf(diff / 3.1); diff ≥ 0.5 carreras; EV ≥ 4%

    Por juego: hasta 2 picks de CATEGORÍAS DISTINTAS (ML/RL vs Total).
    El segundo pick requiere EV ≥ 6% y edge ≥ 4%.
    """
    import math as _math

    def _norm_cdf(x):
        """CDF normal estándar usando math.erf (sin scipy)."""
        return 0.5 * (1.0 + _math.erf(x / _math.sqrt(2.0)))

    def _one_run_pct(model_total):
        return max(0.18, min(0.35, 0.35 - (model_total - 7.0) * 0.015))

    def _edge_float(c):
        try:
            return float(str(c.get("edge", "0")).replace("%", "").replace("+", "").split()[0])
        except:
            return 0.0

    # ── Umbrales — calibrados para máxima precisión ─────────────────────
    MIN_EV_ML    = 0.05   # EV mínimo para ML      (subido de 0.04 → más selectivo)
    MIN_EDGE_ML  = 5.5    # edge mínimo para ML (%) (subido de 4.5 → menos picks, más calidad)
    ML_MAX_JUICE = -155   # cap de juice para ML — más de -155 destruye el valor real
    MIN_EV_RL    = 0.05   # EV mínimo para RL
    MIN_EDGE_RL  = 4.5    # edge mínimo para RL (%) (subido de 3.5)
    MIN_EV_TOT   = 0.08   # EV mínimo para Total — único gate real (Monte Carlo siempre corre)
    # MIN_DIFF_OVR / MIN_DIFF_UND eliminados: el pre-filtro bloqueaba el Monte Carlo
    # aunque el EV fuera positivo. Ahora solo se requiere abs(diff) >= 0.3 runs.
    SIGMA        = 4.0    # σ de distribución de totales MLB (subido 3.5→4.0: más realismo en la varianza de runs totales)
    MIN_EV_SEC   = 0.09   # EV mínimo para segundo pick del mismo juego (subido de 0.07)
    MIN_EDGE_SEC = 7.0    # edge mínimo para segundo pick (%) (subido de 6.0)
    MAX_DAILY    = 6      # top 6 picks — calidad sobre cantidad (era 8)

    # ── Shrinkage asimétrico para underdogs ─────────────────────────────
    # Problema raíz: wRC+/xFIP son stats de "talento promedio de temporada".
    # Cuando el mercado pone a un equipo como +150, sabe algo contextual que
    # el modelo NO sabe: starter muy inferior esta noche, bullpen agotado,
    # lineup parcial, momentum negativo, etc.
    # Solución: darle más peso al mercado cuando el equipo es underdog.
    # No se bloquea — se calibra mejor. Si el edge sigue siendo real después
    # del ajuste, el pick es válido y tiene valor genuino.
    #
    #  Favorito  (odds < 0):    shrinkage normal   → _MODEL_W / _MARKET_W
    #  Ligero    (EVEN a -120): shrinkage +10% mkt → más skeptico
    #  Dog leve  (+101 a +149): shrinkage +20% mkt → bastante más skeptico
    #  Dog real  (+150 a +199): shrinkage +30% mkt → muy skeptico (mercado sabe)
    #  Big dog   (+200 y más):  shrinkage +20% mkt → upside compensa, menos mkt
    #
    ML_DOG_MKT_EXTRA = {
        # (min_odds, max_odds): extra_market_weight
        (-120, -100): 0.10,   # favorito ligero
        ( 101,  149): 0.20,   # underdog leve — mercado probablemente sabe por qué
        ( 150,  199): 0.30,   # underdog real — starter gap, etc.
        ( 200,  999): 0.20,   # big dog — upside alto, menos penalización
    }
    # EV mínimo también sube para underdogs: compensar mayor incertidumbre
    ML_DOG_EV_EXTRA  = {
        (-120, -100): 0.02,   # +2% EV para favorito ligero → necesita 7% total
        ( 101,  149): 0.03,   # +3% EV para dog leve → necesita 8% total
        ( 150,  199): 0.04,   # +4% EV para dog real → necesita 9% total
        ( 200,  999): 0.02,   # +2% EV para big dog → necesita 7% total
    }

    def _dog_params(mkt_o, base_model_w, base_ev):
        """Calcula shrinkage y EV threshold ajustado según odds del equipo."""
        for (lo, hi), extra_mkt in ML_DOG_MKT_EXTRA.items():
            if lo <= mkt_o <= hi:
                adj_model_w = max(0.45, base_model_w - extra_mkt)
                adj_ev      = base_ev + ML_DOG_EV_EXTRA.get((lo, hi), 0)
                return adj_model_w, 1.0 - adj_model_w, adj_ev
        return base_model_w, 1.0 - base_model_w, base_ev

    # ── Dynamic market shrinkage (calculado ANTES del loop) ──────────────
    # Usa sample_q para ajustar confianza en el modelo según punto de temporada
    # Early April: 70% modelo + 30% mercado  (poca muestra histórica)
    # Mid-season:  90% modelo + 10% mercado  (modelo más calibrado)
    _g_cur    = _season_games_estimate(TARGET_DATE)
    _sq       = min(1.0, 0.50 + _g_cur / 120)   # sample quality: 0.5 → 1.0
    _MODEL_W  = round(0.70 + 0.20 * _sq, 3)
    _MARKET_W = round(1.0 - _MODEL_W, 3)
    print(f"  📐 Shrinkage: {_MODEL_W:.0%} modelo / {_MARKET_W:.0%} mercado  (sample_q={_sq:.2f})")

    picks = []
    _near_misses = []   # candidatos que no alcanzaron el threshold — para diagnóstico
    _skipped_started = 0   # contador de juegos ya iniciados
    _now_utc = _math.inf  # fallback: no filtrar si no se puede parsear
    try:
        from datetime import timezone as _tz
        _now_utc = datetime.now(_tz.utc).replace(tzinfo=None)
    except Exception:
        pass

    # Fetch market signals — silencioso si no hay datos disponibles
    try:
        _mlb_market_sigs = _fetch_mlb_market_signals(sport="mlb")
        if _mlb_market_sigs:
            print(f"  📡 Market signals MLB: {len(_mlb_market_sigs)} juegos")
    except Exception:
        _mlb_market_sigs = {}

    _skipped_no_odds  = 0   # sin odds en el API para este juego
    _skipped_lineup   = 0   # sin lineup (solo con --confirmed)
    _games_evaluated  = 0   # juegos que pasaron todos los filtros y se evaluaron

    for r in results:
        away, home = r["away"], r["home"]
        lines  = r["lines"]
        books  = _get_game_books(odds, away, home)
        if not books:
            _skipped_no_odds += 1
            continue

        # ── Suffix para doubleheaders ─────────────────────────────────────
        _dh   = r.get("double_header", "N")
        _gnum = r.get("game_number", 1)
        _dh_suffix = f" (DH-G{_gnum})" if _dh not in ("N", "", None) else ""
        _game_str  = f"{away} @ {home}{_dh_suffix}"

        # ── Filtrar juegos ya iniciados ───────────────────────────────────
        # Si game_time_utc está en el pasado, el juego ya empezó —
        # los odds en vivo no son válidos para picks pre-game.
        _gut = r.get("game_time_utc", "")
        if _gut and not isinstance(_now_utc, float):
            try:
                _clean = _gut.replace("Z", "").split("+")[0]
                _gdt   = datetime.strptime(_clean, "%Y-%m-%dT%H:%M:%S")
                if _now_utc > _gdt:
                    _skipped_started += 1
                    continue   # juego en curso o ya terminado — omitir
            except Exception:
                pass

        # ── Filtrar juegos sin lineups confirmados (SIEMPRE requerido) ──────
        # Sin lineup confirmado el wRC+ usa baseline — el pick puede cambiar
        # radicalmente cuando salga la alineación real. No se recomienda nada
        # hasta tener ambos lineups. --confirmed sigue siendo válido para
        # compatibilidad, pero ahora el comportamiento default ES confirmado.
        _lu_a = r.get("lineup_used_away", False)
        _lu_h = r.get("lineup_used_home", False)
        if not _lu_a or not _lu_h:
            _gtime_disp = _to_et(r.get("game_time_local", "?"))
            _missing = []
            if not _lu_a: _missing.append(away)
            if not _lu_h: _missing.append(home)
            print(f"  ⏳ {_game_str} [{_gtime_disp}] — sin lineup: {', '.join(_missing)} → watchlist")
            _skipped_lineup += 1
            continue

        _games_evaluated += 1

        # ── Debut SP filter: suprimir picks cuando un SP debuta ──────────
        # Un debutante no tiene xFIP válido. El nombre puede colisionar con
        # otro pitcher en el cache (ej: ELMER RODRIGUEZ → EDUARDO RODRIGUEZ).
        # Si _debut_away o _debut_home está marcado en r, omitir picks de este juego.
        _debut_away_flag = r.get("_debut_away", False)
        _debut_home_flag = r.get("_debut_home", False)
        if _debut_away_flag or _debut_home_flag:
            _who = []
            if _debut_away_flag: _who.append(f"{r.get('away_sp','?')} ({away})")
            if _debut_home_flag: _who.append(f"{r.get('home_sp','?')} ({home})")
            print(f"  🆕 {_game_str} — DEBUT SP: {', '.join(_who)} → picks suprimidos (xFIP no confiable)")
            _near_misses.append({
                "game": _game_str, "pick": "DEBUT SP", "odds": "—",
                "ev": 0, "edge": 0,
                "gap": f"⛔ SP debutante: {', '.join(_who)} — sin historial MLB confiable"
            })
            continue

        gtime       = _to_et(r.get("game_time_local", ""))
        model_total = lines["total"]
        tA          = lines["tA"]   # carreras proyectadas away

        # ── Lineup status para este juego (se guarda en el pick JSON) ────
        _lineup_confirmed = bool(r.get("lineup_used_away")) and bool(r.get("lineup_used_home"))

        # ── Triple Lock check (Under qualifier) ───────────────────────
        # Factor 1: FIP combinado promedio < 3.80 (xFIP total SP+BP ponderado)
        _xfip_a = r.get("xfip_a", 4.0)
        _xfip_b = r.get("xfip_b", 4.0)
        _tc_fip = (_xfip_a + _xfip_b) / 2 < 3.80
        # Factor 2: Park Factor del home < 0.98 (parque favorable al pitcheo)
        _tc_pf  = PARK_FACTORS.get(home, 1.0) < 0.98
        # Factor 3: Ambas ofensas wRC+ < 100 (por debajo del promedio de liga)
        _wrc_a  = r.get("wrc_a", 100)
        _wrc_b  = r.get("wrc_b", 100)
        _tc_wrc = _wrc_a < 100 and _wrc_b < 100
        _triple_lock = _tc_fip and _tc_pf and _tc_wrc

        # ── Perfect Storm check (Over qualifier) ─────────────────────
        # Factor 1: FIP combinado promedio > 4.20 (pitcheo débil de ambos lados)
        _tp_fip = (_xfip_a + _xfip_b) / 2 > 4.20
        # Factor 2: Park Factor del home > 1.02 (parque favorable al bateo)
        _tp_pf  = PARK_FACTORS.get(home, 1.0) > 1.02
        # Factor 3: Ambas ofensas wRC+ > 105 (por encima del promedio de liga)
        _tp_wrc = _wrc_a > 105 and _wrc_b > 105
        _perfect_storm = _tp_fip and _tp_pf and _tp_wrc
        tB          = lines["tB"]   # carreras proyectadas home
        p1r         = _one_run_pct(model_total)

        # ── Mejores odds por tipo ─────────────────────────────────────────
        best_ml = {}
        best_rl = {}
        for bk in books.values():
            for t, mk in [(away, f"ML_{away}"), (home, f"ML_{home}")]:
                v = bk.get(mk)
                if v and (t not in best_ml or v > best_ml[t]):
                    best_ml[t] = v
            for t in (away, home):
                sp = bk.get(f"Spread_{t}")
                if sp and sp.get("odds"):
                    prev = best_rl.get(t)
                    if not prev or sp["odds"] > prev["odds"]:
                        best_rl[t] = sp

        total_implied = sum(_am_to_prob(best_ml[t]) for t in [away, home] if t in best_ml)

        mrl_candidates = []   # candidatos ML + RL
        tot_candidates = []   # candidatos Total
        ml_edge_map    = {}   # team → edge ML (para filtro RL)

        # ── 1. ML ────────────────────────────────────────────────────────
        for team, wk in [(away, "winA"), (home, "winB")]:
            mkt_o = best_ml.get(team)
            if not mkt_o: continue

            model_p_raw = lines[wk] / 100
            market_p    = _am_to_prob(mkt_o) / (total_implied or 1)

            # ── Shrinkage asimétrico: underdogs reciben más peso de mercado ──
            # El mercado sabe por qué pone a un equipo como dog esta noche
            # (starter gap, bullpen agotado, lineup parcial…) — el modelo no.
            # Usamos más peso del mercado cuando más underdog es el equipo.
            _mw, _mkw, _ev_thr = _dog_params(mkt_o, _MODEL_W, MIN_EV_ML)
            model_p = model_p_raw * _mw + market_p * _mkw
            edge    = (model_p - market_p) * 100
            ev      = model_p * _am_to_payout(mkt_o) - (1 - model_p)
            ml_edge_map[team] = edge

            # Indicador para display / near-miss
            _dog_label = ""
            if mkt_o >= 150:  _dog_label = " [dog+150]"
            elif mkt_o >= 101: _dog_label = " [dog+101]"
            elif -120 <= mkt_o <= -100: _dog_label = " [fav-lig]"

            if not (ev >= _ev_thr and edge >= MIN_EDGE_ML and mkt_o >= ML_MAX_JUICE):
                # Near-miss: guardar si estuvo cerca (dentro de 3% del threshold)
                ev_gap   = _ev_thr - ev
                edge_gap = MIN_EDGE_ML - edge
                if ev_gap < 0.03 or edge_gap < 3.0:
                    _near_misses.append({
                        "game": _game_str, "pick": f"{team} ML{_dog_label}", "odds": _fmt_odds(mkt_o),
                        "ev": ev, "edge": edge,
                        "gap": f"EV falta {ev_gap*100:+.1f}%  edge falta {edge_gap:+.1f}%"
                              if ev_gap > 0 or edge_gap > 0 else "juice cap (-155)",
                    })
            if ev >= _ev_thr and edge >= MIN_EDGE_ML and mkt_o >= ML_MAX_JUICE:
                # ── Filtro de contradicción SP: si el ERA diferencial del SP
                # propio vs oponente es > 3.5 (propio peor), el pick es
                # contradictorio — el modelo mismo lo flagea en el spread.
                # Protege contra picks como NATIONALS +160 con Littell ERA 7.56
                # vs Holmes ERA 2.15 (diferencial 5.4 — el mercado sabe).
                # recent_era_away/home son ahora dicts {"era", "ip", "n"} o None
                _era_stats_own = r.get("recent_era_away") if team == away else r.get("recent_era_home")
                _era_stats_opp = r.get("recent_era_home") if team == away else r.get("recent_era_away")
                _r_era_own = _era_stats_own["era"] if isinstance(_era_stats_own, dict) else None
                _r_era_opp = _era_stats_opp["era"] if isinstance(_era_stats_opp, dict) else None
                _sp_era_diff = None
                if _r_era_own is not None and _r_era_opp is not None:
                    _sp_era_diff = _r_era_own - _r_era_opp
                # Block: propio SP >= 1.5 ERA peor que oponente Y propio ERA reciente > 5.5
                _sp_contradiction = (
                    _sp_era_diff is not None
                    and _sp_era_diff > 1.5
                    and _r_era_own > 5.5
                )
                if _sp_contradiction:
                    _near_misses.append({
                        "game": _game_str,
                        "pick": f"{team} ML{_dog_label}",
                        "odds": _fmt_odds(mkt_o),
                        "ev": ev, "edge": edge,
                        "gap": f"⚠️ Contradicción SP: ERA propio {_r_era_own:.2f} vs oponente {_r_era_opp:.2f} (diff {_sp_era_diff:+.2f}) — pick bloqueado",
                    })
                    print(f"  🚫 {_game_str} | {team} ML — BLOQUEADO: ERA propio {_r_era_own:.2f} vs oponente {_r_era_opp:.2f} (diff {_sp_era_diff:+.2f} ≥ 1.5 & ERA > 5.5)")
                    continue

                # Nota extra para picks de underdog (muestra shrinkage aplicado)
                _dog_note = (f"mkt_w={_mkw:.0%}" if _dog_label else "")
                mrl_candidates.append({
                    "game":    _game_str, "time": gtime,
                    "pick":    f"{team} ML", "odds": _fmt_odds(mkt_o),
                    "modelo":  f"{model_p*100:.1f}%",   # probabilidad de ganar (blended)
                    "mercado": _fmt_odds(mkt_o),
                    "edge": f"{edge:+.1f}%", "ev": f"{ev*100:+.1f}%",
                    "_ev": ev, "_edge_val": edge,
                    "away": away, "home": home, "team": team, "_type": "ML",
                    "_model_p_raw": round(model_p_raw * 100, 1),  # raw para debug
                    "_dog_label": _dog_label.strip("[] ") or None,  # e.g. "dog+150"
                    "_dog_note": _dog_note,  # shrinkage info
                    "lineup_confirmed": _lineup_confirmed,
                })

        # ── 2. Run Line (-1.5 favoritoes proyectados ≥2 carreras) ────────
        for team, wk in [(away, "winA"), (home, "winB")]:
            rl = best_rl.get(team)
            if not rl: continue
            rl_line = float(rl.get("line", 0))
            rl_odds = rl["odds"]

            # Solo lado favorito (-1.5)
            if rl_line >= 0: continue
            # Odds razonables: -125 a +115
            if rl_odds < -125 or rl_odds > 115: continue
            # El equipo debe ser proyectado favorito por ≥2 carreras
            proj_diff = (tA - tB) if team == away else (tB - tA)
            if proj_diff < 2.0: continue
            # No apilar RL si ML mismo equipo ya tiene edge ≥ MIN_EDGE_ML
            if ml_edge_map.get(team, 0) >= MIN_EDGE_ML: continue

            model_win_p_raw = lines[wk] / 100
            mkt_ml_p = best_ml.get(team)
            mkt_win_anchor = (_am_to_prob(mkt_ml_p) / (total_implied or 1)
                              if mkt_ml_p else model_win_p_raw)
            # Shrinkage en win probability antes de calcular cover_p
            model_win_p = model_win_p_raw * _MODEL_W + mkt_win_anchor * _MARKET_W
            # -1.5: debe ganar por 2+ (= ganar - gana por 1 exactamente)
            cover_p = max(0.0, min(1.0, model_win_p - model_win_p * p1r * 0.6))
            mkt_impl = _am_to_prob(rl_odds)
            edge_pct = (cover_p - mkt_impl) * 100
            ev       = cover_p * _am_to_payout(rl_odds) - (1 - cover_p)
            rl_str   = f"{rl_line:.1f}".replace(".0", "")

            if ev >= MIN_EV_RL and edge_pct >= MIN_EDGE_RL:
                mrl_candidates.append({
                    "game":    _game_str, "time": gtime,
                    "pick":    f"{team} {rl_str}", "odds": _fmt_odds(rl_odds),
                    "modelo":  f"{cover_p*100:.1f}%",
                    "mercado": rl_str,
                    "edge": f"{edge_pct:+.1f}%", "ev": f"{ev*100:+.1f}%",
                    "_ev": ev, "_edge_val": edge_pct,
                    "away": away, "home": home, "team": team, "_type": "RL",
                    "lineup_confirmed": _lineup_confirmed,
                })

        # ── 2b. Run Line Underdog (+1.5) ─────────────────────────────────
        # P(cover +1.5) = P(dog wins outright) + P(fav wins by exactly 1)
        # Solo cuando el mercado ofrece odds razonables (+100 a +145)
        # y el modelo proyecta un juego competitivo (diferencia ≤ 2.5 runs)
        for dog_team, dog_wk, fav_wk in [(away,"winA","winB"),(home,"winB","winA")]:
            rl = best_rl.get(dog_team)
            if not rl: continue
            rl_line = float(rl.get("line", 0))
            rl_odds = rl["odds"]

            # Solo el lado underdog (+1.5)
            if rl_line <= 0: continue
            # Odds con valor real: entre +100 y +145
            # (más alto → el libro ya lo descuenta; muy juiced → no vale)
            if rl_odds < 100 or rl_odds > 145: continue
            # El modelo no debe proyectar al dog perdiendo por más de 2.5 runs
            dog_proj_diff = (tA - tB) if dog_team == away else (tB - tA)
            if dog_proj_diff < -2.5: continue   # demasiado underdog proyectado
            # No apilar con ML del mismo equipo si ya tiene edge sólido
            if ml_edge_map.get(dog_team, 0) >= MIN_EDGE_ML: continue

            dog_win_p_raw = lines[dog_wk] / 100
            fav_win_p_raw = lines[fav_wk] / 100
            # Shrinkage usando ML del mercado como anchor
            mkt_dog_ml = best_ml.get(dog_team)
            mkt_dog_anchor = (_am_to_prob(mkt_dog_ml) / (total_implied or 1)
                              if mkt_dog_ml else dog_win_p_raw)
            dog_win_p = dog_win_p_raw * _MODEL_W + mkt_dog_anchor * _MARKET_W
            fav_win_p = 1.0 - dog_win_p
            # P(fav gana por exactamente 1) ≈ P(fav wins) × p1r × 0.6
            p_fav_by1 = fav_win_p * p1r * 0.6
            cover_p   = min(0.98, dog_win_p + p_fav_by1)
            mkt_impl  = _am_to_prob(rl_odds)
            edge_pct  = (cover_p - mkt_impl) * 100
            ev        = cover_p * _am_to_payout(rl_odds) - (1 - cover_p)

            if ev >= MIN_EV_RL and edge_pct >= MIN_EDGE_RL:
                mrl_candidates.append({
                    "game":    _game_str, "time": gtime,
                    "pick":    f"{dog_team} +1.5", "odds": _fmt_odds(rl_odds),
                    "modelo":  f"{cover_p*100:.1f}%",
                    "mercado": "+1.5",
                    "edge": f"{edge_pct:+.1f}%", "ev": f"{ev*100:+.1f}%",
                    "_ev": ev, "_edge_val": edge_pct,
                    "away": away, "home": home, "team": dog_team, "_type": "RL+",
                    "lineup_confirmed": _lineup_confirmed,
                })

        # ── 3. Total — línea de BetMGM como referencia primaria ──────────
        # El usuario apuesta en BetMGM, así que la comparación modelo vs mercado
        # debe hacerse contra la línea de BetMGM, no la mediana de todos los books.
        # Pinnacle sirve como señal sharp pero NO como línea de referencia para EV.
        # Fallback: FanDuel → DraftKings → mediana (si BetMGM no disponible).
        _USER_BOOK_PRIORITY = ["BetMGM", "FanDuel", "DraftKings"]
        mkt_line = None
        _mkt_line_source = "—"

        # 1) Intentar books en orden de prioridad
        for _priority_book in _USER_BOOK_PRIORITY:
            _pb = books.get(_priority_book, {})
            _pto = _pb.get("Total_Over")
            if _pto and _pto.get("line"):
                try:
                    _v = float(_pto["line"])
                    if 5.5 <= _v <= 14.5:
                        mkt_line = _v
                        _mkt_line_source = _priority_book
                        break
                except (TypeError, ValueError):
                    pass

        # 2) Fallback: mediana de todos los books disponibles
        if mkt_line is None:
            _all_lines = []
            for _bk in books.values():
                _to = _bk.get("Total_Over")
                if _to and _to.get("line"):
                    try:
                        _v = float(_to["line"])
                        if 5.5 <= _v <= 14.5:
                            _all_lines.append(_v)
                    except (TypeError, ValueError):
                        pass
            if _all_lines:
                _all_lines.sort()
                mkt_line = _all_lines[len(_all_lines) // 2]
                _mkt_line_source = "mediana"

        if mkt_line is None:
            pass   # sin datos de total para este juego
        else:
            # Advertir si Pinnacle difiere significativamente de BetMGM (señal sharp)
            _pinnacle_line = None
            _pinn_bk = books.get("Pinnacle", {})
            _pinn_to = _pinn_bk.get("Total_Over", {})
            if _pinn_to and _pinn_to.get("line"):
                try: _pinnacle_line = float(_pinn_to["line"])
                except: pass
            _pinn_note = ""
            if _pinnacle_line and abs(_pinnacle_line - mkt_line) >= 0.5:
                _pinn_note = f"  ⚡ Pinnacle={_pinnacle_line:.1f} (diff {_pinnacle_line-mkt_line:+.1f} vs {_mkt_line_source})"

            diff = model_total - mkt_line   # > 0 = Over, < 0 = Under

            # Sanity check mínimo (0.3 runs) — evita casos donde modelo ≈ mercado.
            # El umbral real de calidad es MIN_EV_TOT (8%) — calculado por Monte Carlo.
            # Removido el pre-filtro MIN_DIFF_OVR/MIN_DIFF_UND: el Monte Carlo Poisson
            # corre siempre que haya al menos 0.3 runs de diferencia, y el EV decide.
            _diff_ok = abs(diff) >= 0.3
            if _diff_ok:

                # Mejores odds para ese mkt_line — BetMGM primero, luego mejor disponible
                # Usamos BetMGM como referencia porque es el book del usuario.
                # Si BetMGM no tiene odds para esta línea, tomamos el mejor de los demás.
                best_over_odds  = None
                best_under_odds = None
                # Paso 1: intentar BetMGM en esa línea exacta
                _bmgm = books.get("BetMGM", {})
                _bmgm_to = _bmgm.get("Total_Over", {})
                _bmgm_tu = _bmgm.get("Total_Under", {})
                if _bmgm_to and _bmgm_to.get("line") and abs(float(_bmgm_to["line"]) - mkt_line) < 0.26:
                    best_over_odds = _bmgm_to.get("odds")
                if _bmgm_tu and _bmgm_tu.get("line") and abs(float(_bmgm_tu["line"]) - mkt_line) < 0.26:
                    best_under_odds = _bmgm_tu.get("odds")
                # Paso 2: si BetMGM no tiene esa línea, tomar el mejor odds disponible
                # (excluye Pinnacle — odds no accesibles para el usuario)
                _SKIP_BOOKS = {"Pinnacle", "Betfair"}
                for _bk_name, _bk in books.items():
                    if _bk_name in _SKIP_BOOKS: continue
                    _to = _bk.get("Total_Over")
                    _tu = _bk.get("Total_Under")
                    if _to and _to.get("line") and abs(float(_to["line"]) - mkt_line) < 0.26:
                        if best_over_odds is None or _to["odds"] > best_over_odds:
                            best_over_odds = _to["odds"]
                    if _tu and _tu.get("line") and abs(float(_tu["line"]) - mkt_line) < 0.26:
                        if best_under_odds is None or _tu["odds"] > best_under_odds:
                            best_under_odds = _tu["odds"]

                # ── Monte Carlo (Poisson) — P(OVER) y P(UNDER) reales ──────────
                # Reemplaza norm_cdf(diff/σ) con 50k simulaciones Poisson.
                # tA/tB son lambdas independientes → distribución correcta para runs.
                _mc = _monte_carlo_totals(tA, tB, mkt_line)

                if diff > 0 and best_over_odds is not None:
                    cover_p_raw = _mc["p_over"]   # P(total > mkt_line) por Poisson
                    mkt_ods     = best_over_odds
                    # Shrinkage: ancla al 50% de mercado (totals son ~50/50 sin vig)
                    mkt_tot_p   = _am_to_prob(mkt_ods)
                    cover_p     = cover_p_raw * _MODEL_W + mkt_tot_p * _MARKET_W
                    pick_str    = f"Over {mkt_line}"
                    edge_disp   = f"+{diff:.1f}"
                    edge_pct    = (cover_p - mkt_tot_p) * 100
                    ev          = cover_p * _am_to_payout(mkt_ods) - (1 - cover_p)
                    if ev >= MIN_EV_TOT:
                        tot_candidates.append({
                            "game":    _game_str, "time": gtime,
                            "pick":    pick_str, "odds": _fmt_odds(mkt_ods),
                            "modelo":  f"Proj {model_total:.1f}",
                            "mercado": f"Line {mkt_line}",
                            "edge": edge_disp, "ev": f"{ev*100:+.1f}%",
                            "_ev": ev, "_edge_val": edge_pct,
                            "away": away, "home": home, "team": None, "_type": "TOT",
                            "tormenta": _perfect_storm,
                            "_tp_fip": _tp_fip, "_tp_pf": _tp_pf, "_tp_wrc": _tp_wrc,
                            "lineup_confirmed": _lineup_confirmed,
                            "p_over":  _mc["p_over"],
                            "p_under": _mc["p_under"],
                            "p_push":  _mc["p_push"],
                            "market_total": mkt_line,
                        })
                    else:
                        ev_gap = MIN_EV_TOT - ev
                        if ev_gap < 0.04:
                            _near_misses.append({
                                "game": _game_str, "pick": pick_str, "odds": _fmt_odds(mkt_ods),
                                "ev": ev, "edge": edge_pct,
                                "gap": f"EV falta {ev_gap*100:+.1f}%  (proyección {model_total:.1f} vs línea {mkt_line})  MC:{_mc['p_over']*100:.1f}%",
                            })
                elif diff < 0 and best_under_odds is not None:
                    cover_p_raw = _mc["p_under"]  # P(total < mkt_line) por Poisson
                    mkt_ods     = best_under_odds
                    mkt_tot_p   = _am_to_prob(mkt_ods)
                    cover_p     = cover_p_raw * _MODEL_W + mkt_tot_p * _MARKET_W
                    pick_str    = f"Under {mkt_line}"
                    edge_disp   = f"+{abs(diff):.1f}"
                    edge_pct    = (cover_p - mkt_tot_p) * 100
                    ev          = cover_p * _am_to_payout(mkt_ods) - (1 - cover_p)
                    if ev >= MIN_EV_TOT:
                        tot_candidates.append({
                            "game":    _game_str, "time": gtime,
                            "pick":    pick_str, "odds": _fmt_odds(mkt_ods),
                            "modelo":  f"Proj {model_total:.1f}",
                            "mercado": f"Line {mkt_line}",
                            "edge": edge_disp, "ev": f"{ev*100:+.1f}%",
                            "_ev": ev, "_edge_val": edge_pct,
                            "away": away, "home": home, "team": None, "_type": "TOT",
                            "candado": _triple_lock,
                            "_tc_fip": _tc_fip, "_tc_pf": _tc_pf, "_tc_wrc": _tc_wrc,
                            "lineup_confirmed": _lineup_confirmed,
                            "p_over":  _mc["p_over"],
                            "p_under": _mc["p_under"],
                            "p_push":  _mc["p_push"],
                            "market_total": mkt_line,
                        })
                    else:
                        ev_gap = MIN_EV_TOT - ev
                        if ev_gap < 0.04:
                            _near_misses.append({
                                "game": _game_str, "pick": pick_str, "odds": _fmt_odds(mkt_ods),
                                "ev": ev, "edge": edge_pct,
                                "gap": f"EV falta {ev_gap*100:+.1f}%  (proyección {model_total:.1f} vs línea {mkt_line})  MC:{_mc['p_under']*100:.1f}%",
                            })

        # ── 3b. Alternate Totals — líneas alternas de BetMGM ────────────
        # Monte Carlo evalúa cada línea alterna. Si tiene mejor EV que la
        # línea principal, compite directamente en tot_candidates.
        _bmgm_alt_totals = books.get("BetMGM", {}).get("Alt_Totals", [])
        if _bmgm_alt_totals:
            for _alt in _bmgm_alt_totals:
                _alt_line = _alt.get("line")
                if _alt_line is None: continue
                _alt_line = float(_alt_line)
                # Evitar duplicar la línea principal (ya evaluada arriba)
                if mkt_line is not None and abs(_alt_line - mkt_line) < 0.1: continue
                _alt_diff = model_total - _alt_line
                if abs(_alt_diff) < 0.3: continue   # sanity check mínimo
                _alt_mc = _monte_carlo_totals(tA, tB, _alt_line)
                if _alt_diff > 0:   # Over la línea alterna
                    _alt_odds = _alt.get("over_odds")
                    if _alt_odds is None: continue
                    _cp_raw = _alt_mc["p_over"]
                    _mkt_p  = _am_to_prob(_alt_odds)
                    _cp     = _cp_raw * _MODEL_W + _mkt_p * _MARKET_W
                    _ev     = _cp * _am_to_payout(_alt_odds) - (1 - _cp)
                    _ep     = (_cp - _mkt_p) * 100
                    if _ev >= MIN_EV_TOT:
                        tot_candidates.append({
                            "game":    _game_str, "time": gtime,
                            "pick":    f"Over {_alt_line}",
                            "odds":    _fmt_odds(_alt_odds),
                            "modelo":  f"Proj {model_total:.1f}",
                            "mercado": f"Alt {_alt_line} (BetMGM)",
                            "edge":    f"+{_alt_diff:.1f}",
                            "ev":      f"{_ev*100:+.1f}%",
                            "_ev": _ev, "_edge_val": _ep,
                            "away": away, "home": home, "team": None, "_type": "TOT",
                            "tormenta": _perfect_storm,
                            "_tp_fip": _tp_fip, "_tp_pf": _tp_pf, "_tp_wrc": _tp_wrc,
                            "lineup_confirmed": _lineup_confirmed,
                            "p_over":  _alt_mc["p_over"],
                            "p_under": _alt_mc["p_under"],
                            "p_push":  _alt_mc["p_push"],
                            "market_total": _alt_line,
                            "_is_alt_line": True,
                        })
                    elif MIN_EV_TOT - _ev < 0.04:
                        _near_misses.append({
                            "game": _game_str, "pick": f"Over {_alt_line} (Alt BetMGM)",
                            "odds": _fmt_odds(_alt_odds), "ev": _ev, "edge": _ep,
                            "gap": f"EV falta {(MIN_EV_TOT-_ev)*100:+.1f}%  MC:{_cp_raw*100:.1f}%",
                        })
                elif _alt_diff < 0:   # Under la línea alterna
                    _alt_odds = _alt.get("under_odds")
                    if _alt_odds is None: continue
                    _cp_raw = _alt_mc["p_under"]
                    _mkt_p  = _am_to_prob(_alt_odds)
                    _cp     = _cp_raw * _MODEL_W + _mkt_p * _MARKET_W
                    _ev     = _cp * _am_to_payout(_alt_odds) - (1 - _cp)
                    _ep     = (_cp - _mkt_p) * 100
                    if _ev >= MIN_EV_TOT:
                        tot_candidates.append({
                            "game":    _game_str, "time": gtime,
                            "pick":    f"Under {_alt_line}",
                            "odds":    _fmt_odds(_alt_odds),
                            "modelo":  f"Proj {model_total:.1f}",
                            "mercado": f"Alt {_alt_line} (BetMGM)",
                            "edge":    f"+{abs(_alt_diff):.1f}",
                            "ev":      f"{_ev*100:+.1f}%",
                            "_ev": _ev, "_edge_val": _ep,
                            "away": away, "home": home, "team": None, "_type": "TOT",
                            "candado": _triple_lock,
                            "_tc_fip": _tc_fip, "_tc_pf": _tc_pf, "_tc_wrc": _tc_wrc,
                            "lineup_confirmed": _lineup_confirmed,
                            "p_over":  _alt_mc["p_over"],
                            "p_under": _alt_mc["p_under"],
                            "p_push":  _alt_mc["p_push"],
                            "market_total": _alt_line,
                            "_is_alt_line": True,
                        })
                    elif MIN_EV_TOT - _ev < 0.04:
                        _near_misses.append({
                            "game": _game_str, "pick": f"Under {_alt_line} (Alt BetMGM)",
                            "odds": _fmt_odds(_alt_odds), "ev": _ev, "edge": _ep,
                            "gap": f"EV falta {(MIN_EV_TOT-_ev)*100:+.1f}%  MC:{_cp_raw*100:.1f}%",
                        })

        # ── 2b. Alternate Spreads — líneas alternas de BetMGM ────────────
        # Usa Monte Carlo de margen de victoria para evaluar cada spread alterno.
        # Útil cuando -0.5 tiene mejor precio que ML, o 2.5 crea valor real.
        _bmgm_alt_spreads = books.get("BetMGM", {}).get("Alt_Spreads", [])
        if _bmgm_alt_spreads:
            # Correr MC una sola vez para este juego (compartido entre todos los spreads)
            _mc_sp_fn = _monte_carlo_spreads(tA, tB)
            for _asp in _bmgm_alt_spreads:
                _asp_team  = _asp.get("team")
                _asp_line  = _asp.get("line")
                _asp_odds  = _asp.get("odds")
                if not _asp_team or _asp_line is None or not _asp_odds: continue
                # Evitar duplicar el spread estándar -1.5/+1.5 (ya evaluado arriba)
                if abs(abs(_asp_line) - 1.5) < 0.1: continue
                _team_is_away = (_asp_team == away)
                _cp_raw = _mc_sp_fn(_asp_line, _team_is_away)
                _mkt_p  = _am_to_prob(_asp_odds)
                # Shrinkage — mismo peso que ML (el mercado entiende el margen)
                _dog_w, _mkt_w, _ev_thr = _dog_params(_asp_odds, _MODEL_W, MIN_EV_RL)
                _cp     = _cp_raw * _dog_w + _mkt_p * _mkt_w
                _ev     = _cp * _am_to_payout(_asp_odds) - (1 - _cp)
                _ep     = (_cp - _mkt_p) * 100
                if _ev >= MIN_EV_RL and _ep >= MIN_EDGE_RL and _asp_odds >= ML_MAX_JUICE:
                    mrl_candidates.append({
                        "game":    _game_str, "time": gtime,
                        "pick":    f"{_asp_team} {_fmt_odds(_asp_line)} (Alt RL)",
                        "odds":    _fmt_odds(_asp_odds),
                        "modelo":  f"MC cover {_cp_raw*100:.1f}%",
                        "mercado": f"Alt Spread {_fmt_odds(_asp_line)} (BetMGM)",
                        "edge":    f"{_ep:+.1f}%",
                        "ev":      f"{_ev*100:+.1f}%",
                        "_ev": _ev, "_edge_val": _ep,
                        "away": away, "home": home, "team": _asp_team, "_type": "RL+",
                        "lineup_confirmed": _lineup_confirmed,
                        "_is_alt_line": True,
                    })

        # ── Selección: hasta 2 picks por juego (categorías distintas) ────
        mrl_candidates.sort(key=lambda x: -x["_ev"])
        tot_candidates.sort(key=lambda x: -x["_ev"])

        best_mrl = mrl_candidates[0] if mrl_candidates else None
        best_tot = tot_candidates[0] if tot_candidates else None

        game_picks = []
        if best_mrl and best_tot:
            # Determinar primary / secondary por EV
            if best_mrl["_ev"] >= best_tot["_ev"]:
                primary, secondary = best_mrl, best_tot
            else:
                primary, secondary = best_tot, best_mrl
            primary["alt_picks"] = []
            game_picks.append(primary)
            # Segundo pick: barra más alta
            if (secondary["_ev"] >= MIN_EV_SEC and
                    secondary["_edge_val"] >= MIN_EDGE_SEC):
                secondary["alt_picks"] = []
                game_picks.append(secondary)
            else:
                # Mostrar como alternativa en la card del primary
                primary["alt_picks"] = [
                    f"{secondary['pick']} ({secondary['odds']}) EV:{secondary['_ev']*100:.1f}%"
                ]
        elif best_mrl:
            best_mrl["alt_picks"] = []
            game_picks.append(best_mrl)
        elif best_tot:
            best_tot["alt_picks"] = []
            game_picks.append(best_tot)

        # ── Market signal confirmation ────────────────────────────────────
        # Añadir señal a cada pick del juego y filtrar FADEs fuertes
        for gp in game_picks:
            p_type = gp.get("_type", "") or gp.get("type", "")
            away_abb = r.get("away", "")
            home_abb = r.get("home", "")
            game_key = f"{away_abb}_{home_abb}"

            if p_type in ("ML", "RL"):
                mkt_side = "HOME" if gp.get("team") == home_abb else "AWAY"
                mkt_bet  = "ml" if p_type == "ML" else "spread"
            elif "Over" in gp.get("pick", "") or p_type == "OVER":
                mkt_side, mkt_bet = "OVER", "over"
            else:
                mkt_side, mkt_bet = "UNDER", "under"

            mkt_conf = _mlb_sharp_confirm(_mlb_market_sigs, game_key, mkt_bet, mkt_side)
            gp["market_signal"] = mkt_conf
            gp["market_label"]  = _format_mlb_market_signal(mkt_conf)

        picks.extend(game_picks)

    # ── Confidence scoring: edge × EV × sample quality ───────────────
    # sample_q ya calculado arriba como _sq (mismo valor)
    sample_q = _sq

    for p in picks:
        edge_v = abs(p["_edge_val"])
        ev_v   = p["_ev"] * 100
        # Boost picks where we have high edge AND high EV (multiplicative)
        # Penalize when sample is small (early season uncertainty)
        base_conf = (edge_v * 0.45 + ev_v * 0.40) * sample_q
        # Triple Lock boost: +15% confidence para Unders que cumplen los 3 factores
        # Perfect Storm boost: +15% confidence para Overs que cumplen los 3 factores
        tc_boost = 1.15 if p.get("candado") else (1.15 if p.get("tormenta") else 1.0)
        p["_conf"] = round(base_conf * tc_boost, 3)
        # ── Fractional Kelly bet sizing (Quarter Kelly) ─────────────────
        try:
            _odds_raw = str(p.get("odds", "0"))
            # Normalize: remove +, replace unicode minus (−) and em-dash (–) with ASCII -
            _odds_clean = (_odds_raw
                           .replace("+", "")
                           .replace("\u2212", "-")   # unicode minus
                           .replace("\u2013", "-")   # en-dash
                           .replace("\u2014", "-")   # em-dash
                           .strip())
            _odds_int = int(float(_odds_clean))      # float first handles "−110.0" etc.
            _model_p  = float(str(p.get("modelo", "50%")).replace("%","").strip()) / 100
            p["kelly"] = _kelly_fraction(_model_p, _odds_int, fraction=0.25)
        except Exception:
            p["kelly"] = "—"

    # ── Top MAX_DAILY por confianza, sin cuotas por tipo ─────────────
    picks.sort(key=lambda x: -x["_conf"])

    # ── Daily dog cap: máximo 2 picks de underdog ML (odds > +100) ───
    # Tener 4+ underdogs en el mismo día crea riesgo de barrida correlacionada:
    # si el mercado tuvo razón al poner a todos como dogs, el daño se multiplica.
    # Se mantienen los mejores 2 por confianza; el resto se elimina del card.
    MAX_DOG_ML = 2
    _dog_count = 0
    _filtered_picks = []
    for _p in picks:
        _is_dog_ml = (
            _p.get("_type") == "ML"
            and _p.get("_dog_label") is not None   # tiene label de dog
        )
        if _is_dog_ml:
            _dog_count += 1
            if _dog_count > MAX_DOG_ML:
                print(f"  🐶 Dog cap ({MAX_DOG_ML}/día): {_p['pick']} {_p['odds']} descartado (conf {_p['_conf']:.2f})")
                continue
        _filtered_picks.append(_p)
    picks = _filtered_picks

    picks = picks[:MAX_DAILY]

    # ── Re-ordenar por horario para el output ─────────────────────────
    picks.sort(key=lambda x: (_parse_time_sort(x["time"]), -x["_conf"]))

    # Exponer contadores como atributos de la función para que el caller los pueda leer
    _compute_picks._skipped_no_odds  = _skipped_no_odds
    _compute_picks._skipped_started  = _skipped_started
    _compute_picks._skipped_lineup   = _skipped_lineup
    _compute_picks._games_evaluated  = _games_evaluated
    _compute_picks._near_misses      = _near_misses

    return picks


# ──────────────────────────────────────────────────────
# MODEL PICKS HISTORY — auto-save + auto-grade
# ──────────────────────────────────────────────────────

def _load_model_picks():
    if os.path.exists(MODEL_PICKS_FILE):
        with open(MODEL_PICKS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def _save_model_picks(picks):
    with open(MODEL_PICKS_FILE, "w", encoding="utf-8") as f:
        json.dump(picks, f, indent=2, ensure_ascii=False)

def _save_debug_state(picks):
    """Snapshot simple del último --picks run. Sin locking, sin historial.
    Siempre sobreescribe. --export-debug lee de aquí."""
    try:
        with open(DEBUG_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"date": TARGET_DATE, "session": CURRENT_SESSION, "picks": picks}, f, ensure_ascii=False)
    except Exception:
        pass

def _load_debug_state():
    """Carga el snapshot del último --picks run."""
    try:
        if os.path.exists(DEBUG_STATE_FILE):
            with open(DEBUG_STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_log_state(picks):
    """Snapshot del último --export-debug HTML. Solo se escribe al generar el debug HTML.
    --log lee de aquí para mostrar EXACTAMENTE los picks que aparecieron en el último HTML."""
    try:
        with open(LOG_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"date": TARGET_DATE, "session": CURRENT_SESSION, "picks": picks}, f, ensure_ascii=False)
    except Exception:
        pass

def _load_log_state():
    """Carga el snapshot del último --export-debug HTML.
    Si no existe, retorna {} — cmd_log_pick hará fallback a debug state."""
    try:
        if os.path.exists(LOG_STATE_FILE):
            with open(LOG_STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

# ──────────────────────────────────────────────────────
# RESULTS CACHE — guarda resultados del modelo por fecha
# para poder generar debug cards al loggear un pick
# ──────────────────────────────────────────────────────

def _save_results_cache(results, date_str=None):
    """
    Guarda los resultados completos del modelo (lista de game-result dicts)
    en un cache JSON por fecha.  Retiene las últimas 7 fechas.
    Silencioso en caso de error para no interrumpir el flujo principal.
    """
    d = date_str or TARGET_DATE
    try:
        existing = {}
        if os.path.exists(RESULTS_CACHE_FILE):
            with open(RESULTS_CACHE_FILE, "r", encoding="utf-8") as f:
                existing = json.load(f)
        existing[d] = results
        # Retener solo las últimas 7 fechas para no crecer indefinidamente
        keys = sorted(existing.keys())[-7:]
        existing = {k: existing[k] for k in keys}
        with open(RESULTS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False)
    except Exception:
        pass


def _load_results_cache(date_str=None):
    """Carga los resultados del modelo cacheados para una fecha dada."""
    d = date_str or TARGET_DATE
    try:
        if os.path.exists(RESULTS_CACHE_FILE):
            with open(RESULTS_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            return cache.get(d, [])
    except Exception:
        pass
    return []


def _model_picks_save_today(picks_list):
    """
    Persiste los picks del modelo para TARGET_DATE + CURRENT_SESSION en MODEL_PICKS_FILE.

    Lógica de locking por sesión:
    - Si ya existen picks para (date, session) y NO hay --force-repick, muestra aviso y NO sobreescribe.
    - Con --force-repick, borra los picks de esa (date, session) y guarda los nuevos.
    - Sesiones independientes: guardar --day no afecta --night ni --full.
    """
    all_picks = _load_model_picks()

    # Picks ya guardados para esta fecha + sesión
    existing = [p for p in all_picks
                if p.get("date") == TARGET_DATE and p.get("session", "full") == CURRENT_SESSION]

    if existing and not FORCE_REPICK:
        _lock_time = existing[0].get("saved_at", "?")
        _lock_lu   = sum(1 for ep in existing if ep.get("lineup_confirmed"))
        print(f"\n  🔒 PICKS BLOQUEADOS (guardados a las {_lock_time}) — "
              f"{len(existing)} picks · {_lock_lu} con lineup confirmado · "
              f"sesión '{CURRENT_SESSION}' {TARGET_DATE}")
        print(f"     Usa --force-repick para sobreescribir (rompe calibración de --grade-picks).")
        print(f"     Picks en historial:")
        for ep in existing:
            _lu_icon = "📋" if ep.get("lineup_confirmed") else "⚠️ "
            print(f"       {_lu_icon}  {ep['game']:<32}  {ep['pick']:<14}  {ep.get('odds',''):<7}  edge:{ep.get('edge','?')}")
        return

    if existing and FORCE_REPICK:
        print(f"  ⚠️  --force-repick activo: borrando {len(existing)} picks previos de "
              f"{TARGET_DATE} sesión '{CURRENT_SESSION}'.")

    # Conserva todo EXCEPTO (date, session) actual
    history = [p for p in all_picks
               if not (p.get("date") == TARGET_DATE and
                       p.get("session", "full") == CURRENT_SESSION)]

    import datetime as _dtsave
    _saved_at = _dtsave.datetime.now().strftime("%H:%M")

    for p in picks_list:
        history.append({
            "date":              TARGET_DATE,
            "session":           CURRENT_SESSION,
            "saved_at":          _saved_at,          # hora del primer lock
            "game":              p["game"],
            "pick":              p["pick"],
            "odds":              p["odds"],
            "modelo":            p.get("modelo",""),
            "mercado":           p.get("mercado",""),
            "edge":              p.get("edge",""),
            "ev":                p.get("ev",""),
            "time":              p.get("time",""),
            "lineup_confirmed":  p.get("lineup_confirmed", False),
            # Campos de debug — preservar para que export-debug sea fiel al run
            "away":              p.get("away",""),
            "home":              p.get("home",""),
            "team":              p.get("team",""),
            "_type":             p.get("_type",""),
            "candado":           p.get("candado", False),
            "tormenta":          p.get("tormenta", False),
            "result":            None,   # se llena con --grade-picks
            "actual":            None,   # score real: "5-3", "OVER 9.5", etc.
        })
    _save_model_picks(history)
    _lu_count = sum(1 for p in picks_list if p.get("lineup_confirmed"))
    _dirty    = len(picks_list) - _lu_count
    _dirty_s  = f"  ⚠️  {_dirty} sin lineup confirmado" if _dirty else ""
    print(f"  💾 {len(picks_list)} picks guardados a las {_saved_at} — {TARGET_DATE} sesión '{CURRENT_SESSION}'.{_dirty_s}")


def _fetch_mlb_scores(date_str):
    """
    Obtiene scores MLB del día desde ESPN API (pública, sin key).
    Retorna lista de {away, home, away_model, home_model, away_score, home_score, status}
    away_model / home_model = nombre del equipo en el modelo (e.g. "D-BACKS") vía TEAM_ABB.
    """
    try:
        ymd = date_str.replace("-", "")
        url = f"https://site.api.espn.com/apis/site/v2/sports/baseball/mlb/scoreboard?dates={ymd}"
        import urllib.request
        with urllib.request.urlopen(url, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        scores = []
        for ev in data.get("events", []):
            comp   = ev.get("competitions", [{}])[0]
            status = comp.get("status", {}).get("type", {}).get("name", "")
            teams  = {t["homeAway"]: t for t in comp.get("competitors", [])}
            away_t = teams.get("away", {})
            home_t = teams.get("home", {})
            away_abb = away_t.get("team", {}).get("abbreviation", "").upper()
            home_abb = home_t.get("team", {}).get("abbreviation", "").upper()
            scores.append({
                "away":       away_abb,
                "home":       home_abb,
                # Model names for direct matching against game_label (e.g. "D-BACKS @ ORIOLES")
                "away_model": TEAM_ABB.get(away_abb, ""),
                "home_model": TEAM_ABB.get(home_abb, ""),
                "away_score": int(away_t.get("score", 0) or 0),
                "home_score": int(home_t.get("score", 0) or 0),
                "status":     status,
            })
        return scores
    except Exception as e:
        print(f"  ⚠️  ESPN API error: {e}")
        return []


def _grade_model_pick(pick_entry, scores):
    """
    Determina W/L/P para un pick del modelo dado los scores reales.
    pick_entry["pick"]: "RED SOX ML", "O 8.5", "U 9.0", "RED SOX -1.5"
    Retorna ("W"|"L"|"P"|None, descripción)
    """
    pick_str = pick_entry["pick"].upper().strip()
    game_str = pick_entry["game"].upper().strip()
    odds_raw = pick_entry["odds"]

    # Encontrar el juego correspondiente en scores
    game_score = None
    for s in scores:
        a = s["away"].upper(); h = s["home"].upper()
        if a in game_str or h in game_str:
            game_score = s
            break

    if not game_score:
        return None, "juego no encontrado en scores"
    if game_score["status"] not in ("STATUS_FINAL", "Final", "STATUS_FINAL_INNINGS"):
        return None, f"juego no finalizado ({game_score['status']})"

    away_s = game_score["away_score"]
    home_s = game_score["home_score"]
    total  = away_s + home_s
    actual_desc = f"{game_score['away']} {away_s} – {home_s} {game_score['home']}"

    # ── Totals ────────────────────────────────────────
    if pick_str.startswith("O ") or pick_str == "OVER":
        line = float(re.sub(r"[^0-9.]", "", pick_str.split()[-1])) if " " in pick_str else 0
        if total > line:   return "W", actual_desc
        if total < line:   return "L", actual_desc
        return "P", actual_desc

    if pick_str.startswith("U ") or pick_str == "UNDER":
        line = float(re.sub(r"[^0-9.]", "", pick_str.split()[-1])) if " " in pick_str else 0
        if total < line:   return "W", actual_desc
        if total > line:   return "L", actual_desc
        return "P", actual_desc

    # ── Moneyline ─────────────────────────────────────
    for t in list(TEAM_ABB.values()) + list(TEAM_ABB.keys()):
        t_up = t.upper()
        if t_up in pick_str and "ML" in pick_str:
            # Determinar si el pick team es away o home
            if game_score["away"].upper() in t_up or t_up in game_score["away"].upper():
                won = away_s > home_s
            else:
                won = home_s > away_s
            if away_s == home_s: return "P", actual_desc
            return ("W" if won else "L"), actual_desc

    # ── Run Line / Spread ─────────────────────────────
    spread_m = re.search(r"([+-]?\d+\.?\d*)\s*$", pick_str)
    if spread_m:
        spread = float(spread_m.group(1))
        for t in list(TEAM_ABB.values()) + list(TEAM_ABB.keys()):
            t_up = t.upper()
            if t_up in pick_str:
                if game_score["away"].upper() in t_up or t_up in game_score["away"].upper():
                    diff = (away_s - home_s) + spread
                else:
                    diff = (home_s - away_s) + spread
                if diff > 0:  return "W", actual_desc
                if diff < 0:  return "L", actual_desc
                return "P", actual_desc

    return None, f"pick no reconocido: {pick_str}"


def cmd_grade_picks():
    """
    --grade-picks [DATE|URL] [--publish]
    Parsea picks desde 'Laboy Picks {DATE}.html' (local) o desde una URL pública.
    Descarga scores MLB del día (ESPN API), actualiza el HTML con W/L/P + genera card de resumen.
    Con --publish: sube el Picks HTML actualizado + Model Card a GitHub Pages.
    Uso:
      python3 mlb.py --grade-picks                           (usa fecha de hoy)
      python3 mlb.py --grade-picks 2026-04-12
      python3 mlb.py --grade-picks 2026-04-12 --publish
      python3 mlb.py --grade-picks URL --publish
    """
    try:
        gi  = sys.argv.index("--grade-picks")
        arg = sys.argv[gi+1] if gi+1 < len(sys.argv) and not sys.argv[gi+1].startswith("-") else None
    except (ValueError, IndexError):
        arg = None

    # Determinar si es URL o fecha
    url_source = None
    if arg and arg.startswith("http"):
        url_source = arg
        # Extraer fecha del URL si posible
        m = re.search(r"(\d{4}-\d{2}-\d{2})", arg)
        gdate = m.group(1) if m else TARGET_DATE
    else:
        gdate = arg if arg else TARGET_DATE

    print(f"\n  📡 Descargando scores MLB para {gdate}...")
    scores = _fetch_mlb_scores(gdate)
    if not scores:
        print("  ❌ No se pudieron obtener scores. ¿El día ya terminó?")
        return

    print(f"  ✅ {len(scores)} juegos encontrados.\n")

    # Obtener HTML de URL o archivo local
    from bs4 import BeautifulSoup

    if url_source:
        print(f"  🌐 Descargando HTML desde URL...")
        import urllib.request
        try:
            req = urllib.request.Request(url_source, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
            })
            with urllib.request.urlopen(req, timeout=15) as resp:
                html_content = resp.read().decode("utf-8")
        except Exception as e:
            print(f"  ❌ Error al descargar URL: {e}\n")
            return
        soup = BeautifulSoup(html_content, "html.parser")
        # Guardar localmente para modificar (con token para consistencia)
        html_filename = f"Laboy Picks {gdate}-{_url_token(gdate)}.html"
        html_path = os.path.join(SCRIPT_DIR, html_filename)
        print(f"  💾 Guardando como: {html_filename}")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
    else:
        import glob as _glob
        # 1. Buscar con token actual
        html_filename = f"Laboy Picks {gdate}-{_url_token(gdate)}.html"
        html_path     = os.path.join(SCRIPT_DIR, html_filename)

        # 2. Fallback: cualquier archivo "Laboy Picks {date}*.html" (token distinto o sin token)
        if not os.path.exists(html_path):
            candidates = sorted(
                _glob.glob(os.path.join(SCRIPT_DIR, f"Laboy Picks {gdate}*.html")),
                key=os.path.getmtime, reverse=True
            )
            if candidates:
                html_path     = candidates[0]
                html_filename = os.path.basename(html_path)
                print(f"  ⚠️  Archivo con token no encontrado — usando: {html_filename}")

        # 3. Buscar también en el repo de GitHub Pages local
        if not os.path.exists(html_path):
            repo_candidates = sorted(
                _glob.glob(os.path.join(GITHUB_PAGES_REPO, f"Laboy Picks {gdate}*.html")),
                key=os.path.getmtime, reverse=True
            )
            if repo_candidates:
                html_path     = repo_candidates[0]
                html_filename = os.path.basename(html_path)
                print(f"  📁 Encontrado en repo: {html_filename}")

        if not os.path.exists(html_path):
            print(f"  ❌ Archivo no encontrado para {gdate}")
            print(f"     Prueba con la URL directa:")
            print(f"     python3 mlb.py --grade-picks {GITHUB_PAGES_URL}/Laboy%20Picks%20{gdate}.html\n")
            return

        print(f"  📄 Leyendo picks desde: {html_filename}")
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")

    pick_cards = soup.find_all("div", class_="pick-card")
    print(f"  🃏 {len(pick_cards)} cards encontrados.\n")

    # Limpiar badges / score-divs de corridas anteriores para evitar duplicados
    for card in pick_cards:
        for bd in card.find_all("div", style=lambda s: s and "justify-content:flex-end" in (s or "")):
            bd.decompose()
        for sd in card.find_all("div", string=lambda t: t and str(t).startswith("Score:")):
            sd.decompose()

    # Helper: short team name → ESPN abbreviation (uppercase)
    def _short_to_espn(name):
        n = name.upper().strip()
        # Direct lookup (e.g. "TWINS" → "min")
        v = ESPN_ABB.get(n)
        if v: return v.upper()
        # Partial/substring match
        for short, espn in ESPN_ABB.items():
            if n in short or short in n:
                return espn.upper()
        return n  # fallback: use as-is

    picks_with_results = []
    rows = []

    for card in pick_cards:
        # ── Extraer texto de divs ──────────────────────────────────────────
        game_label_div = card.find("div", class_="game-label")
        pick_label_div = card.find("div", class_="pick-label")

        if not game_label_div or not pick_label_div:
            continue

        game_label = game_label_div.get_text(strip=True)
        # Pick text solo (sin badge de odds) — buscar nodo de texto directo
        pick_text = pick_label_div.find(string=True, recursive=False)
        if pick_text:
            pick_label = pick_text.strip()
        else:
            pick_label = pick_label_div.get_text(strip=True)

        odds_badge = card.find("span", class_="odds-badge")
        odds_str = odds_badge.get_text(strip=True) if odds_badge else "—"

        pick_upper = pick_label.upper()

        # ── Parsear teams del game_label: "AWAY @ HOME" ──────────────────
        gl_away_espn = ""
        gl_home_espn = ""
        gl_away_name = ""
        gl_home_name = ""
        if " @ " in game_label:
            parts = game_label.split(" @ ", 1)
            gl_away_name = parts[0].strip().upper()
            gl_home_name = parts[1].strip().upper()
            gl_away_espn = _short_to_espn(gl_away_name)
            gl_home_espn = _short_to_espn(gl_home_name)
        else:
            # Intentar buscar ambos equipos como fallback
            for short in ESPN_ABB:
                if short in game_label.upper():
                    if not gl_away_espn: gl_away_espn = ESPN_ABB[short].upper()
                    elif not gl_home_espn: gl_home_espn = ESPN_ABB[short].upper()

        # ── Encontrar score (prioridad: nombre del modelo > ESPN abb) ────
        game_score = None
        for s in scores:
            sa = s["away"].upper(); sh = s["home"].upper()
            sa_m = s.get("away_model", "").upper()
            sh_m = s.get("home_model", "").upper()
            # Primero: comparar por nombre del modelo (más confiable)
            if gl_away_name and sa_m and sa_m == gl_away_name: game_score = s; break
            if gl_home_name and sh_m and sh_m == gl_home_name: game_score = s; break
            # Fallback: ESPN abbreviation match
            if gl_away_espn and sa == gl_away_espn: game_score = s; break
            if gl_home_espn and sh == gl_home_espn: game_score = s; break
            # Extra fallback: cruce (e.g. doubleheader or home/away swap)
            if gl_away_name and sh_m and sh_m == gl_away_name: game_score = s; break
            if gl_home_name and sa_m and sa_m == gl_home_name: game_score = s; break

        result = None    # None = pendiente (no terminó o no encontrado)
        score_str = "—"
        color = "#f07820"  # naranja = pendiente

        if game_score:
            status = game_score.get("status", "")
            # Acepta cualquier variante de FINAL (STATUS_FINAL, STATUS_FINAL_INNINGS, etc.)
            game_finished = (
                "FINAL" in status.upper() or
                "COMPLETE" in status.upper() or
                "POSTPONED" in status.upper() or
                status in ("Final", "Final/Extra Innings")
            )

            if not game_finished:
                score_str = status or "En curso"
            else:
                away_s = game_score["away_score"]
                home_s = game_score["home_score"]
                total  = away_s + home_s
                score_str = f"{away_s}-{home_s}"

                # ── Totals ──────────────────────────────────────────────
                if "OVER" in pick_upper or re.match(r"^O[\s]", pick_upper):
                    m = re.search(r"(?:OVER|O)\s+([\d.]+)", pick_upper)
                    line = float(m.group(1)) if m else 0.0
                    if total > line:   result, color = "W", "#22c55e"
                    elif total < line: result, color = "L", "#ef4444"
                    else:              result, color = "P", "#94a3b8"

                elif "UNDER" in pick_upper or re.match(r"^U[\s]", pick_upper):
                    m = re.search(r"(?:UNDER|U)\s+([\d.]+)", pick_upper)
                    line = float(m.group(1)) if m else 0.0
                    if total < line:   result, color = "W", "#22c55e"
                    elif total > line: result, color = "L", "#ef4444"
                    else:              result, color = "P", "#94a3b8"

                elif "ML" in pick_upper:
                    # Determinar si el pick es por el equipo visitante o local
                    # gl_away_name/gl_home_name ya están definidos del game_label
                    pick_team_is_away = None
                    if gl_away_name and gl_away_name in pick_upper:
                        pick_team_is_away = True
                    elif gl_home_name and gl_home_name in pick_upper:
                        pick_team_is_away = False
                    else:
                        # Fallback: buscar en ESPN_ABB y comparar con API
                        for short, espn in ESPN_ABB.items():
                            if short in pick_upper:
                                espn_u = espn.upper()
                                sa_u   = game_score["away"].upper()
                                sh_u   = game_score["home"].upper()
                                if sa_u == espn_u:
                                    pick_team_is_away = True
                                elif sh_u == espn_u:
                                    pick_team_is_away = False
                                break

                    if pick_team_is_away is None:
                        result, color = "P", "#94a3b8"   # no se pudo determinar
                    elif away_s == home_s:
                        result, color = "P", "#94a3b8"   # empate (imposible en MLB)
                    else:
                        if pick_team_is_away:
                            pick_won = away_s > home_s
                        else:
                            pick_won = home_s > away_s
                        result = "W" if pick_won else "L"
                        color  = "#22c55e" if result == "W" else "#ef4444"

                else:
                    # ── Run Line / Spread (e.g. "ORIOLES -1.5", "RED SOX +1.5") ──
                    spread_m = re.search(r"([+-]\d+\.?\d*)\s*$", pick_upper)
                    if spread_m:
                        spread = float(spread_m.group(1))
                        pick_team_is_away = None
                        if gl_away_name and gl_away_name in pick_upper:
                            pick_team_is_away = True
                        elif gl_home_name and gl_home_name in pick_upper:
                            pick_team_is_away = False
                        else:
                            # Fallback: buscar en ESPN_ABB
                            for short, espn in ESPN_ABB.items():
                                if short in pick_upper:
                                    espn_u = espn.upper()
                                    if game_score["away"].upper() == espn_u:
                                        pick_team_is_away = True
                                    elif game_score["home"].upper() == espn_u:
                                        pick_team_is_away = False
                                    break

                        if pick_team_is_away is not None:
                            if pick_team_is_away:
                                diff = (away_s - home_s) + spread
                            else:
                                diff = (home_s - away_s) + spread
                            if diff > 0:   result, color = "W", "#22c55e"
                            elif diff < 0: result, color = "L", "#ef4444"
                            else:          result, color = "P", "#94a3b8"

        # Garantizar que color siempre refleje el resultado final
        if result == "W":   color = "#22c55e"
        elif result == "L": color = "#ef4444"
        elif result == "P": color = "#94a3b8"
        # (None = pending → color queda naranja #f07820)

        # ── Actualizar card en HTML ─────────────────────────────────────
        if result == "W":
            badge_html = '<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✅ WIN</span>'
            card_bg = "background:linear-gradient(135deg,#0d1f14 0%,#222222 60%)"
        elif result == "L":
            badge_html = '<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">❌ LOSS</span>'
            card_bg = "background:linear-gradient(135deg,#1f0d0d 0%,#222222 60%)"
        elif result == "P":
            badge_html = '<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
            card_bg = "background:linear-gradient(135deg,#14161a 0%,#222222 60%)"
        else:
            # Pendiente: no tocar el color del card
            badge_html = '<span style="background:#f0782022;color:#f07820;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">⏳ PENDING</span>'
            card_bg = ""

        # Insertar badge en card (antes del content)
        badge_div = soup.new_tag("div", style="display:flex;justify-content:flex-end;margin-bottom:10px")
        badge_div.append(BeautifulSoup(badge_html, "html.parser"))

        # Insertar score bajo pick-label (solo si el juego ya tiene resultado real)
        _score_skip = ("—", "STATUS_SCHEDULED", "STATUS_PRE_GAME",
                       "STATUS_IN_PROGRESS", "En curso")
        score_div = None
        if score_str not in _score_skip:
            score_div = soup.new_tag("div", style="font-size:0.75rem;color:#94a3b8;margin-top:4px")
            score_div.string = f"Score: {score_str}"

        if card_bg:
            card["style"] = f"border-left:4px solid {color};{card_bg}"
        else:
            # Pending: only update border, keep existing background
            existing = card.get("style", "")
            cleaned = re.sub(r"border-left[^;]*;?", "", existing).strip().strip(";")
            card["style"] = f"border-left:4px solid {color};{cleaned}"

        card.insert(0, badge_div)
        if score_div is not None:
            pick_label_div.insert_after(score_div)

        result_str = result if result is not None else "⏳"
        picks_with_results.append({
            "game": game_label,
            "pick": pick_label,
            "odds": odds_str,
            "modelo": "—",
            "edge": "—",
            "result": result_str,
            "score": score_str,
            "color": color,
        })

        res_sym = {"W":"✅","L":"❌","P":"—"}.get(result_str,"⏳")
        rows.append([pick_label, odds_str, res_sym + result_str, score_str])

    # Guardar HTML actualizado
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(str(soup))

    print(tab(rows, ["Pick","Odds","Resultado","Score"], fmt="simple"))
    print()

    # ── Stats separadas: calibración limpia vs sucia ─────────────────────
    # Cruzar con JSON para saber si el pick tenía lineup confirmado al hacerse
    _hist_for_grade = {hp.get("game",""): hp
                       for hp in _load_model_picks() if hp.get("date") == gdate}

    clean_results = []
    dirty_results = []
    for pr in picks_with_results:
        _hp = _hist_for_grade.get(pr["game"], {})
        # lineup_confirmed: True = pick hecho con lineup; False/None = sin lineup
        # Default True para picks viejos sin el campo (no penalizar historial anterior)
        if _hp.get("lineup_confirmed", True):
            clean_results.append(pr)
        else:
            dirty_results.append(pr)

    def _stats_line(label, subset):
        w  = sum(1 for p in subset if p["result"]=="W")
        l  = sum(1 for p in subset if p["result"]=="L")
        pu = sum(1 for p in subset if p["result"]=="P")
        pd = sum(1 for p in subset if p["result"] not in ("W","L","P"))
        tot = w + l + pu
        wp  = f"{w/tot*100:.0f}%" if tot else "—"
        pnd = f"  ⏳ {pd} pend." if pd else ""
        return f"  {label}: {w}W / {l}L / {pu}P  →  Win% {wp}{pnd}"

    pending_total = sum(1 for p in picks_with_results if p["result"] not in ("W","L","P"))
    print(f"\n  📊 {gdate} — Resultados:")
    if clean_results:
        print(_stats_line("📋 Con lineup    (calibración válida)", clean_results))
    if dirty_results:
        print(_stats_line("⚠️  Sin lineup   (excluir de calibración)", dirty_results))
    if not clean_results and not dirty_results:
        print("  (sin picks calificados)")
    print()

    # ── Actualizar historial JSON con resultados reales ──────────────────────
    # Mapear game_label → resultado para cruzar con el historial guardado
    result_map = {pr["game"]: pr for pr in picks_with_results if pr["result"] in ("W","L","P")}
    if result_map:
        all_hist = _load_model_picks()
        updated_count = 0
        for hp in all_hist:
            if hp.get("date") != gdate:
                continue
            gkey = hp.get("game", "")
            if gkey in result_map:
                pr = result_map[gkey]
                hp["result"] = pr["result"]
                hp["actual"] = pr.get("score", "")
                updated_count += 1
        if updated_count:
            _save_model_picks(all_hist)
            print(f"  📝 {updated_count} pick(s) actualizados en historial JSON ({gdate}).")

    # Exportar tarjeta de resumen
    card_path = export_daily_picks_card(gdate, picks_with_results)
    if card_path:
        print(f"  📄 Card: {os.path.basename(card_path)}")

    # ── Publicar en GitHub Pages si se pidió ────────────────────────────────
    if PUBLISH_MODE:
        to_publish = []
        # Picks HTML actualizado con W/L/P badges
        if html_path and os.path.isfile(html_path):
            to_publish.append(html_path)
        # Model Card HTML (resumen diario)
        if card_path and os.path.isfile(card_path):
            to_publish.append(card_path)
        if to_publish:
            cmd_publish(to_publish)
        else:
            print("  ⚠️  No hay HTMLs para publicar.\n")

    print()


def _over_under_logo_html(size=44):
    """Retorna <img> con over_under.png como data-URI base64."""
    import base64 as _b64
    p = os.path.join(SCRIPT_DIR, "over_under.png")
    if os.path.exists(p):
        with open(p, "rb") as f:
            data = _b64.b64encode(f.read()).decode()
        return (f'<img src="data:image/png;base64,{data}" alt="O/U" '
                f'width="{size}" height="{size}" style="object-fit:contain;border-radius:4px">')
    return ""


def _is_total_pick(pick_str):
    """True si el pick es un total (Over/Under). Maneja formatos: 'OVER 8.5', 'O 8.5', 'O8.5', 'U 9'."""
    pu = pick_str.strip().upper()
    return bool(
        re.match(r'^O[\s]?[\d.]', pu) or
        re.match(r'^U[\s]?[\d.]', pu) or
        "OVER" in pu or "UNDER" in pu
    )


def _pick_logo(pick_str, game_str, size=44):
    """
    Devuelve un <img> HTML con el logo del equipo del pick.
    Para OVER/UNDER usa over_under.png.
    Para ML picks busca el nombre del equipo en pick_str.
    Para totals usa el equipo visitante del game_label como fallback.
    """
    pu = pick_str.upper()
    # Over/Under → logo especial
    if _is_total_pick(pick_str):
        logo = _over_under_logo_html(size)
        if logo: return logo
    for short in sorted(ESPN_ABB.keys(), key=len, reverse=True):
        if short in pu:
            url = logo_url(short)
            if url:
                return (f'<img src="{url}" alt="{short}" width="{size}" height="{size}" '
                        f'style="object-fit:contain;border-radius:4px" onerror="this.style.display=\'none\'">')
    # Fallback: equipo visitante del game_label ("AWAY @ HOME")
    if " @ " in game_str:
        away = game_str.split(" @ ")[0].strip().upper()
        url = logo_url(away)
        if url:
            return (f'<img src="{url}" alt="{away}" width="{size}" height="{size}" '
                    f'style="object-fit:contain;border-radius:4px" onerror="this.style.display=\'none\'">')
    return ""


def export_daily_picks_card(date_str, picks_data):
    """
    Genera 'Laboy Model Card {DATE}.html' — tarjeta de resumen diario de picks con resultados.

    picks_data: lista de dicts con {game, pick, odds, modelo, edge, result, score, color}
    Retorna path del HTML generado.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    if not picks_data:
        return None

    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
    except:
        dt = datetime.now()

    dstr = dt.strftime("%A, %B %d").upper()
    yr = dt.strftime("%Y")

    # Calcular stats
    w = sum(1 for p in picks_data if p["result"]=="W")
    l = sum(1 for p in picks_data if p["result"]=="L")
    pu = sum(1 for p in picks_data if p["result"]=="P")
    total = w + l + pu
    win_pct = f"{w/total*100:.0f}%" if total else "—"

    # Construir picks rows
    picks_html = ""
    for p in picks_data:
        result = p["result"]
        color  = p.get("color", "#94a3b8")

        if result == "W":
            result_badge = '<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✅ WIN</span>'
            card_bg = "background:linear-gradient(135deg,#0d1f14 0%,#222222 60%)"
        elif result == "L":
            result_badge = '<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">❌ LOSS</span>'
            card_bg = "background:linear-gradient(135deg,#1f0d0d 0%,#222222 60%)"
        elif result == "P":
            result_badge = '<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
            card_bg = "background:linear-gradient(135deg,#14161a 0%,#222222 60%)"
        else:
            result_badge = '<span style="background:#f0782022;color:#f07820;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">⏳ PENDING</span>'
            card_bg = ""

        logo_html = _pick_logo(p["pick"], p["game"])

        picks_html += f"""
        <div class="pick-card" style="border-left:4px solid {color};{card_bg}">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
            <div style="font-size:0.75rem;color:var(--muted)">{esc(p['game'][:36])}</div>
            {result_badge}
          </div>
          <div class="teams-row" style="margin-bottom:10px">
            {logo_html}
            <div>
              <div style="font-size:1.0rem;font-weight:700">{esc(p['pick'])} <span style="background:#f0782022;color:#f07820;border-radius:6px;padding:2px 8px;font-size:0.9rem">{esc(p['odds'])}</span></div>
              {f'<div style="font-size:0.85rem;color:var(--muted);margin-top:6px;font-weight:600">Score: {esc(p["score"])}</div>' if p.get("score") and p["score"] not in ("—", "STATUS_SCHEDULED", "STATUS_PRE_GAME", "STATUS_IN_PROGRESS") else ""}
            </div>
          </div>
        </div>
        """

    win_col = "#22c55e" if w >= l else "#ef4444"
    body = f"""
    <div class="section-title">PICKS DEL MODELO</div>
    {picks_html}

    <div style="background:#1a1a1a;border-radius:12px;padding:16px;margin-top:24px;text-align:center">
      <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">RESUMEN</div>
      <div style="font-size:1.5rem;font-weight:900;color:{win_col}">{w}W · {l}L · {pu}P</div>
      <div style="font-size:1rem;color:var(--muted);margin-top:4px">Win% {win_pct}</div>
    </div>
    """

    html = _html_wrap(f"Laboy Model Card {date_str}", "MLB", dstr, yr, body)

    # Guardar HTML (token para URL impredecible)
    _tok = _url_token(date_str)
    html_file = f"Laboy Model Card {date_str}-{_tok}.html"
    html_path = os.path.join(SCRIPT_DIR, html_file)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    return html_path


def export_record_card(date_str=None):
    """
    Genera 'Laboy Record Card {DATE}.html' con los picks logueados (desde _load_log())
    para una fecha específica o all-time si date_str es None.

    Retorna path del HTML generado.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    log = _load_log()
    if not log:
        return None

    # Filtrar por fecha si se especifica
    if date_str:
        entries = [e for e in log if e.get("date") == date_str]
    else:
        entries = log

    if not date_str:
        date_str = "All-Time"

    try:
        if date_str != "All-Time":
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            dstr = dt.strftime("%A, %B %d").upper()
            yr = dt.strftime("%Y")
        else:
            dstr = "ALL-TIME"
            yr = datetime.now().strftime("%Y")
    except:
        dstr = "ALL-TIME"
        yr = datetime.now().strftime("%Y")

    # Calcular running balance y stats
    running_balance = 0.0
    picks_html = ""
    w, l, p = 0, 0, 0

    for entry in entries:
        stake = _entry_stake(entry)
        pnl = entry.get("pnl")
        result = entry.get("result") or "⏳"

        # Calcular balance
        if result == "W":
            running_balance += pnl if pnl is not None else stake
            w += 1
        elif result == "L":
            running_balance -= stake
            l += 1
        elif result == "P":
            p += 1
        # else: pending

        _pnl = entry.get("pnl")
        if _pnl is not None:
            bal_fmt = f"+${_pnl:.2f}" if _pnl >= 0 else f"-${abs(_pnl):.2f}"
        elif result == "W":
            bal_fmt = f"+${stake:.2f}"
        elif result == "L":
            bal_fmt = f"-${stake:.2f}"
        else:
            bal_fmt = "—"
        logo_html = _pick_logo(entry.get("pick",""), entry.get("game",""), size=44)
        odds_str  = _fmt_odds(entry.get("odds",""))
        book_str  = entry.get("book","")
        _rc_cls   = {"W":"win","L":"loss","P":"push"}.get(result,"pending")
        _rc_bt    = {"W":"WIN","L":"LOSS","P":"PUSH"}.get(result,"PENDING")
        _meta     = " · ".join(x for x in [book_str, f"Stake: ${stake:.2f}"] if x)

        picks_html += f"""
<div class="rc-pick {_rc_cls}">
  <div class="rc-row">
    {logo_html}
    <div class="rc-main">
      <div class="rc-pick-name">{esc(entry.get("pick",""))}<span class="rc-odds">{esc(odds_str)}</span></div>
      <div class="rc-game">{esc(entry.get("game","")[:40])}</div>
      <div class="rc-meta">{esc(_meta)}</div>
    </div>
    <div class="rc-result-col">
      <span class="rc-badge {_rc_cls}">{_rc_bt}</span>
      <div class="rc-pnl {_rc_cls}">{esc(bal_fmt)}</div>
    </div>
  </div>
</div>"""

    total = w + l + p
    total_pnl = sum(e.get("pnl", 0) for e in entries if e.get("result") in ("W", "L", "P"))
    total_wagered = sum(_entry_stake(e) for e in entries if e.get("result") in ("W", "L", "P"))
    roi = (total_pnl / total_wagered * 100) if total_wagered > 0 else 0
    win_pct = f"{w/total*100:.0f}%" if total else "—"

    pnl_fmt = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    bal_fmt = f"+${running_balance:.2f}" if running_balance >= 0 else f"-${abs(running_balance):.2f}"

    win_col = "#22c55e" if w >= l else "#ef4444"
    pnl_col = "#22c55e" if total_pnl >= 0 else "#ef4444"
    bal_col = "#22c55e" if running_balance >= 0 else "#ef4444"
    roi_col = "#22c55e" if roi >= 0 else "#ef4444"

    _pnl_fmt2 = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    _DIAS_EN  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    _MESES_EN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    if date_str:
        try:
            from datetime import date as _date
            _dobj = _date.fromisoformat(date_str)
            _rc_date_lbl = f"{_DIAS_EN[_dobj.weekday()]}, {_MESES_EN[_dobj.month-1]} {_dobj.day}"
        except Exception:
            _rc_date_lbl = date_str
    else:
        _rc_date_lbl = "All-Time"
    _roi_str = f"{roi:+.1f}%"
    _RC_CSS = """<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',system-ui,sans-serif;background:#080c12;color:#e2e8f0}
.rc-title{text-align:center;padding:12px 0 24px}
.rc-sport-lbl{font-size:1.6rem;font-weight:800;letter-spacing:.03em;color:#fff;margin-bottom:5px}
.rc-date-full{font-size:.78rem;color:#64748b;letter-spacing:.08em;text-transform:uppercase}
.rc-pick{border-radius:14px;padding:14px 16px;margin-bottom:10px;position:relative;
  overflow:hidden;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07)}
.rc-pick.win{border-left:4px solid #22c55e;background:linear-gradient(135deg,rgba(34,197,94,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.loss{border-left:4px solid #ef4444;background:linear-gradient(135deg,rgba(239,68,68,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.push{border-left:4px solid #94a3b8;background:rgba(255,255,255,.04)}
.rc-pick.pending{border-left:4px solid #f59e0b;background:rgba(255,255,255,.04)}
.rc-row{display:flex;align-items:center;gap:12px}
.rc-logo{flex-shrink:0}
.rc-main{flex:1;min-width:0}
.rc-pick-name{font-size:.95rem;font-weight:600;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-odds{font-size:.78rem;color:#64748b;margin-left:6px;font-weight:400}
.rc-game{font-size:.78rem;color:#64748b;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-meta{font-size:.72rem;color:#475569;margin-top:4px;letter-spacing:.02em}
.rc-result-col{display:flex;flex-direction:column;align-items:flex-end;gap:5px;flex-shrink:0}
.rc-badge{font-size:.68rem;font-weight:700;letter-spacing:.1em;padding:3px 9px;border-radius:20px}
.rc-badge.win{background:rgba(34,197,94,.18);color:#22c55e}
.rc-badge.loss{background:rgba(239,68,68,.18);color:#ef4444}
.rc-badge.push{background:rgba(148,163,184,.12);color:#94a3b8}
.rc-badge.pending{background:rgba(245,158,11,.12);color:#f59e0b}
.rc-pnl{font-size:.9rem;font-weight:700}
.rc-pnl.win{color:#22c55e}
.rc-pnl.loss{color:#ef4444}
.rc-pnl.push{color:#94a3b8}
.rc-summary{margin-top:18px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);
  border-radius:14px;padding:18px}
.rc-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;text-align:center}
.rc-stat-lbl{font-size:.65rem;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:5px}
.rc-stat-val{font-size:1.15rem;font-weight:700;line-height:1}
</style>"""
    body = f"""{_RC_CSS}
<div class="rc-title"><div class="rc-sport-lbl">⚾ MLB</div><div class="rc-date-full">{_rc_date_lbl}</div></div>
{picks_html}
<div class="rc-summary">
  <div class="rc-stats">
    <div class="rc-stat">
      <div class="rc-stat-lbl">Record</div>
      <div class="rc-stat-val" style="color:{win_col}">{w}-{l}-{p}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">Win Rate</div>
      <div class="rc-stat-val" style="color:{win_col}">{win_pct}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">Profit/Loss</div>
      <div class="rc-stat-val" style="color:{pnl_col}">{_pnl_fmt2}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">ROI</div>
      <div class="rc-stat-val" style="color:{roi_col}">{_roi_str}</div>
    </div>
  </div>
</div>
"""

    html = _html_wrap(f"Laboy Record Card {date_str}", "MLB", dstr, yr, body)

    # Guardar HTML (token para URL impredecible)
    _tok = _url_token(date_str)
    html_file = f"Laboy Record Card {date_str}-{_tok}.html"
    html_path = os.path.join(SCRIPT_DIR, html_file)

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    return html_path


def export_model_calibration_html():
    """
    Genera 'Laboy Model Record.html' con todos los picks históricos del modelo,
    win%, ROI por edge bucket y tabla completa de picks.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    history = _load_model_picks()
    graded  = [p for p in history if p.get("result") in ("W","L","P")]

    # Calcular stats
    total  = len(graded)
    wins   = sum(1 for p in graded if p["result"]=="W")
    losses = sum(1 for p in graded if p["result"]=="L")
    pushes = sum(1 for p in graded if p["result"]=="P")
    win_pct = wins / (wins + losses) if (wins + losses) else 0

    # ROI estimado (asumiendo todos a -110 si no hay odds guardados)
    def _pnl(p):
        try:
            o = float(str(p["odds"]).replace("+",""))
            r = p["result"]
            if r == "W":
                return (100/abs(o) if o < 0 else o/100)
            if r == "L": return -1.0
            return 0.0
        except: return 0.0
    roi_units = sum(_pnl(p) for p in graded)

    # Edge buckets
    buckets = {"0-2%":[], "2-4%":[], "4-6%":[], "6%+":[] }
    for p in graded:
        try:
            e = abs(float(p.get("edge","0").replace("%","").replace("+","")))
            if e < 2:   buckets["0-2%"].append(p)
            elif e < 4: buckets["2-4%"].append(p)
            elif e < 6: buckets["4-6%"].append(p)
            else:        buckets["6%+"].append(p)
        except: pass

    def _bucket_row(label, items):
        if not items: return ""
        w  = sum(1 for x in items if x["result"]=="W")
        l  = sum(1 for x in items if x["result"]=="L")
        wp = f"{w/(w+l)*100:.0f}%" if (w+l) else "—"
        col = "#22c55e" if (w+l) and w/(w+l) > 0.5 else "#ef4444"
        return f"""<tr>
          <td style="padding:8px 12px;color:var(--muted)">{esc(label)}</td>
          <td style="padding:8px 12px;text-align:center">{len(items)}</td>
          <td style="padding:8px 12px;text-align:center">{w}W / {l}L</td>
          <td style="padding:8px 12px;text-align:center;color:{col};font-weight:700">{wp}</td>
        </tr>"""

    bucket_rows = "".join(_bucket_row(k, v) for k, v in buckets.items())

    # Tabla de picks históricos
    pick_rows = ""
    for p in sorted(history, key=lambda x: x.get("date",""), reverse=True):
        res = p.get("result")
        if res == "W":   res_html = '<span style="color:#22c55e;font-weight:700">✅ W</span>'
        elif res == "L": res_html = '<span style="color:#ef4444;font-weight:700">❌ L</span>'
        elif res == "P": res_html = '<span style="color:#94a3b8;font-weight:700">— P</span>'
        else:            res_html = '<span style="color:#f07820">⏳</span>'
        actual = esc(p.get("actual") or "—")
        pick_rows += f"""<tr>
          <td style="padding:7px 10px;color:var(--muted);font-size:0.8rem">{esc(p.get('date',''))}</td>
          <td style="padding:7px 10px;font-size:0.85rem">{esc(p.get('game',''))}</td>
          <td style="padding:7px 10px;font-weight:600">{esc(p.get('pick',''))}</td>
          <td style="padding:7px 10px;text-align:center;color:var(--accent)">{esc(str(p.get('odds','')))}
          <td style="padding:7px 10px;text-align:center">{esc(p.get('edge',''))}</td>
          <td style="padding:7px 10px;text-align:center">{res_html}</td>
          <td style="padding:7px 10px;color:var(--muted);font-size:0.8rem">{actual}</td>
        </tr>"""

    win_col  = "#22c55e" if win_pct > 0.5 else "#ef4444"
    roi_col  = "#22c55e" if roi_units >= 0 else "#ef4444"
    roi_sign = "+" if roi_units >= 0 else ""

    body = f"""
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:24px">
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">PICKS</div>
        <div style="font-size:1.8rem;font-weight:900">{total}</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">RECORD</div>
        <div style="font-size:1.3rem;font-weight:900">{wins}W / {losses}L / {pushes}P</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">WIN %</div>
        <div style="font-size:1.8rem;font-weight:900;color:{win_col}">{win_pct*100:.0f}%</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">ROI (u)</div>
        <div style="font-size:1.8rem;font-weight:900;color:{roi_col}">{roi_sign}{roi_units:.2f}u</div>
      </div>
    </div>

    <div class="section-title" style="margin-top:0">Calibración por Edge</div>
    <div style="background:#1a1a1a;border-radius:12px;overflow:hidden;margin-bottom:24px">
      <table style="width:100%;border-collapse:collapse">
        <thead><tr style="border-bottom:1px solid #333">
          <th style="padding:10px 12px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">EDGE</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">PICKS</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">W/L</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">WIN%</th>
        </tr></thead>
        <tbody>{bucket_rows}</tbody>
      </table>
    </div>

    <div class="section-title">Histórico de Picks</div>
    <div style="background:#1a1a1a;border-radius:12px;overflow:hidden;overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:0.85rem">
        <thead><tr style="border-bottom:1px solid #333">
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">FECHA</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">JUEGO</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">PICK</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">ODDS</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">EDGE</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">RES</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">SCORE</th>
        </tr></thead>
        <tbody>{pick_rows}</tbody>
      </table>
    </div>"""

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")
    html = _html_wrap("Laboy Model Record — MLB", "MLB", dstr, yr, body)

    fpath = os.path.join(SCRIPT_DIR, "Laboy Model Record - MLB.html")
    with open(fpath, "w", encoding="utf-8") as f:
        f.write(html)
    return fpath


def export_picks_html(picks_ignored, results, odds):
    """Genera Laboy Picks YYYY-MM-DD.html — solo sección EV+ Picks."""
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    def logo(team, size=52):
        url = logo_url(team)
        if not url: return ""
        return (f'<img src="{url}" alt="{esc(team)}" width="{size}" height="{size}" '
                f'style="object-fit:contain" onerror="this.style.display=\'none\'">')

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")

    picks = _compute_picks(results, odds)

    # Separar picks reales (lineup confirmado) de watch list (sin lineup)
    confirmed_picks = [p for p in picks if p.get("lineup_confirmed")]
    watchlist_picks = [p for p in picks if not p.get("lineup_confirmed")]

    # Auto-guardar en historial SOLO picks con lineup confirmado
    if confirmed_picks:
        _model_picks_save_today(confirmed_picks)
    elif watchlist_picks:
        # Si no hay ninguno confirmado, guardar igual pero marcados
        _model_picks_save_today(watchlist_picks)
    # Snapshot para --export-debug: siempre sobreescribe
    _save_debug_state(picks)

    def _pick_card_html(p, watchlist=False):
        """Genera el HTML de una pick-card."""
        is_total  = _is_total_pick(p["pick"])
        disp_team = p["team"] or p["away"]

        if is_total:
            logo_h = _over_under_logo_html(52)
            color  = "#f97316" if not watchlist else "#78716c"
        else:
            logo_h = logo(disp_team, 52)
            color  = _team_color_hex(disp_team) if not watchlist else "#78716c"

        ev_col   = "#22c55e" if not p["ev"].startswith("-") else "#ef4444"
        edge_col = "#22c55e" if not p["edge"].startswith("-") else "#ef4444"

        if watchlist:
            warn_banner = ('<div style="font-size:0.7rem;color:#f59e0b;background:#1c1400;'
                           'border-radius:4px;padding:3px 8px;margin-bottom:6px;display:inline-block">'
                           '⚠️ LINEUP PENDIENTE — esperar antes de apostar</div>')
            opacity = 'opacity:0.75;'
        else:
            warn_banner = ""
            opacity = ""

        return f"""
            <div class="pick-card" style="border-left:4px solid {color};{opacity}">
              {warn_banner}
              <div class="pick-time">{esc(p['time'])}</div>
              <div class="teams-row">
                {logo_h}
                <div class="pick-main">
                  <div class="game-label">{esc(p['game'])}</div>
                  <div class="pick-label">{esc(p['pick'])} <span class="odds-badge">{esc(p['odds'])}</span></div>
                </div>
              </div>
              <div class="stats-grid">
                <div class="stat"><div class="stat-label">Modelo</div><div class="stat-val">{esc(p['modelo'])}</div></div>
                <div class="stat"><div class="stat-label">Mercado</div><div class="stat-val">{esc(p['mercado'])}</div></div>
                <div class="stat"><div class="stat-label">Edge</div><div class="stat-val" style="color:{edge_col}">{esc(p['edge'])}</div></div>
                <div class="stat"><div class="stat-label">EV</div><div class="stat-val" style="color:{ev_col}">{esc(p['ev'])}</div></div>
              </div>
            </div>"""

    if not picks:
        body = '<div class="no-picks">No se encontraron picks EV+ para hoy. Verifica que ODDS_API_KEY esté configurado.</div>'
    else:
        body = ""

        # ── Picks reales (lineup confirmado) ─────────────────────────────
        if confirmed_picks:
            body += '<div class="section-title"><i class="fa-solid fa-bullseye fa-icon"></i>Picks EV+ del Modelo</div>\n'
            for p in confirmed_picks:
                body += _pick_card_html(p, watchlist=False)

        # ── Watch List (sin lineup confirmado) ───────────────────────────
        if watchlist_picks:
            body += ('<div class="section-title" style="color:#f59e0b;margin-top:28px">'
                     '<i class="fa-solid fa-clock fa-icon"></i>'
                     f'Watch List — Lineup Pendiente ({len(watchlist_picks)})</div>\n'
                     '<div style="font-size:0.78rem;color:#94a3b8;margin:-10px 0 14px 0">'
                     'Estos picks tienen edge pero el lineup aún no está confirmado. '
                     'Pueden cambiar. Espera el lineup antes de apostar.</div>\n')
            for p in watchlist_picks:
                body += _pick_card_html(p, watchlist=True)

    # Session suffix en filename y título
    if DAY_SESSION:
        _sess_label  = "DAY"
        _sess_emoji  = "☀️"
        _sess_title  = f"☀️ SESIÓN DÍA · {dstr}"
    elif PM_SESSION:
        _sess_label  = "PM"
        _sess_emoji  = "🌤️"
        _sess_title  = f"🌤️ SESIÓN TARDE · {dstr}"
    elif NIGHT_SESSION:
        _sess_label  = "NIGHT"
        _sess_emoji  = "🌙"
        _sess_title  = f"🌙 SESIÓN NOCHE · {dstr}"
    else:
        _sess_label  = ""
        _sess_emoji  = ""
        _sess_title  = dstr

    _title_tag = f"Laboy Picks {_sess_emoji} · {dstr}".strip() if _sess_emoji else f"Laboy Picks · {dstr}"

    # ── Slate auto-numbering ──────────────────────────────────────────────
    # Si ya existe algún archivo de picks para esta fecha+sesión, genera un
    # nombre único con número de slate (S2, S3…) para nunca sobreescribir.
    import glob as _g, hashlib as _hs
    _prefix = f"Laboy Picks {_sess_label + ' ' if _sess_label else ''}{TARGET_DATE}"
    _existing = sorted(_g.glob(os.path.join(SCRIPT_DIR, f"{_prefix}*.html")))

    if not _existing or FORCE_EXPORT:
        # Primera exportación del día (o force): token basado en fecha (reproducible)
        _tok = _url_token(TARGET_DATE)
        _slate_sfx = ""
    else:
        # Exportaciones subsiguientes: slate N + hash de tiempo (siempre único)
        _slate_n   = len(_existing) + 1
        _tok       = _hs.md5(datetime.now().isoformat().encode()).hexdigest()[:7]
        _slate_sfx = f" S{_slate_n}"

    if _sess_label:
        fname = f"Laboy Picks {_sess_label} {TARGET_DATE}{_slate_sfx}-{_tok}.html"
    else:
        fname = f"Laboy Picks {TARGET_DATE}{_slate_sfx}-{_tok}.html"

    html  = _html_wrap(_title_tag, "MLB", _sess_title, yr, body)
    fpath = os.path.join(SCRIPT_DIR, fname)

    # ── Protección de sobreescritura (solo aplica al primer archivo) ──────
    if os.path.exists(fpath) and not FORCE_EXPORT:
        print(f"  🔒 Picks HTML ya existe para {TARGET_DATE} — protegido de sobreescritura.")
        print(f"     → {fname}")
        print(f"     Usa --force-export para regenerar (ej: --export-picks --publish --force-export)")
        return fpath
    with open(fpath, "w", encoding="utf-8") as f: f.write(html)
    print(f"  🎯 Picks HTML: {fname}")
    return fpath


def export_debug_picks_html(results, odds={}, _dashboard_mode=False):
    """
    Genera Laboy Debug YYYY-MM-DD.html — tarjetas detalladas para cada pick
    recomendado, mostrando todos los inputs del modelo. Fiel a la paleta del
    diseño existente: negro #0a0a0a, cards #222222, stats #181818, acento naranja.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    def logo(team, size=44):
        url = logo_url(team)
        if not url: return ""
        return (f'<img src="{url}" alt="{esc(team)}" width="{size}" height="{size}" '
                f'style="object-fit:contain" onerror="this.style.display=\'none\'">')

    # ── Colores del sistema de diseño ─────────────────────────────────────
    BG_INNER  = "#181818"   # igual que .stat
    BG_DEEP   = "#111111"   # pick badge / more depth
    ACCENT    = "#f07820"   # naranja principal
    ACCENT_BG = "#f0782018" # naranja translúcido para badges
    MUTED     = "#94a3b8"
    DIVIDER   = "#2a2a2a"
    TEXT      = "#f1f5f9"

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")

    # ── Cargar snapshot del último --picks run (fiel a lo que el usuario vio) ──
    # mlb_debug_state.json se sobreescribe cada vez que corres --picks o --export-picks.
    # Sin locking, sin sesiones: siempre refleja la última corrida.
    def _enrich(p):
        """Asegura que el pick tenga away/home/team aunque sea un pick viejo."""
        if not p.get("away") and " @ " in p.get("game", ""):
            parts = p["game"].split(" @ ", 1)
            p = dict(p)
            p.setdefault("away", parts[0].strip())
            p.setdefault("home", parts[1].strip())
        p = dict(p)
        p.setdefault("team", p.get("away", ""))
        return p

    _debug_state = _load_debug_state()
    _state_picks = _debug_state.get("picks", [])

    if _state_picks:
        _sess_state = _debug_state.get("session", "full").lower()
        # Verificar que el state corresponde a la sesión activa
        _curr_sess  = ("day" if DAY_SESSION else "night" if NIGHT_SESSION else
                       "pm" if PM_SESSION else "full")
        _sess_match = (_curr_sess == _sess_state) or \
                      (_curr_sess == "full" and _sess_state in ("full", ""))

        if _sess_match:
            # State es de la misma sesión → usar directamente
            picks = [_enrich(p) for p in _state_picks]
            # Aplicar filtro de lineup confirmado:
            # - --confirmed flag: siempre filtra
            # - _dashboard_mode (llamado desde --picks): filtra si hay picks confirmados;
            #   si no hay ninguno aún, muestra todos como watchlist
            if REQUIRE_LINEUPS:
                picks = [p for p in picks if p.get("lineup_confirmed")]
            elif _dashboard_mode:
                _confirmed = [p for p in picks if p.get("lineup_confirmed")]
                if _confirmed:
                    picks = _confirmed   # solo los confirmados
                # else: no hay ninguno aún → mostrar todos (watchlist) con warning
            # Si el filtro deja vacío, recomputar frescos
            if not picks:
                picks = [_enrich(p) for p in _compute_picks(results, odds)]
        else:
            # State es de sesión DIFERENTE (ej: state=day, exportando night)
            # Ignorar el state completamente y recomputar con los flags activos
            picks = [_enrich(p) for p in _compute_picks(results, odds)]
    else:
        # Sin state → recomputar frescos
        picks = [_enrich(p) for p in _compute_picks(results, odds)]

    game_index = {}
    for r in results:
        game_index[f"{r['away']} @ {r['home']}"] = r

    # ── Mini helpers visuales ─────────────────────────────────────────────
    def _bar(val, lo, hi, col):
        pct = max(0, min(100, (val - lo) / max(hi - lo, 0.01) * 100))
        return (f'<div style="background:{DIVIDER};border-radius:3px;height:5px;margin-top:4px">'
                f'<div style="background:{col};height:5px;border-radius:3px;width:{pct:.0f}%"></div></div>')

    def _xfip_col(v):
        if v is None: return MUTED
        return "#22c55e" if v < 3.5 else ("#ef4444" if v > 4.5 else TEXT)

    def _wrc_col(v):
        return "#22c55e" if v >= 110 else ("#ef4444" if v <= 90 else TEXT)

    def _form_badge(ff):
        if ff >= 1.04: return f'<span style="background:#22c55e20;color:#22c55e;border-radius:4px;padding:1px 6px;font-size:0.7rem;font-weight:700">🔥 HOT</span>'
        if ff <= 0.96: return f'<span style="background:#ef444420;color:#ef4444;border-radius:4px;padding:1px 6px;font-size:0.7rem;font-weight:700">❄️ COLD</span>'
        return f'<span style="background:{DIVIDER};color:{MUTED};border-radius:4px;padding:1px 6px;font-size:0.7rem;font-weight:700">— AVG</span>'

    def _hand_pill(hand):
        # Neutral — solo muestra L/R en muted, sin azul ni naranja
        if not hand: return ""
        return f'<span style="background:{DIVIDER};color:{MUTED};border-radius:3px;padding:1px 5px;font-size:0.68rem;font-weight:700;margin-left:4px">{hand}</span>'

    def _section(icon, title, content):
        return (f'<div style="margin-top:16px;padding-top:14px;'
                f'border-top:1px solid rgba(0,220,255,0.12);'
                f'position:relative">'
                # glowing divider cap
                f'<div style="position:absolute;top:-1px;left:0;width:48px;height:1px;'
                f'background:linear-gradient(90deg,rgba(0,220,255,0.6),transparent)"></div>'
                f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:10px">'
                f'<span style="font-size:0.85rem">{icon}</span>'
                f'<span style="font-size:0.6rem;font-weight:800;letter-spacing:0.14em;'
                f'text-transform:uppercase;'
                f'background:linear-gradient(90deg,#00dcff,#7c9eb8);'
                f'-webkit-background-clip:text;-webkit-text-fill-color:transparent">{title}</span>'
                f'</div>'
                f'{content}</div>')

    def _inner(content, cols=1, gap=8):
        grid = f"grid-template-columns:{'1fr ' * cols}".strip()
        return (f'<div style="display:grid;{grid};gap:{gap}px">{content}</div>')

    def _cell(content):
        return (f'<div style="background:linear-gradient(135deg,#13131e 0%,#0e0e14 100%);'
                f'border-radius:8px;padding:10px 11px;'
                f'border:1px solid rgba(0,220,255,0.07);'
                f'box-shadow:0 2px 8px rgba(0,0,0,0.6) inset,0 0 0 1px rgba(0,220,255,0.03)">'
                f'{content}</div>')

    def _label(t):
        return (f'<div style="font-size:0.6rem;font-weight:800;color:rgba(0,220,255,0.55);'
                f'margin-bottom:5px;text-transform:uppercase;letter-spacing:0.12em;'
                f'text-shadow:0 0 8px rgba(0,220,255,0.25)">{t}</div>')

    def _val(v, col=None):
        c = col or TEXT
        return (f'<div style="font-size:0.9rem;font-weight:700;color:{c};'
                f'text-shadow:0 0 10px rgba(255,255,255,0.08)">{v}</div>')

    if not picks:
        body = '<div class="no-picks">No se encontraron picks EV+ para hoy.</div>'
    else:
        _confirmed_count = sum(1 for p in picks
                               if game_index.get(p["game"],{}).get("lineup_used_away")
                               and game_index.get(p["game"],{}).get("lineup_used_home"))
        _unconfirmed_count = len(picks) - _confirmed_count
        _lineup_summary = ""
        if _unconfirmed_count:
            _lineup_summary = (f'<span style="background:#ef444418;color:#ef4444;border-radius:6px;'
                               f'padding:2px 10px;font-size:0.72rem;font-weight:700;margin-left:10px">'
                               f'⚠️ {_unconfirmed_count} sin lineup confirmado</span>')
        if _confirmed_count:
            _lineup_summary += (f'<span style="background:#22c55e18;color:#22c55e;border-radius:6px;'
                                f'padding:2px 10px;font-size:0.72rem;font-weight:700;margin-left:6px">'
                                f'✅ {_confirmed_count} con lineup</span>')
        body = (f'<div class="section-title" style="display:flex;align-items:center;flex-wrap:wrap;gap:4px">'
                f'<span><i class="fa-solid fa-microscope fa-icon"></i>MODEL DEBUG · {len(picks)} Picks</span>'
                f'{_lineup_summary}</div>\n')

        # Agrupar picks por juego → una card por juego (no una card por pick)
        from collections import OrderedDict as _OD
        _game_groups = _OD()
        for _p in picks:
            _game_groups.setdefault(_p["game"], []).append(_p)

        _gi = 0   # índice para IDs únicos por card
        for _game_key, _game_picks in _game_groups.items():
            _gi += 1
            p       = _game_picks[0]   # primer pick del juego para datos game-level
            r       = game_index.get(_game_key, {})
            away    = p["away"]; home = p["home"]
            lines   = r.get("lines", {})
            weather = r.get("weather", {})

            # Border: naranja si múltiples picks del mismo juego, color de equipo si uno solo
            if len(_game_picks) > 1:
                tc_border = ACCENT
            else:
                tc_border = _team_color_hex(p["team"] or away) if not _is_total_pick(p["pick"]) else ACCENT

            away_logo = logo(away); home_logo = logo(home)

            # Variables para compatibilidad (primer pick)
            ev_col   = "#22c55e" if not p["ev"].startswith("-") else "#ef4444"
            edge_col = "#22c55e" if not p["edge"].startswith("-") else "#ef4444"
            ptype    = p.get("_type","")
            type_lbl = {"ML":"ML","RL":"RL","RL+":"RL+","TOT":"TOT"}.get(ptype, "—")
            candado  = "🔒 " if p.get("candado") else ""
            tormenta = "⛈️ " if p.get("tormenta") else ""

            # ── SP ─────────────────────────────────────────────────────────
            away_sp = r.get("away_sp","TBD"); home_sp = r.get("home_sp","TBD")
            fip_a   = r.get("fip_a");         fip_b   = r.get("fip_b")
            xfip_a  = r.get("xfip_a",4.2);   xfip_b  = r.get("xfip_b",4.2)
            hand_a  = r.get("away_sp_hand");  hand_b  = r.get("home_sp_hand")
            _op_a   = r.get("_opener_away", False)
            _op_b   = r.get("_opener_home", False)

            fip_a_s = f' · FIP {fip_a:.2f}' if fip_a else ""
            fip_b_s = f' · FIP {fip_b:.2f}' if fip_b else ""
            _opener_badge = ('<span style="background:#7c3aed22;color:#a78bfa;border-radius:4px;'
                             'padding:1px 5px;font-size:0.65rem;font-weight:700;margin-left:4px">'
                             'OPENER</span>')
            _rp_fip_note  = '&nbsp;·&nbsp;<span style="color:#a78bfa">RP FIP</span>'

            def _xfip_glow(v, col): return f'text-shadow:0 0 8px {col}60'
            sp_html = _inner(
                _cell(f'{_label(f"{esc(away)} SP {_hand_pill(hand_a)}")}'
                      f'{_val(esc(away_sp[:24]) + (_opener_badge if _op_a else ""))}'
                      f'<div style="font-size:0.75rem;color:{MUTED};margin-top:4px">'
                      f'xFIP <span style="color:{_xfip_col(xfip_a)};font-weight:800;{_xfip_glow(xfip_a,_xfip_col(xfip_a))}">{xfip_a:.2f}</span>{fip_a_s}'
                      + (_rp_fip_note if _op_a else "") + '</div>') +
                _cell(f'{_label(f"{esc(home)} SP {_hand_pill(hand_b)}")}'
                      f'{_val(esc(home_sp[:24]) + (_opener_badge if _op_b else ""))}'
                      f'<div style="font-size:0.75rem;color:{MUTED};margin-top:4px">'
                      f'xFIP <span style="color:{_xfip_col(xfip_b)};font-weight:800;{_xfip_glow(xfip_b,_xfip_col(xfip_b))}">{xfip_b:.2f}</span>{fip_b_s}'
                      + (_rp_fip_note if _op_b else "") + '</div>'),
                cols=2)

            # ── Batting Order ──────────────────────────────────────────────
            _lu_names_a = r.get("lineup_names_away", [])
            _lu_names_b = r.get("lineup_names_home", [])

            def _batting_col(names, team_name, confirmed):
                if not confirmed or not names:
                    return _cell(
                        f'{_label(esc(team_name))}'
                        f'<div style="color:{MUTED};font-size:0.73rem;padding-top:4px">⚠️ Sin lineup confirmado</div>')
                rows = ""
                for i, nm in enumerate(names[:9], 1):
                    _nm = esc(nm[:22] + "…" if len(nm) > 22 else nm)
                    # Top batters (1-3) get a slightly brighter name
                    _nm_col = "#e2e8f0" if i <= 3 else TEXT
                    rows += (
                        f'<div style="display:flex;gap:8px;align-items:center;'
                        f'padding:3px 0;border-bottom:1px solid rgba(0,220,255,0.05);font-size:0.73rem">'
                        f'<span style="color:rgba(0,220,255,0.55);min-width:14px;text-align:right;'
                        f'font-size:0.6rem;font-weight:900;letter-spacing:0.04em">{i}</span>'
                        f'<span style="color:{_nm_col};font-weight:{"600" if i<=3 else "400"}">{_nm}</span>'
                        f'</div>'
                    )
                return _cell(f'{_label(esc(team_name))}{rows}')

            lineup_html = _inner(
                _batting_col(_lu_names_a, away, r.get("lineup_used_away")) +
                _batting_col(_lu_names_b, home, r.get("lineup_used_home")),
                cols=2)

            # ── Offense + Platoon ──────────────────────────────────────────
            wrc_a_base = r.get("wrc_a_base", r.get("wrc_a",100))
            wrc_b_base = r.get("wrc_b_base", r.get("wrc_b",100))
            wrc_a      = r.get("wrc_a", wrc_a_base)
            wrc_b      = r.get("wrc_b", wrc_b_base)
            tA         = lines.get("tA","—"); tB = lines.get("tB","—")
            bp_a       = r.get("bp_a",4.2);  bp_b = r.get("bp_b",4.2)

            def _platoon_str(base, adj, opp_hand):
                if opp_hand and round(adj,1) != round(base,1):
                    return (f'<span style="color:{TEXT};font-weight:700">{base:.0f}</span>'
                            f'<span style="color:{MUTED}"> → </span>'
                            f'<span style="color:{ACCENT};font-weight:700">{adj:.0f}</span>'
                            f'<span style="color:{MUTED};font-size:0.68rem"> vs {opp_hand}HP</span>')
                return f'<span style="color:{TEXT};font-weight:700">{base:.0f}</span>'

            lineup_badge_a = (' <span style="color:#22c55e;font-size:0.65rem">📋 LINEUP</span>'
                               if r.get("lineup_used_away")
                               else ' <span style="color:#ef4444;font-size:0.65rem">⚠️ SIN LINEUP</span>')
            lineup_badge_b = (' <span style="color:#22c55e;font-size:0.65rem">📋 LINEUP</span>'
                               if r.get("lineup_used_home")
                               else ' <span style="color:#ef4444;font-size:0.65rem">⚠️ SIN LINEUP</span>')
            rolling_a    = r.get("rolling_wrc_away")    # float o None
            rolling_b    = r.get("rolling_wrc_home")    # float o None
            rolling_a_pa = r.get("rolling_wrc_away_pa", 0)
            rolling_b_pa = r.get("rolling_wrc_home_pa", 0)
            roll_s_a = (f'&nbsp;·&nbsp;14d <span style="color:{"#22c55e" if float(rolling_a)>=wrc_a_base else "#ef4444"};font-weight:700">'
                        f'{float(rolling_a):.0f}</span>'
                        f'<span style="color:#475569;font-size:0.65rem">&thinsp;PA={rolling_a_pa}</span>'
                        if rolling_a is not None else "")
            roll_s_b = (f'&nbsp;·&nbsp;14d <span style="color:{"#22c55e" if float(rolling_b)>=wrc_b_base else "#ef4444"};font-weight:700">'
                        f'{float(rolling_b):.0f}</span>'
                        f'<span style="color:#475569;font-size:0.65rem">&thinsp;PA={rolling_b_pa}</span>'
                        if rolling_b is not None else "")
            off_html = _inner(
                _cell(f'{_label(f"{esc(away)} Offense")}'
                      f'<div style="font-size:0.78rem;margin-bottom:2px">wRC+ {_platoon_str(wrc_a_base,wrc_a,hand_b)}{lineup_badge_a}</div>'
                      f'{_bar(wrc_a,70,135,_wrc_col(wrc_a))}'
                      f'<div style="font-size:0.73rem;color:{MUTED};margin-top:5px">'
                      f'BP xFIP <span style="color:{_xfip_col(bp_a)};font-weight:700">{bp_a:.2f}</span>'
                      f'&nbsp;·&nbsp;Proj <span style="color:{ACCENT};font-weight:700">{tA}R</span>'
                      f'{roll_s_a}</div>') +
                _cell(f'{_label(f"{esc(home)} Offense")}'
                      f'<div style="font-size:0.78rem;margin-bottom:2px">wRC+ {_platoon_str(wrc_b_base,wrc_b,hand_a)}{lineup_badge_b}</div>'
                      f'{_bar(wrc_b,70,135,_wrc_col(wrc_b))}'
                      f'<div style="font-size:0.73rem;color:{MUTED};margin-top:5px">'
                      f'BP xFIP <span style="color:{_xfip_col(bp_b)};font-weight:700">{bp_b:.2f}</span>'
                      f'&nbsp;·&nbsp;Proj <span style="color:{ACCENT};font-weight:700">{tB}R</span>'
                      f'{roll_s_b}</div>'),
                cols=2)

            # ── Recent Form ────────────────────────────────────────────────
            rf_a = r.get("recent_form_away",{}); rf_b = r.get("recent_form_home",{})
            fa   = r.get("form_a",1.0);          fb   = r.get("form_b",1.0)

            def _form_cell(team, rf, ff):
                if not rf:
                    return _cell(f'{_label(esc(team))}<span style="color:{MUTED}">N/A</span>')
                n    = rf.get("games",0); wp = rf.get("wp",0)
                wins = round(wp * n);    rd = rf.get("run_diff",0)
                rest = rf.get("rest_days",1)
                rd_c = "#22c55e" if rd >= 0 else "#ef4444"
                rd_s = f"+{rd:.1f}" if rd >= 0 else f"{rd:.1f}"
                rest_s = "B2B" if rest == 0 else (f"{rest}d rest" if rest < 3 else f"{rest}d rest 💤")
                return _cell(
                    f'{_label(esc(team))}'
                    f'<span style="color:{TEXT};font-weight:700;font-size:0.85rem">{wins}-{n-wins}</span>'
                    f'<span style="color:{MUTED};font-size:0.75rem"> L{n}</span>'
                    f'<div style="font-size:0.73rem;margin-top:4px;display:flex;gap:8px;flex-wrap:wrap">'
                    f'<span style="color:{rd_c};font-weight:600">RD {rd_s}</span>'
                    f'<span style="color:{MUTED}">{rest_s}</span>'
                    f'{_form_badge(ff)}</div>')

            form_html = _inner(_form_cell(away,rf_a,fa) + _form_cell(home,rf_b,fb), cols=2)

            # ── Context (Park / Weather / Umpire) ─────────────────────────
            pf_base = PARK_FACTORS.get(home, 1.0)
            pf_comb = round(calc_pf_combined(home, weather.get("dir",""), weather.get("mph",0),
                                              weather.get("temp",70), weather.get("humidity",50)), 3)
            pf_col  = "#ef4444" if pf_comb > 1.03 else ("#22c55e" if pf_comb < 0.97 else TEXT)
            _w_raw_dir = weather.get("raw_dir","")
            _w_raw_deg = weather.get("raw_deg","")
            _w_deg_s   = f"({_w_raw_deg}°)" if _w_raw_deg != "" else ""
            _w_src     = weather.get("source","Open-Meteo")
            if weather.get("dir") == "DOME":
                _wi = STADIUM_ROOF.get(home, {})
                w_str = ("🔒 Techo Fijo" if _wi.get("roof") == "fixed_dome" else "🏟️ Retráctil")
                if _wi.get("name"): w_str += f" · {_wi['name']}"
            else:
                w_str = (f"{weather.get('temp','?')}°F  {weather.get('dir','')} {weather.get('mph',0)}mph"
                         f"  ·  {_w_raw_dir}{_w_deg_s}")
            ump_hp  = r.get("ump_hp","—"); ump_f = r.get("ump_factor",1.0)
            ump_s   = "avg" if ump_f == 1.0 else (f"+{(ump_f-1)*100:.1f}% R" if ump_f > 1.0 else f"{(ump_f-1)*100:.1f}% R")
            ump_col = "#ef4444" if ump_f > 1.01 else ("#22c55e" if ump_f < 0.99 else MUTED)

            # standings / BP fatigue / recent ERA
            wp_a   = r.get("win_pct_away", 0.500); wp_b = r.get("win_pct_home", 0.500)
            std_fa = r.get("standings_a", 1.0);    std_fb = r.get("standings_b", 1.0)
            bp_fa  = r.get("bp_fatigue_away", 0.0); bp_fb = r.get("bp_fatigue_home", 0.0)
            r_era_a = r.get("recent_era_away"); r_era_b = r.get("recent_era_home")
            def _std_col(f): return "#ef4444" if f < 0.97 else ("#22c55e" if f > 1.03 else MUTED)
            def _bp_col(ip): return "#ef4444" if ip > 12 else ("#f59e0b" if ip > 8 else "#22c55e")
            def _era_col(era_stats, xfip):
                # era_stats = {"era": float, "ip": float, "n": int} or None
                if not isinstance(era_stats, dict): return MUTED
                era = era_stats["era"]
                return "#ef4444" if era > xfip * 1.15 else ("#22c55e" if era < xfip * 0.85 else MUTED)

            ctx_html = _inner(
                _cell(f'{_label("Park Factor")}'
                      f'<div style="font-size:1rem;font-weight:800;color:{pf_col}">{pf_comb}</div>'
                      f'<div style="font-size:0.67rem;color:{MUTED}">base {pf_base}</div>') +
                _cell(f'{_label("Clima")}'
                      f'<div style="font-size:0.78rem;font-weight:600;color:{TEXT}">{esc(w_str)}</div>') +
                _cell(f'{_label("Umpire HP")}'
                      f'<div style="font-size:0.75rem;font-weight:700;color:{TEXT};line-height:1.3">{esc((ump_hp or "?")[:18])}</div>'
                      f'<div style="font-size:0.7rem;color:{ump_col}">{esc(ump_s)}</div>'),
                cols=3) + _inner(
                _cell(f'{_label(f"Standings {esc(away)}")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_std_col(std_fa)}">{wp_a*100:.1f}%</div>'
                      f'<div style="font-size:0.67rem;color:{MUTED}">factor {std_fa:.3f}</div>') +
                _cell(f'{_label(f"Standings {esc(home)}")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_std_col(std_fb)}">{wp_b*100:.1f}%</div>'
                      f'<div style="font-size:0.67rem;color:{MUTED}">factor {std_fb:.3f}</div>') +
                _cell(f'{_label("BP Fatigue Away")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_bp_col(bp_fa)}">{bp_fa:.1f} IP</div>'
                      f'<div style="font-size:0.67rem;color:{MUTED}">3 días</div>') +
                _cell(f'{_label("BP Fatigue Home")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_bp_col(bp_fb)}">{bp_fb:.1f} IP</div>'
                      f'<div style="font-size:0.67rem;color:{MUTED}">3 días</div>'),
                cols=4) + _inner(
                _cell(f'{_label(f"ERA Reciente {esc(away[:3])} SP")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_era_col(r_era_a, r.get("fip_a") or 4.2)}">'
                      + (f'<div>—</div>' if not isinstance(r_era_a, dict) else
                         f'<div>{r_era_a["era"]:.2f}</div>')
                      + f'<div style="font-size:0.67rem;color:{MUTED}">'
                      + ("—" if not isinstance(r_era_a, dict) else f'{r_era_a["n"]}gs/{r_era_a["ip"]:.0f}IP')
                      + '</div></div>') +          # ← cierra también el div 0.9rem (antes faltaba)
                _cell(f'{_label(f"ERA Reciente {esc(home[:3])} SP")}'
                      f'<div style="font-size:0.9rem;font-weight:800;color:{_era_col(r_era_b, r.get("fip_b") or 4.2)}">'
                      + (f'<div>—</div>' if not isinstance(r_era_b, dict) else
                         f'<div>{r_era_b["era"]:.2f}</div>')
                      + f'<div style="font-size:0.67rem;color:{MUTED}">'
                      + ("—" if not isinstance(r_era_b, dict) else f'{r_era_b["n"]}gs/{r_era_b["ip"]:.0f}IP')
                      + '</div></div>'),  # ← cierra también el div 0.9rem (antes faltaba)
                cols=2)

            # ── Model Output ───────────────────────────────────────────────
            total_proj = lines.get("total","—")
            mtotals    = lines.get("mTotals","—")
            mspread    = lines.get("mSpread","—").replace("AWAY",away).replace("HOME",home)
            win_a      = lines.get("winA","—"); win_b = lines.get("winB","—")
            ml_a_str   = lines.get("ml_pct","—").split(" / ")[0] if lines.get("ml_pct") else "—"
            ml_b_str   = lines.get("ml_pct","—").split(" / ")[1] if lines.get("ml_pct") and "/" in lines.get("ml_pct","") else "—"

            model_html = _inner(
                _cell(f'{_label("Total Proyectado")}'
                      f'<div style="font-size:1.15rem;font-weight:900;color:{ACCENT}">{total_proj}</div>'
                      f'<div style="font-size:0.7rem;color:{MUTED}">{esc(mtotals)}</div>') +
                _cell(f'{_label("Spread Modelo")}'
                      f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{esc(mspread)}</div>') +
                _cell(f'{_label(f"ML {esc(away)}")}'
                      f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{ml_a_str}</div>'
                      f'<div style="font-size:0.68rem;color:{MUTED}">{win_a}% win</div>') +
                _cell(f'{_label(f"ML {esc(home)}")}'
                      f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{ml_b_str}</div>'
                      f'<div style="font-size:0.68rem;color:{MUTED}">{win_b}% win</div>'),
                cols=4)

            # ── Pick rows (una por cada pick del juego) ────────────────────
            pick_rows_html = ""
            special_html   = ""
            for _pp in _game_picks:
                _ev_col   = "#22c55e" if not _pp["ev"].startswith("-")   else "#ef4444"
                _edge_col = "#22c55e" if not _pp["edge"].startswith("-") else "#ef4444"
                _ptype    = _pp.get("_type","")
                _type_lbl = {"ML":"ML","RL":"RL","RL+":"RL+","TOT":"TOT"}.get(_ptype, "—")
                _cand     = "🔒 " if _pp.get("candado")  else ""
                _torm     = "⛈️ " if _pp.get("tormenta") else ""
                # Color de pick individual (barra izquierda interior)
                _pc = (_team_color_hex(_pp["team"] or away)
                       if not _is_total_pick(_pp["pick"]) else ACCENT)
                pick_rows_html += f"""
              <div style="background:linear-gradient(135deg,#0c0c14 0%,#080810 100%);
                          border-left:3px solid {_pc};border-radius:10px;
                          padding:12px 15px;margin-bottom:8px;
                          border:1px solid rgba(0,220,255,0.08);
                          box-shadow:0 2px 12px rgba(0,0,0,0.7) inset,
                                     0 0 0 1px rgba(0,220,255,0.03)">
                <div style="display:flex;justify-content:space-between;align-items:center;gap:10px">
                  <div style="flex:1;min-width:0">
                    <div style="font-size:0.58rem;color:rgba(0,220,255,0.5);letter-spacing:0.14em;
                                font-weight:800;margin-bottom:4px;text-transform:uppercase">{esc(_type_lbl)}</div>
                    <div style="font-size:1.18rem;font-weight:900;color:{TEXT};
                                text-shadow:0 0 18px rgba(255,255,255,0.12);
                                display:flex;align-items:center;gap:8px;flex-wrap:wrap">
                      <span>{_cand}{_torm}{esc(_pp['pick'])}</span>
                      <span style="font-size:0.9rem;font-weight:800;
                                   background:rgba(0,220,255,0.08);
                                   border:1px solid rgba(0,220,255,0.2);
                                   color:#00dcff;border-radius:6px;padding:1px 9px;
                                   letter-spacing:0.03em;
                                   text-shadow:0 0 10px rgba(0,220,255,0.5)">{esc(_pp['odds'])}</span>
                    </div>
                  </div>
                  <div style="display:flex;gap:8px;text-align:center;flex-shrink:0">
                    <div style="background:rgba(0,0,0,0.4);border:1px solid rgba(0,220,255,0.1);
                                border-radius:8px;padding:6px 10px;min-width:52px">
                      <div style="font-size:0.55rem;font-weight:800;letter-spacing:0.1em;
                                  color:rgba(0,220,255,0.5);margin-bottom:3px">EDGE</div>
                      <div style="font-size:0.85rem;font-weight:800;color:{_edge_col};
                                  text-shadow:0 0 10px {_edge_col}60">{esc(_pp['edge'])}</div>
                    </div>
                    <div style="background:rgba(0,0,0,0.4);border:1px solid rgba(0,220,255,0.1);
                                border-radius:8px;padding:6px 10px;min-width:52px">
                      <div style="font-size:0.55rem;font-weight:800;letter-spacing:0.1em;
                                  color:rgba(0,220,255,0.5);margin-bottom:3px">EV</div>
                      <div style="font-size:0.85rem;font-weight:800;color:{_ev_col};
                                  text-shadow:0 0 10px {_ev_col}60">{esc(_pp['ev'])}</div>
                    </div>
                    <div style="background:rgba(0,0,0,0.4);border:1px solid rgba(0,220,255,0.1);
                                border-radius:8px;padding:6px 10px;min-width:56px">
                      <div style="font-size:0.55rem;font-weight:800;letter-spacing:0.1em;
                                  color:rgba(0,220,255,0.5);margin-bottom:3px">MODELO</div>
                      <div style="font-size:0.8rem;font-weight:700;color:{TEXT}">{esc(_pp.get('modelo','—'))}</div>
                    </div>
                  </div>
                </div>
              </div>"""
                # Triple Lock / Perfect Storm badges por pick
                def _ck(k, _x=_pp): return "✅" if _x.get(k) else "❌"
                if _pp.get("candado"):
                    special_html += (f'<div style="background:#22c55e0e;border:1px solid #22c55e33;'
                                     f'border-radius:8px;padding:8px 12px;margin-top:8px;font-size:0.76rem;color:#22c55e">'
                                     f'🔒 <strong>TRIPLE LOCK</strong>&nbsp;&nbsp;'
                                     f'FIP&lt;3.80: {_ck("_tc_fip")}&nbsp;·&nbsp;PF&lt;98: {_ck("_tc_pf")}&nbsp;·&nbsp;wRC+&lt;100: {_ck("_tc_wrc")}</div>')
                if _pp.get("tormenta"):
                    special_html += (f'<div style="background:#f078200e;border:1px solid #f0782033;'
                                     f'border-radius:8px;padding:8px 12px;margin-top:8px;font-size:0.76rem;color:{ACCENT}">'
                                     f'⛈️ <strong>PERFECT STORM</strong>&nbsp;&nbsp;'
                                     f'FIP&gt;4.20: {_ck("_tp_fip")}&nbsp;·&nbsp;PF&gt;102: {_ck("_tp_pf")}&nbsp;·&nbsp;wRC+&gt;105: {_ck("_tp_wrc")}</div>')

            # ── MC button data para este juego ───────────────────────────
            _tA_f  = lines.get("tA") or 0
            _tB_f  = lines.get("tB") or 0
            _dbg_books = _get_game_books(odds, away, home)
            _dbg_bmgm  = _dbg_books.get("BetMGM", {})
            _mc_away_js = away.replace("'", "\\'").replace('"', '\\"')
            _mc_home_js = home.replace("'", "\\'").replace('"', '\\"')

            # Detectar tipo del pick primario para modo MC dinámico
            _primary_pick = _game_picks[0]
            _primary_type = _primary_pick.get("_type", "TOT")

            def _parse_ods(s, default=-110):
                try: return int(str(s or default).replace("+", ""))
                except: return default

            if _primary_type == "ML":
                _mc_mode = "ML"
                # Obtener odds ML de BetMGM; fallback: usar odds del pick para el lado elegido
                _ml_away_raw = _dbg_bmgm.get(f"ML_{away}")
                _ml_home_raw = _dbg_bmgm.get(f"ML_{home}")
                _pick_ods_int = _parse_ods(_primary_pick.get("odds"), -110)
                _is_away_pick = 1 if _primary_pick.get("team") == away else 0
                _ml_away_ods = int(_ml_away_raw) if _ml_away_raw else (_pick_ods_int if _is_away_pick else -110)
                _ml_home_ods = int(_ml_home_raw) if _ml_home_raw else (_pick_ods_int if not _is_away_pick else -110)
                _mc_p1_js = str(_ml_away_ods)
                _mc_p2_js = str(_ml_home_ods)
                _mc_p3_js = str(_is_away_pick)

            elif _primary_type in ("RL", "RL+"):
                _mc_mode = _primary_type
                _pick_team = _primary_pick.get("team", away)
                _is_away_pick = 1 if _pick_team == away else 0
                _is_plus = 1 if _primary_type == "RL+" else 0
                # Obtener RL odds de BetMGM (Spread_{TEAM}); fallback: odds del pick
                _spread_data = _dbg_bmgm.get(f"Spread_{_pick_team}", {})
                _pick_ods_int = _parse_ods(_primary_pick.get("odds"), -110)
                _rl_ods = int(_spread_data.get("odds", _pick_ods_int)) if isinstance(_spread_data, dict) else _pick_ods_int
                _mc_p1_js = str(_rl_ods)
                _mc_p2_js = str(_is_away_pick)
                _mc_p3_js = str(_is_plus)

            else:  # TOT
                _mc_mode = "TOT"
                _dbg_to    = _dbg_bmgm.get("Total_Over", {})
                _dbg_tu    = _dbg_bmgm.get("Total_Under", {})
                _dbg_ods_o = int(_dbg_to.get("odds", -110)) if isinstance(_dbg_to, dict) and _dbg_to.get("odds") else -110
                _dbg_ods_u = int(_dbg_tu.get("odds", -110)) if isinstance(_dbg_tu, dict) and _dbg_tu.get("odds") else -110
                _dbg_mkt = None
                for _pp2 in _game_picks:
                    if _pp2.get("market_total") is not None:
                        _dbg_mkt = _pp2["market_total"]; break
                _mc_p1_js = f"{_dbg_mkt:.1f}" if _dbg_mkt is not None else "null"
                _mc_p2_js = str(_dbg_ods_o)
                _mc_p3_js = str(_dbg_ods_u)

            # Badge de tipo: "2 PICKS" si múltiples, tipo si uno
            # — clickable para toggle del debug detail
            _dbg_id    = f"dbg-detail-{_gi}"
            _multi_badge = (
                f'<span class="dbg-toggle-badge" '
                f'onclick="var _c=this.closest(\'.pick-card\');'
                f'var _d=_c?_c.querySelector(\'[id^=dbg-detail]\'):null;'
                f'if(_d){{_d.style.display=_d.style.display===\'none\'?\'block\':\'none\';}}" '
                f'title="Click para ver/ocultar debug" '
                f'style="background:{ACCENT_BG};color:{ACCENT};border-radius:6px;'
                f'padding:3px 10px;font-size:0.78rem;font-weight:700;letter-spacing:0.05em;'
                f'cursor:pointer;user-select:none">'
                + (f'{len(_game_picks)} PICKS' if len(_game_picks) > 1 else esc(type_lbl))
                + '</span>')

            # Botón Monte Carlo — modo dinámico según tipo de pick
            _mc_btn = (
                f'<button onclick="openMC({_tA_f:.3f},{_tB_f:.3f},&quot;{_mc_away_js}&quot;,&quot;{_mc_home_js}&quot;,&quot;{_mc_mode}&quot;,{_mc_p1_js},{_mc_p2_js},{_mc_p3_js})" '
                f'style="background:rgba(0,220,255,0.1);border:1px solid rgba(0,220,255,0.3);'
                f'color:#00dcff;border-radius:6px;padding:3px 10px;font-size:0.78rem;'
                f'font-weight:700;letter-spacing:0.04em;cursor:pointer;'
                f'transition:background 0.15s">Monte Carlo</button>'
            )

            body += f"""
            <div class="pick-card" style="border-left:4px solid {tc_border}">

              <!-- ── Lineup warning banner ── -->
              {"" if r.get("lineup_used_away") and r.get("lineup_used_home") else
               ('<div style="background:#ef444412;border:1px solid #ef444440;border-radius:8px;'
                'padding:7px 12px;margin-bottom:10px;display:flex;align-items:center;gap:8px">'
                '<span style="font-size:0.95rem">⚠️</span>'
                '<span style="font-size:0.72rem;color:#ef4444;font-weight:700">LINEUP NO CONFIRMADO — '
                + ("AMBOS equipos" if not r.get("lineup_used_away") and not r.get("lineup_used_home")
                   else (away if not r.get("lineup_used_away") else home))
                + " · wRC+ usando baseline"
                + ("  ·  Este pick sería omitido con --confirmed" if not (r.get("lineup_used_away") and r.get("lineup_used_home")) else "")
                + '</span></div>')}

              <!-- ── Game header ── -->
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
                <div style="display:flex;align-items:center;gap:10px">
                  {away_logo}
                  <div>
                    <div class="pick-time">{esc(p['time'])}</div>
                    <div style="font-size:0.92rem;font-weight:700;color:{TEXT}">{esc(p['game'])}</div>
                  </div>
                  {home_logo}
                </div>
                <div style="display:flex;align-items:center;gap:8px">
                  {_mc_btn}
                  {_multi_badge}
                </div>
              </div>

              <!-- ── Pick rows (una por pick del juego) ── -->
              {pick_rows_html}

              <!-- ── Debug sections (colapsables) ── -->
              <div id="{_dbg_id}" style="display:none">
                {_section("⚾", "Starting Pitchers", sp_html)}
                {_section("📋", "Batting Order", lineup_html)}
                {_section("🏏", "Offense · wRC+ &amp; Platoon", off_html)}
                {_section("📈", "Recent Form · L10", form_html)}
                {_section("🌤️", "Contexto", ctx_html)}
                {_section("🎯", "Model Output", model_html)}
              </div>
              {special_html}
            </div>"""

    # ── Guardar fragmento del body para el dashboard (siempre sobreescribe) ──
    # mlb_debug_body_current.html: solo el body + CSS mínimo, sin full-page wrap.
    # serve.py lo inyecta directamente en el panel de picks del dashboard.
    _FRAG_CSS = """
<style>
/* ── Debug fragment — inyectado en el dashboard ── */
.dbg-frag-wrap {
  font-family: inherit;
  color: #f1f5f9;
  --muted: #94a3b8;
  --cyan: #00dcff;
  --accent: #f07820;
}

/* Card */
.dbg-frag-wrap .pick-card {
  background: linear-gradient(160deg,#0d0d12 0%,#090910 100%);
  border: 1px solid rgba(0,220,255,.13);
  border-radius: 14px;
  padding: 18px 20px;
  margin-bottom: 20px;
  box-shadow:
    0 0 0 1px rgba(0,220,255,.04),
    0 6px 40px rgba(0,0,0,.85),
    inset 0 1px 0 rgba(255,255,255,.04);
  position: relative;
  overflow: hidden;
}
/* Subtle top glow line on each card */
.dbg-frag-wrap .pick-card::before {
  content: '';
  position: absolute;
  top: 0; left: 10%; right: 10%;
  height: 1px;
  background: linear-gradient(90deg,transparent,rgba(0,220,255,.35),rgba(240,120,32,.25),transparent);
}

/* Game header time */
.dbg-frag-wrap .pick-time {
  font-size: .65rem;
  color: rgba(0,220,255,.55);
  letter-spacing: .08em;
  font-weight: 700;
  margin-bottom: 2px;
  text-transform: uppercase;
}

/* Section header row (MODEL DEBUG · N PICKS) */
.dbg-frag-wrap .section-title {
  font-size: .62rem;
  font-weight: 900;
  letter-spacing: .15em;
  text-transform: uppercase;
  background: linear-gradient(90deg,#00dcff,#7c3aed,#00dcff);
  background-size: 200% auto;
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  animation: dbg-frag-grad 5s linear infinite;
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 8px;
}
@keyframes dbg-frag-grad {
  0%   { background-position: 0% center }
  100% { background-position: 200% center }
}

/* Inner stat cells (EDGE/EV/MODELO) */
.dbg-frag-wrap .stat {
  background: rgba(0,0,0,.45);
  border: 1px solid rgba(0,220,255,.1);
  border-radius: 8px;
  padding: 7px 11px;
}
.dbg-frag-wrap .stat-label {
  font-size: .55rem;
  font-weight: 800;
  letter-spacing: .1em;
  color: rgba(0,220,255,.5);
  margin-bottom: 3px;
  text-transform: uppercase;
}
.dbg-frag-wrap .stat-val {
  font-size: .85rem;
  font-weight: 800;
}

/* Odds badge used in some places */
.dbg-frag-wrap .odds-badge {
  background: rgba(0,220,255,.08);
  border: 1px solid rgba(0,220,255,.22);
  color: #00dcff;
  border-radius: 6px;
  padding: 1px 9px;
  font-size: .9rem;
  font-weight: 800;
  letter-spacing: .03em;
  text-shadow: 0 0 10px rgba(0,220,255,.5);
}

/* Game title text */
.dbg-frag-wrap [style*="font-size:0.92rem;font-weight:700"] {
  text-shadow: 0 0 20px rgba(255,255,255,.07);
}

/* Team logo shimmer glow */
.dbg-frag-wrap img[alt] {
  filter: drop-shadow(0 0 6px rgba(0,220,255,.18));
}

/* wRC+ bar */
.dbg-frag-wrap [style*="border-radius:3px;height:5px"] {
  box-shadow: 0 0 6px rgba(0,220,255,.1);
}

/* Collapsible badge */
.dbg-frag-wrap .dbg-toggle-badge:hover {
  background: rgba(240,120,32,.25) !important;
  box-shadow: 0 0 8px rgba(240,120,32,.3);
}
</style>"""
    _frag_content = f'{_FRAG_CSS}<div class="dbg-frag-wrap">{body}</div>'
    _frag_path = os.path.join(SCRIPT_DIR, "mlb_debug_body_current.html")
    try:
        with open(_frag_path, "w", encoding="utf-8") as _ff:
            _ff.write(_frag_content)
    except Exception as _fe:
        pass  # non-critical

    # Incluir sesión DAY/PM/NIGHT en el nombre — igual que export_picks_html
    _dbg_sess = "DAY " if DAY_SESSION else ("PM " if PM_SESSION else ("NIGHT " if NIGHT_SESSION else ""))

    # ── Slate auto-numbering (mismo patrón que export_picks_html) ─────────
    import glob as _g, hashlib as _hs
    _dbg_prefix = f"Laboy Debug {_dbg_sess}{TARGET_DATE}"
    _dbg_existing = sorted(_g.glob(os.path.join(SCRIPT_DIR, f"{_dbg_prefix}*.html")))

    if not _dbg_existing or FORCE_EXPORT:
        _dbg_tok      = _url_token(TARGET_DATE)
        _dbg_slate    = ""
    else:
        _dbg_slate_n  = len(_dbg_existing) + 1
        _dbg_tok      = _hs.md5(datetime.now().isoformat().encode()).hexdigest()[:7]
        _dbg_slate    = f" S{_dbg_slate_n}"

    html  = _debug_html_wrap(f"Laboy Debug · {dstr}", dstr, yr, body)
    fname = f"Laboy Debug {_dbg_sess}{TARGET_DATE}{_dbg_slate}-{_dbg_tok}.html"
    fpath = os.path.join(SCRIPT_DIR, fname)
    if os.path.exists(fpath) and not FORCE_EXPORT:
        print(f"  🔒 Debug HTML ya existe para {TARGET_DATE} ({_dbg_sess.strip() or 'FULL'}) — protegido de sobreescritura.")
        print(f"     → {fname}")
        print(f"     Usa --force-export para regenerar.")
        return fpath
    with open(fpath, "w", encoding="utf-8") as f: f.write(html)
    print(f"  🔍 Debug HTML: {fname}")
    # ── Guardar fuente de verdad para --log (solo se actualiza al generar el HTML) ──
    _save_log_state(picks)
    return fpath


def export_lines_html(results, odds={}):
    """Genera Laboy Lines YYYY-MM-DD.html con stat-boxes y odds de casas."""
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    def logo(team, size=36):
        url = logo_url(team)
        if not url: return ""
        return (f'<img src="{url}" alt="{esc(team)}" width="{size}" height="{size}" '
                f'style="object-fit:contain" onerror="this.style.display=\'none\'">')

    # Sportsbooks a mostrar (sin Pinnacle — ese es interno)
    BOOKS_DISPLAY = [
        ("BetMGM",     "#b59a2a"),   # dorado
        ("FanDuel",    "#1493ff"),   # azul
        ("DraftKings", "#53d337"),   # verde
    ]

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")

    sorted_results = sorted(results, key=lambda r: _parse_time_sort(r.get("game_time_local","")))

    body = '<div class="section-title"><i class="fa-solid fa-chart-simple fa-icon"></i>Data Model Lines — Todos los Juegos</div>\n'

    for r in sorted_results:
        away, home = r["away"], r["home"]
        ld    = r["lines"]
        wth   = r["weather"]
        gtime = _to_et(r.get("game_time_local", ""))

        # Weather
        if wth.get("dir") == "DOME":
            _roof_info = STADIUM_ROOF.get(r.get("home",""), {})
            _roof_type = _roof_info.get("roof", "retractable")
            _roof_name = _roof_info.get("name", "")
            weather_s  = ("🔒 Techo Fijo" if _roof_type == "fixed_dome" else "🏟️ Techo Retráctil")
            if _roof_name:
                weather_s += f" · {_roof_name}"
        elif wth.get("temp") is not None and wth.get("dir"):
            _raw_dir = wth.get("raw_dir","")
            _raw_deg = wth.get("raw_deg","")
            _deg_s   = f"({_raw_deg}°)" if _raw_deg != "" else ""
            _src     = wth.get("source","Open-Meteo")
            weather_s = (f"🌤️ {wth['temp']}°F · {wth['dir']} {wth['mph']}mph"
                         f"  ·  {_raw_dir}{_deg_s} desde {_src}")
        else:
            weather_s = ""
        # Umpire
        ump_hp = r.get("ump_hp", "")
        ump_f  = r.get("ump_factor", 1.0)
        if ump_hp:
            ump_tend = ("⚖️ avg" if ump_f == 1.0
                        else (f"🔴 +{(ump_f-1)*100:.1f}% run-friendly" if ump_f > 1.0
                              else f"🟢 {(ump_f-1)*100:.1f}% pitcher-friendly"))
            ump_s = f"👨‍⚖️ HP: {ump_hp.title()} · {ump_tend}"
        else:
            ump_s = ""

        sp_row = f"{r.get('away_sp','TBD')} vs {r.get('home_sp','TBD')}"

        # ── Stat boxes (modelo) ────────────────────────────────────────────
        win_a = ld["winA"]; win_b = ld["winB"]
        ml_parts  = ld["ml_pct"].split(" / ")
        ml_a_mod  = ml_parts[0] if len(ml_parts) > 0 else "—"
        ml_b_mod  = ml_parts[1] if len(ml_parts) > 1 else "—"
        total_s   = ld["mTotals"]                                     # "O 8.5" / "U 8.5"
        spread_s  = ld["mSpread"].replace("AWAY", away).replace("HOME", home)

        # Color: verde para el favorito del modelo, gris para el underdog
        a_col = "#22c55e" if win_a >= win_b else "#94a3b8"
        b_col = "#22c55e" if win_b >  win_a else "#94a3b8"
        tot_col = "#f59e0b" if total_s.startswith("O") else "#60a5fa"  # amber=over, blue=under

        # Spread: resalta equipo favorito
        spread_col = "#f07820"

        stats_html = f"""
          <div class="stats-grid" style="margin-top:10px">
            <div class="stat">
              <div class="stat-label">{esc(away[:12])}</div>
              <div class="stat-val" style="color:{a_col}">{esc(ml_a_mod)}</div>
              <div style="font-size:0.62rem;color:var(--muted);margin-top:3px">{win_a:.0f}% win</div>
            </div>
            <div class="stat">
              <div class="stat-label">Total</div>
              <div class="stat-val" style="color:{tot_col}">{esc(total_s)}</div>
            </div>
            <div class="stat">
              <div class="stat-label">Spread</div>
              <div class="stat-val" style="font-size:0.8rem;color:{spread_col}">{esc(spread_s)}</div>
            </div>
            <div class="stat">
              <div class="stat-label">{esc(home[:12])}</div>
              <div class="stat-val" style="color:{b_col}">{esc(ml_b_mod)}</div>
              <div style="font-size:0.62rem;color:var(--muted);margin-top:3px">{win_b:.0f}% win</div>
            </div>
          </div>"""

        # ── Sportsbook ML odds ─────────────────────────────────────────────
        books_data = _get_game_books(odds, away, home)

        # Favicons de las casas (pequeño logo top-left del chip)
        BOOK_FAVICON = {
            "BetMGM":     "https://www.betmgm.com/favicon.ico",
            "FanDuel":    "https://www.fanduel.com/favicon.ico",
            "DraftKings": "https://www.draftkings.com/favicon.ico",
        }

        def _book_logo_img(bname, bcolor):
            fav = BOOK_FAVICON.get(bname, "")
            fallback_txt = esc(bname)
            if fav:
                # Solo oculta la imagen si falla — el texto de bname siempre está al lado
                return (
                    f'<img src="{fav}" width="14" height="14" '
                    f'style="object-fit:contain;border-radius:2px;vertical-align:middle" '
                    f'onerror="this.style.display=\'none\'">'
                )
            return ""  # sin favicon, solo el texto que ya está en el header

        def _odds_side(team, ml_val):
            """Logo equipo + odds, layout vertical centrado."""
            t_url  = logo_url(team)
            t_abb  = team[:3].upper()
            ml_s   = _fmt_odds(ml_val)
            ml_col = "#22c55e" if ml_val < 0 else "#94a3b8"
            if t_url:
                img_html = (
                    f'<img src="{t_url}" alt="{esc(t_abb)}" width="26" height="26" '
                    f'style="object-fit:contain" '
                    f'onerror="this.outerHTML=\'<span style=&quot;font-size:0.58rem;'
                    f'color:#888;font-weight:700&quot;>{esc(t_abb)}</span>\'">'
                )
            else:
                img_html = f'<span style="font-size:0.58rem;color:#888;font-weight:700">{esc(t_abb)}</span>'
            return (
                f'<div style="display:flex;flex-direction:column;align-items:center;gap:3px">'
                f'  {img_html}'
                f'  <span style="font-family:monospace;font-size:0.88rem;font-weight:700;'
                f'color:{ml_col};letter-spacing:-0.3px">{esc(ml_s)}</span>'
                f'</div>'
            )

        book_rows = []
        for bname, bcolor in BOOKS_DISPLAY:
            bk    = books_data.get(bname, {})
            bml_a = bk.get(f"ML_{away}"); bml_h = bk.get(f"ML_{home}")
            if bml_a and bml_h:
                side_a = _odds_side(away, bml_a)
                side_h = _odds_side(home, bml_h)
                blogo  = _book_logo_img(bname, bcolor)
                book_rows.append(
                    f'<div style="background:#161616;border:1px solid {bcolor}33;'
                    f'border-radius:9px;padding:8px 14px">'
                    # ── Book name header ──
                    f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:7px">'
                    f'  {blogo}'
                    f'  <span style="color:{bcolor};font-weight:800;font-size:0.62rem;'
                    f'letter-spacing:0.6px;vertical-align:middle">{esc(bname)}</span>'
                    f'</div>'
                    # ── Teams + odds ──
                    f'<div style="display:flex;align-items:center;justify-content:space-around">'
                    f'  {side_a}'
                    f'  <span style="color:#444;font-size:0.7rem;padding:0 6px">vs</span>'
                    f'  {side_h}'
                    f'</div>'
                    f'</div>'
                )
        books_html = ""
        if book_rows:
            books_html = (
                '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-top:10px">'
                + "".join(book_rows) +
                '</div>'
            )

        body += f"""
        <div class="line-card">
          <div class="line-header">
            <div class="team-logo">{logo(away)}<span>{esc(away)}</span></div>
            <div class="line-time">{esc(gtime)}</div>
            <div class="team-logo">{logo(home)}<span>{esc(home)}</span></div>
          </div>
          <div class="sp-row">{esc(sp_row)}</div>
          {stats_html}
          {f'<div class="weather-row" style="margin-top:8px">{esc(weather_s)}</div>' if weather_s else ""}
          {f'<div class="weather-row" style="margin-top:4px;color:#94a3b8;font-size:0.78rem">{esc(ump_s)}</div>' if ump_s else ""}
          {books_html}
        </div>"""

    html  = _html_wrap(f"Laboy Lines · {dstr}", "MLB", dstr, yr, body)
    fname = f"Laboy Lines {TARGET_DATE}-{_url_token(TARGET_DATE)}.html"
    fpath = os.path.join(SCRIPT_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f: f.write(html)
    print(f"  📊 Lines HTML: {fname}")
    return fpath


def export_html(results, odds={}):
    """Genera ambos HTMLs (picks + lines) y retorna los dos paths."""
    p = export_picks_html(None, results, odds)
    l = export_lines_html(results, odds)
    return p, l


# ──────────────────────────────────────────────────────
# HTML → JPG  (playwright)
# ──────────────────────────────────────────────────────

def html_to_jpg(html_path, width=800, scale=2):
    """
    Convierte un HTML file a JPG usando playwright (Chromium headless).
    Retorna el path del JPG o None si playwright no está disponible.

    Instalación (una sola vez):
      pip install playwright --break-system-packages
      playwright install chromium
    """
    jpg_path = html_path.replace(".html", ".jpg")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  💡 Para generar JPG: pip install playwright --break-system-packages && playwright install chromium")
        return None
    try:
        # Leer HTML como texto — evita el bug de '#' en filename con file:// URLs
        with open(html_path, "r", encoding="utf-8") as _f:
            html_content = _f.read()

        with sync_playwright() as pw:
            browser = pw.chromium.launch(args=["--no-sandbox","--disable-dev-shm-usage"])
            # device_scale_factor=2 → doble resolución (Retina), texto e imágenes nítidos
            page    = browser.new_page(
                viewport={"width": width, "height": 900},
                device_scale_factor=scale
            )
            page.set_content(html_content, wait_until="domcontentloaded")
            # Espera a que carguen imágenes remotas (logos de equipos)
            try:
                page.wait_for_load_state("networkidle", timeout=6000)
            except Exception:
                pass
            # Ajusta la altura al contenido real
            height = page.evaluate("document.body.scrollHeight")
            page.set_viewport_size({"width": width, "height": max(height, 400)})
            # Screenshot full-page → PNG en memoria → JPG guardado
            png_bytes = page.screenshot(full_page=True)
            browser.close()

        from PIL import Image
        import io
        img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
        img.save(jpg_path, "JPEG", quality=95, optimize=True)
        return jpg_path
    except Exception as e:
        print(f"  ⚠️  html_to_jpg falló: {e}")
        return None


# ──────────────────────────────────────────────────────
# DEBUG CARD — para adjuntar al pick logueado
# ──────────────────────────────────────────────────────

def _build_single_game_debug_html(r, logged_pick_str, pick_date=None):
    """
    Genera el HTML del debug card para un solo juego.
    `r`               — el result dict del modelo (de _load_results_cache)
    `logged_pick_str` — el pick tal como quedó en el log (normalizado, ej: "UNDER 9.0")
    `pick_date`       — fecha del pick (YYYY-MM-DD) para que _compute_picks use fecha correcta
    Solo muestra la fila del pick que coincide con logged_pick_str (descarta otros picks del juego).
    Retorna HTML string o "" si falla.
    """
    try:
        def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

        # ── AI palette ────────────────────────────────────────────────────────
        BG_INNER  = "#0e0e14"     # deep navy-black for cells
        BG_DEEP   = "#08080d"     # pick row background
        ACCENT    = "#f07820"     # orange primary
        CYAN      = "#00dcff"     # AI cyan accent
        MUTED     = "#4a6272"     # muted text
        MUTED2    = "#7a8fa0"     # slightly brighter muted
        DIVIDER   = "#151520"     # dividers
        TEXT      = "#e8eef4"     # primary text

        def _bar(val, lo, hi, col):
            pct = max(0, min(100, (val - lo) / max(hi - lo, 0.01) * 100))
            return (f'<div style="background:{DIVIDER};border-radius:3px;height:5px;margin-top:4px">'
                    f'<div style="background:linear-gradient(90deg,{col},{col}aa);height:5px;'
                    f'border-radius:3px;width:{pct:.0f}%;box-shadow:0 0 4px {col}66"></div></div>')

        def _xfip_col(v):
            if v is None: return MUTED
            return "#22c55e" if v < 3.5 else ("#ef4444" if v > 4.5 else TEXT)

        def _wrc_col(v):
            return "#22c55e" if v >= 110 else ("#ef4444" if v <= 90 else TEXT)

        def _section(icon, title, content):
            return (f'<div style="margin-top:14px;padding-top:11px;'
                    f'border-top:1px solid {DIVIDER}">'
                    f'<div style="font-size:0.58rem;font-weight:800;letter-spacing:0.18em;'
                    f'text-transform:uppercase;margin-bottom:8px;'
                    f'background:linear-gradient(90deg,{CYAN},{MUTED2});'
                    f'-webkit-background-clip:text;-webkit-text-fill-color:transparent;">'
                    f'{icon}&nbsp;{title}</div>'
                    f'{content}</div>')

        def _inner(content, cols=1, gap=8):
            grid = f"grid-template-columns:{'1fr ' * cols}".strip()
            return (f'<div style="display:grid;{grid};gap:{gap}px">{content}</div>')

        def _cell(content):
            return (f'<div style="background:{BG_INNER};border:1px solid rgba(0,220,255,0.07);'
                    f'border-radius:8px;padding:9px;'
                    f'box-shadow:0 0 0 1px rgba(0,220,255,0.03),0 2px 8px rgba(0,0,0,0.6) inset">'
                    f'{content}</div>')

        def _label(t):
            return f'<div style="font-size:0.59rem;color:{MUTED};margin-bottom:3px;text-transform:uppercase;letter-spacing:0.1em;font-family:monospace">{t}</div>'

        def _val(v, col=None):
            c = col or TEXT
            return f'<div style="font-size:0.88rem;font-weight:700;color:{c};font-family:\'JetBrains Mono\',\'Fira Code\',monospace">{v}</div>'

        def logo_img(team, size=36):
            url = logo_url(team)
            if not url: return ""
            return (f'<img src="{url}" alt="{esc(team)}" width="{size}" height="{size}" '
                    f'style="object-fit:contain" onerror="this.style.display:\'none\'">')

        away = r.get("away", ""); home = r.get("home", "")
        lines   = r.get("lines", {})
        weather = r.get("weather", {})

        # ── Buscar pick en mlb_model_picks.json (ya tiene EDGE/EV/MODELO) ────
        logged_norm = _normalize_pick_str(logged_pick_str).upper().strip()
        game_key    = f"{away} @ {home}".upper()
        matching_pick = None

        # 1) Buscar en historial de picks del modelo (más rápido, no necesita API)
        try:
            _hist = _load_model_picks()
            _date_filter = pick_date or TARGET_DATE
            for _hp in _hist:
                if (_hp.get("date") == _date_filter and
                        _hp.get("game","").upper() == game_key and
                        _normalize_pick_str(_hp.get("pick","")).upper().strip() == logged_norm):
                    # Convertir formato de historial al formato que espera el HTML
                    matching_pick = {
                        "pick":   _hp.get("pick",""),
                        "odds":   _hp.get("odds",""),
                        "edge":   _hp.get("edge","—"),
                        "ev":     _hp.get("ev","—"),
                        "modelo": _hp.get("modelo","—"),
                        "_type":  ("TOT" if _is_total_pick(_hp.get("pick","")) else "ML"),
                        "team":   None,
                        "away":   away,
                        "home":   home,
                        "candado": False,
                        "tormenta": False,
                    }
                    break
        except Exception:
            pass

        # 2) Fallback: intentar _compute_picks con TARGET_DATE correcto
        if matching_pick is None:
            try:
                import mlb as _self; _saved_td = _self.TARGET_DATE
                if pick_date and len(pick_date) == 10:
                    _self.TARGET_DATE = pick_date
                if pick_date and len(pick_date) == 10:
                    TARGET_DATE = pick_date
                # Build minimal odds so _compute_picks doesn't skip the game
                _mock_odds = {
                    f"{away} vs {home}": {"books": {"BetMGM": {
                        "ml_away": r.get("lines",{}).get("ml_away",-143),
                        "ml_home": r.get("lines",{}).get("ml_home", 143),
                        "total":   r.get("lines",{}).get("total_line", 9.0),
                        "over_odds":  r.get("lines",{}).get("over_odds",-110),
                        "under_odds": r.get("lines",{}).get("under_odds",-110),
                    }}}
                }
                all_picks = _compute_picks([r], _mock_odds)
                for p in all_picks:
                    if _normalize_pick_str(p.get("pick","")).upper().strip() == logged_norm:
                        matching_pick = p
                        break
                if matching_pick is None and all_picks:
                    matching_pick = all_picks[0]
            except Exception:
                pass
            finally:
                try: _self.TARGET_DATE = _saved_td
                except: pass

        if matching_pick is None:
            # Pick manual no recomendado por el modelo — igual mostramos el debug card
            # con todos los datos del juego (SP, lineup, wRC+), EV/EDGE como "—"
            _pick_type = "TOT" if _is_total_pick(logged_pick_str) else \
                         ("RL" if any(x in logged_norm for x in ["-1.5", "+1.5", "-2.5", "+2.5"]) else "ML")
            matching_pick = {
                "pick":    logged_pick_str,
                "odds":    "—",
                "edge":    "—",
                "ev":      "—",
                "modelo":  "—",
                "_type":   _pick_type,
                "team":    None,
                "away":    away,
                "home":    home,
                "candado": False,
                "tormenta": False,
                "_manual": True,   # flag: pick fuera del modelo
            }

        # ── Pick row ───────────────────────────────────────────────────────
        _pp = matching_pick
        _ev_col   = "#22c55e" if not str(_pp.get("ev","")).startswith("-") else "#ef4444"
        _edge_col = "#22c55e" if not str(_pp.get("edge","")).startswith("-") else "#ef4444"
        _ptype    = _pp.get("_type","")
        _type_lbl = {"ML":"ML","RL":"RL","RL+":"RL+","TOT":"TOT"}.get(_ptype, "—")
        _pc = (_team_color_hex(_pp.get("team") or away)
               if not _is_total_pick(_pp.get("pick","")) else ACCENT)
        pick_row_html = f"""
          <div style="background:{BG_DEEP};border-left:3px solid {_pc};border-radius:10px;
                      padding:11px 14px;margin-bottom:8px;
                      box-shadow:0 2px 8px rgba(0,0,0,0.6) inset">
            <div style="display:flex;justify-content:space-between;align-items:center">
              <div>
                <div style="font-size:0.6rem;color:{MUTED};letter-spacing:0.1em;margin-bottom:3px">{esc(_type_lbl)}</div>
                <div style="font-size:1.2rem;font-weight:900;color:{TEXT}">
                  {esc(_pp.get('pick',''))}&nbsp;<span class="odds-badge">{esc(str(_pp.get('odds','')))}</span>
                </div>
              </div>
              <div style="display:flex;gap:10px;text-align:center">
                <div class="stat" style="min-width:52px">
                  <div class="stat-label">EDGE</div>
                  <div class="stat-val" style="color:{_edge_col}">{esc(str(_pp.get('edge','—')))}</div>
                </div>
                <div class="stat" style="min-width:52px">
                  <div class="stat-label">EV</div>
                  <div class="stat-val" style="color:{_ev_col}">{esc(str(_pp.get('ev','—')))}</div>
                </div>
                <div class="stat" style="min-width:52px">
                  <div class="stat-label">MODELO</div>
                  <div class="stat-val" style="font-size:0.78rem">{esc(str(_pp.get('modelo','—')))}</div>
                </div>
              </div>
            </div>
          </div>"""

        # ── SP section ─────────────────────────────────────────────────────
        away_sp = r.get("away_sp","TBD"); home_sp = r.get("home_sp","TBD")
        xfip_a  = r.get("xfip_a", 4.2);  xfip_b  = r.get("xfip_b", 4.2)
        hand_a  = r.get("away_sp_hand");  hand_b  = r.get("home_sp_hand")
        fip_a   = r.get("fip_a");         fip_b   = r.get("fip_b")
        fip_a_s = f' · FIP {fip_a:.2f}' if fip_a else ""
        fip_b_s = f' · FIP {fip_b:.2f}' if fip_b else ""
        def _hand_pill(hand):
            if not hand: return ""
            return f'<span style="background:{DIVIDER};color:{MUTED};border-radius:3px;padding:1px 5px;font-size:0.68rem;font-weight:700;margin-left:4px">{hand}</span>'
        sp_html = _inner(
            _cell(f'{_label(f"{esc(away)} SP {_hand_pill(hand_a)}")}'
                  f'{_val(esc(away_sp[:24]))}'
                  f'<div style="font-size:0.75rem;color:{MUTED};margin-top:3px">'
                  f'xFIP <span style="color:{_xfip_col(xfip_a)};font-weight:700">{xfip_a:.2f}</span>{fip_a_s}</div>') +
            _cell(f'{_label(f"{esc(home)} SP {_hand_pill(hand_b)}")}'
                  f'{_val(esc(home_sp[:24]))}'
                  f'<div style="font-size:0.75rem;color:{MUTED};margin-top:3px">'
                  f'xFIP <span style="color:{_xfip_col(xfip_b)};font-weight:700">{xfip_b:.2f}</span>{fip_b_s}</div>'),
            cols=2)

        # ── Batting Order section ──────────────────────────────────────────
        _lu_names_a = r.get("lineup_names_away", [])
        _lu_names_b = r.get("lineup_names_home", [])
        def _batting_list(names, team, confirmed):
            if not confirmed or not names:
                return _cell(f'{_label(esc(team))}'
                             f'<div style="color:{MUTED};font-size:0.7rem;padding-top:3px">⚠️ Sin lineup</div>')
            rows = ""
            for i, nm in enumerate(names[:9], 1):
                _nm = esc(nm[:20] + "…" if len(nm) > 20 else nm)
                rows += (f'<div style="display:flex;gap:6px;align-items:center;padding:2px 0;'
                         f'border-bottom:1px solid {DIVIDER};font-size:0.7rem">'
                         f'<span style="color:{CYAN};min-width:12px;font-size:0.6rem;'
                         f'font-weight:800;font-family:monospace">{i}</span>'
                         f'<span style="color:{TEXT}">{_nm}</span></div>')
            return _cell(f'{_label(esc(team))}{rows}')
        lineup_html = _inner(
            _batting_list(_lu_names_a, away, r.get("lineup_used_away")) +
            _batting_list(_lu_names_b, home, r.get("lineup_used_home")),
            cols=2)

        # ── Offense section ────────────────────────────────────────────────
        wrc_a = r.get("wrc_a", 100); wrc_b = r.get("wrc_b", 100)
        bp_a  = r.get("bp_a", 4.2);  bp_b  = r.get("bp_b", 4.2)
        tA    = lines.get("tA","—"); tB = lines.get("tB","—")
        off_html = _inner(
            _cell(f'{_label(f"{esc(away)} Offense")}'
                  f'<div style="font-size:0.78rem;margin-bottom:2px">wRC+ <span style="color:{_wrc_col(wrc_a)};font-weight:700">{wrc_a:.0f}</span></div>'
                  f'{_bar(wrc_a,70,135,_wrc_col(wrc_a))}'
                  f'<div style="font-size:0.73rem;color:{MUTED};margin-top:5px">'
                  f'BP xFIP <span style="color:{_xfip_col(bp_a)};font-weight:700">{bp_a:.2f}</span>'
                  f'&nbsp;·&nbsp;Proj <span style="color:{ACCENT};font-weight:700">{tA}R</span></div>') +
            _cell(f'{_label(f"{esc(home)} Offense")}'
                  f'<div style="font-size:0.78rem;margin-bottom:2px">wRC+ <span style="color:{_wrc_col(wrc_b)};font-weight:700">{wrc_b:.0f}</span></div>'
                  f'{_bar(wrc_b,70,135,_wrc_col(wrc_b))}'
                  f'<div style="font-size:0.73rem;color:{MUTED};margin-top:5px">'
                  f'BP xFIP <span style="color:{_xfip_col(bp_b)};font-weight:700">{bp_b:.2f}</span>'
                  f'&nbsp;·&nbsp;Proj <span style="color:{ACCENT};font-weight:700">{tB}R</span></div>'),
            cols=2)

        # ── Context section ────────────────────────────────────────────────
        pf_comb = round(calc_pf_combined(home, weather.get("dir",""), weather.get("mph",0),
                                         weather.get("temp",70), weather.get("humidity",50)), 3)
        pf_col  = "#ef4444" if pf_comb > 1.03 else ("#22c55e" if pf_comb < 0.97 else TEXT)
        if weather.get("dir") == "DOME":
            _wi2 = STADIUM_ROOF.get(home, {})
            w_str = ("🔒 Techo Fijo" if _wi2.get("roof") == "fixed_dome" else "🏟️ Retráctil")
            if _wi2.get("name"): w_str += f" · {_wi2['name']}"
        else:
            w_str = (f"{weather.get('temp','?')}°F  {weather.get('dir','')} {weather.get('mph',0)}mph"
                     if weather.get("dir") else "—")
        ump_hp  = r.get("ump_hp","—"); ump_f = r.get("ump_factor",1.0)
        ump_s   = "avg" if ump_f == 1.0 else (f"+{(ump_f-1)*100:.1f}% R" if ump_f > 1.0 else f"{(ump_f-1)*100:.1f}% R")
        ump_col = "#ef4444" if ump_f > 1.01 else ("#22c55e" if ump_f < 0.99 else MUTED)

        ctx_html = _inner(
            _cell(f'{_label("Park Factor")}'
                  f'{_val(f"{pf_comb}", pf_col)}'
                  f'<div style="font-size:0.67rem;color:{MUTED}">base {PARK_FACTORS.get(home,1.0)}</div>') +
            _cell(f'{_label("Clima")}'
                  f'<div style="font-size:0.78rem;font-weight:600;color:{TEXT}">{esc(w_str)}</div>') +
            _cell(f'{_label("Umpire HP")}'
                  f'<div style="font-size:0.75rem;font-weight:700;color:{TEXT};line-height:1.3">{esc((ump_hp or "?")[:18])}</div>'
                  f'<div style="font-size:0.7rem;color:{ump_col}">{esc(ump_s)}</div>'),
            cols=3)

        # ── Model Output section ───────────────────────────────────────────
        total_proj = lines.get("total","—")
        mtotals    = lines.get("mTotals","—")
        mspread    = lines.get("mSpread","—").replace("AWAY", away).replace("HOME", home)
        win_a      = lines.get("winA","—"); win_b = lines.get("winB","—")
        ml_pct     = lines.get("ml_pct","—")
        ml_a_str   = ml_pct.split(" / ")[0] if ml_pct and "/" in ml_pct else ml_pct
        ml_b_str   = ml_pct.split(" / ")[1] if ml_pct and "/" in ml_pct else "—"

        model_html = _inner(
            _cell(f'{_label("Total Proyectado")}'
                  f'<div style="font-size:1.15rem;font-weight:900;color:{ACCENT}">{total_proj}</div>'
                  f'<div style="font-size:0.7rem;color:{MUTED}">{esc(mtotals)}</div>') +
            _cell(f'{_label("Spread Modelo")}'
                  f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{esc(mspread)}</div>') +
            _cell(f'{_label(esc(away))}'
                  f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{ml_a_str}</div>'
                  f'<div style="font-size:0.68rem;color:{MUTED}">{win_a}% win</div>') +
            _cell(f'{_label(esc(home))}'
                  f'<div style="font-size:0.88rem;font-weight:700;color:{TEXT}">{ml_b_str}</div>'
                  f'<div style="font-size:0.68rem;color:{MUTED}">{win_b}% win</div>'),
            cols=4)

        # ── Assemble card ──────────────────────────────────────────────────
        return f"""
        <div class="pick-card" style="
          border-left:3px solid {CYAN};
          background:linear-gradient(160deg,#070710 0%,#08080d 100%);
          box-shadow:0 0 0 1px rgba(0,220,255,0.08),0 0 24px rgba(0,220,255,0.04),0 6px 32px rgba(0,0,0,0.8)">
          <!-- AI debug header -->
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;
                      padding-bottom:10px;border-bottom:1px solid {DIVIDER}">
            <div style="display:flex;align-items:center;gap:8px">
              {logo_img(away, 30)}
              <div>
                <div style="font-size:0.58rem;font-weight:800;letter-spacing:0.18em;
                            text-transform:uppercase;
                            background:linear-gradient(90deg,{CYAN},{MUTED2});
                            -webkit-background-clip:text;-webkit-text-fill-color:transparent">
                  &#9632;&nbsp;Model Report Output
                </div>
                <div style="font-size:0.65rem;color:{MUTED};margin-top:1px;font-family:monospace">
                  {esc(away)} @ {esc(home)}
                </div>
              </div>
              {logo_img(home, 30)}
            </div>
            <span style="background:rgba(0,220,255,0.07);border:1px solid rgba(0,220,255,0.18);
                         border-radius:20px;padding:2px 9px;font-size:0.55rem;font-weight:700;
                         letter-spacing:2px;color:{CYAN}88;text-transform:uppercase;
                         font-family:monospace">AI · ANALYSIS</span>
          </div>
          {_section("⚾", "Starting Pitchers", sp_html)}
          {_section("📋", "Batting Order", lineup_html)}
          {_section("🏏", "Offense · wRC+", off_html)}
          {_section("🌤️", "Contexto", ctx_html)}
          {_section("🎯", "Model Output", model_html)}
        </div>"""

    except Exception as _e:
        return ""  # silencioso — no interrumpe el flujo de --log


# ──────────────────────────────────────────────────────
# HTML EXPORT — pick logueado + análisis
# ──────────────────────────────────────────────────────

def export_log_pick_html(entry):
    """
    Genera 'Laboy Pick YYYY-MM-DD #N.html' con:
      Card 1 — pick card (mismo diseño que picks/lines HTML)
      Card 2 — análisis (si hay texto)
    Retorna el path o None si falla.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    try:
        pick_date  = entry.get("date", TARGET_DATE)
        pick_id    = entry.get("id", 0)
        game       = entry.get("game", "")
        pick       = _normalize_pick_str(entry.get("pick", ""))  # "O 8.5" → "OVER 8.5"
        odds_v     = entry.get("odds", 0)
        analysis   = entry.get("analysis", "")
        result     = entry.get("result")
        pnl        = entry.get("pnl")
        book       = entry.get("book", "BetMGM")   # sportsbook donde se jugó
        stake_disp = _fmt_stake(entry)    # "$15.00" o "1u" (legacy)

        # Detectar equipo para logo y color
        pick_upper = pick.upper()
        is_total   = _is_total_pick(pick)
        team_name  = None
        if not is_total:
            for t in TEAM_ABB.values():
                if t in pick_upper:
                    team_name = t; break
        if not team_name and not is_total:
            parts = game.replace(" VS ", " @ ").replace(" VS.", " @ ").split(" @ ")
            team_name = parts[0].strip() if parts else None

        if is_total:
            logo_html = _over_under_logo_html(64)
            color     = "#f97316"   # naranja O/U
        elif team_name:
            url = logo_url(team_name)
            logo_html = (f'<img src="{url}" alt="{esc(team_name)}" '
                         f'width="64" height="64" style="object-fit:contain" '
                         f'onerror="this.style.display=\'none\'">' if url else "")
            color = _team_color_hex(team_name)
        else:
            logo_html = ""
            color     = "#f07820"

        odds_fmt = _fmt_odds(odds_v)
        dt        = datetime.strptime(pick_date, "%Y-%m-%d")
        dstr      = dt.strftime("%A, %B %d · %Y").upper()
        yr        = dt.strftime("%Y")

        # Result color + badge
        if result == "W":
            result_color = "#22c55e"
            result_html  = f'<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✓ WIN</span>'
        elif result == "L":
            result_color = "#ef4444"
            result_html  = f'<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✗ LOSS</span>'
        elif result == "P":
            result_color = "#94a3b8"
            result_html  = f'<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
        else:
            result_color = color   # equipo color mientras está pendiente
            result_html  = f'<span style="background:#f0782022;color:#f07820;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">⏳ PENDING</span>'

        # Card border: resultado cuando hay, color equipo si pendiente
        card_border = result_color

        # Card background tint cuando hay resultado (AI-dark edition)
        if result == "W":
            card_bg_style = "background:linear-gradient(140deg,#061410 0%,#07080d 60%)"
            _glow_col = "#22c55e"
        elif result == "L":
            card_bg_style = "background:linear-gradient(140deg,#14060a 0%,#07080d 60%)"
            _glow_col = "#ef4444"
        elif result == "P":
            card_bg_style = "background:linear-gradient(140deg,#090a10 0%,#07080d 60%)"
            _glow_col = "#94a3b8"
        else:
            card_bg_style = "background:linear-gradient(140deg,#090910 0%,#07080d 100%)"
            _glow_col = color

        pnl_html = ""
        if pnl is not None:
            pnl_col  = "#22c55e" if pnl >= 0 else "#ef4444"
            pnl_disp = _fmt_pnl(entry)
            pnl_html = f'<div class="stat"><div class="stat-label">P&amp;L</div><div class="stat-val" style="color:{pnl_col}">{esc(pnl_disp)}</div></div>'

        # Card 1 — pick (AI style)
        card1 = f"""
        <div class="pick-card" style="border-left:4px solid {card_border};{card_bg_style};
             border-top:1px solid {_glow_col}22;
             box-shadow:0 0 0 1px {_glow_col}10,0 0 28px {_glow_col}0a,0 6px 32px rgba(0,0,0,.85)">
          <!-- pick header row -->
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
            <div style="font-size:0.62rem;color:#4a6272;font-family:monospace;letter-spacing:0.08em">
              <i class="fa-solid fa-ticket fa-icon"></i>PICK&nbsp;#{pick_id}&nbsp;&nbsp;·&nbsp;&nbsp;{esc(pick_date)}
            </div>
            {result_html}
          </div>
          <!-- game + pick -->
          <div class="teams-row">
            {logo_html}
            <div>
              <div class="game-label" style="font-family:monospace;font-size:0.72rem;letter-spacing:0.04em">
                <i class="fa-solid fa-baseball fa-icon"></i>{esc(game)}
              </div>
              <div class="pick-label" style="font-size:1.3rem;letter-spacing:0.02em">
                {esc(pick)}&nbsp;<span class="odds-badge">{esc(odds_fmt)}</span>
              </div>
            </div>
          </div>
          <!-- stat chips -->
          <div class="stats-grid">
            <div class="stat"><div class="stat-label">{esc(book)}</div>
              <div class="stat-val" style="color:#f07820">{esc(odds_fmt)}</div></div>
            <div class="stat"><div class="stat-label">Apostado</div>
              <div class="stat-val">{esc(stake_disp)}</div></div>
            <div class="stat"><div class="stat-label">Fecha</div>
              <div class="stat-val" style="font-size:0.72rem">{esc(pick_date)}</div></div>
            {pnl_html if pnl_html else '<div class="stat"><div class="stat-label">Resultado</div><div class="stat-val">—</div></div>'}
          </div>
        </div>"""

        # Card 2 — análisis (border = resultado, AI dark)
        card2 = ""
        if analysis and analysis.strip():
            analysis_html = esc(analysis).replace("\n", "<br>")
            card2 = f"""
        <div class="pick-card" style="border-left:4px solid {card_border};
             background:linear-gradient(150deg,#070812 0%,#06060a 100%);
             box-shadow:0 0 0 1px rgba(0,220,255,0.05),0 4px 24px rgba(0,0,0,.8)">
          <div class="section-title" style="margin-top:0;margin-bottom:12px">
            <i class="fa-solid fa-magnifying-glass-chart fa-icon"></i>Análisis
          </div>
          <div style="font-size:0.88rem;line-height:1.85;color:#c8d4e0;
                      letter-spacing:0.015em;white-space:pre-wrap">{analysis_html}</div>
        </div>"""

        # Card 3 — debug card (si hay datos del modelo cacheados para esta fecha)
        card3 = ""
        try:
            _cached_results = _load_results_cache(pick_date)
            if _cached_results:
                # Buscar el resultado que coincide con este juego — fuzzy match robusto
                # Normaliza: quita espacios, convierte "@" / "vs" / "VS." / " - " a "@"
                def _norm_gkey(s):
                    s = s.upper().strip()
                    s = s.replace(" VS. ", " @ ").replace(" VS ", " @ ").replace(" - ", " @ ")
                    return s.replace(" ", "")  # "DODGERS@ASTROS"

                _entry_norm = _norm_gkey(game)
                _matching_r = None
                for _r in _cached_results:
                    # Coincidencia exacta primero
                    _gk = f"{_r.get('away','')} @ {_r.get('home','')}".upper()
                    if _gk == game.upper().strip():
                        _matching_r = _r; break
                    # Coincidencia normalizada (quita espacios y variantes de "@")
                    if _norm_gkey(_gk) == _entry_norm:
                        _matching_r = _r; break
                    # Coincidencia por equipos individuales (cualquiera de los dos equipos en el pick)
                    _away_up = _r.get("away", "").upper()
                    _home_up = _r.get("home", "").upper()
                    _pick_up = pick.upper()
                    if (_away_up and _away_up in _entry_norm) or (_home_up and _home_up in _entry_norm):
                        # Ambos equipos deben estar en el game string para evitar falsos positivos
                        if _away_up in _entry_norm and _home_up in _entry_norm:
                            _matching_r = _r; break
                if _matching_r:
                    card3 = _build_single_game_debug_html(_matching_r, pick, pick_date)
        except Exception:
            pass

        body = card1 + card2 + card3

        html  = _debug_html_wrap(f"Laboy Pick #{pick_id} · {dstr}", dstr, yr, body)
        fname = f"Laboy Pick {pick_date} #{pick_id}.html"
        fpath = os.path.join(SCRIPT_DIR, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(html)

        # Auto-generar JPG del mismo HTML (ancho 480 × escala 3 → 1440px — AI detail visible)
        jpg_path = html_to_jpg(fpath, width=480, scale=3)
        if jpg_path:
            print(f"  🖼️  JPG: {os.path.basename(jpg_path)}")

        return fpath

    except Exception as e:
        print(f"     ⚠️  HTML export error: {e}")
        return None


# ──────────────────────────────────────────────────────
# JPG EXPORT — pick logueado + análisis (2 cards)
# ──────────────────────────────────────────────────────

def export_log_pick_jpg(entry):
    """
    Genera 'Laboy Pick YYYY-MM-DD #N.jpg' con:
      Card 1 — pick card (mismo diseño que el HTML)
      Card 2 — análisis
    Retorna el path del archivo o None si falla.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para export JPG")
        return None
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io, urllib.request, textwrap
    except ImportError:
        return None

    # ── Colores (mismo que HTML CSS) ─────────────────────────────────────
    C_BG      = (10, 10, 10)       # --bg   #0a0a0a
    C_CARD    = (34, 34, 34)       # --card #222222
    C_ACCENT  = (240, 120, 32)     # --accent #f07820
    C_TEXT    = (241, 245, 249)    # --text
    C_MUTED   = (148, 163, 184)    # --muted
    C_BORDER  = (42, 42, 42)       # --border
    C_GREEN   = (34, 197, 94)
    C_RED     = (239, 68, 68)
    C_STAT_BG = (24, 24, 24)

    W     = 1080
    PAD   = 52
    CARD_R = 18    # border-radius equivalent

    # ── Font loader (reuse existing helper) ──────────────────────────────
    def fnt(name, size):
        path = os.path.join(_FONTS_DIR, name)
        if os.path.exists(path):
            try: return ImageFont.truetype(path, size)
            except: pass
        try: return ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", size)
        except: pass
        return ImageFont.load_default()

    F_TITLE  = fnt("BigShoulders-Bold.ttf", 52)
    F_PICK   = fnt("BigShoulders-Bold.ttf", 72)
    F_ODDS   = fnt("IBMPlexMono-Bold.ttf",  44)
    F_LABEL  = fnt("GeistMono-Regular.ttf", 22)
    F_STAT   = fnt("IBMPlexMono-Bold.ttf",  34)
    F_DATE   = fnt("GeistMono-Regular.ttf", 24)
    F_GAME   = fnt("IBMPlexMono-Bold.ttf",  30)
    F_BODY   = fnt("IBMPlexMono-Regular.ttf", 28)
    F_ANHEAD = fnt("BigShoulders-Bold.ttf", 42)

    def text_w(d, txt, font):
        bb = d.textbbox((0,0), txt, font=font)
        return bb[2]-bb[0]

    def cx(d, txt, font, y, color, img_w=W):
        tw = text_w(d, txt, font)
        d.text(((img_w - tw) // 2, y), txt, font=font, fill=color)

    def rounded_rect(d, xy, radius, fill=None, outline=None, width=1):
        x0,y0,x1,y1 = xy
        d.rounded_rectangle([(x0,y0),(x1,y1)], radius=radius, fill=fill,
                             outline=outline, width=width)

    # ── Download team logo ────────────────────────────────────────────────
    def get_logo(team, size=80):
        url = logo_url(team)
        if not url: return None
        try:
            req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=6) as resp:
                return Image.open(io.BytesIO(resp.read())).convert("RGBA").resize((size,size), Image.LANCZOS)
        except: return None

    # ── Parse entry ──────────────────────────────────────────────────────
    game     = entry.get("game","").upper()
    pick     = entry.get("pick","").upper()
    odds_v   = entry.get("odds", 0)
    odds_s   = _fmt_odds(odds_v)
    units    = entry.get("units", 1)
    analysis = entry.get("analysis","").strip()
    edate    = entry.get("date", TARGET_DATE)
    idx      = entry.get("id", 0)
    is_pos   = odds_v >= 0

    try:
        dt   = datetime.strptime(edate, "%Y-%m-%d")
        dstr = dt.strftime("%A, %B %d · %Y").upper()
    except:
        dstr = edate.upper()

    # Extract teams from game string (e.g. "CUBS VS PIRATES")
    teams_in_game = re.split(r'\s+(?:VS\.?|@|AT)\s+', game, flags=re.IGNORECASE)
    away_team = teams_in_game[0].strip() if teams_in_game else ""
    home_team = teams_in_game[1].strip() if len(teams_in_game) > 1 else ""

    # ── Build canvas height dynamically ──────────────────────────────────
    # Header: ~140px, Card1 (pick): ~420px, Card2 (analysis): ~240+, Footer: ~80px
    analysis_lines = textwrap.wrap(analysis, width=42) if analysis else []
    ANALYSIS_H = 60 + len(analysis_lines) * 36 + 40 if analysis_lines else 0
    CARD2_H    = max(ANALYSIS_H, 120) if analysis else 0
    H = 140 + 30 + 420 + (30 + CARD2_H if CARD2_H else 0) + 80
    H = max(H, 1080)

    img = Image.new("RGB", (W, H), C_BG)
    d   = ImageDraw.Draw(img)

    # ── Orange top stripe ─────────────────────────────────────────────────
    d.rectangle([(0,0),(W,6)], fill=C_ACCENT)

    y = 18

    # ── Logo or title ─────────────────────────────────────────────────────
    logo_src = _logo_b64()  # check if file exists
    logo_path = os.path.join(SCRIPT_DIR, "laboy_logo.png")
    if os.path.exists(logo_path):
        try:
            logo_img = Image.open(logo_path).convert("RGBA")
            # Remove black bg
            pix = logo_img.load()
            for py in range(logo_img.height):
                for px in range(logo_img.width):
                    r,g,b,a = pix[px,py]
                    if r<35 and g<35 and b<35:
                        pix[px,py] = (0,0,0,0)
            logo_h = 110
            ratio  = logo_h / logo_img.height
            logo_w = int(logo_img.width * ratio)
            logo_img = logo_img.resize((logo_w, logo_h), Image.LANCZOS)
            logo_x = (W - logo_w) // 2
            img.paste(logo_img, (logo_x, y), logo_img)
            y += logo_h + 4
        except:
            cx(d, "LABOY PICKS", F_TITLE, y, C_ACCENT); y += 60
    else:
        cx(d, "LABOY PICKS", F_TITLE, y, C_ACCENT); y += 60

    # Date
    cx(d, dstr, F_DATE, y, C_MUTED); y += 32

    # Orange separator
    d.rectangle([(0, y+8), (W, y+10)], fill=C_ACCENT); y += 26

    # ── Card 1: Pick ──────────────────────────────────────────────────────
    C1_X, C1_Y = PAD, y
    C1_W = W - PAD*2
    C1_H = 390
    rounded_rect(d, (C1_X, C1_Y, C1_X+C1_W, C1_Y+C1_H), CARD_R,
                 fill=C_CARD, outline=C_BORDER, width=1)
    # Left accent stripe
    d.rounded_rectangle((C1_X, C1_Y, C1_X+5, C1_Y+C1_H), radius=CARD_R, fill=C_ACCENT)

    iy = C1_Y + 22

    # Game time row
    cx(d, game, F_GAME, iy, C_MUTED); iy += 38

    # Team logos
    logo_y = iy
    away_logo = get_logo(away_team, 72)
    home_logo = get_logo(home_team, 72)
    logo_spacing = 200
    lx_base = W//2 - logo_spacing
    for lx, lteam, limg in [(lx_base, away_team, away_logo),
                             (lx_base + logo_spacing*2 - 72, home_team, home_logo)]:
        if limg:
            bg = Image.new("RGBA", limg.size, (0,0,0,0))
            img.paste(limg, (lx, logo_y), limg)
    # VS
    cx(d, "VS", fnt("IBMPlexMono-Bold.ttf",26), logo_y+24, C_MUTED)
    iy += 84

    # Pick label
    cx(d, pick, F_PICK, iy, C_TEXT); iy += 80

    # Odds badge
    ob_w = text_w(d, odds_s, F_ODDS)
    bx = (W - ob_w - 48) // 2
    bg_col = (8,40,22) if is_pos else (30,20,20)
    ol_col = C_GREEN if is_pos else C_RED
    rounded_rect(d, (bx, iy, bx+ob_w+48, iy+52), 10, fill=bg_col, outline=ol_col, width=2)
    d.text((bx+24, iy+9), odds_s, font=F_ODDS, fill=C_GREEN if is_pos else C_RED)
    iy += 62

    # Units
    cx(d, f"{units}u", F_DATE, iy, C_MUTED); iy += 30

    # ── 4-stat grid ───────────────────────────────────────────────────────
    STAT_LABELS = ["JUEGO", "PICK", "ODDS", "UNITS"]
    STAT_VALS   = [game[:14], pick[:14], odds_s, f"{units}u"]
    sgw = (C1_W - 32) // 4
    sx0 = C1_X + 16
    # place stats at bottom of card
    sy = C1_Y + C1_H - 90
    for i,(lbl,val) in enumerate(zip(STAT_LABELS, STAT_VALS)):
        sx = sx0 + i*sgw
        rounded_rect(d, (sx+2, sy, sx+sgw-4, sy+80), 8, fill=C_STAT_BG)
        lw = text_w(d, lbl, F_LABEL)
        d.text((sx + (sgw-lw)//2, sy+8), lbl, font=F_LABEL, fill=C_MUTED)
        vw = text_w(d, val, F_STAT)
        d.text((sx + (sgw-vw)//2, sy+34), val, font=F_STAT,
               fill=C_GREEN if (i==2 and is_pos) else C_TEXT)

    y = C1_Y + C1_H + 24

    # ── Card 2: Analysis (only if analysis provided) ──────────────────────
    if analysis_lines:
        C2_X, C2_Y = PAD, y
        C2_H = CARD2_H
        C2_W = C1_W
        rounded_rect(d, (C2_X, C2_Y, C2_X+C2_W, C2_Y+C2_H), CARD_R,
                     fill=C_CARD, outline=C_BORDER, width=1)
        d.rounded_rectangle((C2_X, C2_Y, C2_X+5, C2_Y+C2_H), radius=CARD_R, fill=C_MUTED)

        ay = C2_Y + 18
        cx(d, "ANÁLISIS", F_ANHEAD, ay, C_ACCENT); ay += 48
        d.line([(C2_X+16, ay), (C2_X+C2_W-16, ay)], fill=C_BORDER, width=1); ay += 14

        for line in analysis_lines:
            lw = text_w(d, line, F_BODY)
            d.text(((W - lw)//2, ay), line, font=F_BODY, fill=C_TEXT)
            ay += 36

        y = C2_Y + C2_H + 24

    # ── Footer ────────────────────────────────────────────────────────────
    d.rectangle([(0, H-6), (W, H)], fill=C_ACCENT)
    cx(d, "Laboy Picks · dubclub.win", F_DATE, H-44, C_MUTED)

    # ── Save as JPG ───────────────────────────────────────────────────────
    fname = f"Laboy Pick {edate} #{idx}.jpg"
    fpath = os.path.join(SCRIPT_DIR, fname)
    img.convert("RGB").save(fpath, "JPEG", quality=92)
    print(f"  🖼️  Pick JPG: {fname}")
    return fpath


# ──────────────────────────────────────────────────────
# PNG EXPORTS (Pillow) — separados por tipo
# ──────────────────────────────────────────────────────

# Paleta compartida
_BG       = (6,  8, 15)
_PANEL    = (12, 15, 24)
_STAT_BG  = (8,  10, 18)
_AMBER    = (240, 95,  8)
_AMBER_D  = (70,  28,  3)
_AMBER_M  = (140, 52,  4)
_ICE      = (228, 238, 255)
_ICE_D    = (120, 135, 168)
_MUTED    = (58,  68, 98)
_GREEN    = (52,  211, 108)
_GREEN_BG = (8,   40,  22)
_RED      = (239, 68,  68)
_RULE     = (16,  20,  36)
_TEAM_COLORS = {
    "CUBS":(14,51,134),"PIRATES":(253,184,39),"YANKEES":(0,48,135),"RAYS":(9,44,92),
    "DODGERS":(0,90,156),"BRAVES":(206,17,65),"RED SOX":(189,48,57),"METS":(0,45,114),
    "PHILLIES":(232,24,40),"BREWERS":(255,197,47),"CARDINALS":(196,30,58),"REDS":(198,1,31),
    "GIANTS":(253,90,30),"PADRES":(47,36,29),"ATHLETICS":(0,56,49),"MARINERS":(0,92,92),
    "ASTROS":(0,45,98),"RANGERS":(0,50,120),"ANGELS":(186,0,33),"D-BACKS":(167,25,48),
    "ROCKIES":(51,0,111),"TIGERS":(12,35,64),"ROYALS":(0,70,135),"TWINS":(0,43,92),
    "WHITE SOX":(39,37,31),"GUARDIANS":(227,25,55),"ORIOLES":(223,70,1),
    "NATIONALS":(171,0,3),"MARLINS":(0,163,224),"BLUE JAYS":(19,74,142),
}

def _png_dot_bg(d, W, H):
    for gy in range(0, H, 54):
        for gx in range(0, W, 54):
            d.ellipse([(gx-1,gy-1),(gx+1,gy+1)], fill=(14,18,32))

def _png_vignette(img, W, H):
    vgn = Image.new("RGBA",(W,H),(0,0,0,0))
    vd  = ImageDraw.Draw(vgn)
    for s in range(0,90,6):
        vd.rectangle([(s,s),(W-s,H-s)], outline=(0,0,0,int(s*1.6)), width=6)
    return Image.alpha_composite(img.convert("RGBA"), vgn).convert("RGB")

def _tw(d, text, fnt):
    bb = d.textbbox((0,0), text, font=fnt)
    return bb[2]-bb[0], bb[3]-bb[1]

def _cx(d, text, fnt, W, y, col):
    bb = d.textbbox((0,0), text, font=fnt)
    d.text(((W-(bb[2]-bb[0]))//2, y), text, font=fnt, fill=col)


def _png_header(d, W, dstr, subtitle, F_HERO, F_SUB, F_DATE, y_start=16):
    """Dibuja el encabezado estandar en un ImageDraw. Retorna la y después del header."""
    y = y_start
    _cx(d, "LABOY PICKS", F_HERO, W, y, _ICE)
    y += 108
    _cx(d, subtitle, F_SUB, W, y, _AMBER)
    y += 34
    _cx(d, dstr, F_DATE, W, y, _MUTED)
    y += 24
    d.line([(56,y),(W-56,y)], fill=_AMBER_D, width=2)
    y += 10
    return y


def _draw_picks_page(picks_page, dt, dstr, page_num, total_pages):
    """
    Dibuja una página de picks (del LOG del usuario) con diseño limpio y profesional.
    Cada pick ocupa una card grande con pick prominente, odds, units y análisis.
    """
    PAD   = 56
    C_H   = 200       # altura base de cada card
    GAP   = 12
    W     = 1080
    C_W   = W - PAD*2
    HDR_H = 196
    FTR_H = 70

    n = max(1, len(picks_page))
    H = HDR_H + n*(C_H+GAP) + FTR_H

    img = Image.new("RGB", (W, H), _BG)
    d   = ImageDraw.Draw(img)

    # Barras accent top/bottom
    d.rectangle([(0,0),(W,6)], fill=_AMBER)
    d.rectangle([(0,H-6),(W,H)], fill=_AMBER)

    F_HERO  = _fnt("BigShoulders-Bold.ttf", 100)
    F_PICK  = _fnt("BigShoulders-Bold.ttf",  80)
    F_ODDS  = _fnt("IBMPlexMono-Bold.ttf",   38)
    F_LBL   = _fnt("GeistMono-Regular.ttf",  22)
    F_SMALL = _fnt("GeistMono-Regular.ttf",  18)
    F_ANLZ  = _fnt("IBMPlexMono-Regular.ttf",22)

    page_s = f"{dstr}" if total_pages==1 else f"{dstr}   |   {page_num}/{total_pages}"
    y = _png_header(d, W, page_s, "MY PICKS", F_HERO, F_LBL, F_SMALL)

    for p in picks_page:
        # ── Card background ───────────────────────────────
        rbg   = (11, 14, 24)
        d.rounded_rectangle([(PAD,y),(PAD+C_W,y+C_H)], radius=14, fill=rbg)

        # Left accent: amber for active, green for WIN, red for LOSS
        result = str(p.get("result","") or "").upper()
        acc = _GREEN if result=="W" else _RED if result=="L" else _AMBER
        d.rounded_rectangle([(PAD,y),(PAD+5,y+C_H)], radius=3, fill=acc)

        # ── Row 1: game + time ───────────────────────────
        yt = y + 18
        game_txt = p.get("game","").upper()
        time_txt = p.get("time","")
        d.text((PAD+20, yt), game_txt, font=F_SMALL, fill=_ICE_D)
        # Triple Lock / Perfect Storm badge
        tc_x_end = PAD + C_W - 20
        if p.get("candado"):
            tc_lbl  = "TRIPLE LOCK"
            tw_tc,th_tc = _tw(d, tc_lbl, F_SMALL)
            tc_bx   = (tc_x_end - tw_tc - 16, yt - 2,
                       tc_x_end,               yt + th_tc + 6)
            d.rounded_rectangle([tc_bx[:2], tc_bx[2:]], radius=5,
                                 fill=(60, 42, 0))
            d.text((tc_x_end - tw_tc - 8, yt + 2), tc_lbl, font=F_SMALL, fill=_AMBER)
            tc_x_end = tc_bx[0] - 10
        elif p.get("tormenta"):
            tp_lbl  = "PERFECT STORM"
            tw_tp,th_tp = _tw(d, tp_lbl, F_SMALL)
            tp_bx   = (tc_x_end - tw_tp - 16, yt - 2,
                       tc_x_end,               yt + th_tp + 6)
            d.rounded_rectangle([tp_bx[:2], tp_bx[2:]], radius=5,
                                 fill=(10, 28, 55))
            d.text((tc_x_end - tw_tp - 8, yt + 2), tp_lbl, font=F_SMALL, fill=(100, 160, 255))
            tc_x_end = tp_bx[0] - 10
        if time_txt:
            tw_t,_ = _tw(d, time_txt, F_SMALL)
            d.text((tc_x_end - tw_t, yt), time_txt, font=F_SMALL, fill=_AMBER)

        # ── Row 2: pick (big) + odds badge + units ───────
        yt += 30
        pick_txt = p.get("pick","").upper()
        d.text((PAD+20, yt), pick_txt, font=F_PICK, fill=_ICE)

        pk_w,pk_h = _tw(d, pick_txt, F_PICK)

        odds_v = p.get("odds",0)
        odds_s = _fmt_odds(odds_v) if isinstance(odds_v,(int,float)) else str(odds_v)
        is_pos = not odds_s.startswith("-")
        ox = PAD+20 + pk_w + 16
        oy = yt + 6
        ob_w,ob_h = _tw(d, odds_s, F_ODDS)
        d.rounded_rectangle([(ox,oy),(ox+ob_w+24,oy+ob_h+12)], radius=8,
                              fill=_GREEN_BG if is_pos else (20,22,42))
        d.text((ox+12, oy+6), odds_s, font=F_ODDS,
               fill=_GREEN if is_pos else _ICE_D)

        units = p.get("units",1)
        u_txt = f"{units}u"
        u_x   = ox + ob_w + 40
        d.text((u_x, oy+10), u_txt, font=F_LBL, fill=_MUTED)

        # Result badge
        if result in ("W","L","P"):
            res_lbl = {"W":"WIN","L":"LOSS","P":"PUSH"}[result]
            res_col = {"W":_GREEN,"L":_RED,"P":_ICE_D}[result]
            res_bg  = {"W":_GREEN_BG,"L":(40,8,8),"P":(20,22,42)}[result]
            rw,rh   = _tw(d, res_lbl, F_LBL)
            rx      = PAD+C_W - rw - 36
            ry      = ot = oy
            d.rounded_rectangle([(rx-10,ry),(rx+rw+10,ry+rh+10)], radius=6, fill=res_bg)
            d.text((rx, ry+5), res_lbl, font=F_LBL, fill=res_col)

        # ── Row 3: analysis text (word-wrap) ─────────────
        analysis = str(p.get("analysis","") or "").strip()
        if analysis:
            yd = yt + pk_h + 14
            # Word-wrap dentro del ancho disponible
            max_w = C_W - 44
            words_a = analysis.split()
            lines_a = []; cur = ""
            for w in words_a:
                test = (cur+" "+w).strip()
                if _tw(d, test, F_ANLZ)[0] <= max_w:
                    cur = test
                else:
                    if cur: lines_a.append(cur)
                    cur = w
            if cur: lines_a.append(cur)
            for ln in lines_a[:3]:
                d.text((PAD+20, yd), ln, font=F_ANLZ, fill=(80,95,128))
                yd += 26

        y += C_H + GAP

    # Footer
    y += 8
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE, width=1)
    y += 12
    _cx(d, "dubclub.win  ·  Laboy Picks", F_SMALL, W, y, _MUTED)

    return img


def export_picks_png(picks, max_per_page=5):
    """
    PNG(s) con picks EV+ del modelo (ordenados por hora) — uso interno.
    Para tus picks del log usa export_log_picks_png().
    Retorna lista de paths generados.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export")
        return []

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A %B %d · %Y").upper()

    pages  = [picks[i:i+max_per_page] for i in range(0, max(1,len(picks)), max_per_page)] \
             if picks else [[]]
    total  = len(pages)
    paths  = []

    for pi, page in enumerate(pages, 1):
        fname = f"Laboy Picks {TARGET_DATE}.png" if total==1 else f"Laboy Picks {TARGET_DATE} p{pi}.png"
        fpath = os.path.join(SCRIPT_DIR, fname)
        img   = _draw_picks_page(page, dt, dstr, pi, total)
        img   = _png_vignette(img, *img.size)
        img.save(fpath, "PNG", dpi=(300,300))
        print(f"  🖼️  Picks PNG ({pi}/{total}): {fname}")
        paths.append(fpath)

    return paths


def export_log_picks_png(date_str=None):
    """
    --export-picks: Genera PNG con los picks que el usuario logueó para TARGET_DATE.
    Lee del archivo de log local — NO del modelo.
    Retorna lista de paths generados.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export")
        return []

    td   = date_str or TARGET_DATE
    log  = _load_log()
    # Picks del día: activos (sin resultado) + settled del mismo día
    today_picks = [e for e in log if e.get("date","") == td]

    if not today_picks:
        print(f"\n  ℹ️  No hay picks logueados para {td}.")
        print(f"  Usa: python3 mlb.py --log  para registrar un pick.")
        return []

    # Ordenar por hora si disponible, luego por id
    today_picks.sort(key=lambda e: (_parse_time_sort(e.get("time","")), e.get("id",0)))

    dt   = datetime.strptime(td, "%Y-%m-%d")
    dstr = dt.strftime("%A %B %d · %Y").upper()

    MAX_PER = 5
    pages   = [today_picks[i:i+MAX_PER] for i in range(0, len(today_picks), MAX_PER)]
    total   = len(pages)
    paths   = []

    for pi, page in enumerate(pages, 1):
        fname = f"Laboy MyPicks {td}.png" if total==1 else f"Laboy MyPicks {td} p{pi}.png"
        fpath = os.path.join(SCRIPT_DIR, fname)
        img   = _draw_picks_page(page, dt, dstr, pi, total)
        img   = _png_vignette(img, *img.size)
        img.save(fpath, "PNG", dpi=(300,300))
        print(f"  🖼️  My Picks PNG ({pi}/{total}): {fname}")
        paths.append(fpath)

    return paths


def pngs_to_pdf(png_paths, out_path):
    """
    Convierte una lista de PNGs a un PDF usando img2pdf.
    Instala img2pdf si no está disponible.
    Retorna el path del PDF generado, o None si falla.
    """
    try:
        import img2pdf as _img2pdf
    except ImportError:
        import subprocess as _sp, importlib as _il, importlib.util as _ilu
        _sp.run([sys.executable, "-m", "pip", "install", "img2pdf", "-q"], check=True)
        _il.invalidate_caches()
        import img2pdf as _img2pdf

    existing = [p for p in png_paths if os.path.exists(p)]
    if not existing:
        print("  ⚠️  pngs_to_pdf: ningún PNG encontrado.")
        return None

    with open(out_path, "wb") as f:
        f.write(_img2pdf.convert(existing))
    print(f"  📄 PDF generado: {os.path.basename(out_path)}")
    return out_path


def crop_story_to_post(story_path, out_path=None):
    """
    Toma un PNG 1080×1920 (Instagram Story) y lo recorta/rellena a 1080×1080.
    Centrado verticalmente sobre la zona de contenido principal.
    Retorna path del PNG resultante.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG crop")
        return None

    img = Image.open(story_path).convert("RGB")
    w, h = img.size

    # Si ya es cuadrado, guardar directamente
    if w == h:
        out_path = out_path or story_path.replace(".png", "_post.png")
        img.save(out_path, "PNG", dpi=(300, 300))
        return out_path

    # Recortar 1080×1080 centrado verticalmente (toma zona central del story)
    size = min(w, h, 1080)
    left = (w - size) // 2
    # Offset vertical: preferimos zona superior-media (donde está el contenido)
    top  = max(0, (h - size) // 3)
    cropped = img.crop((left, top, left + size, top + size))

    out_path = out_path or story_path.replace(".png", "_post.png")
    cropped.save(out_path, "PNG", dpi=(300, 300))
    print(f"  🟥 Post 1080×1080: {os.path.basename(out_path)}")
    return out_path


def _draw_lines_page(rows_page, dt, dstr, page_num, total_pages, odds_dict={}):
    """
    Dibuja una página de model lines — diseño card-per-game, clean y profesional.
    Cada juego tiene: HORA · MATCHUP · PITCHER · TOTAL (grande, coloreado) · SPREAD · ML%
    """
    PAD   = 48
    W     = 1080
    C_W   = W - PAD*2
    RH    = 116      # altura de cada card de juego
    GAP   = 10
    HDR_H = 196
    FTR_H = 66
    n     = max(1, len(rows_page))
    H     = HDR_H + n*(RH+GAP) + FTR_H

    img = Image.new("RGB", (W, H), _BG)
    d   = ImageDraw.Draw(img)

    # Accent bars
    d.rectangle([(0,0),(W,6)],   fill=_AMBER)
    d.rectangle([(0,H-6),(W,H)], fill=_AMBER)

    F_HERO  = _fnt("BigShoulders-Bold.ttf", 100)
    F_MATCH = _fnt("BigShoulders-Bold.ttf",  52)
    F_TOTAL = _fnt("BigShoulders-Bold.ttf",  68)
    F_MONOB = _fnt("IBMPlexMono-Bold.ttf",   28)
    F_SMALL = _fnt("GeistMono-Regular.ttf",  17)
    F_LBL   = _fnt("GeistMono-Regular.ttf",  22)

    page_s = f"{dstr}" if total_pages==1 else f"{dstr}   |   {page_num}/{total_pages}"
    y = _png_header(d, W, page_s, "MODEL LINES", F_HERO, F_LBL, F_SMALL)

    for i, r in enumerate(rows_page):
        away   = r["away"]
        home   = r["home"]
        ld     = r["lines"]
        wth    = r["weather"]
        gtime  = r.get("game_time_local","")
        mspread= ld["mSpread"].replace("AWAY",away).replace("HOME",home)
        ml_pct = ld.get("ml_pct","")
        books  = _get_game_books(odds_dict, away, home)
        pin    = books.get("Pinnacle",{})
        pin_ml_a = pin.get(f"ML_{away}"); pin_ml_h = pin.get(f"ML_{home}")
        mkt_s  = f"Pin {_fmt_odds(pin_ml_a)} / {_fmt_odds(pin_ml_h)}" \
                 if (pin_ml_a and pin_ml_h) else ""

        is_over = ld["mTotals"].startswith("O")
        tot_col = _GREEN if is_over else _RED
        tot_bg  = _GREEN_BG if is_over else (40,8,8)

        ry  = y + i*(RH+GAP)
        rbg = (11,14,24) if i%2==0 else (7,9,18)
        d.rounded_rectangle([(PAD,ry),(PAD+C_W,ry+RH)], radius=12, fill=rbg)

        # Left accent: green=Over, red=Under
        d.rounded_rectangle([(PAD,ry),(PAD+5,ry+RH)], radius=3, fill=tot_bg)

        # ── Left section: Time + Matchup + SP/Weather ─────
        # Time
        d.text((PAD+20, ry+12), gtime, font=F_SMALL, fill=_AMBER)

        # Matchup
        matchup = f"{away}  @  {home}"
        d.text((PAD+20, ry+34), matchup, font=F_MATCH, fill=_ICE)

        # SP (pitcher matchup)
        asp = r.get("away_sp","TBD")[:15]
        hsp = r.get("home_sp","TBD")[:15]
        sp_txt = f"{asp}  vs  {hsp}"
        d.text((PAD+20, ry+84), sp_txt, font=F_SMALL, fill=(75,90,130))

        # Weather (right of SP)
        if wth and wth.get("dir"):
            if wth["dir"]=="DOME":
                w_txt = "DOME"
            elif wth.get("temp"):
                w_txt = f"{wth['temp']}F  {wth['dir']}  {wth['mph']}mph"
            else:
                w_txt = ""
            if w_txt:
                tw_w,_ = _tw(d, w_txt, F_SMALL)
                d.text((PAD+C_W-tw_w-20, ry+84), w_txt, font=F_SMALL, fill=(50,65,100))

        # ── Right section: Total (big) + Spread + ML% ─────
        SPLIT    = PAD + C_W - 380    # x where right block starts
        tot_txt  = ld["mTotals"]
        tot_w,_  = _tw(d, tot_txt, F_TOTAL)
        d.text((SPLIT, ry+14), tot_txt, font=F_TOTAL, fill=tot_col)

        # Spread and ML below total
        spr_txt  = f"{mspread}"
        d.text((SPLIT, ry+80), spr_txt, font=F_MONOB, fill=_ICE_D)

        ml_x = SPLIT + 220
        d.text((ml_x, ry+80), ml_pct, font=F_MONOB, fill=_ICE_D)

        # Market odds (Pinnacle) — small below
        if mkt_s:
            d.text((SPLIT, ry+104), mkt_s, font=F_SMALL, fill=(55,72,110))

    # Footer
    y += n*(RH+GAP) + 12
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE, width=1)
    y += 12
    _cx(d, "dubclub.win  ·  Laboy Picks Data Model", F_SMALL, W, y, _MUTED)

    return img


def export_lines_png(results, odds={}, max_per_page=7):
    """
    PNG(s) con todas las model lines del día, ordenadas por hora.
    Si hay más de max_per_page juegos, genera múltiples imágenes paginadas.
    Retorna lista de paths generados.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export")
        return []

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A %B %d · %Y").upper()

    # Ordenar por hora
    sorted_results = sorted(results,
                            key=lambda r: _parse_time_sort(r.get("game_time_local","")))

    if not sorted_results:
        fname = f"Laboy Lines {TARGET_DATE}.png"
        fpath = os.path.join(SCRIPT_DIR, fname)
        img   = _draw_lines_page([], dt, dstr, 1, 1, odds)
        img   = _png_vignette(img, *img.size)
        img.save(fpath, "PNG", dpi=(300,300))
        print(f"  📊 Lines PNG (sin juegos): {fname}")
        return [fpath]

    pages  = [sorted_results[i:i+max_per_page]
              for i in range(0, len(sorted_results), max_per_page)]
    total  = len(pages)
    paths  = []

    for pi, page in enumerate(pages, 1):
        if total == 1:
            fname = f"Laboy Lines {TARGET_DATE}.png"
        else:
            fname = f"Laboy Lines {TARGET_DATE} p{pi}.png"
        fpath = os.path.join(SCRIPT_DIR, fname)
        img   = _draw_lines_page(page, dt, dstr, pi, total, odds)
        img   = _png_vignette(img, *img.size)
        img.save(fpath, "PNG", dpi=(300,300))
        print(f"  📊 Lines PNG ({pi}/{total}): {fname}")
        paths.append(fpath)

    return paths


def gen_pick_card_png(entry, analysis=""):
    """PNG card 1080×1080 para un pick individual con análisis."""
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para pick card")
        return None

    W, H = 1080, 1080
    PAD  = 60

    img = Image.new("RGB", (W, H), _BG)
    d   = ImageDraw.Draw(img)
    _png_dot_bg(d, W, H)
    d.rectangle([(0,0),(W,8)], fill=_AMBER)
    d.rectangle([(0,H-8),(W,H)], fill=_AMBER)

    F_HERO  = _fnt("BigShoulders-Bold.ttf", 100)
    F_BIG   = _fnt("BigShoulders-Bold.ttf",  72)
    F_MED   = _fnt("WorkSans-Bold.ttf",       40) if os.path.exists(
              os.path.join(_FONTS_DIR,"WorkSans-Bold.ttf")) else _fnt("IBMPlexMono-Bold.ttf",38)
    F_MONOB = _fnt("IBMPlexMono-Bold.ttf",    32)
    F_MONO  = _fnt("IBMPlexMono-Regular.ttf", 26)
    F_LBL   = _fnt("GeistMono-Regular.ttf",   22)
    F_MICRO = _fnt("GeistMono-Regular.ttf",   18)

    odds_v  = entry.get("odds", 0)
    odds_s  = _fmt_odds(odds_v)
    game    = entry.get("game","").upper()
    pick    = entry.get("pick","").upper()
    edate   = entry.get("date", TARGET_DATE)
    units   = entry.get("units", 1)
    idx     = entry.get("id", 0)

    try:
        dt   = datetime.strptime(edate, "%Y-%m-%d")
        dstr = dt.strftime("%A, %B %d %Y").upper()
    except:
        dstr = edate.upper()

    # Corner marks
    M, L = 40, 28
    for (x1,y1),(x2,y2) in [
        [(M,M),(M+L,M)],[(M,M),(M,M+L)],
        [(W-M-L,M),(W-M,M)],[(W-M,M),(W-M,M+L)],
        [(M,H-M),(M+L,H-M)],[(M,H-M-L),(M,H-M)],
        [(W-M-L,H-M),(W-M,H-M)],[(W-M,H-M-L),(W-M,H-M)],
    ]:
        d.line([(x1,y1),(x2,y2)], fill=_AMBER_M, width=2)

    y = 40
    _cx(d, "LABOY PICKS", F_HERO, W, y, _ICE)
    y += 106
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE, width=1)
    y += 16
    _cx(d, f"MLB  ·  {dstr}", F_MICRO, W, y, _MUTED)
    y += 34

    # Game matchup banner
    d.rounded_rectangle([(PAD,y),(W-PAD,y+70)], radius=12, fill=_PANEL)
    d.rounded_rectangle([(PAD+1,y+1),(W-PAD-1,y+69)], radius=11, outline=(22,28,50), width=1)
    d.rounded_rectangle([(PAD,y),(PAD+4,y+70)], radius=2, fill=_AMBER)
    _cx(d, game, F_MONOB, W, y+18, _ICE)
    y += 86

    # Pick label
    _cx(d, pick, F_BIG, W, y, _ICE)
    y += 78

    # Odds badge centered
    ob_w, ob_h = _tw(d, odds_s, F_MONOB)
    is_pos = not odds_s.startswith("-")
    ox = (W - ob_w - 32) // 2
    d.rounded_rectangle([(ox,y),(ox+ob_w+32,y+ob_h+16)], radius=10,
                          fill=_GREEN_BG if is_pos else (18,20,38))
    d.rounded_rectangle([(ox,y),(ox+ob_w+32,y+ob_h+16)], radius=10,
                          outline=(_GREEN if is_pos else (40,50,80)), width=1)
    d.text((ox+16, y+8), odds_s, font=F_MONOB, fill=_GREEN if is_pos else _ICE)
    y += ob_h + 34

    # Units row
    u_txt = f"UNIDADES: {units}u"
    _cx(d, u_txt, F_LBL, W, y, _MUTED)
    y += 40

    d.line([(PAD,y),(W-PAD,y)], fill=_RULE, width=1)
    y += 20

    # Analysis box
    if analysis:
        d.rounded_rectangle([(PAD,y),(W-PAD,y+220)], radius=14, fill=_PANEL)
        d.rounded_rectangle([(PAD+1,y+1),(W-PAD-1,y+219)], radius=13, outline=(22,28,50), width=1)
        d.rounded_rectangle([(PAD,y),(PAD+4,y+220)], radius=2, fill=_AMBER)
        # Word-wrap analysis text
        words = analysis.split()
        lines_out = []; line_cur = ""
        max_w = W - PAD*2 - 36
        for w_word in words:
            test = (line_cur + " " + w_word).strip()
            tw_test,_ = _tw(d, test, F_MONO)
            if tw_test <= max_w:
                line_cur = test
            else:
                if line_cur: lines_out.append(line_cur)
                line_cur = w_word
        if line_cur: lines_out.append(line_cur)
        ty = y + 18
        for ln in lines_out[:7]:
            d.text((PAD+20, ty), ln, font=F_MONO, fill=_ICE_D)
            ty += 28
        y += 238
    else:
        y += 10

    # Footer
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE, width=1)
    y += 16
    _cx(d, f"Pick #{idx}  ·  dubclub.win  ·  Laboy Picks", F_MICRO, W, y, _MUTED)

    img = _png_vignette(img, W, H)
    fname = f"Laboy Pick #{idx} {edate}.png"
    fpath = os.path.join(SCRIPT_DIR, fname)
    img.save(fpath, "PNG", dpi=(300,300))
    print(f"\n  🃏 Pick card: {fname}")
    return fpath


# ──────────────────────────────────────────────────────
# ACTUALIZAR EXCEL
# ──────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────
# ACTUALIZAR EXCEL
# ──────────────────────────────────────────────────────

def copy_formula(f, from_row, to_row):
    if not isinstance(f, str) or not f.startswith("="): return f
    return re.sub(r'([A-Z]+)(\d+)',
                  lambda m: f"{m.group(1)}{int(m.group(2))+(to_row-from_row)}", f)


def update_excel(wb, games_with_lines):
    ws = wb[LINES_SHEET]
    target_dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and isinstance(row[0], datetime) and row[0].date()==target_dt.date():
            print(f"\n⚠️  Ya existen juegos para {TARGET_DATE}. Borra y corre de nuevo.")
            return False
    last_row = max((r[0].row for r in ws.iter_rows(min_row=2, max_row=ws.max_row)
                    if r[0].value is not None), default=None)
    if not last_row:
        print("❌ No hay fila de referencia en MLB Lines."); return False
    next_row = ws.max_row + 1
    for g in games_with_lines:
        r = next_row
        ws.cell(r,1).value  = target_dt; ws.cell(r,1).number_format = "MM/DD/YYYY"
        ws.cell(r,2).value  = g["away"];  ws.cell(r,4).value  = g["away_sp"]
        ws.cell(r,5).value  = g["fip_a"]; ws.cell(r,8).value  = g["home"]
        ws.cell(r,10).value = g["home_sp"]; ws.cell(r,11).value = g["fip_b"]
        ws.cell(r,21).value = g["weather"]["dir"]
        ws.cell(r,22).value = g["weather"]["mph"]
        ws.cell(r,23).value = g["weather"]["temp"]
        for col in [3,6,7,9,12,13,15,16,17,19,20,25,26,27,28,29,30,31,32,33]:
            src = ws.cell(last_row, col); dst = ws.cell(r, col)
            dst.value = copy_formula(src.value, last_row, r)
            if src.has_style:
                dst.font=copy(src.font); dst.fill=copy(src.fill)
                dst.border=copy(src.border); dst.alignment=copy(src.alignment)
                dst.number_format=src.number_format
        next_row += 1
    return True


def refresh_mlb_data(wb, fg):
    """
    Fetcha wRC+ y Bullpen xFIP de FanGraphs y:
      1. Guarda a JSON cache (mlb_fg_cache.json) — fuente principal del modelo
      2. También escribe a Excel MLB Data (cols K,L,N,O) como referencia visual
    """
    print("\n🔄 Actualizando datos FanGraphs...")
    ws = wb[DATA_SHEET]
    def safe(fn, *a):
        try: return fn(*a)
        except requests.exceptions.HTTPError as e:
            if "403" in str(e):
                print(f"  ❌ 403 — necesitas cookie de FanGraphs")
                print(f"     export FANGRAPHS_COOKIE_NAME='wordpress_logged_in_...'")
                print(f"     export FANGRAPHS_COOKIE='tu_valor_de_cookie'")
            else: print(f"  ⚠️  {e}")
            return {}
        except Exception as e: print(f"  ⚠️  {e}"); return {}

    wrc26  = safe(get_team_wrc,          fg, SEASON)
    wrc25  = safe(get_team_wrc,          fg, "2025")
    bp26   = safe(get_bullpen_xfip,      fg, SEASON)
    bp25   = safe(get_bullpen_xfip,      fg, "2025")
    sp26   = safe(get_sp_xfip,           fg, SEASON)
    sp25   = safe(get_sp_xfip,           fg, "2025")
    splits = safe(get_team_wrc_splits,   fg, SEASON)

    if not any([wrc26, wrc25, bp26, bp25, sp26, sp25]):
        print("  ❌ No se obtuvo ningún dato. Verifica cookie."); return

    # ── 1. Guardar a JSON cache (fuente principal) ─────────────────────────
    raw = load_fg_cache()
    if wrc26: raw["wrc_2026"]     = wrc26
    if wrc25: raw["wrc_2025"]     = wrc25
    if bp26:  raw["bp_2026"]      = bp26
    if bp25:  raw["bp_2025"]      = bp25
    if sp26:  raw["sp_xfip_2026"] = sp26
    if sp25:  raw["sp_xfip_2025"] = sp25
    if splits.get("vs_rhp"): raw["wrc_vs_rhp"] = splits["vs_rhp"]
    if splits.get("vs_lhp"): raw["wrc_vs_lhp"] = splits["vs_lhp"]
    save_fg_cache(raw)
    blended = blend_fg_data(raw)
    n_sp = len(raw.get("sp_xfip_2025", {}))
    print(f"  ✅ JSON caché guardada: {os.path.basename(FG_CACHE_FILE)} ({len(blended)} equipos, {n_sp} SP)")

    # ── 2. También actualizar Excel como referencia visual ─────────────────
    updated = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        abb  = str(row[2].value or "").strip().upper()
        team = TEAM_ABB.get(abb)
        if not team: continue
        r = row[2].row
        if team in bp26:  ws.cell(r, 11).value = bp26[team]
        if team in wrc26: ws.cell(r, 12).value = wrc26[team]
        if team in wrc25: ws.cell(r, 14).value = wrc25[team]
        if team in bp25:  ws.cell(r, 15).value = bp25[team]
        updated += 1
    print(f"  ✅ {updated} equipos también actualizados en MLB Data (Excel)")


# ──────────────────────────────────────────────────────
# COMPUTE LINES FROM EXCEL
# ──────────────────────────────────────────────────────

def compute_lines_from_api(silent=False, skip_lineups=False):
    """
    Jala schedule + weather del API y calcula líneas del modelo para TARGET_DATE.
    No lee ni escribe el Excel — idéntico al flujo diario pero sin guardar.
    skip_lineups=True: omite _get_confirmed_lineups (mucho más rápido, útil para --log).
    Retorna (games_with_lines, odds).
    """
    global FG_FIP_AVAILABLE
    fg       = fg_session()
    mlb_data = load_fg_blended(fg)
    sp_xfip  = load_sp_xfip_blended()
    FG_FIP_AVAILABLE = len(sp_xfip) > 0
    if not FG_FIP_AVAILABLE and not silent:
        print("  ⚠️  SP xFIP no en caché — corre --refresh para actualizar")

    # Umpire data (tendencias + asignaciones de hoy)
    ump_tendencies = load_ump_cache()
    ump_assignments = get_game_umpires(TARGET_DATE) if ump_tendencies else {}

    games = get_mlb_schedule(TARGET_DATE)
    odds  = get_market_odds() if ODDS_API_KEY else {}

    if not games:
        return [], odds

    import concurrent.futures as _cf

    # ── Factores de contexto en PARALELO ─────────────────────────────────
    # Antes: secuencial (~8-15s). Ahora: concurrente → tiempo del más lento.
    # Todos son independientes entre sí y del schedule, así que se pueden
    # lanzar simultáneamente.
    _ctx_futures = {}
    with _cf.ThreadPoolExecutor(max_workers=6) as _ctx_pool:
        _ctx_futures["recent_form"]   = _ctx_pool.submit(_get_mlb_recent_form,    TARGET_DATE)
        _ctx_futures["wrc_splits"]    = _ctx_pool.submit(load_wrc_splits)
        _ctx_futures["standings"]     = _ctx_pool.submit(_get_mlb_standings)
        _ctx_futures["bp_fatigue"]    = _ctx_pool.submit(_get_bullpen_fatigue,     TARGET_DATE)
        _ctx_futures["rolling_wrc"]   = _ctx_pool.submit(_get_rolling_wrc,         TARGET_DATE)
        _ctx_futures["home_away_spl"] = _ctx_pool.submit(_get_home_away_splits,    TARGET_DATE[:4])
        # Resultados se bloquean al leer (ya terminaron para cuando los usamos)

    recent_form   = _ctx_futures["recent_form"].result()
    wrc_splits    = _ctx_futures["wrc_splits"].result()
    standings     = _ctx_futures["standings"].result()
    bp_fatigue    = _ctx_futures["bp_fatigue"].result()
    rolling_wrc   = _ctx_futures["rolling_wrc"].result()
    home_away_spl = _ctx_futures["home_away_spl"].result()

    # ── Pre-fetch lineups + SP FIP + player stats en paralelo ────────────
    # Batch A (concurrente): Lineups + SP FIP by ID (independientes entre sí)
    # Batch B (después): Player OPS (necesita IDs de Batch A)
    if not skip_lineups:
        _pre_pks    = [g.get("game_pk")    for g in games if g.get("game_pk")]
        _sp_ids_all = [sp_id for g in games
                       for sp_id in (g.get("away_sp_id"), g.get("home_sp_id"))
                       if sp_id is not None]
        _season_pre = TARGET_DATE[:4]

        # Batch A: lineups + SP FIP simultáneamente
        with _cf.ThreadPoolExecutor(max_workers=min(len(_pre_pks) + len(_sp_ids_all), 20)) as _poolA:
            _lu_futs  = {pk: _poolA.submit(_get_confirmed_lineups, pk)  for pk in _pre_pks}
            _sp_futs  = {sid: _poolA.submit(_get_sp_season_fip_by_id, sid) for sid in _sp_ids_all}
            # Esperamos a que terminen antes del Batch B
            for fut in _lu_futs.values():  fut.result()
            for fut in _sp_futs.values():  fut.result()

        # Batch B: Player OPS (depende de Batch A — IDs ya en _LINEUP_CACHE)
        _all_pids = set()
        for _pk in _pre_pks:
            for _side_ids in _LINEUP_CACHE.get(_pk, {}).values():
                _all_pids.update(_side_ids)
        if _all_pids:
            def _prefetch_ops(pid):
                _get_player_ops(pid, _season_pre)
            with _cf.ThreadPoolExecutor(max_workers=min(len(_all_pids), 20)) as _poolB:
                list(_poolB.map(_prefetch_ops, _all_pids))

        # Persistir a disco para que próximas corridas del mismo día sean instantáneas
        _save_disk_cache()

    games_with_lines = []
    for g in games:
        away, home       = g["away"], g["home"]
        away_sp, home_sp = g["away_sp"], g["home_sp"]
        game_utc         = g.get("game_time_utc", "")

        def _blend_fg_fullname(name):
            """FanGraphs xFIP — SOLO por nombre completo. Sin fallback de apellido."""
            if not name or name == "TBD":
                return None, "TBD"
            keys = name_keys(name)
            for k in keys:
                if " " in k:          # nombre completo únicamente
                    v = sp_xfip.get(k)
                    if v:
                        return v, f"FG xFIP ({v:.2f})"
            return None, "FG: sin match nombre completo"

        # ════════════════════════════════════════════════════════════════
        # SP FIP — FUENTE PRIMARIA: statsapi por MLB pitcher ID (sin nombre)
        # FUENTE SECUNDARIA: FanGraphs xFIP por nombre completo (refinement)
        # ════════════════════════════════════════════════════════════════
        away_sp_id = g.get("away_sp_id")
        home_sp_id = g.get("home_sp_id")

        sp_id_stats_a = _get_sp_season_fip_by_id(away_sp_id)
        sp_id_stats_b = _get_sp_season_fip_by_id(home_sp_id)

        fg_xfip_a, fg_src_a = _blend_fg_fullname(away_sp) if away_sp != "TBD" else (None, "TBD")
        fg_xfip_b, fg_src_b = _blend_fg_fullname(home_sp) if home_sp != "TBD" else (None, "TBD")

        def _merge_sp_fip(sp_name, sp_id_stats, fg_xfip, fg_src):
            """
            Combina statsapi FIP (por ID) con FanGraphs xFIP (por nombre completo).
            Prioridad: statsapi ID > FanGraphs > fallback.
            Si GS=0 → debut → retorna (None, debut_flag=True).
            """
            if sp_id_stats is not None:
                gs         = sp_id_stats["gs"]
                g          = sp_id_stats.get("g", 0)
                fip        = sp_id_stats["fip"]
                ip         = sp_id_stats["ip"]
                is_opener  = sp_id_stats.get("is_opener", False)

                if gs == 0:
                    if is_opener:
                        # ── Opener / bullpen game — NOT a debut ─────────────
                        # Este pitcher es un RP fungiendo como opener hoy.
                        # Usar su FIP de RP si existe; si no, usar BP proxy del equipo.
                        # No suprimir picks — ajustar xFIP como "bullpen game".
                        if fip is not None:
                            # FIP calculado de sus IP como RP
                            src = (f"🔀 OPENER (RP {g}ap/{ip:.1f}IP) — "
                                   f"usando FIP reliever {fip:.2f}")
                            return fip, False, src
                        elif fg_xfip is not None:
                            src = (f"🔀 OPENER (RP {g}ap) — "
                                   f"FG xFIP fallback {fg_xfip:.2f}")
                            return fg_xfip, False, src
                        else:
                            # Sin FIP ni xFIP → usar BP proxy (None, debut=False)
                            return None, False, f"🔀 OPENER (RP {g}ap) — sin FIP, usando BP proxy"
                    else:
                        # Verdadero debut (0 GS, 0 o pocas apariciones)
                        return None, True, f"DEBUT ({sp_name}) — 0 GS en MLB esta temporada"

                if fip is None:
                    # Tiene GS pero IP mínima — usar FG si existe
                    if fg_xfip is not None:
                        return fg_xfip, False, f"statsapi GS={gs} IP mínima → FG xFIP ({fg_xfip:.2f})"
                    return None, False, f"SP data insuficiente (GS={gs}, IP={ip:.1f})"
                # Tiene FIP real de statsapi — blend ponderado por GS (no 50/50 fijo)
                # Con pocos GS, el FIP de statsapi es ruidoso: 1 mala salida puede
                # hacer el FIP = 8.0 con 8 IP. Dar más peso al FG xFIP (más estable)
                # hasta que el pitcher acumule sample suficiente (~15 GS).
                #   1 GS  →  10% statsapi, 90% FG
                #   5 GS  →  33% statsapi, 67% FG
                #   10 GS →  50% statsapi, 50% FG
                #   15 GS →  67% statsapi, 33% FG
                #   20+ GS → 80% statsapi, 20% FG
                if fg_xfip is not None:
                    _gs_w = min(0.80, gs / 25.0)   # peso statsapi: sube con GS, cap 80%
                    _fg_w = 1.0 - _gs_w
                    blended = round(_gs_w * fip + _fg_w * fg_xfip, 2)
                    src = (f"✅ GS-blend statsapi FIP={fip:.2f}×{_gs_w:.0%} (GS={gs},IP={ip:.1f})"
                           f" + FG={fg_xfip:.2f}×{_fg_w:.0%} = {blended:.2f}")
                    return blended, False, src
                else:
                    src = f"✅ statsapi-ID FIP={fip:.2f} (GS={gs}, IP={ip:.1f}) — FG sin match"
                    return fip, False, src
            else:
                # statsapi falló (sin internet o pitcher sin data) → FG fallback
                if fg_xfip is not None:
                    if fg_xfip < 3.20:
                        # xFIP muy bajo sin data MLB = probable proyección/MiLB (debut pitcher)
                        # Regresar hacia la media de la liga para no inflar el Under artificialmente
                        LEAGUE_AVG_XFIP = 4.15
                        blended = round(0.35 * fg_xfip + 0.65 * LEAGUE_AVG_XFIP, 2)
                        return blended, False, f"⚠️ statsapi fail → FG xFIP {fg_xfip:.2f} (proyección/debut) → regressed {blended:.2f}"
                    return fg_xfip, False, f"⚠️ statsapi fail → FG xFIP ({fg_xfip:.2f})"
                return None, False, "⚠️ Sin data SP — usando BP como proxy"

        fip_a, _debut_a, src_a = _merge_sp_fip(away_sp, sp_id_stats_a, fg_xfip_a, fg_src_a)
        fip_b, _debut_b, src_b = _merge_sp_fip(home_sp, sp_id_stats_b, fg_xfip_b, fg_src_b)

        # ── DATA AUDIT: imprime fuente de cada SP antes de continuar ─────
        print(f"  📋 SP AUDIT  {away:12s}  {away_sp[:22]:<22}  {src_a}")
        print(f"  📋 SP AUDIT  {home:12s}  {home_sp[:22]:<22}  {src_b}")

        # ── Debut / Opener detection ──────────────────────────────────────
        # _debut_*: True = verdadero debut (sin historial MLB), picks suprimidos
        # is_opener: False debut, se usa FIP de RP, picks NO suprimidos
        if _debut_a:
            print(f"  🆕 DEBUT: {away_sp} ({away}) — picks suprimidos (0 GS MLB)")
        elif sp_id_stats_a and sp_id_stats_a.get("is_opener"):
            print(f"  🔀 OPENER: {away_sp} ({away}) — RP como opener, usando FIP reliever (picks NO suprimidos)")
        if _debut_b:
            print(f"  🆕 DEBUT: {home_sp} ({home}) — picks suprimidos (0 GS MLB)")
        elif sp_id_stats_b and sp_id_stats_b.get("is_opener"):
            print(f"  🔀 OPENER: {home_sp} ({home}) — RP como opener, usando FIP reliever (picks NO suprimidos)")

        # ── SP forma reciente: blend FIP temporada + ERA últimas 5 salidas ─
        # _get_sp_recent_era también popula _SP_GS_CACHE (conteo total GS)
        recent_era_a = _get_sp_recent_era(away_sp_id)
        recent_era_b = _get_sp_recent_era(home_sp_id)

        if fip_a is not None:
            fip_a = _sp_recent_adj_xfip(fip_a, recent_era_a)
        if fip_b is not None:
            fip_b = _sp_recent_adj_xfip(fip_b, recent_era_b)

        # ── SP H2H vs equipo específico ────────────────────────────────────
        # El SP away enfrenta el lineup del home — necesito el H2H del SP away vs home.
        # El SP home enfrenta el lineup del away — necesito el H2H del SP home vs away.
        # Ej: Pablo López (MIN) vs Yankees → si tiene 5.40 ERA en 4 starts vs NYY,
        #     el modelo ahora ajusta su xFIP al alza (beneficia a los Yankees en el ML).
        away_sp_id = g.get("away_sp_id")
        home_sp_id = g.get("home_sp_id")
        h2h_a = _get_sp_vs_team(away_sp_id, home)   # SP away vs equipo home
        h2h_b = _get_sp_vs_team(home_sp_id, away)   # SP home vs equipo away
        _h2h_note_a = _h2h_note_b = None
        if fip_a is not None and h2h_a:
            _fip_a_pre = fip_a
            fip_a = _h2h_xfip_adj(fip_a, h2h_a)
            if abs(fip_a - _fip_a_pre) >= 0.05:
                _h2h_note_a = (f"H2H {away_sp} vs {home}: {h2h_a['era']:.2f} ERA "
                               f"en {h2h_a['gs']} GS → xFIP {_fip_a_pre:.2f}→{fip_a:.2f}")
        if fip_b is not None and h2h_b:
            _fip_b_pre = fip_b
            fip_b = _h2h_xfip_adj(fip_b, h2h_b)
            if abs(fip_b - _fip_b_pre) >= 0.05:
                _h2h_note_b = (f"H2H {home_sp} vs {away}: {h2h_b['era']:.2f} ERA "
                               f"en {h2h_b['gs']} GS → xFIP {_fip_b_pre:.2f}→{fip_b:.2f}")
        if _h2h_note_a: print(f"  ⚔️  {_h2h_note_a}")
        if _h2h_note_b: print(f"  ⚔️  {_h2h_note_b}")

        d_a = mlb_data.get(away, {}); d_b = mlb_data.get(home, {})
        wrc_a_base = d_a.get("wrc", 100);  wrc_b_base = d_b.get("wrc", 100)
        # Bullpen xFIP ajustado por fatiga de los últimos 3 días
        bp_a  = _bullpen_fatigue_adj(away, d_a.get("bp_xfip", 4.2), bp_fatigue)
        bp_b  = _bullpen_fatigue_adj(home, d_b.get("bp_xfip", 4.2), bp_fatigue)

        # ── Home/Away split: equipos visita rinden menos ─────────────────
        wrc_a_base = _apply_home_away_split(away, False, home_away_spl, wrc_a_base)
        wrc_b_base = _apply_home_away_split(home, True,  home_away_spl, wrc_b_base)

        # ── Rolling 14-day wRC+ (forma ofensiva reciente) ─────────────────
        wrc_a_base = _rolling_wrc_factor(away, rolling_wrc, wrc_a_base)
        wrc_b_base = _rolling_wrc_factor(home, rolling_wrc, wrc_b_base)

        # ── Confirmed lineup: recalcular wRC+ con el batting order real ───
        season_yr = TARGET_DATE[:4]
        game_pk   = g.get("game_pk")
        lineups   = {} if skip_lineups else (_get_confirmed_lineups(game_pk) if game_pk else {})
        wrc_a_lineup = None
        wrc_b_lineup = None
        # lineup_used = True si la API devolvió bateadores confirmados (independiente de si cambió wRC+)
        lineup_used_a = bool(lineups.get("away"))
        lineup_used_b = bool(lineups.get("home"))
        if lineups.get("away"):
            wrc_a_lineup = _lineup_weighted_wrc(lineups["away"], wrc_a_base, season_yr)
            if wrc_a_lineup != wrc_a_base:
                wrc_a_base = wrc_a_lineup
        if lineups.get("home"):
            wrc_b_lineup = _lineup_weighted_wrc(lineups["home"], wrc_b_base, season_yr)
            if wrc_b_lineup != wrc_b_base:
                wrc_b_base = wrc_b_lineup

        # ── Impact log: muestra cómo cada ajuste movió el wRC+ ──────────────
        _wrc_raw_a = d_a.get("wrc", 100)
        _wrc_raw_b = d_b.get("wrc", 100)
        _roll_entry_a = rolling_wrc.get(away, {})
        _roll_entry_b = rolling_wrc.get(home, {})
        print(f"  🔬 {away:12s} wRC+ base:{_wrc_raw_a:.0f}"
              f"  →HA:{wrc_a_base:.0f}"   # after home/away + rolling but not lineup yet
              + (f"  rolling14:{_roll_entry_a.get('wrc',0):.0f}(PA={_roll_entry_a.get('pa',0)})" if _roll_entry_a else "  rolling14:N/A")
              + (f"  lineup:✅" if lineup_used_a else ""))
        print(f"  🔬 {home:12s} wRC+ base:{_wrc_raw_b:.0f}"
              f"  →HA:{wrc_b_base:.0f}"
              + (f"  rolling14:{_roll_entry_b.get('wrc',0):.0f}(PA={_roll_entry_b.get('pa',0)})" if _roll_entry_b else "  rolling14:N/A")
              + (f"  lineup:✅" if lineup_used_b else ""))

        # ── Platoon adjustment: wRC+ vs la mano del SP rival ─────────────
        # Away team batea contra el SP del home (home_sp_hand)
        # Home team batea contra el SP del away (away_sp_hand)
        home_sp_hand = _get_pitcher_hand(g.get("home_sp_id"))  # mano del SP home
        away_sp_hand = _get_pitcher_hand(g.get("away_sp_id"))  # mano del SP away

        def _platoon_wrc(team, sp_hand, base_wrc):
            """Usa wRC+ platoon si disponible; si no, usa base."""
            if sp_hand == "L":
                split_val = wrc_splits["vs_lhp"].get(team)
            elif sp_hand in ("R", "S"):
                split_val = wrc_splits["vs_rhp"].get(team)
            else:
                split_val = None
            if split_val:
                # Blend 50/50 platoon split vs base (splits tienen menos sample)
                return round((float(split_val) * 0.5 + base_wrc * 0.5), 1), sp_hand
            return base_wrc, None

        wrc_a, platoon_used_a = _platoon_wrc(away, home_sp_hand, wrc_a_base)
        wrc_b, platoon_used_b = _platoon_wrc(home, away_sp_hand, wrc_b_base)

        # ── Recent form + rest + standings factors ────────────────────────
        form_a = _form_factor(away, recent_form)
        form_b = _form_factor(home, recent_form)
        rest_a = _rest_factor(away, recent_form)
        rest_b = _rest_factor(home, recent_form)
        std_a  = _standings_factor(away, standings)
        std_b  = _standings_factor(home, standings)

        # SP xFIP fallback: si no está en caché usar 3.90 (SP MLB promedio ligeramente mejor que liga)
        SP_FALLBACK = 3.90

        def _expected_ip(sp_xfip):
            """Aces van más innings. Ajuste adaptativo reduce el peso del bullpen
            para pitchers de calidad — antes era fijo a 5.0 para todos."""
            if sp_xfip is None:        return 5.0   # desconocido → promedio
            if sp_xfip < 3.2:          return 6.5   # Ace elite (Sanchez, Sale, Scherzer)
            elif sp_xfip < 3.6:        return 6.0   # Buen abridor
            elif sp_xfip < 4.0:        return 5.5   # Promedio
            else:                      return 5.0   # Débil / volátil

        ip_a = _expected_ip(fip_a)
        ip_b = _expected_ip(fip_b)
        xfip_a = calc_xfip_tot(fip_a if fip_a else SP_FALLBACK, bp_a, ip_a)
        xfip_b = calc_xfip_tot(fip_b if fip_b else SP_FALLBACK, bp_b, ip_b)

        weather = get_weather(home, game_time_utc=game_utc)
        time.sleep(0.3)
        pf    = calc_pf_combined(home, weather["dir"], weather["mph"], weather["temp"], weather.get("humidity", 50))

        # Extraer hora local para ajuste día/noche
        try:
            from datetime import datetime, timedelta as _td
            _tz = STADIUM_TZ_OFFSET.get(home, -5)
            _clean = game_utc.replace("Z","").split("+")[0]
            _utc_dt = datetime.strptime(_clean, "%Y-%m-%dT%H:%M:%S")
            _local_h = (_utc_dt + _td(hours=_tz)).hour
        except Exception:
            _local_h = None

        # Ajuste de umpire HP
        ump_name   = ump_assignments.get((away, home), "")
        _ump_f     = ump_total_factor(ump_name, ump_tendencies)

        # GS qualifier para ace suppressor — sólo pitcher con ≥8 GS probados cuenta
        gs_a = sp_id_stats_a["gs"] if sp_id_stats_a else None
        gs_b = sp_id_stats_b["gs"] if sp_id_stats_b else None

        lns   = calc_lines(wrc_a, wrc_b, xfip_a, xfip_b, pf,
                           game_hour_local=_local_h, ump_factor=_ump_f,
                           form_a=form_a, form_b=form_b,
                           rest_a=rest_a, rest_b=rest_b,
                           standings_a=std_a, standings_b=std_b,
                           sp_xfip_raw_away=fip_a, sp_xfip_raw_home=fip_b,
                           gs_away=gs_a, gs_home=gs_b,
                           sp_name_away=away_sp, sp_name_home=home_sp)
        gtime = format_game_time(game_utc, home)

        games_with_lines.append({
            **g,
            "fip_a": fip_a, "fip_b": fip_b,
            "weather": weather, "lines": lns, "game_time_local": gtime,
            "fip_source": "" if FG_FIP_AVAILABLE else "⚠️ FG no disponible (BP como proxy)",
            "fip_source_a": src_a, "fip_source_b": src_b,
            "wrc_a": wrc_a, "wrc_b": wrc_b,
            "wrc_a_base": wrc_a_base, "wrc_b_base": wrc_b_base,
            "platoon_a": platoon_used_a, "platoon_b": platoon_used_b,
            "home_sp_hand": home_sp_hand, "away_sp_hand": away_sp_hand,
            "bp_a": bp_a, "bp_b": bp_b,
            "xfip_a": xfip_a, "xfip_b": xfip_b,
            "ump_hp": ump_name, "ump_factor": _ump_f,
            "form_a": form_a, "form_b": form_b,
            "rest_a": rest_a, "rest_b": rest_b,
            "standings_a": std_a, "standings_b": std_b,
            "win_pct_away": standings.get(away, 0.500),
            "win_pct_home": standings.get(home, 0.500),
            "bp_fatigue_away": bp_fatigue.get(away, 0.0),
            "bp_fatigue_home": bp_fatigue.get(home, 0.0),
            "bp_a_adj": bp_a, "bp_b_adj": bp_b,
            "recent_era_away": recent_era_a,
            "recent_era_home": recent_era_b,
            "_debut_away": _debut_a,   # True si SP away tiene 0 GS esta temporada (verdadero debut)
            "_debut_home": _debut_b,   # True si SP home tiene 0 GS esta temporada (verdadero debut)
            "_opener_away": sp_id_stats_a.get("is_opener", False) if sp_id_stats_a else False,
            "_opener_home": sp_id_stats_b.get("is_opener", False) if sp_id_stats_b else False,
            "h2h_away_sp": h2h_a,    # {era, ip, gs} del SP away vs equipo home
            "h2h_home_sp": h2h_b,    # {era, ip, gs} del SP home vs equipo away
            "recent_form_away": recent_form.get(away, {}),
            "recent_form_home": recent_form.get(home, {}),
            "lineup_used_away": lineup_used_a,
            "lineup_used_home": lineup_used_b,
            "lineup_names_away": _LINEUP_NAMES_CACHE.get(game_pk, {}).get("away", []),
            "lineup_names_home": _LINEUP_NAMES_CACHE.get(game_pk, {}).get("home", []),
            "rolling_wrc_away":    rolling_wrc.get(away, {}).get("wrc"),   # float o None
            "rolling_wrc_away_pa": rolling_wrc.get(away, {}).get("pa", 0),
            "rolling_wrc_home":    rolling_wrc.get(home, {}).get("wrc"),   # float o None
            "rolling_wrc_home_pa": rolling_wrc.get(home, {}).get("pa", 0),
            "home_away_split_away": home_away_spl.get(away, {}).get("away"),
            "home_away_split_home": home_away_spl.get(home, {}).get("home"),
        })

    return games_with_lines, odds


def compute_lines_from_excel():
    global FG_FIP_AVAILABLE
    wb        = load_workbook(EXCEL_FILE, data_only=True, keep_links=False)
    mlb_data  = load_fg_blended()       # wRC+/BP xFIP desde caché FanGraphs
    sp_xfip_c = load_sp_xfip_blended()  # SP xFIP blended desde caché FanGraphs
    FG_FIP_AVAILABLE = len(sp_xfip_c) > 0   # actualiza flag para display_lines/show_picks
    ws        = wb[LINES_SHEET]
    target_dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d").date()

    # Intentar obtener horas del schedule (silencioso)
    time_map = {}
    try:
        sched = get_mlb_schedule(TARGET_DATE, silent=True)
        time_map = {f"{g['away']} vs {g['home']}":
                    format_game_time(g["game_time_utc"], g["home"]) for g in sched}
    except Exception: pass

    games_with_lines = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        try:
            row_date = row[0].date() if isinstance(row[0], datetime) \
                       else datetime.strptime(str(row[0])[:10],"%Y-%m-%d").date()
        except: continue
        if row_date != target_dt: continue

        away   = str(row[1]  or "").strip(); home   = str(row[7]  or "").strip()
        away_sp= str(row[3]  or "TBD").strip(); home_sp= str(row[9] or "TBD").strip()
        wind_d = str(row[20] or "").strip(); wind_m = row[21]; temp = row[22]
        if not away or not home: continue

        d_away=mlb_data.get(away,{}); d_home=mlb_data.get(home,{})
        wrc_a=d_away.get("wrc",100); wrc_b=d_home.get("wrc",100)
        bp_a=d_away.get("bp_xfip",4.2); bp_b=d_home.get("bp_xfip",4.2)

        # SP xFIP desde caché FanGraphs (no Excel)
        def _sp(name):
            for k in name_keys(name.strip().upper()):
                v = sp_xfip_c.get(k)
                if v: return v
            return None

        f_a = _sp(away_sp) if away_sp not in ("TBD","") else None
        f_b = _sp(home_sp) if home_sp not in ("TBD","") else None
        SP_FALLBACK = 3.90
        f_a = f_a if f_a else SP_FALLBACK
        f_b = f_b if f_b else SP_FALLBACK
        # Adaptive IP: aces go deeper, reducing bullpen inflation
        def _exp_ip(x): return 6.5 if x<3.2 else (6.0 if x<3.6 else (5.5 if x<4.0 else 5.0))
        xfip_a=calc_xfip_tot(f_a,bp_a,_exp_ip(f_a)); xfip_b=calc_xfip_tot(f_b,bp_b,_exp_ip(f_b))
        pf=calc_pf_combined(home, wind_d or None, int(wind_m) if wind_m else 0,
                            int(temp) if temp else None)
        # Extraer hora local desde gtime para ajuste día/noche
        _gtime_raw = time_map.get(f"{away} vs {home}", time_map.get(f"{home} vs {away}", ""))
        try:
            import re as _re
            _hm = _re.search(r'(\d{1,2}):(\d{2})\s*(AM|PM)', str(_gtime_raw), _re.IGNORECASE)
            if _hm:
                _h, _m, _ap = int(_hm.group(1)), int(_hm.group(2)), _hm.group(3).upper()
                if _ap == "PM" and _h != 12: _h += 12
                if _ap == "AM" and _h == 12: _h = 0
                _local_h_ex = _h
            else:
                _local_h_ex = None
        except Exception:
            _local_h_ex = None
        lns=calc_lines(wrc_a,wrc_b,xfip_a,xfip_b,pf,game_hour_local=_local_h_ex,
                       sp_xfip_raw_away=f_a, sp_xfip_raw_home=f_b,
                       sp_name_away=away_sp, sp_name_home=home_sp)
        wthr={"dir":wind_d or None,"mph":int(wind_m) if wind_m else 0,
              "temp":int(temp) if temp else None,"raw_dir":wind_d or None}
        gtime = _gtime_raw

        games_with_lines.append({
            "away":away,"home":home,"away_sp":away_sp,"home_sp":home_sp,
            "weather":wthr,"lines":lns,"fip_a":f_a,"fip_b":f_b,
            "game_time_local":gtime,
            "wrc_a":wrc_a,"wrc_b":wrc_b,"bp_a":bp_a,"bp_b":bp_b,
            "xfip_a":xfip_a,"xfip_b":xfip_b,
        })

    odds = get_market_odds() if ODDS_API_KEY else {}
    return games_with_lines, odds, wb


# ──────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n{'═'*52}")
    print(f"  LABOY PICKS — MLB")
    if   REFRESH_MODE:      mode = "REFRESH DATA"
    elif LINES_MODE:        mode = "SHOW LINES"
    elif PICKS_MODE:        mode = "EV PICKS (terminal)"
    elif EXPORT_LINES:      mode = "EXPORT HTML"
    elif PICKS_EXPORT:      mode = "EXPORT HTML"         # legacy alias
    elif STATS_MODE:        mode = "TEAM STATS"
    elif STATS_RAW_MODE:    mode = "STATS RAW"
    elif EXPORT_MODE:       mode = "EXPORT LINES TXT"
    elif RECORD_MODE:       mode = "RECORD"
    elif FEEDBACK_MODE:     mode = "FEEDBACK (with AI)"
    elif LOG_MODE:          mode = "LOG PICK"
    elif GRADE_MODE:        mode = "GRADE PICK"
    elif EXPORT_RECORD_MODE:mode = "EXPORT RECORD"
    elif REMOVE_MODE:       mode = "REMOVE PICK"
    elif EXPORT_DEBUG_MODE: mode = "EXPORT DEBUG HTML"
    elif DEBUG_GAME_MODE:   mode = "DEBUG GAME"
    elif GRADE_PICKS_MODE:  mode = "GRADE PICKS"
    elif WEATHER_MODE:      mode = "🌤️  WEATHER CHECK"
    else:                   mode = f"DAILY — {TARGET_DATE}"
    # Session suffix
    if   DAY_SESSION:    mode += "  ☀️  SESIÓN DÍA   (< 5 PM ET)"
    elif NIGHT_SESSION:  mode += "  🌙  SESIÓN NOCHE (≥ 5 PM ET)"
    if REQUIRE_LINEUPS:  mode += "  ✅ --confirmed"
    print(f"  {mode}")
    print(f"{'═'*52}\n")

    try:
        # ── Tracker (no Excel) ────────────────────────────
        if GRADE_PICKS_MODE: cmd_grade_picks();  sys.exit(0)
        if LOG_MODE:         cmd_log_pick();    sys.exit(0)
        if EXPORT_LOG_MODE:  cmd_export_log();  sys.exit(0)
        if EXPORT_RECORD_MODE:
            card_path = export_record_card(TARGET_DATE)
            if card_path:
                print(f"\n  📄 HTML: {os.path.basename(card_path)}")
                if PUBLISH_MODE:
                    try:
                        cmd_publish([card_path])
                    except Exception as _ep:
                        print(f"  ⚠️  Publish falló: {_ep}")
            else:
                print("\n  ⚠️  No hay picks en el log para esa fecha.")
            sys.exit(0)
        if GRADE_MODE:       cmd_grade_pick();  sys.exit(0)
        if REMOVE_MODE:      cmd_remove_pick(); sys.exit(0)
        if RECORD_MODE:      cmd_record();      sys.exit(0)
        if FEEDBACK_MODE:    cmd_feedback();    sys.exit(0)

        # ── Stats desde caché FanGraphs (sin Excel) ───────
        if STATS_MODE or STATS_RAW_MODE:
            if STATS_MODE:     show_stats()
            if STATS_RAW_MODE: show_stats_raw()
            sys.exit(0)

        if LINES_MODE or PICKS_MODE or PICKS_EXPORT or EXPORT_LINES or EXPORT_PICKS or EXPORT_MODE or DEBUG_GAME_MODE or EXPORT_DEBUG_MODE or WEATHER_MODE:
            # --lines / --picks / --export-lines / --export-debug / --weather: jala desde API directamente (sin Excel)
            print("Jalando data...\n")
            games, odds = compute_lines_from_api()
            _save_results_cache(games)   # cache para debug cards en --log
            games = _filter_by_session(games)
            if not games:
                _sess = "de día (< 5 PM ET)" if DAY_SESSION else "de noche (≥ 5 PM ET)" if NIGHT_SESSION else ""
                print(f"❌ No hay juegos MLB {_sess}para {TARGET_DATE}.")
                sys.exit(0)
            if DAY_SESSION or NIGHT_SESSION:
                _lbl = "☀️  DÍA" if DAY_SESSION else "🌙  NOCHE"
                print(f"  ⚡ Sesión {_lbl}: {len(games)} juego(s) filtrado(s)\n")
            if WEATHER_MODE:            cmd_weather(games); sys.exit(0)
            if LINES_MODE:              display_lines(games, odds)
            if PICKS_MODE:
                show_picks(games, odds)
                # Auto-generate the rich debug fragment for the dashboard.
                # export_debug_picks_html reads from mlb_debug_state.json (already
                # written by show_picks) so no API re-fetch happens here.
                if not EXPORT_DEBUG_MODE:   # avoid double-call if both flags present
                    try:
                        export_debug_picks_html(games, odds, _dashboard_mode=True)
                    except Exception as _de:
                        print(f"  ⚠️  Debug fragment: {_de}")
            _published_htmls = []
            if EXPORT_LINES or PICKS_EXPORT:
                lp = export_lines_html(games, odds)
                if lp: _published_htmls.append(lp)
            if EXPORT_PICKS:
                pp = export_picks_html(None, games, odds)
                if pp: _published_htmls.append(pp)
            if EXPORT_DEBUG_MODE:
                dp = export_debug_picks_html(games, odds)
                if dp: _published_htmls.append(dp)
            if EXPORT_MODE:
                fp = export_lines(games, odds)
                print(f"  TXT: {fp}\n")
            if DEBUG_GAME_MODE:         cmd_debug_game(games, odds)
            if PUBLISH_MODE:            cmd_publish(_published_htmls or [])
            sys.exit(0)

        # ── Refresh FanGraphs data ──────────────────────
        if REFRESH_MODE:
            wb = load_workbook(EXCEL_FILE, keep_links=False)
            fg = fg_session()
            refresh_mlb_data(wb, fg)
            wb.save(EXCEL_FILE)
            print(f"\n💾 Guardado: {EXCEL_FILE}")
            sys.exit(0)

        # ── Daily mode (jala API + guarda en Excel) ─────────
        print("Jalando data...\n")
        games_with_lines, odds = compute_lines_from_api()
        _save_results_cache(games_with_lines)   # cache para debug cards en --log
        games_with_lines = _filter_by_session(games_with_lines)

        if not games_with_lines:
            _sess = "de día (< 5 PM ET)" if DAY_SESSION else "de noche (≥ 5 PM ET)" if NIGHT_SESSION else ""
            print(f"❌ No hay juegos MLB {_sess}para esta fecha."); sys.exit(0)

        if DAY_SESSION or NIGHT_SESSION:
            _lbl = "☀️  DÍA" if DAY_SESSION else "🌙  NOCHE"
            print(f"  ⚡ Sesión {_lbl}: {len(games_with_lines)} juego(s) filtrado(s)\n")

        display_lines(games_with_lines, odds)
        show_picks(games_with_lines, odds)

        # Exportar TXT de líneas
        export_lines(games_with_lines, odds)

        # Guardar en Excel
        wb = load_workbook(EXCEL_FILE, keep_links=False)
        saved = update_excel(wb, games_with_lines)
        if saved:
            wb.save(EXCEL_FILE)
            print(f"💾 Excel actualizado: {EXCEL_FILE}")

    except FileNotFoundError:
        print(f"❌ No encontré el Excel en:\n   {EXCEL_FILE}")
    except requests.exceptions.ConnectionError:
        print("❌ Sin conexión a internet.")
    except Exception as e:
        print(f"❌ Error: {e}")
        if DEBUG: raise
