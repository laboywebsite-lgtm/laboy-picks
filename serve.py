#!/usr/bin/env python3
"""
serve.py — Laboy Picks · Servidor móvil unificado
Corre: python3 serve.py
Expón con ngrok: ngrok http 5001
"""
import os, sys, json, re, socket, subprocess, threading, shlex
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import parse_qs, urlparse, unquote_plus

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BSN_DIR  = os.path.join(BASE_DIR, "BSN")
NBA_DIR  = os.path.join(BASE_DIR, "NBA")
MLB_DIR  = os.path.join(BASE_DIR, "MLB")

BSN_LOG  = os.path.join(BSN_DIR, "bsn_picks_log.json")
NBA_LOG  = os.path.join(NBA_DIR, "nba_picks_log.json")
MLB_LOG  = os.path.join(MLB_DIR, "laboy_picks_log.json")
NBA_IR   = os.path.join(NBA_DIR, "nba_injuries.json")
BSN_GP   = os.path.join(BSN_DIR, "bsn_gp.json")
MLB_PICKS = os.path.join(MLB_DIR, "mlb_model_picks.json")

PORT = int(os.environ.get("PORT", 5001))   # Railway/cloud usa $PORT; local usa 5001
for i, a in enumerate(sys.argv):
    if a in ("--port","-p") and i+1 < len(sys.argv):
        try: PORT = int(sys.argv[i+1])
        except: pass

# ── Teams ──────────────────────────────────────────────────────────────
BSN_TEAMS = sorted(["ATLETICOS","CANGREJEROS","CAPITANES","CRIOLLOS","GIGANTES",
                    "INDIOS","LEONES","METS","OSOS","PIRATAS","SANTEROS","VAQUEROS"])
MLB_TEAMS = sorted(["D-BACKS","BRAVES","ORIOLES","RED SOX","CUBS","WHITE SOX",
                    "REDS","GUARDIANS","ROCKIES","TIGERS","ASTROS","ROYALS",
                    "ANGELS","DODGERS","MARLINS","BREWERS","TWINS","METS",
                    "YANKEES","ATHLETICS","PHILLIES","PIRATES","PADRES","GIANTS",
                    "MARINERS","CARDINALS","RAYS","RANGERS","BLUE JAYS","NATIONALS"])
NBA_TEAMS = {
    "ATL":"Hawks","BOS":"Celtics","BKN":"Nets","CHA":"Hornets","CHI":"Bulls",
    "CLE":"Cavaliers","DAL":"Mavericks","DEN":"Nuggets","DET":"Pistons",
    "GSW":"Warriors","HOU":"Rockets","IND":"Pacers","LAC":"Clippers",
    "LAL":"Lakers","MEM":"Grizzlies","MIA":"Heat","MIL":"Bucks",
    "MIN":"Timberwolves","NOP":"Pelicans","NYK":"Knicks","OKC":"Thunder",
    "ORL":"Magic","PHI":"76ers","PHX":"Suns","POR":"Trail Blazers",
    "SAC":"Kings","SAS":"Spurs","TOR":"Raptors","UTA":"Jazz","WAS":"Wizards",
}

# ── Team Colors (bg, fg) ───────────────────────────────────────────────
_MLB_COLORS = {
    "D-BACKS":("#A71930","#E3D4AD"),"BRAVES":("#CE1141","#fff"),
    "ORIOLES":("#DF4601","#000"),"RED SOX":("#BD3039","#fff"),
    "CUBS":("#0E3386","#CC3433"),"WHITE SOX":("#27251F","#C4CED4"),
    "REDS":("#C6011F","#fff"),"GUARDIANS":("#0C2340","#E31937"),
    "ROCKIES":("#33006F","#C4CED4"),"TIGERS":("#0C2340","#FA4616"),
    "ASTROS":("#002D62","#EB6E1F"),"ROYALS":("#004687","#BD9B60"),
    "ANGELS":("#BA0021","#fff"),"DODGERS":("#005A9C","#fff"),
    "MARLINS":("#00A3E0","#EF3340"),"BREWERS":("#FFC52F","#000"),
    "TWINS":("#002B5C","#D31145"),"METS":("#002D72","#FF5910"),
    "YANKEES":("#003087","#fff"),"ATHLETICS":("#003831","#EFB21E"),
    "PHILLIES":("#E81828","#fff"),"PIRATES":("#FDB827","#000"),
    "PADRES":("#2F241D","#FFC52F"),"GIANTS":("#FD5A1E","#000"),
    "MARINERS":("#0C2C56","#C4CED4"),"CARDINALS":("#C41E3A","#fff"),
    "RAYS":("#092C5C","#8FBCE6"),"RANGERS":("#003278","#C0111F"),
    "BLUE JAYS":("#134A8E","#E8291C"),"NATIONALS":("#AB0003","#14225A"),
}
_MLB_ABB = {
    "D-BACKS":"ARI","BRAVES":"ATL","ORIOLES":"BAL","RED SOX":"BOS",
    "CUBS":"CHC","WHITE SOX":"CWS","REDS":"CIN","GUARDIANS":"CLE",
    "ROCKIES":"COL","TIGERS":"DET","ASTROS":"HOU","ROYALS":"KC",
    "ANGELS":"LAA","DODGERS":"LAD","MARLINS":"MIA","BREWERS":"MIL",
    "TWINS":"MIN","METS":"NYM","YANKEES":"NYY","ATHLETICS":"OAK",
    "PHILLIES":"PHI","PIRATES":"PIT","PADRES":"SD","GIANTS":"SF",
    "MARINERS":"SEA","CARDINALS":"STL","RAYS":"TB","RANGERS":"TEX",
    "BLUE JAYS":"TOR","NATIONALS":"WSH",
}
_BSN_COLORS = {
    "ATLETICOS":("#006633","#fff"),"CANGREJEROS":("#CC0000","#fff"),
    "CAPITANES":("#1a2550","#FFD700"),"CRIOLLOS":("#8B0000","#fff"),
    "GIGANTES":("#FF6B00","#fff"),"INDIOS":("#4B0082","#fff"),
    "LEONES":("#DAA520","#000"),"METS":("#1565C0","#fff"),
    "OSOS":("#2E4057","#C4CED4"),"PIRATAS":("#1a1a1a","#C4A400"),
    "SANTEROS":("#7B3F00","#fff"),"VAQUEROS":("#003366","#C0C0C0"),
}
_NBA_COLORS = {
    "ATL":("#E03A3E","#fff"),"BOS":("#007A33","#fff"),"BKN":("#000","#fff"),
    "CHA":("#1D1160","#00788C"),"CHI":("#CE1141","#fff"),"CLE":("#860038","#FDBB30"),
    "DAL":("#00538C","#fff"),"DEN":("#0E2240","#FEC524"),"DET":("#C8102E","#1D42BA"),
    "GSW":("#1D428A","#FFC72C"),"HOU":("#CE1141","#fff"),"IND":("#002D62","#FDBB30"),
    "LAC":("#C8102E","#1D428A"),"LAL":("#552583","#FDB927"),"MEM":("#5D76A9","#12173F"),
    "MIA":("#98002E","#F9A01B"),"MIL":("#00471B","#EEE1C6"),"MIN":("#0C2340","#236192"),
    "NOP":("#0C2340","#85714D"),"NYK":("#006BB6","#F58426"),"OKC":("#007AC1","#EF3B24"),
    "ORL":("#0077C0","#C4CED4"),"PHI":("#006BB6","#ED174C"),"PHX":("#1D1160","#E56020"),
    "POR":("#E03A3E","#000"),"SAC":("#5A2D81","#63727A"),"SAS":("#C4CED4","#000"),
    "TOR":("#CE1141","#000"),"UTA":("#002B5C","#F9A01B"),"WAS":("#002B5C","#E31837"),
}

# ── JSON helpers ──────────────────────────────────────────────────────
def _rj(path):
    try:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f: return json.load(f)
    except: pass
    return []

def _wj(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def _parse_odds(s):
    s = str(s).strip().replace(" ","")
    try:
        v = float(re.sub(r"[^\d.\-+]","",s))
        return int(v) if v == int(v) else v
    except: return 0

def _fmt_odds(o):
    try:
        o = float(o)
        if o > 0: return f"+{int(o)}"
        return str(int(o))
    except: return str(o) if o else "—"

def _entry_stake(e):
    try: return float(e.get("stake") or e.get("units") or 1)
    except: return 1.0

def _quick_stats(log_path):
    log = _rj(log_path)
    settled = [e for e in log if e.get("result") in ("W","L","P")]
    w = sum(1 for e in settled if e.get("result")=="W")
    l = sum(1 for e in settled if e.get("result")=="L")
    pend = len([e for e in log if not e.get("result")])
    total_pnl = sum(e.get("pnl",0) or 0 for e in settled)
    total_wager = sum(_entry_stake(e) for e in settled)
    roi = (total_pnl/total_wager*100) if total_wager>0 else 0
    return w, l, pend, total_pnl, roi

# ── Run subprocess ─────────────────────────────────────────────────────
def _run(cmd, cwd=None, timeout=180):
    try:
        r = subprocess.run(
            cmd, cwd=cwd or BASE_DIR,
            capture_output=True, text=True, timeout=timeout
        )
        out = (r.stdout or "") + (r.stderr or "")
        return out.strip() or "(sin salida)"
    except subprocess.TimeoutExpired:
        return "⏱ Timeout — el comando tardó demasiado."
    except Exception as e:
        return f"Error: {e}"

# ── Background tasks (for long-running publish operations) ─────────────
_BG_TASKS  = {}   # task_id -> {"status": "running"|"done"|"error", "out": str}
_BG_LOCK   = threading.Lock()
_BG_COUNTER = [0]

# ── NBA Injury Report auto-refresh (cada 15 min) ────────────────────────
_NBA_IR_LAST_REFRESH  = [0.0]   # timestamp del último refresh
_NBA_IR_LOCK          = threading.Lock()
_NBA_IR_INTERVAL      = 15 * 60  # 15 minutos en segundos

def _nba_ir_auto_refresh():
    """
    Refresca nba_injuries.json desde el PDF oficial de la NBA cada 15 min.
    Corre en background — no bloquea el servidor.
    Solo activo entre 9 AM y 2 AM ET (horario de juegos NBA).
    """
    import time as _time
    from datetime import datetime as _dt, timedelta as _td
    while True:
        _time.sleep(_NBA_IR_INTERVAL)
        try:
            # Verificar horario ET (solo durante horas de juego)
            now_et = _dt.utcnow() - _td(hours=4)
            hour_et = now_et.hour
            # Activo de 9 AM a 2 AM del día siguiente (2 = 02:00)
            if not (9 <= hour_et or hour_et < 2):
                continue
            with _NBA_IR_LOCK:
                _NBA_IR_LAST_REFRESH[0] = _time.time()
            out = _run(["python3", "nba.py", "--ir", "refresh"],
                       cwd=NBA_DIR, timeout=60)
            ts = now_et.strftime("%H:%M ET")
            print(f"  🏥 [NBA IR auto-refresh {ts}] {out.strip()[:120]}")
        except Exception as e:
            print(f"  ⚠️  [NBA IR auto-refresh] error: {e}")

# Iniciar hilo de auto-refresh al cargar el servidor
_nba_ir_thread = threading.Thread(target=_nba_ir_auto_refresh, daemon=True)
_nba_ir_thread.start()

def _new_task_id():
    with _BG_LOCK:
        _BG_COUNTER[0] += 1
        return str(_BG_COUNTER[0])

def _bg_run_multi(task_id, commands):
    """Run [(cmd_list, label, cwd, timeout), ...] sequentially in a thread."""
    parts = []
    try:
        for cmd, label, cwd, timeout in commands:
            out = _run(cmd, cwd=cwd, timeout=timeout)
            parts.append(f"━━━ {label} ━━━\n{out}")
        combined = "\n\n".join(parts)
        with _BG_LOCK:
            _BG_TASKS[task_id] = {"status": "done", "out": combined}
    except Exception as ex:
        with _BG_LOCK:
            _BG_TASKS[task_id] = {"status": "error", "out": f"Error inesperado: {ex}"}

# ── HTML helpers ──────────────────────────────────────────────────────
def _esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

def _team_opts(teams, label="Equipo"):
    if isinstance(teams, dict):
        return "\n".join(f'<option value="{k}">{k} — {v}</option>' for k,v in sorted(teams.items()))
    return "\n".join(f'<option value="{t}">{t.title()}</option>' for t in teams)

# ── Grade helpers ──────────────────────────────────────────────────────
def _calc_pnl(odds, stake, result):
    if result == "P": return 0.0
    if result == "L": return round(-float(stake), 2)
    odds_f = float(odds)
    stake_f = float(stake)
    if odds_f > 0:
        return round(stake_f * odds_f / 100, 2)
    else:
        return round(stake_f * 100 / abs(odds_f), 2)

def _grade_rows(log_path, endpoint):
    log = _rj(log_path)
    pending = [e for e in log if not e.get("result")]
    if not pending:
        return '<p style="font-size:0.82rem;color:#64748b;text-align:center;padding:14px 0">No hay picks pendientes ✓</p>'
    rows = ""
    for e in pending:
        if e.get("type") == "parlay":
            game_label = f"Parlay ({len(e.get('legs',[]))} legs)"
            pick_label = " | ".join(lg.get("pick","") for lg in e.get("legs",[])[:2])
        else:
            game_label = e.get("game","—")
            pick_label = e.get("pick","—")
        odds_s = _fmt_odds(e.get("odds",0))
        stake_s = f"${e.get('stake',0)}"
        eid = e.get("id","")
        rows += f"""<div style="background:#0a0a0a;border-radius:8px;padding:10px;border:1px solid #1e293b;margin-bottom:8px">
  <div style="font-size:0.68rem;color:#94a3b8;margin-bottom:3px">#{eid} · {e.get('date','')} · {odds_s} · {stake_s}</div>
  <div style="font-size:0.83rem;font-weight:700;margin-bottom:8px">{_esc(game_label)}<br><span style="color:#f07820">{_esc(pick_label)}</span></div>
  <div style="display:flex;gap:7px">
    <button class="btn green" style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick('{endpoint}',{eid},'W')">✅ WIN</button>
    <button class="btn red" style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick('{endpoint}',{eid},'L')">❌ LOSS</button>
    <button class="btn gray" style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick('{endpoint}',{eid},'P')">🔄 PUSH</button>
  </div>
</div>"""
    return rows

def _grade_rows_mlb(log_path, endpoint):
    """Grade rows para MLB con campos CLV: closing_line y actual_runs."""
    log = _rj(log_path)
    pending = [e for e in log if not e.get("result")]
    if not pending:
        return '<p style="font-size:0.82rem;color:#64748b;text-align:center;padding:14px 0">No hay picks pendientes ✓</p>'
    rows = ""
    for e in pending:
        game_label = e.get("game","—")
        pick_label = e.get("pick","—")
        odds_s  = _fmt_odds(e.get("odds",0))
        stake_s = f"${e.get('stake',0)}"
        eid     = e.get("id","")
        is_total = any(w in pick_label.upper() for w in ("OVER","UNDER"))
        extra_fields = ""
        if is_total:
            extra_fields = (
                f'<div style="display:flex;gap:6px;margin-bottom:7px">'
                f'<input id="cl-{eid}" type="number" step="0.5" placeholder="Closing line ej: 9.0"'
                f' style="flex:1;background:#111827;border:1px solid #334155;border-radius:6px;'
                f'color:#e2e8f0;padding:5px 8px;font-size:0.72rem">'
                f'<input id="ar-{eid}" type="number" step="1" placeholder="Runs reales ej: 11"'
                f' style="flex:1;background:#111827;border:1px solid #334155;border-radius:6px;'
                f'color:#e2e8f0;padding:5px 8px;font-size:0.72rem">'
                f'</div>'
            )
        _cl_arg = f"'cl-{eid}'" if is_total else "null"
        rows += (
            f'<div style="background:#0a0a0a;border-radius:8px;padding:10px;border:1px solid #1e293b;margin-bottom:8px">'
            f'<div style="font-size:0.68rem;color:#94a3b8;margin-bottom:3px">#{eid} · {e.get("date","")} · {odds_s} · {stake_s}</div>'
            f'<div style="font-size:0.83rem;font-weight:700;margin-bottom:8px">{_esc(game_label)}<br>'
            f'<span style="color:#f07820">{_esc(pick_label)}</span></div>'
            f'{extra_fields}'
            f'<div style="display:flex;gap:7px">'
            f'<button class="btn green" style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick(\'{endpoint}\',{eid},\'W\',{_cl_arg})">✅ WIN</button>'
            f'<button class="btn red"   style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick(\'{endpoint}\',{eid},\'L\',{_cl_arg})">❌ LOSS</button>'
            f'<button class="btn gray"  style="flex:1;padding:7px 4px;font-size:0.75rem" onclick="gradePick(\'{endpoint}\',{eid},\'P\',{_cl_arg})">🔄 PUSH</button>'
            f'</div></div>'
        )
    return rows


def _grade_pick(log_path, data):
    id_s   = data.get("id","").strip()
    result = data.get("result","").strip().upper()
    if not id_s:
        return False, "⚠️ ID requerido."
    if result not in ("W","L","P"):
        return False, "⚠️ Resultado inválido (W/L/P)."
    try:
        pick_id = int(id_s)
        log     = _rj(log_path)
        entry   = next((e for e in log if e.get("id") == pick_id), None)
        if not entry:
            return False, f"⚠️ Pick #{pick_id} no encontrado."
        entry["result"] = result
        entry["pnl"]    = _calc_pnl(entry.get("odds",0), entry.get("stake",1), result)

        # ── CLV (Closing Line Value) ──────────────────────────────────────
        # Para O/U: CLV = closing_line - bet_line si OVER (bueno si +),
        #                 bet_line - closing_line si UNDER (bueno si +).
        # Refleja si apostaste antes de que el mercado se moviera a tu favor.
        cl_s = data.get("closing_line","").strip()
        ar_s = data.get("actual_runs","").strip()
        if cl_s:
            try:
                closing_line = float(cl_s)
                pick_str = entry.get("pick","").upper()
                m = re.match(r'(OVER|UNDER)\s+([\d.]+)', pick_str)
                if m:
                    bet_line = float(m.group(2))
                    side     = m.group(1)
                    clv      = (closing_line - bet_line) if side == "OVER" else (bet_line - closing_line)
                    entry["closing_line"] = closing_line
                    entry["clv"]          = round(clv, 2)
            except (ValueError, TypeError):
                pass

        # ── Actual runs (para predictions calibration log) ───────────────
        if ar_s:
            try:
                actual_runs = float(ar_s)
                entry["actual_runs"] = actual_runs
                # Actualizar predictions log si existe
                pred_log_path = os.path.join(MLB_DIR, "mlb_predictions_log.json")
                if os.path.exists(pred_log_path):
                    pred_log = _rj(pred_log_path)
                    game_key = entry.get("game","")
                    pick_date = entry.get("date","")
                    for row in pred_log:
                        if row.get("date") == pick_date and row.get("game") == game_key:
                            row["actual_runs"] = actual_runs
                            break
                    _wj(pred_log_path, pred_log)
            except (ValueError, TypeError):
                pass

        _wj(log_path, log)
        pnl_s = f"+${entry['pnl']:.2f}" if entry["pnl"] >= 0 else f"-${abs(entry['pnl']):.2f}"
        clv_s = f" · CLV {entry['clv']:+.1f}" if entry.get("clv") is not None else ""
        return True, f"✅ Pick #{pick_id} → {result} ({pnl_s}{clv_s})"
    except Exception as ex:
        return False, f"⚠️ Error: {ex}"

# ══════════════════════════════════════════════════════════════════════
# AUTO-GRADE — jala scores reales y gradúa picks pendientes automáticamente
# ══════════════════════════════════════════════════════════════════════

# MLB full team name → our internal key (same as _MLB_COLORS keys)
_MLB_FULLNAME_MAP = {
    "Arizona Diamondbacks":"D-BACKS","Atlanta Braves":"BRAVES",
    "Baltimore Orioles":"ORIOLES","Boston Red Sox":"RED SOX",
    "Chicago Cubs":"CUBS","Chicago White Sox":"WHITE SOX",
    "Cincinnati Reds":"REDS","Cleveland Guardians":"GUARDIANS",
    "Colorado Rockies":"ROCKIES","Detroit Tigers":"TIGERS",
    "Houston Astros":"ASTROS","Kansas City Royals":"ROYALS",
    "Los Angeles Angels":"ANGELS","Los Angeles Dodgers":"DODGERS",
    "Miami Marlins":"MARLINS","Milwaukee Brewers":"BREWERS",
    "Minnesota Twins":"TWINS","New York Mets":"METS",
    "New York Yankees":"YANKEES","Oakland Athletics":"ATHLETICS",
    "Athletics":"ATHLETICS",
    "Philadelphia Phillies":"PHILLIES","Pittsburgh Pirates":"PIRATES",
    "San Diego Padres":"PADRES","San Francisco Giants":"GIANTS",
    "Seattle Mariners":"MARINERS","St. Louis Cardinals":"CARDINALS",
    "Tampa Bay Rays":"RAYS","Texas Rangers":"RANGERS",
    "Toronto Blue Jays":"BLUE JAYS","Washington Nationals":"NATIONALS",
}

# NBA API team name → our ABB (matches NBA_TEAMS keys)
_NBA_FULLNAME_MAP = {
    "Atlanta Hawks":"ATL","Boston Celtics":"BOS","Brooklyn Nets":"BKN",
    "Charlotte Hornets":"CHA","Chicago Bulls":"CHI","Cleveland Cavaliers":"CLE",
    "Dallas Mavericks":"DAL","Denver Nuggets":"DEN","Detroit Pistons":"DET",
    "Golden State Warriors":"GSW","Houston Rockets":"HOU","Indiana Pacers":"IND",
    "Los Angeles Clippers":"LAC","Los Angeles Lakers":"LAL","Memphis Grizzlies":"MEM",
    "Miami Heat":"MIA","Milwaukee Bucks":"MIL","Minnesota Timberwolves":"MIN",
    "New Orleans Pelicans":"NOP","New York Knicks":"NYK","Oklahoma City Thunder":"OKC",
    "Orlando Magic":"ORL","Philadelphia 76ers":"PHI","Phoenix Suns":"PHX",
    "Portland Trail Blazers":"POR","Sacramento Kings":"SAC","San Antonio Spurs":"SAS",
    "Toronto Raptors":"TOR","Utah Jazz":"UTA","Washington Wizards":"WAS",
}

# Reverse NBA: ABB → nickname (for matching pick strings)
_NBA_ABB_TO_NICK = {
    "ATL":"Hawks","BOS":"Celtics","BKN":"Nets","CHA":"Hornets","CHI":"Bulls",
    "CLE":"Cavaliers","DAL":"Mavericks","DEN":"Nuggets","DET":"Pistons",
    "GSW":"Warriors","HOU":"Rockets","IND":"Pacers","LAC":"Clippers",
    "LAL":"Lakers","MEM":"Grizzlies","MIA":"Heat","MIL":"Bucks",
    "MIN":"Timberwolves","NOP":"Pelicans","NYK":"Knicks","OKC":"Thunder",
    "ORL":"Magic","PHI":"76ers","PHX":"Suns","POR":"Trail Blazers",
    "SAC":"Kings","SAS":"Spurs","TOR":"Raptors","UTA":"Jazz","WAS":"Wizards",
}

def _fetch_mlb_scores(date_str):
    """Fetch final MLB scores for a given date. Returns {game_key: (away_score, home_score, away_team, home_team)}."""
    import urllib.request as _ur
    try:
        # Remove gameType=R restriction so playoffs/non-regular games are included too
        url = (f"https://statsapi.mlb.com/api/v1/schedule?sportId=1"
               f"&date={date_str}&hydrate=linescore")
        req = _ur.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with _ur.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
        scores = {}
        for date_entry in data.get("dates", []):
            for game in date_entry.get("games", []):
                state = game.get("status", {}).get("abstractGameState", "")
                detailed = game.get("status", {}).get("detailedState", "")
                # Only accept genuinely finished games — "Final" state or detailedState contains "final"
                # Do NOT accept games just because they have scores (in-progress games also have scores)
                is_final = (state == "Final" or "final" in detailed.lower())
                if not is_final:
                    continue
                away_fn = game["teams"]["away"]["team"]["name"]
                home_fn = game["teams"]["home"]["team"]["name"]
                away_key = _MLB_FULLNAME_MAP.get(away_fn, away_fn.upper())
                home_key = _MLB_FULLNAME_MAP.get(home_fn, home_fn.upper())
                away_score_raw = game["teams"]["away"].get("score")
                home_score_raw = game["teams"]["home"].get("score")
                away_score = int(away_score_raw or 0)
                home_score = int(home_score_raw or 0)
                scores[f"{away_key} @ {home_key}"] = (away_score, home_score, away_key, home_key)
        return scores
    except Exception as ex:
        return {"_error": str(ex)}

def _fetch_nba_scores(date_str):
    """Fetch final NBA scores for a given date using NBA Stats API.
    Returns {game_key: (away_score, home_score, away_abb, home_abb)}."""
    import urllib.request as _ur
    try:
        # NBA scoreboard endpoint (public, no auth needed)
        _dt = date_str.replace("-", "")  # YYYYMMDD
        url = (f"https://stats.nba.com/stats/scoreboardv2?"
               f"DayOffset=0&LeagueID=00&gameDate={date_str[:5].replace('-','/')}{date_str[5:]}")
        # Fallback: use cdn scoreboard for dates in the past
        url2 = (f"https://cdn.nba.com/static/json/staticData/scheduleLeagueV2.json")
        # Primary: stats.nba.com scoreboard
        hdrs = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.nba.com/",
                "x-nba-stats-origin": "stats", "x-nba-stats-token": "true"}
        req = _ur.Request(
            f"https://stats.nba.com/stats/scoreboardv2?DayOffset=0&LeagueID=00&gameDate={date_str}",
            headers=hdrs)
        with _ur.urlopen(req, timeout=12) as resp:
            data = json.loads(resp.read())
        scores = {}
        # Parse resultSets[1] = LineScore
        result_sets = {rs["name"]: rs for rs in data.get("resultSets", [])}
        line_score  = result_sets.get("LineScore", {})
        headers_ls  = line_score.get("headers", [])
        rows_ls     = line_score.get("rowSet", [])
        # Find column indices
        def _col(name): return headers_ls.index(name) if name in headers_ls else None
        _i_game  = _col("GAME_ID")
        _i_team  = _col("TEAM_ABBREVIATION")
        _i_fn    = _col("TEAM_CITY_NAME")
        _i_pts   = _col("PTS")
        _i_status= _col("GAME_STATUS_TEXT")
        # Group rows by GAME_ID
        games_data = {}
        for row in rows_ls:
            if _i_game is None or _i_pts is None: continue
            gid  = row[_i_game]
            abb  = row[_i_team] if _i_team is not None else ""
            pts  = row[_i_pts]
            if pts is None: continue
            if gid not in games_data:
                games_data[gid] = []
            games_data[gid].append({"abb": abb.upper(), "pts": int(pts)})
        # Build score dict
        gs_set = result_sets.get("GameHeader", {})
        gh_hdrs = gs_set.get("headers", [])
        gh_rows = gs_set.get("rowSet", [])
        _gi_gid    = gh_hdrs.index("GAME_ID")    if "GAME_ID"    in gh_hdrs else None
        _gi_status = gh_hdrs.index("GAME_STATUS_ID") if "GAME_STATUS_ID" in gh_hdrs else None
        final_gids = set()
        for gh_row in gh_rows:
            if _gi_gid and _gi_status:
                gid_  = gh_row[_gi_gid]
                stat_ = gh_row[_gi_status]
                if stat_ == 3:  # 3 = Final
                    final_gids.add(gid_)
        for gid, teams in games_data.items():
            if gid not in final_gids: continue
            if len(teams) < 2: continue
            # teams[0]=away, teams[1]=home (NBA convention)
            away_abb, away_pts = teams[0]["abb"], teams[0]["pts"]
            home_abb, home_pts = teams[1]["abb"], teams[1]["pts"]
            key = f"{away_abb} @ {home_abb}"
            scores[key] = (away_pts, home_pts, away_abb, home_abb)
        return scores
    except Exception as ex:
        return {"_error": str(ex)}

def _determine_mlb_result(pick_str, game_key, away_score, home_score):
    """Determina W/L/P para un pick de MLB basado en scores finales."""
    pick_upper = pick_str.upper().strip()
    away_key, home_key = (game_key.split(" @ ") + ["",""])[:2]
    total = away_score + home_score

    # Over / Under
    m = re.match(r'(OVER|UNDER|O|U)\s+([\d.]+)', pick_upper)
    if m:
        side = "OVER" if m.group(1) in ("OVER","O") else "UNDER"
        line = float(m.group(2))
        if total > line:   return "W" if side == "OVER" else "L"
        elif total < line: return "L" if side == "OVER" else "W"
        else:              return "P"

    # ML
    if "ML" in pick_upper:
        team = pick_upper.replace("ML","").strip()
        if team in (away_key, away_key[:3]):
            if away_score > home_score: return "W"
            elif away_score < home_score: return "L"
            else: return "P"
        elif team in (home_key, home_key[:3]):
            if home_score > away_score: return "W"
            elif home_score < away_score: return "L"
            else: return "P"

    # Run Line — skip auto (harder to detect which side)
    return None

def _determine_nba_result(pick_str, game_key, away_score, home_score):
    """Determina W/L/P para un pick de NBA basado en scores finales."""
    pick_upper = pick_str.upper().strip()
    away_abb, home_abb = (game_key.split(" @ ") + ["",""])[:2]
    total = away_score + home_score

    # Over / Under
    m = re.match(r'(OVER|UNDER|O|U)\s+([\d.]+)', pick_upper)
    if m:
        side = "OVER" if m.group(1) in ("OVER","O") else "UNDER"
        line = float(m.group(2))
        if total > line:   return "W" if side == "OVER" else "L"
        elif total < line: return "L" if side == "OVER" else "W"
        else:              return "P"

    # ML — match by nickname or ABB
    if "ML" in pick_upper:
        team_part = pick_upper.replace("ML","").replace("[MKT~EST]","").strip()
        away_nick = _NBA_ABB_TO_NICK.get(away_abb,"").upper()
        home_nick = _NBA_ABB_TO_NICK.get(home_abb,"").upper()
        is_away = (team_part in (away_abb, away_nick) or
                   any(w in team_part for w in away_nick.split()))
        is_home = (team_part in (home_abb, home_nick) or
                   any(w in team_part for w in home_nick.split()))
        if is_away:
            if away_score > home_score: return "W"
            elif away_score < home_score: return "L"
            else: return "P"
        if is_home:
            if home_score > away_score: return "W"
            elif home_score < away_score: return "L"
            else: return "P"

    # Spread — skip auto
    return None

def _auto_grade_mlb():
    """Jala scores reales de statsapi y gradúa picks MLB pendientes."""
    log     = _rj(MLB_LOG)
    pending = [e for e in log if not e.get("result")]
    if not pending:
        return True, "No hay picks MLB pendientes."

    dates  = sorted(set(e.get("date","") for e in pending if e.get("date")))
    graded = 0
    skipped = []

    for date_str in dates:
        scores = _fetch_mlb_scores(date_str)
        if "_error" in scores:
            skipped.append(f"{date_str}: {scores['_error']}")
            continue
        for e in pending:
            if e.get("result") or e.get("date") != date_str:
                continue
            game_str = e.get("game","").upper().strip()
            # Normalize "RAYS vs. BLUE JAYS" / "RAYS VS BLUE JAYS" → extract both teams
            import re as _re
            _sep = _re.split(r'\s+(?:VS\.?|@)\s+', game_str, flags=_re.IGNORECASE)
            _teams = [t.strip() for t in _sep if t.strip()]
            # Try exact key match first
            score_entry = scores.get(game_str)
            if not score_entry:
                # Try both orderings: away@home and home@away
                for k, v in scores.items():
                    k_teams = [t.strip() for t in _re.split(r'\s+(?:VS\.?|@)\s+', k)]
                    if len(_teams) >= 2 and len(k_teams) >= 2:
                        if (_teams[0] in k_teams[0] or k_teams[0] in _teams[0]) and \
                           (_teams[1] in k_teams[1] or k_teams[1] in _teams[1]):
                            score_entry = v; break
                        if (_teams[0] in k_teams[1] or k_teams[1] in _teams[0]) and \
                           (_teams[1] in k_teams[0] or k_teams[0] in _teams[1]):
                            score_entry = v; break
                    elif len(_teams) == 1:
                        # Fallback: single token, check if it appears in any key
                        if _teams[0] in k:
                            score_entry = v; break
            if not score_entry:
                skipped.append(f"#{e['id']} {game_str} (no score found)")
                continue
            away_score, home_score, away_key_norm, home_key_norm = score_entry
            # Use normalized "AWAY @ HOME" key so _determine_mlb_result can split on " @ " correctly
            game_key_norm = f"{away_key_norm} @ {home_key_norm}"
            result = _determine_mlb_result(e.get("pick",""), game_key_norm, away_score, home_score)
            if not result:
                skipped.append(f"#{e['id']} pick type not auto-gradable")
                continue
            e["result"] = result
            e["pnl"]    = _calc_pnl(e.get("odds",0), e.get("stake",1), result)
            # Also update actual_runs in predictions log
            pred_log_path = os.path.join(MLB_DIR, "mlb_predictions_log.json")
            if os.path.exists(pred_log_path):
                pred_log = _rj(pred_log_path)
                total = away_score + home_score
                for row in pred_log:
                    if row.get("date") == date_str and row.get("game","").upper() == game_str:
                        row["actual_runs"] = total
                        break
                _wj(pred_log_path, pred_log)
            graded += 1

    _wj(MLB_LOG, log)
    msg = f"✅ {graded} pick(s) MLB gradeados."
    if skipped:
        msg += f" ⚠️ {len(skipped)} omitidos: {'; '.join(skipped[:3])}"
    return True, msg

def _auto_grade_nba():
    """Jala scores NBA y gradúa picks pendientes."""
    log     = _rj(NBA_LOG)
    pending = [e for e in log if not e.get("result")]
    if not pending:
        return True, "No hay picks NBA pendientes."

    dates  = sorted(set(e.get("date","") for e in pending if e.get("date")))
    graded = 0
    skipped = []

    for date_str in dates:
        scores = _fetch_nba_scores(date_str)
        if "_error" in scores:
            skipped.append(f"{date_str}: {scores['_error']}")
            continue
        for e in pending:
            if e.get("result") or e.get("date") != date_str:
                continue
            game_str = e.get("game","").upper().strip()
            # Strip series suffix: "Hawks @ Celtics [0-2]" → "Hawks @ Celtics"
            clean_game = re.sub(r'\s*\[.*?\]', '', game_str).strip()
            # Split on either "@" or "vs." so both formats work
            parts = re.split(r'\s+(?:VS\.?|@)\s+', clean_game, flags=re.IGNORECASE)
            # Map nicknames to ABBs for matching
            def _nick_to_abb(nick):
                nick = nick.strip().upper()
                for abb, n in _NBA_ABB_TO_NICK.items():
                    if n.upper() == nick or abb == nick:
                        return abb
                return nick
            if len(parts) == 2:
                away_abb = _nick_to_abb(parts[0])
                home_abb = _nick_to_abb(parts[1])
                key = f"{away_abb} @ {home_abb}"
            else:
                away_abb = home_abb = ""
                key = clean_game

            score_entry = scores.get(key)
            if not score_entry:
                for k, v in scores.items():
                    a, h = (k.split(" @ ") + ["",""])[:2]
                    if away_abb and home_abb and away_abb in (a,) and home_abb in (h,):
                        score_entry = v; break
                    # Also try reversed (home vs away ordering)
                    if away_abb and home_abb and away_abb in (h,) and home_abb in (a,):
                        score_entry = v; break
            if not score_entry:
                skipped.append(f"#{e['id']} {clean_game} (no score)")
                continue
            away_score, home_score, _, _ = score_entry
            result = _determine_nba_result(e.get("pick",""), key, away_score, home_score)
            if not result:
                skipped.append(f"#{e['id']} tipo no auto-gradable")
                continue
            e["result"] = result
            e["pnl"]    = _calc_pnl(e.get("odds",0), e.get("stake",1), result)
            graded += 1

    _wj(NBA_LOG, log)
    msg = f"✅ {graded} pick(s) NBA gradeados."
    if skipped:
        msg += f" ⚠️ {len(skipped)} omitidos: {'; '.join(skipped[:3])}"
    return True, msg


# ── MLB today's picks helper ────────────────────────────────────────────
def _bsn_today_games():
    """Return today's manual BSN games from manual_games.json."""
    path = os.path.join(BSN_DIR, "manual_games.json")
    today_str = date.today().strftime("%Y-%m-%d")
    entries = _rj(path)
    if not isinstance(entries, list):
        return []
    return [e for e in entries if e.get("date") == today_str]


def _mlb_today_games():
    """Return list of (away, home, display_label) for today's MLB model picks."""
    today_str = date.today().strftime("%Y-%m-%d")
    picks = _rj(MLB_PICKS)
    seen = set()
    games = []
    for p in picks:
        if p.get("date","") != today_str:
            continue
        game = p.get("game","")  # e.g. "WHITE SOX @ ROYALS"
        if game in seen:
            continue
        seen.add(game)
        if " @ " in game:
            away, home = [t.strip() for t in game.split(" @ ", 1)]
        else:
            continue
        games.append((away, home, game))
    return games


# ══════════════════════════════════════════════════════════════════════
# HTML PAGE
# ══════════════════════════════════════════════════════════════════════
CSS = """
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}

body{
  background:#060a0f;
  background-image:radial-gradient(rgba(255,255,255,.018) 1px,transparent 1px);
  background-size:24px 24px;background-color:#060a0f;
  font-family:'Inter',system-ui,sans-serif;
  color:#e2e8f0;min-height:100vh;
}

/* ── Site Header ── */
.site-hdr{
  position:sticky;top:0;z-index:100;
  background:rgba(6,10,15,.93);
  backdrop-filter:blur(18px);-webkit-backdrop-filter:blur(18px);
  border-bottom:1px solid rgba(255,255,255,.07);
}
.hdr-inner{
  max-width:860px;margin:0 auto;padding:13px 20px;
  display:flex;align-items:center;justify-content:space-between;
}
.brand{display:flex;align-items:center;gap:9px}
.brand-dot{
  width:8px;height:8px;border-radius:50%;
  background:#f07820;box-shadow:0 0 14px rgba(240,120,32,.65);
}
.brand-name{font-size:.8rem;font-weight:900;letter-spacing:.2em;text-transform:uppercase}
.brand-name em{color:#f07820;font-style:normal}
.hdr-right{display:flex;align-items:center;gap:12px}
.live-badge{
  display:flex;align-items:center;gap:5px;
  background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.2);
  border-radius:99px;padding:3px 10px;
  font-size:.57rem;font-weight:800;letter-spacing:.1em;color:#22c55e;
}
.live-dot{
  width:5px;height:5px;border-radius:50%;
  background:#22c55e;box-shadow:0 0 7px rgba(34,197,94,.9);
  animation:blink 2s ease-in-out infinite;
}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.15}}
.hdr-date{font-size:.62rem;color:#475569;font-weight:500;letter-spacing:.04em}

/* ── League Nav ── */
.lg-nav{
  position:sticky;top:50px;z-index:99;
  background:rgba(6,10,15,.9);
  backdrop-filter:blur(12px);-webkit-backdrop-filter:blur(12px);
  border-bottom:1px solid rgba(255,255,255,.07);
}
.lg-nav-inner{max-width:860px;margin:0 auto;display:flex}
.tab{
  flex:1;padding:12px 8px;
  display:flex;align-items:center;justify-content:center;gap:7px;
  cursor:pointer;border-bottom:2px solid transparent;transition:all .2s;
}
.tab:hover{background:rgba(255,255,255,.03)}
.tab.active{border-bottom-color:#f07820}
.tab-icon{font-size:.9rem}
.tab-name{
  font-size:.68rem;font-weight:800;letter-spacing:.14em;text-transform:uppercase;
  color:#475569;transition:color .2s;
}
.tab.active .tab-name{color:#f07820}
.tab-rec{
  font-size:.6rem;font-weight:700;
  background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);
  border-radius:99px;padding:2px 8px;color:#334155;transition:all .2s;
}
.tab.active .tab-rec{
  background:rgba(240,120,32,.1);border-color:rgba(240,120,32,.25);color:#f07820;
}

/* ── Content ── */
.site-main{max-width:860px;margin:0 auto;padding:22px 20px 60px}
.panel{display:none}
.panel.active{display:block}

/* ── League Hero Stats ── */
.lg-hero{
  background:rgba(255,255,255,.03);
  border:1px solid rgba(255,255,255,.07);
  border-radius:16px;padding:20px 22px;
  margin-bottom:18px;position:relative;overflow:hidden;
}
.lg-hero::before{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--acc,#f07820),transparent);
}
.lg-hero::after{
  content:'';position:absolute;top:-70px;left:-30px;
  width:220px;height:220px;border-radius:50%;
  background:radial-gradient(circle,var(--glow,rgba(240,120,32,.1)) 0%,transparent 70%);
  pointer-events:none;
}
.hero-top{
  display:flex;align-items:center;justify-content:space-between;
  margin-bottom:14px;position:relative;z-index:1;
}
.hero-league-lbl{
  font-size:.57rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase;
  color:var(--acc,#f07820);
}
.hero-pending{
  font-size:.57rem;font-weight:700;
  color:#f59e0b;background:rgba(245,158,11,.1);
  border:1px solid rgba(245,158,11,.2);
  border-radius:99px;padding:2px 9px;
}
.hero-rec{
  display:flex;align-items:baseline;gap:2px;
  margin-bottom:16px;position:relative;z-index:1;
}
.r-big{font-size:3rem;font-weight:900;line-height:1}
.r-w{color:#22c55e;text-shadow:0 0 32px rgba(34,197,94,.3)}
.r-l{color:#ef4444;text-shadow:0 0 32px rgba(239,68,68,.25)}
.r-sep{font-size:2rem;font-weight:300;color:#1e293b;margin:0 6px;align-self:center}
.r-push{font-size:1.2rem;font-weight:700;color:#475569;margin-left:10px;align-self:flex-end;margin-bottom:3px}
.stats-4{
  display:grid;grid-template-columns:repeat(4,1fr);gap:7px;
  position:relative;z-index:1;
}
@media(max-width:500px){.stats-4{grid-template-columns:repeat(2,1fr)}}
.s4-cell{
  background:rgba(0,0,0,.3);border:1px solid rgba(255,255,255,.06);
  border-radius:10px;padding:10px 8px;text-align:center;
}
.s4-val{font-size:.95rem;font-weight:800;margin-bottom:3px}
.s4-lbl{font-size:.5rem;font-weight:700;letter-spacing:.11em;text-transform:uppercase;color:#475569}

/* ── Section labels & dividers ── */
.sec-lbl{
  font-size:.56rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase;
  color:#334155;padding:6px 0 8px;margin-top:4px;
}
.sec-div{height:1px;background:rgba(255,255,255,.06);margin:12px 0 14px}

/* ── Command cards ── */
.cmds{display:grid;gap:7px}
.cmd-card{
  background:rgba(255,255,255,.03);
  border:1px solid rgba(255,255,255,.07);
  border-radius:12px;padding:13px 16px;
  display:flex;align-items:center;justify-content:space-between;gap:14px;
  transition:background .15s,border-color .15s;
}
.cmd-card:hover{background:rgba(255,255,255,.055);border-color:rgba(255,255,255,.12)}
.cmd-desc{font-size:.75rem;color:#64748b;flex:1;min-width:0;line-height:1.4}

/* ── Buttons ── */
.btn{
  background:#f07820;color:#fff;border:none;border-radius:8px;
  padding:9px 18px;font-size:.76rem;font-weight:700;
  cursor:pointer;white-space:nowrap;flex-shrink:0;
  font-family:inherit;letter-spacing:.03em;
  transition:opacity .15s,box-shadow .15s;
}
.btn:hover{box-shadow:0 0 16px rgba(240,120,32,.35)}
.btn:active{opacity:.82}
.btn.green{background:#16a34a}.btn.green:hover{box-shadow:0 0 16px rgba(22,163,74,.35)}
.btn.blue{background:#2563eb}.btn.blue:hover{box-shadow:0 0 16px rgba(37,99,235,.35)}
.btn.red{background:#dc2626}.btn.red:hover{box-shadow:0 0 12px rgba(220,38,38,.3)}
.btn.gray{background:#0f172a;color:#64748b;border:1px solid rgba(255,255,255,.07)}
.btn.gray:hover{box-shadow:none;background:#1e293b}

/* ── Modal ── */
.modal-bg{
  display:none;position:fixed;inset:0;
  background:rgba(0,0,0,.72);
  backdrop-filter:blur(6px);-webkit-backdrop-filter:blur(6px);
  z-index:200;overflow-y:auto;padding:20px 16px;
}
.modal-bg.open{display:flex;align-items:flex-start;justify-content:center}
.modal{
  background:#080d12;border:1px solid rgba(255,255,255,.09);
  border-radius:16px;width:100%;max-width:520px;padding:22px;margin:auto 0;
}
.modal h2{font-size:.88rem;font-weight:800;margin-bottom:18px;color:#f07820;letter-spacing:.04em}
label{
  display:block;font-size:.6rem;color:#475569;font-weight:700;
  letter-spacing:.1em;text-transform:uppercase;margin-bottom:5px;
}
input,select,textarea{
  width:100%;background:rgba(0,0,0,.4);color:#e2e8f0;
  border:1px solid rgba(255,255,255,.08);
  border-radius:8px;padding:11px 13px;font-size:.92rem;
  margin-bottom:12px;-webkit-appearance:none;font-family:inherit;
  transition:border-color .2s;
}
input:focus,select:focus,textarea:focus{outline:none;border-color:#f07820}
textarea{height:72px;resize:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px}
.btn-row{display:flex;gap:10px;margin-top:4px}
.btn-row .btn{flex:1}
.alert{padding:11px 14px;border-radius:8px;margin-bottom:12px;font-weight:600;font-size:.85rem}
.alert.ok{background:#14532d;color:#86efac}
.alert.err{background:#7f1d1d;color:#fca5a5}

/* ── Output panel ── */
.out-bg{
  display:none;position:fixed;inset:0;
  background:rgba(0,0,0,.75);
  backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px);
  z-index:300;align-items:flex-end;
}
.out-bg.open{display:flex}
.out-panel{
  background:#060a0f;border-top:2px solid #f07820;
  border-radius:16px 16px 0 0;
  width:100%;max-height:75vh;display:flex;flex-direction:column;
}
.out-header{
  padding:13px 20px;
  display:flex;align-items:center;justify-content:space-between;
  border-bottom:1px solid rgba(255,255,255,.07);flex-shrink:0;
}
.out-hdr-left{display:flex;align-items:center;gap:8px}
.out-dot{
  width:6px;height:6px;border-radius:50%;
  background:#f07820;box-shadow:0 0 8px rgba(240,120,32,.8);
  animation:blink 1.4s ease-in-out infinite;
}
.out-title-lbl{font-size:.75rem;font-weight:700;color:#f07820;letter-spacing:.06em}
.out-body{
  overflow-y:auto;padding:14px 20px;
  font-family:'SF Mono','Fira Code',Menlo,monospace;
  font-size:.72rem;color:#a3e635;
  white-space:pre-wrap;flex:1;line-height:1.7;
}
.spinner{
  display:inline-block;width:12px;height:12px;
  border:2px solid #f07820;border-top-color:transparent;
  border-radius:50%;animation:spin .7s linear infinite;
  margin-right:8px;vertical-align:middle;
}
@keyframes spin{to{transform:rotate(360deg)}}

/* ── Leg builder ── */
.leg{background:rgba(0,0,0,.3);border-radius:8px;padding:12px;border:1px solid rgba(255,255,255,.07);margin-bottom:10px}
.leg-title{font-size:.65rem;font-weight:700;color:#f07820;margin-bottom:8px;letter-spacing:.1em}
.section-sep{height:1px;background:rgba(255,255,255,.07);margin:14px 0}

/* ── Detail Panel overlay ── */
.detail-overlay{
  position:fixed;inset:0;background:rgba(0,0,0,.6);
  backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px);
  z-index:150;display:none;
}
.detail-overlay.open{display:block}

/* ── Pick cards ── */
.pick-card{
  background:rgba(255,255,255,.035);border:1px solid rgba(255,255,255,.08);
  border-radius:12px;padding:15px 16px;margin-bottom:10px;
  position:relative;overflow:hidden;
}
.pick-card::before{
  content:'';position:absolute;left:0;top:0;bottom:0;width:3px;
  background:var(--acc,#f07820);
}
.pc-game{font-size:.63rem;font-weight:700;color:#475569;letter-spacing:.1em;text-transform:uppercase;margin-bottom:6px}
.pc-pick{font-size:1rem;font-weight:800;color:#e2e8f0;margin-bottom:8px}
.pc-row{display:flex;align-items:center;gap:7px;flex-wrap:wrap}
.pc-odds{
  font-size:.7rem;font-weight:700;
  background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.1);
  border-radius:6px;padding:3px 8px;color:#94a3b8;
}
.pc-ev{
  font-size:.68rem;font-weight:700;color:#22c55e;
  background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.22);
  border-radius:6px;padding:3px 8px;
}
.pc-tag{
  font-size:.64rem;font-weight:700;color:#f59e0b;
  background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.22);
  border-radius:6px;padding:3px 8px;
}
.pc-note{
  font-size:.7rem;color:#64748b;margin-top:9px;line-height:1.5;
  padding-top:9px;border-top:1px solid rgba(255,255,255,.05);
}

/* ── Log history entries ── */
.log-entry{
  background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.06);
  border-radius:10px;padding:12px 14px;margin-bottom:8px;
  display:flex;align-items:flex-start;gap:11px;
}
.le-result{
  width:28px;height:28px;border-radius:50%;flex-shrink:0;
  display:flex;align-items:center;justify-content:center;
  font-size:.62rem;font-weight:900;
}
.le-w{background:rgba(34,197,94,.12);color:#22c55e;border:1px solid rgba(34,197,94,.28)}
.le-l{background:rgba(239,68,68,.12);color:#ef4444;border:1px solid rgba(239,68,68,.28)}
.le-p{background:rgba(100,116,139,.12);color:#94a3b8;border:1px solid rgba(100,116,139,.28)}
.le-n{background:rgba(245,158,11,.08);color:#f59e0b;border:1px solid rgba(245,158,11,.22)}
.le-body{flex:1;min-width:0}
.le-game{font-size:.65rem;color:#475569;margin-bottom:3px;letter-spacing:.03em}
.le-pick{font-size:.82rem;font-weight:700;color:#e2e8f0}
.le-meta{font-size:.64rem;color:#334155;margin-top:4px}
.le-pos{color:#22c55e;font-weight:700}
.le-neg{color:#ef4444;font-weight:700}

/* ── Terminal in detail panel ── */
.terminal{
  font-family:'SF Mono','Fira Code',Menlo,monospace;
  font-size:.7rem;color:#a3e635;line-height:1.75;white-space:pre-wrap;
  background:rgba(0,0,0,.35);border:1px solid rgba(255,255,255,.06);
  border-radius:10px;padding:14px 16px;
}
.t-ok{color:#22c55e}.t-err{color:#ef4444}
.t-warn{color:#f59e0b}.t-hi{color:#f07820}.t-dim{color:#475569}

/* ── Visual command output (.vw-*) ── */
.vw-output{display:flex;flex-direction:column;gap:2px}
.vw-hdr{
  font-size:.55rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase;
  color:#f07820;padding:16px 0 7px;
  border-bottom:1px solid rgba(255,255,255,.06);margin-bottom:4px;margin-top:4px;
}
.vw-hdr:first-child{padding-top:4px}
.vw-table{display:flex;flex-direction:column;gap:3px;margin-bottom:4px}
.vt-row{
  display:flex;align-items:center;justify-content:space-between;
  background:rgba(255,255,255,.025);border-radius:8px;
  padding:8px 12px;gap:8px;transition:background .15s;
}
.vt-row:hover{background:rgba(255,255,255,.04)}
.vt-team{font-size:.74rem;font-weight:800;color:#e2e8f0;min-width:72px;flex-shrink:0}
.vt-vals{display:flex;gap:9px;align-items:center;flex-wrap:wrap}
.vt-cell{font-size:.72rem;font-weight:700}
.vt-pos{color:#22c55e}.vt-neg{color:#ef4444}.vt-neu{color:#94a3b8}
.vt-label{color:#475569;font-size:.68rem;font-weight:500}
.vw-pick{
  background:rgba(255,255,255,.03);border-radius:10px;padding:12px 14px;
  border-left:3px solid #f07820;
  display:flex;align-items:center;justify-content:space-between;gap:10px;
  margin-bottom:4px;transition:background .15s;
}
.vw-pick:hover{background:rgba(255,255,255,.045)}
.vw-pick-star{border-left-color:#f59e0b;background:rgba(245,158,11,.05)}
.vw-star{color:#f59e0b;font-size:.9rem;margin-right:5px;flex-shrink:0}
.vw-pick-text{font-size:.82rem;font-weight:700;color:#e2e8f0;flex:1;min-width:0}
.vw-pick-badges{display:flex;gap:6px;align-items:center;flex-shrink:0}
.vw-odds{
  font-size:.68rem;font-weight:700;
  background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.1);
  border-radius:6px;padding:3px 8px;color:#94a3b8;
}
.vw-ev{
  font-size:.66rem;font-weight:800;color:#22c55e;
  background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.22);
  border-radius:6px;padding:3px 8px;
}
.vw-stat-line{
  font-size:.76rem;font-weight:600;color:#94a3b8;
  padding:4px 2px;border-bottom:1px solid rgba(255,255,255,.04);
}
.vw-text{font-size:.74rem;color:#64748b;padding:2px 2px;line-height:1.55}
.vw-section-gap{height:8px}

/* ══ WEBSITE LAYOUT ══════════════════════════════════════════════════ */

/* ── Section header row ── */
.section-hdr{display:flex;align-items:center;justify-content:space-between;margin:18px 0 10px}
.section-title{font-size:.6rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase;color:#e2e8f0}
.section-btn{
  font-size:.6rem;font-weight:700;color:#f07820;cursor:pointer;
  background:rgba(240,120,32,.08);border:1px solid rgba(240,120,32,.2);
  border-radius:7px;padding:4px 11px;font-family:inherit;letter-spacing:.04em;
  transition:background .15s;
}
.section-btn:hover{background:rgba(240,120,32,.16)}

/* ── Today empty state ── */
.today-empty{
  text-align:center;padding:24px 16px;
  background:rgba(255,255,255,.02);border:1px dashed rgba(255,255,255,.07);
  border-radius:12px;color:#334155;font-size:.76rem;line-height:1.8;
}

/* ── Team badge ── */
.team-badge{
  display:inline-flex;align-items:center;justify-content:center;
  border-radius:8px;font-weight:900;font-size:.65rem;
  white-space:nowrap;flex-shrink:0;padding:5px 10px;
  letter-spacing:.05em;line-height:1;font-family:inherit;
}
.matchup-vs{font-size:.56rem;color:#334155;font-weight:700;letter-spacing:.08em}
.game-time{font-size:.6rem;color:#475569;margin-left:auto;white-space:nowrap;font-weight:600}

/* ── Game card ── */
.game-card{
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:14px;padding:14px 16px;margin-bottom:8px;transition:border-color .15s;
}
.game-card:hover{border-color:rgba(255,255,255,.13)}
.game-matchup{display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap}
.game-picks{display:flex;flex-direction:column;gap:5px}
.gpick{
  display:flex;align-items:center;gap:8px;padding:8px 11px;
  background:rgba(255,255,255,.025);border-radius:8px;
  border-left:2px solid var(--acc,#f07820);transition:background .12s;
}
.gpick:hover{background:rgba(255,255,255,.04)}
.gpick.star{border-left-color:#f59e0b;background:rgba(245,158,11,.04)}
.gpick-ico{font-size:.7rem;flex-shrink:0;color:#f07820;min-width:10px;text-align:center}
.gpick.star .gpick-ico{color:#f59e0b}
.gpick-pick{font-size:.8rem;font-weight:700;color:#e2e8f0;flex:1;min-width:0}
.gpick-odds{
  font-size:.67rem;font-weight:700;background:rgba(255,255,255,.07);
  border:1px solid rgba(255,255,255,.1);border-radius:5px;
  padding:2px 7px;color:#94a3b8;flex-shrink:0;
}
.gpick-ev{
  font-size:.62rem;font-weight:800;color:#22c55e;
  background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.22);
  border-radius:5px;padding:2px 7px;flex-shrink:0;
}

/* ── IR / Schedule mini rows ── */
.mini-table{display:flex;flex-direction:column;gap:4px}
.mt-row{
  display:flex;align-items:center;gap:10px;
  background:rgba(255,255,255,.025);border-radius:9px;padding:9px 12px;
}
.mt-name{font-size:.78rem;font-weight:700;color:#e2e8f0;flex:1;min-width:0}
.mt-sub{font-size:.63rem;color:#475569;flex-shrink:0}
.mt-status{font-size:.6rem;font-weight:800;border-radius:5px;padding:2px 8px;flex-shrink:0}
.mt-out{background:rgba(239,68,68,.15);color:#ef4444;border:1px solid rgba(239,68,68,.22)}
.mt-dbt{background:rgba(249,115,22,.12);color:#f97316;border:1px solid rgba(249,115,22,.2)}
.mt-qst{background:rgba(234,179,8,.1);color:#eab308;border:1px solid rgba(234,179,8,.18)}

/* ── Action icon row ── */
.action-row{display:flex;gap:7px;flex-wrap:wrap;margin:18px 0 4px}
.act-btn{
  display:flex;flex-direction:column;align-items:center;gap:4px;
  background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);
  border-radius:11px;padding:11px 14px;cursor:pointer;font-family:inherit;
  color:#e2e8f0;transition:all .15s;min-width:52px;flex-shrink:0;
}
.act-btn:hover{background:rgba(255,255,255,.08);border-color:rgba(255,255,255,.15);transform:translateY(-1px)}
.act-btn:active{transform:translateY(0)}
.act-btn.accent{background:rgba(240,120,32,.1);border-color:rgba(240,120,32,.22)}
.act-btn.accent:hover{background:rgba(240,120,32,.18)}
.act-btn.green{background:rgba(34,197,94,.08);border-color:rgba(34,197,94,.2)}
.act-btn.green:hover{background:rgba(34,197,94,.16)}
.act-btn.blue{background:rgba(37,99,235,.08);border-color:rgba(37,99,235,.2)}
.act-btn.blue:hover{background:rgba(37,99,235,.16)}
.act-icon{font-size:1.05rem;line-height:1}
.act-lbl{font-size:.48rem;font-weight:800;letter-spacing:.1em;text-transform:uppercase;color:#475569}

/* ── Tools accordion ── */
.tools-group{
  background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.06);
  border-radius:12px;overflow:hidden;margin-top:14px;
}
.tools-toggle{
  padding:13px 16px;display:flex;align-items:center;justify-content:space-between;
  cursor:pointer;font-size:.6rem;font-weight:700;color:#475569;
  letter-spacing:.12em;text-transform:uppercase;user-select:none;transition:background .15s;
}
.tools-toggle:hover{background:rgba(255,255,255,.03)}
.tools-chevron{transition:transform .22s;font-style:normal;font-size:.75rem;display:inline-block}
.tools-chevron.open{transform:rotate(90deg)}
.tools-body{display:none;padding:6px 10px 12px;border-top:1px solid rgba(255,255,255,.06)}
.tools-body.open{display:block}

/* ══ SIDE PANEL UPGRADE ══════════════════════════════════════════════ */
.detail-panel{
  position:fixed;top:0;right:0;
  width:min(600px,100vw);height:100vh;
  background:linear-gradient(160deg,#0b1018 0%,#060a0f 100%);
  border-left:1px solid rgba(255,255,255,.08);
  z-index:151;display:flex;flex-direction:column;
  transform:translateX(100%);
  transition:transform .34s cubic-bezier(.4,0,.2,1);
  box-shadow:-40px 0 100px rgba(0,0,0,.75);
}
.detail-panel.open{transform:translateX(0)}
.detail-hdr{
  padding:20px 24px;
  display:flex;align-items:center;justify-content:space-between;
  border-bottom:1px solid rgba(255,255,255,.07);flex-shrink:0;
  background:rgba(0,0,0,.2);
}
.detail-title{
  font-size:.9rem;font-weight:900;color:#fff;
  letter-spacing:.12em;text-transform:uppercase;
}
.detail-body{overflow-y:auto;padding:20px 24px;flex:1}

/* Side panel section headers */
.vw-hdr{
  display:flex;align-items:center;gap:9px;
  font-size:.6rem;font-weight:900;letter-spacing:.2em;text-transform:uppercase;
  color:#f07820;padding:18px 0 9px;
  border-bottom:1px solid rgba(240,120,32,.12);
  margin-bottom:8px;margin-top:6px;
}
.vw-hdr::before{
  content:'';width:3px;height:14px;
  background:linear-gradient(180deg,#f07820,rgba(240,120,32,.3));
  border-radius:2px;flex-shrink:0;
}
.vw-hdr:first-child{padding-top:2px}
.vw-text{font-size:.76rem;color:#64748b;padding:3px 2px;line-height:1.6}
.vw-stat-line{
  font-size:.74rem;font-weight:600;color:#64748b;
  padding:5px 4px;border-bottom:1px solid rgba(255,255,255,.04);line-height:1.5;
}

/* Pipe-delimited stats table */
.vpt{display:flex;flex-direction:column;gap:3px;margin-bottom:8px}
.vpt-hdr{
  display:flex;align-items:center;gap:6px;
  padding:6px 12px;margin-bottom:2px;
  background:rgba(255,255,255,.02);border-radius:7px;
}
.vpt-th{
  font-size:.5rem;font-weight:900;letter-spacing:.14em;text-transform:uppercase;
  color:#334155;flex:1;text-align:center;
}
.vpt-th:first-child{flex:2;text-align:left}
.vpt-row{
  display:flex;align-items:center;gap:8px;
  padding:9px 12px;
  background:rgba(255,255,255,.028);
  border:1px solid rgba(255,255,255,.05);
  border-radius:10px;transition:all .12s;
  position:relative;
}
.vpt-row:hover{background:rgba(255,255,255,.05);border-color:rgba(255,255,255,.1)}
.vpt-rank{
  font-size:.55rem;font-weight:900;color:#1e293b;min-width:16px;text-align:right;flex-shrink:0;
}
.vpt-team{flex:2;display:flex;align-items:center}
.vpt-td{
  font-size:.82rem;font-weight:800;flex:1;text-align:center;
}
.vpt-td.vt-pos{color:#22c55e;text-shadow:0 0 18px rgba(34,197,94,.2)}
.vpt-td.vt-neg{color:#ef4444;text-shadow:0 0 18px rgba(239,68,68,.15)}
.vpt-td.vt-neu{color:#94a3b8}
.vpt-team-txt{font-size:.76rem;font-weight:800;color:#e2e8f0}

/* Pick cards in side panel */
.vw-pick{
  background:rgba(255,255,255,.03);border-radius:12px;
  padding:13px 15px;border-left:3px solid #f07820;
  display:flex;align-items:center;justify-content:space-between;gap:10px;
  margin-bottom:5px;transition:background .15s;
  border:1px solid rgba(255,255,255,.07);border-left-width:3px;
}
.vw-pick:hover{background:rgba(255,255,255,.05)}
.vw-pick-star{border-left-color:#f59e0b;background:rgba(245,158,11,.05)}
.vw-star{color:#f59e0b;font-size:.9rem;margin-right:5px;flex-shrink:0}
.vw-pick-text{font-size:.88rem;font-weight:800;color:#e2e8f0;flex:1;min-width:0}
.vw-pick-badges{display:flex;gap:6px;align-items:center;flex-shrink:0}
.vw-odds{
  font-size:.7rem;font-weight:700;
  background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.12);
  border-radius:6px;padding:3px 9px;color:#94a3b8;
}
.vw-ev{
  font-size:.68rem;font-weight:900;color:#22c55e;
  background:rgba(34,197,94,.12);border:1px solid rgba(34,197,94,.25);
  border-radius:6px;padding:3px 9px;
  text-shadow:0 0 12px rgba(34,197,94,.3);
}
.detail-empty{
  text-align:center;padding:60px 20px;color:#1e293b;font-size:.82rem;
  letter-spacing:.06em;line-height:1.8;
}

/* ══════════════════════════════════════════════════════════════════════
   PREMIUM OVERHAUL — overrides everything above
   ══════════════════════════════════════════════════════════════════════ */

/* Body */
body{
  background:#05080d;
  background-image:
    radial-gradient(rgba(240,120,32,.006) 1px,transparent 1px),
    radial-gradient(rgba(255,255,255,.015) 1px,transparent 1px);
  background-size:40px 40px,10px 10px;
  background-position:0 0,5px 5px;
  font-family:'Inter',system-ui,sans-serif;
  color:#f1f5f9;min-height:100vh;
}

/* ── Header ── */
.site-hdr{
  position:sticky;top:0;z-index:100;
  background:rgba(5,8,13,.95);
  backdrop-filter:blur(24px);-webkit-backdrop-filter:blur(24px);
  border-bottom:1px solid rgba(255,255,255,.06);
}
.hdr-inner{max-width:900px;margin:0 auto;padding:14px 24px;display:flex;align-items:center;justify-content:space-between}
.brand-dot{width:10px;height:10px;border-radius:50%;background:#f07820;box-shadow:0 0 20px rgba(240,120,32,.8),0 0 40px rgba(240,120,32,.3)}
.brand-name{font-size:.85rem;font-weight:900;letter-spacing:.25em;text-transform:uppercase}
.brand-name em{color:#f07820;font-style:normal}
.live-badge{
  background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.2);
  border-radius:99px;padding:4px 12px;font-size:.58rem;font-weight:800;letter-spacing:.12em;color:#10b981;
  display:flex;align-items:center;gap:6px;
}
.live-dot{width:6px;height:6px;border-radius:50%;background:#10b981;box-shadow:0 0 10px rgba(16,185,129,1);animation:blink 2s ease-in-out infinite}
.hdr-date{font-size:.64rem;color:#475569;font-weight:600;letter-spacing:.06em}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.1}}

/* ── League Nav ── */
.lg-nav{
  position:sticky;top:52px;z-index:99;
  background:rgba(5,8,13,.92);
  backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);
  border-bottom:1px solid rgba(255,255,255,.06);
}
.lg-nav-inner{max-width:900px;margin:0 auto;display:flex}
.tab{flex:1;padding:14px 10px;display:flex;align-items:center;justify-content:center;gap:8px;cursor:pointer;border-bottom:3px solid transparent;transition:all .2s}
.tab:hover{background:rgba(255,255,255,.025)}
.tab.active{border-bottom-color:#f07820;background:rgba(240,120,32,.04)}
.tab-icon{font-size:1rem}
.tab-name{font-size:.72rem;font-weight:900;letter-spacing:.18em;text-transform:uppercase;color:#475569;transition:color .2s}
.tab.active .tab-name{color:#f07820}
.tab-rec{font-size:.65rem;font-weight:800;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);border-radius:99px;padding:3px 10px;color:#334155;transition:all .2s}
.tab.active .tab-rec{background:rgba(240,120,32,.12);border-color:rgba(240,120,32,.3);color:#f07820}

/* ── Main ── */
.site-main{max-width:900px;margin:0 auto;padding:28px 24px 80px}
.panel{display:none}.panel.active{display:block}

/* ── League Hero ── */
.lg-hero{
  background:linear-gradient(140deg,rgba(255,255,255,.055) 0%,rgba(255,255,255,.015) 100%);
  border:1px solid rgba(255,255,255,.1);border-radius:24px;
  padding:30px 32px;margin-bottom:24px;position:relative;overflow:hidden;
  box-shadow:0 16px 64px rgba(0,0,0,.5),inset 0 1px 0 rgba(255,255,255,.07);
}
.lg-hero::before{
  content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,var(--acc,#f07820) 0%,rgba(240,120,32,.3) 50%,transparent 100%);
}
.lg-hero::after{
  content:'';position:absolute;top:-100px;left:-50px;
  width:320px;height:320px;border-radius:50%;
  background:radial-gradient(circle,var(--glow,rgba(240,120,32,.14)) 0%,transparent 65%);
  pointer-events:none;
}
/* Second glow orb top-right */
.lg-hero .hero-rec::after{
  content:'';position:absolute;top:-60px;right:-60px;
  width:200px;height:200px;border-radius:50%;
  background:radial-gradient(circle,rgba(255,255,255,.02) 0%,transparent 70%);
  pointer-events:none;
}
.hero-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:20px;position:relative;z-index:1}
.hero-league-lbl{font-size:.62rem;font-weight:900;letter-spacing:.24em;text-transform:uppercase;color:var(--acc,#f07820)}
.hero-pending{
  font-size:.58rem;font-weight:800;color:#f59e0b;
  background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.25);
  border-radius:99px;padding:3px 12px;
  box-shadow:0 0 16px rgba(245,158,11,.1);
}
.hero-rec{display:flex;align-items:baseline;gap:2px;margin-bottom:22px;position:relative;z-index:1}
.r-big{font-size:5.5rem;font-weight:900;line-height:.85;letter-spacing:-.03em}
.r-w{color:#10b981;text-shadow:0 0 70px rgba(16,185,129,.5),0 0 120px rgba(16,185,129,.2),0 2px 0 rgba(0,0,0,.5)}
.r-l{color:#f43f5e;text-shadow:0 0 70px rgba(244,63,94,.4),0 0 120px rgba(244,63,94,.15),0 2px 0 rgba(0,0,0,.5)}
.r-sep{font-size:3rem;font-weight:300;color:#0d1520;margin:0 12px;align-self:center}
.r-push{font-size:1.5rem;font-weight:700;color:#334155;margin-left:16px;align-self:flex-end;margin-bottom:6px}
.stats-4{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;position:relative;z-index:1}
@media(max-width:480px){.stats-4{grid-template-columns:repeat(2,1fr)}}
.s4-cell{
  background:rgba(0,0,0,.45);border:1px solid rgba(255,255,255,.07);
  border-radius:14px;padding:14px 10px;text-align:center;
  transition:border-color .15s;
}
.s4-cell:hover{border-color:rgba(255,255,255,.14)}
.s4-val{font-size:1.1rem;font-weight:900;margin-bottom:4px;letter-spacing:-.02em}
.s4-lbl{font-size:.5rem;font-weight:800;letter-spacing:.14em;text-transform:uppercase;color:#334155}

/* ── Section headers on page ── */
.section-hdr{display:flex;align-items:center;justify-content:space-between;margin:28px 0 14px}
.section-title{
  font-size:.68rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;color:#f1f5f9;
  display:flex;align-items:center;gap:10px;
}
.section-title::before{
  content:'';width:4px;height:18px;
  background:linear-gradient(180deg,var(--acc,#f07820),rgba(240,120,32,.3));
  border-radius:3px;box-shadow:0 0 10px rgba(240,120,32,.4);
}
.section-btn{
  font-size:.62rem;font-weight:700;color:#f07820;cursor:pointer;
  background:rgba(240,120,32,.08);border:1px solid rgba(240,120,32,.25);
  border-radius:9px;padding:6px 16px;font-family:inherit;letter-spacing:.05em;
  transition:all .18s;
}
.section-btn:hover{
  background:rgba(240,120,32,.2);
  box-shadow:0 0 20px rgba(240,120,32,.25);
  transform:translateY(-1px);
}

/* ── Today empty state ── */
.today-empty{
  text-align:center;padding:28px 16px;
  background:rgba(255,255,255,.02);border:1px dashed rgba(255,255,255,.07);
  border-radius:14px;color:#334155;font-size:.78rem;line-height:2;
}

/* ── Team badge ── */
.team-badge{
  display:inline-flex;align-items:center;justify-content:center;
  border-radius:10px;font-weight:900;font-size:.7rem;
  white-space:nowrap;flex-shrink:0;padding:7px 13px;
  letter-spacing:.06em;line-height:1;font-family:inherit;
  box-shadow:0 2px 8px rgba(0,0,0,.35);
}
.matchup-vs{font-size:.65rem;color:#1e293b;font-weight:900;letter-spacing:.1em}
.game-time{font-size:.63rem;color:#475569;margin-left:auto;white-space:nowrap;font-weight:700}

/* ── Game card ── */
.game-card{
  background:linear-gradient(140deg,rgba(255,255,255,.05) 0%,rgba(255,255,255,.018) 100%);
  border:1px solid rgba(255,255,255,.1);
  border-radius:20px;padding:20px 22px;margin-bottom:12px;
  position:relative;overflow:hidden;
  box-shadow:0 6px 32px rgba(0,0,0,.35),inset 0 1px 0 rgba(255,255,255,.04);
  transition:all .22s;
}
.game-card:hover{
  border-color:rgba(255,255,255,.18);
  box-shadow:0 12px 48px rgba(0,0,0,.5),inset 0 1px 0 rgba(255,255,255,.06);
  transform:translateY(-2px);
}
.game-card::before{
  content:'';position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,var(--acc,#f07820) 0%,rgba(240,120,32,.4) 40%,transparent 80%);
}
.game-card::after{
  content:'';position:absolute;top:-60px;right:-30px;width:180px;height:180px;border-radius:50%;
  background:radial-gradient(circle,rgba(255,255,255,.015) 0%,transparent 70%);pointer-events:none;
}
.game-matchup{display:flex;align-items:center;gap:10px;margin-bottom:14px;flex-wrap:wrap;position:relative;z-index:1}
.game-picks{display:flex;flex-direction:column;gap:7px;position:relative;z-index:1}
.gpick{
  display:flex;align-items:center;gap:10px;padding:11px 14px;
  background:rgba(255,255,255,.035);border-radius:12px;
  border-left:3px solid var(--acc,#f07820);
  transition:all .18s;
}
.gpick:hover{background:rgba(255,255,255,.065);transform:translateX(2px)}
.gpick.star{border-left-color:#f59e0b;background:rgba(245,158,11,.06)}
.gpick.star:hover{background:rgba(245,158,11,.1)}
.gpick-ico{font-size:.85rem;flex-shrink:0;color:#f07820;min-width:13px;text-align:center}
.gpick.star .gpick-ico{color:#f59e0b;text-shadow:0 0 12px rgba(245,158,11,.6)}
.gpick-pick{font-size:.88rem;font-weight:800;color:#f1f5f9;flex:1;min-width:0}
.gpick-odds{
  font-size:.72rem;font-weight:700;background:rgba(255,255,255,.08);
  border:1px solid rgba(255,255,255,.12);border-radius:7px;
  padding:4px 11px;color:#94a3b8;flex-shrink:0;
}
.gpick-ev{
  font-size:.68rem;font-weight:900;color:#10b981;
  background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.28);
  border-radius:7px;padding:4px 11px;flex-shrink:0;
  text-shadow:0 0 16px rgba(16,185,129,.5);
}

/* ── IR / Schedule mini rows ── */
.mini-table{display:flex;flex-direction:column;gap:5px}
.mt-row{
  display:flex;align-items:center;gap:12px;
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:12px;padding:11px 14px;
}
.mt-name{font-size:.82rem;font-weight:700;color:#f1f5f9;flex:1;min-width:0}
.mt-sub{font-size:.66rem;color:#475569;flex-shrink:0}
.mt-status{font-size:.62rem;font-weight:800;border-radius:7px;padding:3px 10px;flex-shrink:0}
.mt-out{background:rgba(244,63,94,.15);color:#f43f5e;border:1px solid rgba(244,63,94,.25)}
.mt-dbt{background:rgba(249,115,22,.12);color:#fb923c;border:1px solid rgba(249,115,22,.22)}
.mt-qst{background:rgba(234,179,8,.1);color:#eab308;border:1px solid rgba(234,179,8,.2)}

/* ── Action row — legacy (kept for compatibility) ── */
.action-row{display:flex;gap:8px;flex-wrap:wrap;margin:16px 0 4px}
.act-btn{display:flex;flex-direction:row;align-items:center;gap:6px;
  background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.09);
  border-radius:99px;padding:9px 18px;cursor:pointer;font-family:inherit;
  color:#cbd5e1;transition:all .18s;white-space:nowrap}
.act-btn:hover{background:rgba(255,255,255,.1);border-color:rgba(255,255,255,.18);transform:translateY(-1px)}
.act-btn.green{background:rgba(16,185,129,.08);border-color:rgba(16,185,129,.22);color:#10b981}
.act-btn.blue{background:rgba(59,130,246,.08);border-color:rgba(59,130,246,.22);color:#60a5fa}
.act-icon{font-size:.9rem;line-height:1}
.act-lbl{font-size:.68rem;font-weight:700;letter-spacing:.06em;text-transform:uppercase}

/* ── NEW Panel Actions Layout ── */
.pa-wrap{margin:18px 0 4px}
/* Primary CTA row */
.pa-primary{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}
.pa-cta{
  display:flex;align-items:center;justify-content:center;gap:8px;
  padding:14px 10px;border-radius:14px;cursor:pointer;font-family:inherit;
  border:1.5px solid transparent;transition:all .18s;font-weight:900;
  letter-spacing:.06em;text-transform:uppercase;font-size:.72rem;
}
.pa-cta.green{
  background:linear-gradient(135deg,rgba(16,185,129,.18),rgba(16,185,129,.08));
  border-color:rgba(16,185,129,.4);color:#34d399;
  box-shadow:0 0 20px rgba(16,185,129,.1),inset 0 1px 0 rgba(255,255,255,.04);
}
.pa-cta.green:hover{background:linear-gradient(135deg,rgba(16,185,129,.3),rgba(16,185,129,.15));box-shadow:0 6px 28px rgba(16,185,129,.3);transform:translateY(-1px)}
.pa-cta.blue{
  background:linear-gradient(135deg,rgba(59,130,246,.15),rgba(59,130,246,.07));
  border-color:rgba(59,130,246,.38);color:#93c5fd;
  box-shadow:0 0 20px rgba(59,130,246,.08),inset 0 1px 0 rgba(255,255,255,.04);
}
.pa-cta.blue:hover{background:linear-gradient(135deg,rgba(59,130,246,.28),rgba(59,130,246,.15));box-shadow:0 6px 28px rgba(59,130,246,.28);transform:translateY(-1px)}
.pa-cta.orange{
  background:linear-gradient(135deg,rgba(245,166,35,.18),rgba(245,166,35,.07));
  border-color:rgba(245,166,35,.4);color:#f5a623;
  box-shadow:0 0 20px rgba(245,166,35,.1),inset 0 1px 0 rgba(255,255,255,.04);
}
.pa-cta.orange:hover{background:linear-gradient(135deg,rgba(245,166,35,.3),rgba(245,166,35,.15));box-shadow:0 6px 28px rgba(245,166,35,.28);transform:translateY(-1px)}
.pa-cta-icon{font-size:1.1rem;line-height:1}
/* Secondary icon grid */
.pa-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}
.pa-item{
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  gap:5px;padding:12px 6px 10px;
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:14px;cursor:pointer;font-family:inherit;color:#94a3b8;
  transition:all .16s;
}
.pa-item:hover{background:rgba(255,255,255,.07);border-color:rgba(255,255,255,.14);color:#f1f5f9;transform:translateY(-1px)}
.pa-item:active{transform:translateY(0)}
.pa-item-icon{font-size:1.15rem;line-height:1}
.pa-item-lbl{font-size:.46rem;font-weight:800;letter-spacing:.1em;text-transform:uppercase;text-align:center}
/* Section separator inside pa-wrap */
.pa-sep{height:1px;background:rgba(255,255,255,.05);margin:10px 0}

/* ── Cmd-cards inside tools ── */
.sec-lbl{font-size:.56rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase;color:#334155;padding:6px 0 8px;margin-top:4px}
.sec-div{height:1px;background:rgba(255,255,255,.05);margin:10px 0 12px}
.cmds{display:grid;gap:7px}
.cmd-card{
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:12px;padding:12px 15px;
  display:flex;align-items:center;justify-content:space-between;gap:12px;
  transition:background .15s;
}
.cmd-card:hover{background:rgba(255,255,255,.05)}
.cmd-desc{font-size:.75rem;color:#64748b;flex:1;min-width:0;line-height:1.4}
.btn{
  background:#f07820;color:#fff;border:none;border-radius:9px;
  padding:9px 18px;font-size:.74rem;font-weight:700;cursor:pointer;
  white-space:nowrap;flex-shrink:0;font-family:inherit;letter-spacing:.03em;
  transition:all .15s;
}
.btn:hover{box-shadow:0 0 20px rgba(240,120,32,.4);opacity:.9}
.btn.green{background:#059669}.btn.green:hover{box-shadow:0 0 20px rgba(5,150,105,.4)}
.btn.blue{background:#2563eb}.btn.blue:hover{box-shadow:0 0 20px rgba(37,99,235,.4)}
.btn.red{background:#e11d48}.btn.red:hover{box-shadow:0 0 16px rgba(225,29,72,.35)}
.btn.gray{background:#0f172a;color:#64748b;border:1px solid rgba(255,255,255,.07)}
.btn.gray:hover{background:#1e293b;box-shadow:none}

/* ── Tools accordion ── */
.tools-group{
  background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
  border-radius:16px;overflow:hidden;margin-top:18px;
}
.tools-toggle{
  padding:15px 20px;display:flex;align-items:center;justify-content:space-between;
  cursor:pointer;font-size:.62rem;font-weight:800;color:#475569;
  letter-spacing:.14em;text-transform:uppercase;user-select:none;transition:background .15s;
}
.tools-toggle:hover{background:rgba(255,255,255,.025)}
.tools-chevron{transition:transform .22s;font-style:normal;font-size:.8rem;display:inline-block}
.tools-chevron.open{transform:rotate(90deg)}
.tools-body{display:none;padding:8px 12px 14px;border-top:1px solid rgba(255,255,255,.06)}
.tools-body.open{display:block}

/* ── Modals ── */
.modal-bg{
  display:none;position:fixed;inset:0;
  background:rgba(0,0,0,.78);
  backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);
  z-index:200;overflow-y:auto;padding:20px 16px;
}
.modal-bg.open{display:flex;align-items:flex-start;justify-content:center}
.modal{
  background:#080e16;border:1px solid rgba(255,255,255,.1);
  border-radius:20px;width:100%;max-width:520px;padding:26px;margin:auto 0;
  box-shadow:0 24px 80px rgba(0,0,0,.6);
}
.modal h2{font-size:.9rem;font-weight:900;margin-bottom:20px;color:#f07820;letter-spacing:.06em}
label{display:block;font-size:.6rem;color:#475569;font-weight:800;letter-spacing:.12em;text-transform:uppercase;margin-bottom:5px}
input,select,textarea{
  width:100%;background:rgba(0,0,0,.5);color:#f1f5f9;
  border:1px solid rgba(255,255,255,.09);
  border-radius:10px;padding:12px 14px;font-size:.92rem;
  margin-bottom:14px;-webkit-appearance:none;font-family:inherit;transition:border-color .2s;
}
input:focus,select:focus,textarea:focus{outline:none;border-color:#f07820;box-shadow:0 0 0 3px rgba(240,120,32,.08)}
textarea{height:72px;resize:none}
.row2{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px}
.btn-row{display:flex;gap:10px;margin-top:6px}.btn-row .btn{flex:1}
.alert{padding:12px 14px;border-radius:10px;margin-bottom:14px;font-weight:600;font-size:.84rem}
.alert.ok{background:#052e16;color:#86efac;border:1px solid rgba(134,239,172,.2)}
.alert.err{background:#1c0a0a;color:#fca5a5;border:1px solid rgba(252,165,165,.2)}
.leg{background:rgba(0,0,0,.35);border-radius:10px;padding:12px;border:1px solid rgba(255,255,255,.07);margin-bottom:10px}
.leg-title{font-size:.65rem;font-weight:700;color:#f07820;margin-bottom:8px;letter-spacing:.1em}
.section-sep{height:1px;background:rgba(255,255,255,.07);margin:14px 0}

/* ── Output panel (bottom terminal — for background tasks) ── */
.out-bg{
  display:none;position:fixed;inset:0;
  background:rgba(0,0,0,.8);backdrop-filter:blur(4px);-webkit-backdrop-filter:blur(4px);
  z-index:300;align-items:flex-end;
}
.out-bg.open{display:flex}
.out-panel{
  background:#06090e;border-top:2px solid #f07820;
  border-radius:18px 18px 0 0;width:100%;max-height:70vh;display:flex;flex-direction:column;
  box-shadow:0 -20px 60px rgba(0,0,0,.6);
}
.out-header{
  padding:14px 22px;display:flex;align-items:center;justify-content:space-between;
  border-bottom:1px solid rgba(255,255,255,.07);flex-shrink:0;
}
.out-hdr-left{display:flex;align-items:center;gap:8px}
.out-dot{width:6px;height:6px;border-radius:50%;background:#f07820;box-shadow:0 0 10px rgba(240,120,32,.9);animation:blink 1.4s ease-in-out infinite}
.out-title-lbl{font-size:.76rem;font-weight:800;color:#f07820;letter-spacing:.08em}
.out-body{overflow-y:auto;padding:16px 22px;font-family:'SF Mono','Fira Code',Menlo,monospace;font-size:.72rem;color:#a3e635;white-space:pre-wrap;flex:1;line-height:1.8}
.spinner{display:inline-block;width:13px;height:13px;border:2px solid #f07820;border-top-color:transparent;border-radius:50%;animation:spin .7s linear infinite;margin-right:8px;vertical-align:middle}
@keyframes spin{to{transform:rotate(360deg)}}
.t-ok{color:#10b981}.t-err{color:#f43f5e}.t-warn{color:#f59e0b}.t-hi{color:#f07820}.t-dim{color:#334155}

/* ── Pick cards (log historial) ── */
.pick-card{
  background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.09);
  border-radius:14px;padding:17px 18px;margin-bottom:10px;position:relative;overflow:hidden;
}
.pick-card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;background:var(--acc,#f07820)}
.pc-game{font-size:.63rem;font-weight:800;color:#475569;letter-spacing:.1em;text-transform:uppercase;margin-bottom:6px}
.pc-pick{font-size:1.05rem;font-weight:900;color:#f1f5f9;margin-bottom:9px}
.pc-row{display:flex;align-items:center;gap:7px;flex-wrap:wrap}
.pc-odds{font-size:.72rem;font-weight:700;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.1);border-radius:7px;padding:3px 9px;color:#94a3b8}
.pc-ev{font-size:.68rem;font-weight:900;color:#10b981;background:rgba(16,185,129,.1);border:1px solid rgba(16,185,129,.22);border-radius:7px;padding:3px 9px}
.pc-tag{font-size:.64rem;font-weight:700;color:#f59e0b;background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.22);border-radius:7px;padding:3px 9px}
.pc-note{font-size:.72rem;color:#64748b;margin-top:10px;line-height:1.6;padding-top:10px;border-top:1px solid rgba(255,255,255,.05)}

/* ── Log history entries ── */
.log-entry{
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:12px;padding:13px 15px;margin-bottom:8px;display:flex;align-items:flex-start;gap:12px;
}
.le-result{width:32px;height:32px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:.66rem;font-weight:900}
.le-w{background:rgba(16,185,129,.12);color:#10b981;border:1px solid rgba(16,185,129,.3)}
.le-l{background:rgba(244,63,94,.12);color:#f43f5e;border:1px solid rgba(244,63,94,.3)}
.le-p{background:rgba(100,116,139,.12);color:#94a3b8;border:1px solid rgba(100,116,139,.28)}
.le-n{background:rgba(245,158,11,.08);color:#f59e0b;border:1px solid rgba(245,158,11,.22)}
.le-body{flex:1;min-width:0}
.le-game{font-size:.65rem;color:#475569;margin-bottom:3px;letter-spacing:.03em}
.le-pick{font-size:.85rem;font-weight:700;color:#f1f5f9}
.le-meta{font-size:.64rem;color:#334155;margin-top:4px}
.le-pos{color:#10b981;font-weight:700}.le-neg{color:#f43f5e;font-weight:700}

/* ── Grade rows ── */
.btn.green{background:#059669}.btn.blue{background:#2563eb}

/* ── Detail overlay ── */
.detail-overlay{position:fixed;inset:0;background:rgba(0,0,0,.65);backdrop-filter:blur(5px);-webkit-backdrop-filter:blur(5px);z-index:150;display:none}
.detail-overlay.open{display:block}

/* ── SIDE PANEL ── */
.detail-panel{
  position:fixed;top:0;right:0;width:min(620px,100vw);height:100%;height:100dvh;
  background:linear-gradient(160deg,#0a1018 0%,#060a10 100%);
  border-left:1px solid rgba(255,255,255,.09);
  z-index:151;display:flex;flex-direction:column;
  transform:translateX(100%);transition:transform .35s cubic-bezier(.4,0,.2,1);
  box-shadow:-50px 0 120px rgba(0,0,0,.8);
  overscroll-behavior:contain;
}
.detail-panel.open{transform:translateX(0)}
.detail-hdr{
  padding:22px 28px;display:flex;align-items:center;justify-content:space-between;
  border-bottom:1px solid rgba(255,255,255,.07);flex-shrink:0;
  background:linear-gradient(90deg,rgba(0,0,0,.3),rgba(0,0,0,.1));
}
.detail-title{font-size:1rem;font-weight:900;color:#fff;letter-spacing:.14em;text-transform:uppercase}
.detail-body{
  overflow-y:auto;padding:22px 28px;flex:1;
  -webkit-overflow-scrolling:touch;
  overscroll-behavior:contain;
}
.detail-empty{text-align:center;padding:64px 20px;color:#1e293b;font-size:.84rem;letter-spacing:.06em;line-height:2}

/* ── Side panel content ── */
.vw-output{display:flex;flex-direction:column;gap:3px}
.vw-hdr{
  display:flex;align-items:center;gap:10px;
  font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
  color:#f07820;padding:20px 0 10px;
  border-bottom:1px solid rgba(240,120,32,.14);
  margin-bottom:10px;margin-top:8px;
}
.vw-hdr::before{content:'';width:4px;height:16px;background:linear-gradient(180deg,#f07820,rgba(240,120,32,.2));border-radius:2px;flex-shrink:0}
.vw-hdr:first-child{padding-top:4px;margin-top:0}
.vw-text{font-size:.78rem;color:#64748b;padding:4px 2px;line-height:1.65}
.vw-stat-line{font-size:.76rem;font-weight:600;color:#475569;padding:6px 4px;border-bottom:1px solid rgba(255,255,255,.04);line-height:1.6}

/* ── Space-table ── */
.vw-table{display:flex;flex-direction:column;gap:3px;margin-bottom:6px}
.vt-row{display:flex;align-items:center;justify-content:space-between;background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);border-radius:10px;padding:9px 13px;gap:8px;transition:background .15s}
.vt-row:hover{background:rgba(255,255,255,.05)}
.vt-team{font-size:.76rem;font-weight:800;color:#f1f5f9;min-width:72px;flex-shrink:0}
.vt-vals{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
.vt-cell{font-size:.76rem;font-weight:700}
.vt-pos{color:#10b981}.vt-neg{color:#f43f5e}.vt-neu{color:#94a3b8}
.vt-label{color:#475569;font-size:.68rem;font-weight:500}

/* ── Pipe-table stats (the premium rankings table) ── */
.vpt{display:flex;flex-direction:column;gap:5px;margin-bottom:12px}
.vpt-hdr{
  display:flex;align-items:center;padding:8px 16px;margin-bottom:2px;
  background:rgba(255,255,255,.015);border-radius:8px;
}
.vpt-th{font-size:.5rem;font-weight:900;letter-spacing:.18em;text-transform:uppercase;color:#1e293b;flex:1;text-align:center}
.vpt-th:first-child{flex:.4;text-align:center}
.vpt-th:nth-child(2){flex:2.4;text-align:left}
.vpt-row{
  display:flex;align-items:center;gap:10px;
  padding:12px 16px;
  background:rgba(255,255,255,.03);
  border:1px solid rgba(255,255,255,.06);
  border-radius:12px;transition:all .18s;
  position:relative;overflow:hidden;
}
.vpt-row:hover{
  background:rgba(255,255,255,.06);
  border-color:rgba(255,255,255,.13);
  transform:translateX(-3px);
  box-shadow:4px 0 20px rgba(0,0,0,.3);
}
/* Top-3 rank highlights */
.vpt-row:nth-child(2){border-left:3px solid #f07820;background:rgba(240,120,32,.04)}
.vpt-row:nth-child(3){border-left:3px solid rgba(240,120,32,.55)}
.vpt-row:nth-child(4){border-left:3px solid rgba(240,120,32,.25)}
.vpt-rank{
  font-size:.58rem;font-weight:900;color:#1e293b;
  min-width:20px;text-align:center;flex-shrink:0;
}
.vpt-row:nth-child(2) .vpt-rank{color:#f07820;font-size:.65rem}
.vpt-team{flex:2.4;display:flex;align-items:center;gap:7px}
.vpt-td{
  font-size:.9rem;font-weight:900;flex:1;text-align:center;
  position:relative;padding-bottom:5px;
}
.vpt-td.vt-pos{color:#10b981;text-shadow:0 0 28px rgba(16,185,129,.35)}
.vpt-td.vt-neg{color:#f43f5e;text-shadow:0 0 28px rgba(244,63,94,.25)}
.vpt-td.vt-neu{color:#94a3b8}
.vpt-bar{
  position:absolute;bottom:0;left:50%;transform:translateX(-50%);
  height:2px;width:var(--pct,0%);background:currentColor;
  opacity:.4;border-radius:1px;min-width:6px;
  transition:width .4s ease;
}
.vpt-team-txt{font-size:.82rem;font-weight:900;color:#f1f5f9}

/* ── Pick cards in side panel ── */
.vw-pick{
  background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.09);
  border-left:3px solid #f07820;
  border-radius:14px;padding:15px 17px;
  display:flex;align-items:center;justify-content:space-between;gap:10px;
  margin-bottom:7px;transition:all .18s;
  box-shadow:0 2px 12px rgba(0,0,0,.2);
}
.vw-pick:hover{
  background:rgba(255,255,255,.08);
  transform:translateX(-3px);
  box-shadow:0 6px 24px rgba(0,0,0,.35);
  border-color:rgba(240,120,32,.4);
}
.vw-pick-star{
  border-left-color:#f59e0b;background:rgba(245,158,11,.07);
}
.vw-pick-star:hover{border-color:rgba(245,158,11,.5)}
.vw-star{color:#f59e0b;font-size:1rem;margin-right:6px;flex-shrink:0;text-shadow:0 0 14px rgba(245,158,11,.6)}
.vw-pick-text{font-size:.9rem;font-weight:800;color:#f1f5f9;flex:1;min-width:0;line-height:1.3}
.vw-pick-badges{display:flex;gap:7px;align-items:center;flex-shrink:0}
.vw-odds{
  font-size:.72rem;font-weight:700;background:rgba(255,255,255,.08);
  border:1px solid rgba(255,255,255,.12);border-radius:7px;padding:4px 11px;color:#94a3b8;
}
.vw-ev{
  font-size:.7rem;font-weight:900;color:#10b981;
  background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.3);
  border-radius:7px;padding:4px 11px;
  text-shadow:0 0 18px rgba(16,185,129,.5);
}

/* ── Fade-in animation for side panel content ── */
@keyframes fadeSlideIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:none}}
.vw-output>*{animation:fadeSlideIn .25s ease both}
.vw-output>*:nth-child(2){animation-delay:.04s}
.vw-output>*:nth-child(3){animation-delay:.08s}
.vw-output>*:nth-child(4){animation-delay:.12s}
.vw-output>*:nth-child(5){animation-delay:.16s}
.vw-output>*:nth-child(6){animation-delay:.2s}
.vw-output>*:nth-child(n+7){animation-delay:.24s}

/* ═══════════════════════════════════════════════════
   MOBILE  (iPhone / ≤ 430 px)
   ═══════════════════════════════════════════════════ */
@media(max-width:430px){

  /* ── Header ── */
  .hdr-inner{padding:10px 14px;gap:8px}
  .brand{gap:7px}
  .brand-dot{width:8px;height:8px}
  .brand-name{font-size:.72rem;letter-spacing:.18em}
  .hdr-right{gap:6px}
  .hdr-date{display:none}
  .live-badge{padding:3px 8px;font-size:.5rem;gap:4px}
  .live-dot{width:5px;height:5px}
  /* shrink the 📊 Record header button */
  .hdr-right .btn.gray{padding:4px 9px;font-size:.56rem}

  /* ── League nav ── */
  .lg-nav-inner{gap:0}
  .tab{padding:10px 4px;gap:4px}
  .tab-icon{font-size:.85rem}
  .tab-name{font-size:.58rem;letter-spacing:.1em}
  .tab-rec{font-size:.54rem;padding:2px 6px;display:none}

  /* ── Main content ── */
  .site-main{padding:14px 12px 70px}

  /* ── League hero card ── */
  .lg-hero{padding:18px 16px;border-radius:18px}
  .hero-top{margin-bottom:14px}
  .hero-rec{margin-bottom:16px;flex-wrap:wrap;gap:0}
  .r-big{font-size:3.8rem}
  .r-sep{font-size:2rem;margin:0 6px}
  .r-push{font-size:1.1rem;margin-left:10px;margin-bottom:3px}
  .stats-4{gap:6px}
  .s4-cell{padding:10px 6px;border-radius:10px}
  .s4-val{font-size:.95rem}
  .s4-lbl{font-size:.44rem;letter-spacing:.1em}

  /* ── Game cards ── */
  .game-card{padding:14px 14px;border-radius:16px}
  .game-matchup{gap:6px;margin-bottom:10px}
  .team-badge{padding:5px 9px;font-size:.62rem;border-radius:8px}
  .matchup-vs{font-size:.58rem}
  .game-time{font-size:.52rem}
  .gpick{padding:9px 11px;gap:8px}
  .gpick-pick{font-size:.8rem}
  .gpick-odds{font-size:.65rem;padding:3px 8px}
  .gpick-ev{font-size:.62rem;padding:3px 8px}

  /* ── Action row ── */
  .action-row{gap:6px;margin:16px 0 4px}
  /* big hero buttons: each takes ~half row */
  .act-btn.big{flex:1;min-width:calc(50% - 3px);justify-content:center;padding:11px 12px}
  .act-btn.big .act-lbl{font-size:.65rem}
  /* small buttons: slightly smaller */
  .act-btn{padding:7px 13px}
  .act-lbl{font-size:.6rem}

  /* ── Detail side panel (stats tables, logs) ── */
  .detail-panel{width:100%!important;max-width:100%!important;border-radius:18px 18px 0 0}
  .detail-body{padding:14px 12px;overflow-x:auto}
  .detail-hdr{padding:14px 16px}

  /* ── Modals ── */
  .modal{padding:20px 16px;border-radius:16px}
  .modal-bg{padding:10px 8px}
  .row2{grid-template-columns:1fr}
  .row3{grid-template-columns:1fr 1fr}

  /* ── Output/terminal panel ── */
  .out-body{font-size:.65rem;padding:12px 14px}

  /* ── Tools accordion ── */
  .tools-toggle{padding:12px 16px;font-size:.58rem}
  .cmd-card{padding:10px 12px;gap:8px;flex-wrap:wrap}
  .cmd-desc{font-size:.68rem}

  /* ── Log/historial cards ── */
  .pick-card{padding:12px 14px;border-radius:14px}
  .le-ico{width:34px;height:34px;font-size:.72rem}

  /* ── Section headers ── */
  .section-hdr{margin:20px 0 10px}
  .section-title{font-size:.6rem;letter-spacing:.16em}

  /* ── Mini table rows (IR) ── */
  .mt-row{padding:9px 11px;gap:8px}
  .mt-name{font-size:.75rem}
}
"""

# ── Monte Carlo Modal — injected globally into every page ─────────────────
_MC_MODAL_HTML = """<!-- ── Monte Carlo Modal (global) ──────────────────────── -->
<div id="mc-modal" style="display:none;position:fixed;inset:0;z-index:9999;
  background:rgba(0,0,0,.82);backdrop-filter:blur(8px);overflow-y:auto;padding:20px 12px">
  <div style="max-width:460px;margin:0 auto;background:#0f172a;border:1px solid rgba(255,255,255,.12);
    border-radius:20px;padding:20px 18px;position:relative">
    <button onclick="closeMC()" style="position:absolute;top:12px;right:14px;background:none;
      border:none;color:#475569;font-size:1.1rem;cursor:pointer;line-height:1;padding:4px 6px">&#x2715;</button>
    <div style="font-size:.58rem;font-weight:900;letter-spacing:.2em;
      text-transform:uppercase;color:#a78bfa;margin-bottom:3px">Monte Carlo &middot; Poisson &middot; 20,000 sims</div>
    <div id="mc-matchup" style="font-size:1rem;font-weight:900;color:#f1f5f9;margin-bottom:14px;letter-spacing:.02em"></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:14px">
      <div style="background:rgba(56,189,248,.07);border:1px solid rgba(56,189,248,.2);border-radius:10px;padding:10px 8px;text-align:center">
        <div style="font-size:.48rem;color:#38bdf8;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:4px" id="mc-lbl-away-stat">Away proj.</div>
        <div id="mc-tA-disp" style="font-size:1.15rem;font-weight:900;color:#f1f5f9">&#x2014;</div>
        <div style="font-size:.5rem;color:#475569;margin-top:2px">runs (&lambda;)</div>
      </div>
      <div style="background:rgba(249,115,22,.07);border:1px solid rgba(249,115,22,.2);border-radius:10px;padding:10px 8px;text-align:center">
        <div style="font-size:.48rem;color:#f97316;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:4px" id="mc-lbl-home-stat">Home proj.</div>
        <div id="mc-tB-disp" style="font-size:1.15rem;font-weight:900;color:#f1f5f9">&#x2014;</div>
        <div style="font-size:.5rem;color:#475569;margin-top:2px">runs (&lambda;)</div>
      </div>
      <div style="background:rgba(167,139,250,.07);border:1px solid rgba(167,139,250,.2);border-radius:10px;padding:10px 8px;text-align:center">
        <div id="mc-lbl-stat3" style="font-size:.48rem;color:#a78bfa;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:4px">BetMGM Line</div>
        <div id="mc-line-disp" style="font-size:1.15rem;font-weight:900;color:#f1f5f9">&#x2014;</div>
        <div id="mc-sub-stat3" style="font-size:.5rem;color:#475569;margin-top:2px">market total</div>
      </div>
      <div style="background:rgba(16,185,129,.06);border:1px solid rgba(16,185,129,.18);border-radius:10px;padding:10px 8px;text-align:center">
        <div id="mc-lbl-stat4" style="font-size:.48rem;color:#10b981;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:4px">Model Total</div>
        <div id="mc-model-disp" style="font-size:1.15rem;font-weight:900;color:#f1f5f9">&#x2014;</div>
        <div id="mc-sub-stat4" style="font-size:.5rem;color:#475569;margin-top:2px">projected runs</div>
      </div>
    </div>
    <div style="height:3px;background:rgba(255,255,255,.07);border-radius:2px;margin-bottom:14px;overflow:hidden">
      <div id="mc-prog" style="height:100%;background:linear-gradient(90deg,#a78bfa,#38bdf8);width:0%;transition:width .025s"></div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:8px;margin-bottom:16px">
      <div id="mc-card-over" style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.2);border-radius:10px;padding:10px 6px;text-align:center">
        <div id="mc-lbl-result1" style="font-size:.46rem;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.08em">P(Over)</div>
        <div id="mc-p-over" style="font-size:1.05rem;font-weight:900;color:#10b981;margin-top:3px">&#x2014;</div>
      </div>
      <div id="mc-card-under" style="background:rgba(124,58,237,.08);border:1px solid rgba(124,58,237,.2);border-radius:10px;padding:10px 6px;text-align:center">
        <div id="mc-lbl-result2" style="font-size:.46rem;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.08em">P(Under)</div>
        <div id="mc-p-under" style="font-size:1.05rem;font-weight:900;color:#7c3aed;margin-top:3px">&#x2014;</div>
      </div>
      <div style="background:rgba(100,116,139,.06);border:1px solid rgba(100,116,139,.15);border-radius:10px;padding:10px 6px;text-align:center">
        <div id="mc-lbl-result3" style="font-size:.46rem;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.08em">P(Push)</div>
        <div id="mc-p-push" style="font-size:1.05rem;font-weight:900;color:#475569;margin-top:3px">&#x2014;</div>
      </div>
      <div style="background:rgba(167,139,250,.07);border:1px solid rgba(167,139,250,.2);border-radius:10px;padding:10px 6px;text-align:center">
        <div style="font-size:.46rem;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.08em">Exp. Value</div>
        <div id="mc-ev-odds" style="font-size:.48rem;font-weight:700;color:#475569;margin-top:2px">&#x2014;</div>
        <div id="mc-ev" style="font-size:.95rem;font-weight:900;color:#a78bfa;margin-top:1px">&#x2014;</div>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
      <div>
        <div style="font-size:.48rem;color:#38bdf8;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px" id="mc-lbl-away">Away dist.</div>
        <div id="mc-dist-away" style="font-size:.56rem"></div>
      </div>
      <div>
        <div style="font-size:.48rem;color:#f97316;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px" id="mc-lbl-home">Home dist.</div>
        <div id="mc-dist-home" style="font-size:.56rem"></div>
      </div>
    </div>
    <div style="margin-bottom:14px">
      <div id="mc-lbl-combo" style="font-size:.48rem;color:#a78bfa;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px">Combined total</div>
      <div id="mc-dist-combo" style="font-size:.56rem"></div>
    </div>
    <button id="mc-run-btn" onclick="runMC()" style="width:100%;padding:9px;background:rgba(167,139,250,.1);border:1px solid rgba(167,139,250,.25);border-radius:10px;color:#a78bfa;font-size:.65rem;font-weight:800;letter-spacing:.12em;text-transform:uppercase;cursor:pointer">&#x21BA; Re-run simulation</button>
  </div>
</div>
<script>
(function(){
  function poissonSample(lam){var L=Math.exp(-lam),k=0,p=1;do{k++;p*=Math.random();}while(p>L);return k-1;}
  function renderDist(id,counts,total,maxBin,colorFn){
    var el=document.getElementById(id);if(!el)return;
    var maxC=1;for(var i=0;i<=maxBin;i++)if((counts[i]||0)>maxC)maxC=counts[i];
    var h='';
    for(var i=0;i<=maxBin;i++){
      var c=counts[i]||0,w=(c/maxC*100).toFixed(0),pct=total>0?(c/total*100).toFixed(1):'0.0';
      h+='<div style="display:flex;align-items:center;gap:4px;margin-bottom:2px">'
        +'<span style="width:16px;text-align:right;color:#334155;font-size:.5rem">'+i+'</span>'
        +'<div style="flex:1;background:rgba(255,255,255,.05);border-radius:2px;height:10px;overflow:hidden">'
        +'<div style="height:100%;border-radius:2px;width:'+w+'%;background:'+colorFn(i)+'"></div></div>'
        +'<span style="width:30px;color:#475569;font-size:.48rem;text-align:right">'+pct+'%</span></div>';
    }
    el.innerHTML=h;
  }
  function _odsFmt(o){return o>=100?'+'+o:''+o;}
  function _amPayout(o){return o>=100?o/100:100/Math.abs(o);}
  var _mcRunning=false,_mcTA=4.5,_mcTB=4.5,_mcMode='TOT',_mcP1=8.5,_mcP2=-110,_mcP3=-110,_mcLine=8.5;
  window.runMC=function(){
    if(_mcRunning)return;
    _mcRunning=true;
    var btn=document.getElementById('mc-run-btn');
    btn.disabled=true;btn.textContent='Simulating…';
    document.getElementById('mc-prog').style.width='0%';
    var tA=_mcTA,tB=_mcTB,N=20000,BATCH=600;
    var awayD={},homeD={},comboD={},over=0,under=0,push=0,done=0;
    var prog=document.getElementById('mc-prog');
    function batch(){
      var end=Math.min(done+BATCH,N);
      for(var i=done;i<end;i++){
        var a=poissonSample(tA),b=poissonSample(tB),tot=a+b;
        awayD[a]=(awayD[a]||0)+1;homeD[b]=(homeD[b]||0)+1;comboD[tot]=(comboD[tot]||0)+1;
        if(_mcMode==='TOT'){
          if(tot>_mcLine)over++;else if(tot<_mcLine)under++;else push++;
        }else if(_mcMode==='ML'){
          if(a>b)over++;else if(b>a)under++;else push++;
        }else{
          // RL / RL+: _mcP2=isAway(1=away picked), _mcP3=isPlus(1=+1.5, 0=-1.5)
          var margin=_mcP2?(a-b):(b-a);
          if(_mcP3){if(margin>=-1)over++;else under++;}
          else{if(margin>=2)over++;else under++;}
        }
      }
      done=end;prog.style.width=(done/N*100).toFixed(0)+'%';
      if(done<N){requestAnimationFrame(batch);return;}
      var pO=over/N,pU=under/N,pP=push/N;
      var ev,evSide,evOds;
      if(_mcMode==='TOT'){
        var evO=pO*_amPayout(_mcP2)-(1-pO),evU=pU*_amPayout(_mcP3)-(1-pU);
        ev=evO>=evU?evO:evU;evSide=evO>=evU?'Over':'Under';evOds=evO>=evU?_mcP2:_mcP3;
      }else if(_mcMode==='ML'){
        // _mcP1=awayOds, _mcP2=homeOds, _mcP3=isAway(1=away pick)
        if(_mcP3){ev=pO*_amPayout(_mcP1)-(1-pO);evSide='Away';evOds=_mcP1;}
        else{ev=pU*_amPayout(_mcP2)-(1-pU);evSide='Home';evOds=_mcP2;}
      }else{
        // RL / RL+: _mcP1=pickedOds, cover probability = pO
        ev=pO*_amPayout(_mcP1)-(1-pO);evSide=_mcP3?'+1.5':'-1.5';evOds=_mcP1;
      }
      var overEl=document.getElementById('mc-p-over');
      overEl.textContent=(pO*100).toFixed(1)+'%';overEl.style.color=pO>0.5?'#10b981':'#94a3b8';
      document.getElementById('mc-card-over').style.background=pO>0.5?'rgba(16,185,129,.12)':'rgba(16,185,129,.04)';
      var underEl=document.getElementById('mc-p-under');
      underEl.textContent=(pU*100).toFixed(1)+'%';underEl.style.color=pU>0.5?'#10b981':'#94a3b8';
      document.getElementById('mc-card-under').style.background=pU>0.5?'rgba(16,185,129,.12)':'rgba(124,58,237,.06)';
      document.getElementById('mc-p-push').textContent=(pP*100).toFixed(1)+'%';
      document.getElementById('mc-ev-odds').textContent=evSide+' \xb7 '+_odsFmt(evOds);
      var evEl=document.getElementById('mc-ev');
      evEl.textContent=(ev>=0?'+':'')+(ev*100).toFixed(1)+'%';
      evEl.style.color=ev>=0.08?'#10b981':ev>=0?'#f59e0b':'#ef4444';
      var maxA=0,maxH=0,maxC=0;
      for(var k in awayD)if(+k>maxA)maxA=+k;for(var k in homeD)if(+k>maxH)maxH=+k;for(var k in comboD)if(+k>maxC)maxC=+k;
      renderDist('mc-dist-away',awayD,N,Math.min(maxA,12),function(){return'#0ea5e9';});
      renderDist('mc-dist-home',homeD,N,Math.min(maxH,12),function(){return'#f97316';});
      if(_mcMode==='TOT'){
        var lineBin=Math.round(_mcLine);
        renderDist('mc-dist-combo',comboD,N,Math.min(maxC,20),function(i){return i>lineBin?'#10b981':i<lineBin?'#7c3aed':'#475569';});
      }else{
        renderDist('mc-dist-combo',comboD,N,Math.min(maxC,20),function(){return'#475569';});
      }
      btn.disabled=false;btn.textContent='↺ Re-run simulation';_mcRunning=false;
    }
    requestAnimationFrame(batch);
  };
  window.openMC=function(tA,tB,awayName,homeName,mode,p1,p2,p3){
    var modal=document.getElementById('mc-modal');if(!modal)return;
    _mcTA=tA;_mcTB=tB;_mcMode=mode||'TOT';_mcP1=p1;_mcP2=p2;_mcP3=p3;
    document.getElementById('mc-matchup').textContent=awayName+' @ '+homeName;
    document.getElementById('mc-lbl-away-stat').textContent=awayName+' proj.';
    document.getElementById('mc-lbl-home-stat').textContent=homeName+' proj.';
    document.getElementById('mc-lbl-away').textContent=awayName+' dist.';
    document.getElementById('mc-lbl-home').textContent=homeName+' dist.';
    document.getElementById('mc-tA-disp').textContent=parseFloat(tA).toFixed(2);
    document.getElementById('mc-tB-disp').textContent=parseFloat(tB).toFixed(2);
    document.getElementById('mc-lbl-combo').textContent='Combined total';
    if(mode==='ML'){
      // p1=awayOds, p2=homeOds, p3=isAway(1=away picked)
      document.getElementById('mc-lbl-stat3').textContent=awayName+' ML';
      document.getElementById('mc-line-disp').textContent=_odsFmt(p1);
      document.getElementById('mc-sub-stat3').textContent='away odds';
      document.getElementById('mc-lbl-stat4').textContent=homeName+' ML';
      document.getElementById('mc-model-disp').textContent=_odsFmt(p2);
      document.getElementById('mc-sub-stat4').textContent='home odds';
      document.getElementById('mc-lbl-result1').textContent='P(Away Win)';
      document.getElementById('mc-lbl-result2').textContent='P(Home Win)';
      document.getElementById('mc-lbl-result3').textContent='P(Tie)';
    }else if(mode==='RL'||mode==='RL+'){
      // p1=pickedOds, p2=isAway(1=away), p3=isPlus(1=+1.5, 0=-1.5)
      var pickedName=p2?awayName:homeName;
      var spreadStr=p3?'+1.5':'-1.5';
      document.getElementById('mc-lbl-stat3').textContent=pickedName+' RL';
      document.getElementById('mc-line-disp').textContent=spreadStr;
      document.getElementById('mc-sub-stat3').textContent=_odsFmt(p1);
      document.getElementById('mc-lbl-stat4').textContent='Model Total';
      document.getElementById('mc-model-disp').textContent=(parseFloat(tA)+parseFloat(tB)).toFixed(2);
      document.getElementById('mc-sub-stat4').textContent='proj. runs';
      document.getElementById('mc-lbl-result1').textContent='P(Cover '+spreadStr+')';
      document.getElementById('mc-lbl-result2').textContent='P(No Cover)';
      document.getElementById('mc-lbl-result3').textContent='P(Push)';
    }else{
      // TOT: p1=mktLine, p2=odsO, p3=odsU
      var noOdds=(!p1||p1<=0);
      _mcLine=noOdds?(tA+tB):p1;
      document.getElementById('mc-lbl-stat3').textContent='BetMGM Line';
      document.getElementById('mc-line-disp').textContent=noOdds?'⚠️ Re-run Lines':parseFloat(p1).toFixed(1);
      document.getElementById('mc-sub-stat3').textContent='market total';
      document.getElementById('mc-lbl-stat4').textContent='Model Total';
      document.getElementById('mc-model-disp').textContent=(parseFloat(tA)+parseFloat(tB)).toFixed(2);
      document.getElementById('mc-sub-stat4').textContent='projected runs';
      document.getElementById('mc-lbl-result1').textContent='P(Over)';
      document.getElementById('mc-lbl-result2').textContent='P(Under)';
      document.getElementById('mc-lbl-result3').textContent='P(Push)';
    }
    document.getElementById('mc-ev-odds').textContent='—';
    document.getElementById('mc-p-over').style.color='#10b981';
    document.getElementById('mc-p-under').style.color='#7c3aed';
    document.getElementById('mc-p-over').textContent='—';
    document.getElementById('mc-p-under').textContent='—';
    document.getElementById('mc-p-push').textContent='—';
    document.getElementById('mc-ev').textContent='—';
    document.getElementById('mc-ev').style.color='#a78bfa';
    document.getElementById('mc-prog').style.width='0%';
    document.getElementById('mc-dist-away').innerHTML='';
    document.getElementById('mc-dist-home').innerHTML='';
    document.getElementById('mc-dist-combo').innerHTML='';
    document.getElementById('mc-run-btn').disabled=false;
    document.getElementById('mc-run-btn').textContent='↺ Re-run simulation';
    modal.style.display='block';document.body.style.overflow='hidden';
    runMC();
  };
  window.closeMC=function(){
    var modal=document.getElementById('mc-modal');
    if(modal)modal.style.display='none';
    document.body.style.overflow='';
  };
  document.addEventListener('keydown',function(e){if(e.key==='Escape')closeMC();});
})();
</script>"""

JS = """
// Tab switching
function showTab(t){
  document.querySelectorAll('.tab').forEach(e=>e.classList.toggle('active',e.dataset.tab===t));
  document.querySelectorAll('.panel').forEach(e=>e.classList.toggle('active',e.id==='panel-'+t));
  localStorage.setItem('tab',t);
}
window.addEventListener('load',()=>{
  const saved=localStorage.getItem('tab')||'bsn';
  showTab(saved);
});

// Modal
function openModal(id){document.getElementById(id).classList.add('open');}
function closeModal(id){
  document.getElementById(id).classList.remove('open');
  const a=document.querySelector('#'+id+' .alert');
  if(a){a.remove();}
}
document.addEventListener('keydown',e=>{
  if(e.key==='Escape'){
    document.querySelectorAll('.modal-bg.open').forEach(m=>m.classList.remove('open'));
    closeDetail();
  }
});

// Output panel
function showOutput(title){
  document.getElementById('out-title').textContent=title;
  document.getElementById('out-body').innerHTML='<span class="spinner"></span>Running...';
  document.getElementById('out-bg').classList.add('open');
}
function closeOutput(){document.getElementById('out-bg').classList.remove('open');}
function setOutput(text){document.getElementById('out-body').textContent=text;}

// Run command
async function runCmd(cmd,cwd,title){
  showOutput(title||cmd);
  try{
    const r=await fetch('/api/run',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({cmd,cwd})});
    const d=await r.json();
    setOutput(d.out||'(sin salida)');
  }catch(e){setOutput('Error: '+e);}
}

// Form submit helper
async function submitForm(formId,endpoint,modalId){
  const form=document.getElementById(formId);
  const fd=new FormData(form);
  // Build URL-encoded body manually (avoids URLSearchParams(FormData) Safari bug)
  var pairs=[];
  fd.forEach(function(v,k){pairs.push(encodeURIComponent(k)+'='+encodeURIComponent(v));});
  var body=pairs.join('&');
  try{
    const r=await fetch(endpoint,{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},body:body});
    const d=await r.json();
    const old=form.querySelector('.alert');if(old)old.remove();
    const div=document.createElement('div');
    div.className='alert '+(d.ok?'ok':'err');
    div.textContent=d.msg;
    form.prepend(div);
    if(d.ok){form.reset();setTimeout(()=>location.reload(),1200);}
  }catch(e){alert('Error: '+e);}
}

// Grade pick
async function autoGrade(endpoint, modalId){
  if(!confirm('Auto-gradear picks pendientes jalando scores reales?')) return;
  try{
    const r=await fetch(endpoint,{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},body:'action=auto'});
    const d=await r.json();
    alert(d.msg);
    if(d.ok) location.reload();
  }catch(e){alert('Error: '+e);}
}

async function gradePick(endpoint, id, result, closingInputId){
  const labels={'W':'WIN','L':'LOSS','P':'PUSH'};
  if(!confirm('Marcar pick #'+id+' como '+labels[result]+'?')) return;
  let extra = '';
  if(closingInputId){
    const el = document.getElementById(closingInputId);
    if(el && el.value.trim()) extra += '&closing_line='+encodeURIComponent(el.value.trim());
    // actual_runs input (optional companion field)
    const el2 = document.getElementById(closingInputId.replace('cl-','ar-'));
    if(el2 && el2.value.trim()) extra += '&actual_runs='+encodeURIComponent(el2.value.trim());
  }
  try{
    const r=await fetch(endpoint,{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},
      body:'id='+id+'&result='+result+extra});
    const d=await r.json();
    alert(d.msg);
    if(d.ok) location.reload();
  }catch(e){alert('Error: '+e);}
}

// Add parlay leg
let legCount=1;
function addLeg(){
  legCount++;
  const c=document.getElementById('legs-container');
  const div=document.createElement('div');div.className='leg';div.id='leg-'+legCount;
  div.innerHTML=`<div class="leg-title">LEG ${legCount} <button type="button" onclick="removeLeg(${legCount})" style="float:right;background:#7f1d1d;color:#fca5a5;border:none;border-radius:4px;padding:1px 7px;cursor:pointer;font-size:0.7rem">✕</button></div>`+legHtml(legCount);
  c.appendChild(div);
}
function removeLeg(n){const el=document.getElementById('leg-'+n);if(el)el.remove();}
function legHtml(n,teams){
  return `<input name="leg${n}_game" placeholder="Juego (Ej: LEONES vs OSOS)" autocapitalize="characters" required>
<input name="leg${n}_pick" placeholder="Pick (Ej: LEONES ML)" autocapitalize="characters" required>`;
}

// BSN game management
async function bsnEditTime(away, home, newTime){
  if(!newTime){alert('Ingresa la hora');return;}
  try{
    const r=await fetch('/api/bsn/edit-game-time',{method:'POST',
      headers:{'Content-Type':'application/x-www-form-urlencoded'},
      body:'away='+encodeURIComponent(away)+'&home='+encodeURIComponent(home)+'&time='+encodeURIComponent(newTime)});
    const d=await r.json();
    alert(d.msg);
    if(d.ok) location.reload();
  }catch(e){alert('Error: '+e);}
}
async function bsnRemoveGame(away, home){
  if(!confirm('¿Remover juego '+away+' @ '+home+'?')) return;
  try{
    const r=await fetch('/api/bsn/remove-game',{method:'POST',
      headers:{'Content-Type':'application/x-www-form-urlencoded'},
      body:'away='+encodeURIComponent(away)+'&home='+encodeURIComponent(home)});
    const d=await r.json();
    alert(d.msg);
    if(d.ok) location.reload();
  }catch(e){alert('Error: '+e);}
}

// MLB debug game — shows in side panel
async function runDebugGame(away, home){
  openDetail('MLB · Debug: '+away+' @ '+home,
    '<div style="padding:48px;text-align:center"><span class="spinner"></span><div style="color:#475569;font-size:.72rem;margin-top:10px">Analizando matchup...</div></div>');
  try{
    const r=await fetch('/api/mlb/debug-game',{method:'POST',
      headers:{'Content-Type':'application/x-www-form-urlencoded'},
      body:'away='+encodeURIComponent(away)+'&home='+encodeURIComponent(home)});
    const d=await r.json();
    document.getElementById('detail-body').innerHTML=d.html||`<div class="terminal">${fmtOut(d.out||d.msg||'(sin salida)')}</div>`;
  }catch(e){document.getElementById('detail-body').innerHTML=`<div style="color:#ef4444;padding:20px">Error: ${e}</div>`;}
}

// MLB combined publish: export-picks + export-debug/lines en background
async function mlbPublishAll(){
  showOutput('MLB \u2192 Publish Picks + Debug Lines');
  document.getElementById('out-body').innerHTML='<span class="spinner"></span>Iniciando publish en background...';
  try{
    const r=await fetch('/api/mlb/publish-all',{method:'POST',
      headers:{'Content-Type':'application/x-www-form-urlencoded'},body:''});
    const d=await r.json();
    if(!d.task_id){
      // fallback: respuesta síncrona (no debería pasar)
      setOutput(d.out||d.msg||'(sin salida)');
      return;
    }
    const taskId=d.task_id;
    let dots=0;
    const poll=async()=>{
      try{
        const r2=await fetch('/api/task-status?id='+taskId);
        const s=await r2.json();
        if(s.status==='done'||s.status==='error'){
          setOutput(s.out||'(sin salida)');
        }else{
          dots=(dots+1)%4;
          const spinner='<span class="spinner"></span>';
          document.getElementById('out-body').innerHTML=spinner+'Publicando'+'.'.repeat(dots+1)+' (puede tardar 2-5 min)';
          setTimeout(poll,3000);
        }
      }catch(e){setOutput('Error polling: '+e);}
    };
    setTimeout(poll,3000);
  }catch(e){setOutput('Error: '+e);}
}

// MLB publish log en background (igual que mlbPublishAll pero para log picks)
async function mlbPublishLog(){
  showOutput('MLB → Publish Log');
  document.getElementById('out-body').innerHTML='<span class="spinner"></span>Publicando log en background...';
  try{
    const r=await fetch('/api/mlb/publish-log',{method:'POST',
      headers:{'Content-Type':'application/x-www-form-urlencoded'},body:''});
    const d=await r.json();
    if(!d.task_id){setOutput(d.msg||d.out||'(sin salida)');return;}
    const taskId=d.task_id;
    let dots=0;
    const poll=async()=>{
      try{
        const r2=await fetch('/api/task-status?id='+taskId);
        const s=await r2.json();
        if(s.status==='done'||s.status==='error'){
          setOutput(s.out||'(sin salida)');
        }else{
          dots=(dots+1)%4;
          document.getElementById('out-body').innerHTML='<span class="spinner"></span>Publicando'+'.'.repeat(dots+1);
          setTimeout(poll,3000);
        }
      }catch(e){setOutput('Error polling: '+e);}
    };
    setTimeout(poll,3000);
  }catch(e){setOutput('Error: '+e);}
}

// ── Detail Panel ──────────────────────────────────────
// iOS Safari fix: overflow:hidden alone doesn't stop body scroll on iPhone.
// We use position:fixed trick to truly lock the background.
var _panelScrollY=0;
function _lockBodyScroll(){
  _panelScrollY=window.scrollY||window.pageYOffset;
  document.body.style.position='fixed';
  document.body.style.top='-'+_panelScrollY+'px';
  document.body.style.left='0';
  document.body.style.right='0';
  document.body.style.overflow='hidden';
}
function _unlockBodyScroll(){
  document.body.style.position='';
  document.body.style.top='';
  document.body.style.left='';
  document.body.style.right='';
  document.body.style.overflow='';
  window.scrollTo(0,_panelScrollY);
}
function openDetail(title,html){
  document.getElementById('detail-title').textContent=title;
  document.getElementById('detail-body').innerHTML=html;
  document.getElementById('detail-panel').classList.add('open');
  document.getElementById('detail-overlay').classList.add('open');
  _lockBodyScroll();
}
function closeDetail(){
  document.getElementById('detail-panel').classList.remove('open');
  document.getElementById('detail-overlay').classList.remove('open');
  _unlockBodyScroll();
}

// Fetch a pre-rendered HTML view from the server
async function openView(endpoint,title){
  openDetail(title,'<div style="padding:48px;text-align:center"><span class="spinner"></span></div>');
  try{
    const r=await fetch(endpoint);
    const d=await r.json();
    const body=document.getElementById('detail-body');
    body.innerHTML=d.html||'<div class="detail-empty">Sin datos disponibles.</div>';
    // innerHTML doesn't execute <script> tags — re-create them so they run
    body.querySelectorAll('script').forEach(old=>{
      const s=document.createElement('script');
      s.textContent=old.textContent;
      old.parentNode.replaceChild(s,old);
    });
  }catch(e){document.getElementById('detail-body').innerHTML=`<div style="color:#ef4444;padding:20px">Error: ${e}</div>`;}
}

// Run a command and render output visually in the detail panel
async function runInView(cmd,cwd,title){
  openDetail(title,
    '<div style="padding:48px;text-align:center"><span class="spinner"></span><div style="color:#475569;font-size:.72rem;margin-top:12px">Ejecutando modelo...</div></div>');
  try{
    const r=await fetch('/api/run-view',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({cmd,cwd})});
    const d=await r.json();
    document.getElementById('detail-body').innerHTML=d.html||'<div class="detail-empty">Sin datos disponibles.</div>';
  }catch(e){document.getElementById('detail-body').innerHTML=`<div style="color:#ef4444;padding:20px">Error: ${e}</div>`;}
}

// Run a command (to write/refresh JSON), then open a view endpoint in the detail panel
async function runThenView(cmd, cwd, viewEndpoint, title){
  openDetail(title,
    '<div style="padding:48px;text-align:center"><span class="spinner"></span><div style="color:#475569;font-size:.72rem;margin-top:12px">Generando líneas...</div></div>');
  try{
    await fetch('/api/run',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({cmd,cwd})});
    const r=await fetch(viewEndpoint);
    const d=await r.json();
    const body=document.getElementById('detail-body');
    body.innerHTML=d.html||'<div class="detail-empty">Sin datos disponibles.</div>';
    // innerHTML doesn't execute <script> tags — re-create them so they run
    body.querySelectorAll('script').forEach(old=>{
      const s=document.createElement('script');
      s.textContent=old.textContent;
      old.parentNode.replaceChild(s,old);
    });
  }catch(e){document.getElementById('detail-body').innerHTML=`<div style="color:#ef4444;padding:20px">Error: ${e}</div>`;}
}

// Tools accordion toggle
function toggleTools(id){
  const body=document.getElementById('tools-'+id);
  const chev=document.getElementById('chev-'+id);
  if(!body) return;
  const open=body.classList.toggle('open');
  if(chev) chev.classList.toggle('open',open);
}

// Refresh picks block in-page (run model → update JSON → re-render from JSON)
async function refreshPicksBlock(blockId, modelCmd, cwd, viewEndpoint){
  const block=document.getElementById(blockId);
  if(!block) return;
  block.innerHTML='<div style="padding:38px;text-align:center"><span class="spinner"></span><div style="color:#475569;font-size:.72rem;margin-top:12px;letter-spacing:.05em">Corriendo modelo...</div></div>';
  try{
    // Step 1: run the model to regenerate picks JSON
    await fetch('/api/run',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({cmd:modelCmd,cwd:cwd})});
    // Step 2: fetch rendered view from the updated JSON (cache-busted)
    if(viewEndpoint){
      const r=await fetch(viewEndpoint,{cache:'no-store'});
      const d=await r.json();
      block.innerHTML=d.html||'<div class="today-empty">Sin picks generados hoy.</div>';
    } else {
      // For leagues with no view endpoint, re-fetch rendered output
      const r=await fetch('/api/run-view',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({cmd:modelCmd,cwd:cwd})});
      const d=await r.json();
      block.innerHTML=d.html||'<div class="today-empty">Sin picks.</div>';
    }
  }catch(e){
    block.innerHTML='<div class="today-empty" style="color:#ef4444">Error al correr el modelo.<br><span style="font-size:.7rem">'+e+'</span></div>';
  }
}

// Color-code terminal output
function fmtOut(text){
  let s=text.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  // Highlight positive/negative money values
  s=s.replace(/(\+\$[\d,.]+)/g,'<span class="t-ok">$1</span>');
  s=s.replace(/(\-\$[\d,.]+)/g,'<span class="t-err">$1</span>');
  // Highlight EV+ and W/L markers in text
  s=s.replace(/(EV\+\S*)/g,'<span class="t-hi">$1</span>');
  s=s.replace(/(\bWIN\b|\bW \d|\d-\d)/g,'<span class="t-ok">$1</span>');
  // Dim separator lines
  s=s.replace(/(^[=\-]{4,}.*$)/gm,'<span class="t-dim">$1</span>');
  return s;
}

// ── Record Date Picker ─────────────────────────────────────────────────
var _recScript = '';
var _recCwd    = '';
var _recAccent = '#f07820';

const _CHIP_IDS  = ['chip-hoy','chip-ayer','chip-d2','chip-d3','chip-d4','chip-d5'];
const _CHIP_OFFS = [0,-1,-2,-3,-4,-5];

function _recDateOf(offset){
  const d = new Date(); d.setDate(d.getDate()+offset);
  return d.toISOString().slice(0,10);
}
function _recFmtChip(offset){
  if(offset===0)  return 'Hoy';
  if(offset===-1) return 'Ayer';
  const d = new Date(); d.setDate(d.getDate()+offset);
  return d.toLocaleDateString('es',{weekday:'short',day:'numeric'}).replace('.',', ');
}
function recSetDate(offset){
  document.getElementById('rec-date').value = _recDateOf(offset);
  _CHIP_IDS.forEach((id,i)=>{
    document.getElementById(id).classList.toggle('active', _CHIP_OFFS[i]===offset);
  });
}
function openRecordModal(league, script, cwd, accent){
  _recScript = script;
  _recCwd    = cwd;
  _recAccent = accent || '#f07820';
  document.getElementById('rec-accent-bar').style.background = _recAccent;
  document.getElementById('rec-league-label').style.color    = _recAccent;
  document.getElementById('rec-league-label').textContent    = league + ' · Record Card';
  document.getElementById('rec-submit-btn').style.background = _recAccent;
  document.documentElement.style.setProperty('--rec-accent', _recAccent);
  _CHIP_IDS.forEach((id,i)=>{
    document.getElementById(id).textContent = _recFmtChip(_CHIP_OFFS[i]);
  });
  recSetDate(-1);
  openModal('modal-record');
}
function submitRecordModal(){
  const date = document.getElementById('rec-date').value;
  if(!date){ alert('Selecciona una fecha'); return; }
  closeModal('modal-record');
  const cmd = 'python3 ' + _recScript + ' --export-record ' + date + ' --publish';
  runCmd(cmd, _recCwd, _recCwd + ' → Record ' + date);
}

// ── BSN Set Lines ─────────────────────────────────────────────────────
function _bslV(id){ const el=document.getElementById(id); return el?el.value.trim():''; }

function _bsnLineCard(g, i){
  const ex = g.existing || {};
  const hasSaved = ex.ml1 || ex.total;
  const badge = hasSaved
    ? '<span class="bsl-saved-badge">✓ líneas guardadas</span>'
    : '';
  // populate existing or blank
  const v = (field, def) => ex[field]!=null ? ex[field] : (def||'');
  // spread fav select options
  const t1 = g.team1 || ''; const t2 = g.team2 || '';
  const selFav = (team) => `<option value="${t1}"${v('spread_fav')==t1?' selected':''}>` + t1 + `</option>` +
                            `<option value="${t2}"${v('spread_fav')==t2?' selected':''}>` + t2 + `</option>`;
  return `
<div class="bsl-card" id="bsl-card-${i}">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:11px">
    <div style="font-size:.78rem;font-weight:800;color:#e2e8f0;letter-spacing:.04em">${t1} <span style="color:#475569;font-weight:500">vs</span> ${t2}</div>
    ${badge}
  </div>

  <div class="bsl-section-lbl">💵 Moneyline</div>
  <div class="bsl-2col">
    <div>
      <span class="bsl-field-lbl">${t1} ML</span>
      <input id="bsl-${i}-ml1" placeholder="-115" value="${v('ml1')}">
    </div>
    <div>
      <span class="bsl-field-lbl">${t2} ML</span>
      <input id="bsl-${i}-ml2" placeholder="+105" value="${v('ml2')}">
    </div>
  </div>

  <div class="bsl-section-lbl">📊 Spread</div>
  <div class="bsl-3col">
    <div>
      <span class="bsl-field-lbl">Favorito</span>
      <select id="bsl-${i}-sfav">
        <option value="">—</option>
        <option value="${t1}"${v('spread_fav')==t1?' selected':''}>${t1}</option>
        <option value="${t2}"${v('spread_fav')==t2?' selected':''}>${t2}</option>
      </select>
    </div>
    <div>
      <span class="bsl-field-lbl">Línea</span>
      <input id="bsl-${i}-sline" placeholder="-1.5" value="${v('spread_line')}">
    </div>
    <div>
      <span class="bsl-field-lbl">Odds Fav</span>
      <input id="bsl-${i}-sodds" placeholder="-110" value="${v('spread_odds')}">
    </div>
  </div>
  <div style="margin-bottom:10px">
    <span class="bsl-field-lbl">Odds Perro</span>
    <input id="bsl-${i}-sdog" placeholder="-130" value="${v('spread_dog_odds')}">
  </div>

  <div class="bsl-section-lbl">🎯 Total (O/U)</div>
  <div class="bsl-3col">
    <div>
      <span class="bsl-field-lbl">Línea</span>
      <input id="bsl-${i}-tot" placeholder="182.5" value="${v('total')}">
    </div>
    <div>
      <span class="bsl-field-lbl">Over odds</span>
      <input id="bsl-${i}-over" placeholder="-110" value="${v('over_odds')}">
    </div>
    <div>
      <span class="bsl-field-lbl">Under odds</span>
      <input id="bsl-${i}-under" placeholder="-110" value="${v('under_odds')}">
    </div>
  </div>
</div>`;
}

// ── BSN Log Pick — game-aware modal ──────────────────────────────────────
let _bsnLogGames = [];

function openBsnLogPick(){
  // Reset form
  const sel = document.getElementById('bsn-log-game');
  if(sel) sel.value = '';
  const t1h = document.getElementById('bsn-log-t1');
  const t2h = document.getElementById('bsn-log-t2');
  if(t1h) t1h.value = '';
  if(t2h) t2h.value = '';
  const pi = document.getElementById('bsn-log-pick-inp');
  if(pi) pi.value = '';
  const oi = document.getElementById('bsn-log-odds-inp');
  if(oi) oi.value = '';
  const manual = document.getElementById('bsn-log-manual');
  if(manual) manual.style.display = 'none';
  const optsEl = document.getElementById('bsn-log-pick-opts');
  if(optsEl){ optsEl.innerHTML=''; optsEl.style.display='none'; }
  openModal('bsn-log');
}

function bsnLogFillGame(sel){
  const val = sel.value;
  const t1h = document.getElementById('bsn-log-t1');
  const t2h = document.getElementById('bsn-log-t2');
  const manual = document.getElementById('bsn-log-manual');
  const optsEl = document.getElementById('bsn-log-pick-opts');
  if(val === '__manual__'){
    if(manual) manual.style.display = '';
    if(t1h) t1h.value = '';
    if(t2h) t2h.value = '';
    if(optsEl) optsEl.innerHTML = '';
    return;
  }
  if(manual) manual.style.display = 'none';
  const [t1, t2] = (val||'').split('|');
  if(t1h) t1h.value = t1||'';
  if(t2h) t2h.value = t2||'';
  // Suggest pick options — ML, Spread, Totals
  if(optsEl && t1 && t2){
    const _p = (pick, label, accent, suffix) =>
      `<span style="display:inline-block;cursor:pointer;padding:5px 10px;border-radius:8px;
        font-size:.62rem;font-weight:700;letter-spacing:.04em;border:1px solid ${accent}40;
        color:${accent};background:${accent}12;transition:all .15s;font-family:inherit"
        onmouseover="this.style.background='${accent}28'"
        onmouseout="this.style.background='${accent}12'"
        onclick="bsnLogSetPick('${pick.replace(/'/g,"\\'")}','')">${label||pick}${suffix||''}</span>`;
    const _row = (lbl, ...pills) =>
      `<div style="display:flex;gap:5px;flex-wrap:wrap;align-items:center;margin-bottom:5px">
        <span style="font-size:.44rem;font-weight:900;color:#475569;letter-spacing:.1em;min-width:28px;flex-shrink:0">${lbl}</span>
        ${pills.join('')}
      </div>`;
    optsEl.innerHTML =
      _row('ML',  _p(`${t1} ML`,  `${t1} ML`,  '#4f8ef7'),
                  _p(`${t2} ML`,  `${t2} ML`,  '#4f8ef7')) +
      _row('SPR', _p(`${t1} -`,   `${t1}`,     '#f07820', ' <span style="opacity:.55">-▸</span>'),
                  _p(`${t2} -`,   `${t2}`,     '#f07820', ' <span style="opacity:.55">-▸</span>')) +
      _row('TOT', _p('O ', 'OVER',  '#f97316'),
                  _p('U ', 'UNDER', '#a78bfa'));
    optsEl.style.display='block';
    optsEl.style.marginBottom='8px';
  }
}

function bsnLogSetPick(pick, odds){
  const pi = document.getElementById('bsn-log-pick-inp');
  if(pi){ pi.value = pick; pi.focus(); }
  const oi = document.getElementById('bsn-log-odds-inp');
  if(oi && odds){ oi.value = odds; }
}

// ── BSN IR ─────────────────────────────────────────
const _RATE_FACTOR = {1:0.70, 2:0.75, 3:0.80};
const _RATE_COLOR  = {1:'#ef4444', 2:'#fbbf24', 3:'#94a3b8'};
const _RATE_LBL    = {1:'OUT', 2:'Doubtful', 3:'Limited'};

async function openBsnIrModal(){
  openModal('bsn-ir');
  await _bsnIrLoadList();
  // Live impact preview
  ['bsn-ir-ppg','bsn-ir-usg'].forEach(id => {
    const el = document.getElementById(id);
    if(el) el.addEventListener('input', _bsnIrPreview);
  });
  document.querySelector('#f-bsn-ir select[name="rate"]')
    ?.addEventListener('change', _bsnIrPreview);
}

function _bsnIrPreview(){
  const ppg  = parseFloat(document.getElementById('bsn-ir-ppg')?.value) || 0;
  const usg  = parseFloat(document.getElementById('bsn-ir-usg')?.value) || 0;
  const rate = parseInt(document.querySelector('#f-bsn-ir select[name="rate"]')?.value) || 2;
  const rf   = _RATE_FACTOR[rate] || 0.75;
  const imp  = ppg * (usg/100) * rf;
  const el   = document.getElementById('bsn-ir-impact');
  if(el && ppg && usg)
    el.innerHTML = `📊 Impact estimado: <b style="color:#f5a623">${imp.toFixed(2)} pts</b> (${ppg.toFixed(1)} × ${usg.toFixed(1)}% × ${rf})`;
  else if(el) el.innerHTML = '';
}

async function _bsnIrLoadList(){
  const el = document.getElementById('bsn-ir-list');
  if(!el) return;
  el.innerHTML = '<div style="font-size:.55rem;color:#475569;text-align:center;padding:8px">Cargando...</div>';
  try{
    const r = await fetch('/api/bsn/ir-list');
    const d = await r.json();
    const entries = d.entries || [];
    if(!entries.length){
      el.innerHTML = '<div style="font-size:.6rem;color:#334155;text-align:center;padding:10px 0">Sin jugadores en IR</div>';
      return;
    }
    // Sort by rate then team
    entries.sort((a,b) => a.rate - b.rate || a.team.localeCompare(b.team));
    const rows = entries.map(e => {
      const rc = _RATE_COLOR[e.rate] || '#94a3b8';
      const rl = _RATE_LBL[e.rate] || '?';
      const imp = e.impact ? e.impact.toFixed(2) : '—';
      return `<div style="display:flex;align-items:center;gap:8px;padding:7px 10px;
        background:rgba(255,255,255,.03);border-radius:8px;margin-bottom:4px;
        border:1px solid rgba(255,255,255,.06)">
        <span style="font-size:.48rem;font-weight:900;padding:2px 6px;border-radius:4px;
          background:${rc}22;color:${rc};border:1px solid ${rc}44;min-width:50px;text-align:center">${rl}</span>
        <div style="flex:1;min-width:0">
          <div style="font-size:.58rem;font-weight:900;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${e.player}</div>
          <div style="font-size:.5rem;color:#475569">${e.team} · ${e.ppg.toFixed(1)} PPG · ${(e.usg*100).toFixed(1)}% USG</div>
        </div>
        <div style="text-align:right">
          <div style="font-size:.52rem;color:#f5a623;font-weight:700">${imp} pts</div>
          <button onclick="_bsnIrPrefill('${e.team}','${e.player}',${e.rate},${e.ppg},${(e.usg*100).toFixed(1)})"
            style="font-size:.45rem;padding:2px 6px;border-radius:4px;background:rgba(239,68,68,.1);
            color:#ef4444;border:1px solid rgba(239,68,68,.3);cursor:pointer;margin-top:2px">Editar / 🗑</button>
        </div>
      </div>`;
    }).join('');
    el.innerHTML = `<div style="font-size:.52rem;font-weight:900;letter-spacing:.1em;text-transform:uppercase;
      color:#334155;margin-bottom:6px">Jugadores en IR (${entries.length})</div>${rows}
      <div style="height:1px;background:rgba(255,255,255,.06);margin:10px 0"></div>`;
  }catch(e){
    el.innerHTML = `<div style="font-size:.6rem;color:#ef4444;padding:6px">Error: ${e}</div>`;
  }
}

function _bsnIrPrefill(team, player, rate, ppg, usg){
  const f = document.getElementById('f-bsn-ir');
  if(!f) return;
  f.querySelector('select[name="team"]').value  = team;
  f.querySelector('input[name="player"]').value = player;
  f.querySelector('select[name="rate"]').value  = String(rate);
  document.getElementById('bsn-ir-ppg').value   = ppg;
  document.getElementById('bsn-ir-usg').value   = usg;
  _bsnIrPreview();
}

function _bsnIrRemove(){
  document.getElementById('bsn-ir-action').value = 'remove';
  submitForm('f-bsn-ir', '/api/bsn/ir', 'bsn-ir');
}

function _bsnIrSubmit(){
  document.getElementById('bsn-ir-action').value = 'add';
  submitForm('f-bsn-ir', '/api/bsn/ir', 'bsn-ir');
}

let _bslGames = [];

async function openBsnLines(){
  document.getElementById('bsn-lines-body').innerHTML =
    '<div style="text-align:center;padding:30px;color:#475569">Cargando juegos...</div>';
  openModal('bsn-lines');
  try{
    const r = await fetch('/api/bsn/games-today');
    const d = await r.json();
    _bslGames = d.games || [];
    if(!_bslGames.length){
      document.getElementById('bsn-lines-body').innerHTML =
        '<div style="text-align:center;padding:30px;color:#475569">Sin juegos programados para hoy.<br><span style="font-size:.7rem">Agrega juegos primero con el botón 📅.</span></div>';
      return;
    }
    document.getElementById('bsn-lines-body').innerHTML =
      _bslGames.map((g,i) => _bsnLineCard(g,i)).join('');
  }catch(e){
    document.getElementById('bsn-lines-body').innerHTML =
      '<div style="text-align:center;padding:30px;color:#ef4444">Error cargando juegos: '+e+'</div>';
  }
}

async function submitBsnLines(){
  const entries = _bslGames.map((g,i) => ({
    team1:          g.team1,
    team2:          g.team2,
    ml1:            _bslV('bsl-'+i+'-ml1'),
    ml2:            _bslV('bsl-'+i+'-ml2'),
    spread_fav:     _bslV('bsl-'+i+'-sfav'),
    spread_line:    _bslV('bsl-'+i+'-sline'),
    spread_odds:    _bslV('bsl-'+i+'-sodds'),
    spread_dog_odds:_bslV('bsl-'+i+'-sdog'),
    total:          _bslV('bsl-'+i+'-tot'),
    over_odds:      _bslV('bsl-'+i+'-over'),
    under_odds:     _bslV('bsl-'+i+'-under'),
  })).filter(e => e.ml1||e.ml2||e.spread_fav||e.total);

  if(!entries.length){
    alert('Ingresa al menos una línea antes de guardar.');
    return;
  }
  const btn = document.querySelector('#bsn-lines .modal button[onclick="submitBsnLines()"]');
  if(btn){ btn.disabled=true; btn.textContent='Guardando...'; }
  try{
    const r = await fetch('/api/bsn/lines',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({entries})
    });
    const d = await r.json();
    if(btn){ btn.disabled=false; btn.textContent='✅ GUARDAR TODO'; }
    if(d.ok){
      closeModal('bsn-lines');
      // show toast
      const t=document.createElement('div');
      t.style.cssText='position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:#059669;color:#fff;padding:10px 22px;border-radius:99px;font-size:.82rem;font-weight:700;z-index:9999;box-shadow:0 4px 20px rgba(0,0,0,.4)';
      t.textContent=d.msg;
      document.body.appendChild(t);
      setTimeout(()=>t.remove(),2800);
    } else {
      alert(d.msg||'Error guardando líneas');
    }
  }catch(e){
    if(btn){ btn.disabled=false; btn.textContent='✅ GUARDAR TODO'; }
    alert('Error: '+e);
  }
}
"""

def _stats_bar(log_path, color, league_name="", accent="#f07820", glow="rgba(240,120,32,.1)"):
    log     = _rj(log_path)
    settled = [e for e in log if e.get("result") in ("W","L","P")]
    w       = sum(1 for e in settled if e.get("result")=="W")
    l       = sum(1 for e in settled if e.get("result")=="L")
    p       = sum(1 for e in settled if e.get("result")=="P")
    pend    = len([e for e in log if not e.get("result")])
    total_pnl   = sum(e.get("pnl",0) or 0 for e in settled)
    total_wager = sum(_entry_stake(e) for e in settled)
    roi     = (total_pnl / total_wager * 100) if total_wager > 0 else 0.0
    decided = w + l
    pct     = w / decided * 100 if decided else 0.0

    wc  = "#22c55e" if pct >= 55 else ("#86efac" if pct >= 50 else ("#f59e0b" if pct >= 45 else "#ef4444"))
    pc  = "#22c55e" if total_pnl >= 0 else "#ef4444"
    rc  = "#22c55e" if roi >= 0 else "#ef4444"
    pnl_s = f"+${total_pnl:,.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):,.2f}"

    push_frag    = f'<span class="r-push">P{p}</span>' if p else ""
    pending_frag = f'<span class="hero-pending">⏳ {pend} pending</span>' if pend else ""

    return f"""<div class="lg-hero" style="--acc:{accent};--glow:{glow}">
  <div class="hero-top">
    <span class="hero-league-lbl">{league_name} · All-Time Record</span>
    {pending_frag}
  </div>
  <div class="hero-rec">
    <span class="r-big r-w">{w}</span>
    <span class="r-sep">—</span>
    <span class="r-big r-l">{l}</span>
    {push_frag}
  </div>
  <div class="stats-4">
    <div class="s4-cell">
      <div class="s4-val" style="color:{wc}">{pct:.1f}%</div>
      <div class="s4-lbl">Win Rate</div>
    </div>
    <div class="s4-cell">
      <div class="s4-val" style="color:{pc}">{pnl_s}</div>
      <div class="s4-lbl">P&amp;L</div>
    </div>
    <div class="s4-cell">
      <div class="s4-val" style="color:{rc}">{roi:+.1f}%</div>
      <div class="s4-lbl">ROI</div>
    </div>
    <div class="s4-cell">
      <div class="s4-val" style="color:#64748b">{w+l+p}</div>
      <div class="s4-lbl">Total Picks</div>
    </div>
  </div>
</div>"""


def _cmd_card(desc, btn_label, onclick, btn_class=""):
    """Render a simple action card with description and button."""
    return f"""<div class="cmd-card">
  <span class="cmd-desc">{desc}</span>
  <button class="btn {btn_class}" onclick="{onclick}">{btn_label}</button>
</div>"""


# ── Team badge helpers ─────────────────────────────────────────────────
# ── ESPN CDN logo slugs ───────────────────────────────────────────────
# MLB: https://a.espncdn.com/i/teamlogos/mlb/500/{slug}.png
_MLB_ESPN = {
    "ARI":"ari","ATL":"atl","BAL":"bal","BOS":"bos","CHC":"chc",
    "CWS":"cws","CIN":"cin","CLE":"cle","COL":"col","DET":"det",
    "HOU":"hou","KC":"kc","LAA":"laa","LAD":"lad","MIA":"mia",
    "MIL":"mil","MIN":"min","NYM":"nym","NYY":"nyy","OAK":"oak",
    "PHI":"phi","PIT":"pit","SD":"sd","SF":"sf","SEA":"sea",
    "STL":"stl","TB":"tb","TEX":"tex","TOR":"tor","WSH":"wsh",
}
# NBA: https://a.espncdn.com/i/teamlogos/nba/500/{slug}.png
_NBA_ESPN = {
    "ATL":"atl","BOS":"bos","BKN":"bkn","CHA":"cha","CHI":"chi",
    "CLE":"cle","DAL":"dal","DEN":"den","DET":"det","GSW":"gs",
    "HOU":"hou","IND":"ind","LAC":"lac","LAL":"lal","MEM":"mem",
    "MIA":"mia","MIL":"mil","MIN":"min","NOP":"no","NYK":"ny",
    "OKC":"okc","ORL":"orl","PHI":"phi","PHX":"phx","POR":"por",
    "SAC":"sac","SAS":"sa","TOR":"tor","UTA":"utah","WAS":"wsh",
}

def _logo_img(url, abb, bg, fg, size=32):
    """Return a logo circle: colored ring → white inner circle → ESPN CDN logo.
    White inner background ensures visibility for all team colors (Dodgers, Braves, etc.)
    onerror uses DOM methods only — no nested quotes that would break the HTML parser.
    """
    ring  = size
    inner = max(int(size * .78), size - 6)   # white circle inside the ring
    img   = max(int(size * .62), size - 10)  # logo image inside the white circle
    fs    = int(size * .28)
    safe  = _esc(abb)
    # onerror: hide img, reset inner bg to team color, inject text node via DOM (no inner quotes)
    onerror = (
        f"var p=this.parentElement;"
        f"p.style.background='{bg}';"
        f"this.style.display='none';"
        f"var s=document.createElement('span');"
        f"s.style.cssText='font-size:{fs}px;font-weight:900;color:{fg};line-height:1';"
        f"s.textContent='{safe}';"
        f"p.appendChild(s)"
    )
    return (
        f'<span style="display:inline-flex;align-items:center;justify-content:center;'
        f'flex-shrink:0;width:{ring}px;height:{ring}px;border-radius:50%;'
        f'background:{bg};border:2px solid rgba(255,255,255,.18)">'
        f'<span style="display:flex;align-items:center;justify-content:center;'
        f'width:{inner}px;height:{inner}px;border-radius:50%;background:#fff;overflow:hidden">'
        f'<img src="{url}" width="{img}" height="{img}" '
        f'style="object-fit:contain;display:block" '
        f'onerror="{onerror}" '
        f'alt="{safe}">'
        f'</span>'
        f'</span>'
    )

def _badge_mlb(team, size=32):
    key = team.upper()
    bg, fg = _MLB_COLORS.get(key, ("#1e293b","#94a3b8"))
    abb = _MLB_ABB.get(key, key[:3])
    slug = _MLB_ESPN.get(abb, abb.lower())
    url  = f"https://a.espncdn.com/i/teamlogos/mlb/500/{slug}.png"
    return _logo_img(url, abb, bg, fg, size)

_BSN_LOGO_CACHE = {}  # team_lower → "data:image/png;base64,..." or None

def _strip_accents_bsn(s):
    """Normaliza acentos: 'ATLÉTICOS' → 'ATLETICOS'"""
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")

def _bsn_logo_b64_url(team):
    """Return base64 data-URL for a BSN team logo, or None if not found."""
    # Try accent-stripped version first (handles ATLÉTICOS → atleticos.png)
    key_raw   = team.lower()
    key_ascii = _strip_accents_bsn(key_raw)
    for key in dict.fromkeys([key_ascii, key_raw]):  # try stripped first, then original
        if key not in _BSN_LOGO_CACHE:
            import base64 as _b64
            logo_path = os.path.join(BSN_DIR, "logos", f"{key}.png")
            if os.path.exists(logo_path):
                try:
                    with open(logo_path, "rb") as _lf:
                        _BSN_LOGO_CACHE[key] = "data:image/png;base64," + _b64.b64encode(_lf.read()).decode()
                except Exception:
                    _BSN_LOGO_CACHE[key] = None
            else:
                _BSN_LOGO_CACHE[key] = None
        if _BSN_LOGO_CACHE.get(key):
            return _BSN_LOGO_CACHE[key]
    return None

def _badge_bsn(team, size=32):
    key = team.upper()
    bg, fg = _BSN_COLORS.get(key, ("#1e293b","#94a3b8"))
    abb = key[:3]
    logo_url = _bsn_logo_b64_url(key)
    if logo_url:
        # Use actual logo with team-color ring background
        inner = max(int(size * .78), size - 6)
        return (f'<span style="width:{size}px;height:{size}px;border-radius:50%;'
                f'background:{bg};display:inline-flex;align-items:center;'
                f'justify-content:center;flex-shrink:0;overflow:hidden">'
                f'<img src="{logo_url}" alt="{_esc(abb)}" '
                f'width="{inner}" height="{inner}" style="object-fit:contain">'
                f'</span>')
    # Fallback: colored text badge
    return (f'<span class="team-badge" style="background:{bg};color:{fg};'
            f'width:{size}px;height:{size}px;border-radius:50%;padding:0;'
            f'display:inline-flex;align-items:center;justify-content:center;'
            f'font-size:{int(size*.31)}px;flex-shrink:0">{_esc(abb)}</span>')

def _stats_logo_nba(abb, size=36):
    """Bare logo img for stats tables — no ring, no circle."""
    slug = _NBA_ESPN.get(abb.upper(), abb.lower())
    url  = f"https://a.espncdn.com/i/teamlogos/nba/500/{slug}.png"
    return (f'<img src="{url}" alt="{_esc(abb)}" width="{size}" height="{size}" '
            f'style="object-fit:contain;flex-shrink:0" '
            f'onerror="this.style.visibility=\'hidden\'">')

def _stats_logo_bsn(team, size=36):
    """Bare logo img for stats tables — no ring, no circle."""
    logo_url = _bsn_logo_b64_url(team.upper())
    if logo_url:
        return (f'<img src="{logo_url}" alt="{_esc(team)}" width="{size}" height="{size}" '
                f'style="object-fit:contain;flex-shrink:0">')
    bg, fg = _BSN_COLORS.get(team.upper(), ("#1e293b","#94a3b8"))
    return (f'<span style="width:{size}px;height:{size}px;border-radius:50%;'
            f'background:{bg};color:{fg};display:inline-flex;align-items:center;'
            f'justify-content:center;font-size:{int(size*.35)}px;font-weight:900;'
            f'flex-shrink:0">{_esc(team[:3].upper())}</span>')

def _stats_logo_mlb(team, size=36):
    """Bare logo img for stats tables — no ring, no circle."""
    key  = team.upper()
    abb  = _MLB_ABB.get(key, key[:3])
    slug = _MLB_ESPN.get(abb, abb.lower())
    url  = f"https://a.espncdn.com/i/teamlogos/mlb/500/{slug}.png"
    return (f'<img src="{url}" alt="{_esc(abb)}" width="{size}" height="{size}" '
            f'style="object-fit:contain;flex-shrink:0" '
            f'onerror="this.style.visibility=\'hidden\'">')

_OU_ICON_CACHE = {}   # "" → "data:image/png;base64,..." or None

def _ou_icon_b64_url():
    """Return base64 data-URL for over_under.png, or None if not found."""
    if "" not in _OU_ICON_CACHE:
        import base64 as _b64
        # over_under.png lives in the MLB folder
        p = os.path.join(MLB_DIR, "over_under.png")
        if os.path.exists(p):
            try:
                with open(p, "rb") as f:
                    _OU_ICON_CACHE[""] = "data:image/png;base64," + _b64.b64encode(f.read()).decode()
            except Exception:
                _OU_ICON_CACHE[""] = None
        else:
            _OU_ICON_CACHE[""] = None
    return _OU_ICON_CACHE.get("")

def _badge_nba(team_abb, size=32):
    key = team_abb.upper()
    bg, fg = _NBA_COLORS.get(key, ("#1e293b","#94a3b8"))
    slug = _NBA_ESPN.get(key, key.lower())
    url  = f"https://a.espncdn.com/i/teamlogos/nba/500/{slug}.png"
    return _logo_img(url, key, bg, fg, size)


def _act(icon, label, onclick, cls="", big=False):
    """Render a single action icon button (legacy)."""
    extra = " big" if big else ""
    return (f'<button class="act-btn {cls}{extra}" onclick="{onclick}">'
            f'<span class="act-icon">{icon}</span>'
            f'<span class="act-lbl">{label}</span>'
            f'</button>')


def _pa_cta(icon, label, onclick, cls="green"):
    """Primary CTA button for new panel layout."""
    return (f'<button class="pa-cta {cls}" onclick="{onclick}">'
            f'<span class="pa-cta-icon">{icon}</span>{label}</button>')


def _pa_item(icon, label, onclick):
    """Secondary icon-grid item for new panel layout."""
    return (f'<button class="pa-item" onclick="{onclick}">'
            f'<span class="pa-item-icon">{icon}</span>'
            f'<span class="pa-item-lbl">{label}</span></button>')


def _panel_actions(primary, secondary):
    """
    Render the new structured panel actions block.
    primary  = list of (icon, label, onclick, cls) — max 2, shown as big CTAs
    secondary = list of (icon, label, onclick) — shown as 4-col icon grid
    """
    pri_html = "".join(_pa_cta(i, l, o, c) for i, l, o, c in primary)
    # pad secondary to multiples of 4
    items = list(secondary)
    while len(items) % 4 != 0:
        items.append(None)
    grid_html = ""
    for item in items:
        if item:
            grid_html += _pa_item(item[0], item[1], item[2])
        else:
            grid_html += '<div></div>'
    return (f'<div class="pa-wrap">'
            f'<div class="pa-primary">{pri_html}</div>'
            f'<div class="pa-grid">{grid_html}</div>'
            f'</div>')


# ── Today sections (rendered server-side into the panel) ───────────────
def _mlb_today_section():
    """Game cards for today's MLB picks. Shows rich debug fragment if available."""
    today_str = date.today().strftime("%Y-%m-%d")

    # Use the HTML fragment ONLY if it's from today AND is at least as recent
    # as mlb_debug_state.json (the source of truth for the latest model run).
    frag_path        = os.path.join(MLB_DIR, "mlb_debug_body_current.html")
    debug_state_path = os.path.join(MLB_DIR, "mlb_debug_state.json")
    mlb_py_path      = os.path.join(MLB_DIR, "mlb.py")
    if os.path.exists(frag_path):
        try:
            from datetime import datetime as _dt2
            frag_mtime    = os.path.getmtime(frag_path)
            frag_date     = _dt2.fromtimestamp(frag_mtime).strftime("%Y-%m-%d")
            state_mtime   = os.path.getmtime(debug_state_path) if os.path.exists(debug_state_path) else 0
            mlb_py_mtime  = os.path.getmtime(mlb_py_path) if os.path.exists(mlb_py_path) else 0
            serve_py_path2 = os.path.join(os.path.dirname(MLB_DIR), "serve.py")
            serve_mtime2  = os.path.getmtime(serve_py_path2) if os.path.exists(serve_py_path2) else 0
            # Only trust the HTML if it's from today, at least as fresh as the state JSON,
            # AND neither mlb.py nor serve.py has been updated since (catches code fixes mid-day)
            if frag_date == today_str and frag_mtime >= state_mtime and frag_mtime >= mlb_py_mtime and frag_mtime >= serve_mtime2:
                with open(frag_path, "r", encoding="utf-8") as f:
                    return f.read()
        except Exception:
            pass

    # Prefer mlb_debug_state.json (latest model run) over mlb_model_picks.json
    today_p = []
    state_data = _rj(debug_state_path)
    if isinstance(state_data, dict) and state_data.get("date","") == today_str:
        today_p = state_data.get("picks", [])
    if not today_p:
        picks = _rj(MLB_PICKS)
        today_p = [p for p in picks if p.get("date","") == today_str]

    if not today_p:
        return ('<div class="today-empty">'
                'Sin picks del modelo para hoy.<br>'
                '<span style="font-size:.7rem;color:#475569">'
                'Corre el modelo con ⚙ Regenerar para generar picks.</span>'
                '</div>')

    # Add stale-fragment notice banner
    try:
        from datetime import datetime as _dt2b
        _st = _dt2b.fromtimestamp(os.path.getmtime(debug_state_path)).strftime("%H:%M")
        _stale_notice = (
            f'<div style="font-size:.63rem;color:#f59e0b;background:#1c1500;border-radius:6px;'
            f'padding:6px 12px;margin-bottom:12px;letter-spacing:.03em">'
            f'⚠️ Vista básica · última corrida {_st}. '
            f'Presiona <b>⚙ Regenerar</b> para ver cards completas con clima y análisis.</div>'
        )
    except Exception:
        _stale_notice = ""

    # Group by game
    games_map = {}
    for p in today_p:
        g = p.get("game","?")
        games_map.setdefault(g, []).append(p)

    html = _stale_notice
    for game, gpicks in games_map.items():
        # Parse team names
        if " @ " in game:
            away, home = [t.strip() for t in game.split(" @ ", 1)]
        elif re.search(r'\s+vs\.?\s+', game, re.IGNORECASE):
            pts = re.split(r'\s+vs\.?\s+', game, flags=re.IGNORECASE)
            away, home = pts[0].strip(), (pts[1].strip() if len(pts) > 1 else "")
        else:
            away, home = game, ""

        away_b = _badge_mlb(away)
        home_b = _badge_mlb(home) if home else ""

        picks_html = ""
        any_watchlist = any(not p.get("lineup_confirmed", True) for p in gpicks)
        for p in gpicks:
            pick     = p.get("pick","")
            odds     = _fmt_odds(p.get("odds", 0))
            ev       = p.get("edge","") or p.get("ev","") or p.get("expected_value","")
            is_watch = not p.get("lineup_confirmed", True)
            try:
                is_star = float(str(ev).replace("%","").strip()) >= 5.0 if ev else False
            except Exception:
                is_star = False

            ev_b = f'<span class="gpick-ev">EV+ {_esc(str(ev))}%</span>' if ev else ""
            star_cls = " star" if (is_star and not is_watch) else ""
            icon = "⏳" if is_watch else ("★" if is_star else "·")
            watch_style = 'opacity:.65;' if is_watch else ''
            picks_html += (
                f'<div class="gpick{star_cls}" style="{watch_style}">'
                f'<span class="gpick-ico">{icon}</span>'
                f'<span class="gpick-pick">{_esc(pick)}</span>'
                f'<span class="gpick-odds">{odds}</span>'
                f'{ev_b}'
                f'</div>'
            )

        # watchlist banner: shown once per game if any pick lacks confirmed lineup
        watch_banner = (
            '<div style="font-size:.65rem;color:#f59e0b;background:#1c1400;border-radius:4px;'
            'padding:2px 8px;margin-bottom:6px;letter-spacing:.03em">'
            '⏳ LINEUP PENDIENTE — no apostar hasta confirmar</div>'
        ) if any_watchlist else ""

        html += (
            f'<div class="game-card" style="--acc:#e05252">'
            f'<div class="game-matchup">{away_b}'
            f'<span class="matchup-vs">@</span>'
            f'{home_b}</div>'
            f'{watch_banner}'
            f'<div class="game-picks">{picks_html}</div>'
            f'</div>'
        )
    return html


def _bsn_today_section():
    """Today's BSN games rendered as game-card rows."""
    today_games = _bsn_today_games()
    if not today_games:
        return ('<div class="today-empty">'
                'Sin juegos cargados para hoy.<br>'
                '<span style="font-size:.7rem;color:#475569">'
                'Agrega juegos con 📅 Games.</span>'
                '</div>')

    html = ""
    for g in today_games:
        away  = g.get("team1","")
        home  = g.get("team2","")
        gtime = g.get("game_time","")
        away_b = _badge_bsn(away)
        home_b = _badge_bsn(home)
        time_frag = f'<span class="game-time">🕐 {_esc(gtime)}</span>' if gtime else ""
        html += (
            f'<div class="game-card" style="--acc:#f5a623">'
            f'<div class="game-matchup">{away_b}'
            f'<span class="matchup-vs">@</span>'
            f'{home_b}{time_frag}</div>'
            f'</div>'
        )
    return html


def _nba_ir_section():
    """Active NBA injury report mini table."""
    ACTIVE = {"out","doubtful","questionable"}
    all_ir = _rj(NBA_IR)
    entries = [e for e in all_ir if e.get("status","").lower() in ACTIVE]
    if not entries:
        return '<div class="today-empty" style="padding:14px">Sin lesionados activos reportados.</div>'

    html = '<div class="mini-table">'
    for e in sorted(entries, key=lambda x: (x.get("team_abb",""), x.get("player",""))):
        status = e.get("status","").lower()
        if status == "out":
            sc, sl = "mt-out", "OUT"
        elif status == "doubtful":
            sc, sl = "mt-dbt", "DBT"
        else:
            sc, sl = "mt-qst", "QST"
        abb  = e.get("team_abb","")
        badge = _badge_nba(abb) if abb else ""
        ppg = e.get("ppg","")
        sub = f"{ppg} PPG" if ppg else ""
        html += (
            f'<div class="mt-row">'
            f'{badge}'
            f'<span class="mt-name">{_esc(e.get("player",""))}</span>'
            f'<span class="mt-sub">{_esc(sub)}</span>'
            f'<span class="mt-status {sc}">{sl}</span>'
            f'</div>'
        )
    html += '</div>'
    return html


# ── BSN PANEL ─────────────────────────────────────────────────────────
def bsn_panel():
    team_opts  = _team_opts(BSN_TEAMS)
    log        = _rj(BSN_LOG)
    gp_data    = _rj(BSN_GP) if os.path.exists(BSN_GP) else {}
    gp_opts    = "\n".join(f'<option value="{t}">{t.title()}</option>' for t in BSN_TEAMS)
    books = ["BetMGM","DraftKings","FanDuel","Caesars","Bet365","PointsBet","Otro"]
    book_opts = "\n".join(f'<option value="{b}">{b}</option>' for b in books)
    today = date.today().strftime("%Y-%m-%d")
    grade_html  = _grade_rows(BSN_LOG, "/api/bsn/grade")
    today_games = _bsn_today_games()

    # Pre-compute game options for Log Pick dropdown (no backslashes inside f-string)
    _bsn_game_opts = ""
    for _g in today_games:
        _t1 = _esc(_g.get("team1", ""))
        _t2 = _esc(_g.get("team2", ""))
        _gt = _g.get("game_time", "")
        _time_bit = f" · {_gt}" if _gt else ""
        _bsn_game_opts += f'<option value="{_t1}|{_t2}">{_t1} @ {_t2}{_time_bit}</option>\n'

    # Build games list for management modal
    if today_games:
        game_rows = ""
        for g in today_games:
            away = g.get("team1","")
            home = g.get("team2","")
            gtime = g.get("game_time","—")
            label = f"{away} @ {home}"
            game_rows += f"""<div style="background:#0a0a0a;border-radius:8px;padding:10px;border:1px solid #1e293b;margin-bottom:8px">
  <div style="font-size:0.85rem;font-weight:700;margin-bottom:8px">{_esc(label)}</div>
  <div style="display:flex;gap:8px;align-items:center;margin-bottom:6px">
    <input id="gt-{away}-{home}" value="{_esc(gtime)}" style="flex:1;margin-bottom:0;padding:7px 10px;font-size:0.82rem" placeholder="8:00 PM">
    <button class="btn" style="padding:7px 12px;font-size:0.75rem;flex-shrink:0" onclick="bsnEditTime('{away}','{home}',document.getElementById('gt-{away}-{home}').value)">💾</button>
  </div>
  <button class="btn red" style="width:100%;padding:7px;font-size:0.75rem" onclick="bsnRemoveGame('{away}','{home}')">🗑 Remover juego</button>
</div>"""
        games_section = f'<p style="font-size:0.72rem;color:#94a3b8;margin-bottom:10px">{len(today_games)} juego(s) hoy</p>' + game_rows
    else:
        games_section = '<p style="font-size:0.82rem;color:#64748b;text-align:center;padding:10px 0">Sin juegos manuales para hoy</p>'

    return f"""
{_stats_bar(BSN_LOG, '#f5a623', 'BSN', '#f5a623', 'rgba(245,166,35,.1)')}

<!-- Today's Games -->
<div class="section-hdr">
  <span class="section-title">🏀 Juegos de Hoy</span>
  <button class="section-btn" onclick="openModal('bsn-games')">📅 Gestionar</button>
</div>
{_bsn_today_section()}

{_panel_actions(
  primary=[
    ('+',  'Log Pick',  "openBsnLogPick()",                                                                                          'green'),
    ('▶',  'Run Modelo',"refreshPicksBlock('bsn-picks-block','python3 bsn.py --picks','BSN','/api/view/bsn/picks')", 'orange'),
  ],
  secondary=[
    ('📋', 'Lines',    "openBsnLines()"),
    ('📊', 'Stats',    "openView('/api/view/bsn/stats','BSN · Stats')"),
    ('✓',  'Grade',   "openModal('bsn-grade')"),
    ('📜', 'Historial',"openView('/api/view/bsn/log','BSN · Historial')"),
    ('🖼', 'Record',  "openRecordModal('BSN','bsn.py','BSN','#f5a623')"),
    ('🃏', 'Parlay',  "openModal('bsn-parlay')"),
    ('🏥', 'IR',      "openBsnIrModal()"),
    ('🎯', 'GP',      "openModal('bsn-gp')"),
  ]
)}

<!-- BSN picks placeholder block -->
<div id="bsn-picks-block" style="margin-top:8px"></div>

<!-- Tools & Publish Accordion -->
<div class="tools-group">
  <div class="tools-toggle" onclick="toggleTools('bsn')">
    <span>⚙ Tools &amp; Publishing</span>
    <span class="tools-chevron" id="chev-bsn">›</span>
  </div>
  <div class="tools-body" id="tools-bsn">
    {_cmd_card('Generar HTML y publicar a GitHub Pages', '🌐 Publish', "runCmd('python3 bsn.py --export-html --publish','BSN','BSN → Export & Publish')", 'blue')}
    {_cmd_card('Record diario JPG (hoy) y publicar', f'🖼 Record {today}', f"runCmd('python3 bsn.py --export-record {today} --publish','BSN','BSN → Record Diario')", 'blue')}
    {_cmd_card('Record all-time JPG y publicar', '🏆 Record All-Time', "runCmd('python3 bsn.py --export-record --publish','BSN','BSN → Record All-Time')", 'blue')}
  </div>
</div>

<!-- BSN Log Modal -->
<div class="modal-bg" id="bsn-log">
<div class="modal">
  <h2>BSN — Log Pick</h2>
  <form id="f-bsn-log" onsubmit="event.preventDefault();submitForm('f-bsn-log','/api/bsn/log','bsn-log')">
    <input type="hidden" name="date" value="{today}">
    <input type="hidden" name="team1" id="bsn-log-t1">
    <input type="hidden" name="team2" id="bsn-log-t2">
    <label>Juego</label>
    <select id="bsn-log-game" onchange="bsnLogFillGame(this)" required>
      <option value="">— Selecciona juego —</option>
      {_bsn_game_opts}
      <option value="__manual__">✏️ Otro juego...</option>
    </select>
    <div id="bsn-log-manual" style="display:none">
      <div class="row2" style="margin-top:6px">
        <div><label>Equipo A</label><input id="bsn-log-t1-inp" placeholder="LEONES" autocapitalize="characters" oninput="document.getElementById('bsn-log-t1').value=this.value"></div>
        <div><label>Equipo B</label><input id="bsn-log-t2-inp" placeholder="INDIOS" autocapitalize="characters" oninput="document.getElementById('bsn-log-t2').value=this.value"></div>
      </div>
    </div>
    <label style="margin-top:10px">Pick</label>
    <div class="mlg-pick-opts" id="bsn-log-pick-opts"></div>
    <input id="bsn-log-pick-inp" name="pick" placeholder="Ej: LEONES ML / O 155.5" autocapitalize="characters" required>
    <div class="row2">
      <div><label>Odds</label><input name="odds" id="bsn-log-odds-inp" placeholder="-110" required></div>
      <div><label>Apuesta $</label><input type="number" name="stake" placeholder="15" step="0.01" min="1" required></div>
    </div>
    <label>Sportsbook</label>
    <select name="book">{book_opts}</select>
    <label>Análisis</label>
    <textarea name="analysis" placeholder="Razón del pick..."></textarea>
    <div class="btn-row">
      <button type="button" class="btn gray" onclick="closeModal('bsn-log')">Cancelar</button>
      <button type="submit" class="btn green">✅ Loguear</button>
    </div>
  </form>
</div></div>

<!-- BSN Parlay Modal -->
<div class="modal-bg" id="bsn-parlay">
<div class="modal">
  <h2>BSN — Log Parlay</h2>
  <form id="f-bsn-parlay" onsubmit="event.preventDefault();submitForm('f-bsn-parlay','/api/bsn/log-parlay','bsn-parlay')">
    <label>Fecha</label>
    <input type="date" name="date" value="{today}" required>
    <div class="row2">
      <div><label>Odds parlay</label><input name="odds" placeholder="+350" required></div>
      <div><label>Apuesta $</label><input type="number" name="stake" placeholder="10" step="0.01" min="1" required></div>
    </div>
    <label>Sportsbook</label>
    <select name="book">{book_opts}</select>
    <div class="section-sep"></div>
    <div id="legs-container">
      <div class="leg" id="leg-1">
        <div class="leg-title">LEG 1</div>
        <input name="leg1_game" placeholder="Juego (Ej: LEONES vs OSOS)" autocapitalize="characters" required>
        <input name="leg1_pick" placeholder="Pick (Ej: LEONES ML)" autocapitalize="characters" required>
      </div>
    </div>
    <button type="button" class="btn gray" style="width:100%;margin-bottom:12px" onclick="addLeg()">+ Agregar Leg</button>
    <label>Análisis</label>
    <textarea name="analysis" placeholder="Razón del parlay..."></textarea>
    <div class="btn-row">
      <button type="button" class="btn gray" onclick="closeModal('bsn-parlay')">Cancelar</button>
      <button type="submit" class="btn green">✅ Loguear Parlay</button>
    </div>
  </form>
</div></div>

<!-- BSN Add Game Modal -->
<div class="modal-bg" id="bsn-addgame">
<div class="modal">
  <h2>BSN — Agregar Juego</h2>
  <p style="font-size:0.72rem;color:#64748b;margin-bottom:14px">Agrega un juego al calendario de hoy para que el modelo lo considere.</p>
  <form id="f-bsn-addgame" onsubmit="event.preventDefault();submitForm('f-bsn-addgame','/api/bsn/add-game','bsn-addgame')">
    <label>Equipo Visitante (Away)</label>
    <select name="away" required><option value="">—</option>{gp_opts}</select>
    <label>Equipo Local (Home)</label>
    <select name="home" required><option value="">—</option>{gp_opts}</select>
    <label>Hora del juego</label>
    <input name="time" placeholder="8:00 PM" value="8:00 PM">
    <div class="btn-row" style="margin-top:8px">
      <button type="button" class="btn gray" onclick="closeModal('bsn-addgame')">Cancelar</button>
      <button type="submit" class="btn green">➕ Agregar Juego</button>
    </div>
  </form>
</div></div>

<!-- BSN Grade Modal -->
<div class="modal-bg" id="bsn-grade">
<div class="modal">
  <h2>BSN — Grade Picks</h2>
  <div style="max-height:420px;overflow-y:auto">
    {grade_html}
  </div>
  <div class="btn-row" style="margin-top:12px">
    <button type="button" class="btn gray" onclick="closeModal('bsn-grade')">Cerrar</button>
  </div>
</div></div>

<!-- BSN IR Modal -->
<div class="modal-bg" id="bsn-ir">
<div class="modal" style="max-width:480px">
  <!-- Header AI style -->
  <div style="background:linear-gradient(135deg,rgba(239,68,68,.12),rgba(245,166,35,.08));
    border:1px solid rgba(239,68,68,.25);border-radius:14px;padding:14px 16px;margin-bottom:16px;
    position:relative;overflow:hidden">
    <div style="position:absolute;top:0;left:0;right:0;height:3px;
      background:linear-gradient(90deg,#ef4444,#f5a623,transparent)"></div>
    <div style="font-size:.58rem;font-weight:900;letter-spacing:.2em;text-transform:uppercase;
      color:#ef4444;margin-bottom:2px">🏥 BSN — INJURY REPORT</div>
    <div style="font-size:.6rem;color:#64748b">Rate: <b style="color:#f87171">1=OUT</b> · <b style="color:#fbbf24">2=Doubtful</b> · <b style="color:#94a3b8">3=Limited</b></div>
  </div>

  <!-- Live IR list -->
  <div id="bsn-ir-list" style="margin-bottom:14px">
    <div style="font-size:.55rem;color:#475569;text-align:center;padding:8px">Cargando IR...</div>
  </div>

  <!-- Add / Remove form -->
  <div style="background:rgba(239,68,68,.04);border:1px solid rgba(239,68,68,.15);
    border-radius:12px;padding:14px 16px">
    <div style="font-size:.55rem;font-weight:900;letter-spacing:.15em;text-transform:uppercase;
      color:#ef4444;margin-bottom:10px">➕ Agregar / Actualizar Jugador</div>
    <form id="f-bsn-ir" onsubmit="event.preventDefault();_bsnIrSubmit()">
      <input type="hidden" name="action" id="bsn-ir-action" value="add">
      <div class="row2">
        <div>
          <label>Equipo</label>
          <select name="team" required>
            <option value="">— Selecciona —</option>{gp_opts}
          </select>
        </div>
        <div>
          <label>Jugador</label>
          <input name="player" placeholder="WATERS" autocapitalize="characters" required>
        </div>
      </div>
      <label>Rate</label>
      <select name="rate" required>
        <option value="1">1 — OUT (mayor impacto)</option>
        <option value="2" selected>2 — Doubtful</option>
        <option value="3">3 — Limited</option>
      </select>
      <div class="row2" style="margin-top:8px">
        <div>
          <label>PPG</label>
          <input name="ppg" id="bsn-ir-ppg" type="number" step="0.1" min="0" placeholder="18.5" required>
        </div>
        <div>
          <label>USG%</label>
          <input name="usg" id="bsn-ir-usg" type="number" step="0.1" min="0" max="100" placeholder="22.3" required>
        </div>
      </div>
      <div id="bsn-ir-impact" style="font-size:.58rem;color:#94a3b8;margin-top:6px;min-height:16px"></div>
      <div class="btn-row" style="margin-top:10px">
        <button type="button" class="btn gray" onclick="closeModal('bsn-ir')">Cancelar</button>
        <button type="button" class="btn red" onclick="_bsnIrRemove()">🗑 Remover</button>
        <button type="submit" class="btn green">➕ Guardar</button>
      </div>
    </form>
  </div>
</div></div>

<!-- BSN GP Modal -->
<div class="modal-bg" id="bsn-gp">
<div class="modal">
  <h2>BSN — Actualizar GP</h2>
  <form id="f-bsn-gp" onsubmit="event.preventDefault();submitForm('f-bsn-gp','/api/bsn/gp','bsn-gp')">
    <div class="row2">
      <div><label>Equipo</label><select name="team" required><option value="">—</option>{gp_opts}</select></div>
      <div><label>Juegos (GP)</label><input type="number" name="gp" placeholder="12" min="0" required></div>
    </div>
    <div class="btn-row">
      <button type="button" class="btn gray" onclick="closeModal('bsn-gp')">Cancelar</button>
      <button type="submit" class="btn green">✅ Actualizar</button>
    </div>
  </form>
</div></div>

<!-- BSN Games Modal -->
<div class="modal-bg" id="bsn-games">
<div class="modal">
  <h2>BSN — Juegos de Hoy</h2>
  <div style="max-height:320px;overflow-y:auto;margin-bottom:14px">
    {games_section}
  </div>
  <div class="section-sep"></div>
  <p style="font-size:0.72rem;font-weight:700;color:#94a3b8;letter-spacing:1px;margin:10px 0 10px">AGREGAR JUEGO</p>
  <form id="f-bsn-games-add" onsubmit="event.preventDefault();submitForm('f-bsn-games-add','/api/bsn/add-game','bsn-games')">
    <div class="row2">
      <div><label>Visitante (Away)</label><select name="away" required><option value="">—</option>{gp_opts}</select></div>
      <div><label>Local (Home)</label><select name="home" required><option value="">—</option>{gp_opts}</select></div>
    </div>
    <label>Hora del juego</label>
    <input name="time" placeholder="8:00 PM" value="8:00 PM">
    <div class="btn-row">
      <button type="button" class="btn gray" onclick="closeModal('bsn-games')">Cerrar</button>
      <button type="submit" class="btn green">➕ Agregar</button>
    </div>
  </form>
</div></div>

<!-- BSN Set Lines Modal -->
<div class="modal-bg" id="bsn-lines" onclick="if(event.target===this)closeModal('bsn-lines')">
<div class="modal" style="max-width:500px;padding:0;overflow:hidden;border-radius:20px">
  <div style="background:linear-gradient(90deg,rgba(245,166,35,.07),transparent);border-bottom:1px solid rgba(245,166,35,.1);padding:16px 20px 14px;display:flex;align-items:center;justify-content:space-between">
    <div style="font-size:.65rem;font-weight:900;letter-spacing:.2em;color:#f5a623">📋 LÍNEAS DE MERCADO</div>
    <button type="button" onclick="closeModal('bsn-lines')" style="background:rgba(245,166,35,.08);border:1px solid rgba(245,166,35,.2)!important;border-radius:50%!important;width:26px;height:26px;color:#f5a623!important;cursor:pointer;font-size:.7rem;display:flex;align-items:center;justify-content:center;padding:0;margin-bottom:0;flex-shrink:0">✕</button>
  </div>
  <div style="padding:16px 20px 20px">
    <div id="bsn-lines-body" style="max-height:60vh;overflow-y:auto;padding-right:4px">
      <div style="text-align:center;padding:30px;color:#475569">Cargando juegos...</div>
    </div>
    <div style="display:flex;gap:8px;margin-top:14px">
      <button type="button" onclick="closeModal('bsn-lines')" style="background:rgba(255,255,255,.04)!important;border:1px solid rgba(255,255,255,.1)!important;border-radius:10px!important;color:#475569!important;font-size:.78rem;font-weight:700;padding:11px 16px;cursor:pointer;font-family:inherit">Cancelar</button>
      <button type="button" onclick="submitBsnLines()" style="flex:1;padding:12px;background:linear-gradient(90deg,#059669,#10b981);border:none!important;border-radius:10px!important;color:#fff!important;font-size:.8rem;font-weight:800;letter-spacing:.08em;cursor:pointer;font-family:inherit">✅ GUARDAR TODO</button>
    </div>
  </div>
</div></div>
<style>
.bsl-section-lbl{{font-size:.52rem;font-weight:900;color:rgba(245,166,35,.7);letter-spacing:.14em;text-transform:uppercase;margin-bottom:5px}}
.bsl-card{{background:#0d0d18;border:1px solid rgba(245,166,35,.1);border-radius:12px;padding:14px;margin-bottom:10px}}
.bsl-card input,.bsl-card select{{background:rgba(0,0,0,.5)!important;border:1px solid rgba(245,166,35,.1)!important;border-radius:8px!important;color:#e2e8f0!important;padding:8px 11px!important;font-size:.83rem!important;width:100%;margin-bottom:0;transition:border-color .2s}}
.bsl-card input:focus,.bsl-card select:focus{{outline:none!important;border-color:rgba(245,166,35,.4)!important}}
.bsl-field-lbl{{font-size:.58rem;color:#64748b;margin-bottom:3px;display:block}}
.bsl-3col{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:7px;margin-bottom:10px}}
.bsl-2col{{display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-bottom:10px}}
.bsl-saved-badge{{font-size:.55rem;background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.3);color:#22c55e;padding:2px 7px;border-radius:99px}}
</style>
"""


# ── NBA PANEL ─────────────────────────────────────────────────────────
def nba_panel():
    nba_team_opts = _team_opts(NBA_TEAMS)
    today = date.today().strftime("%Y-%m-%d")
    books = ["DraftKings","FanDuel","BetMGM","Caesars","Bet365","PointsBet","Otro"]
    book_opts = "\n".join(f'<option value="{b}">{b}</option>' for b in books)

    # Current IR — only show ACTIVE injuries (out / doubtful / questionable)
    ACTIVE_STATUSES = {"out", "doubtful", "questionable"}
    all_ir = _rj(NBA_IR)
    ir_entries = [e for e in all_ir if e.get("status","").lower() in ACTIVE_STATUSES]
    ir_rows = ""
    for e in sorted(ir_entries, key=lambda x: (x.get("team_abb",""), x.get("player",""))):
        status = e.get("status","").upper()
        sc = "#ef4444" if status=="OUT" else ("#f97316" if status=="DOUBTFUL" else "#eab308")
        ir_rows += (f'<tr>'
                    f'<td>{e.get("team_abb","")}</td>'
                    f'<td>{e.get("player","")}</td>'
                    f'<td style="color:{sc}">{status}</td>'
                    f'<td>{e.get("ppg","—")}</td>'
                    f'</tr>')

    grade_html = _grade_rows(NBA_LOG, "/api/nba/grade")

    last_ir = ('—' if _NBA_IR_LAST_REFRESH[0] == 0.0 else
               __import__('datetime').datetime.fromtimestamp(_NBA_IR_LAST_REFRESH[0]).strftime('%H:%M'))

    return f"""
{_stats_bar(NBA_LOG, '#4f8ef7', 'NBA', '#4f8ef7', 'rgba(79,142,247,.1)')}

<!-- Picks Section -->
<div class="section-hdr">
  <span class="section-title">🏀 Picks de Hoy</span>
</div>
<div id="nba-picks-block">
  <div class="today-empty">
    Corre el modelo para ver los picks de hoy.<br>
    <span style="font-size:.7rem;color:#475569">Presiona ▶ Run Modelo para generar picks.</span>
  </div>
</div>

{_panel_actions(
  primary=[
    ('+', 'Log Pick',  "nbaLogOpen()",                                                                                                         'green'),
    ('▶', 'Run Modelo',"refreshPicksBlock('nba-picks-block','bash -c &quot;rm -f nba_picks_body_current.html &amp;&amp; python3 nba.py --picks&quot;','NBA','/api/view/nba/picks')", 'blue'),
  ],
  secondary=[
    ('📐', 'Lines',    "runThenView('python3 nba.py --lines','NBA','/api/view/nba/lines','NBA · Lines')"),
    ('📊', 'Stats',    "openView('/api/view/nba/stats','NBA · Stats')"),
    ('✓',  'Grade',   "openModal('nba-grade')"),
    ('📜', 'Historial',"openView('/api/view/nba/log','NBA · Historial')"),
    ('🖼', 'Record',  "openRecordModal('NBA','nba.py','NBA','#3b82f6')"),
    ('🏥', 'IR',      "openModal('nba-ir')"),
    ('🔄', 'Refresh IR',"runCmd('python3 nba.py --ir refresh','NBA','NBA → IR')"),
    ('📤', 'Publish', "runCmd('python3 nba.py --picks --publish','NBA','NBA → Publish')"),
  ]
)}

<!-- Tools & Publish Accordion -->
<div class="tools-group">
  <div class="tools-toggle" onclick="toggleTools('nba')">
    <span>⚙ Tools &amp; Publishing</span>
    <span class="tools-chevron" id="chev-nba">›</span>
  </div>
  <div class="tools-body" id="tools-nba">
    {_cmd_card('Picks + publicar HTML a GitHub Pages', '🚀 Publish Picks', "runCmd('python3 nba.py --picks --publish','NBA','NBA → Picks + Publish')", 'blue')}
    {_cmd_card('Exportar lines + picks HTML y publicar', '🌐 Export HTML', "runCmd('python3 nba.py --export-html --publish','NBA','NBA → Export HTML')", 'blue')}
    {_cmd_card('Exportar log de picks y publicar', '📤 Publish Log', "runCmd('python3 nba.py --export-log --publish','NBA','NBA → Export Log')", 'blue')}
  </div>
</div>

<style>
#nba-log .modal{{background:linear-gradient(160deg,#070710 0%,#05050d 100%);border:1px solid rgba(79,142,247,.18);border-radius:20px;padding:0;overflow:hidden;box-shadow:0 0 0 1px rgba(79,142,247,.04),0 24px 80px rgba(0,0,0,.95),inset 0 1px 0 rgba(255,255,255,.05);max-width:480px}}
#nba-log .mlg-hdr{{background:linear-gradient(90deg,rgba(79,142,247,.07),rgba(124,58,237,.05),transparent);border-bottom:1px solid rgba(79,142,247,.1);padding:16px 20px 14px;display:flex;align-items:center;justify-content:space-between}}
#nba-log .mlg-title{{font-size:.7rem;font-weight:900;letter-spacing:.2em;background:linear-gradient(90deg,#4f8ef7,#7c3aed);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
#nba-log .mlg-close{{background:rgba(79,142,247,.08);border:1px solid rgba(79,142,247,.2)!important;border-radius:50%!important;width:26px;height:26px;color:#4f8ef7!important;cursor:pointer;font-size:.7rem;display:flex;align-items:center;justify-content:center;padding:0;margin-bottom:0;flex-shrink:0}}
#nba-log .mlg-body{{padding:16px 20px 20px}}
#nba-log .mlg-lbl{{font-size:.55rem;font-weight:800;color:rgba(79,142,247,.75);letter-spacing:.14em;text-transform:uppercase;margin-bottom:5px;text-shadow:0 0 8px rgba(79,142,247,.2);display:block}}
#nba-log select,#nba-log input:not([type=hidden]),#nba-log textarea{{background:rgba(0,0,0,.5)!important;border:1px solid rgba(79,142,247,.12)!important;border-radius:9px!important;color:#e2e8f0!important;padding:10px 13px!important;font-size:.88rem!important;width:100%;margin-bottom:0;transition:border-color .2s,box-shadow .2s;font-family:inherit}}
#nba-log select:focus,#nba-log input:focus,#nba-log textarea:focus{{outline:none!important;border-color:rgba(79,142,247,.4)!important;box-shadow:0 0 0 3px rgba(79,142,247,.07)!important}}
#nba-log .mlg-field{{margin-bottom:13px}}
#nba-log .mlg-pick-card{{background:linear-gradient(135deg,rgba(79,142,247,.05),rgba(124,58,237,.04));border:1px solid rgba(79,142,247,.2);border-radius:12px;padding:12px 14px;margin-bottom:13px;display:none;position:relative;overflow:hidden}}
#nba-log .mlg-pick-card::before{{content:'';position:absolute;top:0;left:10%;right:10%;height:1px;background:linear-gradient(90deg,transparent,rgba(79,142,247,.5),transparent)}}
#nba-log .mlg-pick-name{{font-size:1.05rem;font-weight:800;color:#f1f5f9;letter-spacing:.02em;margin-bottom:7px}}
#nba-log .mlg-chips{{display:flex;gap:5px;flex-wrap:wrap}}
#nba-log .mlg-chip{{padding:3px 9px;border-radius:99px;font-size:.6rem;font-weight:700;letter-spacing:.05em}}
#nba-log .mlg-pick-opts{{display:none;flex-direction:column;gap:6px;margin-bottom:13px}}
#nba-log .mlg-pick-opt{{
  display:flex;align-items:center;justify-content:space-between;gap:10px;
  background:rgba(79,142,247,.04);border:1px solid rgba(79,142,247,.12);
  border-radius:10px;padding:10px 13px;cursor:pointer;transition:all .15s;
}}
#nba-log .mlg-pick-opt:hover{{background:rgba(79,142,247,.1);border-color:rgba(79,142,247,.3)}}
#nba-log .mlg-pick-opt.selected{{
  background:rgba(79,142,247,.14);border-color:rgba(79,142,247,.45);
  box-shadow:0 0 16px rgba(79,142,247,.12);
}}
#nba-log .mlg-opt-pick{{font-size:.88rem;font-weight:800;color:#f1f5f9;flex:1;min-width:0}}
#nba-log .mlg-opt-meta{{display:flex;gap:5px;align-items:center;flex-shrink:0}}
#nba-log .mlg-opt-odds{{font-size:.72rem;font-weight:700;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.1);border-radius:6px;padding:2px 8px;color:#94a3b8}}
#nba-log .mlg-opt-ev{{font-size:.68rem;font-weight:800;color:#10b981;background:rgba(16,185,129,.1);border:1px solid rgba(16,185,129,.25);border-radius:6px;padding:2px 8px}}
#nba-log .mlg-opt-type{{font-size:.58rem;font-weight:900;color:#4f8ef7;background:rgba(79,142,247,.1);border:1px solid rgba(79,142,247,.2);border-radius:6px;padding:2px 7px;letter-spacing:.06em;text-transform:uppercase}}
.nba-chip-odds{{background:rgba(79,142,247,.1);border:1px solid rgba(79,142,247,.3);color:#4f8ef7}}
.nba-chip-edge{{background:rgba(240,120,32,.1);border:1px solid rgba(240,120,32,.25);color:#f07820}}
.nba-chip-modelo{{background:rgba(124,58,237,.1);border:1px solid rgba(124,58,237,.25);color:#a78bfa}}
.nba-chip-ev{{background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.25);color:#22c55e}}
#nba-log .mlg-row2{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:13px}}
#nba-log .mlg-books{{display:flex;gap:6px;flex-wrap:wrap;margin-top:5px}}
#nba-log .mlg-book-btn{{background:rgba(0,0,0,.4);border:1px solid rgba(255,255,255,.1)!important;border-radius:8px!important;color:#64748b;font-size:.65rem;font-weight:700;letter-spacing:.04em;padding:7px 11px;cursor:pointer;transition:all .18s;font-family:inherit;margin-bottom:0}}
#nba-log .mlg-book-btn.active{{background:rgba(79,142,247,.12)!important;border-color:rgba(79,142,247,.4)!important;color:#4f8ef7!important;box-shadow:0 0 14px rgba(79,142,247,.12)}}
#nba-log .mlg-book-btn:hover:not(.active){{border-color:rgba(79,142,247,.2)!important;color:#e2e8f0}}
#nba-log-odds-manual-wrap{{display:none;margin-top:8px}}
#nba-log .mlg-footer{{display:flex;gap:8px;margin-top:4px;align-items:stretch}}
#nba-log .mlg-submit{{flex:1;padding:12px;background:linear-gradient(90deg,#059669,#10b981);border:none!important;border-radius:10px!important;color:#fff!important;font-size:.8rem;font-weight:800;letter-spacing:.1em;cursor:pointer;font-family:inherit;transition:opacity .2s}}
#nba-log .mlg-submit:hover{{opacity:.85}}
#nba-log .mlg-cancel{{background:rgba(255,255,255,.04)!important;border:1px solid rgba(255,255,255,.1)!important;border-radius:10px!important;color:#475569!important;font-size:.78rem;font-weight:700;padding:11px 16px;cursor:pointer;font-family:inherit;white-space:nowrap}}
.mlg-alt-pill{{display:inline-block;cursor:pointer;padding:5px 10px;border-radius:8px;font-size:.62rem;font-weight:700;letter-spacing:.04em;transition:filter .15s,opacity .15s;font-family:inherit}}
.mlg-alt-pill:hover{{filter:brightness(1.35);opacity:.9}}
#nba-log textarea{{height:60px!important;resize:none}}
</style>
<script>
var nbaLogPicksData={{}};

function nbaLogOpen(){{
  fetch('/api/nba/picks-today')
    .then(function(r){{return r.json();}})
    .then(function(d){{
      nbaLogPicksData = d.picks || {{}};
      var sel = document.getElementById('nba-log-game');
      if(!sel) return;
      var games = Object.keys(nbaLogPicksData);
      sel.innerHTML = '<option value="">— Selecciona juego —</option>';
      games.forEach(function(game){{
        if(game.indexOf(' @ ')<0) return;
        var parts = game.split(' @ ');
        var opt = document.createElement('option');
        opt.value = parts[0]+'|'+parts[1];
        opt.textContent = game;
        sel.appendChild(opt);
      }});
      if(games.length === 0){{
        sel.innerHTML = '<option value="">Sin picks del modelo para hoy</option>';
      }}
    }})
    .catch(function(){{ /* sin conexión — usa lo que hay */ }});
  // Reset on open
  var optsEl=document.getElementById('nba-log-pick-opts');
  if(optsEl){{ optsEl.innerHTML=''; optsEl.style.display='none'; }}
  var pickInp=document.getElementById('nba-log-pick-inp');
  if(pickInp) pickInp.value='';
  var oddsInp=document.getElementById('nba-log-odds-inp');
  if(oddsInp) oddsInp.value='';
  openModal('nba-log');
}}

var _nbaCurrentPicks = [];

function _nbaAltPicksHtml(away, home){{
  var accent = '#4f8ef7';
  var _sp = function(pick, label, ac, suffix){{
    var a = ac||accent;
    return '<span class="mlg-alt-pill" style="border:1px solid '+a+'40;color:'+a+';background:'+a+'12" '+
      'onclick="nbaLogSetAltPick(\\\''+pick+'\\\')">'+(label||pick)+(suffix||'')+'</span>';
  }};
  var _row = function(lbl, pills){{
    return '<div style="display:flex;gap:5px;flex-wrap:wrap;align-items:center;margin-bottom:5px">'+
      '<span style="font-size:.44rem;font-weight:900;color:#475569;letter-spacing:.1em;min-width:28px;flex-shrink:0">'+lbl+'</span>'+
      pills+'</div>';
  }};
  return '<div style="padding-top:10px;margin-top:8px;border-top:1px solid rgba(79,142,247,.1)">'+
    '<div style="font-size:.44rem;font-weight:900;color:#334155;letter-spacing:.12em;margin-bottom:6px;text-transform:uppercase">Otros Picks</div>'+
    _row('ML',  _sp(away+' ML', away+' ML', '#4f8ef7') + _sp(home+' ML', home+' ML', '#4f8ef7')) +
    _row('SPR', _sp(away+' -', away, '#f07820', ' <span style="opacity:.55">-▸</span>') +
                _sp(home+' -', home, '#f07820', ' <span style="opacity:.55">-▸</span>')) +
    _row('TOT', _sp('O ', 'OVER', '#f97316') + _sp('U ', 'UNDER', '#a78bfa')) +
  '</div>';
}}

function nbaLogFillPick(sel){{
  var v=sel.value; if(!v) return;
  var parts=v.split('|'); var away=parts[0], home=parts[1];
  document.querySelector('#f-nba-log [name=away]').value=away;
  document.querySelector('#f-nba-log [name=home]').value=home;

  // Find game key
  var gameKey='';
  Object.keys(nbaLogPicksData).forEach(function(k){{
    var kp=k.split(' @ ');
    if(kp[0]===away&&kp[1]===home) gameKey=k;
  }});
  _nbaCurrentPicks = gameKey ? (nbaLogPicksData[gameKey]||[]) : [];

  var optsEl = document.getElementById('nba-log-pick-opts');
  if(_nbaCurrentPicks.length===0){{
    // No model picks — show only alt picks
    optsEl.innerHTML = _nbaAltPicksHtml(away, home);
    optsEl.style.display='flex';
    optsEl.style.flexDirection='column';
    document.getElementById('nba-log-pick-inp').value='';
    document.getElementById('nba-log-odds-inp').value='';
    return;
  }}

  // Render all model picks as selectable rows + alt picks below
  var html='';
  _nbaCurrentPicks.forEach(function(p,i){{
    var typeTag = p.type ? '<span class="mlg-opt-type">'+p.type+'</span>' : '';
    var evTag   = p.ev   ? '<span class="mlg-opt-ev">EV '+p.ev+'</span>'  : '';
    var oddsTag = p.odds ? '<span class="mlg-opt-odds">'+p.odds+'</span>' : '';
    html += '<div class="mlg-pick-opt" id="nba-pick-opt-'+i+'" onclick="nbaLogSelectPick('+i+')">'
          +   '<span class="mlg-opt-pick">'+(p.pick||'—')+'</span>'
          +   '<span class="mlg-opt-meta">'+typeTag+evTag+oddsTag+'</span>'
          + '</div>';
  }});
  optsEl.innerHTML = html + _nbaAltPicksHtml(away, home);
  optsEl.style.display='flex';
  optsEl.style.flexDirection='column';

  // Auto-select first pick
  nbaLogSelectPick(0);
}}

function nbaLogSetAltPick(pick){{
  var inp = document.getElementById('nba-log-pick-inp');
  if(inp){{ inp.value=pick; inp.focus(); inp.setSelectionRange(pick.length, pick.length); }}
  document.getElementById('nba-log-odds-inp').value='';
  // Deselect any highlighted model pick
  _nbaCurrentPicks.forEach(function(_,i){{
    var el=document.getElementById('nba-pick-opt-'+i);
    if(el) el.classList.remove('selected');
  }});
}}

function nbaLogSelectPick(idx){{
  _nbaCurrentPicks.forEach(function(_,i){{
    var el=document.getElementById('nba-pick-opt-'+i);
    if(el) el.classList.toggle('selected', i===idx);
  }});
  var p=_nbaCurrentPicks[idx];
  if(!p) return;
  document.getElementById('nba-log-pick-inp').value=p.pick||'';
  document.getElementById('nba-log-odds-inp').value=p.odds||'';
}}

function nbaLogSelectBook(btn,book){{
  document.querySelectorAll('#nba-log .mlg-book-btn').forEach(function(b){{b.classList.remove('active');}});
  btn.classList.add('active');
  document.querySelector('#f-nba-log [name=book]').value=book;
  document.getElementById('nba-log-odds-manual-wrap').style.display=(book==='Otro'?'block':'none');
}}
</script>
<!-- NBA Log Modal -->
<div class="modal-bg" id="nba-log">
<div class="modal">
  <div class="mlg-hdr">
    <div class="mlg-title">🏀 &nbsp;LOG PICK</div>
    <button type="button" class="mlg-close" onclick="closeModal('nba-log')">✕</button>
  </div>
  <div class="mlg-body">
  <form id="f-nba-log" onsubmit="event.preventDefault();var g=document.getElementById('nba-log-game');if(g&&!g.value){{alert('Selecciona un juego');return;}}submitForm('f-nba-log','/api/nba/log','nba-log')">
    <input type="hidden" name="date" value="{today}">
    <input type="hidden" name="away">
    <input type="hidden" name="home">
    <input type="hidden" name="book" value="DraftKings">
    <div class="mlg-field">
      <div class="mlg-lbl">⚡ Juego del Modelo</div>
      <select id="nba-log-game" onchange="nbaLogFillPick(this)">
        <option value="">Cargando picks...</option>
      </select>
    </div>
    <!-- Model picks selector: shows all picks for selected game -->
    <div class="mlg-pick-opts" id="nba-log-pick-opts"></div>
    <div class="mlg-field">
      <div class="mlg-lbl">🎯 Pick</div>
      <input id="nba-log-pick-inp" name="pick" placeholder="KNICKS ML / OVER 224.5 / WOLVES -5.5" required autocapitalize="characters">
    </div>
    <div class="mlg-row2">
      <div>
        <div class="mlg-lbl">💰 Odds</div>
        <input id="nba-log-odds-inp" name="odds" placeholder="-110" required>
      </div>
      <div>
        <div class="mlg-lbl">💵 Stake $</div>
        <input type="number" name="stake" placeholder="15" step="0.01" min="1" required>
      </div>
    </div>
    <div class="mlg-field">
      <div class="mlg-lbl">📚 Sportsbook</div>
      <div class="mlg-books">
        <button type="button" class="mlg-book-btn active" onclick="nbaLogSelectBook(this,'DraftKings')">DraftKings</button>
        <button type="button" class="mlg-book-btn" onclick="nbaLogSelectBook(this,'FanDuel')">FanDuel</button>
        <button type="button" class="mlg-book-btn" onclick="nbaLogSelectBook(this,'BetMGM')">BetMGM</button>
        <button type="button" class="mlg-book-btn" onclick="nbaLogSelectBook(this,'Caesars')">Caesars</button>
        <button type="button" class="mlg-book-btn" onclick="nbaLogSelectBook(this,'Bet365')">Bet365</button>
        <button type="button" class="mlg-book-btn" onclick="nbaLogSelectBook(this,'Otro')">Otro</button>
      </div>
      <div id="nba-log-odds-manual-wrap">
        <input name="book_manual" placeholder="Nombre del libro...">
      </div>
    </div>
    <div class="mlg-field">
      <div class="mlg-lbl">📝 Análisis</div>
      <textarea name="analysis" placeholder="Razón del pick..."></textarea>
    </div>
    <div class="mlg-footer">
      <button type="button" class="mlg-cancel" onclick="closeModal('nba-log')">Cancelar</button>
      <button type="submit" class="mlg-submit">✅ LOGUEAR PICK</button>
    </div>
  </form>
  </div>
</div></div>

<!-- NBA Grade Modal -->
<div class="modal-bg" id="nba-grade">
<div class="modal">
  <h2>NBA — Grade Picks</h2>
  <div style="max-height:420px;overflow-y:auto">
    {grade_html}
  </div>
  <div class="btn-row" style="margin-top:12px">
    <button type="button" class="btn gray" onclick="closeModal('nba-grade')">Cerrar</button>
    <button type="button" class="btn" style="background:#6366f1"
      onclick="autoGrade('/api/nba/auto-grade','nba-grade')">⚡ Auto-Grade</button>
  </div>
</div></div>

<!-- NBA IR Modal -->
<div class="modal-bg" id="nba-ir">
<div class="modal">
  <h2>NBA — Injury Report</h2>
  <div style="overflow-x:auto;margin-bottom:14px">
    <table style="width:100%;border-collapse:collapse;font-size:0.78rem">
      <thead><tr style="border-bottom:1px solid #1e293b">
        <th style="padding:6px 8px;color:#64748b;text-align:left">Team</th>
        <th style="padding:6px 8px;color:#64748b;text-align:left">Player</th>
        <th style="padding:6px 8px;color:#64748b;text-align:left">Status</th>
        <th style="padding:6px 8px;color:#64748b;text-align:left">PPG</th>
      </tr></thead>
      <tbody>{ir_rows if ir_rows else '<tr><td colspan="4" style="color:#334155;padding:10px;text-align:center">Sin lesionados activos</td></tr>'}</tbody>
    </table>
  </div>
  <div class="section-sep"></div>
  <p style="font-size:0.72rem;color:#64748b;margin-bottom:12px">Stats se buscan automáticamente desde basketball-reference al agregar.</p>
  <form id="f-nba-ir" onsubmit="event.preventDefault();submitForm('f-nba-ir','/api/nba/ir','nba-ir')">
    <input type="hidden" name="action" id="nba-ir-action" value="add">
    <div class="row2">
      <div><label>Team</label><select name="team" required><option value="">—</option>{nba_team_opts}</select></div>
      <div><label>Player (nombre completo)</label><input name="player" placeholder="LEBRON JAMES" autocapitalize="characters" required></div>
    </div>
    <label>Rate</label>
    <select name="rate">
      <option value="1">1 — Superstar (TOP impacto)</option>
      <option value="2">2 — Star</option>
      <option value="3" selected>3 — Role Player</option>
    </select>
    <label>Status</label>
    <select name="status">
      <option value="out">OUT</option>
      <option value="doubtful">Doubtful</option>
      <option value="questionable">Questionable</option>
    </select>
    <div class="btn-row" style="margin-top:4px">
      <button type="button" class="btn gray" onclick="closeModal('nba-ir')">Cerrar</button>
      <button type="button" class="btn red" onclick="document.getElementById('nba-ir-action').value='remove';document.getElementById('f-nba-ir').dispatchEvent(new Event('submit'))">🗑 Remover</button>
      <button type="submit" class="btn green" onclick="document.getElementById('nba-ir-action').value='add'">➕ Agregar</button>
    </div>
  </form>
</div></div>

<!-- NBA Add Game Modal -->
<div class="modal-bg" id="nba-add-game">
<div class="modal">
  <h2>NBA — Agregar Juego</h2>
  <p style="font-size:.72rem;color:#64748b;margin-bottom:14px">
    Usa esto cuando el modelo no detecta un juego automáticamente.<br>
    Abreviaciones: MIN, SAS, NYK, PHI, BOS, MIA, etc.
  </p>
  <form id="f-nba-add-game" onsubmit="event.preventDefault();nbaAddGame()">
    <div class="row2">
      <div>
        <label>Visitante (Away)</label>
        <select name="away" id="nba-ag-away" required>
          <option value="">—</option>{nba_team_opts}
        </select>
      </div>
      <div>
        <label>Local (Home)</label>
        <select name="home" id="nba-ag-home" required>
          <option value="">—</option>{nba_team_opts}
        </select>
      </div>
    </div>
    <label>Hora (opcional)</label>
    <input type="text" id="nba-ag-time" placeholder="7:30 PM ET">
    <div class="btn-row" style="margin-top:14px">
      <button type="button" class="btn gray" onclick="closeModal('nba-add-game')">Cerrar</button>
      <button type="submit" class="btn green">➕ Agregar Juego</button>
    </div>
  </form>
  <div id="nba-ag-result" style="margin-top:10px;font-size:.75rem;color:#10b981"></div>
</div></div>

<script>
async function nbaAddGame(){{
  const away = document.getElementById('nba-ag-away').value;
  const home = document.getElementById('nba-ag-home').value;
  const time = document.getElementById('nba-ag-time').value.trim();
  const res = document.getElementById('nba-ag-result');
  if(!away || !home){{ res.style.color='#ef4444'; res.textContent='Selecciona ambos equipos.'; return; }}
  if(away === home){{ res.style.color='#ef4444'; res.textContent='El visitante y local no pueden ser el mismo equipo.'; return; }}
  res.style.color='#94a3b8'; res.textContent='Agregando...';
  const cmd = time ? `python3 nba.py --add-game ${{away}} ${{home}} "${{time}}"` : `python3 nba.py --add-game ${{away}} ${{home}}`;
  const r = await fetch('/api/run',{{method:'POST',headers:{{'Content-Type':'application/json'}},
    body:JSON.stringify({{cmd,cwd:'NBA'}})}});
  const d = await r.json();
  res.style.color='#10b981';
  res.textContent = `✅ ${{away}} @ ${{home}} agregado. Corre ▶ Run Modelo para generar picks.`;
}}
</script>
"""


# ── MLB PANEL ─────────────────────────────────────────────────────────
def mlb_panel():
    mlb_team_opts = _team_opts(MLB_TEAMS)
    today = date.today().strftime("%Y-%m-%d")
    books = ["DraftKings","FanDuel","BetMGM","Caesars","Bet365","PointsBet","Otro"]
    book_opts = "\n".join(f'<option value="{b}">{b}</option>' for b in books)
    grade_html = _grade_rows_mlb(MLB_LOG, "/api/mlb/grade")

    # Build log pick modal: picks data + game selector + full AI-styled HTML
    import json as _log_json
    today_games = _mlb_today_games()
    _today_lg = date.today().strftime("%Y-%m-%d")
    _picks_by_game = {}
    for _p in _rj(MLB_PICKS):
        if _p.get("date","") != _today_lg:
            continue
        _g = _p.get("game","")
        if _g not in _picks_by_game:
            _picks_by_game[_g] = []
        _picks_by_game[_g].append({
            "pick":   _p.get("pick",""),
            "odds":   _p.get("odds",""),
            "edge":   _p.get("edge",""),
            "ev":     _p.get("ev",""),
            "modelo": _p.get("modelo",""),
            "time":   _p.get("time",""),
        })
    _picks_js = _log_json.dumps(_picks_by_game)

    if today_games:
        _log_game_opts = "\n".join(
            f'<option value="{aw}|{hm}">{gm}</option>'
            for aw, hm, gm in today_games
        )
        log_game_section = (
            '<select id="mlb-log-game" onchange="mlbLogFillPick(this)">'
            '<option value="">— Selecciona juego —</option>'
            + _log_game_opts + '</select>'
        )
        _log_hidden_teams = '<input type="hidden" name="away"><input type="hidden" name="home">'
    else:
        log_game_section = (
            '<div class="mlg-row2" style="margin-bottom:0">'
            '<div><div class="mlg-lbl">Away</div>'
            '<select name="away" required><option value="">—</option>' + mlb_team_opts + '</select></div>'
            '<div><div class="mlg-lbl">Home</div>'
            '<select name="home" required><option value="">—</option>' + mlb_team_opts + '</select></div>'
            '</div>'
        )
        _log_hidden_teams = ''

    _log_modal_html = f"""<style>
#mlb-log .modal{{background:linear-gradient(160deg,#070710 0%,#05050d 100%);border:1px solid rgba(0,220,255,.18);border-radius:20px;padding:0;overflow:hidden;box-shadow:0 0 0 1px rgba(0,220,255,.04),0 24px 80px rgba(0,0,0,.95),inset 0 1px 0 rgba(255,255,255,.05);max-width:480px}}
#mlb-log .mlg-hdr{{background:linear-gradient(90deg,rgba(0,220,255,.07),rgba(124,58,237,.05),transparent);border-bottom:1px solid rgba(0,220,255,.1);padding:16px 20px 14px;display:flex;align-items:center;justify-content:space-between}}
#mlb-log .mlg-title{{font-size:.7rem;font-weight:900;letter-spacing:.2em;background:linear-gradient(90deg,#00dcff,#7c3aed);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
#mlb-log .mlg-close{{background:rgba(0,220,255,.08);border:1px solid rgba(0,220,255,.2)!important;border-radius:50%!important;width:26px;height:26px;color:#00dcff!important;cursor:pointer;font-size:.7rem;display:flex;align-items:center;justify-content:center;padding:0;margin-bottom:0;flex-shrink:0}}
#mlb-log .mlg-body{{padding:16px 20px 20px}}
#mlb-log .mlg-lbl{{font-size:.55rem;font-weight:800;color:rgba(0,220,255,.65);letter-spacing:.14em;text-transform:uppercase;margin-bottom:5px;text-shadow:0 0 8px rgba(0,220,255,.2);display:block}}
#mlb-log select,#mlb-log input:not([type=hidden]),#mlb-log textarea{{background:rgba(0,0,0,.5)!important;border:1px solid rgba(0,220,255,.12)!important;border-radius:9px!important;color:#e2e8f0!important;padding:10px 13px!important;font-size:.88rem!important;width:100%;margin-bottom:0;transition:border-color .2s,box-shadow .2s;font-family:inherit}}
#mlb-log select:focus,#mlb-log input:focus,#mlb-log textarea:focus{{outline:none!important;border-color:rgba(0,220,255,.4)!important;box-shadow:0 0 0 3px rgba(0,220,255,.07)!important}}
#mlb-log .mlg-field{{margin-bottom:13px}}
#mlb-log .mlg-pick-card{{background:linear-gradient(135deg,rgba(0,220,255,.05),rgba(124,58,237,.04));border:1px solid rgba(0,220,255,.2);border-radius:12px;padding:12px 14px;margin-bottom:13px;display:none;position:relative;overflow:hidden}}
#mlb-log .mlg-pick-card::before{{content:'';position:absolute;top:0;left:10%;right:10%;height:1px;background:linear-gradient(90deg,transparent,rgba(0,220,255,.5),transparent)}}
#mlb-log .mlg-pick-name{{font-size:1.05rem;font-weight:800;color:#f1f5f9;letter-spacing:.02em;margin-bottom:7px}}
#mlb-log .mlg-chips{{display:flex;gap:5px;flex-wrap:wrap}}
#mlb-log .mlg-chip{{padding:3px 9px;border-radius:99px;font-size:.6rem;font-weight:700;letter-spacing:.05em}}
.mlg-chip-odds{{background:rgba(0,220,255,.1);border:1px solid rgba(0,220,255,.3);color:#00dcff}}
.mlg-chip-edge{{background:rgba(240,120,32,.1);border:1px solid rgba(240,120,32,.25);color:#f07820}}
.mlg-chip-modelo{{background:rgba(124,58,237,.1);border:1px solid rgba(124,58,237,.25);color:#a78bfa}}
.mlg-chip-ev{{background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.25);color:#22c55e}}
.mlg-chip-time{{background:rgba(100,116,139,.1);border:1px solid rgba(100,116,139,.2);color:#94a3b8}}
#mlb-log .mlg-row2{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:13px}}
#mlb-log .mlg-books{{display:flex;gap:6px;flex-wrap:wrap;margin-top:5px}}
#mlb-log .mlg-book-btn{{background:rgba(0,0,0,.4);border:1px solid rgba(255,255,255,.1)!important;border-radius:8px!important;color:#64748b;font-size:.65rem;font-weight:700;letter-spacing:.04em;padding:7px 11px;cursor:pointer;transition:all .18s;font-family:inherit;margin-bottom:0}}
#mlb-log .mlg-book-btn.active{{background:rgba(0,220,255,.12)!important;border-color:rgba(0,220,255,.4)!important;color:#00dcff!important;box-shadow:0 0 14px rgba(0,220,255,.12)}}
#mlb-log .mlg-book-btn:hover:not(.active){{border-color:rgba(0,220,255,.2)!important;color:#e2e8f0}}
#mlg-odds-manual-wrap{{display:none;margin-top:8px}}
#mlb-log .mlg-footer{{display:flex;gap:8px;margin-top:4px;align-items:stretch}}
#mlb-log .mlg-submit{{flex:1;padding:12px;background:linear-gradient(90deg,#059669,#10b981);border:none!important;border-radius:10px!important;color:#fff!important;font-size:.8rem;font-weight:800;letter-spacing:.1em;cursor:pointer;font-family:inherit;transition:opacity .2s}}
#mlb-log .mlg-submit:hover{{opacity:.85}}
#mlb-log .mlg-cancel{{background:rgba(255,255,255,.04)!important;border:1px solid rgba(255,255,255,.1)!important;border-radius:10px!important;color:#475569!important;font-size:.78rem;font-weight:700;padding:11px 16px;cursor:pointer;font-family:inherit;white-space:nowrap}}
#mlb-log textarea{{height:60px!important;resize:none}}
</style>
<script>
var mlbLogPicksData={{}};

function mlbLogOpen(){{
  // Fetch picks frescos cada vez que se abre el modal
  fetch('/api/mlb/picks-today')
    .then(function(r){{return r.json();}})
    .then(function(d){{
      mlbLogPicksData = d.picks || {{}};
      var sel = document.getElementById('mlb-log-game');
      if(!sel) return;
      var games = Object.keys(mlbLogPicksData);
      // Rebuild options
      sel.innerHTML = '<option value="">— Selecciona juego —</option>';
      games.forEach(function(game){{
        if(game.indexOf(' @ ')<0) return;
        var parts = game.split(' @ ');
        var opt = document.createElement('option');
        opt.value = parts[0]+'|'+parts[1];
        opt.textContent = game;
        sel.appendChild(opt);
      }});
      if(games.length === 0){{
        sel.innerHTML = '<option value="">Sin picks del modelo para hoy</option>';
      }}
    }})
    .catch(function(){{ /* sin conexión — usa lo que está */ }});
  // Reset pick card and alt picks on open
  var card=document.getElementById('mlg-pick-card');
  if(card) card.style.display='none';
  var altEl=document.getElementById('mlb-log-alt-picks');
  if(altEl){{ altEl.innerHTML=''; altEl.style.display='none'; }}
  var pickInp=document.getElementById('mlb-log-pick-inp');
  if(pickInp) pickInp.value='';
  var oddsInp=document.getElementById('mlb-log-odds-inp');
  if(oddsInp) oddsInp.value='';
  openModal('mlb-log');
}}

function mlbLogFillPick(sel){{
  var v=sel.value;if(!v)return;
  var parts=v.split('|');var away=parts[0],home=parts[1];
  document.querySelector('#f-mlb-log [name=away]').value=away;
  document.querySelector('#f-mlb-log [name=home]').value=home;
  var game=away+' @ '+home;
  var picks=mlbLogPicksData[game];
  var card=document.getElementById('mlg-pick-card');
  if(!picks||!picks.length){{card.style.display='none';}}
  else{{
    var p=picks[0];
    var pi=document.getElementById('mlb-log-pick-inp');if(pi)pi.value=p.pick||'';
    var oi=document.getElementById('mlb-log-odds-inp');if(oi)oi.value=p.odds||'';
    document.getElementById('mlg-pick-name').textContent=p.pick||'';
    var chips='';
    if(p.odds)chips+='<span class="mlg-chip mlg-chip-odds">'+p.odds+'</span>';
    if(p.edge)chips+='<span class="mlg-chip mlg-chip-edge">EDGE '+p.edge+'</span>';
    if(p.modelo)chips+='<span class="mlg-chip mlg-chip-modelo">MOD '+p.modelo+'</span>';
    if(p.ev)chips+='<span class="mlg-chip mlg-chip-ev">EV '+p.ev+'</span>';
    if(p.time)chips+='<span class="mlg-chip mlg-chip-time">'+p.time+'</span>';
    document.getElementById('mlg-pick-chips').innerHTML=chips;
    card.style.display='block';
  }}
  // Always render alt pick pills for ML, RL, O/U
  var altEl=document.getElementById('mlb-log-alt-picks');
  if(altEl && away && home){{
    var ac='#00dcff';
    var _sp=function(pick,label,a,suffix){{
      return '<span class="mlg-alt-pill" style="border:1px solid '+a+'40;color:'+a+';background:'+a+'12" '+
        'onclick="mlbLogSetAltPick(\\\''+pick+'\\\')">'+(label||pick)+(suffix||'')+'</span>';
    }};
    var _row=function(lbl,pills){{
      return '<div style="display:flex;gap:5px;flex-wrap:wrap;align-items:center;margin-bottom:5px">'+
        '<span style="font-size:.44rem;font-weight:900;color:#475569;letter-spacing:.1em;min-width:28px;flex-shrink:0">'+lbl+'</span>'+
        pills+'</div>';
    }};
    altEl.innerHTML =
      '<div style="font-size:.44rem;font-weight:900;color:#334155;letter-spacing:.12em;margin-bottom:6px;text-transform:uppercase">Picks Alternativos</div>'+
      _row('ML',  _sp(away+' ML',   away+' ML',   ac) + _sp(home+' ML',   home+' ML',   ac)) +
      _row('RL',  _sp(away+' -1.5', away+' RL',   '#f07820') + _sp(home+' -1.5', home+' RL',   '#f07820')) +
      _row('TOT', _sp('O ',  'OVER',  '#f97316') + _sp('U ',  'UNDER', '#a78bfa'));
    altEl.style.display='block';
  }}
}}

function mlbLogSetAltPick(pick){{
  var inp=document.getElementById('mlb-log-pick-inp');
  if(inp){{ inp.value=pick; inp.focus(); inp.setSelectionRange(pick.length, pick.length); }}
  document.getElementById('mlb-log-odds-inp').value='';
}}
function mlgSelectBook(btn,book){{
  document.querySelectorAll('#mlb-log .mlg-book-btn').forEach(function(b){{b.classList.remove('active');}});
  btn.classList.add('active');
  document.querySelector('#f-mlb-log [name=book]').value=book;
  document.getElementById('mlg-odds-manual-wrap').style.display=(book==='Otro'?'block':'none');
}}
</script>
<!-- MLB Log Modal -->
<div class="modal-bg" id="mlb-log">
<div class="modal">
  <div class="mlg-hdr">
    <div class="mlg-title">⚾ &nbsp;LOG PICK</div>
    <button type="button" class="mlg-close" onclick="closeModal('mlb-log')">✕</button>
  </div>
  <div class="mlg-body">
  <form id="f-mlb-log" onsubmit="event.preventDefault();var g=document.getElementById('mlb-log-game');if(g&&!g.value){{alert('Selecciona un juego');return;}}submitForm('f-mlb-log','/api/mlb/log','mlb-log')">
    <input type="hidden" name="date" value="{today}">
    <input type="hidden" name="away">
    <input type="hidden" name="home">
    <input type="hidden" name="book" value="DraftKings">
    <div class="mlg-field">
      <div class="mlg-lbl">⚡ Juego del Modelo</div>
      <select id="mlb-log-game" onchange="mlbLogFillPick(this)">
        <option value="">Cargando picks...</option>
      </select>
    </div>
    <div class="mlg-pick-card" id="mlg-pick-card">
      <div class="mlg-pick-name" id="mlg-pick-name"></div>
      <div class="mlg-chips" id="mlg-pick-chips"></div>
    </div>
    <div id="mlb-log-alt-picks" style="display:none;background:rgba(0,220,255,.03);border:1px solid rgba(0,220,255,.1);border-radius:12px;padding:12px 14px;margin-bottom:13px"></div>
    <div class="mlg-field">
      <div class="mlg-lbl">🎯 Pick</div>
      <input id="mlb-log-pick-inp" name="pick" placeholder="ROYALS ML / OVER 8.5 / YANKEES -1.5" required autocapitalize="characters">
    </div>
    <div class="mlg-row2">
      <div>
        <div class="mlg-lbl">💰 Odds</div>
        <input id="mlb-log-odds-inp" name="odds" placeholder="-115" required>
      </div>
      <div>
        <div class="mlg-lbl">💵 Stake $</div>
        <input type="number" name="stake" placeholder="15" step="0.01" min="1" required>
      </div>
    </div>
    <div class="mlg-field">
      <div class="mlg-lbl">🏦 Sportsbook</div>
      <div class="mlg-books">
        <button type="button" class="mlg-book-btn active" onclick="mlgSelectBook(this,'DraftKings')">DraftKings</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'FanDuel')">FanDuel</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'BetMGM')">BetMGM</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'Caesars')">Caesars</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'Bet365')">Bet365</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'PointsBet')">PointsBet</button>
        <button type="button" class="mlg-book-btn" onclick="mlgSelectBook(this,'Otro')">✎ Otro</button>
      </div>
      <div id="mlg-odds-manual-wrap">
        <input placeholder="Nombre del sportsbook..." oninput="document.querySelector('#f-mlb-log [name=book]').value=this.value">
      </div>
    </div>
    <div class="mlg-field">
      <div class="mlg-lbl">📝 Análisis (opcional)</div>
      <textarea name="analysis" placeholder="Pitcher matchup, tendencias..."></textarea>
    </div>
    <div class="mlg-footer">
      <button type="button" class="mlg-cancel" onclick="closeModal('mlb-log')">Cancelar</button>
      <button type="submit" class="mlg-submit">✅ &nbsp;LOGUEAR PICK</button>
    </div>
  </form>
  </div>
</div></div>"""

    # Build debug game options from today's model picks
    if today_games:
        debug_opts = "\n".join(
            f'<option value="{away}|{home}">{game}</option>'
            for away, home, game in today_games
        )
        debug_note = f"Juegos recomendados hoy ({len(today_games)})"
        debug_select_html = (
            '<select id="mlb-debug-game">'
            '<option value="">— Selecciona un juego —</option>'
            + debug_opts +
            '</select>'
        )
        debug_onclick = (
            "var sel=document.getElementById('mlb-debug-game').value;"
            "if(!sel){alert('Selecciona un juego');return;}"
            "var parts=sel.split('|');var away=parts[0];var home=parts[1];"
            "closeModal('mlb-debug');"
            "runDebugGame(away,home)"
        )
    else:
        debug_note = "Sin picks para hoy. Selecciona equipos manualmente:"
        debug_select_html = (
            '<div class="row2">'
            '<div><label>Away</label><select id="mlb-debug-away"><option value="">—</option>'
            + mlb_team_opts +
            '</select></div>'
            '<div><label>Home</label><select id="mlb-debug-home"><option value="">—</option>'
            + mlb_team_opts +
            '</select></div></div>'
        )
        debug_onclick = (
            "var a=document.getElementById('mlb-debug-away').value;"
            "var h=document.getElementById('mlb-debug-home').value;"
            "if(!a||!h){alert('Selecciona ambos equipos');return;}"
            "closeModal('mlb-debug');"
            "runDebugGame(a,h)"
        )

    return f"""
{_stats_bar(MLB_LOG, '#e05252', 'MLB', '#e05252', 'rgba(224,82,82,.1)')}

<!-- Today's Picks -->
<div class="section-hdr">
  <span class="section-title">⚾ Picks de Hoy</span>
</div>
<div id="mlb-picks-block">{_mlb_today_section()}</div>

{_panel_actions(
  primary=[
    ('+', 'Log Pick',  "mlbLogOpen()",                                                                                                                                      'green'),
    ('▶', 'Run Modelo',"refreshPicksBlock('mlb-picks-block','bash -c &quot;rm -f mlb_debug_body_current.html &amp;&amp; python3 mlb.py --picks&quot;','MLB','/api/view/mlb/picks')", 'blue'),
  ],
  secondary=[
    ('📋', 'Lines',    "runThenView('python3 mlb.py --lines','MLB','/api/view/mlb/lines','MLB · Lines')"),
    ('📊', 'Stats',    "openView('/api/view/mlb/stats','MLB · Stats')"),
    ('⛅', 'Weather',  "openView('/api/view/mlb/weather','MLB · Weather')"),
    ('✓',  'Grade',   "openModal('mlb-grade')"),
    ('📜', 'Historial',"openView('/api/view/mlb/log','MLB · Historial')"),
    ('🖼', 'Record',  "openRecordModal('MLB','mlb.py','MLB','#22c55e')"),
    ('🔍', 'Debug',   "openModal('mlb-debug')"),
    ('📈', 'Feedback', "runInView('python3 mlb.py --feedback','MLB','MLB · Feedback')"),
  ]
)}

<!-- Tools & Publish Accordion -->
<div class="tools-group">
  <div class="tools-toggle" onclick="toggleTools('mlb')">
    <span>⚙ Tools &amp; Publishing</span>
    <span class="tools-chevron" id="chev-mlb">›</span>
  </div>
  <div class="tools-body" id="tools-mlb">
    {_cmd_card('Correr modelo y ver output completo', '⚙ Run Modelo', "runInView('python3 mlb.py --picks','MLB','MLB · Modelo')", 'gray')}
    {_cmd_card('Actualizar datos FanGraphs (wRC+/xFIP) antes de picks', '🔄 Refresh FG', "runCmd('python3 mlb.py --refresh','MLB','MLB → Refresh')", 'gray')}
    {_cmd_card('Publicar picks + debug a GitHub Pages (un click)', '🚀 Publish All', "mlbPublishAll()", 'blue')}
    {_cmd_card('Publicar líneas del modelo', '🌐 Publish Lines', "runCmd('python3 mlb.py --export-lines --publish','MLB','MLB → Lines')", 'blue')}
    {_cmd_card('Publicar log de picks', '📤 Publish Log', "mlbPublishLog()", 'blue')}
  </div>
</div>

{_log_modal_html}

<!-- MLB Grade Modal -->
<div class="modal-bg" id="mlb-grade">
<div class="modal">
  <h2>MLB — Grade Picks</h2>
  <p style="font-size:0.72rem;color:#64748b;margin-bottom:10px">Para O/U: ingresa <b>closing line</b> y <b>runs reales</b> para tracking CLV. O usa ⚡ Auto-Grade.</p>
  <div style="max-height:420px;overflow-y:auto">
    {grade_html}
  </div>
  <div class="btn-row" style="margin-top:12px">
    <button type="button" class="btn gray" onclick="closeModal('mlb-grade')">Cerrar</button>
    <button type="button" class="btn" style="background:#6366f1"
      onclick="autoGrade('/api/mlb/auto-grade','mlb-grade')">⚡ Auto-Grade</button>
  </div>
</div></div>

<!-- MLB Debug Modal -->
<div class="modal-bg" id="mlb-debug">
<div class="modal">
  <h2>MLB — Debug Game</h2>
  <p style="font-size:0.8rem;color:#64748b;margin-bottom:14px">{debug_note}</p>
  {debug_select_html}
  <div class="btn-row" style="margin-top:8px">
    <button type="button" class="btn gray" onclick="closeModal('mlb-debug')">Cancelar</button>
    <button type="button" class="btn" onclick="{debug_onclick}">▶ Analizar</button>
  </div>
</div></div>
"""


# ══════════════════════════════════════════════════════════════════════
# FULL PAGE
# ══════════════════════════════════════════════════════════════════════
def full_page(alert="", alert_type=""):
    bsn_w, bsn_l, _, _, _ = _quick_stats(BSN_LOG)
    nba_w, nba_l, _, _, _ = _quick_stats(NBA_LOG)
    mlb_w, mlb_l, _, _, _ = _quick_stats(MLB_LOG)

    today_str  = date.today().strftime("%B %d, %Y").upper()
    alert_html = (f'<div style="max-width:860px;margin:12px auto 0;padding:0 20px">'
                  f'<div class="alert {alert_type}">{alert}</div></div>') if alert else ""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Laboy Picks · Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<style>{CSS}</style>
</head>
<body>

<!-- ── Header ── -->
<header class="site-hdr">
  <div class="hdr-inner">
    <div class="brand">
      <div class="brand-dot"></div>
      <span class="brand-name">LABOY <em>PICKS</em></span>
    </div>
    <div class="hdr-right">
      <button class="btn gray" style="padding:5px 11px;font-size:.65rem;font-weight:800;letter-spacing:.05em"
        onclick="openView('/api/view/calendar','CALENDARIO P&amp;L')">P&amp;L CAL</button>
      <button class="btn gray" style="padding:5px 11px;font-size:.65rem;font-weight:800;letter-spacing:.05em"
        onclick="openView('/api/view/record/all','RECORD ALL-TIME')">RECORD</button>
      <div class="live-badge"><div class="live-dot"></div>LIVE</div>
      <span class="hdr-date">{today_str}</span>
    </div>
  </div>
</header>

<!-- ── League Nav ── -->
<nav class="lg-nav">
  <div class="lg-nav-inner">
    <div class="tab active" data-tab="bsn" onclick="showTab('bsn')">
      <span class="tab-icon">🏀</span>
      <span class="tab-name">BSN</span>
      <span class="tab-rec">{bsn_w}-{bsn_l}</span>
    </div>
    <div class="tab" data-tab="nba" onclick="showTab('nba')">
      <span class="tab-icon">🏀</span>
      <span class="tab-name">NBA</span>
      <span class="tab-rec">{nba_w}-{nba_l}</span>
    </div>
    <div class="tab" data-tab="mlb" onclick="showTab('mlb')">
      <span class="tab-icon">⚾</span>
      <span class="tab-name">MLB</span>
      <span class="tab-rec">{mlb_w}-{mlb_l}</span>
    </div>
  </div>
</nav>

{alert_html}

<!-- ── Main Content ── -->
<main class="site-main">
  <div class="panel active" id="panel-bsn">{bsn_panel()}</div>
  <div class="panel" id="panel-nba">{nba_panel()}</div>
  <div class="panel" id="panel-mlb">{mlb_panel()}</div>
</main>

<!-- ── Detail Panel (slide-in) ── -->
<div id="detail-overlay" class="detail-overlay" onclick="closeDetail()"></div>
<div id="detail-panel" class="detail-panel">
  <div class="detail-hdr">
    <span class="detail-title" id="detail-title">Detail</span>
    <button class="btn gray" style="padding:6px 14px;font-size:.7rem" onclick="closeDetail()">✕ Cerrar</button>
  </div>
  <div class="detail-body" id="detail-body"></div>
</div>

<!-- ── Output Panel (Terminal — for background tasks) ── -->
<div id="out-bg" class="out-bg" onclick="if(event.target===this)closeOutput()">
  <div class="out-panel">
    <div class="out-header">
      <div class="out-hdr-left">
        <div class="out-dot"></div>
        <span class="out-title-lbl" id="out-title">Output</span>
      </div>
      <button class="btn gray" style="padding:6px 14px;font-size:.7rem" onclick="closeOutput()">✕ Cerrar</button>
    </div>
    <div class="out-body" id="out-body"></div>
  </div>
</div>

<!-- ── Record Date Picker Modal (shared: BSN / NBA / MLB) ── -->
<div class="modal-bg" id="modal-record" onclick="if(event.target===this)closeModal('modal-record')">
<div class="modal" style="max-width:440px;position:relative;overflow:hidden">

  <!-- accent gradient bar (color set dynamically) -->
  <div id="rec-accent-bar" style="position:absolute;top:0;left:0;right:0;height:3px;background:#f07820"></div>

  <!-- header -->
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:22px">
    <div>
      <div id="rec-league-label" style="font-size:.52rem;font-weight:900;letter-spacing:.22em;color:#f07820;text-transform:uppercase;margin-bottom:4px">BSN</div>
      <div style="font-size:.9rem;font-weight:900;color:#f1f5f9;letter-spacing:.04em">🖼 Record Card</div>
      <div style="font-size:.62rem;color:#475569;margin-top:3px">Genera el JPG del record y publica</div>
    </div>
    <button onclick="closeModal('modal-record')" style="background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.09);border-radius:8px;padding:7px 12px;color:#475569;cursor:pointer;font-size:.72rem">✕</button>
  </div>

  <!-- date input -->
  <label>Fecha del record</label>
  <input type="date" id="rec-date" style="font-size:1rem;font-weight:700;text-align:center;cursor:pointer;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.12);border-radius:12px;padding:14px">

  <!-- quick shortcuts -->
  <div style="margin:-4px 0 18px;display:flex;gap:7px;flex-wrap:wrap">
    <button onclick="recSetDate(0)" class="rec-chip" id="chip-hoy">Hoy</button>
    <button onclick="recSetDate(-1)" class="rec-chip" id="chip-ayer">Ayer</button>
    <button onclick="recSetDate(-2)" class="rec-chip" id="chip-d2"></button>
    <button onclick="recSetDate(-3)" class="rec-chip" id="chip-d3"></button>
    <button onclick="recSetDate(-4)" class="rec-chip" id="chip-d4"></button>
    <button onclick="recSetDate(-5)" class="rec-chip" id="chip-d5"></button>
  </div>

  <!-- action buttons -->
  <div class="btn-row">
    <button class="btn gray" onclick="closeModal('modal-record')">Cancelar</button>
    <button id="rec-submit-btn" class="btn accent" style="flex:2;font-size:.82rem;font-weight:900"
      onclick="submitRecordModal()">⚡ Generar & Publicar</button>
  </div>

</div>
</div>

<style>
.rec-chip{{
  background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.09);
  border-radius:99px;padding:5px 13px;font-size:.62rem;font-weight:800;
  letter-spacing:.06em;text-transform:uppercase;color:#64748b;cursor:pointer;
  font-family:inherit;transition:all .15s;
}}
.rec-chip:hover{{background:rgba(255,255,255,.09);color:#cbd5e1}}
.rec-chip.active{{border-color:var(--rec-accent,#f07820);color:var(--rec-accent,#f07820);background:rgba(240,120,32,.08)}}
</style>

<script>{JS}</script>
{_MC_MODAL_HTML}
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════
# API HANDLERS
# ══════════════════════════════════════════════════════════════════════
def _log_pick(log_path, data):
    """Write pick entry to JSON log. Returns (ok, msg)."""
    t1  = data.get("team1","").strip().upper() or data.get("away","").strip().upper()
    t2  = data.get("team2","").strip().upper() or data.get("home","").strip().upper()
    pick= data.get("pick","").strip().upper()
    odds_s = data.get("odds","")
    stake_s= data.get("stake","")
    book   = data.get("book","BetMGM")
    analysis = data.get("analysis","")
    d_val  = data.get("date", date.today().strftime("%Y-%m-%d"))

    if not t1 or not t2: return False, "⚠️ Selecciona ambos equipos."
    if t1 == t2:         return False, "⚠️ Los equipos no pueden ser iguales."
    if not pick:         return False, "⚠️ El pick es requerido."
    if not odds_s:       return False, "⚠️ Los odds son requeridos."
    if not stake_s:      return False, "⚠️ La apuesta es requerida."

    try:
        odds_v  = _parse_odds(odds_s)
        stake_v = float(re.sub(r"[^\d.]","", stake_s.split()[0]))
        game    = f"{t1} vs. {t2}"
        log     = _rj(log_path)
        entry   = {"id": len(log), "date": d_val, "game": game,
                   "pick": pick, "odds": odds_v, "stake": stake_v,
                   "book": book, "result": None, "pnl": None, "analysis": analysis}
        log.append(entry)
        _wj(log_path, log)
        return True, f"✅ Pick #{entry['id']} logueado — {game} | {pick} | {_fmt_odds(odds_v)}"
    except Exception as ex:
        return False, f"⚠️ Error: {ex}"


def _log_parlay(log_path, data):
    odds_s = data.get("odds","")
    stake_s= data.get("stake","")
    book   = data.get("book","BetMGM")
    d_val  = data.get("date", date.today().strftime("%Y-%m-%d"))
    analysis = data.get("analysis","")

    if not odds_s:  return False, "⚠️ Los odds son requeridos."
    if not stake_s: return False, "⚠️ La apuesta es requerida."

    # Parse legs
    legs = []
    i = 1
    while True:
        game = data.get(f"leg{i}_game","").strip()
        pick = data.get(f"leg{i}_pick","").strip().upper()
        if not game and not pick: break
        if game or pick:
            legs.append({"game": game, "pick": pick})
        i += 1
    if len(legs) < 2: return False, "⚠️ Se necesitan al menos 2 legs."

    try:
        odds_v  = _parse_odds(odds_s)
        stake_v = float(re.sub(r"[^\d.]","", stake_s.split()[0]))
        log     = _rj(log_path)
        entry   = {"id": len(log), "date": d_val, "type": "parlay",
                   "legs": legs, "odds": odds_v, "stake": stake_v,
                   "book": book, "result": None, "pnl": None, "analysis": analysis}
        log.append(entry)
        _wj(log_path, log)
        return True, f"✅ Parlay #{entry['id']} ({len(legs)} legs) logueado"
    except Exception as ex:
        return False, f"⚠️ Error: {ex}"


# ══════════════════════════════════════════════════════════════════════
# RICH VIEW RENDERERS  (return HTML strings for the side panel)
# ══════════════════════════════════════════════════════════════════════

def _mlb_picks_debug_view():
    """Return the rich debug fragment (export_debug_picks_html body) if available for today.
    Falls back to simple picks cards if not yet generated."""
    today_str = date.today().strftime("%Y-%m-%d")
    frag_path = os.path.join(MLB_DIR, "mlb_debug_body_current.html")
    if os.path.exists(frag_path):
        try:
            mtime = os.path.getmtime(frag_path)
            from datetime import datetime as _dt
            mdate = _dt.fromtimestamp(mtime).strftime("%Y-%m-%d")
            if mdate == today_str:
                with open(frag_path, "r", encoding="utf-8") as f:
                    return f.read()
        except Exception:
            pass
    # Fallback: simple pick cards
    return _render_picks_html(MLB_PICKS, "#e05252")


def _render_picks_html(picks_path, accent="#f07820"):
    """Render today's model picks from a JSON file as pick cards."""
    today_str = date.today().strftime("%Y-%m-%d")
    picks     = _rj(picks_path) if picks_path and os.path.exists(picks_path) else []
    today_p   = [p for p in picks if p.get("date","") == today_str]

    if not today_p:
        return ('<div class="detail-empty">'
                'No hay picks del modelo para hoy.<br>'
                '<span style="font-size:.7rem">Ejecuta el modelo primero.</span>'
                '</div>')

    seen = set()
    html = (f'<div class="sec-lbl" style="color:{accent};margin-bottom:12px">'
            f'{len(today_p)} PICK(S) · {today_str}</div>')

    for p in today_p:
        game = p.get("game","")
        if game in seen: continue
        seen.add(game)

        pick  = p.get("pick","")
        odds  = _fmt_odds(p.get("odds",0))
        ev    = p.get("edge","") or p.get("ev","") or p.get("expected_value","")
        sp    = p.get("sp","") or p.get("pitcher","") or p.get("starter","")
        note  = p.get("analysis","") or p.get("note","") or p.get("rationale","")

        ev_html   = f'<span class="pc-ev">EV+ {_esc(str(ev))}%</span>' if ev else ""
        sp_html   = f'<span class="pc-tag">{_esc(str(sp))}</span>' if sp else ""
        note_html = f'<div class="pc-note">{_esc(str(note))}</div>' if note else ""

        html += (f'<div class="pick-card" style="--acc:{accent}">'
                 f'<div class="pc-game">{_esc(game)}</div>'
                 f'<div class="pc-pick">{_esc(pick)}</div>'
                 f'<div class="pc-row"><span class="pc-odds">{odds}</span>{ev_html}{sp_html}</div>'
                 f'{note_html}'
                 f'</div>')
    return html


def _render_log_html(log_path, limit=50):
    """Render recent pick history from a log JSON file."""
    log = _rj(log_path)
    if not log:
        return '<div class="detail-empty">Sin picks en el historial.</div>'

    recent = list(reversed(log[-limit:]))
    total  = len(log)
    html   = (f'<div class="sec-lbl" style="margin-bottom:12px">'
              f'HISTORIAL · {min(limit,total)} DE {total} PICKS</div>')

    for e in recent:
        res = e.get("result")
        if   res == "W": rc, rl = "le-w", "W"
        elif res == "L": rc, rl = "le-l", "L"
        elif res == "P": rc, rl = "le-p", "P"
        else:            rc, rl = "le-n", "·"

        pnl = e.get("pnl")
        if pnl is not None and res in ("W","L","P"):
            pc      = "le-pos" if pnl >= 0 else "le-neg"
            sign    = "+" if pnl >= 0 else ""
            pnl_s   = f' · <span class="{pc}">{sign}${abs(pnl):.2f}</span>'
        else:
            pnl_s = ""

        if e.get("type") == "parlay":
            legs      = e.get("legs",[])
            pick_text = f"Parlay ({len(legs)} legs)"
        else:
            pick_text = e.get("pick","—")

        odds_s  = _fmt_odds(e.get("odds",0))
        stake_s = f"${e.get('stake',0):.2f}"
        book    = e.get("book","")
        book_s  = f" · {book}" if book else ""

        html += (f'<div class="log-entry">'
                 f'<div class="le-result {rc}">{rl}</div>'
                 f'<div class="le-body">'
                 f'<div class="le-game">{_esc(e.get("game","—"))} · {e.get("date","")}</div>'
                 f'<div class="le-pick">{_esc(pick_text)}</div>'
                 f'<div class="le-meta">{odds_s} · {stake_s}{book_s}{pnl_s}</div>'
                 f'</div></div>')
    return html


def _pnl_calendar_html():
    """P&L calendar — daily profit/loss across all leagues with filters."""
    import json as _json

    def _load(path):
        try:
            d = _rj(path)
            return d if isinstance(d, list) else []
        except Exception:
            return []

    bsn_raw = _load(BSN_LOG)
    nba_raw = _load(NBA_LOG)
    mlb_raw = _load(MLB_LOG)

    def _norm(entries, league):
        out = []
        for e in entries:
            r = e.get("result","")
            if r not in ("W","L","P"): continue
            out.append({
                "date":   e.get("date",""),
                "pnl":    float(e.get("pnl") or 0),
                "result": r,
                "pick":   str(e.get("pick","")),
                "game":   str(e.get("game","")),
                "odds":   e.get("odds",""),
                "stake":  float(e.get("stake") or 0),
                "league": league,
            })
        return out

    all_picks = _norm(bsn_raw,"BSN") + _norm(nba_raw,"NBA") + _norm(mlb_raw,"MLB")
    picks_json = _json.dumps(all_picks, ensure_ascii=False)

    # Build HTML — JS is in a separate <script> block that openView will re-execute
    html = """
<style>
#calRoot{padding:0 0 24px}
.cal-toolbar{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-bottom:18px;flex-wrap:wrap}
.cal-filter-row{display:flex;gap:5px;flex-wrap:wrap}
.cal-pill{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.09);color:#64748b;
  font-size:.58rem;font-weight:800;letter-spacing:.09em;padding:5px 14px;border-radius:4px;
  cursor:pointer;transition:all .14s;text-transform:uppercase;user-select:none}
.cal-pill:hover{border-color:#f07820;color:#f07820;background:rgba(240,120,32,.07)}
.cal-pill.active{background:#f07820;border-color:#f07820;color:#000;box-shadow:0 0 14px rgba(240,120,32,.35)}
.cal-nav{display:flex;align-items:center;gap:12px;margin-bottom:14px}
.cal-nav-btn{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.09);color:#94a3b8;
  width:32px;height:32px;border-radius:4px;cursor:pointer;font-size:1.1rem;display:flex;
  align-items:center;justify-content:center;transition:all .14s;flex-shrink:0}
.cal-nav-btn:hover{background:rgba(240,120,32,.12);border-color:#f07820;color:#f07820}
.cal-month-label{font-size:.82rem;font-weight:900;letter-spacing:.14em;text-transform:uppercase;
  color:#f1f5f9;flex:1;text-align:center}
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:3px}
.cal-dow{text-align:center;font-size:.48rem;font-weight:800;letter-spacing:.12em;color:#334155;
  padding:7px 0 5px;text-transform:uppercase}
.cal-cell{border-radius:6px;padding:7px 6px 7px;min-height:76px;
  border:1px solid rgba(255,255,255,.05);background:rgba(255,255,255,.02);
  cursor:pointer;transition:all .15s;display:flex;flex-direction:column;position:relative}
.cal-cell:hover{border-color:rgba(240,120,32,.35);background:rgba(240,120,32,.05);transform:translateY(-1px);z-index:1}
.cal-cell.empty{background:transparent;border-color:transparent;cursor:default;pointer-events:none}
.cal-cell.is-today{border-color:rgba(240,120,32,.45)!important;box-shadow:inset 0 0 0 1px rgba(240,120,32,.15)}
.cal-cell.is-selected{outline:2px solid #f07820;outline-offset:-1px}
.cal-cell.win-day{background:linear-gradient(145deg,rgba(34,197,94,.11) 0%,rgba(34,197,94,.02) 100%);border-color:rgba(34,197,94,.22)}
.cal-cell.loss-day{background:linear-gradient(145deg,rgba(239,68,68,.11) 0%,rgba(239,68,68,.02) 100%);border-color:rgba(239,68,68,.22)}
.cal-cell.push-day{border-color:rgba(148,163,184,.15)}
.cal-day-num{font-size:.6rem;font-weight:700;color:#334155;margin-bottom:5px;line-height:1}
.cal-cell.is-today .cal-day-num{color:#f07820}
.cal-pnl{font-size:.8rem;font-weight:900;line-height:1;margin-bottom:2px}
.cal-pnl.pos{color:#22c55e}.cal-pnl.neg{color:#ef4444}.cal-pnl.zero{color:#64748b}
.cal-rec{font-size:.5rem;font-weight:700;color:#475569;letter-spacing:.03em}
.cal-ldots{display:flex;gap:3px;margin-top:auto;padding-top:4px;flex-wrap:wrap}
.cal-ldot{width:4px;height:4px;border-radius:50%}
.cal-summary{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-top:14px}
.cal-sum-card{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);
  border-radius:8px;padding:13px 10px;text-align:center}
.cal-sum-lbl{font-size:.48rem;color:#475569;text-transform:uppercase;letter-spacing:.12em;margin-bottom:6px;font-weight:800}
.cal-sum-val{font-size:1.1rem;font-weight:900;line-height:1}
.cal-detail{margin-top:14px;border:1px solid rgba(255,255,255,.07);border-radius:8px;overflow:hidden}
.cal-det-hdr{padding:10px 14px;background:rgba(255,255,255,.04);border-bottom:1px solid rgba(255,255,255,.06);
  font-size:.6rem;font-weight:800;letter-spacing:.1em;color:#64748b;text-transform:uppercase;
  display:flex;align-items:center;justify-content:space-between}
.cal-det-hdr-pnl{font-size:.75rem;font-weight:900}
.cal-det-row{display:flex;align-items:center;gap:9px;padding:9px 14px;border-bottom:1px solid rgba(255,255,255,.04)}
.cal-det-row:last-child{border-bottom:none}
.cal-det-badge{font-size:.5rem;font-weight:800;letter-spacing:.1em;padding:2px 8px;border-radius:3px;flex-shrink:0;min-width:26px;text-align:center}
.cal-det-badge.W{background:rgba(34,197,94,.16);color:#22c55e}
.cal-det-badge.L{background:rgba(239,68,68,.16);color:#ef4444}
.cal-det-badge.P{background:rgba(148,163,184,.1);color:#94a3b8}
.cal-det-info{flex:1;min-width:0}
.cal-det-pick-name{font-size:.7rem;font-weight:700;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cal-det-game{font-size:.58rem;color:#475569;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px}
.cal-det-lg{font-size:.46rem;font-weight:800;letter-spacing:.1em;padding:2px 6px;border-radius:3px;
  background:rgba(240,120,32,.12);color:#f07820;flex-shrink:0}
.cal-det-pnl{font-size:.72rem;font-weight:800;flex-shrink:0;min-width:54px;text-align:right}
</style>

<div id="calRoot">
  <div class="cal-toolbar">
    <div class="cal-filter-row" id="calLeagueFilters">
      <span class="cal-pill active" data-league="ALL">ALL</span>
      <span class="cal-pill" data-league="BSN">BSN</span>
      <span class="cal-pill" data-league="NBA">NBA</span>
      <span class="cal-pill" data-league="MLB">MLB</span>
    </div>
    <div class="cal-filter-row" id="calRangeFilters">
      <span class="cal-pill" data-range="7">7D</span>
      <span class="cal-pill" data-range="30">30D</span>
      <span class="cal-pill" data-range="90">90D</span>
      <span class="cal-pill active" data-range="0">TODO</span>
    </div>
  </div>
  <div class="cal-nav">
    <button class="cal-nav-btn" id="calPrev">&#8249;</button>
    <div class="cal-month-label" id="calMonthLabel">—</div>
    <button class="cal-nav-btn" id="calNext">&#8250;</button>
  </div>
  <div class="cal-grid" id="calGrid">
    <div class="cal-dow">DOM</div><div class="cal-dow">LUN</div>
    <div class="cal-dow">MAR</div><div class="cal-dow">MIE</div>
    <div class="cal-dow">JUE</div><div class="cal-dow">VIE</div>
    <div class="cal-dow">SAB</div>
  </div>
  <div class="cal-summary" id="calSummary"></div>
  <div class="cal-detail" id="calDetail" style="display:none"></div>
</div>
""" + "<script>(function(){\n" + f"var _PICKS={picks_json};\n" + r"""
var _league='ALL', _range=0, _selDate=null;
var _td=new Date(); _td.setHours(0,0,0,0);
var _vy=_td.getFullYear(), _vm=_td.getMonth();
var _LC={BSN:'#a855f7',NBA:'#f07820',MLB:'#22c55e'};
var _MO=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'];

function _filtered(){
  var p=_PICKS.filter(function(x){return _league==='ALL'||x.league===_league;});
  if(_range>0){
    var co=new Date(_td); co.setDate(co.getDate()-_range+1);
    var cs=co.toISOString().slice(0,10);
    p=p.filter(function(x){return x.date>=cs;});
  }
  return p;
}
function _byDate(picks){
  var m={};
  picks.forEach(function(p){if(p.date){if(!m[p.date])m[p.date]=[];m[p.date].push(p);}});
  return m;
}
function _ps(v){return(v>=0?'+$':'-$')+Math.abs(v).toFixed(2);}

function _render(){
  var picks=_filtered(), bd=_byDate(picks);
  document.getElementById('calMonthLabel').textContent=_MO[_vm]+' '+_vy;
  var tdStr=_td.toISOString().slice(0,10);
  var fd=new Date(_vy,_vm,1).getDay();
  var dim=new Date(_vy,_vm+1,0).getDate();
  var grid=document.getElementById('calGrid');
  while(grid.children.length>7)grid.removeChild(grid.lastChild);
  for(var i=0;i<fd;i++){var e=document.createElement('div');e.className='cal-cell empty';grid.appendChild(e);}
  for(var d=1;d<=dim;d++){
    var ds=_vy+'-'+String(_vm+1).padStart(2,'0')+'-'+String(d).padStart(2,'0');
    var dp=bd[ds]||[];
    var cell=document.createElement('div');
    var cls='cal-cell';
    if(ds===tdStr)cls+=' is-today';
    if(ds===_selDate)cls+=' is-selected';
    var tot=0,w=0,l=0,pu=0,lg={};
    dp.forEach(function(p){tot+=p.pnl;if(p.result==='W')w++;else if(p.result==='L')l++;else pu++;lg[p.league]=1;});
    if(dp.length>0){cls+=tot>0?' win-day':tot<0?' loss-day':' push-day';}
    cell.className=cls;
    var dots=Object.keys(lg).map(function(k){return '<div class="cal-ldot" style="background:'+_LC[k]+'"></div>';}).join('');
    var pc=tot>0?'pos':tot<0?'neg':'zero';
    cell.innerHTML='<div class="cal-day-num">'+d+'</div>'+
      (dp.length?'<div class="cal-pnl '+pc+'">'+_ps(tot)+'</div>':'')+
      (dp.length?'<div class="cal-rec">'+w+'W-'+l+'L'+(pu?'-'+pu+'P':'')+'</div>':'')+
      '<div class="cal-ldots">'+dots+'</div>';
    (function(date,dpicks){cell.addEventListener('click',function(){_selDate=date;_renderDetail(date,dpicks);_render();});})(ds,dp);
    grid.appendChild(cell);
  }
  var tp=0,tw=0,tl=0,tpu=0,ts=0;
  picks.forEach(function(p){tp+=p.pnl;ts+=p.stake;if(p.result==='W')tw++;else if(p.result==='L')tl++;else tpu++;});
  var roi=ts>0?(tp/ts*100):0;
  var pc=tp>0?'#22c55e':tp<0?'#ef4444':'#64748b';
  var rc=roi>0?'#22c55e':roi<0?'#ef4444':'#64748b';
  document.getElementById('calSummary').innerHTML=
    '<div class="cal-sum-card"><div class="cal-sum-lbl">P&L</div><div class="cal-sum-val" style="color:'+pc+'">'+_ps(tp)+'</div></div>'+
    '<div class="cal-sum-card"><div class="cal-sum-lbl">Record</div><div class="cal-sum-val" style="color:#f1f5f9;font-size:.88rem">'+tw+'W &middot; '+tl+'L'+(tpu?' &middot; '+tpu+'P':'')+'</div></div>'+
    '<div class="cal-sum-card"><div class="cal-sum-lbl">ROI</div><div class="cal-sum-val" style="color:'+rc+'">'+roi.toFixed(1)+'%</div></div>'+
    '<div class="cal-sum-card"><div class="cal-sum-lbl">Picks</div><div class="cal-sum-val" style="color:#f07820">'+(tw+tl+tpu)+'</div></div>';
}

function _renderDetail(ds,picks){
  var el=document.getElementById('calDetail');
  if(!picks||!picks.length){el.style.display='none';return;}
  el.style.display='block';
  var tot=picks.reduce(function(a,p){return a+p.pnl;},0);
  var pc=tot>0?'#22c55e':tot<0?'#ef4444':'#64748b';
  var rows=picks.map(function(p){
    var ppc=p.pnl>0?'#22c55e':p.pnl<0?'#ef4444':'#64748b';
    return '<div class="cal-det-row">'+
      '<div class="cal-det-badge '+p.result+'">'+p.result+'</div>'+
      '<div class="cal-det-info">'+
        '<div class="cal-det-pick-name">'+p.pick+'</div>'+
        '<div class="cal-det-game">'+p.game+' &nbsp;&middot;&nbsp; '+p.odds+'</div>'+
      '</div>'+
      '<div class="cal-det-lg">'+p.league+'</div>'+
      '<div class="cal-det-pnl" style="color:'+ppc+'">'+_ps(p.pnl)+'</div>'+
    '</div>';
  }).join('');
  el.innerHTML='<div class="cal-det-hdr"><span>'+ds+' &nbsp;&middot;&nbsp; '+picks.length+' pick'+(picks.length!==1?'s':'')+'</span>'+
    '<span class="cal-det-hdr-pnl" style="color:'+pc+'">'+_ps(tot)+'</span></div>'+rows;
}

// Wire filters
document.querySelectorAll('#calLeagueFilters .cal-pill').forEach(function(el){
  el.addEventListener('click',function(){
    _league=this.dataset.league;
    document.querySelectorAll('#calLeagueFilters .cal-pill').forEach(function(b){b.classList.remove('active');});
    this.classList.add('active');
    _selDate=null;document.getElementById('calDetail').style.display='none';_render();
  });
});
document.querySelectorAll('#calRangeFilters .cal-pill').forEach(function(el){
  el.addEventListener('click',function(){
    _range=parseInt(this.dataset.range||'0');
    document.querySelectorAll('#calRangeFilters .cal-pill').forEach(function(b){b.classList.remove('active');});
    this.classList.add('active');
    _vy=_td.getFullYear();_vm=_td.getMonth();
    _selDate=null;document.getElementById('calDetail').style.display='none';_render();
  });
});
document.getElementById('calPrev').addEventListener('click',function(){
  _vm--;if(_vm<0){_vm=11;_vy--;}
  _selDate=null;document.getElementById('calDetail').style.display='none';_render();
});
document.getElementById('calNext').addEventListener('click',function(){
  _vm++;if(_vm>11){_vm=0;_vy++;}
  _selDate=null;document.getElementById('calDetail').style.display='none';_render();
});
_render();
})();</script>"""
    return html

def _alltime_record_html():
    """All-leagues all-time record — styled like the league record cards."""

    _RC_CSS = """<style>
.rc-pick{border-radius:12px;padding:11px 14px;margin-bottom:8px;position:relative;
  overflow:hidden;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07)}
.rc-pick.win{border-left:4px solid #22c55e;background:linear-gradient(135deg,rgba(34,197,94,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.loss{border-left:4px solid #ef4444;background:linear-gradient(135deg,rgba(239,68,68,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.push{border-left:4px solid #94a3b8;background:rgba(255,255,255,.04)}
.rc-pick.pending{border-left:4px solid #f59e0b;background:rgba(255,255,255,.04)}
.rc-row{display:flex;align-items:center;gap:10px}
.rc-main{flex:1;min-width:0}
.rc-pick-name{font-size:.78rem;font-weight:700;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-odds{font-size:.68rem;color:#64748b;margin-left:5px;font-weight:400}
.rc-game{font-size:.65rem;color:#64748b;margin-top:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-result-col{display:flex;flex-direction:column;align-items:flex-end;gap:3px;flex-shrink:0}
.rc-badge{font-size:.58rem;font-weight:700;letter-spacing:.1em;padding:2px 7px;border-radius:20px}
.rc-badge.win{background:rgba(34,197,94,.18);color:#22c55e}
.rc-badge.loss{background:rgba(239,68,68,.18);color:#ef4444}
.rc-badge.push{background:rgba(148,163,184,.12);color:#94a3b8}
.rc-badge.pending{background:rgba(245,158,11,.12);color:#f59e0b}
.rc-pnl{font-size:.75rem;font-weight:700}
.rc-pnl.win{color:#22c55e}.rc-pnl.loss{color:#ef4444}.rc-pnl.push{color:#94a3b8}
.rc-stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;text-align:center;margin-top:10px}
.rc-stat-lbl{font-size:.52rem;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:3px}
.rc-stat-val{font-size:.95rem;font-weight:800;line-height:1}
</style>"""

    def _league_stats(log_path):
        log = _rj(log_path)
        if not isinstance(log, list): log = []
        settled = [e for e in log if e.get("result") in ("W","L","P")]
        w   = sum(1 for e in settled if e["result"]=="W")
        l   = sum(1 for e in settled if e["result"]=="L")
        p   = sum(1 for e in settled if e["result"]=="P")
        pnl = sum(e.get("pnl") or 0 for e in settled)
        stk = sum(_entry_stake(e) for e in settled)
        roi = (pnl/stk*100) if stk else 0
        return w, l, p, pnl, stk, roi, log

    def _col(v):  return "#22c55e" if v > 0 else ("#ef4444" if v < 0 else "#94a3b8")
    def _pnl_s(v): return (f"+${v:.2f}" if v >= 0 else f"-${abs(v):.2f}")

    def _pick_logo(league, entry, size=36):
        pick = str(entry.get("pick","")).strip().upper()
        game = str(entry.get("game","")).strip().upper()
        import re as _re
        # Totals — O/U with or without number → use over_under.png like pick cards
        if _re.match(r'^(O|U|OVER|UNDER)(\s|$)', pick):
            _ou_url = _ou_icon_b64_url()
            if _ou_url:
                return (f'<img src="{_ou_url}" alt="O/U" width="{size}" height="{size}" '
                        f'style="object-fit:contain;flex-shrink:0">')
            # Fallback if png not found
            _ou = 'O' if pick[0] == 'O' else 'U'
            _c  = '#f97316' if _ou == 'O' else '#a78bfa'
            return (f'<span style="width:{size}px;height:{size}px;border-radius:50%;'
                    f'background:{_c}22;color:{_c};display:inline-flex;align-items:center;'
                    f'justify-content:center;font-size:{int(size*.44)}px;font-weight:900;'
                    f'flex-shrink:0;border:1.5px solid {_c}44">{_ou}</span>')
        # Strip bet-type suffixes to isolate team: "PISTONS +3.5" → "PISTONS"
        # Also handles "PISTONS ML", "PISTONS -110" etc.
        tokens = pick.split()
        team_tokens = []
        for tok in tokens:
            # Stop at numbers, odds (+/-digits), or ML/RL/SPR keyword
            if _re.match(r'^[+-]?\d', tok) or tok in ('ML','RL','SPR','SPREAD','PK','PICK'):
                break
            team_tokens.append(tok)
        team = " ".join(team_tokens) if team_tokens else (tokens[0] if tokens else "")
        if league == "BSN":
            # Try exact match; normalize accents so ATLÉTICOS → atleticos.png works
            team_ascii = _strip_accents_bsn(team)
            logo = _stats_logo_bsn(team_ascii, size)
            if not logo or 'img' not in logo:
                logo = _stats_logo_bsn(team, size)
            if not logo or 'img' not in logo:
                first_ascii = _strip_accents_bsn(tokens[0]) if tokens else team_ascii
                logo = _stats_logo_bsn(first_ascii, size)
            return logo
        elif league == "NBA":
            # Resolve full team name to abbreviation if needed
            # e.g. "SPURS" → "SAS", "KNICKS" → "NYK"; "BOS" stays as "BOS"
            nba_abb = team
            if team not in _NBA_ESPN:
                nba_abb = _NBA_NICK_REV.get(team, team)
            # Also try matching from the game string (e.g. "BOS @ NYK")
            if nba_abb not in _NBA_ESPN:
                g_parts = _re.split(r'\s+(?:@|VS\.?)\s+', game, flags=_re.IGNORECASE)
                for gp in g_parts:
                    gp = gp.strip().upper()
                    if gp in _NBA_ESPN:
                        # Pick likely refers to one of the game teams
                        # Choose the one that matches the pick token best
                        if gp == team or gp in pick:
                            nba_abb = gp; break
                        nba_abb = gp  # last resort
            return _stats_logo_nba(nba_abb, size)
        else:  # MLB — use game teams as fallback if pick token doesn't match
            parts = _re.split(r'\s+(?:@|VS\.?)\s+', game, flags=_re.IGNORECASE)
            for pt in parts:
                pt2 = pt.strip().upper()
                if pt2 and (pt2 in pick or pt2 == team):
                    team = pt2; break
            return _stats_logo_mlb(team, size)

    bsn_w,bsn_l,bsn_p,bsn_pnl,bsn_stk,bsn_roi,bsn_log = _league_stats(BSN_LOG)
    nba_w,nba_l,nba_p,nba_pnl,nba_stk,nba_roi,nba_log = _league_stats(NBA_LOG)
    mlb_w,mlb_l,mlb_p,mlb_pnl,mlb_stk,mlb_roi,mlb_log = _league_stats(MLB_LOG)

    tot_w   = bsn_w + nba_w + mlb_w
    tot_l   = bsn_l + nba_l + mlb_l
    tot_p   = bsn_p + nba_p + mlb_p
    tot_pnl = bsn_pnl + nba_pnl + mlb_pnl
    tot_stk = bsn_stk + nba_stk + mlb_stk
    tot_roi = (tot_pnl / tot_stk * 100) if tot_stk else 0
    tot_total = tot_w + tot_l + tot_p
    tot_wpct  = f"{tot_w/tot_total*100:.0f}%" if tot_total else "—"

    html = _RC_CSS

    # ── Master header ────────────────────────────────────────────────────
    html += f"""
<div style="background:linear-gradient(135deg,rgba(99,102,241,.12),rgba(168,85,247,.06));
  border:1px solid rgba(99,102,241,.3);border-radius:18px;
  padding:16px 18px;margin-bottom:18px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#6366f1,#a855f7,#ec4899,transparent)"></div>
  <div style="font-size:.56rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
    color:#a5b4fc;margin-bottom:12px">🏆 LABOY PICKS — ALL-TIME</div>
  <div class="rc-stat-grid">
    <div>
      <div class="rc-stat-lbl">Record</div>
      <div class="rc-stat-val" style="color:{_col(tot_w-tot_l)}">{tot_w}-{tot_l}{f'-{tot_p}' if tot_p else ''}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">Win%</div>
      <div class="rc-stat-val" style="color:{_col(tot_w-tot_l)}">{tot_wpct}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">Net Profit</div>
      <div class="rc-stat-val" style="color:{_col(tot_pnl)}">{_pnl_s(tot_pnl)}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">ROI</div>
      <div class="rc-stat-val" style="color:{_col(tot_roi)}">{tot_roi:+.1f}%</div>
    </div>
  </div>
</div>"""

    # ── Per-league sections ──────────────────────────────────────────────
    leagues = [
        ("BSN", "🏀 BSN", "#f5a623", bsn_w, bsn_l, bsn_p, bsn_pnl, bsn_stk, bsn_roi, bsn_log),
        ("NBA", "🏀 NBA", "#3b82f6", nba_w, nba_l, nba_p, nba_pnl, nba_stk, nba_roi, nba_log),
        ("MLB", "⚾ MLB", "#22c55e", mlb_w, mlb_l, mlb_p, mlb_pnl, mlb_stk, mlb_roi, mlb_log),
    ]

    for (key, lbl, accent, w, l, p, pnl, stk, roi, log) in leagues:
        total   = w + l + p
        wpct    = f"{w/total*100:.0f}%" if total else "—"
        pending = [e for e in log if not e.get("result")]
        settled = [e for e in log if e.get("result") in ("W","L","P")]
        # Last 5 settled picks
        recent  = list(reversed(settled))[:5]

        # CLV stats (MLB only — where closing_line tracking is active)
        _clv_entries = [e for e in settled if e.get("clv") is not None]
        _clv_html = ""
        if key == "MLB" and _clv_entries:
            _clv_avg  = sum(e["clv"] for e in _clv_entries) / len(_clv_entries)
            _clv_pos  = sum(1 for e in _clv_entries if e["clv"] > 0)
            _clv_pct  = _clv_pos / len(_clv_entries) * 100
            _clv_col  = "#22c55e" if _clv_avg > 0 else "#ef4444"
            _clv_html = (
                f'<div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.2);'
                f'border-radius:10px;padding:8px 12px;margin-bottom:10px;display:flex;'
                f'justify-content:space-between;align-items:center">'
                f'<div style="font-size:.55rem;color:#a5b4fc;font-weight:700;letter-spacing:.1em">CLV TRACKING</div>'
                f'<div style="display:flex;gap:14px">'
                f'<div style="text-align:center"><div style="font-size:.5rem;color:#64748b">Avg CLV</div>'
                f'<div style="font-size:.82rem;font-weight:800;color:{_clv_col}">{_clv_avg:+.2f}</div></div>'
                f'<div style="text-align:center"><div style="font-size:.5rem;color:#64748b">+CLV%</div>'
                f'<div style="font-size:.82rem;font-weight:800;color:{_clv_col}">{_clv_pct:.0f}%</div></div>'
                f'<div style="text-align:center"><div style="font-size:.5rem;color:#64748b">Samples</div>'
                f'<div style="font-size:.82rem;font-weight:800;color:#94a3b8">{len(_clv_entries)}</div></div>'
                f'</div></div>'
            )

        # Streak
        streak_n, streak_t = 0, ""
        for e in reversed(settled):
            r = e.get("result")
            if streak_n == 0: streak_t = r
            if r == streak_t: streak_n += 1
            else: break
        streak_s = (f'<span style="font-size:.52rem;font-weight:900;padding:2px 7px;border-radius:99px;'
                    f'background:{"rgba(34,197,94,.15)" if streak_t=="W" else "rgba(239,68,68,.15)"};'
                    f'color:{"#22c55e" if streak_t=="W" else "#ef4444"}">'
                    f'{"🔥" if streak_t=="W" else "❄️"} {streak_n} en racha</span>') if streak_n >= 2 else ""

        picks_html = ""
        for e in recent:
            res    = e.get("result","")
            cls    = {"W":"win","L":"loss","P":"push"}.get(res,"pending")
            badge  = {"W":"WIN","L":"LOSS","P":"PUSH"}.get(res,"PEND")
            ep     = e.get("pnl")
            estk   = _entry_stake(e)
            pnl_e  = ep if ep is not None else (-estk if res=="L" else estk)
            pnl_s2 = _pnl_s(pnl_e)
            raw_pick = str(e.get("pick",""))
            odds_v   = e.get("odds",0)
            odds_s   = (f"+{odds_v}" if isinstance(odds_v,(int,float)) and odds_v>0 else str(odds_v)) if odds_v else "—"
            logo_html = _pick_logo(key, e, 32)
            picks_html += f"""<div class="rc-pick {cls}">
  <div class="rc-row">
    {logo_html}
    <div class="rc-main">
      <div class="rc-pick-name">{_esc(raw_pick)}<span class="rc-odds">{_esc(odds_s)}</span></div>
      <div class="rc-game">{_esc(e.get("game",""))}</div>
    </div>
    <div class="rc-result-col">
      <span class="rc-badge {cls}">{badge}</span>
      <div class="rc-pnl {cls}">{pnl_s2}</div>
    </div>
  </div>
</div>"""

        html += f"""
<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.08);
  border-radius:16px;padding:14px 16px;margin-bottom:14px;
  border-top:3px solid {accent};position:relative;overflow:hidden">
  <!-- League header -->
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
    <div style="font-size:.58rem;font-weight:900;letter-spacing:.18em;text-transform:uppercase;color:{accent}">{lbl}</div>
    <div style="display:flex;gap:6px;align-items:center">
      {streak_s}
      {f'<span style="font-size:.5rem;color:#f59e0b;font-weight:700">{len(pending)} pend.</span>' if pending else ''}
    </div>
  </div>
  <!-- Stats row -->
  <div class="rc-stat-grid" style="margin-top:0;margin-bottom:12px">
    <div>
      <div class="rc-stat-lbl">Record</div>
      <div class="rc-stat-val" style="color:{_col(w-l)};font-size:.85rem">{w}-{l}{f'-{p}' if p else ''}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">Win%</div>
      <div class="rc-stat-val" style="color:{_col(w-l)};font-size:.85rem">{wpct}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">Profit</div>
      <div class="rc-stat-val" style="color:{_col(pnl)};font-size:.85rem">{_pnl_s(pnl)}</div>
    </div>
    <div>
      <div class="rc-stat-lbl">ROI</div>
      <div class="rc-stat-val" style="color:{_col(roi)};font-size:.85rem">{roi:+.1f}%</div>
    </div>
  </div>
  <!-- Divider -->
  <div style="height:1px;background:rgba(255,255,255,.06);margin-bottom:10px"></div>
  {_clv_html}
  <!-- Recent picks -->
  {''.join([picks_html]) if picks_html else '<div style="font-size:.6rem;color:#334155;text-align:center;padding:8px 0">Sin picks registrados</div>'}
  {f'<div style="font-size:.52rem;color:#334155;text-align:center;margin-top:4px">Últimos 5 picks · {total} total</div>' if total > 5 else ''}
</div>"""

    return html


def _bsn_daily_record_html():
    """BSN daily record view — pick-by-pick breakdown grouped by date."""
    log = _rj(BSN_LOG)
    if not log:
        return '<div class="detail-empty">Sin picks en el log BSN.</div>'

    # Group by date (most recent first)
    from collections import OrderedDict
    by_date = OrderedDict()
    for e in log:
        d = e.get("date","?")
        if d not in by_date:
            by_date[d] = []
        by_date[d].append(e)

    # Overall totals (settled only)
    all_settled = [e for e in log if e.get("result") in ("W","L","P")]
    tot_w   = sum(1 for e in all_settled if e["result"] == "W")
    tot_l   = sum(1 for e in all_settled if e["result"] == "L")
    tot_p   = sum(1 for e in all_settled if e["result"] == "P")
    tot_pnl = sum(e.get("pnl") or 0 for e in all_settled)
    tot_stk = sum(e.get("stake") or 0 for e in all_settled)
    roi     = (tot_pnl / tot_stk * 100) if tot_stk else 0
    tot_pnl_col = "#22c55e" if tot_pnl >= 0 else "#ef4444"
    roi_col     = "#22c55e" if roi >= 0 else "#ef4444"
    pnl_sign    = "+" if tot_pnl >= 0 else ""

    # ── Summary header ──────────────────────────────────────────────────
    html = f"""<div style="
      background:linear-gradient(135deg,rgba(245,166,35,.08),rgba(245,166,35,.03));
      border:1px solid rgba(245,166,35,.2);border-radius:16px;
      padding:18px 20px;margin-bottom:16px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#f5a623,transparent)"></div>
  <div style="font-size:.58rem;font-weight:900;letter-spacing:.2em;
    color:#f5a623;margin-bottom:12px;text-transform:uppercase">🏀 BSN · Record Acumulado</div>
  <div style="display:flex;gap:20px;flex-wrap:wrap;align-items:center">
    <div>
      <div style="font-size:1.6rem;font-weight:900;color:#f1f5f9;line-height:1">
        {tot_w}<span style="color:#475569;font-size:1rem">-</span>{tot_l}{'<span style="color:#475569;font-size:.8rem"> · '+str(tot_p)+'P</span>' if tot_p else ''}
      </div>
      <div style="font-size:.58rem;color:#475569;margin-top:3px;letter-spacing:.08em">W — L</div>
    </div>
    <div style="width:1px;height:36px;background:rgba(255,255,255,.06)"></div>
    <div>
      <div style="font-size:1.4rem;font-weight:900;color:{tot_pnl_col};line-height:1">
        {pnl_sign}${abs(tot_pnl):.2f}
      </div>
      <div style="font-size:.58rem;color:#475569;margin-top:3px;letter-spacing:.08em">NET P&L</div>
    </div>
    <div style="width:1px;height:36px;background:rgba(255,255,255,.06)"></div>
    <div>
      <div style="font-size:1.4rem;font-weight:900;color:{roi_col};line-height:1">
        {roi:+.1f}%
      </div>
      <div style="font-size:.58rem;color:#475569;margin-top:3px;letter-spacing:.08em">ROI</div>
    </div>
    <div style="width:1px;height:36px;background:rgba(255,255,255,.06)"></div>
    <div>
      <div style="font-size:1.4rem;font-weight:900;color:#94a3b8;line-height:1">
        ${tot_stk:.0f}
      </div>
      <div style="font-size:.58rem;color:#475569;margin-top:3px;letter-spacing:.08em">STAKED</div>
    </div>
  </div>
</div>"""

    # ── Per-day rows (most recent first) ────────────────────────────────
    for d in sorted(by_date.keys(), reverse=True):
        picks = by_date[d]
        settled = [e for e in picks if e.get("result") in ("W","L","P")]
        pending = [e for e in picks if not e.get("result")]
        dw = sum(1 for e in settled if e["result"]=="W")
        dl = sum(1 for e in settled if e["result"]=="L")
        dp = sum(1 for e in settled if e["result"]=="P")
        dpnl = sum(e.get("pnl") or 0 for e in settled)
        dstk = sum(e.get("stake") or 0 for e in picks)
        dpnl_col = "#22c55e" if dpnl > 0 else ("#ef4444" if dpnl < 0 else "#94a3b8")
        dpnl_s   = f'+${dpnl:.2f}' if dpnl >= 0 else f'-${abs(dpnl):.2f}'

        # Day header
        rec_s = f"{dw}W-{dl}L" + (f"-{dp}P" if dp else "")
        if pending:
            rec_s += f" · <span style='color:#f5a623;font-size:.62rem'>{len(pending)} pendiente{'s' if len(pending)>1 else ''}</span>"

        html += f"""<div style="margin-bottom:10px">
  <div style="display:flex;justify-content:space-between;align-items:center;
    padding:8px 14px;background:rgba(255,255,255,.03);
    border:1px solid rgba(255,255,255,.07);border-radius:10px 10px 0 0;
    border-bottom:1px solid rgba(245,166,35,.12)">
    <div style="display:flex;align-items:center;gap:10px">
      <span style="font-size:.65rem;font-weight:900;color:#f5a623;letter-spacing:.06em">{_esc(d)}</span>
      <span style="font-size:.65rem;color:#94a3b8">{rec_s}</span>
    </div>
    <span style="font-size:.75rem;font-weight:800;color:{dpnl_col}">{dpnl_s}</span>
  </div>"""

        # Individual picks
        for e in picks:
            res  = e.get("result")
            if   res == "W": rc, rl, rb = "#22c55e", "W", "rgba(34,197,94,.1)"
            elif res == "L": rc, rl, rb = "#ef4444", "L", "rgba(239,68,68,.08)"
            elif res == "P": rc, rl, rb = "#94a3b8", "P", "rgba(148,163,184,.08)"
            else:            rc, rl, rb = "#f5a623", "·", "rgba(245,166,35,.05)"

            pnl   = e.get("pnl")
            pnl_s = ""
            if pnl is not None and res in ("W","L","P"):
                pc    = "#22c55e" if pnl >= 0 else "#ef4444"
                sign  = "+" if pnl >= 0 else ""
                pnl_s = f'<span style="color:{pc};font-weight:700">{sign}${abs(pnl):.2f}</span>'

            is_parlay = e.get("type") == "parlay"
            pick_text = f"Parlay ({len(e.get('legs',[]))} legs)" if is_parlay else e.get("pick","—")
            game_text = e.get("game","—")
            odds_s    = _fmt_odds(e.get("odds", 0))
            stake_s   = f'${e.get("stake",0):.2f}'
            book_s    = e.get("book","")

            html += f"""  <div style="display:flex;align-items:center;gap:10px;
    padding:9px 14px;background:{rb};
    border:1px solid rgba(255,255,255,.04);border-top:none">
    <div style="width:22px;height:22px;border-radius:6px;background:{rc}1a;
      border:1px solid {rc}44;display:flex;align-items:center;justify-content:center;
      font-size:.6rem;font-weight:900;color:{rc};flex-shrink:0">{rl}</div>
    <div style="flex:1;min-width:0">
      <div style="font-size:.72rem;font-weight:700;color:#f1f5f9;
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{_esc(pick_text)}</div>
      <div style="font-size:.62rem;color:#475569;margin-top:2px;
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{_esc(game_text)}</div>
    </div>
    <div style="text-align:right;flex-shrink:0;font-size:.65rem;color:#64748b;line-height:1.7">
      <div>{_esc(odds_s)} · {_esc(stake_s)}{(' · '+_esc(book_s)) if book_s else ''}</div>
      <div>{pnl_s}</div>
    </div>
  </div>"""

        html += '</div>'  # close day block

    return f'<div class="vw-output">{html}</div>'


def _nba_picks_html():
    """Build NBA picks panel — reads fragment saved by nba.py --picks (same as MLB pattern)."""
    today_str = date.today().strftime("%Y-%m-%d")
    frag_path   = os.path.join(NBA_DIR, "nba_picks_body_current.html")
    picks_path  = os.path.join(NBA_DIR, "nba_model_picks.json")
    nba_py_path = os.path.join(NBA_DIR, "nba.py")
    if os.path.exists(frag_path):
        try:
            from datetime import datetime as _dt
            frag_mtime    = os.path.getmtime(frag_path)
            frag_date     = _dt.fromtimestamp(frag_mtime).strftime("%Y-%m-%d")
            picks_mtime   = os.path.getmtime(picks_path) if os.path.exists(picks_path) else 0
            nba_py_mtime  = os.path.getmtime(nba_py_path) if os.path.exists(nba_py_path) else 0
            serve_py_path = os.path.join(os.path.dirname(NBA_DIR), "serve.py")
            serve_mtime   = os.path.getmtime(serve_py_path) if os.path.exists(serve_py_path) else 0
            # Only use the cached HTML if it's from today, at least as fresh as the picks JSON,
            # AND neither nba.py nor serve.py has been updated since (catches code fixes mid-day)
            if frag_date == today_str and frag_mtime >= picks_mtime and frag_mtime >= nba_py_mtime and frag_mtime >= serve_mtime:
                with open(frag_path, "r", encoding="utf-8") as f:
                    return f.read()
        except Exception:
            pass
    # Fallback: grid view from JSON
    NBA_PICKS_FILE = os.path.join(NBA_DIR, "nba_model_picks.json")
    data = _rj(NBA_PICKS_FILE)
    if not data or not isinstance(data, dict):
        return ('<div class="detail-empty">Sin picks del modelo.<br>'
                '<span style="font-size:.7rem;color:#475569">Corre ▶ Run Modelo primero.</span></div>')

    today_str = date.today().strftime("%Y-%m-%d")
    picks = data.get(today_str, [])
    display_date = today_str

    if not picks:
        # Fall back to most recent date
        dates = sorted(data.keys(), reverse=True)
        if dates:
            display_date = dates[0]
            picks = data[display_date]
        if not picks:
            return '<div class="detail-empty">Sin picks disponibles.</div>'

    # Sort by EV descending
    picks = sorted(picks, key=lambda p: p.get('_ev', 0) or 0, reverse=True)
    n = len(picks)
    best_ev = max((p.get('_ev', 0) or 0 for p in picks), default=0)

    # ── Helpers ────────────────────────────────────────────────────────
    def _fmt_odds_nba(v):
        if v is None: return '—'
        try:
            vi = int(float(v))
            return f'+{vi}' if vi > 0 else str(vi)
        except Exception:
            return str(v)

    def _ev_style(ev_str):
        try:
            v = float(str(ev_str).rstrip('%').lstrip('+'))
            if v >= 40: return 'color:#10b981;font-weight:900;text-shadow:0 0 18px rgba(16,185,129,.5)'
            if v >= 15: return 'color:#10b981;font-weight:800'
            if v >= 0:  return 'color:#86efac;font-weight:700'
            return 'color:#f43f5e;font-weight:700'
        except Exception:
            return 'color:#94a3b8'

    def _edge_style(edge_str):
        s = str(edge_str)
        if s.startswith('+') and s not in ('+0', '+0.0'):
            return 'color:#10b981;font-weight:700'
        if s.startswith('-'):
            return 'color:#f43f5e;font-weight:700'
        return 'color:#94a3b8'

    def _modelo_style(modelo_str, pick_type):
        """Color model odds/projection."""
        s = str(modelo_str)
        if pick_type in ('OVER','UNDER') or 'Proj' in s:
            return 'color:#60a5fa;font-weight:700'  # blue for totals
        try:
            v = float(s.lstrip('+'))
            # Model ML: negative = model favorite (strong signal)
            return 'color:#f43f5e;font-weight:800' if v < -150 else 'color:#94a3b8;font-weight:700'
        except Exception:
            return 'color:#94a3b8'

    # ── Title card ────────────────────────────────────────────────────
    ev_pct = f'{best_ev*100:.0f}%' if best_ev < 5 else f'{best_ev:.1f}%'
    # (if _ev is stored as decimal 0-1 or percentage)
    ev_display = ev_pct if best_ev < 5 else f'{best_ev:.1f}%'

    title_html = f"""<div style="
      background:rgba(79,142,247,.07);border:1px solid rgba(79,142,247,.18);
      border-radius:18px;padding:18px 22px;margin-bottom:16px;
      position:relative;overflow:hidden;">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#4f8ef7,transparent)"></div>
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px">
    <span style="font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;color:#4f8ef7">
      🏀 LABOY PICKS — NBA · {_esc(display_date)}
    </span>
    <span style="font-size:.65rem;font-weight:900;color:#10b981;
      background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.28);
      border-radius:6px;padding:3px 10px;text-shadow:0 0 14px rgba(16,185,129,.5)">
      EV+ {_esc(ev_display)}
    </span>
  </div>
  <div style="font-size:.66rem;color:#475569;font-weight:600">
    Fuente: MODELO + MERCADO &nbsp;·&nbsp; {n} pick{'s' if n != 1 else ''}
  </div>
</div>"""

    # ── Table header ──────────────────────────────────────────────────
    _th = lambda txt, w='': (
        f'<span style="font-size:.48rem;font-weight:900;letter-spacing:.14em;'
        f'text-transform:uppercase;color:#1e293b;{w}">{txt}</span>'
    )
    hdr_html = f"""<div style="
      display:grid;
      grid-template-columns:28px 1fr 110px 52px 70px 70px 65px 65px;
      gap:6px;align-items:center;
      padding:8px 14px;margin-bottom:4px;
      background:rgba(255,255,255,.015);border-radius:8px;">
  {_th('#', 'text-align:center')}
  {_th('JUEGO')}
  {_th('PICK')}
  {_th('ODDS', 'text-align:center')}
  {_th('MODELO', 'text-align:center')}
  {_th('MERCADO', 'text-align:center')}
  {_th('EDGE', 'text-align:center')}
  {_th('EV', 'text-align:center')}
</div>"""

    # ── Data rows ─────────────────────────────────────────────────────
    rows_html = ""
    for i, p in enumerate(picks, 1):
        game      = p.get("game", "—")
        pick_txt  = p.get("pick", "—")
        pick_type = p.get("type", "")
        odds_v    = p.get("odds")
        modelo_v  = p.get("modelo", "—")
        mercado_v = p.get("mercado", "—")
        edge_v    = p.get("edge", "—")
        ev_v      = p.get("ev", "—")
        away_abb  = p.get("away_abb", "")
        home_abb  = p.get("home_abb", "")

        odds_s    = _fmt_odds_nba(odds_v)
        modelo_s  = str(modelo_v) if modelo_v else "—"
        mercado_s = str(mercado_v) if mercado_v else "—"
        edge_s    = str(edge_v) if edge_v else "—"
        ev_s      = str(ev_v) if ev_v else "—"

        # Team badges for game column
        away_b = _badge_nba(away_abb) if away_abb else f'<span style="font-size:.7rem;font-weight:800;color:#94a3b8">{_esc(away_abb)}</span>'
        home_b = _badge_nba(home_abb) if home_abb else f'<span style="font-size:.7rem;font-weight:800;color:#94a3b8">{_esc(home_abb)}</span>'

        # Top-3 EV gets accent border
        if i == 1:   brd = "border-left:3px solid #10b981;background:rgba(16,185,129,.04)"
        elif i == 2: brd = "border-left:3px solid rgba(16,185,129,.45)"
        elif i == 3: brd = "border-left:3px solid rgba(16,185,129,.2)"
        else:        brd = ""

        ev_sty   = _ev_style(ev_s)
        edge_sty = _edge_style(edge_s)
        mdl_sty  = _modelo_style(modelo_s, pick_type)

        # Market conflict detection + series missing detection
        _conflict      = p.get("_market_conflict", False)
        _conflict_disc = p.get("_market_conflict_disc", 0)
        _model_data    = p.get("model", {})
        _series_miss   = _model_data.get("series_missing", False) if _model_data else False
        _series_note   = _model_data.get("series_note", None) if _model_data else None

        if _conflict:
            brd = "border-left:3px solid rgba(139,92,246,.5);background:rgba(139,92,246,.03)"
        if _series_miss and not _conflict:
            brd = "border-left:3px solid rgba(239,68,68,.45);background:rgba(239,68,68,.02)"

        conflict_banner = ""
        if _series_miss:
            # Sin datos de la serie — esta es la raíz del problema modelo vs mercado en playoffs
            conflict_banner = (
                f'<div style="margin-top:6px;padding:8px 14px;'
                f'background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.25);'
                f'border-radius:8px;display:flex;align-items:flex-start;gap:8px">'
                f'<span style="font-size:.9rem;flex-shrink:0">🚨</span>'
                f'<div>'
                f'<div style="font-size:.62rem;font-weight:900;color:#ef4444;letter-spacing:.08em;margin-bottom:2px">'
                f'SIN DATOS DE SERIE — PROYECCIÓN CIEGA</div>'
                f'<div style="font-size:.6rem;color:#7f1d1d;line-height:1.5">'
                f'El modelo no pudo obtener los scores de juegos anteriores de esta serie. '
                f'Está proyectando con solo stats de temporada + playoffs acumulados (incluye R1 vs otro rival). '
                f'El mercado usa estos datos — esta es la mayor fuente de error en playoff. '
                f'<strong>Ingresa los scores manualmente:</strong> '
                f'<code style="background:rgba(0,0,0,.2);padding:1px 4px;border-radius:3px">'
                f'python3 nba.py --add-series-game AWAY HOME PTS_A PTS_H FECHA</code></div>'
                f'</div></div>'
            )
        elif _conflict:
            conflict_banner = (
                f'<div style="margin-top:6px;padding:8px 14px;'
                f'background:rgba(139,92,246,.06);border:1px solid rgba(139,92,246,.2);'
                f'border-radius:8px;display:flex;align-items:flex-start;gap:8px">'
                f'<span style="font-size:.9rem;flex-shrink:0">⚠️</span>'
                f'<div>'
                f'<div style="font-size:.62rem;font-weight:900;color:#a78bfa;letter-spacing:.08em;margin-bottom:2px">'
                f'CONFLICTO MODELO vs MERCADO · {_conflict_disc:.1f} PTS DE DISCREPANCIA</div>'
                f'<div style="font-size:.6rem;color:#6d28d9;line-height:1.5">'
                f'El modelo y el mercado apuntan en <strong>lados opuestos</strong>. '
                f'Verifica score de la serie, lesiones, y noticias del día antes de apostar.</div>'
                f'</div></div>'
            )

        rows_html += f"""<div style="margin-bottom:8px">
<div style="
  display:grid;
  grid-template-columns:28px 1fr 110px 52px 70px 70px 65px 65px;
  gap:6px;align-items:center;
  padding:11px 14px;
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);
  border-radius:12px {'12px 12px 0 0' if conflict_banner else '12px'};
  transition:all .15s;{brd}">
  <span style="font-size:.6rem;font-weight:900;color:{'#10b981' if i==1 and not _conflict else '#a78bfa' if _conflict else '#1e293b'};text-align:center">#{i}</span>
  <div style="display:flex;align-items:center;gap:5px;min-width:0">
    {away_b}
    <span style="font-size:.55rem;color:#1e293b;font-weight:800">@</span>
    {home_b}
  </div>
  <span style="font-size:.78rem;font-weight:800;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{_esc(pick_txt)}</span>
  <span style="font-size:.78rem;font-weight:700;color:#94a3b8;text-align:center">{_esc(odds_s)}</span>
  <span style="font-size:.78rem;text-align:center;{_esc(mdl_sty)}">{_esc(modelo_s)}</span>
  <span style="font-size:.78rem;font-weight:700;color:#94a3b8;text-align:center">{_esc(mercado_s)}</span>
  <span style="font-size:.76rem;text-align:center;{_esc(edge_sty)}">{_esc(edge_s)}</span>
  <span style="font-size:.8rem;text-align:center;{_esc(ev_sty)}">{_esc(ev_s)}</span>
</div>{conflict_banner}</div>"""

    # ── Alt picks (if any) ────────────────────────────────────────────
    alt_html = ""
    for p in picks:
        alts = p.get("alt_picks", [])
        if alts:
            game = p.get("game","")
            alt_html += f'<div style="margin-top:4px;padding:7px 14px;background:rgba(255,255,255,.015);border-radius:9px">'
            alt_html += f'<span style="font-size:.52rem;color:#334155;font-weight:800;letter-spacing:.1em;text-transform:uppercase">ALT · {_esc(game)}</span><br>'
            for a in alts:
                alt_html += f'<span style="font-size:.72rem;color:#475569;font-weight:600">· {_esc(str(a))}</span><br>'
            alt_html += '</div>'

    # ── Game Projections (deduplicated) ───────────────────────────────
    seen_games = {}
    for p in picks:
        game = p.get("game", "")
        if game and game not in seen_games and p.get("model"):
            seen_games[game] = {
                "model": p["model"],
                "away_abb": p.get("away_abb",""),
                "home_abb": p.get("home_abb",""),
            }

    proj_html = ""
    if seen_games:
        proj_html += f"""<div style="margin-top:20px">
  <div style="font-size:.5rem;font-weight:900;letter-spacing:.2em;text-transform:uppercase;
    color:#334155;margin-bottom:10px;padding-left:4px">📐 PROYECCIONES DEL MODELO</div>"""
        for game, gd in seen_games.items():
            mdl = gd["model"]
            away_abb = gd["away_abb"]
            home_abb = gd["home_abb"]
            pts_a   = mdl.get("pts_a", 0)
            pts_h   = mdl.get("pts_h", 0)
            wp_a    = mdl.get("wp_a",  0)
            wp_h    = mdl.get("wp_h",  0)
            ml_a    = mdl.get("ml_a",  0)
            ml_h    = mdl.get("ml_h",  0)
            spread  = mdl.get("spread", 0)
            total   = mdl.get("total",  0)
            away_b  = _badge_nba(away_abb) if away_abb else f'<span style="font-size:.72rem;font-weight:800;color:#94a3b8">{_esc(away_abb)}</span>'
            home_b  = _badge_nba(home_abb) if home_abb else f'<span style="font-size:.72rem;font-weight:800;color:#94a3b8">{_esc(home_abb)}</span>'
            ml_a_s  = f'+{int(ml_a)}' if ml_a > 0 else str(int(ml_a))
            ml_h_s  = f'+{int(ml_h)}' if ml_h > 0 else str(int(ml_h))
            sprd_s  = f'{spread:+.1f}' if spread != 0 else '0'
            fav_a   = wp_a > wp_h
            # color the favored team score
            sc_a    = '#10b981' if fav_a else '#64748b'
            sc_h    = '#10b981' if not fav_a else '#64748b'
            proj_html += f"""<div style="
    background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
    border-radius:14px;padding:13px 16px;margin-bottom:8px;
    display:flex;flex-direction:column;gap:8px">
  <!-- Teams + scores -->
  <div style="display:flex;align-items:center;gap:0;justify-content:space-between">
    <div style="display:flex;align-items:center;gap:7px;flex:1">
      {away_b}
      <span style="font-size:.62rem;color:#1e293b;font-weight:700">AWAY</span>
    </div>
    <div style="display:flex;align-items:center;gap:12px;font-size:.9rem;font-weight:900">
      <span style="color:{sc_a}">{pts_a:.1f}</span>
      <span style="font-size:.58rem;color:#1e293b;font-weight:700">vs</span>
      <span style="color:{sc_h}">{pts_h:.1f}</span>
    </div>
    <div style="display:flex;align-items:center;gap:7px;flex:1;justify-content:flex-end">
      <span style="font-size:.62rem;color:#1e293b;font-weight:700">HOME</span>
      {home_b}
    </div>
  </div>
  <!-- Stats row -->
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;text-align:center">
    <div style="background:rgba(255,255,255,.03);border-radius:8px;padding:6px 4px">
      <div style="font-size:.48rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">ML AWAY</div>
      <div style="font-size:.76rem;font-weight:900;color:{'#10b981' if fav_a else '#64748b'}">{ml_a_s}</div>
    </div>
    <div style="background:rgba(255,255,255,.03);border-radius:8px;padding:6px 4px">
      <div style="font-size:.48rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">ML HOME</div>
      <div style="font-size:.76rem;font-weight:900;color:{'#10b981' if not fav_a else '#64748b'}">{ml_h_s}</div>
    </div>
    <div style="background:rgba(255,255,255,.03);border-radius:8px;padding:6px 4px">
      <div style="font-size:.48rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">SPREAD</div>
      <div style="font-size:.76rem;font-weight:900;color:#60a5fa">{sprd_s}</div>
    </div>
    <div style="background:rgba(255,255,255,.03);border-radius:8px;padding:6px 4px">
      <div style="font-size:.48rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">TOTAL</div>
      <div style="font-size:.76rem;font-weight:900;color:#a78bfa">{total:.1f}</div>
    </div>
  </div>
  <!-- Win probabilities bar -->
  <div style="display:flex;align-items:center;gap:6px">
    <span style="font-size:.6rem;font-weight:800;color:{sc_a};white-space:nowrap">{wp_a:.1f}%</span>
    <div style="flex:1;height:5px;background:rgba(255,255,255,.06);border-radius:3px;overflow:hidden">
      <div style="height:100%;width:{wp_a:.1f}%;background:linear-gradient(90deg,#10b981,#3b82f6);border-radius:3px;transition:width .4s ease"></div>
    </div>
    <span style="font-size:.6rem;font-weight:800;color:{sc_h};white-space:nowrap">{wp_h:.1f}%</span>
  </div>
</div>"""
        proj_html += '</div>'

    # ── Injury Report subsection ─────────────────────────────────────────
    ACTIVE = {"out","doubtful","questionable"}
    all_ir = _rj(NBA_IR)
    ir_entries = [e for e in all_ir if e.get("status","").lower() in ACTIVE]
    ir_html = ""
    if ir_entries:
        # Collect ALL teams playing today from both picks AND no-pick games (lines file)
        today_teams = set()
        for g_data in seen_games.values():
            if g_data.get("away_abb"): today_teams.add(g_data["away_abb"])
            if g_data.get("home_abb"): today_teams.add(g_data["home_abb"])
        # Also pull teams from lines file (catches games with no edge pick)
        NBA_LINES_FILE_IR = os.path.join(NBA_DIR, "nba_model_lines.json")
        _lines_ir = _rj(NBA_LINES_FILE_IR)
        if _lines_ir and isinstance(_lines_ir, dict):
            _ir_date = display_date if 'display_date' in dir() else date.today().strftime("%Y-%m-%d")
            for _gl in _lines_ir.get(_ir_date, []):
                if _gl.get("away_abb"): today_teams.add(_gl["away_abb"])
                if _gl.get("home_abb"): today_teams.add(_gl["home_abb"])
        # Filter to today's teams if we have them, otherwise show all
        ir_show = [e for e in ir_entries if not today_teams or e.get("team_abb","") in today_teams]
        if not ir_show:
            ir_show = ir_entries  # fallback: show all

        ir_rows_html = ""
        for e in sorted(ir_show, key=lambda x: (x.get("team_abb",""), -float(x.get("ppg",0) or 0))):
            status = (e.get("status") or "").lower()
            if status == "out":
                sc, sl = "#ef4444", "OUT"
            elif status == "doubtful":
                sc, sl = "#f97316", "DBT"
            else:
                sc, sl = "#eab308", "QST"
            abb   = e.get("team_abb","")
            badge = _badge_nba(abb) if abb else ""
            ppg   = e.get("ppg","")
            impact= e.get("impact","")
            ppg_s = f"{ppg} PPG" if ppg else ""
            imp_s = f"  {impact} pts impact" if impact else ""
            ir_rows_html += f"""<div style="
  display:flex;align-items:center;gap:9px;
  padding:8px 12px;
  background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.05);
  border-radius:10px;margin-bottom:5px">
  {badge}
  <div style="flex:1;min-width:0">
    <div style="font-size:.72rem;font-weight:800;color:#e2e8f0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{_esc(e.get("player",""))}</div>
    <div style="font-size:.56rem;color:#475569;font-weight:600">{_esc(ppg_s)}{_esc(imp_s)}</div>
  </div>
  <span style="font-size:.62rem;font-weight:900;padding:3px 9px;
    background:rgba(255,255,255,.05);border-radius:6px;color:{sc};
    border:1px solid {sc}40">{sl}</span>
</div>"""

        last_ir_ts = ('—' if _NBA_IR_LAST_REFRESH[0] == 0.0 else
                      __import__('datetime').datetime.fromtimestamp(_NBA_IR_LAST_REFRESH[0]).strftime('%H:%M'))
        ir_html = f"""<div style="margin-top:22px">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
    <span style="font-size:.5rem;font-weight:900;letter-spacing:.2em;text-transform:uppercase;color:#334155">🏥 INJURY REPORT</span>
    <span style="font-size:.52rem;color:#1e293b;font-weight:600">Auto-refresh · {last_ir_ts}</span>
  </div>
  {ir_rows_html}
</div>"""

    # ── Juegos sin pick (modelo los calculó pero no hubo edge suficiente) ──────
    no_pick_html = ""
    NBA_LINES_FILE = os.path.join(NBA_DIR, "nba_model_lines.json")
    lines_data = _rj(NBA_LINES_FILE)
    if lines_data and isinstance(lines_data, dict):
        lines_today = lines_data.get(display_date, [])
        # Games that have picks already — match by away+home abbreviations since game label may differ
        picked_pairs = {(p.get("away_abb",""), p.get("home_abb","")) for p in picks}
        no_pick_games = [
            g for g in lines_today
            if (g.get("away_abb",""), g.get("home_abb","")) not in picked_pairs
            and g.get("model")
        ]
        if no_pick_games:
            no_pick_html = f"""<div style="margin-top:20px;border-top:1px solid rgba(255,255,255,.06);padding-top:16px">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:12px">
  <div style="flex:1;height:1px;background:linear-gradient(90deg,rgba(79,142,247,.25),transparent)"></div>
  <span style="font-size:.44rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
    color:#475569;white-space:nowrap">OTROS JUEGOS · NO PICK</span>
  <div style="flex:1;height:1px;background:linear-gradient(270deg,rgba(79,142,247,.25),transparent)"></div>
</div>"""
            for g in no_pick_games:
                mdl   = g.get("model", {})
                ga    = g.get("away_abb","")
                gh    = g.get("home_abb","")
                wp_a  = mdl.get("wp_a", 50)
                wp_h  = mdl.get("wp_h", 50)
                ml_a  = mdl.get("ml_a", 0)
                ml_h  = mdl.get("ml_h", 0)
                sprd  = mdl.get("spread", 0)
                total = mdl.get("total", 0)
                sn    = mdl.get("series_note") or ""
                smiss = mdl.get("series_missing", False)
                fav_a = wp_a >= wp_h
                a_b   = _badge_nba(ga, size=26) if ga else f'<span style="font-size:.7rem;font-weight:800;color:#94a3b8">{_esc(ga)}</span>'
                h_b   = _badge_nba(gh, size=26) if gh else f'<span style="font-size:.7rem;font-weight:800;color:#94a3b8">{_esc(gh)}</span>'
                ml_a_s  = f'+{int(ml_a)}' if ml_a > 0 else str(int(ml_a))
                ml_h_s  = f'+{int(ml_h)}' if ml_h > 0 else str(int(ml_h))
                sprd_s  = f'{sprd:+.1f}'
                total_s = f'{total:.1f}'
                a_ml_c  = '#10b981' if fav_a else '#64748b'
                h_ml_c  = '#10b981' if not fav_a else '#64748b'
                wp_fav  = max(wp_a, wp_h)
                fav_abb = ga if fav_a else gh
                wp_bar_a = wp_a
                smiss_badge = ('<span style="font-size:.42rem;font-weight:900;color:#ef4444;'
                               'padding:2px 5px;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);'
                               'border-radius:4px;margin-left:5px">🚨 SIN DATOS SERIE</span>') if smiss else ''

                no_pick_html += f"""<div style="margin-bottom:7px;
  background:rgba(255,255,255,.018);border:1px solid rgba(255,255,255,.05);
  border-radius:12px;overflow:hidden;transition:all .15s">
  <!-- header row -->
  <div style="display:flex;align-items:center;justify-content:space-between;padding:9px 14px 0">
    <div style="display:flex;align-items:center;gap:7px">
      {a_b}
      <span style="font-size:.5rem;color:#1e293b;font-weight:900">@</span>
      {h_b}
      {smiss_badge}
    </div>
    <div style="display:flex;align-items:center;gap:5px">
      <span style="font-size:.62rem;font-weight:900;color:#a78bfa">{wp_fav:.0f}%</span>
      <span style="font-size:.42rem;color:#334155;font-weight:700">{_esc(fav_abb)}</span>
      <span style="font-size:.4rem;font-weight:900;letter-spacing:.06em;padding:2px 7px;
        background:rgba(30,41,59,.6);border:1px solid rgba(255,255,255,.06);
        border-radius:5px;color:#334155">NO PICK</span>
    </div>
  </div>
  <!-- wp bar -->
  <div style="display:flex;align-items:center;gap:4px;padding:5px 14px 7px">
    <span style="font-size:.5rem;font-weight:700;color:{a_ml_c};width:22px;text-align:right">{wp_bar_a:.0f}%</span>
    <div style="flex:1;height:4px;background:rgba(255,255,255,.07);border-radius:2px;overflow:hidden">
      <div style="height:100%;width:{wp_bar_a:.1f}%;background:linear-gradient(90deg,#3b82f6,#6366f1);border-radius:2px;transition:width .4s ease"></div>
    </div>
    <span style="font-size:.5rem;font-weight:700;color:{h_ml_c};width:22px">{100-wp_bar_a:.0f}%</span>
  </div>
  <!-- stats row -->
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;border-top:1px solid rgba(255,255,255,.04)">
    <div style="text-align:center;padding:6px 4px;border-right:1px solid rgba(255,255,255,.04)">
      <div style="font-size:.36rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase;margin-bottom:2px">Spread</div>
      <div style="font-size:.72rem;font-weight:900;color:#60a5fa">{_esc(sprd_s)}</div>
    </div>
    <div style="text-align:center;padding:6px 4px;border-right:1px solid rgba(255,255,255,.04)">
      <div style="font-size:.36rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase;margin-bottom:2px">Total</div>
      <div style="font-size:.72rem;font-weight:900;color:#a78bfa">{_esc(total_s)}</div>
    </div>
    <div style="text-align:center;padding:6px 4px;border-right:1px solid rgba(255,255,255,.04)">
      <div style="font-size:.36rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase;margin-bottom:2px">{_esc(ga)} ML</div>
      <div style="font-size:.72rem;font-weight:900;color:{a_ml_c}">{_esc(ml_a_s)}</div>
    </div>
    <div style="text-align:center;padding:6px 4px">
      <div style="font-size:.36rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase;margin-bottom:2px">{_esc(gh)} ML</div>
      <div style="font-size:.72rem;font-weight:900;color:{h_ml_c}">{_esc(ml_h_s)}</div>
    </div>
  </div>
  {f'<div style="padding:4px 14px 7px;font-size:.48rem;color:#334155;border-top:1px solid rgba(255,255,255,.04)">📊 {_esc(sn)}</div>' if sn else ''}
</div>"""
            no_pick_html += "</div>"

    return (f'<div class="vw-output">'
            f'{title_html}{hdr_html}'
            f'<div style="display:flex;flex-direction:column">{rows_html}</div>'
            f'{alt_html}'
            f'{proj_html}'
            f'{no_pick_html}'
            f'{ir_html}'
            f'</div>')


def _nba_lines_html():
    """Build full model lines view — all games, not just edge picks.
    Reads from nba_model_lines.json (written by --lines or --picks).
    Falls back to nba_model_picks.json if lines file doesn't exist yet.
    """
    NBA_LINES_FILE = os.path.join(NBA_DIR, "nba_model_lines.json")
    NBA_PICKS_FILE = os.path.join(NBA_DIR, "nba_model_picks.json")

    # Prefer the dedicated lines JSON (has ALL games)
    data = _rj(NBA_LINES_FILE)
    if not data or not isinstance(data, dict):
        # Fall back to picks JSON — fewer games but at least something
        data = _rj(NBA_PICKS_FILE)
    if not data or not isinstance(data, dict):
        return ('<div class="detail-empty">Sin datos del modelo.<br>'
                '<span style="font-size:.7rem;color:#475569">Corre ▶ Run Modelo primero.</span></div>')

    today_str = date.today().strftime("%Y-%m-%d")
    picks = data.get(today_str, [])
    display_date = today_str
    if not picks:
        dates = sorted(data.keys(), reverse=True)
        if dates:
            display_date = dates[0]
            picks = data[display_date]
    if not picks:
        return '<div class="detail-empty">Sin líneas disponibles.</div>'

    # Normalize entries — nba_model_lines.json has {game,away_abb,home_abb,model} directly,
    # while nba_model_picks.json has {game,away_abb,home_abb,model,pick,_ev,...}.
    # We unify both into the same seen-dict structure used below.
    seen = {}
    for p in picks:
        game = p.get("game","")
        if not game or not p.get("model"):
            continue
        if game not in seen:
            seen[game] = {
                "model":    p["model"],
                "away_abb": p.get("away_abb",""),
                "home_abb": p.get("home_abb",""),
                "picks":    [],
            }
        # Only append if it looks like a pick entry (has _ev or pick key)
        if p.get("pick") or "_ev" in p:
            seen[game]["picks"].append(p)

    # Sort games by absolute away win probability spread (most lopsided first)
    def _lopsided(gd):
        wp_a = gd["model"].get("wp_a", 50)
        return abs(wp_a - 50)
    games_sorted = sorted(seen.items(), key=lambda kv: _lopsided(kv[1]), reverse=True)

    # ── Title ─────────────────────────────────────────────────────────
    n_games = len(games_sorted)
    html = f"""<div class="vw-output">
<div style="background:rgba(79,142,247,.07);border:1px solid rgba(79,142,247,.18);
  border-radius:18px;padding:16px 20px;margin-bottom:16px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#4f8ef7,transparent)"></div>
  <div style="font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;color:#4f8ef7;margin-bottom:4px">
    📐 LABOY NBA — MODEL LINES · {_esc(display_date)}
  </div>
  <div style="font-size:.64rem;color:#475569;font-weight:600">
    {n_games} juego{'s' if n_games!=1 else ''} · Proyecciones del modelo
  </div>
</div>"""

    for game, gd in games_sorted:
        mdl     = gd["model"]
        away_ab = gd["away_abb"]
        home_ab = gd["home_abb"]
        pts_a   = mdl.get("pts_a", 0)
        pts_h   = mdl.get("pts_h", 0)
        wp_a    = mdl.get("wp_a",  50)
        wp_h    = mdl.get("wp_h",  50)
        ml_a    = mdl.get("ml_a",  0)
        ml_h    = mdl.get("ml_h",  0)
        spread  = mdl.get("spread", 0)
        total   = mdl.get("total",  0)
        away_b  = _badge_nba(away_ab) if away_ab else f'<span style="font-size:.75rem;font-weight:800;color:#94a3b8">{_esc(away_ab)}</span>'
        home_b  = _badge_nba(home_ab) if home_ab else f'<span style="font-size:.75rem;font-weight:800;color:#94a3b8">{_esc(home_ab)}</span>'
        ml_a_s  = f'+{int(ml_a)}' if ml_a > 0 else str(int(ml_a))
        ml_h_s  = f'+{int(ml_h)}' if ml_h > 0 else str(int(ml_h))
        sprd_s  = f'{spread:+.1f}'
        fav_a   = wp_a >= wp_h
        sc_a    = '#10b981' if fav_a  else '#64748b'
        sc_h    = '#10b981' if not fav_a else '#64748b'

        # Picks badges for this game
        pick_badges = ""
        for p in gd["picks"]:
            ev_raw = p.get("_ev", 0) or 0
            ev_pct = ev_raw * 100 if ev_raw < 5 else ev_raw
            if ev_pct >= 15:
                bg = "rgba(16,185,129,.15)"; border = "rgba(16,185,129,.4)"; tc = "#10b981"
            else:
                bg = "rgba(255,255,255,.05)"; border = "rgba(255,255,255,.1)"; tc = "#94a3b8"
            pick_lbl = _esc(p.get("pick",""))
            ev_lbl   = f'{ev_pct:.0f}% EV'
            pick_badges += (f'<span style="font-size:.6rem;font-weight:800;padding:3px 9px;'
                           f'background:{bg};border:1px solid {border};border-radius:6px;color:{tc}'
                           f'">{pick_lbl} <span style="font-size:.52rem;opacity:.8">{ev_lbl}</span></span> ')

        html += f"""<div style="
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:16px;padding:15px 18px;margin-bottom:10px">
  <!-- Score row -->
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
    <div style="display:flex;align-items:center;gap:8px;flex:1">
      {away_b}
      <div>
        <div style="font-size:.52rem;color:#334155;font-weight:800;text-transform:uppercase;letter-spacing:.08em">Away</div>
        <div style="font-size:1.05rem;font-weight:900;color:{sc_a};line-height:1">{pts_a:.1f}</div>
        <div style="font-size:.55rem;color:{'#10b981' if fav_a else '#475569'};font-weight:700">{wp_a:.1f}% WP</div>
      </div>
    </div>
    <div style="text-align:center;padding:0 10px">
      <div style="font-size:.5rem;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.1em">Total</div>
      <div style="font-size:.82rem;font-weight:900;color:#a78bfa">{total:.1f}</div>
      <div style="font-size:.5rem;color:#334155;font-weight:700">O/U proj.</div>
    </div>
    <div style="display:flex;align-items:center;gap:8px;flex:1;justify-content:flex-end;text-align:right">
      <div>
        <div style="font-size:.52rem;color:#334155;font-weight:800;text-transform:uppercase;letter-spacing:.08em">Home</div>
        <div style="font-size:1.05rem;font-weight:900;color:{sc_h};line-height:1">{pts_h:.1f}</div>
        <div style="font-size:.55rem;color:{'#10b981' if not fav_a else '#475569'};font-weight:700">{wp_h:.1f}% WP</div>
      </div>
      {home_b}
    </div>
  </div>
  <!-- Win prob bar -->
  <div style="display:flex;align-items:center;gap:6px;margin-bottom:10px">
    <span style="font-size:.58rem;font-weight:800;color:{sc_a};white-space:nowrap;min-width:36px">{wp_a:.1f}%</span>
    <div style="flex:1;height:6px;background:rgba(255,255,255,.05);border-radius:3px;overflow:hidden">
      <div style="height:100%;width:{wp_a:.1f}%;background:linear-gradient(90deg,#10b981,#3b82f6);border-radius:3px;transition:width .4s ease"></div>
    </div>
    <span style="font-size:.58rem;font-weight:800;color:{sc_h};white-space:nowrap;min-width:36px;text-align:right">{wp_h:.1f}%</span>
  </div>
  <!-- Stats row -->
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;text-align:center;margin-bottom:{'10px' if pick_badges else '0'}">
    <div style="background:rgba(255,255,255,.04);border-radius:9px;padding:7px 4px">
      <div style="font-size:.45rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">ML Away</div>
      <div style="font-size:.78rem;font-weight:900;color:{'#10b981' if fav_a else '#64748b'}">{ml_a_s}</div>
    </div>
    <div style="background:rgba(255,255,255,.04);border-radius:9px;padding:7px 4px">
      <div style="font-size:.45rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">Spread (Away)</div>
      <div style="font-size:.78rem;font-weight:900;color:#60a5fa">{sprd_s}</div>
    </div>
    <div style="background:rgba(255,255,255,.04);border-radius:9px;padding:7px 4px">
      <div style="font-size:.45rem;color:#475569;font-weight:800;letter-spacing:.1em;text-transform:uppercase;margin-bottom:3px">ML Home</div>
      <div style="font-size:.78rem;font-weight:900;color:{'#10b981' if not fav_a else '#64748b'}">{ml_h_s}</div>
    </div>
  </div>
  {f'<div style="display:flex;flex-wrap:wrap;gap:5px;margin-top:6px">{pick_badges}</div>' if pick_badges else ''}
</div>"""

    html += '</div>'
    return html


def _mlb_weather_html():
    """Fetch weather for today's MLB games via Open-Meteo. Render visual cards."""
    import urllib.request as _ur

    # Stadium coordinates: {internal_abb: (name, city, lat, lon, has_roof)}
    _PARKS = {
        "ARI":("Chase Field","Phoenix, AZ",33.4455,-112.0667,True),
        "ATL":("Truist Park","Atlanta, GA",33.8908,-84.4678,False),
        "BAL":("Oriole Park","Baltimore, MD",39.2839,-76.6218,False),
        "BOS":("Fenway Park","Boston, MA",42.3467,-71.0972,False),
        "CHC":("Wrigley Field","Chicago, IL",41.9484,-87.6553,False),
        "CWS":("Guaranteed Rate Field","Chicago, IL",41.8300,-87.6339,False),
        "CIN":("Great American Ball Park","Cincinnati, OH",39.0979,-84.5081,False),
        "CLE":("Progressive Field","Cleveland, OH",41.4962,-81.6852,False),
        "COL":("Coors Field","Denver, CO",39.7559,-104.9942,False),
        "DET":("Comerica Park","Detroit, MI",42.3390,-83.0485,False),
        "HOU":("Minute Maid Park","Houston, TX",29.7572,-95.3552,True),
        "KC": ("Kauffman Stadium","Kansas City, MO",39.0517,-94.4803,False),
        "LAA":("Angel Stadium","Anaheim, CA",33.8003,-117.8827,False),
        "LAD":("Dodger Stadium","Los Angeles, CA",34.0739,-118.2400,False),
        "MIA":("loanDepot park","Miami, FL",25.7781,-80.2197,True),
        "MIL":("American Family Field","Milwaukee, WI",43.0280,-87.9712,True),
        "MIN":("Target Field","Minneapolis, MN",44.9817,-93.2783,False),
        "NYM":("Citi Field","New York, NY",40.7571,-73.8458,False),
        "NYY":("Yankee Stadium","New York, NY",40.8296,-73.9262,False),
        "OAK":("Oakland Coliseum","Oakland, CA",37.7516,-122.2005,False),
        "PHI":("Citizens Bank Park","Philadelphia, PA",39.9061,-75.1665,False),
        "PIT":("PNC Park","Pittsburgh, PA",40.4469,-80.0057,False),
        "SD": ("Petco Park","San Diego, CA",32.7073,-117.1566,False),
        "SF": ("Oracle Park","San Francisco, CA",37.7786,-122.3893,False),
        "SEA":("T-Mobile Park","Seattle, WA",47.5914,-122.3325,True),
        "STL":("Busch Stadium","St. Louis, MO",38.6226,-90.1928,False),
        "TB": ("Tropicana Field","St. Petersburg, FL",27.7682,-82.6534,True),
        "TEX":("Globe Life Field","Arlington, TX",32.7473,-97.0829,True),
        "TOR":("Rogers Centre","Toronto, ON",43.6414,-79.3894,True),
        "WSH":("Nationals Park","Washington, DC",38.8730,-77.0074,False),
    }
    # MLB team name → internal abb (for schedule API)
    _MLB_NAMEMAP = {
        "arizona diamondbacks":"ARI","atlanta braves":"ATL","baltimore orioles":"BAL",
        "boston red sox":"BOS","chicago cubs":"CHC","chicago white sox":"CWS",
        "cincinnati reds":"CIN","cleveland guardians":"CLE","colorado rockies":"COL",
        "detroit tigers":"DET","houston astros":"HOU","kansas city royals":"KC",
        "los angeles angels":"LAA","los angeles dodgers":"LAD","miami marlins":"MIA",
        "milwaukee brewers":"MIL","minnesota twins":"MIN","new york mets":"NYM",
        "new york yankees":"NYY","athletics":"OAK","oakland athletics":"OAK",
        "philadelphia phillies":"PHI","pittsburgh pirates":"PIT","san diego padres":"SD",
        "san francisco giants":"SF","seattle mariners":"SEA","st. louis cardinals":"STL",
        "tampa bay rays":"TB","texas rangers":"TEX","toronto blue jays":"TOR",
        "washington nationals":"WSH",
    }
    _WMO = {
        0:("☀️","Despejado"),1:("🌤️","Mayormente despejado"),2:("⛅","Parcialmente nublado"),
        3:("☁️","Nublado"),45:("🌫️","Niebla"),48:("🌫️","Niebla con escarcha"),
        51:("🌦️","Llovizna"),53:("🌦️","Llovizna"),55:("🌧️","Llovizna densa"),
        61:("🌧️","Lluvia ligera"),63:("🌧️","Lluvia moderada"),65:("🌧️","Lluvia fuerte"),
        71:("🌨️","Nieve ligera"),73:("🌨️","Nieve"),75:("❄️","Nieve fuerte"),
        80:("🌦️","Aguaceros"),81:("🌧️","Aguaceros"),82:("⛈️","Aguaceros fuertes"),
        95:("⛈️","Tormenta"),96:("⛈️","Tormenta"),99:("⛈️","Tormenta c/granizo"),
    }
    def _wmo(code):
        c = int(code)
        for k in sorted(_WMO.keys(), reverse=True):
            if c >= k: return _WMO.get(k, ("🌡️","Desconocido"))
        return ("🌡️","Desconocido")

    def _compass(deg):
        dirs = ["N","NNE","NE","ENE","E","ESE","SE","SSE",
                "S","SSO","SO","OSO","O","ONO","NO","NNO"]
        return dirs[int((float(deg)+11.25)/22.5) % 16]

    def _wind_impact(speed_mph, direction_deg):
        """Classify wind impact for baseball betting."""
        if speed_mph < 5:
            return ("🟡", "Viento débil — impacto mínimo")
        d = float(direction_deg) % 360
        # "OUT" = blowing toward outfield (roughly 135°-225° from home plate perspective)
        # simplified: wind from ~180° = blowing OUT to CF, wind from ~0° = blowing IN
        if 135 <= d <= 225:
            intensity = "fuerte" if speed_mph > 15 else "moderado"
            return ("🟢", f"💨 Viento SALIENTE {intensity} ({speed_mph:.0f}mph) → favorece OVERS / HRs")
        elif d < 45 or d > 315:
            intensity = "fuerte" if speed_mph > 15 else "moderado"
            return ("🔴", f"💨 Viento ENTRANTE {intensity} ({speed_mph:.0f}mph) → favorece UNDERS")
        else:
            return ("🟡", f"💨 Viento CRUZADO {_compass(d)} {speed_mph:.0f}mph — impacto moderado")

    # ── Fetch today's MLB schedule ────────────────────────────────────
    today_str = date.today().strftime("%Y-%m-%d")
    games_today = []
    try:
        url = f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={today_str}&hydrate=team"
        req = __import__('urllib.request', fromlist=['Request','urlopen']).Request(
            url, headers={"User-Agent":"Mozilla/5.0"})
        with _ur.urlopen(req, timeout=10) as resp:
            sched = json.loads(resp.read())
        for d_entry in sched.get("dates", []):
            for gm in d_entry.get("games", []):
                ht = gm.get("teams",{}).get("home",{}).get("team",{})
                at = gm.get("teams",{}).get("away",{}).get("team",{})
                h_name = (ht.get("name","") or "").lower()
                a_name = (at.get("name","") or "").lower()
                h_abb  = _MLB_NAMEMAP.get(h_name)
                a_abb  = _MLB_NAMEMAP.get(a_name)
                # Fuzzy fallback
                if not h_abb:
                    for k,v in _MLB_NAMEMAP.items():
                        if k in h_name or h_name.endswith(k.split()[-1]):
                            h_abb = v; break
                if not a_abb:
                    for k,v in _MLB_NAMEMAP.items():
                        if k in a_name or a_name.endswith(a_name.split()[-1]):
                            a_abb = v; break
                if h_abb:
                    gtime = gm.get("gameDate","")
                    games_today.append({"home":h_abb,"away":a_abb or "?","game_time":gtime})
    except Exception as ex:
        return (f'<div class="detail-empty">Error al obtener schedule MLB.<br>'
                f'<span style="font-size:.7rem;color:#475569">{_esc(str(ex))}</span></div>')

    if not games_today:
        return '<div class="detail-empty">Sin juegos MLB programados para hoy.</div>'

    # ── Fetch weather per home park ────────────────────────────────────
    def _fetch_wx(lat, lon):
        try:
            params = (f"latitude={lat}&longitude={lon}"
                      f"&current=temperature_2m,apparent_temperature,relative_humidity_2m,"
                      f"wind_speed_10m,wind_direction_10m,precipitation_probability,"
                      f"weather_code"
                      f"&wind_speed_unit=mph&temperature_unit=fahrenheit&timezone=auto")
            url = f"https://api.open-meteo.com/v1/forecast?{params}"
            with _ur.urlopen(url, timeout=8) as r:
                return json.loads(r.read()).get("current",{})
        except Exception:
            return {}

    # Build cards
    cards_html = ""
    for gm in games_today:
        h = gm["home"]
        a = gm.get("away","?")
        park = _PARKS.get(h)
        if not park:
            continue
        park_name, city, lat, lon, has_roof = park
        wx = _fetch_wx(lat, lon)
        if not wx:
            continue

        temp     = wx.get("temperature_2m", "—")
        feels    = wx.get("apparent_temperature", "—")
        humidity = wx.get("relative_humidity_2m", "—")
        wind_spd = wx.get("wind_speed_10m", 0)
        wind_dir = wx.get("wind_direction_10m", 0)
        precip   = wx.get("precipitation_probability", 0)
        wcode    = wx.get("weather_code", 0)
        wx_icon, wx_desc = _wmo(wcode)
        compass  = _compass(wind_dir)
        w_color, w_impact = _wind_impact(wind_spd, wind_dir)

        # Temperature color
        try:
            tf = float(temp)
            tc = "#ef4444" if tf > 90 else ("#10b981" if tf > 65 else "#60a5fa" if tf > 45 else "#818cf8")
        except: tc = "#94a3b8"

        # Precipitation color
        try:
            pf = float(precip)
            pc = "#ef4444" if pf > 70 else ("#f97316" if pf > 40 else "#10b981")
        except: pc = "#94a3b8"

        # Roof note
        roof_note = '<span style="font-size:.52rem;color:#60a5fa;font-weight:700;padding:2px 7px;background:rgba(96,165,250,.1);border-radius:5px;margin-left:6px">🏟️ CUBIERTO</span>' if has_roof else ""

        # Team logos
        h_bg, h_fg = _MLB_COLORS.get(_MLB_ABB_REV.get(h, h), ("#1e293b","#94a3b8"))
        home_logo = _logo_img(f"https://a.espncdn.com/i/teamlogos/mlb/500/{_MLB_ESPN.get(h, h.lower())}.png", h, h_bg, h_fg, 38)
        away_logo = ""
        if a and a != "?":
            a_bg, a_fg = _MLB_COLORS.get(_MLB_ABB_REV.get(a, a), ("#1e293b","#94a3b8"))
            away_logo = _logo_img(f"https://a.espncdn.com/i/teamlogos/mlb/500/{_MLB_ESPN.get(a, a.lower())}.png", a, a_bg, a_fg, 32)

        # Wind rotation arrow
        wind_rot = int(float(wind_dir))

        # Pre-format numeric display strings (avoids invalid f-string format specs)
        def _fmt(v, dec=0):
            try: return f"{float(v):.{dec}f}"
            except: return str(v)
        temp_s   = _fmt(temp)
        feels_s  = _fmt(feels)
        wind_s   = _fmt(wind_spd)
        precip_s = _fmt(precip)
        humid_s  = _fmt(humidity)

        cards_html += f"""<div style="
  background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:16px;padding:16px 18px;margin-bottom:12px;overflow:hidden;position:relative">

  <!-- bg glow based on condition -->
  <div style="position:absolute;top:-20px;right:-20px;width:100px;height:100px;
    border-radius:50%;background:radial-gradient(circle,{'rgba(239,68,68,.08)' if precip > 50 else 'rgba(16,185,129,.06)'} 0%,transparent 70%);pointer-events:none"></div>

  <!-- Header: matchup + park -->
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
    {away_logo}
    <span style="font-size:.55rem;color:#334155;font-weight:700">@</span>
    {home_logo}
    <div style="flex:1;min-width:0">
      <div style="font-size:.72rem;font-weight:900;color:#e2e8f0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
        {_esc(park_name)}{roof_note}
      </div>
      <div style="font-size:.58rem;color:#475569;font-weight:600">{_esc(city)}</div>
    </div>
    <div style="font-size:2rem;line-height:1">{wx_icon}</div>
  </div>

  <!-- Main temp row -->
  <div style="display:flex;align-items:flex-end;gap:18px;margin-bottom:12px">
    <div>
      <div style="font-size:2.6rem;font-weight:900;color:{tc};line-height:1">{temp_s}°</div>
      <div style="font-size:.56rem;color:#475569;font-weight:600;margin-top:2px">Sensación {feels_s}° · {_esc(wx_desc)}</div>
    </div>
    <!-- Wind compass -->
    <div style="text-align:center">
      <div style="position:relative;width:52px;height:52px;border-radius:50%;
        border:2px solid rgba(255,255,255,.1);background:rgba(255,255,255,.04);
        display:flex;align-items:center;justify-content:center;margin:0 auto">
        <div style="position:absolute;width:3px;height:20px;background:#60a5fa;
          border-radius:2px;bottom:50%;left:50%;transform-origin:bottom center;
          transform:translateX(-50%) rotate({wind_rot}deg)"></div>
        <div style="font-size:.52rem;font-weight:900;color:#60a5fa;z-index:1">{_esc(compass)}</div>
      </div>
      <div style="font-size:.6rem;font-weight:800;color:#60a5fa;margin-top:3px">{wind_s} mph</div>
    </div>
  </div>

  <!-- Stats row -->
  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:10px">
    <div style="background:rgba(255,255,255,.04);border-radius:8px;padding:7px;text-align:center">
      <div style="font-size:.45rem;color:#475569;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:3px">Lluvia</div>
      <div style="font-size:.82rem;font-weight:900;color:{pc}">{precip_s}%</div>
    </div>
    <div style="background:rgba(255,255,255,.04);border-radius:8px;padding:7px;text-align:center">
      <div style="font-size:.45rem;color:#475569;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:3px">Humedad</div>
      <div style="font-size:.82rem;font-weight:900;color:#94a3b8">{humid_s}%</div>
    </div>
    <div style="background:rgba(255,255,255,.04);border-radius:8px;padding:7px;text-align:center">
      <div style="font-size:.45rem;color:#475569;font-weight:800;text-transform:uppercase;letter-spacing:.1em;margin-bottom:3px">Viento</div>
      <div style="font-size:.82rem;font-weight:900;color:#60a5fa">{_esc(compass)}</div>
    </div>
  </div>

  <!-- Betting impact -->
  <div style="padding:8px 12px;background:{'rgba(239,68,68,.07)' if w_color=='🔴' else 'rgba(16,185,129,.07)' if w_color=='🟢' else 'rgba(234,179,8,.07)'};
    border:1px solid {'rgba(239,68,68,.2)' if w_color=='🔴' else 'rgba(16,185,129,.2)' if w_color=='🟢' else 'rgba(234,179,8,.2)'};
    border-radius:9px;font-size:.68rem;font-weight:700;
    color:{'#ef4444' if w_color=='🔴' else '#10b981' if w_color=='🟢' else '#eab308'}">
    {_esc(w_impact)}
  </div>
</div>"""

    if not cards_html:
        return '<div class="detail-empty">No se pudo obtener el clima para los juegos de hoy.</div>'

    n = len(games_today)
    return f"""<div class="vw-output">
<div style="background:rgba(96,165,250,.07);border:1px solid rgba(96,165,250,.18);
  border-radius:18px;padding:16px 20px;margin-bottom:16px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#60a5fa,#38bdf8,transparent)"></div>
  <div style="font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
    color:#60a5fa;margin-bottom:4px">⛅ LABOY MLB — WEATHER REPORT · {_esc(today_str)}</div>
  <div style="font-size:.64rem;color:#475569;font-weight:600">
    {n} estadio{'s' if n!=1 else ''} · Open-Meteo live data
  </div>
</div>
{cards_html}</div>"""


# ── Reverse MLB abbreviation lookup (3-letter → full name key) ────────
_MLB_ABB_REV = {v: k for k, v in _MLB_ABB.items()}

_NBA_NICKNAMES = {
    "ATL":"HAWKS",   "BOS":"CELTICS",  "BKN":"NETS",      "CHA":"HORNETS",
    "CHI":"BULLS",   "CLE":"CAVS",     "DAL":"MAVS",      "DEN":"NUGGETS",
    "DET":"PISTONS", "GSW":"WARRIORS", "HOU":"ROCKETS",   "IND":"PACERS",
    "LAC":"CLIPPERS","LAL":"LAKERS",   "MEM":"GRIZZLIES", "MIA":"HEAT",
    "MIL":"BUCKS",   "MIN":"WOLVES",   "NOP":"PELICANS",  "NYK":"KNICKS",
    "OKC":"THUNDER", "ORL":"MAGIC",    "PHI":"76ERS",     "PHX":"SUNS",
    "POR":"BLAZERS", "SAC":"KINGS",    "SAS":"SPURS",     "TOR":"RAPTORS",
    "UTA":"JAZZ",    "WAS":"WIZARDS",
}
# Reverse: nickname → abbreviation (e.g. "SPURS" → "SAS", "KNICKS" → "NYK")
_NBA_NICK_REV = {v: k for k, v in _NBA_NICKNAMES.items()}


def _nba_stats_html():
    """Build premium NBA stats table from nba_stats_cache.json (ORTG / DRTG / PACE / NET)."""
    cache = _rj(os.path.join(NBA_DIR, "nba_stats_cache.json"))
    if not cache or not isinstance(cache, dict):
        return ('<div class="detail-empty">Sin datos NBA.<br>'
                '<span style="font-size:.7rem;color:#475569">Corre las líneas primero.</span></div>')

    # ── Determine which teams are CURRENTLY active (current round) ──────────
    # Strategy: find every unique series (sorted team pair), track the first
    # date each series appeared. The series with the LATEST first-date are the
    # current round. Round start = max of all series' first-dates. Teams in
    # series that started on/after that date are "active".
    _active_teams = set()
    try:
        _pglog = _rj(os.path.join(NBA_DIR, "nba_playoff_game_log.json"))
        _games = _pglog.get("games", []) if isinstance(_pglog, dict) else []
        if _games:
            # Map each series key → earliest date it appeared
            _series_first = {}
            for _g in _games:
                _a, _h, _d = _g.get("away",""), _g.get("home",""), _g.get("date","")
                if not (_a and _h and _d):
                    continue
                _key = tuple(sorted([_a, _h]))
                if _key not in _series_first or _d < _series_first[_key]:
                    _series_first[_key] = _d
            if _series_first:
                # Current round = all series whose first date is within 5 days of
                # the most recently started series (handles staggered round starts)
                from datetime import datetime as _dt2, timedelta as _td2
                _round_start = max(_series_first.values())
                _window_start = (_dt2.fromisoformat(_round_start) - _td2(days=5)).strftime("%Y-%m-%d")
                for (_ta, _th), _fd in _series_first.items():
                    if _fd >= _window_start:
                        _active_teams.add(_ta)
                        _active_teams.add(_th)
        _active_teams.discard("")
    except Exception:
        pass

    rows = []
    for abb, d in cache.items():
        if not isinstance(d, dict):
            continue
        ortg = d.get("ortg")
        drtg = d.get("drtg")
        pace = d.get("pace")
        net  = d.get("net") or ((ortg - drtg) if ortg and drtg else None)
        po_ortg = d.get("po_ortg")
        po_drtg = d.get("po_drtg")
        po_gp   = d.get("po_gp", 0)
        rows.append({
            "abb": abb, "ortg": ortg, "drtg": drtg, "pace": pace,
            "net": net, "po_ortg": po_ortg, "po_drtg": po_drtg, "po_gp": po_gp,
        })

    # Sort by NET rating descending
    rows.sort(key=lambda r: r["net"] or -99, reverse=True)

    # Ranges for progress bars
    ortg_vals = [r["ortg"] for r in rows if r["ortg"]]
    drtg_vals = [r["drtg"] for r in rows if r["drtg"]]
    pace_vals = [r["pace"] for r in rows if r["pace"]]
    net_vals  = [r["net"]  for r in rows if r["net"] is not None]

    o_min, o_max   = (min(ortg_vals), max(ortg_vals)) if ortg_vals else (105, 125)
    d_min, d_max   = (min(drtg_vals), max(drtg_vals)) if drtg_vals else (105, 120)
    p_min, p_max   = (min(pace_vals), max(pace_vals)) if pace_vals else (95, 105)
    n_min, n_max   = (min(net_vals),  max(net_vals))  if net_vals  else (-10, 12)

    def _opct(v): return max(8, int((v - o_min) / max(o_max - o_min, .01) * 100))
    def _dpct(v): return max(8, int((d_max - v) / max(d_max - d_min, .01) * 100))  # lower DRTG = better
    def _ppct(v): return max(8, int((v - p_min) / max(p_max - p_min, .01) * 100))
    def _npct(v): return max(8, int((v - n_min) / max(n_max - n_min, .01) * 100))

    def _ocls(v):  return "vt-pos" if v and v >= 117 else ("vt-neg" if v and v < 112 else "vt-neu")
    def _dcls(v):  return "vt-pos" if v and v <= 111 else ("vt-neg" if v and v > 116 else "vt-neu")
    def _ncls(v):  return "vt-pos" if v and v >= 5   else ("vt-neg" if v and v < 0   else "vt-neu")

    # ── Subtitle bar with last-updated timestamp + refresh button ───────
    cache_path = os.path.join(NBA_DIR, "nba_stats_cache.json")
    try:
        import datetime as _dt_mod
        _mtime = os.path.getmtime(cache_path)
        _ago   = _dt_mod.datetime.fromtimestamp(_mtime).strftime("%d %b %Y %H:%M")
        _ts    = f'Actualizado: <span style="color:#4f8ef7">{_ago}</span>'
    except Exception:
        _ts = ''
    title_html = (f'<div style="display:flex;align-items:center;justify-content:space-between;'
                  f'gap:10px;flex-wrap:wrap;font-size:.62rem;color:#475569;margin-bottom:12px;'
                  f'padding:8px 12px;background:rgba(79,142,247,.04);'
                  f'border-radius:10px;border:1px solid rgba(79,142,247,.1)">'
                  f'<span>ORTG/DRTG blended reg+playoffs &nbsp;·&nbsp; '
                  f'<span style="color:#4f8ef7">{len(rows)} equipos</span>'
                  f'&nbsp;·&nbsp; R2 activos coloreados'
                  f'{(" &nbsp;·&nbsp; " + _ts) if _ts else ""}'
                  f'</span>'
                  f'<button onclick="runThenView(\'python3 nba.py --refresh\',\'NBA\','
                  f'\'/api/view/nba/stats\',\'NBA · Team Stats\')" '
                  f'style="background:rgba(79,142,247,.12);border:1px solid rgba(79,142,247,.3);'
                  f'border-radius:8px;padding:5px 13px;font-size:.6rem;font-weight:800;'
                  f'color:#93c5fd;cursor:pointer;font-family:inherit;letter-spacing:.05em;'
                  f'white-space:nowrap;flex-shrink:0;transition:all .15s"'
                  f' onmouseover="this.style.background=\'rgba(79,142,247,.22)\'"'
                  f' onmouseout="this.style.background=\'rgba(79,142,247,.12)\'">'
                  f'🔄 Refresh</button>'
                  f'</div>')

    # ── Column header ────────────────────────────────────────────────────
    hdr_html = """<div class="vpt-hdr">
  <span class="vpt-th" style="flex:.4">#</span>
  <span class="vpt-th" style="flex:2.4;text-align:left">TEAM</span>
  <span class="vpt-th">ORTG</span>
  <span class="vpt-th">DRTG</span>
  <span class="vpt-th">NET</span>
  <span class="vpt-th">PACE</span>
  <span class="vpt-th">PO GP</span>
</div>"""

    # ── Data rows ─────────────────────────────────────────────────────────
    rows_html = ""
    for i, r in enumerate(rows, 1):
        abb  = r["abb"]
        badge = _stats_logo_nba(abb, size=36)
        nick  = _NBA_NICKNAMES.get(abb, abb)

        ortg = r["ortg"]; drtg = r["drtg"]; pace = r["pace"]; net = r["net"]
        po_gp = r["po_gp"] or 0

        ortg_s = f"{ortg:.1f}" if ortg else "—"
        drtg_s = f"{drtg:.1f}" if drtg else "—"
        pace_s = f"{pace:.1f}" if pace else "—"
        net_s  = (f"+{net:.1f}" if net and net >= 0 else f"{net:.1f}") if net is not None else "—"
        pogp_s = str(int(po_gp)) if po_gp else "—"

        in_po = (abb in _active_teams) if _active_teams else (po_gp > 0)

        # If not currently active: grey out everything
        if not in_po:
            oc = dc = nc = "vt-neu"
            op = dp = pp = np = 0
        else:
            oc = _ocls(ortg); dc = _dcls(drtg); nc = _ncls(net)
            op = _opct(ortg) if ortg else 0
            dp = _dpct(drtg) if drtg else 0
            pp = _ppct(pace)  if pace  else 0
            np = _npct(net)   if net is not None else 0

        # Row styling
        if not in_po:
            brd = "opacity:.38;filter:grayscale(1)"
        elif i == 1: brd = "border-left:3px solid #4f8ef7;background:rgba(79,142,247,.04)"
        elif i == 2: brd = "border-left:3px solid rgba(79,142,247,.5)"
        elif i == 3: brd = "border-left:3px solid rgba(79,142,247,.22)"
        else:        brd = ""

        rank_style = "color:#4f8ef7;font-weight:900" if (in_po and i == 1) else ""
        name_color = "#f1f5f9" if in_po else "#475569"

        def _td(val, cls, pct):
            return (f'<span class="vpt-td {cls}" style="--pct:{pct}%">'
                    f'{_esc(val)}'
                    f'<span class="vpt-bar" style="--pct:{pct}%"></span>'
                    f'</span>')

        rows_html += (
            f'<div class="vpt-row" style="{brd}">'
            f'<span class="vpt-rank" style="{rank_style}">#{i}</span>'
            f'<span class="vpt-team">{badge}'
            f'<span style="font-size:.7rem;font-weight:700;color:{name_color};margin-left:7px">{_esc(nick)}</span>'
            f'</span>'
            f'{_td(ortg_s, oc, op)}'
            f'{_td(drtg_s, dc, dp)}'
            f'{_td(net_s,  nc, np)}'
            f'{_td(pace_s, "vt-neu", pp)}'
            f'<span class="vpt-td vt-neu">{_esc(pogp_s)}</span>'
            f'</div>'
        )

    legend = """<div style="margin-top:14px;padding:12px 14px;
      background:rgba(255,255,255,.02);border-radius:10px;
      font-size:.62rem;color:#334155;line-height:1.9">
  <span style="color:#22c55e;font-weight:700">ORTG ≥ 117</span> = élite ofensivo &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">ORTG &lt;112</span> = débil<br>
  <span style="color:#22c55e;font-weight:700">DRTG ≤ 111</span> = élite defensivo &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">DRTG &gt;116</span> = vulnerable<br>
  <span style="color:#22c55e;font-weight:700">NET ≥ +5</span> = equipo de élite &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">NET &lt;0</span> = equipo below .500<br>
  <span style="color:#475569">Equipos sin playoffs (PO GP = —) aparecen en gris</span>
</div>"""

    return (f'<div class="vw-output">'
            f'{title_html}{hdr_html}'
            f'<div class="vpt">{rows_html}</div>'
            f'{legend}'
            f'</div>')


def _bsn_picks_html():
    """Build BSN picks panel from bsn_model_picks.json — AI style."""
    BSN_PICKS_FILE = os.path.join(BSN_DIR, "bsn_model_picks.json")
    data = _rj(BSN_PICKS_FILE)
    if not data or not isinstance(data, list):
        return ('<div class="detail-empty">Sin picks del modelo.<br>'
                '<span style="font-size:.7rem;color:#475569">Corre ▶ Run Modelo primero.</span></div>')

    today_str = date.today().strftime("%Y-%m-%d")
    picks = [p for p in data if p.get("date") == today_str]
    display_date = today_str
    if not picks:
        dates = sorted({p.get("date","") for p in data if p.get("date")}, reverse=True)
        if dates:
            display_date = dates[0]
            picks = [p for p in data if p.get("date") == display_date]
    if not picks:
        return '<div class="detail-empty">Sin picks disponibles.</div>'

    n = len(picks)
    clr = "#f5a623"
    html = f"""<div class="vw-output">
<div style="background:rgba(245,166,35,.07);border:1px solid rgba(245,166,35,.2);
  border-radius:18px;padding:16px 20px;margin-bottom:16px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#f5a623,transparent)"></div>
  <div style="font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
    color:#f5a623;margin-bottom:4px">🏀 LABOY BSN — PICKS DEL MODELO · {_esc(display_date)}</div>
  <div style="font-size:.64rem;color:#475569;font-weight:600">
    {n} pick{'s' if n!=1 else ''} con edge detectado
  </div>
</div>"""

    for p in picks:
        game      = p.get("game","")
        pick_lbl  = p.get("pick","")
        odds_v    = p.get("odds","")
        edge      = p.get("edge","")
        modelo    = p.get("modelo","")
        reason    = p.get("reason","")
        wp        = p.get("wp","") or modelo

        # Parse teams from game string "T1 @ T2" or "T1 vs T2"
        parts = game.replace(" vs ", " @ ").split(" @ ")
        t1 = parts[0].strip() if parts else game
        t2 = parts[1].strip() if len(parts) > 1 else ""

        logo_a = _bsn_logo_b64_url(t1)
        logo_h = _bsn_logo_b64_url(t2)
        badge_a = (f'<img src="{logo_a}" style="width:32px;height:32px;object-fit:contain;border-radius:6px">'
                   if logo_a else f'<span style="font-size:.62rem;font-weight:900;color:#f5a623">{_esc(t1[:3])}</span>')
        badge_h = (f'<img src="{logo_h}" style="width:32px;height:32px;object-fit:contain;border-radius:6px">'
                   if logo_h else f'<span style="font-size:.62rem;font-weight:900;color:#94a3b8">{_esc(t2[:3])}</span>')

        odds_s  = _fmt_odds(odds_v) if isinstance(odds_v, (int,float)) else str(odds_v)
        edge_s  = str(edge) if edge else "—"
        reason_s = str(reason) if reason else ""

        html += f"""<div style="background:rgba(245,166,35,.04);border:1px solid rgba(245,166,35,.15);
  border-radius:16px;padding:0;margin-bottom:12px;overflow:hidden;
  border-left:3px solid #f5a623">
  <!-- Header -->
  <div style="display:flex;align-items:center;justify-content:space-between;padding:12px 16px 8px">
    <div style="display:flex;align-items:center;gap:8px">
      {badge_a}
      <span style="font-size:.52rem;color:#334155;font-weight:800">@</span>
      {badge_h}
      <div style="margin-left:4px">
        <div style="font-size:.56rem;font-weight:900;color:#f1f5f9">{_esc(t1)} <span style="color:#334155">vs</span> {_esc(t2)}</div>
      </div>
    </div>
    <span style="font-size:.48rem;font-weight:900;padding:3px 8px;
      background:rgba(245,166,35,.15);border:1px solid rgba(245,166,35,.35);
      border-radius:6px;color:#f5a623">🎯 PICK</span>
  </div>
  <!-- Pick block -->
  <div style="margin:0 12px 10px;background:rgba(0,0,0,.3);border-radius:12px;
    border:1px solid rgba(245,166,35,.12);padding:12px 14px">
    <div style="font-size:.52rem;color:#f5a623;font-weight:900;letter-spacing:.1em;
      text-transform:uppercase;margin-bottom:4px">Pick</div>
    <div style="font-size:1.1rem;font-weight:900;color:#f1f5f9;margin-bottom:8px">{_esc(pick_lbl)}</div>
    <div style="display:flex;gap:8px;flex-wrap:wrap">
      <div style="text-align:center;background:rgba(255,255,255,.05);border-radius:8px;padding:5px 10px">
        <div style="font-size:.38rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase">Odds</div>
        <div style="font-size:.78rem;font-weight:900;color:#10b981">{_esc(odds_s)}</div>
      </div>
      <div style="text-align:center;background:rgba(255,255,255,.05);border-radius:8px;padding:5px 10px">
        <div style="font-size:.38rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase">Edge</div>
        <div style="font-size:.78rem;font-weight:900;color:#60a5fa">{_esc(edge_s)}</div>
      </div>
      {f'<div style="text-align:center;background:rgba(255,255,255,.05);border-radius:8px;padding:5px 10px"><div style="font-size:.38rem;color:#334155;font-weight:900;letter-spacing:.08em;text-transform:uppercase">WP%</div><div style="font-size:.78rem;font-weight:900;color:#a78bfa">{_esc(str(wp))}</div></div>' if wp else ''}
    </div>
  </div>
  {f'<div style="padding:0 12px 10px;font-size:.52rem;color:#475569;line-height:1.5">📊 {_esc(reason_s)}</div>' if reason_s else ''}
</div>"""

    html += "</div>"
    return html


def _bsn_stats_html():
    """Build BSN stats table from Excel model (BSN - Advanced sheet)."""
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(
            os.path.join(BSN_DIR, "Laboy Picks - Data Model Module - Last Version.xlsx"),
            data_only=True, read_only=True)
        ws = wb["BSN - Advanced"]
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as ex:
        return (f'<div class="detail-empty">Sin datos BSN.<br>'
                f'<span style="font-size:.7rem;color:#475569">{ex}</span></div>')

    # Parse: header row has 'TEAM','ORtg','DRtg','Pace' at columns H–K (index 7–10)
    # Data rows start at row 5 (index 4) — team short name in col C (index 2)
    stats = {}
    for row in rows_raw:
        short = row[2] if len(row) > 2 else None
        ortg  = row[8] if len(row) > 8 else None
        drtg  = row[9] if len(row) > 9 else None
        pace  = row[10] if len(row) > 10 else None
        if (isinstance(short, str) and short.upper() in _BSN_COLORS
                and isinstance(ortg, (int, float))):
            stats[short.upper()] = {
                "ortg": float(ortg),
                "drtg": float(drtg) if isinstance(drtg, (int, float)) else None,
                "pace": float(pace) if isinstance(pace, (int, float)) else None,
            }

    if not stats:
        return '<div class="detail-empty">Sin datos en el Excel BSN.</div>'

    rows = []
    for abb, d in stats.items():
        ortg = d["ortg"]; drtg = d["drtg"]; pace = d["pace"]
        net  = round(ortg - drtg, 1) if (ortg and drtg) else None
        rows.append({"abb": abb, "ortg": ortg, "drtg": drtg, "pace": pace, "net": net})

    rows.sort(key=lambda r: r["net"] or -99, reverse=True)

    # Ranges
    ortg_vals = [r["ortg"] for r in rows if r["ortg"]]
    drtg_vals = [r["drtg"] for r in rows if r["drtg"]]
    pace_vals = [r["pace"] for r in rows if r["pace"]]
    net_vals  = [r["net"]  for r in rows if r["net"] is not None]
    o_min, o_max = (min(ortg_vals), max(ortg_vals)) if ortg_vals else (108, 122)
    d_min, d_max = (min(drtg_vals), max(drtg_vals)) if drtg_vals else (105, 122)
    p_min, p_max = (min(pace_vals), max(pace_vals)) if pace_vals else (74, 85)
    n_min, n_max = (min(net_vals),  max(net_vals))  if net_vals  else (-8, 8)

    def _opct(v): return max(8, int((v - o_min) / max(o_max - o_min, .01) * 100))
    def _dpct(v): return max(8, int((d_max - v) / max(d_max - d_min, .01) * 100))
    def _ppct(v): return max(8, int((v - p_min) / max(p_max - p_min, .01) * 100))
    def _npct(v): return max(8, int((v - n_min) / max(n_max - n_min, .01) * 100))
    def _ocls(v): return "vt-pos" if v >= 117 else ("vt-neg" if v < 113 else "vt-neu")
    def _dcls(v): return "vt-pos" if v <= 110 else ("vt-neg" if v > 116 else "vt-neu")
    def _ncls(v): return "vt-pos" if v >= 3   else ("vt-neg" if v < 0   else "vt-neu")

    # Subtitle bar
    title_html = (f'<div style="font-size:.62rem;color:#475569;margin-bottom:12px;'
                  f'padding:6px 10px;background:rgba(245,166,35,.04);'
                  f'border-radius:8px;border:1px solid rgba(245,166,35,.12)">'
                  f'ORTG / DRTG / PACE &nbsp;·&nbsp; Ordenado por NET &nbsp;·&nbsp; '
                  f'<span style="color:#f5a623">{len(rows)} equipos BSN 2026</span>'
                  f'</div>')

    hdr_html = """<div class="vpt-hdr">
  <span class="vpt-th" style="flex:.4">#</span>
  <span class="vpt-th" style="flex:2.4;text-align:left">TEAM</span>
  <span class="vpt-th">ORTG</span>
  <span class="vpt-th">DRTG</span>
  <span class="vpt-th">NET</span>
  <span class="vpt-th">PACE</span>
</div>"""

    rows_html = ""
    for i, r in enumerate(rows, 1):
        abb   = r["abb"]
        badge = _stats_logo_bsn(abb, size=36)
        ortg  = r["ortg"]; drtg = r["drtg"]; pace = r["pace"]; net = r["net"]
        ortg_s = f"{ortg:.1f}" if ortg else "—"
        drtg_s = f"{drtg:.1f}" if drtg else "—"
        pace_s = f"{pace:.1f}" if pace else "—"
        net_s  = (f"+{net:.1f}" if net and net >= 0 else f"{net:.1f}") if net is not None else "—"
        oc = _ocls(ortg) if ortg else "vt-neu"
        dc = _dcls(drtg) if drtg else "vt-neu"
        nc = _ncls(net)  if net is not None else "vt-neu"
        op = _opct(ortg) if ortg else 0
        dp = _dpct(drtg) if drtg else 0
        pp = _ppct(pace)  if pace  else 0
        np = _npct(net)   if net is not None else 0
        if   i == 1: brd = "border-left:3px solid #f5a623;background:rgba(245,166,35,.04)"
        elif i == 2: brd = "border-left:3px solid rgba(245,166,35,.5)"
        elif i == 3: brd = "border-left:3px solid rgba(245,166,35,.22)"
        else:        brd = ""
        rank_style = "color:#f5a623;font-weight:900" if i == 1 else ""
        def _td(val, cls, pct):
            return (f'<span class="vpt-td {cls}" style="--pct:{pct}%">'
                    f'{_esc(val)}'
                    f'<span class="vpt-bar" style="--pct:{pct}%"></span>'
                    f'</span>')
        rows_html += (
            f'<div class="vpt-row" style="{brd}">'
            f'<span class="vpt-rank" style="{rank_style}">#{i}</span>'
            f'<span class="vpt-team">{badge}'
            f'<span style="font-size:.7rem;font-weight:700;color:#f1f5f9;margin-left:7px">{_esc(abb)}</span>'
            f'</span>'
            f'{_td(ortg_s, oc, op)}'
            f'{_td(drtg_s, dc, dp)}'
            f'{_td(net_s,  nc, np)}'
            f'{_td(pace_s, "vt-neu", pp)}'
            f'</div>'
        )

    legend = """<div style="margin-top:14px;padding:12px 14px;
      background:rgba(255,255,255,.02);border-radius:10px;
      font-size:.62rem;color:#334155;line-height:1.9">
  <span style="color:#22c55e;font-weight:700">ORTG ≥ 117</span> = élite ofensivo &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">ORTG &lt;113</span> = débil<br>
  <span style="color:#22c55e;font-weight:700">DRTG ≤ 110</span> = élite defensivo &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">DRTG &gt;116</span> = vulnerable
</div>"""

    return (f'<div class="vw-output">'
            f'{title_html}{hdr_html}'
            f'<div class="vpt">{rows_html}</div>'
            f'{legend}'
            f'</div>')


def _mlb_lines_html():
    """Build premium MLB model lines view from mlb_model_lines.json."""
    MLB_LINES_FILE = os.path.join(MLB_DIR, "mlb_model_lines.json")
    data = _rj(MLB_LINES_FILE)
    if not data or not isinstance(data, dict):
        return ('<div class="detail-empty">Sin datos de líneas.<br>'
                '<span style="font-size:.7rem;color:#475569">Corre 📋 Lines primero.</span></div>')

    today_str = date.today().strftime("%Y-%m-%d")
    games = data.get(today_str, [])
    display_date = today_str
    if not games:
        dates = sorted(data.keys(), reverse=True)
        if dates:
            display_date = dates[0]
            games = data[display_date]
    if not games:
        return '<div class="detail-empty">Sin líneas disponibles.</div>'

    # ── Filtrar juegos ya iniciados ──────────────────────────────────────
    # game_time viene como "1:05 PM ET". Comparar contra hora actual ET.
    def _game_started(gt_str):
        """True si el juego ya debió haber empezado (hora actual >= hora del juego)."""
        if not gt_str: return False
        try:
            from datetime import datetime as _dt, timezone as _tz, timedelta as _td
            # Hora actual en ET (UTC-4 en verano, UTC-5 en invierno)
            _utc_now = _dt.now(_tz.utc)
            _et_offset = _td(hours=-4)   # EDT (mayo = verano)
            _et_now = _utc_now + _et_offset
            # Parsear "1:05 PM ET" → hora del juego hoy en ET
            _clean = gt_str.replace(" ET","").strip()
            _gdt   = _dt.strptime(_clean, "%I:%M %p")
            _gdt   = _gdt.replace(year=_et_now.year, month=_et_now.month, day=_et_now.day)
            # Considerar iniciado si ya pasaron 20 min desde la hora de inicio
            _now_min  = _et_now.hour * 60 + _et_now.minute
            _game_min = _gdt.hour * 60 + _gdt.minute
            return _now_min >= _game_min + 20
        except Exception:
            return False

    games = [g for g in games if not _game_started(g.get("game_time",""))]

    if not games:
        return '<div class="detail-empty">Todos los juegos de hoy ya iniciaron. Vuelve mañana 🌙</div>'

    n = len(games)
    html = f"""<div class="vw-output">

<!-- ── Lines header ────────────────────────────────────────────────── -->

<div style="background:rgba(224,82,82,.07);border:1px solid rgba(224,82,82,.18);
  border-radius:18px;padding:16px 20px;margin-bottom:16px;position:relative;overflow:hidden">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#e05252,transparent)"></div>
  <div style="font-size:.62rem;font-weight:900;letter-spacing:.22em;text-transform:uppercase;
    color:#e05252;margin-bottom:4px">⚾ LABOY MLB — MODEL LINES · {_esc(display_date)}</div>
  <div style="font-size:.64rem;color:#475569;font-weight:600">
    {n} juego{'s' if n!=1 else ''} · Proyecciones del modelo
  </div>
</div>"""

    for g in games:
        away     = g.get("away","")
        home     = g.get("home","")
        away_sp  = g.get("away_sp","TBD")
        home_sp  = g.get("home_sp","TBD")
        sp_hand_a = g.get("away_sp_hand","")
        sp_hand_h = g.get("home_sp_hand","")
        gtime    = g.get("game_time","")
        mdl      = g.get("model",{})
        weather  = g.get("weather",{})
        ump      = g.get("ump_hp","")
        ump_f    = g.get("ump_factor",1.0)
        lu_a     = g.get("lineup_away", False)
        lu_h     = g.get("lineup_home", False)
        lu_both  = g.get("lineup_confirmed", False)

        total_v  = mdl.get("total")
        total_s  = f"{total_v:.1f}" if total_v is not None else "—"
        mtotals  = mdl.get("mTotals","—")
        mspread  = mdl.get("mSpread","—")
        ml_pct   = mdl.get("ml_pct","")
        wp_a     = mdl.get("wp_away") or 50.0
        wp_h     = mdl.get("wp_home") or 50.0
        tA_v     = mdl.get("tA") or (total_v * 0.5 if total_v else 4.5)
        tB_v     = mdl.get("tB") or (total_v * 0.5 if total_v else 4.5)
        mkt_books = g.get("mkt", {}).get("books", {})
        mkt_line_v = None
        # Priority: BetMGM → FanDuel → DraftKings → any book
        for _book in ["BetMGM", "FanDuel", "DraftKings"]:
            _bk_v = mkt_books.get(_book, {})
            _to = _bk_v.get("Total_Over", {})
            if _to and _to.get("line"):
                try: mkt_line_v = float(_to["line"]); break
                except: pass
        if mkt_line_v is None:
            for _bk_v in mkt_books.values():
                _to = _bk_v.get("Total_Over", {}) if isinstance(_bk_v, dict) else {}
                if _to and _to.get("line"):
                    try: mkt_line_v = float(_to["line"]); break
                    except: pass
        # Si no hay línea de mercado real, pasar null al widget (no usar modelo como fallback)
        mc_mkt  = f"{mkt_line_v:.1f}" if mkt_line_v else "null"
        # BetMGM actual over/under odds for the main line (for EV calculation in widget)
        _bmgm_bk   = mkt_books.get("BetMGM", {})
        _bmgm_to   = _bmgm_bk.get("Total_Over", {})
        _bmgm_tu   = _bmgm_bk.get("Total_Under", {})
        mc_ods_o   = int(_bmgm_to.get("odds", -110)) if isinstance(_bmgm_to, dict) and _bmgm_to.get("odds") else -110
        mc_ods_u   = int(_bmgm_tu.get("odds", -110)) if isinstance(_bmgm_tu, dict) and _bmgm_tu.get("odds") else -110

        away_b = _stats_logo_mlb(away, size=34) if away else f'<span style="font-weight:800;color:#94a3b8">{_esc(away)}</span>'
        home_b = _stats_logo_mlb(home, size=34) if home else f'<span style="font-weight:800;color:#94a3b8">{_esc(home)}</span>'

        fav_a  = wp_a >= wp_h
        sc_a   = '#10b981' if fav_a  else '#64748b'
        sc_h   = '#10b981' if not fav_a else '#64748b'

        # ML display
        ml_parts = ml_pct.split(" / ") if " / " in ml_pct else ["", ""]
        ml_a_s = ml_parts[0].strip(); ml_h_s = ml_parts[1].strip() if len(ml_parts)>1 else ""

        # Weather badge
        w_badge = ""
        if weather.get("dome"):
            w_badge = '<span style="font-size:.58rem;background:rgba(148,163,184,.1);border:1px solid rgba(148,163,184,.2);border-radius:5px;padding:2px 6px;color:#64748b">🏟️ Dome</span>'
        elif weather.get("temp") is not None:
            d = weather.get("dir","")
            mph = weather.get("mph",0)
            clr = {"OUT":"#f59e0b","IN":"#3b82f6","CROSS":"#94a3b8","L-R":"#94a3b8","R-L":"#94a3b8"}.get(d,"#94a3b8")
            ico = {"OUT":"↑","IN":"↓","CROSS":"↔","L-R":"↔","R-L":"↔"}.get(d,"·")
            w_badge = (f'<span style="font-size:.58rem;background:rgba(255,255,255,.05);'
                      f'border:1px solid rgba(255,255,255,.1);border-radius:5px;padding:2px 6px;color:{clr}">'
                      f'{ico}{d} {mph}mph · {weather.get("temp","")}°F</span>')

        # Umpire badge
        ump_badge = ""
        if ump:
            uf = float(ump_f or 1.0)
            uc = "#f59e0b" if uf > 1.02 else ("#3b82f6" if uf < 0.98 else "#64748b")
            tend = f"+{(uf-1)*100:.0f}% runs" if uf > 1.0 else (f"{(uf-1)*100:.0f}% runs" if uf < 1.0 else "avg")
            ump_badge = (f'<span style="font-size:.58rem;background:rgba(255,255,255,.04);'
                        f'border:1px solid rgba(255,255,255,.08);border-radius:5px;padding:2px 6px;color:{uc}">'
                        f'🧑‍⚖️ {_esc(ump)} ({tend})</span>')

        sp_h_a = f" ({sp_hand_a}HP)" if sp_hand_a else ""
        sp_h_h = f" ({sp_hand_h}HP)" if sp_hand_h else ""

        # Lineup status badge
        if lu_both:
            lu_badge = '<span style="font-size:.58rem;font-weight:800;padding:2px 7px;background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.35);border-radius:5px;color:#10b981">📋 Lineups ✓</span>'
        elif lu_a or lu_h:
            who_missing = []
            if not lu_a: who_missing.append(away[:3])
            if not lu_h: who_missing.append(home[:3])
            lu_badge = f'<span style="font-size:.58rem;font-weight:700;padding:2px 7px;background:rgba(245,158,11,.08);border:1px solid rgba(245,158,11,.3);border-radius:5px;color:#f59e0b">⏳ Falta: {"/".join(who_missing)}</span>'
        else:
            lu_badge = '<span style="font-size:.58rem;font-weight:700;padding:2px 7px;background:rgba(148,163,184,.07);border:1px solid rgba(148,163,184,.18);border-radius:5px;color:#475569">⏳ Sin lineups</span>'

        html += f"""<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);
  border-radius:16px;padding:14px 16px;margin-bottom:10px">
  <!-- Header row: teams + time -->
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
    <div style="display:flex;align-items:center;gap:6px;flex:1">
      {away_b}
      <div>
        <div style="font-size:.5rem;color:#334155;font-weight:800;text-transform:uppercase;letter-spacing:.08em">Away</div>
        <div style="font-size:.82rem;font-weight:900;color:{sc_a};line-height:1">{wp_a:.1f}%</div>
        <div style="font-size:.56rem;color:#94a3b8;font-weight:600">{_esc(ml_a_s)}</div>
      </div>
    </div>
    <div style="text-align:center;padding:0 8px">
      <div style="font-size:.5rem;color:#1e293b;font-weight:800;text-transform:uppercase;letter-spacing:.1em">Total</div>
      <div style="font-size:.88rem;font-weight:900;color:#a78bfa">{total_s}</div>
      <div style="font-size:.58rem;color:#7c3aed;font-weight:700">{_esc(mtotals)}</div>
      {f'<div style="font-size:.52rem;color:#64748b;margin-top:2px">{_esc(gtime)}</div>' if gtime else ''}
      <button onclick="openMC({tA_v:.3f},{tB_v:.3f},'{_esc(away)}','{_esc(home)}','TOT',{mc_mkt},{mc_ods_o},{mc_ods_u})"
        style="margin-top:6px;background:rgba(167,139,250,.12);border:1px solid rgba(167,139,250,.3);
        border-radius:7px;padding:3px 8px;font-size:.5rem;font-weight:800;color:#a78bfa;
        cursor:pointer;letter-spacing:.06em;text-transform:uppercase">🎲 MC</button>
    </div>
    <div style="display:flex;align-items:center;gap:6px;flex:1;justify-content:flex-end">
      <div style="text-align:right">
        <div style="font-size:.5rem;color:#334155;font-weight:800;text-transform:uppercase;letter-spacing:.08em">Home</div>
        <div style="font-size:.82rem;font-weight:900;color:{sc_h};line-height:1">{wp_h:.1f}%</div>
        <div style="font-size:.56rem;color:#94a3b8;font-weight:600">{_esc(ml_h_s)}</div>
      </div>
      {home_b}
    </div>
  </div>
  <!-- Spread + info row -->
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:6px">
    <span style="font-size:.6rem;font-weight:800;padding:3px 9px;
      background:rgba(224,82,82,.1);border:1px solid rgba(224,82,82,.25);
      border-radius:7px;color:#e05252">📊 {_esc(mspread)}</span>
    {lu_badge}
    {w_badge}
    {ump_badge}
  </div>
  <!-- SP row -->
  <div style="font-size:.6rem;color:#475569;border-top:1px solid rgba(255,255,255,.05);padding-top:6px;margin-top:2px">
    <span style="color:#334155">SP:</span>
    <span style="color:#94a3b8">{_esc(away_sp[:22])}{_esc(sp_h_a)}</span>
    <span style="color:#1e293b;margin:0 4px">vs</span>
    <span style="color:#94a3b8">{_esc(home_sp[:22])}{_esc(sp_h_h)}</span>
  </div>
</div>"""

    html += "</div>"
    return html


def _mlb_stats_html():
    """Build premium HTML stats table from mlb_fg_cache.json directly."""
    cache = _rj(os.path.join(MLB_DIR, "mlb_fg_cache.json"))
    if not cache:
        return ('<div class="detail-empty">Sin datos FanGraphs.<br>'
                '<span style="font-size:.7rem;color:#475569">Presiona el botón para cargar.</span><br><br>'
                '<button onclick="runThenView(\'python3 mlb.py --refresh\',\'MLB\','
                '\'/api/view/mlb/stats\',\'MLB · Team Stats\')" '
                'style="font-size:.65rem;font-weight:700;padding:6px 14px;border-radius:10px;'
                'border:1px solid rgba(224,82,82,.4);background:rgba(224,82,82,.15);'
                'color:#e05252;cursor:pointer">🔄 Refresh FG</button></div>')

    wrc_25  = cache.get("wrc_2025", {})
    wrc_26  = cache.get("wrc_2026", {})
    bp_25   = cache.get("bp_2025", {})
    bp_26   = cache.get("bp_2026", {})
    wrc_rhp = cache.get("wrc_vs_rhp", {})
    wrc_lhp = cache.get("wrc_vs_lhp", {})
    fetched = cache.get("fetched_at", "—")

    # Blend: wRC+ = 75% 2025 + 25% 2026 | xFIP = 90% 2025 + 10% 2026
    teams = sorted(set(wrc_25.keys()) | set(wrc_26.keys()))
    rows = []
    for t in teams:
        w25 = wrc_25.get(t); w26 = wrc_26.get(t)
        b25 = bp_25.get(t);  b26 = bp_26.get(t)
        if w25 is None and w26 is None: continue
        wrc = (0.75 * w25 + 0.25 * w26) if (w25 and w26) else (w25 or w26)
        bp  = (0.90 * b25 + 0.10 * b26) if (b25 and b26) else (b25 or b26)
        rhp = wrc_rhp.get(t); lhp = wrc_lhp.get(t)
        rows.append({"team": t, "wrc": wrc, "bp": bp, "rhp": rhp, "lhp": lhp})

    # Sort by wRC+ blend (best offense first)
    rows.sort(key=lambda r: r["wrc"] or 0, reverse=True)

    # Ranges for progress bars
    wrc_vals = [r["wrc"] for r in rows if r["wrc"]]
    bp_vals  = [r["bp"]  for r in rows if r["bp"]]
    wrc_min, wrc_max = (min(wrc_vals), max(wrc_vals)) if wrc_vals else (80, 130)
    bp_min,  bp_max  = (min(bp_vals),  max(bp_vals))  if bp_vals  else (3.0, 5.0)

    def _wpct(v): return max(8, int((v - wrc_min) / max(wrc_max - wrc_min, .01) * 100))
    def _bpct(v): return max(8, int((bp_max - v)  / max(bp_max  - bp_min,  .01) * 100))  # inverted (low = good)
    def _wcls(v): return "vt-pos" if v >= 108 else ("vt-neg" if v < 95 else "vt-neu")
    def _bcls(v): return "vt-pos" if v <= 3.8  else ("vt-neg" if v > 4.5  else "vt-neu")

    # Timestamp display
    try:
        from datetime import datetime as _dt
        ts = _dt.fromisoformat(fetched).strftime("%m/%d %H:%M")
    except Exception:
        ts = str(fetched)[:16]

    # ── Title card ────────────────────────────────────────────────────
    title_html = f"""<div style="
      background:rgba(224,82,82,.07);border:1px solid rgba(224,82,82,.18);
      border-radius:18px;padding:20px 22px;margin-bottom:18px;
      position:relative;overflow:hidden;">
  <div style="position:absolute;top:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#e05252,transparent)"></div>
  <div style="font-size:.62rem;font-weight:900;letter-spacing:.22em;
    text-transform:uppercase;color:#e05252;margin-bottom:8px">⚾ MLB · TEAM STATS BLEND</div>
  <div style="font-size:.74rem;color:#94a3b8;margin-bottom:5px">
    wRC+: <span style="color:#f1f5f9;font-weight:700">75% 2025 + 25% 2026</span>
    &nbsp;·&nbsp; BP xFIP: <span style="color:#f1f5f9;font-weight:700">90% 2025 + 10% 2026</span>
  </div>
  <div style="display:flex;align-items:center;gap:10px;margin-top:6px">
    <div style="font-size:.62rem;color:#334155;flex:1">
      📅 FanGraphs caché — {_esc(ts)} · {len(rows)} equipos
    </div>
    <button onclick="runThenView('python3 mlb.py --refresh','MLB','/api/view/mlb/stats','MLB · Team Stats')"
      style="font-size:.6rem;font-weight:700;letter-spacing:.06em;padding:4px 10px;
      border-radius:8px;border:1px solid rgba(224,82,82,.4);
      background:rgba(224,82,82,.12);color:#e05252;cursor:pointer;
      transition:background .15s" onmouseover="this.style.background='rgba(224,82,82,.22)'"
      onmouseout="this.style.background='rgba(224,82,82,.12)'">
      🔄 Refresh FG
    </button>
  </div>
</div>"""

    # ── Column header ─────────────────────────────────────────────────
    hdr_html = """<div class="vpt-hdr">
  <span class="vpt-th" style="flex:.4">#</span>
  <span class="vpt-th" style="flex:2.4;text-align:left">TEAM</span>
  <span class="vpt-th">wRC+ BLEND</span>
  <span class="vpt-th">BP xFIP</span>
  <span class="vpt-th">vs RHP</span>
  <span class="vpt-th">vs LHP</span>
</div>"""

    # ── Data rows ─────────────────────────────────────────────────────
    rows_html = ""
    for i, r in enumerate(rows, 1):
        t = r["team"]
        badge = _stats_logo_mlb(t, size=36)
        wrc = r["wrc"]; bp = r["bp"]
        rhp = r["rhp"]; lhp = r["lhp"]

        wrc_s = f"{wrc:.1f}" if wrc else "—"
        bp_s  = f"{bp:.3f}" if bp  else "—"
        rhp_s = f"{rhp:.1f}" if rhp else "—"
        lhp_s = f"{lhp:.1f}" if lhp else "—"

        wc = _wcls(wrc) if wrc else "vt-neu"
        bc = _bcls(bp)  if bp  else "vt-neu"
        rc = _wcls(rhp) if rhp else "vt-neu"
        lc = _wcls(lhp) if lhp else "vt-neu"

        wp = _wpct(wrc) if wrc else 0
        bpp= _bpct(bp)  if bp  else 0
        rp = _wpct(rhp) if rhp else 0
        lp = _wpct(lhp) if lhp else 0

        # Top-3 accent
        if   i == 1: brd = "border-left:3px solid #10b981;background:rgba(16,185,129,.04)"
        elif i == 2: brd = "border-left:3px solid rgba(16,185,129,.5)"
        elif i == 3: brd = "border-left:3px solid rgba(16,185,129,.22)"
        else:        brd = ""

        rank_style = "color:#10b981;font-weight:900" if i == 1 else ""

        def _td(val, cls, pct):
            return (f'<span class="vpt-td {cls}" style="--pct:{pct}%">'
                    f'{_esc(val)}'
                    f'<span class="vpt-bar" style="--pct:{pct}%"></span>'
                    f'</span>')

        rows_html += (
            f'<div class="vpt-row" style="{brd}">'
            f'<span class="vpt-rank" style="{rank_style}">#{i}</span>'
            f'<span class="vpt-team">{badge}</span>'
            f'{_td(wrc_s, wc, wp)}'
            f'{_td(bp_s,  bc, bpp)}'
            f'{_td(rhp_s, rc, rp)}'
            f'{_td(lhp_s, lc, lp)}'
            f'</div>'
        )

    # ── Note on coloring ──────────────────────────────────────────────
    legend = """<div style="margin-top:14px;padding:12px 14px;
      background:rgba(255,255,255,.02);border-radius:10px;
      font-size:.62rem;color:#334155;line-height:1.9">
  <span style="color:#10b981;font-weight:700">wRC+ ≥ 108</span> = élite ofensivo &nbsp;·&nbsp;
  <span style="color:#94a3b8">95-107</span> = promedio &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">&lt;95</span> = débil<br>
  <span style="color:#10b981;font-weight:700">xFIP ≤ 3.8</span> = pitcheo élite &nbsp;·&nbsp;
  <span style="color:#f43f5e;font-weight:700">&gt;4.5</span> = vulnerable
</div>"""

    return (f'<div class="vw-output">'
            f'{title_html}{hdr_html}'
            f'<div class="vpt">{rows_html}</div>'
            f'{legend}'
            f'</div>')


def _render_cmd_output(text):
    """Convert raw script stdout into AI-style structured HTML.
    Handles: section headers, pick lines, pipe-tables with team badges,
    space-separated data tables, stats summary lines.
    """
    if not text or not text.strip():
        return '<div class="detail-empty">Sin datos disponibles.</div>'

    def _val_cls(s):
        try:
            v = float(s.rstrip('%').lstrip('+'))
            if '%' in s:
                return 'vt-pos' if v >= 55 else ('vt-neg' if v < 45 else 'vt-neu')
            if 70 <= v <= 200:
                return 'vt-pos' if v > 108 else ('vt-neg' if v < 92 else 'vt-neu')
            if 1.5 <= v <= 7.0:
                return 'vt-pos' if v < 3.5 else ('vt-neg' if v > 4.5 else 'vt-neu')
            if v < 0: return 'vt-neg'
        except Exception: pass
        return 'vt-neu'

    def _auto_badge(name):
        """Return team badge if name matches any known team, else plain text span."""
        u = name.upper()
        if u in _MLB_COLORS: return _badge_mlb(name)
        if u in _BSN_COLORS: return _badge_bsn(name)
        if u in _NBA_COLORS: return _badge_nba(name)
        return '<span class="vpt-team-txt">{}</span>'.format(_esc(name))

    lines = text.split('\n')
    parts = []
    tbuf  = []   # space-table: [(team, [(val, cls)])]
    ptbuf = []   # pipe-table rows: [list_of_col_strings]
    phdr  = []   # pipe-table headers

    def _flush_space_table():
        if not tbuf: return
        rows = ""
        for team, cells in tbuf:
            rows += (
                '<div class="vt-row">'
                '<span class="vt-team">{}</span>'
                '<div class="vt-vals">{}</div>'
                '</div>'.format(
                    _esc(team),
                    "".join('<span class="vt-cell {}">{}</span>'.format(c, _esc(v)) for v, c in cells)
                )
            )
        parts.append('<div class="vw-table">{}</div>'.format(rows))
        tbuf.clear()

    def _flush_pipe_table():
        if not ptbuf: return
        # ── Pre-compute per-column min/max for progress bars ──────────────
        if ptbuf:
            n_cols = max(len(r) for r in ptbuf)
            col_vals = [[] for _ in range(n_cols)]
            for row in ptbuf:
                for ci, cv in enumerate(row):
                    try:
                        col_vals[ci].append(float(cv.rstrip('%').lstrip('+')))
                    except Exception:
                        col_vals[ci].append(None)
            col_min = []
            col_max = []
            for cv_list in col_vals:
                nums = [v for v in cv_list if v is not None]
                col_min.append(min(nums) if nums else 0)
                col_max.append(max(nums) if nums else 1)

        hdr_html = ""
        if phdr:
            hdr_html = '<div class="vpt-hdr">' + ''.join(
                '<span class="vpt-th">{}</span>'.format(_esc(h)) for h in phdr
            ) + '</div>'
        rows_html = ""
        for rank, row_cols in enumerate(ptbuf, 1):
            if not row_cols: continue
            team_name = row_cols[0]
            badge     = _auto_badge(team_name)
            vals      = row_cols[1:]
            val_cells = ""
            for ci, v in enumerate(vals):
                vcls = _val_cls(v)
                # Compute percentage bar width relative to column range
                real_ci = ci + 1  # offset by team col
                try:
                    fv = float(v.rstrip('%').lstrip('+'))
                    cmin = col_min[real_ci] if real_ci < len(col_min) else 0
                    cmax = col_max[real_ci] if real_ci < len(col_max) else 1
                    span = cmax - cmin
                    pct = int((fv - cmin) / span * 100) if span > 0 else 50
                    pct = max(8, min(100, pct))
                except Exception:
                    pct = 0
                bar = '<span class="vpt-bar" style="--pct:{}%"></span>'.format(pct) if pct else ''
                val_cells += '<span class="vpt-td {}" style="--pct:{}%">{}{}</span>'.format(
                    vcls, pct, _esc(v), bar)
            rows_html += (
                '<div class="vpt-row">'
                '<span class="vpt-rank">#{}</span>'
                '<span class="vpt-team">{}</span>'
                '{}'
                '</div>'.format(rank, badge, val_cells)
            )
        parts.append('<div class="vpt">{}{}</div>'.format(hdr_html, rows_html))
        ptbuf.clear()
        phdr.clear()

    for raw in lines:
        s = raw.strip()
        if not s:
            _flush_space_table()
            _flush_pipe_table()
            continue

        # ── Pure separator lines ─────────────────────────────────────────
        core = s.replace(' ','').replace('|','').replace('+','')
        if len(core) > 3 and core and set(core) <= set('-=─━═'):
            _flush_space_table()
            _flush_pipe_table()
            continue

        # ── Pipe-delimited table rows: | A | B | C | ────────────────────
        # Catch BOTH "| A | B |" (starts with |) and "A | B | C" (embedded pipes)
        _has_pipes = s.count('|') >= 2
        _pipe_start = s.startswith('|') and s.count('|') >= 3
        # Also detect "TEAM | val | val" style (no leading pipe)
        _embedded_pipe = (not s.startswith('|') and _has_pipes
                          and re.match(r'^[A-Z][\w\-\s]{1,18}\s*\|', s))
        if _pipe_start or _embedded_pipe:
            if _embedded_pipe:
                s = '|' + s  # normalize to pipe-start format
            pipe_cols = [c.strip() for c in s.split('|') if c.strip()]
            if not pipe_cols: continue
            # Separator row: | --- | --- | or purely dashes
            if all(set(c.replace('-','').replace('=','').replace(' ','')) <= {''} for c in pipe_cols):
                continue
            # Header row: all text (no standalone numbers), no existing data rows yet
            is_all_text = all(not re.match(r'^[+-]?[\d.]+%?$', c) for c in pipe_cols)
            if is_all_text and not ptbuf:
                _flush_space_table()
                phdr.clear()
                phdr.extend(pipe_cols)
            else:
                _flush_space_table()
                ptbuf.append(pipe_cols)
            continue

        # ── Section headers ──────────────────────────────────────────────
        first_ord      = ord(s[0]) if s else 0
        has_emoji_pfx  = first_ord > 0x2500
        # Explicitly exclude lines with pipes (pipe detection should handle those)
        _no_pipes      = '|' not in s
        is_allcaps_hdr = (_no_pipes and s.isupper() and 2 < len(s) < 60
                          and not re.search(r'\d{3,}', s))
        is_colon_label = (_no_pipes and s.endswith(':') and len(s.split()) <= 6
                          and not re.search(r'\d', s))
        is_hdr = has_emoji_pfx or is_allcaps_hdr or is_colon_label

        if is_hdr:
            _flush_space_table()
            _flush_pipe_table()
            parts.append('<div class="vw-hdr">{}</div>'.format(_esc(s)))
            continue

        # ── Pick lines ───────────────────────────────────────────────────
        has_odds   = bool(re.search(r'[+-]\d{3}', s))
        has_ev     = bool(re.search(r'EV\+?\s*[\d.]+', s, re.IGNORECASE))
        is_pick_kw = bool(re.search(r'\bML\b|\bOVER\b|\bUNDER\b|\bO/U\b|RUN LINE|\bRPL\b|\bF5\b', s, re.IGNORECASE))
        is_star    = any(c in s for c in ('⭐','🔥','★','✅'))

        if is_star or has_ev or (has_odds and is_pick_kw):
            _flush_space_table()
            _flush_pipe_table()
            ev_m   = re.search(r'EV\+?\s*([\d.]+)', s, re.IGNORECASE)
            odds_m = re.search(r'([+-]\d{3})', s)
            ev_b   = '<span class="vw-ev">EV+ {}%</span>'.format(ev_m.group(1)) if ev_m else ''
            od_b   = '<span class="vw-odds">{}</span>'.format(odds_m.group(1)) if odds_m else ''
            body   = re.sub(r'EV\+?\s*[\d.]+%?', '', s, flags=re.IGNORECASE)
            body   = re.sub(r'[+-]\d{3}', '', body).strip(' |·-').strip()
            for ch in ('⭐','🔥','★','✅','🏆'): body = body.replace(ch,'')
            body   = body.strip()
            sc = ' vw-pick-star' if is_star else ''
            si = '<span class="vw-star">★</span>' if is_star else ''
            parts.append(
                '<div class="vw-pick{}">{}<span class="vw-pick-text">{}</span>'
                '<div class="vw-pick-badges">{}{}</div></div>'.format(sc, si, _esc(body), od_b, ev_b)
            )
            continue

        # ── Space-separated data table: TEAM num num … ──────────────────
        cols = s.split()
        if (len(cols) >= 2 and cols[0].isupper() and cols[0].isalpha()
                and 2 <= len(cols[0]) <= 12 and not is_hdr):
            nums = [c for c in cols[1:] if re.match(r'^[+-]?[\d.]+%?$', c)]
            if nums:
                cells = [(v, _val_cls(v)) if re.match(r'^[+-]?[\d.]+%?$', v) else (v,'vt-label')
                         for v in cols[1:]]
                tbuf.append((cols[0], cells))
                continue

        _flush_space_table()
        _flush_pipe_table()

        if re.search(r'\d+-\d+|Win%|ROI|P&L|\bROI\b|\bpct\b|record', s, re.IGNORECASE):
            parts.append('<div class="vw-stat-line">{}</div>'.format(_esc(s)))
        else:
            parts.append('<div class="vw-text">{}</div>'.format(_esc(s)))

    _flush_space_table()
    _flush_pipe_table()

    if parts:
        return '<div class="vw-output">{}</div>'.format(''.join(parts))
    return '<div class="detail-empty">Sin datos.</div>'


def handle_api(path, data):
    """Returns (status_code, json_dict)"""
    j = lambda ok, msg: {"ok": ok, "msg": msg}

    # ── BSN ──────────────────────────────────────────────────────
    if path == "/api/bsn/log":
        ok, msg = _log_pick(BSN_LOG, data)
        if ok:
            try:
                log_now = _rj(BSN_LOG)
                pick_idx = len(log_now) - 1
                def _bsn_auto_pub(idx):
                    try:
                        _run(["python3","bsn.py","--export-log", str(idx), "--publish"],
                             cwd=BSN_DIR, timeout=90)
                    except Exception as _e:
                        print(f"  ⚠️  bsn auto-publish pick #{idx}: {_e}")
                threading.Thread(target=_bsn_auto_pub, args=(pick_idx,), daemon=True).start()
            except Exception: pass
        return 200, j(ok, msg)

    if path == "/api/bsn/log-parlay":
        ok, msg = _log_parlay(BSN_LOG, data)
        return 200, j(ok, msg)

    if path == "/api/bsn/grade":
        ok, msg = _grade_pick(BSN_LOG, data)
        return 200, j(ok, msg)

    if path == "/api/bsn/gp":
        team = data.get("team","").strip().upper()
        gp_s = data.get("gp","").strip()
        if not team or not gp_s: return 200, j(False, "⚠️ Equipo y GP requeridos.")
        try:
            gp_data = {}
            if os.path.exists(BSN_GP):
                gp_data = _rj(BSN_GP)
                if isinstance(gp_data, list): gp_data = {}
            gp_data[team] = int(gp_s)
            _wj(BSN_GP, gp_data)
            return 200, j(True, f"✅ {team} → {gp_s} GP guardado")
        except Exception as ex:
            return 200, j(False, f"⚠️ Error: {ex}")

    if path == "/api/bsn/add-game":
        away  = data.get("away","").strip().upper()
        home  = data.get("home","").strip().upper()
        time  = data.get("time","8:00 PM").strip()
        if not away or not home: return 200, j(False, "⚠️ Selecciona ambos equipos.")
        out = _run(["python3","bsn.py","--add-game", away, home, time], cwd=BSN_DIR)
        ok = "error" not in out.lower() and "❌" not in out
        return 200, j(ok, f"{'✅' if ok else '❌'} {out[:200]}")

    if path == "/api/bsn/ir":
        action = data.get("action","add")
        team   = data.get("team","").strip().upper()
        player = data.get("player","").strip().upper()
        rate_s = data.get("rate","1")
        ppg_s  = data.get("ppg","").strip()
        usg_s  = data.get("usg","").strip()
        if not team or not player: return 200, j(False, "⚠️ Equipo y jugador requeridos.")
        if action == "remove":
            out = _run(["python3","bsn.py","--remove-injury", team, player], cwd=BSN_DIR)
        else:
            cmd = ["python3","bsn.py","--add-injury", team, player, rate_s]
            if ppg_s and usg_s:
                cmd += [ppg_s, usg_s]  # skip RealGM scrape
            out = _run(cmd, cwd=BSN_DIR, timeout=60)
        ok = "error" not in out.lower() and "❌" not in out
        return 200, j(ok, f"{'✅' if ok else '❌'} {out[:300]}")

    if path == "/api/bsn/lines":
        home  = data.get("home","").strip().upper()
        away  = data.get("away","").strip().upper()
        total = data.get("total","")
        hml   = data.get("home_ml","")
        aml   = data.get("away_ml","")
        hs    = data.get("home_spread","")
        if not home or not away: return 200, j(False, "⚠️ Equipos requeridos.")
        try:
            lines_file = os.path.join(BSN_DIR, "bsn_market_lines.json")
            lines = _rj(lines_file) if os.path.exists(lines_file) else {}
            if not isinstance(lines, dict): lines = {}
            key = f"{away} vs. {home}"
            today_str = date.today().strftime("%Y-%m-%d")
            if today_str not in lines: lines[today_str] = {}
            entry = {}
            if total: entry["total"] = float(total)
            if hml:   entry["home_ml"] = _parse_odds(hml)
            if aml:   entry["away_ml"] = _parse_odds(aml)
            if hs:    entry["home_spread"] = float(hs)
            lines[today_str][key] = entry
            _wj(lines_file, lines)
            return 200, j(True, f"✅ Líneas guardadas: {key} | Total:{total} ML:{hml}/{aml}")
        except Exception as ex:
            return 200, j(False, f"⚠️ Error: {ex}")

    # ── NBA ──────────────────────────────────────────────────────
    if path == "/api/nba/log":
        ok, msg = _log_pick(NBA_LOG, data)
        return 200, j(ok, msg)

    if path == "/api/nba/grade":
        ok, msg = _grade_pick(NBA_LOG, data)
        return 200, j(ok, msg)

    if path == "/api/nba/auto-grade":
        ok, msg = _auto_grade_nba()
        return 200, j(ok, msg)

    if path == "/api/nba/ir":
        action = data.get("action","add")
        team   = data.get("team","").strip().upper()
        player = data.get("player","").strip().upper()
        rate_s = data.get("rate","3")
        status = data.get("status","out")
        if not team or not player: return 200, j(False, "⚠️ Team y player requeridos.")
        if action == "remove":
            out = _run(["python3","nba.py","--remove-injury", team, player], cwd=NBA_DIR)
        else:
            out = _run(["python3","nba.py","--add-injury", team, player, rate_s, status], cwd=NBA_DIR, timeout=60)
        ok = "error" not in out.lower() and "❌" not in out
        return 200, j(ok, f"{'✅' if ok else '❌'} {out[:300]}")

    # ── BSN game management ────────────────────────────────────
    if path == "/api/bsn/edit-game-time":
        away  = data.get("away","").strip().upper()
        home  = data.get("home","").strip().upper()
        time  = data.get("time","").strip()
        if not away or not home: return 200, j(False, "⚠️ Equipos requeridos.")
        games_path = os.path.join(BSN_DIR, "manual_games.json")
        games = _rj(games_path)
        if not isinstance(games, list): games = []
        today_str = date.today().strftime("%Y-%m-%d")
        found = False
        for g in games:
            if (g.get("date") == today_str
                    and g.get("team1","").upper() == away
                    and g.get("team2","").upper() == home):
                g["game_time"] = time
                found = True
                break
        if not found:
            return 200, j(False, f"⚠️ Juego {away} @ {home} no encontrado para hoy.")
        _wj(games_path, games)
        return 200, j(True, f"✅ {away} @ {home} → {time}")

    if path == "/api/bsn/remove-game":
        away = data.get("away","").strip().upper()
        home = data.get("home","").strip().upper()
        if not away or not home: return 200, j(False, "⚠️ Equipos requeridos.")
        games_path = os.path.join(BSN_DIR, "manual_games.json")
        games = _rj(games_path)
        if not isinstance(games, list): games = []
        today_str = date.today().strftime("%Y-%m-%d")
        new_games = [g for g in games if not (
            g.get("date") == today_str
            and g.get("team1","").upper() == away
            and g.get("team2","").upper() == home
        )]
        if len(new_games) == len(games):
            return 200, j(False, f"⚠️ {away} @ {home} no encontrado para hoy.")
        _wj(games_path, new_games)
        return 200, j(True, f"✅ {away} @ {home} removido")

    # ── MLB ──────────────────────────────────────────────────────
    if path == "/api/mlb/log":
        ok, msg = _log_pick(MLB_LOG, data)
        if ok:
            # Auto-publish en background después de loguear
            try:
                log_now = _rj(MLB_LOG)
                pick_idx = len(log_now) - 1
                def _auto_pub(idx):
                    try:
                        _run(["python3","mlb.py","--export-log", str(idx), "--publish"],
                             cwd=MLB_DIR, timeout=90)
                    except Exception as _e:
                        print(f"  ⚠️  auto-publish pick #{idx}: {_e}")
                threading.Thread(target=_auto_pub, args=(pick_idx,), daemon=True).start()
            except Exception: pass
        return 200, j(ok, msg)

    if path == "/api/mlb/grade":
        ok, msg = _grade_pick(MLB_LOG, data)
        return 200, j(ok, msg)

    if path == "/api/mlb/auto-grade":
        ok, msg = _auto_grade_mlb()
        return 200, j(ok, msg)

    if path == "/api/mlb/scores-debug":
        from datetime import date as _date
        date_str = data.get("date", _date.today().strftime("%Y-%m-%d"))
        raw = _fetch_mlb_scores(date_str)
        return 200, {"ok": True, "date": date_str, "keys": list(raw.keys()), "scores": {k: list(v) for k,v in raw.items() if k != "_error"}, "error": raw.get("_error","")}

    if path == "/api/mlb/debug-game":
        away = data.get("away","").strip().upper()
        home = data.get("home","").strip().upper()
        if not away or not home: return 200, j(False, "⚠️ Selecciona ambos equipos.")
        out = _run(["python3","mlb.py","--debug-game", away, home, "--confirmed"], cwd=MLB_DIR, timeout=120)
        return 200, {"ok": True, "out": out, "html": _render_cmd_output(out)}

    if path == "/api/mlb/publish-all":
        task_id = _new_task_id()
        with _BG_LOCK:
            _BG_TASKS[task_id] = {"status": "running", "out": ""}
        commands = [
            (["python3","mlb.py","--export-picks","--publish"],  "PICKS PUBLICADOS",       MLB_DIR, 180),
            (["python3","mlb.py","--export-lines","--publish"],  "LINES / DEBUG PUBLICADO", MLB_DIR, 180),
        ]
        t = threading.Thread(target=_bg_run_multi, args=(task_id, commands), daemon=True)
        t.start()
        return 200, {"ok": True, "task_id": task_id, "msg": "Publicando en background..."}

    if path == "/api/mlb/publish-log":
        # Publica todos los picks pendientes del log (sin resultado aún no, sino los últimos N HTMLs)
        # Corre --export-log para cada pick sin HTML publicado, luego push.
        task_id = _new_task_id()
        with _BG_LOCK:
            _BG_TASKS[task_id] = {"status": "running", "out": ""}
        # Exporta los últimos 2 picks del log (cubre el caso de 2 picks nuevos)
        log_now = _rj(MLB_LOG)
        if not isinstance(log_now, list): log_now = []
        # Find picks that have a result OR no result (log all recent ones)
        n = len(log_now)
        cmds = []
        # Export last 2 logged picks individually
        for idx in range(max(0, n-2), n):
            cmds.append((["python3","mlb.py","--export-log", str(idx), "--publish"],
                         f"Pick #{idx} publicado", MLB_DIR, 120))
        if not cmds:
            return 200, j(False, "No hay picks en el log.")
        t = threading.Thread(target=_bg_run_multi, args=(task_id, cmds), daemon=True)
        t.start()
        return 200, {"ok": True, "task_id": task_id, "msg": "Publicando log en background..."}

    return 404, j(False, "Not found")


# ══════════════════════════════════════════════════════════════════════
# HTTP SERVER
# ══════════════════════════════════════════════════════════════════════
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def _send_json(self, code, obj):
        body = json.dumps(obj).encode()
        self.send_response(code)
        self.send_header("Content-Type","application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, code, html):
        body = html.encode()
        self.send_response(code)
        self.send_header("Content-Type","text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path == "/favicon.ico":
            self.send_response(204); self.end_headers(); return

        # ── /api/test-autopush — diagnóstico del autopush ─────────────────
        if self.path == "/api/test-autopush":
            import base64 as _b64, urllib.request as _ur, urllib.error as _ue
            token = os.environ.get("GITHUB_TOKEN","")
            user  = os.environ.get("GITHUB_USER","laboywebsite-lgtm")
            repo  = os.environ.get("GITHUB_REPO","laboy-picks")
            result = {"token_set": bool(token), "user": user, "repo": repo, "steps": []}
            if token:
                hdrs = {"Authorization": f"token {token}",
                        "Accept": "application/vnd.github.v3+json"}
                try:
                    req = _ur.Request(
                        f"https://api.github.com/repos/{user}/{repo}/git/ref/heads/main",
                        headers=hdrs)
                    with _ur.urlopen(req, timeout=10) as r:
                        data = json.loads(r.read())
                        result["steps"].append(f"✅ GitHub API OK — SHA: {data['object']['sha'][:7]}")
                except _ue.HTTPError as e:
                    result["steps"].append(f"❌ GitHub API error {e.code}: {e.read().decode()[:200]}")
                except Exception as e:
                    result["steps"].append(f"❌ Exception: {str(e)}")
            else:
                result["steps"].append("❌ GITHUB_TOKEN no está definido en el entorno")
            self._send_json(200, result)
            return
        # ── Rich view endpoints (return {"html": "..."}) ─────────────────
        if self.path.startswith("/api/view/"):
            _clean_path = self.path.split("?")[0]  # strip query string for routing
            vmap = {
                "/api/view/mlb/picks":   lambda: _mlb_picks_debug_view(),
                "/api/view/mlb/lines":   lambda: _mlb_lines_html(),
                "/api/view/mlb/stats":   lambda: _mlb_stats_html(),
                "/api/view/mlb/weather": lambda: _mlb_weather_html(),
                "/api/view/mlb/log":     lambda: _render_log_html(MLB_LOG),
                "/api/view/nba/picks": lambda: _nba_picks_html(),
                "/api/view/nba/lines": lambda: _nba_lines_html(),
                "/api/view/nba/stats": lambda: _nba_stats_html(),
                "/api/view/nba/log":   lambda: _render_log_html(NBA_LOG),
                "/api/view/bsn/picks":  lambda: _bsn_picks_html(),
                "/api/view/bsn/stats":   lambda: _bsn_stats_html(),
                "/api/view/bsn/record": lambda: _bsn_daily_record_html(),
                "/api/view/bsn/log":    lambda: _render_log_html(BSN_LOG),
                "/api/view/record/all": lambda: _alltime_record_html(),
                "/api/view/calendar":   lambda: _pnl_calendar_html(),
            }
            fn = vmap.get(_clean_path)
            if fn:
                try:    self._send_json(200, {"html": fn()})
                except Exception as ex:
                    self._send_json(200, {"html": f'<div style="color:#ef4444;padding:20px">Error: {ex}</div>'})
            else:
                self._send_json(404, {"html": '<div class="detail-empty">Vista no encontrada.</div>'})
            return
        # ── MLB scores debug (GET) ───────────────────────────────────────────
        if self.path.startswith("/api/mlb/scores-debug"):
            from urllib.parse import urlparse, parse_qs as _pqs
            _qs = _pqs(urlparse(self.path).query)
            _d = _qs.get("date", [date.today().strftime("%Y-%m-%d")])[0]
            raw = _fetch_mlb_scores(_d)
            self._send_json(200, {"ok": True, "date": _d, "keys": list(raw.keys()),
                                  "scores": {k: list(v) for k,v in raw.items() if k != "_error"},
                                  "error": raw.get("_error","")})
            return
        # ── MLB today picks (for Log Pick modal live refresh) ───────────────
        if self.path == "/api/mlb/picks-today":
            today_str = date.today().strftime("%Y-%m-%d")
            by_game = {}
            # Prefer mlb_debug_state.json (latest model run) over mlb_model_picks.json
            debug_state_path = os.path.join(MLB_DIR, "mlb_debug_state.json")
            if os.path.exists(debug_state_path):
                try:
                    ds = _rj(debug_state_path)
                    if isinstance(ds, dict) and ds.get("date","") == today_str:
                        for p in ds.get("picks", []):
                            g = p.get("game","")
                            if not g:
                                continue
                            if g not in by_game:
                                by_game[g] = []
                            by_game[g].append({
                                "pick":   p.get("pick",""),
                                "odds":   str(p.get("odds","")),
                                "edge":   p.get("edge",""),
                                "ev":     p.get("ev",""),
                                "modelo": p.get("modelo",""),
                                "time":   p.get("time",""),
                            })
                except Exception:
                    pass
            # Fall back to cumulative picks log if no debug state for today
            if not by_game:
                all_picks = _rj(MLB_PICKS) if os.path.exists(MLB_PICKS) else []
                if not isinstance(all_picks, list):
                    all_picks = []
                today_picks = [p for p in all_picks if p.get("date","") == today_str]
                for p in today_picks:
                    g = p.get("game","")
                    if g not in by_game:
                        by_game[g] = []
                    by_game[g].append({
                        "pick":   p.get("pick",""),
                        "odds":   str(p.get("odds","")),
                        "edge":   p.get("edge",""),
                        "ev":     p.get("ev",""),
                        "modelo": p.get("modelo",""),
                        "time":   p.get("time",""),
                    })
            self._send_json(200, {"picks": by_game, "date": today_str})
            return
        # ── NBA today picks (for Log Pick modal live refresh) ───────────────
        if self.path == "/api/nba/picks-today":
            today_str = date.today().strftime("%Y-%m-%d")
            by_game = {}
            # Load model picks (games with actual picks)
            nba_picks_path = os.path.join(NBA_DIR, "nba_model_picks.json")
            if os.path.exists(nba_picks_path):
                try:
                    nba_data = _rj(nba_picks_path)
                    today_picks = nba_data.get(today_str, []) if isinstance(nba_data, dict) else [p for p in nba_data if p.get("date","") == today_str]
                    for p in today_picks:
                        g = p.get("game","")
                        if not g:
                            continue
                        if g not in by_game:
                            by_game[g] = []
                        by_game[g].append({
                            "pick":      p.get("pick",""),
                            "odds":      str(p.get("odds","")),
                            "edge":      p.get("edge",""),
                            "ev":        p.get("ev",""),
                            "modelo":    p.get("modelo",""),
                            "away_abb":  p.get("away_abb",""),
                            "home_abb":  p.get("home_abb",""),
                        })
                except Exception:
                    pass
            # Also add all today's games from nba_model_lines.json (so SIN PICK games appear too)
            nba_lines_path = os.path.join(NBA_DIR, "nba_model_lines.json")
            if os.path.exists(nba_lines_path):
                try:
                    lines_data = _rj(nba_lines_path)
                    for g_entry in lines_data.get(today_str, []):
                        g = g_entry.get("game","")
                        if g and g not in by_game:
                            by_game[g] = []  # no picks yet, empty list — user fills manually
                except Exception:
                    pass
            self._send_json(200, {"picks": by_game, "date": today_str})
            return

        # ── BSN IR list (GET) ────────────────────────────────────────────
        if self.path == "/api/bsn/ir-list":
            try:
                import openpyxl as _xl
                _wb = _xl.load_workbook(
                    os.path.join(BSN_DIR, "Laboy Picks - Data Model Module - Last Version.xlsx"),
                    data_only=True, read_only=True)
                _ws = _wb["IR - BSN"]
                _entries = []
                for _row in _ws.iter_rows(min_row=2, max_row=_ws.max_row, values_only=True):
                    _rt = _row[1]; _pl = _row[2]; _ra = _row[3]; _pg = _row[4]; _ug = _row[5]; _im = _row[6]
                    if not _rt or not _pl: continue
                    _team = str(_rt).strip().upper()
                    if _team not in BSN_TEAMS: continue   # ← solo equipos BSN
                    _entries.append({
                        "team":   _team,
                        "player": str(_pl).strip().upper(),
                        "rate":   int(_ra) if _ra else 0,
                        "ppg":    float(_pg) if _pg else 0.0,
                        "usg":    float(_ug) if _ug else 0.0,
                        "impact": float(_im) if _im else 0.0,
                    })
                _wb.close()
                self._send_json(200, {"entries": _entries})
            except Exception as _ex:
                self._send_json(200, {"entries": [], "error": str(_ex)})
            return

        # ── BSN games-today (for Set Lines modal) ───────────────────────
        if self.path == "/api/bsn/games-today":
            today_str = date.today().strftime("%Y-%m-%d")
            games_file = os.path.join(BSN_DIR, "manual_games.json")
            lines_file = os.path.join(BSN_DIR, "bsn_market_lines.json")
            all_games = _rj(games_file) if os.path.exists(games_file) else []
            all_lines = _rj(lines_file) if os.path.exists(lines_file) else {}
            today_games = [g for g in all_games if g.get("date") == today_str]
            today_lines = all_lines.get(today_str, []) if isinstance(all_lines, dict) else []
            lines_lookup = {(e.get("team1","").upper(), e.get("team2","").upper()): e
                            for e in today_lines}
            result = []
            for g in today_games:
                t1 = g.get("team1","").upper()
                t2 = g.get("team2","").upper()
                result.append({
                    "team1": t1, "team2": t2,
                    "game_time": g.get("game_time",""),
                    "existing": lines_lookup.get((t1, t2), {})
                })
            self._send_json(200, {"games": result, "date": today_str})
            return

        # ── Task status polling ──────────────────────────────────────────
        if self.path.startswith("/api/task-status"):
            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)
            task_id = params.get("id", [""])[0]
            with _BG_LOCK:
                task = _BG_TASKS.get(task_id)
            if task is None:
                self._send_json(404, {"status": "not_found", "out": "Task no encontrada."})
            else:
                self._send_json(200, task)
            return
        try:
            self._send_html(200, full_page())
        except Exception as ex:
            import traceback
            err = traceback.format_exc()
            self._send_html(500, f"<pre style='color:red;padding:20px'>{err}</pre>")

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length",0))
            raw    = self.rfile.read(length)
            ctype  = self.headers.get("Content-Type","")
            path   = self.path

            # ── /api/run ─────────────────────────────────────────────
            if path == "/api/run":
                try:
                    body = json.loads(raw.decode())
                    cmd_str  = body.get("cmd","")
                    cwd_name = body.get("cwd","")
                    cwd = {"BSN": BSN_DIR, "NBA": NBA_DIR, "MLB": MLB_DIR}.get(cwd_name, BASE_DIR)
                    # Auto-delete stale HTML fragments before any --picks run
                    # so the fresh model output always replaces the old cache.
                    if "--picks" in cmd_str:
                        for _frag_name, _frag_dir in [
                            ("mlb_debug_body_current.html", MLB_DIR),
                            ("nba_picks_body_current.html", NBA_DIR),
                        ]:
                            _fp = os.path.join(_frag_dir, _frag_name)
                            try:
                                if os.path.exists(_fp):
                                    os.remove(_fp)
                            except Exception:
                                pass
                    # Use shlex.split to handle quoted arguments with spaces
                    try:
                        cmd = shlex.split(cmd_str)
                    except ValueError:
                        cmd = cmd_str.split()
                    out = _run(cmd, cwd=cwd)
                    # Autopush si el comando modificó estado (solo en cloud)
                    if any(sc in cmd_str for sc in _GIT_STATE_CMDS):
                        _git_autopush_bg(cmd_str)
                    self._send_json(200, {"out": out})
                except Exception as ex:
                    self._send_json(200, {"out": f"Error: {ex}"})
                return

            # ── /api/run-view (run command → rich visual HTML) ────────────
            if path == "/api/run-view":
                try:
                    body     = json.loads(raw.decode())
                    cmd_str  = body.get("cmd","")
                    cwd_name = body.get("cwd","")
                    cwd = {"BSN": BSN_DIR, "NBA": NBA_DIR, "MLB": MLB_DIR}.get(cwd_name, BASE_DIR)
                    # Auto-delete stale fragments before --picks runs
                    if "--picks" in cmd_str:
                        for _frag_name, _frag_dir in [
                            ("mlb_debug_body_current.html", MLB_DIR),
                            ("nba_picks_body_current.html", NBA_DIR),
                        ]:
                            _fp = os.path.join(_frag_dir, _frag_name)
                            try:
                                if os.path.exists(_fp):
                                    os.remove(_fp)
                            except Exception:
                                pass
                    try:
                        cmd = shlex.split(cmd_str)
                    except ValueError:
                        cmd = cmd_str.split()
                    out = _run(cmd, cwd=cwd)
                    self._send_json(200, {"html": _render_cmd_output(out)})
                except Exception as ex:
                    self._send_json(200, {"html": '<div style="color:#ef4444;padding:20px">Error: {}</div>'.format(_esc(str(ex)))})
                return

            # ── BSN Lines JSON endpoint (new multi-game format) ──────────
            if path == "/api/bsn/lines" and "application/json" in ctype:
                try:
                    body    = json.loads(raw.decode())
                    entries = body.get("entries", [])
                    if not entries:
                        self._send_json(200, {"ok": False, "msg": "Sin entradas"}); return
                    today_str  = date.today().strftime("%Y-%m-%d")
                    lines_file = os.path.join(BSN_DIR, "bsn_market_lines.json")
                    all_lines  = _rj(lines_file) if os.path.exists(lines_file) else {}
                    if not isinstance(all_lines, dict): all_lines = {}
                    # Merge: preserve games already saved that aren't in this submission
                    existing_today = {(e.get("team1",""), e.get("team2","")): e
                                      for e in all_lines.get(today_str, [])}
                    for entry in entries:
                        key = (entry.get("team1",""), entry.get("team2",""))
                        existing_today[key] = entry
                    all_lines[today_str] = list(existing_today.values())
                    _wj(lines_file, all_lines)
                    _git_autopush_bg("--set-lines BSN")
                    self._send_json(200, {"ok": True,
                        "msg": f"✅ Líneas guardadas: {len(entries)} juego(s)"})
                except Exception as ex:
                    self._send_json(200, {"ok": False, "msg": f"Error: {ex}"})
                return

            # ── Form endpoints ────────────────────────────────────────
            if "application/x-www-form-urlencoded" in ctype:
                raw_str = raw.decode("utf-8")
                form_data = {k: v[0] for k,v in parse_qs(raw_str, keep_blank_values=True).items()}
                code, resp = handle_api(path, form_data)
                self._send_json(code, resp)
                return

            self._send_json(404, {"ok": False, "msg": f"Not found: {path}"})
        except Exception as ex:
            import traceback
            try:
                self._send_json(500, {"ok": False, "msg": f"Server error: {ex}"})
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════
# GIT AUTOPUSH — persiste estado JSON en GitHub después de cada acción
# ══════════════════════════════════════════════════════════════════════
#
# Solo activo cuando existe GITHUB_TOKEN en el entorno (Render/cloud).
# En local no hace nada — el usuario hace push manualmente.
#
# Archivos de estado que se persisten:
_GIT_STATE_FILES = [
    "BSN/bsn_picks_log.json",
    "BSN/bsn_gp.json",
    "BSN/bsn_market_lines.json",
    "BSN/bsn_model_picks.json",
    "BSN/manual_games.json",
    "MLB/laboy_picks_log.json",
    "MLB/mlb_model_picks.json",
    "MLB/mlb_log_state.json",
    "MLB/mlb_debug_state.json",
    "NBA/nba_picks_log.json",
    "NBA/nba_injuries.json",
    "NBA/nba_playoff_game_log.json",
]

# Comandos que modifican estado y deben triggear autopush:
_GIT_STATE_CMDS = [
    "--log", "--grade", "--remove", "--edit", "--gp",
    "--refresh", "--set-lines", "--clear-lines",
    "--log-parlay", "--log-special", "--log-retro",
    "--grade-picks", "--add-injury", "--remove-injury",
    "--add-game", "--remove-game",
]

def _setup_git_autopush():
    """No-op — autopush ahora usa GitHub API, no git CLI."""
    token = os.environ.get("GITHUB_TOKEN", "")
    if token:
        print("  ✅ GitHub API autopush activo (modo cloud).")

def _git_autopush_bg(trigger_cmd=""):
    """
    Persiste archivos de estado en GitHub via REST API.
    Un solo commit con todos los archivos cambiados.
    No depende de git CLI — funciona en cualquier entorno cloud.
    """
    def _push():
        import base64 as _b64, urllib.request as _ur, urllib.error as _ue
        token = os.environ.get("GITHUB_TOKEN", "")
        user  = os.environ.get("GITHUB_USER", "laboywebsite-lgtm")
        repo  = os.environ.get("GITHUB_REPO",  "laboy-picks")
        if not token:
            return

        hdrs = {
            "Authorization": f"token {token}",
            "Accept":        "application/vnd.github.v3+json",
            "Content-Type":  "application/json",
        }
        base = f"https://api.github.com/repos/{user}/{repo}"

        def _api(method, path, payload=None, timeout=15):
            data = json.dumps(payload).encode() if payload else None
            req  = _ur.Request(f"{base}{path}", data=data,
                               headers=hdrs, method=method)
            try:
                with _ur.urlopen(req, timeout=timeout) as r:
                    return json.loads(r.read()), r.status
            except _ue.HTTPError as e:
                return json.loads(e.read() or b"{}"), e.code
            except Exception as e:
                return {}, 0

        try:
            # 1. SHA del último commit en main
            ref, st = _api("GET", "/git/ref/heads/main")
            if st != 200:
                print(f"  ⚠️ Autopush: no se pudo leer ref ({st})")
                return
            base_commit_sha = ref["object"]["sha"]

            # 2. SHA del árbol base
            commit, st = _api("GET", f"/git/commits/{base_commit_sha}")
            if st != 200: return
            base_tree_sha = commit["tree"]["sha"]

            # 3. Crear blobs para cada archivo de estado
            tree_items = []
            for rel_path in _GIT_STATE_FILES:
                abs_path = os.path.join(BASE_DIR, rel_path)
                if not os.path.exists(abs_path):
                    continue
                with open(abs_path, "rb") as f:
                    content_b64 = _b64.b64encode(f.read()).decode()
                blob, st = _api("POST", "/git/blobs",
                                {"content": content_b64, "encoding": "base64"})
                if st not in (200, 201):
                    continue
                tree_items.append({
                    "path": rel_path, "mode": "100644",
                    "type": "blob", "sha": blob["sha"]
                })

            if not tree_items:
                return

            # 4. Nuevo árbol
            tree, st = _api("POST", "/git/trees",
                            {"base_tree": base_tree_sha, "tree": tree_items})
            if st not in (200, 201):
                print(f"  ⚠️ Autopush: error creando tree ({st})")
                return

            # 5. Nuevo commit
            short = (trigger_cmd or "state update")[:60]
            new_commit, st = _api("POST", "/git/commits", {
                "message": f"auto: {short}",
                "tree":    tree["sha"],
                "parents": [base_commit_sha]
            })
            if st not in (200, 201):
                print(f"  ⚠️ Autopush: error creando commit ({st})")
                return

            # 6. Actualizar ref de main
            _, st = _api("PATCH", "/git/refs/heads/main",
                         {"sha": new_commit["sha"]})
            if st in (200, 201):
                print(f"  ☁️  GitHub autopush OK: auto: {short}")
            else:
                print(f"  ⚠️ Autopush: error actualizando ref ({st})")

        except Exception as e:
            print(f"  ⚠️  Autopush error: {e}")

    threading.Thread(target=_push, daemon=True).start()


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        local_ip = "127.0.0.1"

    _setup_git_autopush()   # configura git remote con token si estamos en cloud
    server = HTTPServer(("0.0.0.0", PORT), Handler)

    print(f"\n{'═'*58}")
    print(f"  📱  LABOY PICKS — DASHBOARD MÓVIL")
    print(f"{'═'*58}")
    print(f"\n  🏠  Red local:  http://{local_ip}:{PORT}")
    print(f"  💻  Localhost:  http://127.0.0.1:{PORT}")
    print(f"\n  Para acceso REMOTO desde cualquier lugar:")
    print(f"     ngrok http {PORT}   →  copia la URL https://")
    print(f"\n  Ctrl+C para detener\n{'═'*58}\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  🛑  Servidor detenido.\n")
        server.server_close()
