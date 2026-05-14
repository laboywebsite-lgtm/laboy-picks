"""
LABOY PICKS — BSN (Puerto Rico Basketball)
══════════════════════════════════════════════════════
  python3 bsn.py                        → líneas del modelo para hoy
  python3 bsn.py 2026-04-12             → fecha específica
  python3 bsn.py --lines                → líneas detalladas
  python3 bsn.py --picks                → picks con edge de modelo
  python3 bsn.py --stats                → estadísticas blended por equipo
  python3 bsn.py --ir                   → injury report actual
  python3 bsn.py --add-injury TEAM PLAYER RATE
                                        → agrega lesión (scrape PPG/USG% de RealGM)
  python3 bsn.py --remove-injury TEAM PLAYER
                                        → elimina jugador del IR
  python3 bsn.py --refresh              → actualiza stats 2026 desde RealGM
  python3 bsn.py --schedule             → muestra próximos juegos en Excel
  python3 bsn.py --export-html          → genera HTML de picks + lines (mismo diseño que MLB)
  python3 bsn.py --add-game VISIT LOCAL HORA
                                        → agrega juego manual para hoy
  (ej: python3 bsn.py --add-game SANTEROS LEONES '8:00 PM')
  python3 bsn.py --remove-game VISIT LOCAL [FECHA]
                                        → elimina juego manual
  python3 bsn.py --list-games           → lista juegos manuales de hoy
  python3 bsn.py --list-games all       → lista todos los juegos manuales

  RATES de lesión:
    1 = Out / Inactivo      (70% del valor del jugador perdido)
    2 = Dudoso / Doubtful   (75% del valor perdido)
    3 = Limitado / Limited  (80% del valor perdido)

  MODELO:
    Expected Points  = (ORTG_A + DRTG_B) / 2 × PACE / 100
    Blend stats      = 80% 2026 + 20% 2025
    Injury Impact    = PPG × USG% × rate_factor
    Win Prob         = Pythagorean exp=13.91

══════════════════════════════════════════════════════
"""

import sys, os, re, json, warnings, hashlib
import requests
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from copy import copy
import math

warnings.filterwarnings("ignore")

# ── tabulate (opcional) ───────────────────────────────
try:
    from tabulate import tabulate as _tabulate
    def tab(rows, headers, fmt="rounded_outline", **kw):
        return _tabulate(rows, headers=headers, tablefmt=fmt, **kw)
except ImportError:
    def tab(rows, headers, fmt=None, **kw):
        if not rows: return "  ".join(str(h) for h in headers)
        w = [max(len(str(h)), max((len(str(r[i])) for r in rows), default=0))
             for i, h in enumerate(headers)]
        sep = "  ".join(str(h).ljust(w[i]) for i, h in enumerate(headers))
        lines = [sep, "─"*len(sep)]
        for row in rows:
            lines.append("  ".join(str(c).ljust(w[i]) for i, c in enumerate(row)))
        return "\n".join(lines)

# ──────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE   = os.path.join(SCRIPT_DIR, "Laboy Picks - Data Model Module - Last Version.xlsx")
LOG_FILE          = os.path.join(SCRIPT_DIR, "bsn_picks_log.json")
GP_FILE           = os.path.join(SCRIPT_DIR, "bsn_gp.json")          # juegos jugados por equipo
MODEL_PICKS_FILE  = os.path.join(SCRIPT_DIR, "bsn_model_picks.json") # histórico picks del modelo

GITHUB_PAGES_REPO = os.environ.get(
    "BSN_GITHUB_REPO",
    os.path.join(os.path.expanduser("~"), "repos", "bsn-picks")
)
GITHUB_PAGES_URL  = "https://laboywebsite-lgtm.github.io/bsn-picks"

# ── URL token — seguridad por oscuridad ──────────────
# Salt privado. Cambia (o pon env var BSN_TOKEN_SALT) para rotar tokens.
_URL_TOKEN_SALT = os.environ.get("BSN_TOKEN_SALT", "laboyBsnSalt2026")

def _url_token(date_str: str) -> str:
    """Token corto (7 chars) determinístico basado en la fecha + salt."""
    raw = f"{date_str}:{_URL_TOKEN_SALT}"
    return hashlib.sha256(raw.encode()).hexdigest()[:7]

# Dashboard secreto — solo tú tienes la URL.
DASHBOARD_TOKEN = os.environ.get("BSN_DASHBOARD_TOKEN", "bsN7m2Pk")

ODDS_API_KEY = os.environ.get("ODDS_API_KEY", "524c2c3a534298ebbd212c6dc621a458")

BSN_LINES_SHEET    = "BSN Lines"
BSN_ADV_SHEET      = "BSN - Advanced"
IR_SHEET           = "IR - BSN"
INJ_IMPACT_SHEET   = "INJURY IMPACT - BSN"

# ── Help ──────────────────────────────────────────────
if "--help" in sys.argv or "-h" in sys.argv:
    print("""
╔══════════════════════════════════════════════════════════════════╗
║            🏀  LABOY PICKS — BSN  |  Comandos                    ║
╚══════════════════════════════════════════════════════════════════╝

  DATOS Y LÍNEAS
  ──────────────────────────────────────────────────────────────
  python3 bsn.py                     Líneas del modelo para hoy
  python3 bsn.py 2026-04-15          Lo mismo pero para una fecha específica
  python3 bsn.py --lines             Muestra líneas detalladas
  python3 bsn.py --picks             Picks con edge del modelo
  python3 bsn.py --set-lines         Ingresa odds de mercado (ML/Spread/Total) manualmente
  python3 bsn.py --set-lines DATE    Ingresa odds para una fecha específica
  python3 bsn.py --clear-lines       Elimina líneas guardadas de TARGET_DATE
  python3 bsn.py --clear-lines DATE  Elimina líneas de una fecha específica
  python3 bsn.py --clear-lines --all Elimina TODAS las líneas guardadas
  python3 bsn.py --stats             Estadísticas blended por equipo
  python3 bsn.py --ir                Injury report interactivo (ver / agregar / remover jugadores)
  python3 bsn.py --refresh           Actualiza stats 2026

  PICKS — LOG
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --log               Registra una jugada nueva (te pregunta los datos)
  python3 bsn.py --grade IDX W|L|P   Califica jugada por número (W=Win L=Loss P=Push)
                                     Ejemplo: python3 bsn.py --grade 3 W
  python3 bsn.py --remove IDX        Elimina el pick #IDX del log
  python3 bsn.py --remove IDX1 IDX2  Elimina múltiples picks (ej: --remove 2 4 7)
  python3 bsn.py --export-log        Re-exporta el último pick como HTML
  python3 bsn.py --export-log N      Re-exporta el pick #N como HTML

  ESTADÍSTICAS Y RÉCORD
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --record            Muestra récord completo de jugadas (W/L/P + P&L)
  python3 bsn.py --feedback          Análisis de rendimiento por tipo de pick

  EXPORTAR HTML
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --export-html              Genera HTMLs con picks + lines del modelo
  python3 bsn.py --export-html --publish    Exporta y sube a GitHub Pages (bsn-picks)

  MODELO — CALIBRACIÓN HISTÓRICA
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --grade-picks              Parsea HTML + pide scores → W/L/P + card
  python3 bsn.py --grade-picks DATE         Evalúa picks de una fecha específica
  python3 bsn.py --grade-picks DATE --publish   Evalúa Y publica Picks + Model Card
  python3 bsn.py --export-record            Exporta tarjeta de récord (all-time)
  python3 bsn.py --export-record DATE       Exporta récord de una fecha específica
  python3 bsn.py --export-record --publish  Exporta récord y lo publica en GitHub Pages

  BLEND DINÁMICO — JUEGOS JUGADOS
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --gp EQUIPO1 N1 EQUIPO2 N2 ...
                                     Registra juegos jugados por equipo para blend dinámico.
                                     El modelo usa peso_2026 = GP / (GP + 7).
                                     Ejemplo: python3 bsn.py --gp SANTEROS 7 CAPITANES 15
                                     Con 7 GP  → 50% 2026 / 50% 2025
                                     Con 15 GP → 68% 2026 / 32% 2025
                                     Con 20 GP → 100% 2026 (credibilidad plena)

  OTROS
  ──────────────────────────────────────────────────────────────
  python3 bsn.py --help              Muestra este menú
""")
    sys.exit(0)

# ── Parse flags ───────────────────────────────────────
LINES_MODE   = "--lines"          in sys.argv
PICKS_MODE   = "--picks"          in sys.argv
STATS_MODE   = "--stats"          in sys.argv
IR_MODE      = "--ir"             in sys.argv
REFRESH_MODE = "--refresh"        in sys.argv
SCHEDULE_MODE    = "--schedule"        in sys.argv
DEBUG_SCHEDULE   = "--debug-schedule"  in sys.argv

ADD_INJURY_MODE    = "--add-injury"    in sys.argv
REMOVE_INJURY_MODE = "--remove-injury" in sys.argv
ADD_GAME_MODE      = "--add-game"      in sys.argv
REMOVE_GAME_MODE   = "--remove-game"   in sys.argv
EDIT_GAME_MODE     = "--edit-game"     in sys.argv
LIST_GAMES_MODE    = "--list-games"    in sys.argv

EXPORT_LINES     = "--export-lines"      in sys.argv   # PNG de model lines
EXPORT_PICKS     = "--export-picks"      in sys.argv   # PNG de picks del log
EXPORT_LINES_PDF = "--export-lines-pdf"  in sys.argv   # PDF de lines
EXPORT_PICKS_PDF = "--export-picks-pdf"  in sys.argv   # PDF de picks
EXPORT_STORY     = "--export-story"      in sys.argv   # story 1080×1920 (alias picks)
EXPORT_POST      = "--export-post"       in sys.argv   # post 1080×1080 (recortado)
EXPORT_HTML_MODE = "--export-html"       in sys.argv   # HTML lines + picks (mismo diseño que MLB)

# ── Pick tracking flags ────────────────────────────────
LOG_MODE         = "--log"         in sys.argv
LOG_RETRO_MODE   = "--log-retro"   in sys.argv
LOG_PARLAY_MODE  = "--log-parlay"  in sys.argv
LOG_SPECIAL_MODE = "--log-special" in sys.argv   # Combo pick mismo juego (ML + Total, etc.)
SEASON_CARD_MODE = "--season-card" in sys.argv
GRADE_MODE       = "--grade"       in sys.argv
RECORD_MODE      = "--record"      in sys.argv
FEEDBACK_MODE    = "--feedback"    in sys.argv
EXPORT_LOG_MODE  = "--export-log"  in sys.argv
REMOVE_MODE      = "--remove"      in sys.argv
EDIT_MODE        = "--edit"        in sys.argv
PUBLISH_MODE     = "--publish"     in sys.argv
FORCE_EXPORT     = "--force-export" in sys.argv   # sobreescribir picks HTML aunque ya exista
GP_MODE          = "--gp"          in sys.argv
GRADE_PICKS_MODE = "--grade-picks" in sys.argv
EXPORT_RECORD_MODE = "--export-record" in sys.argv
SERVE_MODE       = "--serve"        in sys.argv
SET_LINES_MODE   = "--set-lines"    in sys.argv
CLEAR_LINES_MODE = "--clear-lines"  in sys.argv

# ── Positional args ───────────────────────────────────
_skip = set()
for _flag, _n in [("--add-injury",3), ("--remove-injury",2), ("--add-game",3), ("--remove-game",2),
                   ("--edit-game",4), ("--edit",3),
                   ("--grade",2), ("--remove",100), ("--export-log",1)]:
    if _flag in sys.argv:
        fi = sys.argv.index(_flag)
        if _flag == "--remove":
            # Especial para --remove que puede tener múltiples args
            j = fi + 1
            while j < len(sys.argv) and not sys.argv[j].startswith("--"):
                _skip.add(j)
                j += 1
        else:
            for x in range(1, min(_n+1, len(sys.argv)-fi)):
                if fi+x < len(sys.argv): _skip.add(fi+x)

args        = [a for i,a in enumerate(sys.argv[1:],1)
               if not a.startswith("--") and i not in _skip]
TARGET_DATE = args[0] if args else date.today().strftime("%Y-%m-%d")

# ──────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────

# Nombres cortos de equipos
BSN_TEAMS = [
    "ATLETICOS","CANGREJEROS","CAPITANES","CRIOLLOS","GIGANTES",
    "INDIOS","LEONES","METS","OSOS","PIRATAS","SANTEROS","VAQUEROS",
    "CAVALIERS",
]

# Nombre completo en RealGM → nombre corto
REALGM_NAME_MAP = {
    "atleticos de san german":      "ATLETICOS",
    "cangrejeros de santurce":      "CANGREJEROS",
    "capitanes de arecibo":         "CAPITANES",
    "criollos de caguas":           "CRIOLLOS",
    "gigantes de carolina":         "GIGANTES",
    "indios de mayaguez":           "INDIOS",
    "leones de ponce":              "LEONES",
    "mets de guaynabo":             "METS",
    "osos de manati":               "OSOS",
    "piratas de quebradillas":      "PIRATAS",
    "santeros de aguada":           "SANTEROS",
    "vaqueros de bayamon":          "VAQUEROS",
    "cavaliers de isabela":         "CAVALIERS",
    # aliases
    "atleticos":    "ATLETICOS",
    "cangrejeros":  "CANGREJEROS",
    "capitanes":    "CAPITANES",
    "criollos":     "CRIOLLOS",
    "gigantes":     "GIGANTES",
    "indios":       "INDIOS",
    "leones":       "LEONES",
    "mets":         "METS",
    "osos":         "OSOS",
    "piratas":      "PIRATAS",
    "santeros":     "SANTEROS",
    "vaqueros":     "VAQUEROS",
    "cavaliers":    "CAVALIERS",
}

# Factor por rate de lesión
RATE_FACTOR = {1: 0.70, 2: 0.75, 3: 0.80}

# Pythagorean exponent para baloncesto
PYTH_EXP = 13.91

# Liga — promedio de puntos por juego (referencia)
LEAGUE_AVG_PACE = 78.5

# Canchas por equipo LOCAL (Team2 siempre es el local)
HOME_VENUES = {
    "ATLETICOS":   "Coliseo Edgardo Zayas Hernández, San Germán",
    "CANGREJEROS": "Coliseo Roberto Clemente, Santurce",
    "CAPITANES":   "Coliseo de Arecibo, Arecibo",
    "CRIOLLOS":    "Coliseo Guillermo Angulo, Caguas",
    "GIGANTES":    "Coliseo Osvaldo Martínez, Carolina",
    "INDIOS":      "Coliseo Ruben Rodríguez, Mayagüez",
    "LEONES":      "Coliseo Municipal de Ponce, Ponce",
    "METS":        "Coliseo Héctor Solá Bezares, Guaynabo",
    "OSOS":        "Coliseo Jesús 'Chuíto' Morales, Manatí",
    "PIRATAS":     "Coliseo Municipio de Quebradillas, Quebradillas",
    "SANTEROS":    "Coliseo de Aguada, Aguada",
    "VAQUEROS":    "José M. Agrelot Coliseum, Bayamón",
    "CAVALIERS":   "Coliseo de Isabela, Isabela",
}

# Colores por equipo BSN (para HTML export cards)
BSN_TEAM_COLORS = {
    "ATLETICOS":   "#0057A8",
    "CANGREJEROS": "#CC0000",
    "CAPITANES":   "#1a3a6b",
    "CRIOLLOS":    "#6B2D8B",
    "GIGANTES":    "#CC0000",
    "INDIOS":      "#8B0000",
    "LEONES":      "#c8a000",
    "METS":        "#003087",
    "OSOS":        "#3a3a3a",
    "PIRATAS":     "#111111",
    "SANTEROS":    "#b30000",
    "VAQUEROS":    "#8B0000",
    "CAVALIERS":   "#CC0000",
}

# Mapa de nombres bsnpr.com → nombre corto
BSNPR_NAME_MAP = {
    # Nombres comunes en el sitio oficial
    "atleticos":      "ATLETICOS",
    "cangrejeros":    "CANGREJEROS",
    "capitanes":      "CAPITANES",
    "criollos":       "CRIOLLOS",
    "gigantes":       "GIGANTES",
    "indios":         "INDIOS",
    "leones":         "LEONES",
    "mets":           "METS",
    "osos":           "OSOS",
    "piratas":        "PIRATAS",
    "santeros":       "SANTEROS",
    "vaqueros":       "VAQUEROS",
    "cavaliers":      "CAVALIERS",
    # Con ciudad
    "san german":     "ATLETICOS",
    "santurce":       "CANGREJEROS",
    "arecibo":        "CAPITANES",
    "caguas":         "CRIOLLOS",
    "carolina":       "GIGANTES",
    "mayaguez":       "INDIOS",
    "ponce":          "LEONES",
    "guaynabo":       "METS",
    "manati":         "OSOS",
    "quebradillas":   "PIRATAS",
    "aguada":         "SANTEROS",
    "bayamon":        "VAQUEROS",
    "isabela":        "CAVALIERS",
}

# ──────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────

def norm_team(name):
    """Normaliza nombre de equipo a nombre corto."""
    if not name: return None
    n = str(name).strip().upper()
    if n in BSN_TEAMS: return n
    low = n.lower()
    return REALGM_NAME_MAP.get(low, n)

def prob_to_american(p):
    """Probabilidad → odds americanos."""
    p = max(0.01, min(0.99, p))
    if p >= 0.5:
        return int(round(-(p / (1-p)) * 100))
    else:
        return int(round(((1-p) / p) * 100))

def fmt_odds(o):
    return f"+{o}" if o >= 0 else str(o)

def pyth_win_prob(pts_a, pts_b):
    """Win probability vía Pythagorean exponent basketball."""
    if pts_a <= 0 or pts_b <= 0: return 0.5
    return pts_a**PYTH_EXP / (pts_a**PYTH_EXP + pts_b**PYTH_EXP)

def _parse_time_sort(t):
    if not t: return 9999
    try:
        parts = t.split(); h,m = map(int,parts[0].split(':'))
        ap = parts[1].upper() if len(parts)>1 else 'PM'
        if ap=='PM' and h!=12: h+=12
        if ap=='AM' and h==12: h=0
        return h*100+m
    except: return 9999

# ──────────────────────────────────────────────────────
# GAMES PLAYED (GP) — blend dinámico por equipo
# ──────────────────────────────────────────────────────
#
# BSN temporada regular = 34 juegos.
# Filosofía del blend:
#   - Inicio de temporada: datos previos dominan (alta incertidumbre)
#   - Ramp-up Bayesiano acelerado hasta el juego ~7 (crossover 50/50)
#   - A partir del juego 20 (59% de la temporada): 100% temporada actual
#     porque a mitad de temporada los equipos BSN ya mostraron su nivel real.
#     El año pasado en ese punto empieza a ser ruido, no señal.
#
# Curva resultante (GP_REGRESSION=7):
#   0 GP → 10% actual  (sin GP registrado: casi todo prior year)
#   4 GP → 36% actual
#   7 GP → 50% actual  ← crossover
#  10 GP → 59% actual
#  15 GP → 68% actual
#  17 GP → 71% actual
#  20 GP → 100% actual ← credibilidad plena (59% de temporada)
#  34 GP → 100% actual
#
# Cambio vs anterior (GP_REGRESSION=10, BSN_FULL_CRED_GP=25):
#   A 17GP: 63% → 71% 2026 (+8pp)
#   A 20GP: 67% → 100% 2026 (+33pp)  ← mayor impacto

BSN_SEASON_GP    = 34   # juegos totales temporada regular BSN
BSN_FULL_CRED_GP = 20   # GP mínimo para credibilidad plena (100% temporada actual)
GP_REGRESSION    = 7    # factor Bayesiano: crossover 50/50 en juego 7

def _load_gp():
    """Carga juegos jugados por equipo desde bsn_gp.json."""
    if os.path.exists(GP_FILE):
        with open(GP_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def _save_gp(gp_dict):
    with open(GP_FILE, "w", encoding="utf-8") as f:
        json.dump(gp_dict, f, indent=2, ensure_ascii=False)

def _blend_weight(gp):
    """
    Retorna peso temporada actual (2026) basado en juegos jugados.

    Dos fases:
      1. Ramp-up Bayesiano (0–19 GP): GP / (GP + GP_REGRESSION)
         Crossover 50/50 en el juego 7 (~21% de temporada).
      2. Credibilidad plena (≥20 GP): 100% temporada actual.
         Después de 20 juegos de una temporada de 34, el año pasado
         ya no agrega señal útil.
    """
    if not gp or gp <= 0:
        return 0.10   # inicio de temporada: 90% prior year
    if gp >= BSN_FULL_CRED_GP:
        return 1.0    # ≥20 GP → 100% temporada actual
    return round(gp / (gp + GP_REGRESSION), 4)

def cmd_set_gp():
    """
    --gp TEAM1 N1 TEAM2 N2 ...
    Registra juegos jugados por equipo para blend dinámico.
    Ej: python3 bsn.py --gp SANTEROS 6 CAPITANES 8
    """
    try:
        idx   = sys.argv.index("--gp")
        args  = sys.argv[idx+1:]
    except ValueError:
        print("  ❌ Uso: python3 bsn.py --gp EQUIPO1 GP1 EQUIPO2 GP2 ...")
        return

    if not args or len(args) % 2 != 0:
        print("  ❌ Debes pasar pares EQUIPO + número de juegos.")
        print("     Ej: python3 bsn.py --gp SANTEROS 6 CAPITANES 8")
        return

    gp_dict = _load_gp()
    updated = []
    for i in range(0, len(args), 2):
        raw_team = args[i].upper().strip()
        team     = norm_team(raw_team) or raw_team
        try:
            gp = int(args[i+1])
            assert gp >= 0
        except (ValueError, AssertionError):
            print(f"  ⚠️  Número inválido para {raw_team}: '{args[i+1]}' — ignorado")
            continue
        gp_dict[team] = gp
        w = _blend_weight(gp)
        updated.append((team, gp, w))

    _save_gp(gp_dict)

    print(f"\n  📊 GP actualizado ({len(updated)} equipos):\n")
    print(f"  {'Equipo':<20} {'GP':>4}   {'2026':>6}  {'2025':>6}  Blend")
    print(f"  {'─'*20}  {'─'*4}   {'─'*6}  {'─'*6}  {'─'*18}")
    for team, gp, w in updated:
        pct26 = f"{w*100:.0f}%"
        pct25 = f"{(1-w)*100:.0f}%"
        bar   = "█" * int(w * 20) + "░" * (20 - int(w * 20))
        print(f"  {team:<20} {gp:>4}   {pct26:>6}  {pct25:>6}  {bar}")

    # Mostrar todos los equipos registrados
    print(f"\n  📋 Estado completo de GP:")
    print(f"  {'Equipo':<20} {'GP':>4}   {'2026':>6}  {'2025':>6}")
    print(f"  {'─'*20}  {'─'*4}   {'─'*6}  {'─'*6}")
    for team, gp in sorted(gp_dict.items()):
        w = _blend_weight(gp)
        print(f"  {team:<20} {gp:>4}   {w*100:.0f}%     {(1-w)*100:.0f}%")
    print()


# ──────────────────────────────────────────────────────
# LEER DATOS DE EXCEL
# ──────────────────────────────────────────────────────

def load_bsn_advanced(wb):
    """Lee stats blended desde BSN - Advanced (col C-F, filas 5+).
    El blend es DINÁMICO por equipo según GP registrado en bsn_gp.json:
        0–19 GP : GP / (GP + GP_REGRESSION)  [Bayesiano, crossover en GP=7]
        ≥20 GP  : 100% temporada actual       [credibilidad plena]
    Si D-F tienen None/fórmula sin caché, recalcula blend desde cols I-K/N-P.
    PACE siempre usa 100% 2026 (col K).
    """
    # Load JSON 2026 stats override
    _bsn_stats_json = {}
    _stats_json_path = os.path.join(SCRIPT_DIR, "bsn_team_stats.json")
    if os.path.exists(_stats_json_path):
        try:
            with open(_stats_json_path, "r", encoding="utf-8") as _f:
                _bsn_stats_json = json.load(_f)
        except Exception:
            pass

    ws     = wb[BSN_ADV_SHEET]
    data   = {}
    gp_map = _load_gp()   # {team_short: gp}

    # Leer tabla 2025 (col M=team, N=ortg, O=drtg, P=pace) para fallback
    tbl25 = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        t25 = row[12]  # col M
        if not t25: continue
        try:
            tbl25[str(t25).strip()] = {
                "ortg": float(row[13]) if row[13] else None,
                "drtg": float(row[14]) if row[14] else None,
            }
        except Exception:
            pass

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        raw_team = row[2]   # col C (short name)
        if not raw_team: continue
        team = norm_team(str(raw_team))
        if not team: continue

        # Blend weight — dinámico según GP registrado para este equipo
        gp = gp_map.get(team)
        BL = _blend_weight(gp)   # peso 2026; (1-BL) = peso 2025

        ortg_cached = row[3]   # col D (blend cacheado en Excel)
        drtg_cached = row[4]   # col E
        pace_cached = row[5]   # col F

        # JSON override (bsn_team_stats.json) takes priority for 2026 stats
        _js = _bsn_stats_json.get(team, {})
        o26 = _js.get("ortg") if _js.get("ortg") else row[8]    # col I — ORtg 2026
        d26 = _js.get("drtg") if _js.get("drtg") else row[9]    # col J — DRtg 2026
        p26 = _js.get("pace") if _js.get("pace") else row[10]   # col K — Pace 2026
        h26 = str(row[7]).strip() if row[7] else ""  # col H = nombre completo 2026

        team25 = tbl25.get(h26, {})
        o25    = team25.get("ortg")
        d25    = team25.get("drtg")

        def _get(cached, v26_raw, v25_raw, fallback, use_blend=True):
            """Usa valor cacheado si es float válido; si no, recalcula con BL dinámico."""
            # Intentar usar valor cacheado del Excel
            if cached is not None:
                try:
                    cached_f = float(cached)
                    # Si hay GP registrado, ignoramos el cached (que usa blend fijo Excel)
                    # y recalculamos con el BL dinámico correcto
                    if gp is not None and use_blend:
                        raise ValueError("recalculate with dynamic BL")
                    return cached_f
                except (ValueError, TypeError):
                    pass
            # Recalcular con blend dinámico desde columnas fuente
            if use_blend:
                try:
                    v26 = float(v26_raw) if v26_raw else None
                    v25 = float(v25_raw) if v25_raw else None
                    if v26 is not None and v25 is not None:
                        return round(BL * v26 + (1 - BL) * v25, 2)
                    if v26 is not None:
                        return round(v26, 2)
                except Exception:
                    pass
            else:
                # Sin blend (PACE): usar 100% v26
                try:
                    return float(v26_raw) if v26_raw else fallback
                except Exception:
                    pass
            return fallback

        try:
            data[team] = {
                "ortg": _get(ortg_cached, o26, o25, 100.0),
                "drtg": _get(drtg_cached, d26, d25, 100.0),
                "pace": _get(pace_cached, p26, None, LEAGUE_AVG_PACE, use_blend=False),
                "gp":   gp,    # guardamos GP para mostrarlo en --stats
                "blend": BL,
            }
        except Exception:
            pass
    return data

def load_injury_impact(wb):
    """Lee impacto total por equipo desde INJURY IMPACT - BSN."""
    ws     = wb[INJ_IMPACT_SHEET]
    impact = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        raw_team = row[1]   # col B
        val      = row[2]   # col C
        if not raw_team: continue
        team = norm_team(str(raw_team))
        if not team: continue
        try:    impact[team] = float(val) if val else 0.0
        except: impact[team] = 0.0
    return impact

def load_ir_entries(wb):
    """Lee entradas del IR desde IR - BSN.  Solo equipos BSN (excluye NBA)."""
    ws      = wb[IR_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_team   = row[1]   # col B
        player     = row[2]   # col C
        rate       = row[3]   # col D
        ppg        = row[4]   # col E
        usg        = row[5]   # col F
        impact     = row[6]   # col G
        if not raw_team and not player: continue
        team = norm_team(str(raw_team)) if raw_team else None
        if not team or not player: continue
        # ── Filtrar solo equipos BSN — excluir equipos NBA ─────────────
        if team not in BSN_TEAMS: continue
        try:
            entries.append({
                "team":   team,
                "player": str(player).strip().upper(),
                "rate":   int(rate) if rate else 0,
                "ppg":    float(ppg) if ppg else 0.0,
                "usg":    float(usg) if usg else 0.0,
                "impact": float(impact) if impact else 0.0,
            })
        except: pass
    return entries

def _norm_bsnpr(name):
    """Normaliza nombre de equipo desde bsnpr.com a nombre corto."""
    if not name: return None
    n = name.strip().upper()
    if n in BSN_TEAMS: return n
    low = name.strip().lower()
    # Chequea mapa directo
    if low in BSNPR_NAME_MAP: return BSNPR_NAME_MAP[low]
    # Chequea si alguna keyword del mapa está contenida
    for key, val in BSNPR_NAME_MAP.items():
        if key in low: return val
    # Fallback al mapa general
    return REALGM_NAME_MAP.get(low, n)


def scrape_realgm_schedule(target_date_str):
    """
    Scrape el calendario BSN desde basketball.realgm.com/international/league/62/
    Retorna lista de dicts: {date, team1(away), team2(home), game_time, venue, source}

    Estrategias:
    1. Per-date URL  /schedules/YYYY-MM-DD  — solo trae ese día, no hay que filtrar por fecha
    2. Full schedule /schedules              — filtra por fecha en cada fila
    """
    target_dt  = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    games      = []
    clean      = lambda s: re.sub(r'<[^>]+>', '', s).strip()

    def _norm_realgm(name):
        """Mapea nombre completo RealGM → nombre corto BSN."""
        if not name: return None
        n = name.strip().upper()
        if n in BSN_TEAMS: return n
        low = name.strip().lower()
        # Exact map lookup
        if low in REALGM_NAME_MAP: return REALGM_NAME_MAP[low]
        # Substring match (e.g. "Vaqueros" matches "vaqueros de bayamon")
        for key, val in REALGM_NAME_MAP.items():
            if key in low: return val
        for t in BSN_TEAMS:
            if t.lower() in low: return t
        return None

    def _extract_from_rows(rows, filter_date=None):
        """
        Parsea lista de <tr> HTML.
        filter_date: si se da, solo acepta filas cuya fecha coincida.
        Si es None, acepta todas las filas con ≥2 equipos.
        """
        result = []
        seen   = set()

        # Detect column headers from <th> row
        col_map = {}   # lower_header → column index
        for row_html in rows:
            ths = re.findall(r'<th[^>]*>(.*?)</th>', row_html, re.DOTALL|re.IGNORECASE)
            if ths:
                for i, h in enumerate(ths):
                    col_map[clean(h).lower()] = i
                break

        for row_html in rows:
            cells_raw = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL|re.IGNORECASE)
            if len(cells_raw) < 2:
                continue
            cell_vals = [clean(c) for c in cells_raw]
            raw_row   = " ".join(cell_vals).lower()

            # ── Date filtering (only when scraping the full schedule page) ──
            if filter_date is not None:
                row_dt = None
                # Try to find a date in the row cells
                date_m = re.search(
                    r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+'
                    r'(\d{1,2}),?\s*(\d{4})',
                    raw_row, re.IGNORECASE)
                if date_m:
                    try:
                        ds = f"{date_m.group(1)[:3].capitalize()} {date_m.group(2)} {date_m.group(3)}"
                        row_dt = datetime.strptime(ds, "%b %d %Y").date()
                    except: pass
                if not row_dt:
                    num_m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', raw_row)
                    if num_m:
                        try:
                            row_dt = datetime.strptime(
                                f"{num_m.group(1)}/{num_m.group(2)}/{num_m.group(3)}",
                                "%m/%d/%Y" if len(num_m.group(3))==4 else "%m/%d/%y"
                            ).date()
                        except: pass
                if not row_dt:
                    iso_m = re.search(r'(\d{4})-(\d{2})-(\d{2})', raw_row)
                    if iso_m:
                        try: row_dt = datetime.strptime(iso_m.group(0), "%Y-%m-%d").date()
                        except: pass
                if row_dt is None or row_dt != filter_date:
                    continue

            # ── Team extraction ───────────────────────────────────────────
            # Strategy A: use known column indices (Away, Home, Visitor, etc.)
            t1, t2 = None, None
            away_keys = ["away", "visitor", "away team", "visitante"]
            home_keys = ["home", "home team", "local"]
            for k in away_keys:
                if k in col_map:
                    idx = col_map[k]
                    if idx < len(cell_vals):
                        t1 = _norm_realgm(cell_vals[idx])
                    break
            for k in home_keys:
                if k in col_map:
                    idx = col_map[k]
                    if idx < len(cell_vals):
                        t2 = _norm_realgm(cell_vals[idx])
                    break

            # Strategy B: search all cells for team names
            if not (t1 and t2 and t1 != t2):
                found = []
                for cv in cell_vals:
                    nt = _norm_realgm(cv)
                    if nt and nt not in found:
                        found.append(nt)
                # Fallback: scan raw text for team keywords
                if len(found) < 2:
                    for alias, team in REALGM_NAME_MAP.items():
                        if alias in raw_row and team not in found:
                            found.append(team)
                if len(found) >= 2:
                    t1, t2 = found[0], found[1]

            if not (t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS):
                continue

            # ── Time extraction ───────────────────────────────────────────
            game_time = ""
            for cv in cell_vals:
                tm = re.search(r'(\d{1,2}:\d{2}\s*(?:AM|PM|am|pm))', cv)
                if tm:
                    game_time = tm.group(1).upper()
                    break

            key = tuple(sorted([t1, t2]))
            if key not in seen:
                seen.add(key)
                result.append({
                    "date":      target_dt,
                    "team1":     t1,
                    "team2":     t2,
                    "game_time": game_time,
                    "venue":     HOME_VENUES.get(t2, ""),
                    "game":      f"{t1} vs. {t2}",
                    "source":    "RealGM",
                })
        return result

    # ── Estrategia 1: Per-date URL (solo trae ese día) ────────────────────
    per_date_urls = [
        f"https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/schedules/{target_date_str}",
        f"https://basketball.realgm.com/international/league/62/puerto-rican-bsn/schedules/{target_date_str}",
    ]
    for url in per_date_urls:
        html = _fetch_html(url, silent=True)
        if html and len(html) > 500:
            rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL|re.IGNORECASE)
            result = _extract_from_rows(rows, filter_date=None)   # all rows = target date
            if result:
                return result
            break   # got a page but no games — date probably has none

    # ── Estrategia 2: Full schedule page, filtra por fecha ────────────────
    year = target_dt.year
    full_urls = [
        f"https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/schedules/{year}",
        f"https://basketball.realgm.com/international/league/62/puerto-rican-bsn/schedules/{year}",
        "https://basketball.realgm.com/international/league/62/puerto-rican-bsn/schedules",
        "https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/schedules",
    ]
    for url in full_urls:
        html = _fetch_html(url, silent=True)
        if html and len(html) > 1000:
            rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL|re.IGNORECASE)
            result = _extract_from_rows(rows, filter_date=target_dt)
            if result:
                return result
            break   # parsed but no match for this date

    return games


def _parse_bsnpr_game(event_dict, target_dt):
    """
    Intenta extraer un juego de un dict de evento (WordPress REST / JSON embebido).
    Retorna dict de juego o None.
    """
    raw = json.dumps(event_dict).lower()

    # Buscar equipos en el JSON del evento
    found_teams = []
    for t in BSN_TEAMS:
        if t.lower() in raw:
            found_teams.append(t)
    for alias, team in BSNPR_NAME_MAP.items():
        if alias in raw and team not in found_teams:
            found_teams.append(team)
    found_teams = list(dict.fromkeys(found_teams))
    if len(found_teams) < 2:
        return None

    # Hora
    # Campos comunes en The Events Calendar: start_date, date, start_date_details, etc.
    time_str = ""
    for key in ["start_date","date","startDate","start_date_details","start","hora","time"]:
        val = event_dict.get(key, "")
        if val and isinstance(val, str):
            t_match = re.search(r'(\d{1,2}:\d{2}\s*(?:AM|PM|am|pm)?)', val)
            if t_match:
                time_str = t_match.group(1).upper()
                break
    if not time_str and isinstance(event_dict.get("start_date_details"), dict):
        d = event_dict["start_date_details"]
        h, m = d.get("hour",""), d.get("minutes","")
        if h:
            hr = int(h); mn = int(m or 0)
            ampm = "PM" if hr >= 12 else "AM"
            if hr > 12: hr -= 12
            time_str = f"{hr}:{mn:02d} {ampm}"

    t1 = found_teams[0]
    t2 = found_teams[1]
    return {
        "date":      target_dt,
        "team1":     t1,
        "team2":     t2,
        "game_time": time_str,
        "venue":     HOME_VENUES.get(t2, ""),
        "game":      f"{t1} vs. {t2}",
        "source":    "bsnpr.com",
    }


def scrape_bsnpr_schedule(target_date_str):
    """
    Obtiene juegos BSN del día desde bsnpr.com.

    Estrategias (en orden):
    1. WordPress REST API: /wp-json/tribe/events/v1/events  (The Events Calendar plugin)
    2. WordPress REST API: /wp-json/wp/v2/tribe_events
    3. ESPN API (Puerto Rico basketball)
    4. HTML scraping como último recurso
    """
    target_dt  = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    date_s     = target_dt.strftime("%Y-%m-%d")
    games      = []

    def _dedup(games):
        seen = set()
        out  = []
        for g in games:
            key = tuple(sorted([g["team1"], g["team2"]]))
            if key not in seen:
                seen.add(key); out.append(g)
        return out

    # ── Estrategia 1: The Events Calendar REST API ─────────────────────────
    tribe_urls = [
        f"https://www.bsnpr.com/wp-json/tribe/events/v1/events?start_date={date_s}&per_page=20",
        f"https://bsnpr.com/wp-json/tribe/events/v1/events?start_date={date_s}&per_page=20",
        f"https://www.bsnpr.com/wp-json/tribe/events/v1/events?start_date={date_s}&end_date={date_s}&per_page=20",
    ]
    for url in tribe_urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                data = r.json()
                events = data.get("events", []) if isinstance(data, dict) else (data if isinstance(data, list) else [])
                for ev in events:
                    # Verificar que el evento sea del día
                    ev_date = str(ev.get("start_date","") or ev.get("date",""))[:10]
                    if ev_date != date_s:
                        continue
                    g = _parse_bsnpr_game(ev, target_dt)
                    if g: games.append(g)
                if games:
                    return _dedup(games)
        except Exception:
            pass

    # ── Estrategia 2: WordPress REST API estándar (tribe_events post type) ─
    wp_urls = [
        f"https://www.bsnpr.com/wp-json/wp/v2/tribe_events?per_page=20&before={date_s}T23:59:59&after={date_s}T00:00:00",
        f"https://www.bsnpr.com/wp-json/wp/v2/events?per_page=20&after={date_s}T00:00:00&before={date_s}T23:59:59",
        f"https://www.bsnpr.com/wp-json/wp/v2/posts?per_page=20&after={date_s}T00:00:00&before={date_s}T23:59:59&categories=juegos",
    ]
    for url in wp_urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                events = r.json()
                if isinstance(events, list):
                    for ev in events:
                        g = _parse_bsnpr_game(ev, target_dt)
                        if g: games.append(g)
                    if games:
                        return _dedup(games)
        except Exception:
            pass

    # ── Estrategia 3: ESPN API (Puerto Rico BSN) ───────────────────────────
    espn_urls = [
        f"https://site.api.espn.com/apis/site/v2/sports/basketball/bsn/scoreboard?dates={date_s.replace('-','')}",
        f"http://site.api.espn.com/apis/site/v2/sports/basketball/puerto-rico-bsn/scoreboard?dates={date_s.replace('-','')}",
        f"https://site.api.espn.com/apis/site/v2/sports/basketball/mens-bsn/scoreboard?dates={date_s.replace('-','')}",
    ]
    for url in espn_urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                data = r.json()
                for ev in data.get("events", []):
                    comps = ev.get("competitions", [{}])[0]
                    competitors = comps.get("competitors", [])
                    if len(competitors) < 2: continue
                    # ESPN: homeAway = "home" / "away"
                    away_c = next((c for c in competitors if c.get("homeAway")=="away"), competitors[0])
                    home_c = next((c for c in competitors if c.get("homeAway")=="home"), competitors[1])
                    t1 = _norm_bsnpr(away_c.get("team",{}).get("shortDisplayName","") or
                                     away_c.get("team",{}).get("displayName",""))
                    t2 = _norm_bsnpr(home_c.get("team",{}).get("shortDisplayName","") or
                                     home_c.get("team",{}).get("displayName",""))
                    if t1 and t2 and t1 != t2:
                        date_raw = ev.get("date","")
                        time_s   = ""
                        if date_raw:
                            try:
                                dt_utc = datetime.strptime(date_raw[:19], "%Y-%m-%dT%H:%M:%S")
                                # Convertir UTC→ AST (UTC-4)
                                from datetime import timezone, timedelta as _td
                                dt_ast = dt_utc - _td(hours=4)
                                hr = dt_ast.hour; mn = dt_ast.minute
                                ampm = "PM" if hr >= 12 else "AM"
                                if hr > 12: hr -= 12
                                if hr == 0: hr = 12
                                time_s = f"{hr}:{mn:02d} {ampm}"
                            except: pass
                        games.append({
                            "date": target_dt, "team1": t1, "team2": t2,
                            "game_time": time_s,
                            "venue": HOME_VENUES.get(t2, ""),
                            "game": f"{t1} vs. {t2}",
                            "source": "ESPN",
                        })
                if games:
                    return _dedup(games)
        except Exception:
            pass

    # ── Estrategia 4: RealGM BSN schedule ────────────────────────────────────
    realgm_games = scrape_realgm_schedule(target_date_str)
    if realgm_games:
        return _dedup(realgm_games)

    # ── Estrategia 5: HTML scraping (bsnpr.com puede ser JS-rendered) ──────
    month_year = target_dt.strftime("%Y-%m")
    html_urls  = [
        "https://www.bsnpr.com/calendario/",
        f"https://www.bsnpr.com/calendario/?mes={month_year}",
        "https://bsnpr.com/calendario/",
        "https://www.bsnpr.com/schedule/",
        "https://www.bsnpr.com/",
    ]
    html = None
    for url in html_urls:
        html = _fetch_html(url, silent=True)
        if html and len(html) > 500:
            break
    if not html:
        return []

    date_patterns = [
        date_s,
        target_dt.strftime("%d/%m/%Y"),
        target_dt.strftime("%B %d").lower(),
        target_dt.strftime("%-d de %B").lower() if hasattr(target_dt,"strftime") else "",
    ]

    # JSON embebido en script tags
    for script_match in re.finditer(r'<script[^>]*>(\{.*?})</script>', html, re.DOTALL):
        try:
            data = json.loads(script_match.group(1))
            raw  = json.dumps(data)
            if date_s in raw or any(dp in raw.lower() for dp in date_patterns if dp):
                g = _parse_bsnpr_game(data, target_dt)
                if g: games.append(g)
        except: pass
    if games: return _dedup(games)

    # vs-pattern en texto plano
    html_low = html.lower()
    if any(dp in html_low for dp in date_patterns if dp):
        vs_pat = re.compile(
            r'(' + '|'.join(re.escape(k) for k in list(BSN_TEAMS) + list(BSNPR_NAME_MAP.keys())) + r')'
            r'\s*(?:vs\.?|at|@|contra|-)\s*'
            r'(' + '|'.join(re.escape(k) for k in list(BSN_TEAMS) + list(BSNPR_NAME_MAP.keys())) + r')',
            re.IGNORECASE
        )
        for m in vs_pat.finditer(html):
            t1 = _norm_bsnpr(m.group(1))
            t2 = _norm_bsnpr(m.group(2))
            if t1 and t2 and t1 != t2:
                key = tuple(sorted([t1, t2]))
                if not any(tuple(sorted([g["team1"],g["team2"]]))==key for g in games):
                    games.append({
                        "date": target_dt, "team1": t1, "team2": t2,
                        "game_time": "", "venue": HOME_VENUES.get(t2, ""),
                        "game": f"{t1} vs. {t2}", "source": "bsnpr.com",
                    })

    return _dedup(games)


# ── Flashscore.com scraping ────────────────────────────────────────────────

# Mapping from Flashscore team name fragments → BSN short names
FLASHSCORE_NAME_MAP = {
    "santeros":    "SANTEROS",
    "leones":      "LEONES",
    "cangrejeros": "CANGREJEROS",
    "capitanes":   "CAPITANES",
    "piratas":     "PIRATAS",
    "criollos":    "CRIOLLOS",
    "gigantes":    "GIGANTES",
    "indios":      "INDIOS",
    "mets":        "METS",
    "osos":        "OSOS",
    "atleticos":   "ATLETICOS",
    "vaqueros":    "VAQUEROS",
    "cavaliers":   "CAVALIERS",
    "arecibo":     "CAPITANES",
    "bayamon":     "VAQUEROS",
    "caguas":      "CRIOLLOS",
    "carolina":    "GIGANTES",
    "coamo":       "OSOS",
    "guaynabo":    "METS",
    "humacao":     "INDIOS",
    "juncos":      "ATLETICOS",
    "mayaguez":    "INDIOS",
    "ponce":       "LEONES",
    "quebradilla": "CANGREJEROS",
    "san german":  "CANGREJEROS",
    "santurce":    "CANGREJEROS",
}


def _norm_flashscore(name):
    """Normaliza nombre de equipo Flashscore → BSN short name."""
    if not name: return None
    low = name.lower().strip()
    for k, v in FLASHSCORE_NAME_MAP.items():
        if k in low:
            return v
    # Try generic BSN team names
    for t in BSN_TEAMS:
        if t.lower() in low:
            return t
    return None


def scrape_espn_schedule(target_date_str):
    """
    Fetch BSN schedule from ESPN's public API.
    ESPN covers the BSN under league slug 'bsn'.
    Endpoint: site.api.espn.com/apis/site/v2/sports/basketball/bsn/scoreboard
    No authentication required.
    """
    target_dt = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    date_param = target_dt.strftime("%Y%m%d")   # ESPN uses YYYYMMDD

    espn_headers = {
        **HEADERS,
        "Accept":   "application/json",
        "Referer":  "https://www.espn.com/",
    }

    # ESPN league slugs to try for BSN Puerto Rico
    slugs = ["bsn", "wnba"]   # 'bsn' is the direct BSN slug if ESPN has it
    # Also try the general basketball scoreboard with date filter
    urls = [
        f"https://site.api.espn.com/apis/site/v2/sports/basketball/bsn/scoreboard?dates={date_param}",
        f"https://site.api.espn.com/apis/site/v2/sports/basketball/bsn/scoreboard?date={date_param}",
        # Fallback: general basketball endpoint filtering by date
        f"https://site.api.espn.com/apis/site/v2/sports/basketball/scoreboard?dates={date_param}&league=bsn",
    ]

    games = []
    for url in urls:
        try:
            r = requests.get(url, headers=espn_headers, timeout=15)
            if r.status_code != 200:
                continue
            data  = r.json()
            events = data.get("events", [])
            if not events:
                continue
            for ev in events:
                try:
                    # ESPN event date
                    ev_date_str = ev.get("date", "")[:10]  # "2026-04-10T..."
                    if ev_date_str != target_date_str:
                        continue
                    comps = ev.get("competitions", [{}])[0]
                    competitors = comps.get("competitors", [])
                    if len(competitors) < 2:
                        continue
                    away_c = next((c for c in competitors if c.get("homeAway") == "away"), competitors[0])
                    home_c = next((c for c in competitors if c.get("homeAway") == "home"), competitors[1])
                    raw_t1 = (away_c.get("team", {}).get("shortDisplayName") or
                              away_c.get("team", {}).get("abbreviation") or
                              away_c.get("team", {}).get("displayName", ""))
                    raw_t2 = (home_c.get("team", {}).get("shortDisplayName") or
                              home_c.get("team", {}).get("abbreviation") or
                              home_c.get("team", {}).get("displayName", ""))
                    t1 = norm_team(raw_t1)
                    t2 = norm_team(raw_t2)
                    if not (t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS):
                        continue
                    # Extract game time in AST (UTC-4)
                    hora = ""
                    ev_ts = ev.get("date", "")
                    if ev_ts:
                        try:
                            from datetime import timezone, timedelta as _td
                            dt_utc = datetime.strptime(ev_ts[:19], "%Y-%m-%dT%H:%M:%S").replace(
                                tzinfo=timezone.utc)
                            dt_ast = dt_utc - _td(hours=4)
                            h, mn  = dt_ast.hour, dt_ast.minute
                            ap = "PM" if h >= 12 else "AM"
                            if h > 12: h -= 12
                            if h == 0: h = 12
                            hora = f"{h}:{mn:02d} {ap}"
                        except Exception:
                            pass
                    games.append({
                        "date":      target_dt,
                        "team1":     t1,
                        "team2":     t2,
                        "game_time": hora,
                        "venue":     HOME_VENUES.get(t2, ""),
                        "game":      f"{t1} vs. {t2}",
                        "source":    "ESPN",
                    })
                except Exception:
                    continue
            if games:
                return _dedup_games(games)
        except Exception:
            continue

    return []


def scrape_flashscore_schedule(target_date_str):
    """
    Scrape BSN schedule from flashscore.com.

    Flashscore is a JavaScript-heavy site but its match data is embedded
    in the HTML as a text response from their AJAX endpoints.

    Tries two approaches:
    1. Undocumented Flashscore AJAX API (x-fsign header required)
    2. Plain HTML parse on the BSN league page
    """
    target_dt  = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    games      = []

    # ── Approach 1: Flashscore sports data API ────────────────────────────
    # Flashscore uses a CDN-served binary-ish text protocol, but the league
    # scoreboard endpoint returns parseable text for many leagues.
    date_ts = int(datetime.combine(target_dt, datetime.min.time()).timestamp())

    fs_headers = {
        **HEADERS,
        "x-fsign":    "SW9D1eZo",   # public token used by all flashscore pages
        "Referer":    "https://www.flashscore.com/",
        "Origin":     "https://www.flashscore.com",
        "Accept":     "text/plain",
    }
    # Flashscore tournament IDs for BSN Puerto Rico basketball
    tournament_ids = ["pRGPD5hN", "lfJmMaiC"]  # common IDs used by FS for BSN
    for tid in tournament_ids:
        try:
            url = (f"https://d.flashscore.com/x/feed/f_1_0_{tid}_en_1"
                   f"?_={date_ts}")
            r = requests.get(url, headers=fs_headers, timeout=15)
            if r.status_code == 200 and r.text and len(r.text) > 50:
                # FS text protocol: fields separated by ¬ or ~, records by ÷
                text = r.text
                # Extract matches: look for team names and times
                # Format varies; search for date string or team name fragments
                date_s = target_dt.strftime("%d.%m.%Y")
                if date_s in text or target_date_str in text:
                    # Parse team names from the binary-ish text
                    segs = re.split(r'[÷~]', text)
                    i = 0
                    while i < len(segs) - 1:
                        s = segs[i]
                        t1 = _norm_flashscore(s)
                        t2 = _norm_flashscore(segs[i+1]) if i+1 < len(segs) else None
                        if t1 and t2 and t1 != t2:
                            games.append({
                                "date": target_dt, "team1": t1, "team2": t2,
                                "game_time": "", "venue": HOME_VENUES.get(t2,""),
                                "game": f"{t1} vs. {t2}", "source": "Flashscore",
                            })
                        i += 1
                    if games:
                        return games
        except Exception:
            pass

    # ── Approach 2: Flashscore HTML page scraping ─────────────────────────
    # The HTML page loads data via JS, but the initial HTML sometimes contains
    # event data in <script> blocks or data attributes.
    fs_urls = [
        "https://www.flashscore.com/basketball/puerto-rico/bsn/",
        "https://www.flashscore.es/baloncesto/puerto-rico/bsn/",
        "https://m.flashscore.com/basketball/puerto-rico/bsn/",
    ]
    for url in fs_urls:
        try:
            r = requests.get(url, headers={**HEADERS, "Referer": "https://www.flashscore.com/"}, timeout=20)
            if r.status_code != 200 or len(r.text) < 200:
                continue
            html = r.text

            # Look for JSON-LD or embedded JSON with today's games
            date_s   = target_dt.strftime("%Y-%m-%d")
            date_alt = target_dt.strftime("%d.%m.%Y")

            # Try to find team names adjacent to today's date string
            if date_s in html or date_alt in html:
                # Find all BSN team name mentions in proximity
                all_teams_re = re.compile(
                    r'(' + '|'.join(re.escape(k) for k in FLASHSCORE_NAME_MAP) + r')',
                    re.IGNORECASE
                )
                # Find sections of text near the date
                for m in re.finditer(re.escape(date_s) + r'|' + re.escape(date_alt), html):
                    start = max(0, m.start() - 500)
                    end   = min(len(html), m.end() + 500)
                    snippet = html[start:end]
                    found = all_teams_re.findall(snippet)
                    unique = []
                    for f in found:
                        n = _norm_flashscore(f)
                        if n and n not in unique: unique.append(n)
                    if len(unique) >= 2:
                        for j in range(0, len(unique)-1, 2):
                            t1, t2 = unique[j], unique[j+1]
                            if t1 != t2:
                                games.append({
                                    "date": target_dt, "team1": t1, "team2": t2,
                                    "game_time": "",
                                    "venue": HOME_VENUES.get(t2,""),
                                    "game": f"{t1} vs. {t2}",
                                    "source": "Flashscore",
                                })
            if games:
                # Dedup
                seen = set()
                out  = []
                for g in games:
                    key = tuple(sorted([g["team1"], g["team2"]]))
                    if key not in seen:
                        seen.add(key); out.append(g)
                return out
        except Exception:
            pass

    return []


# ── Manual --add-game command ──────────────────────────────────────────────

MANUAL_GAMES_FILE  = os.path.join(SCRIPT_DIR, "manual_games.json")
MANUAL_LINES_FILE  = os.path.join(SCRIPT_DIR, "bsn_market_lines.json")


# ──────────────────────────────────────────────────────
# MANUAL MARKET LINES — entrada de odds del libro
# ──────────────────────────────────────────────────────

def _load_manual_lines():
    """Carga líneas de mercado guardadas manualmente (bsn_market_lines.json)."""
    if not os.path.exists(MANUAL_LINES_FILE):
        return {}
    try:
        with open(MANUAL_LINES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_manual_lines(data):
    with open(MANUAL_LINES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _get_market_line(team1, team2, date_str=None):
    """Busca línea de mercado para un juego específico. Retorna dict o None."""
    ds = date_str or TARGET_DATE
    all_lines = _load_manual_lines()
    for entry in all_lines.get(ds, []):
        pair_saved = {entry.get("team1","").upper(), entry.get("team2","").upper()}
        pair_query = {team1.upper(), team2.upper()}
        if pair_saved == pair_query:
            return entry
    return None


def _american_to_prob(odds_str):
    """Convierte odds americanas ('+150', '-170') a probabilidad implícita (0-100)."""
    try:
        o = int(str(odds_str).replace("+","").strip())
        if o > 0:
            return 100.0 / (o + 100.0) * 100.0
        else:
            return abs(o) / (abs(o) + 100.0) * 100.0
    except Exception:
        return None


def cmd_set_lines_bsn():
    """
    --set-lines [DATE]
    Entrada interactiva de las líneas del mercado (ML / Spread / Total) para los juegos del día.
    Los datos se guardan en bsn_market_lines.json y --picks los usa para calcular edge real.
    Uso:
      python3 bsn.py --set-lines             (usa TARGET_DATE)
      python3 bsn.py --set-lines 2026-04-14
    """
    try:
        si = sys.argv.index("--set-lines")
        sl_date = (sys.argv[si+1]
                   if si+1 < len(sys.argv) and not sys.argv[si+1].startswith("-")
                   else TARGET_DATE)
    except (ValueError, IndexError):
        sl_date = TARGET_DATE

    print(f"\n  📝 LÍNEAS DE MERCADO BSN — {sl_date}")
    print(f"  {'─'*54}")
    print("  Ingresa las odds del libro para cada juego.")
    print("  Presiona ENTER para omitir un campo (deja en blanco).")
    print(f"  {'─'*54}\n")

    games = _get_manual_games(sl_date)
    if not games:
        print("  ⚠️  No se encontraron juegos para esta fecha.")
        print("  Agrega los juegos primero:  python3 bsn.py --add-game TEAM1 TEAM2 '8:00 PM'\n")
        return

    all_lines  = _load_manual_lines()
    # Partir de las líneas ya guardadas para esta fecha (si existen)
    saved      = {(e["team1"].upper(), e["team2"].upper()): e
                  for e in all_lines.get(sl_date, [])}
    date_lines = []

    for g in games:
        t1 = g["team1"]; t2 = g["team2"]
        key = (t1.upper(), t2.upper())
        existing = saved.get(key)

        print(f"\n  🏀  {t1}  @  {t2}")
        print(f"  {'─'*40}")

        # Si ya hay líneas guardadas, mostrarlas y preguntar si actualizar
        if existing:
            parts = []
            if existing.get("ml1") or existing.get("ml2"):
                parts.append(f"ML {existing.get('ml1','—')}/{existing.get('ml2','—')}")
            if existing.get("spread_fav"):
                parts.append(f"Spread {existing['spread_fav']} {existing.get('spread_line','')} "
                              f"(fav:{existing.get('spread_odds','')} dog:{existing.get('spread_dog_odds','')})")
            if existing.get("total"):
                parts.append(f"Total {existing['total']} "
                              f"(O:{existing.get('over_odds','')} U:{existing.get('under_odds','')})")
            print(f"  ✅ Ya guardado: {' | '.join(parts)}")
            update = input(f"  ¿Actualizar este juego? (S = sí / ENTER = mantener):  ").strip().upper()
            if update != "S":
                date_lines.append(existing)
                print(f"  ↩️  Mantenido sin cambios.")
                continue

        ml1  = input(f"  ML {t1:<16} (ej: +150, -120 o ENTER):  ").strip()
        ml2  = input(f"  ML {t2:<16} (ej: +150, -120 o ENTER):  ").strip()

        print(f"  Spread favorito  (ej: {t2}  ó  ENTER para omitir):")
        sp_fav  = input(f"    Equipo favorito:                          ").strip()
        sp_line = input(f"    Línea            (ej: -4.5 o ENTER):      ").strip()
        sp_odds     = input(f"    Odds spread fav  (ej: -110, ENTER=-110):  ").strip() or ("-110" if sp_line else "")
        sp_dog_odds = input(f"    Odds spread dog  (ej: -110, ENTER=-110):  ").strip() or ("-110" if sp_line else "")

        print(f"  Total:")
        tot_line       = input(f"    Línea total      (ej: 165.5 o ENTER):    ").strip()
        tot_over_odds  = input(f"    Over odds        (ej: -110, ENTER=-110): ").strip() or ("-110" if tot_line else "")
        tot_under_odds = input(f"    Under odds       (ej: -110, ENTER=-110): ").strip() or ("-110" if tot_line else "")

        entry = {"team1": t1, "team2": t2}
        if ml1:            entry["ml1"]             = ml1
        if ml2:            entry["ml2"]             = ml2
        if sp_fav:         entry["spread_fav"]      = sp_fav
        if sp_line:        entry["spread_line"]     = sp_line
        if sp_odds:        entry["spread_odds"]     = sp_odds
        if sp_dog_odds:    entry["spread_dog_odds"] = sp_dog_odds
        if tot_line:       entry["total"]           = tot_line
        if tot_over_odds:  entry["over_odds"]       = tot_over_odds
        if tot_under_odds: entry["under_odds"]      = tot_under_odds

        date_lines.append(entry)
        print(f"  ✅ Guardado: {t1} @ {t2}")

    all_lines[sl_date] = date_lines
    _save_manual_lines(all_lines)

    print(f"\n  ✅ Líneas guardadas para {sl_date} ({len(date_lines)} juego(s)).")
    print(f"  Ejecuta:  python3 bsn.py --picks   para ver picks con edge del mercado.\n")


def cmd_clear_lines():
    """
    --clear-lines [DATE]
    Muestra los juegos con líneas guardadas y deja escoger cuál(es) borrar.
    Sin fecha → usa TARGET_DATE.
    Con --all → borra todos los juegos de esa fecha sin preguntar.

    Ejemplos:
      python3 bsn.py --clear-lines
      python3 bsn.py --clear-lines 2026-04-13
      python3 bsn.py --clear-lines --all
    """
    clear_all = "--all" in sys.argv

    # Determinar fecha
    cl_date = TARGET_DATE
    if "--clear-lines" in sys.argv:
        ci = sys.argv.index("--clear-lines")
        if ci + 1 < len(sys.argv) and not sys.argv[ci + 1].startswith("--"):
            cl_date = sys.argv[ci + 1]

    all_lines = _load_manual_lines()

    if cl_date not in all_lines or not all_lines[cl_date]:
        print(f"\n  ⚠️  No hay líneas guardadas para {cl_date}.\n")
        return

    games = all_lines[cl_date]

    if clear_all:
        del all_lines[cl_date]
        _save_manual_lines(all_lines)
        print(f"\n  🗑️  Todas las líneas de {cl_date} eliminadas ({len(games)} juego(s)).")
        print(f"  Ejecuta:  python3 bsn.py --set-lines   para ingresar nuevas líneas.\n")
        return

    # Mostrar lista de juegos
    print(f"\n  🗑️  Líneas guardadas para {cl_date}:\n")
    for i, g in enumerate(games, 1):
        t1 = g.get("team1","?"); t2 = g.get("team2","?")
        parts = []
        if g.get("ml1") or g.get("ml2"):
            parts.append(f"ML {g.get('ml1','—')}/{g.get('ml2','—')}")
        if g.get("spread_fav"):
            parts.append(f"Spread {g['spread_fav']} {g.get('spread_line','')} ({g.get('spread_odds','')})")
        if g.get("total"):
            parts.append(f"Total {g['total']}")
        detail = "  |  ".join(parts) if parts else "sin detalle"
        print(f"  [{i}] {t1} @ {t2}  —  {detail}")

    print(f"\n  Escribe el número del juego a borrar (o varios separados por coma, ej: 1,3)")
    print(f"  ENTER sin número = cancelar")
    sel = input("  > ").strip()

    if not sel:
        print("  Cancelado.\n")
        return

    # Parsear selección
    indices = set()
    for s in sel.split(","):
        s = s.strip()
        if s.isdigit():
            idx = int(s) - 1
            if 0 <= idx < len(games):
                indices.add(idx)
            else:
                print(f"  ⚠️  Número {s} fuera de rango, ignorado.")

    if not indices:
        print("  No se seleccionó ningún juego válido.\n")
        return

    removed = [games[i] for i in sorted(indices)]
    all_lines[cl_date] = [g for i, g in enumerate(games) if i not in indices]
    if not all_lines[cl_date]:
        del all_lines[cl_date]
    _save_manual_lines(all_lines)

    for g in removed:
        print(f"  🗑️  Eliminado: {g.get('team1','?')} @ {g.get('team2','?')}")
    print(f"\n  Ejecuta:  python3 bsn.py --set-lines   para ingresar nuevas líneas.\n")


def _load_manual_games():
    """Load manually added games from JSON file."""
    if not os.path.exists(MANUAL_GAMES_FILE):
        return []
    try:
        with open(MANUAL_GAMES_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return []


def _save_manual_games(entries):
    with open(MANUAL_GAMES_FILE, "w") as f:
        json.dump(entries, f, indent=2, default=str)


def cmd_add_game():
    """
    --add-game VISITANTE LOCAL HORA
    Agrega un juego manual para TARGET_DATE (hoy por defecto).
    El juego queda guardado en manual_games.json y se usa si la web no encuentra datos.
    """
    idx = sys.argv.index("--add-game")
    try:
        raw_t1   = sys.argv[idx + 1]
        raw_t2   = sys.argv[idx + 2]
        raw_hora = sys.argv[idx + 3] if idx + 3 < len(sys.argv) else ""
    except IndexError:
        print("  ❌ Uso: python3 bsn.py --add-game VISITANTE LOCAL HORA")
        print("  Ejemplo: python3 bsn.py --add-game SANTEROS LEONES '8:00 PM'")
        sys.exit(1)

    t1 = norm_team(raw_t1.upper())
    t2 = norm_team(raw_t2.upper())

    if not t1 or t1 not in BSN_TEAMS:
        # Try partial match
        matches = [t for t in BSN_TEAMS if raw_t1.upper() in t]
        if matches:
            t1 = matches[0]
        else:
            print(f"  ❌ Equipo visitante '{raw_t1}' no reconocido.")
            print(f"  Equipos válidos: {', '.join(BSN_TEAMS)}")
            sys.exit(1)

    if not t2 or t2 not in BSN_TEAMS:
        matches = [t for t in BSN_TEAMS if raw_t2.upper() in t]
        if matches:
            t2 = matches[0]
        else:
            print(f"  ❌ Equipo local '{raw_t2}' no reconocido.")
            print(f"  Equipos válidos: {', '.join(BSN_TEAMS)}")
            sys.exit(1)

    if t1 == t2:
        print(f"  ❌ Visitante y local no pueden ser el mismo equipo.")
        sys.exit(1)

    # Normalize time string
    hora = raw_hora.strip().upper()
    if hora and not re.search(r'AM|PM', hora, re.I):
        hora += " PM"   # default to PM if no meridiem

    entry = {
        "date":      TARGET_DATE,
        "team1":     t1,
        "team2":     t2,
        "game_time": hora,
        "venue":     HOME_VENUES.get(t2, ""),
        "game":      f"{t1} vs. {t2}",
        "source":    "Manual",
    }

    existing = _load_manual_games()
    # Check if already exists
    key_new = tuple(sorted([t1, t2]))
    duplicate = any(
        tuple(sorted([e["team1"], e["team2"]])) == key_new and e["date"] == TARGET_DATE
        for e in existing
    )
    if duplicate:
        print(f"  ⚠️  Juego {t1} vs. {t2} ya existe para {TARGET_DATE}.")
        print(f"      Nada que agregar.")
    else:
        existing.append(entry)
        _save_manual_games(existing)
        print(f"\n  ✅ Juego agregado para {TARGET_DATE}:")
        print(f"     {t1} (visit.) @ {t2} (local)  ·  {hora or 'hora TBD'}")
        print(f"\n  Ahora corre: python3 bsn.py")
        print(f"  (El juego manual se usará si la web no devuelve resultados)\n")


def cmd_list_games():
    """
    --list-games [FECHA]
    Muestra todos los juegos manuales guardados (de TARGET_DATE por defecto,
    o todos si se pasa 'all').
    """
    entries = _load_manual_games()
    show_all = "all" in [a.lower() for a in sys.argv]

    if show_all:
        filtered = entries
        lbl = "todos los juegos manuales"
    else:
        filtered = [e for e in entries if e.get("date") == TARGET_DATE]
        lbl = f"juegos manuales para {TARGET_DATE}"

    print(f"\n  📋 {lbl.upper()}")
    print(f"  {'─'*52}")

    if not filtered:
        print(f"  (Ninguno)\n")
        print(f"  Tip: python3 bsn.py --add-game VISITANTE LOCAL HORA\n")
        return

    for i, e in enumerate(filtered):
        t1   = e.get("team1","?"); t2 = e.get("team2","?")
        hora = e.get("game_time","—") or "—"
        dt   = e.get("date","?")
        print(f"  [{i}]  {dt}  │  {t1:<14} @ {t2:<14}  │  {hora}")

    print(f"\n  Para borrar: python3 bsn.py --remove-game VISITANTE LOCAL [FECHA]\n")


def cmd_remove_game():
    """
    --remove-game VISITANTE LOCAL [FECHA]
    Elimina un juego manual. Si FECHA se omite usa TARGET_DATE.
    Ejemplos:
      python3 bsn.py --remove-game SANTEROS LEONES
      python3 bsn.py --remove-game SANTEROS LEONES 2026-04-11
    """
    idx = sys.argv.index("--remove-game")
    try:
        raw_t1 = sys.argv[idx + 1]
        raw_t2 = sys.argv[idx + 2]
    except IndexError:
        print("  ❌ Uso: python3 bsn.py --remove-game VISITANTE LOCAL [FECHA]")
        sys.exit(1)

    # Fecha opcional como tercer arg (no empieza con --)
    try:
        maybe_date = sys.argv[idx + 3]
        if re.match(r'\d{4}-\d{2}-\d{2}', maybe_date):
            target = maybe_date
        else:
            target = TARGET_DATE
    except IndexError:
        target = TARGET_DATE

    t1 = norm_team(raw_t1.upper())
    t2 = norm_team(raw_t2.upper())

    # Partial match si no reconoce
    if not t1 or t1 not in BSN_TEAMS:
        matches = [t for t in BSN_TEAMS if raw_t1.upper() in t]
        t1 = matches[0] if matches else raw_t1.upper()
    if not t2 or t2 not in BSN_TEAMS:
        matches = [t for t in BSN_TEAMS if raw_t2.upper() in t]
        t2 = matches[0] if matches else raw_t2.upper()

    key = tuple(sorted([t1, t2]))
    entries  = _load_manual_games()
    original = len(entries)
    kept     = [e for e in entries
                if not (e.get("date") == target
                        and tuple(sorted([e["team1"], e["team2"]])) == key)]

    removed = original - len(kept)
    if removed == 0:
        print(f"\n  ⚠️  No se encontró {t1} vs. {t2} para {target}.")
        # Muestra qué hay guardado
        today_games = [e for e in entries if e.get("date") == target]
        if today_games:
            print(f"  Juegos guardados para {target}:")
            for e in today_games:
                print(f"    • {e['team1']} @ {e['team2']}  {e.get('game_time','')}")
        else:
            print(f"  No hay juegos manuales para {target}.")
        print()
    else:
        _save_manual_games(kept)
        print(f"\n  ✅ Eliminado: {t1} @ {t2}  ({target})")
        remaining = [e for e in kept if e.get("date") == target]
        if remaining:
            print(f"  Quedan {len(remaining)} juego(s) para {target}:")
            for e in remaining:
                print(f"    • {e['team1']} @ {e['team2']}  {e.get('game_time','')}")
        else:
            print(f"  No quedan juegos manuales para {target}.")
        print()


def cmd_edit_game():
    """
    --edit-game VISITANTE LOCAL campo valor [FECHA]
    Edita un campo de un juego manual (game_time, date, venue).
    Si FECHA se omite usa TARGET_DATE (hoy).

    Ejemplos:
      python3 bsn.py --edit-game SANTEROS LEONES game_time "9:00 PM"
      python3 bsn.py --edit-game SANTEROS LEONES date 2026-04-20
      python3 bsn.py --edit-game SANTEROS LEONES game_time "8:30 PM" 2026-04-18
    """
    EDITABLE = {"game_time", "date", "venue"}

    try:
        ei     = sys.argv.index("--edit-game")
        raw_t1 = sys.argv[ei + 1]
        raw_t2 = sys.argv[ei + 2]
        field  = sys.argv[ei + 3].lower()
        value  = sys.argv[ei + 4]
    except IndexError:
        print("  ❌ Uso: python3 bsn.py --edit-game VISITANTE LOCAL campo valor [FECHA]")
        print("  Campos: game_time, date, venue")
        print("  Ejemplo: python3 bsn.py --edit-game SANTEROS LEONES game_time '9:00 PM'")
        return

    if field not in EDITABLE:
        print(f"  ❌ Campo '{field}' no válido. Usa: {', '.join(sorted(EDITABLE))}")
        return

    # Fecha opcional como sexto arg
    try:
        maybe = sys.argv[ei + 5]
        target = maybe if re.match(r'\d{4}-\d{2}-\d{2}', maybe) else TARGET_DATE
    except IndexError:
        target = TARGET_DATE

    t1 = norm_team(raw_t1.upper())
    t2 = norm_team(raw_t2.upper())
    if not t1 or t1 not in BSN_TEAMS:
        matches = [t for t in BSN_TEAMS if raw_t1.upper() in t]
        t1 = matches[0] if matches else raw_t1.upper()
    if not t2 or t2 not in BSN_TEAMS:
        matches = [t for t in BSN_TEAMS if raw_t2.upper() in t]
        t2 = matches[0] if matches else raw_t2.upper()

    key     = tuple(sorted([t1, t2]))
    entries = _load_manual_games()
    found   = False

    for e in entries:
        if (e.get("date") == target and
                tuple(sorted([e["team1"], e["team2"]])) == key):
            old_val   = e.get(field)
            e[field]  = value
            # Keep "game" field in sync if date changes
            if field == "date":
                e["date"] = value
            found = True
            print(f"\n  ✅ Juego actualizado ({target}):")
            print(f"     {t1} @ {t2}")
            print(f"     {field}: {repr(old_val)}  →  {repr(value)}\n")
            break

    if not found:
        print(f"\n  ⚠️  No se encontró {t1} vs. {t2} para {target}.")
        today_games = [e for e in entries if e.get("date") == target]
        if today_games:
            print(f"  Juegos manuales para {target}:")
            for e in today_games:
                print(f"    • {e['team1']} @ {e['team2']}  {e.get('game_time','')}")
        else:
            print(f"  No hay juegos manuales para {target}.")
        print()
        return

    _save_manual_games(entries)


def _get_manual_games(target_date_str):
    """Returns manual games for a given date."""
    entries = _load_manual_games()
    result  = []
    for e in entries:
        if e.get("date") == target_date_str:
            result.append({
                "date":      datetime.strptime(target_date_str, "%Y-%m-%d").date(),
                "team1":     e["team1"],
                "team2":     e["team2"],
                "game_time": e.get("game_time",""),
                "venue":     e.get("venue",""),
                "game":      e.get("game",""),
                "source":    "Manual",
            })
    return result


# ── basketball24.com scraper ───────────────────────────────────────────────

# Basketball24 team name fragments → BSN short names
# (same family as Flashscore — identical HTML structure)
B24_NAME_MAP = {
    **FLASHSCORE_NAME_MAP,   # inherit all flashscore mappings
    # basketball24 may use city names instead of team names
    "humacao":      "INDIOS",
    "aguada":       "VAQUEROS",
    "san juan":     "CANGREJEROS",
    "cabo rojo":    "GIGANTES",
    "fajardo":      "CAPITANES",
    "aibonito":     "PIRATAS",
    "salinas":      "ATLETICOS",
    "guayama":      "CRIOLLOS",
    "juana diaz":   "TOROS",
}

def _norm_b24(name):
    if not name: return None
    low = name.lower().strip()
    for k, v in B24_NAME_MAP.items():
        if k in low: return v
    for t in BSN_TEAMS:
        if t.lower() in low: return t
    return None

def scrape_betsapi_schedule(target_date_str):
    """
    Scrape BSN fixtures from betsapi.com/basketball/ls/4479/puerto-rico-superior-nacional
    League ID 4479 = Puerto Rico Superior Nacional (BSN)

    Tries:
    1. Direct HTML page — parses tables and embedded JSON in <script> tags
    2. Internal AJAX endpoints betsapi uses to populate the page
    """
    target_dt  = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    games      = []
    clean      = lambda s: re.sub(r'<[^>]+>', '', s).strip()

    BETS_HEADERS = {
        **HEADERS,
        "Referer":  "https://betsapi.com/",
        "Accept":   "text/html,application/xhtml+xml,application/json,*/*;q=0.9",
    }

    def _norm_bets(name):
        if not name: return None
        n = name.strip().upper()
        if n in BSN_TEAMS: return n
        low = name.strip().lower()
        for key, val in REALGM_NAME_MAP.items():
            if key in low or low in key: return val
        for key, val in BSNPR_NAME_MAP.items():
            if key in low or low in key: return val
        for t in BSN_TEAMS:
            if t.lower() in low: return t
        return None

    def _parse_bets_games(text):
        """Extract games from HTML or JSON text for target_dt."""
        found_games = []
        date_s = target_dt.strftime("%Y-%m-%d")
        # Date formats betsapi may use in the page
        date_alts = [
            target_dt.strftime("%d/%m/%Y"),
            target_dt.strftime("%d.%m.%Y"),
            target_dt.strftime("%b %d, %Y"),
            target_dt.strftime("%B %d, %Y"),
        ]

        # ── JSON in script tags ───────────────────────────────────────────
        for script in re.finditer(r'<script[^>]*>(.*?)</script>', text, re.DOTALL|re.IGNORECASE):
            blob = script.group(1).strip()
            if len(blob) < 20: continue
            # Look for JSON arrays/objects with event data
            for json_m in re.finditer(r'(\{[^{}]{50,}\}|\[[^\[\]]{50,}\])', blob):
                try:
                    obj = json.loads(json_m.group(0))
                    raw = json.dumps(obj).lower()
                    # Must mention the date and at least two BSN teams
                    if date_s not in raw and not any(a in raw for a in date_alts): continue
                    # Try to extract events from the JSON
                    events = []
                    if isinstance(obj, list):
                        events = obj
                    elif isinstance(obj, dict):
                        for key in ["events","matches","games","fixtures","data","items","results"]:
                            if key in obj and isinstance(obj[key], list):
                                events = obj[key]; break
                    for ev in events:
                        if not isinstance(ev, dict): continue
                        ev_raw = json.dumps(ev).lower()
                        if date_s not in ev_raw and not any(a.lower() in ev_raw for a in date_alts):
                            continue
                        # Try known field names for home/away teams
                        t1_raw = (ev.get("away") or ev.get("awayTeam") or ev.get("away_team") or
                                  ev.get("team1") or ev.get("visitor") or "")
                        t2_raw = (ev.get("home") or ev.get("homeTeam") or ev.get("home_team") or
                                  ev.get("team2") or ev.get("local") or "")
                        if isinstance(t1_raw, dict): t1_raw = t1_raw.get("name","")
                        if isinstance(t2_raw, dict): t2_raw = t2_raw.get("name","")
                        t1 = _norm_bets(str(t1_raw))
                        t2 = _norm_bets(str(t2_raw))
                        if not (t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS):
                            continue
                        # Time
                        time_str = ""
                        for tk in ["time","start_time","hora","startTime","kickoff","kickoff_time"]:
                            if ev.get(tk):
                                tm = re.search(r'(\d{1,2}:\d{2}\s*(?:AM|PM)?)', str(ev[tk]), re.I)
                                if tm: time_str = tm.group(1).upper(); break
                        key = tuple(sorted([t1, t2]))
                        if not any(tuple(sorted([g["team1"],g["team2"]]))==key for g in found_games):
                            found_games.append({
                                "date": target_dt, "team1": t1, "team2": t2,
                                "game_time": time_str, "venue": HOME_VENUES.get(t2,""),
                                "game": f"{t1} vs. {t2}", "source": "betsapi",
                            })
                except Exception:
                    continue

        if found_games: return found_games

        # ── HTML table rows ───────────────────────────────────────────────
        # betsapi shows a table with columns like: Date | Home | Away | Time | Result
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL|re.IGNORECASE)
        for row_html in rows:
            cells_raw = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL|re.IGNORECASE)
            if len(cells_raw) < 2: continue
            cell_vals  = [clean(c) for c in cells_raw]
            raw_row    = " ".join(cell_vals).lower()
            # Check date is somewhere in the row
            if date_s not in raw_row and not any(a.lower() in raw_row for a in date_alts):
                continue
            # Extract teams via link text (betsapi wraps team names in <a> tags)
            link_texts = re.findall(r'<a[^>]+href=["\'][^"\']*(?:/t/|/team/)[^"\']*["\'][^>]*>(.*?)</a>',
                                    row_html, re.DOTALL|re.IGNORECASE)
            teams_in_row = []
            for lt in link_texts:
                nt = _norm_bets(clean(lt))
                if nt and nt not in teams_in_row: teams_in_row.append(nt)
            # Fallback: search all cell text
            if len(teams_in_row) < 2:
                for cv in cell_vals:
                    nt = _norm_bets(cv)
                    if nt and nt not in teams_in_row: teams_in_row.append(nt)
            if len(teams_in_row) < 2: continue
            t1, t2 = teams_in_row[0], teams_in_row[1]
            if not (t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS): continue
            time_match = re.search(r'(\d{1,2}:\d{2}\s*(?:AM|PM|am|pm))', " ".join(cell_vals))
            game_time  = time_match.group(1).upper() if time_match else ""
            key = tuple(sorted([t1, t2]))
            if not any(tuple(sorted([g["team1"],g["team2"]]))==key for g in found_games):
                found_games.append({
                    "date": target_dt, "team1": t1, "team2": t2,
                    "game_time": game_time, "venue": HOME_VENUES.get(t2,""),
                    "game": f"{t1} vs. {t2}", "source": "betsapi",
                })

        # ── Team anchor links anywhere on the page ────────────────────────
        if not found_games:
            # betsapi lists matchups as "TeamA vs TeamB" in page text near a date
            for date_pat in [date_s] + date_alts:
                if not date_pat or date_pat.lower() not in text.lower(): continue
                # Find the section near the date
                idx = text.lower().find(date_pat.lower())
                snippet = text[max(0,idx-50):min(len(text),idx+2000)]
                # Look for "vs" pattern with team names from anchor links
                for m in re.finditer(
                    r'([A-Za-záéíóúñÁÉÍÓÚÑ ]{4,40})\s*(?:vs\.?|@|-)\s*([A-Za-záéíóúñÁÉÍÓÚÑ ]{4,40})',
                    snippet):
                    t1 = _norm_bets(m.group(1).strip())
                    t2 = _norm_bets(m.group(2).strip())
                    if t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS:
                        key = tuple(sorted([t1, t2]))
                        if not any(tuple(sorted([g["team1"],g["team2"]]))==key for g in found_games):
                            found_games.append({
                                "date": target_dt, "team1": t1, "team2": t2,
                                "game_time": "", "venue": HOME_VENUES.get(t2,""),
                                "game": f"{t1} vs. {t2}", "source": "betsapi",
                            })
        return found_games

    # ── Fetch the main fixtures page ──────────────────────────────────────
    urls = [
        "https://betsapi.com/basketball/ls/4479/puerto-rico-superior-nacional",
        "https://betsapi.com/l/4479/Puerto-Rico-Superior-Nacional",
        f"https://betsapi.com/basketball/ls/4479/puerto-rico-superior-nacional?date={target_date_str}",
    ]
    for url in urls:
        try:
            r = requests.get(url, headers=BETS_HEADERS, timeout=20)
            if r.status_code == 200 and len(r.text) > 200:
                result = _parse_bets_games(r.text)
                if result:
                    return result
        except Exception:
            pass

    # ── Try betsapi internal AJAX endpoint ────────────────────────────────
    # betsapi uses sport_id=18 for basketball
    ajax_urls = [
        f"https://betsapi.com/api/v1/events/upcoming?sport_id=18&league_id=4479&day={target_date_str.replace('-','')}",
        f"https://api.betsapi.com/v1/events/upcoming?sport_id=18&league_id=4479&day={target_date_str.replace('-','')}",
        f"https://betsapi.com/api/v2/events/upcoming?sport_id=18&league_id=4479",
    ]
    for url in ajax_urls:
        try:
            r = requests.get(url, headers={**BETS_HEADERS, "Accept": "application/json",
                                           "X-Requested-With": "XMLHttpRequest"}, timeout=15)
            if r.status_code == 200:
                try:
                    data = r.json()
                    result = _parse_bets_games(json.dumps(data))
                    if result:
                        return result
                except Exception:
                    pass
        except Exception:
            pass

    return games


def scrape_basketball24_schedule(target_date_str):
    """
    Scrape BSN fixtures from basketball24.com.

    Strategy:
    1. Fetch the 262KB page HTML (confirmed working — status 200)
    2. Extract the real-time server number and tournament ID from the HTML
    3. Call the flashscore.ninja data feed directly with those IDs
    4. Fall back to aggressive HTML parsing if feed fails
    """
    target_dt = datetime.strptime(target_date_str, "%Y-%m-%d").date()

    page_headers = {
        **HEADERS,
        "Referer":         "https://www.basketball24.com/",
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
    }

    # ── Step 1: Fetch the page (confirmed 200 + 262KB) ───────────────────
    html = None
    for url in [
        "https://www.basketball24.com/puerto-rico/bsn/fixtures/",
        "https://www.basketball24.com/puerto-rico/bsn/",
    ]:
        try:
            r = requests.get(url, headers=page_headers, timeout=25)
            if r.status_code == 200 and len(r.text) > 1000:
                html = r.text
                break
        except Exception:
            pass

    if not html:
        return []

    # ── Step 2: Extract server number from HTML ───────────────────────────
    # The HTML has: <link rel="preconnect" href="https://110.flashscore.ninja">
    server_num = "110"   # confirmed from debug output
    m = re.search(r'https://(\d+)\.flashscore\.ninja', html)
    if m:
        server_num = m.group(1)

    # ── Step 3: Extract tournament/unique-tournament ID from HTML ─────────
    # Flashscore embeds the tournament ID in several places.
    # Real IDs are 6-12 chars and ALWAYS contain at least one digit
    # (e.g. VmrBBJo2, pRGPD5hN) — pure-lowercase words are navigation slugs.
    _NAV_SLUGS = {"standings","archive","fixtures","results","schedule",
                  "live","table","news","odds","draw","transfers","squads"}
    tid_candidates = set()

    # Pattern A: URL slug — must contain at least one digit (filters nav words)
    for m in re.finditer(r'/puerto-rico/bsn/([A-Za-z0-9]{6,12})/', html):
        cand = m.group(1)
        if any(c.isdigit() for c in cand) and cand not in _NAV_SLUGS:
            tid_candidates.add(cand)

    # Pattern B: uniqueTournamentId in JSON/JS
    for m in re.finditer(r'"uniqueTournamentId"\s*:\s*"?(\w+)"?', html):
        tid_candidates.add(m.group(1))

    # Pattern C: tournament ID in data attributes
    for m in re.finditer(r'data-(?:id|tournament-id|tid)="([A-Za-z0-9]{4,12})"', html):
        cand = m.group(1)
        if any(c.isdigit() for c in cand):
            tid_candidates.add(cand)

    # Pattern D: numeric tournamentId / uniqueTournamentId
    for m in re.finditer(r'"(?:unique)?[Tt]ournament[Ii]d"\s*:\s*(\d+)', html):
        tid_candidates.add(m.group(1))

    # Pattern E: window.__ or similar JS vars holding the tournament config
    for m in re.finditer(r'["\']tournamentId["\']\s*:\s*["\']?([A-Za-z0-9]{4,12})["\']?', html):
        cand = m.group(1)
        if any(c.isdigit() for c in cand):
            tid_candidates.add(cand)

    # Remove obvious non-IDs
    tid_candidates = {t for t in tid_candidates
                      if len(t) >= 4 and t not in {"true","false","null","undefined"}}

    # ── Step 4: Try flashscore.ninja data feed ────────────────────────────
    feed_headers = {
        **HEADERS,
        "x-fsign":         "SW9D1eZo",
        "Referer":         "https://www.basketball24.com/",
        "Origin":          "https://www.basketball24.com",
        "Accept":          "text/plain, */*",
    }

    date_ts = int(datetime.combine(target_dt, datetime.min.time()).timestamp())

    # Extracted IDs first, then a broader set of known/candidate BSN IDs
    # (flashscore IDs change season-to-season — extend this list when found)
    all_tids = list(tid_candidates) + [
        "YlkXoSjC", "8OyKJY76", "3ZCMZMnC", "gm5QKHB9",
        "UpxqxJnL", "xE2ROAP6", "fSTUBkKp", "WtBxFkjB",
    ]

    for tid in all_tids:
        for feed_type in ["f_1_0", "f_2_0", "f_1_1"]:
            feed_url = (f"https://{server_num}.flashscore.ninja/x/feed/"
                        f"{feed_type}_{tid}_en_1")
            try:
                r = requests.get(feed_url, headers=feed_headers, timeout=12)
                if r.status_code == 200 and len(r.text) > 80:
                    g = _parse_b24_feed(r.text, target_dt)
                    if g:
                        return g
            except Exception:
                pass

    # ── Step 5: Also try d.basketball24.com ──────────────────────────────
    for tid in all_tids:
        try:
            url = (f"https://d.basketball24.com/x/feed/f_1_0_{tid}_en_1"
                   f"?_{date_ts}")
            r = requests.get(url, headers={**feed_headers,
                                           "Host": "d.basketball24.com"},
                             timeout=12)
            if r.status_code == 200 and len(r.text) > 80:
                g = _parse_b24_feed(r.text, target_dt)
                if g:
                    return g
        except Exception:
            pass

    # ── Step 6: Aggressive HTML parse (last resort) ───────────────────────
    # The 262KB HTML may have some SSR data — try every known pattern
    return _parse_b24_html(html, target_dt)


def _b24_extract_date_section(html, target_dt):
    """
    Basketball24 / flashscore HTML has date section headers like:
      "Friday, 10 Apr 2026" or "10.04.2026"

    Returns the slice of HTML that corresponds to target_dt's games block,
    or None if the date is not found in the HTML at all.
    """
    # Build multiple date string candidates
    day_names = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    date_strs = [
        target_dt.strftime("%A, %d %b %Y"),    # "Friday, 10 Apr 2026"
        target_dt.strftime("%A, %d. %b %Y"),   # "Friday, 10. Apr 2026"
        target_dt.strftime("%d.%m.%Y"),         # "10.04.2026"
        target_dt.strftime("%Y-%m-%d"),         # "2026-04-10"
        target_dt.strftime("%d/%m/%Y"),         # "10/04/2026"
        target_dt.strftime("%-d %b %Y"),        # "10 Apr 2026"  (Linux)
    ]

    html_lower = html.lower()
    start_pos = -1
    for ds in date_strs:
        idx = html_lower.find(ds.lower())
        if idx != -1:
            start_pos = idx
            break

    if start_pos == -1:
        return None   # target date not found at all

    # Find the next date header after start_pos (so we don't bleed into tomorrow's games)
    # Flashscore uses day names in titles — next occurrence of a weekday that is NOT
    # within 30 chars of our current marker is the next section.
    end_pos = len(html)
    for dn in day_names:
        search_from = start_pos + 20  # skip past the current date string itself
        idx = html_lower.find(dn.lower(), search_from)
        if idx != -1 and idx < end_pos:
            end_pos = idx

    return html[start_pos:end_pos]


def _parse_b24_html(html, target_dt):
    """
    Parse basketball24.com / flashscore-family HTML.

    Strategy 1: CSS class-based parsing within today's date section.
      <div class="event__match">
        <div class="event__time">20:00</div>
        <div class="event__participant event__participant--away">Santeros</div>
        <div class="event__participant event__participant--home">Leones</div>
      </div>

    Date headers:
      <div class="event__header">
        <div class="event__title">Thursday, 10 Apr 2026</div>
      </div>

    NOTE: Strategy 2 (broad regex) has been DISABLED — it matches team names
    scattered throughout navigation, standings, and full-season data in the 262KB
    HTML, returning 53+ garbage results with wrong venues and no date context.
    """
    games = []

    # ── Isolate today's date section first ───────────────────────────────
    section = _b24_extract_date_section(html, target_dt)
    # If we can isolate today's section, search within it; else search full HTML
    search_html = section if section else html
    found_date_section = section is not None

    # ── Strategy 1: CSS class-based parsing ──────────────────────────────
    away_pat = re.compile(
        r'event__participant--away[^>]*>([^<]+)<', re.IGNORECASE)
    home_pat = re.compile(
        r'event__participant--home[^>]*>([^<]+)<', re.IGNORECASE)
    time_pat = re.compile(
        r'event__time[^>]*>([^<]+)<', re.IGNORECASE)

    away_teams = [m.group(1).strip() for m in away_pat.finditer(search_html)]
    home_teams = [m.group(1).strip() for m in home_pat.finditer(search_html)]
    times      = [m.group(1).strip() for m in time_pat.finditer(search_html)]

    paired = list(zip(away_teams, home_teams))
    for i, (raw_t1, raw_t2) in enumerate(paired):
        t1 = _norm_b24(raw_t1)
        t2 = _norm_b24(raw_t2)
        if t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS:
            hora = times[i] if i < len(times) else ""
            hora = _convert_24h(hora)
            games.append({
                "date":      target_dt,
                "team1":     t1,
                "team2":     t2,
                "game_time": hora,
                "venue":     HOME_VENUES.get(t2, ""),
                "game":      f"{t1} vs. {t2}",
                "source":    "basketball24",
            })

    if games:
        return _dedup_games(games)

    # ── Strategy 2 (DISABLED) ─────────────────────────────────────────────
    # The broad team-name regex across the entire 262KB page returns 53+ fake
    # matchups from navigation links, standings tables, and full-season data.
    # Without a confirmed date section in HTML, this is not usable.
    # If a date section was found but CSS parsing found nothing, today really
    # has no games listed in the SSR portion of the page.
    if found_date_section:
        # Today's section found, CSS found nothing → no games in page for today
        return []

    # No date section at all → page didn't include today; return empty rather
    # than return garbage from full-page scan.
    return []


def _parse_b24_feed(text, target_dt):
    """Parse basketball24/flashscore binary-ish text feed."""
    games  = []
    date_s = target_dt.strftime("%d.%m.%Y")

    # The feed uses ¬ ~ ÷ as separators; team names appear in sequences
    # Filter only sections that contain today's date
    if date_s not in text and target_dt.strftime("%Y-%m-%d") not in text:
        return []

    segs = re.split(r'[÷¬~\x00-\x08\x0b\x0e-\x1f]', text)
    for i in range(len(segs) - 1):
        t1 = _norm_b24(segs[i])
        t2 = _norm_b24(segs[i+1]) if i+1 < len(segs) else None
        if t1 and t2 and t1 != t2 and t1 in BSN_TEAMS and t2 in BSN_TEAMS:
            # Try to find a time string nearby
            nearby = "".join(segs[max(0,i-3):i+4])
            tm = re.search(r'\b(\d{1,2}:\d{2})\b', nearby)
            hora = _convert_24h(tm.group(1)) if tm else ""
            games.append({
                "date":      target_dt,
                "team1":     t1,
                "team2":     t2,
                "game_time": hora,
                "venue":     HOME_VENUES.get(t2, ""),
                "game":      f"{t1} vs. {t2}",
                "source":    "basketball24",
            })
    return _dedup_games(games)


def _convert_24h(t):
    """Convert '20:00' → '8:00 PM', leave '8:00 PM' unchanged."""
    if not t: return ""
    t = t.strip()
    if re.search(r'[AP]M', t, re.I): return t   # already 12h
    m = re.match(r'^(\d{1,2}):(\d{2})$', t)
    if not m: return t
    h, mn = int(m.group(1)), int(m.group(2))
    if h == 0:   return f"12:{mn:02d} AM"
    if h < 12:   return f"{h}:{mn:02d} AM"
    if h == 12:  return f"12:{mn:02d} PM"
    return f"{h-12}:{mn:02d} PM"


def _dedup_games(games):
    seen = set(); out = []
    for g in games:
        key = tuple(sorted([g["team1"], g["team2"]]))
        if key not in seen:
            seen.add(key); out.append(g)
    return out


# ── SofaScore API scraper ──────────────────────────────────────────────────
# SofaScore has a real JSON API — no JS rendering needed.
# Returns all scheduled basketball events for a date, we filter for BSN.

SOFA_NAME_MAP = {
    **B24_NAME_MAP,
    # SofaScore uses full team names like "Santeros de Aguada"
    "aguada":        "SANTEROS",
    "vega alta":     "LEONES",
    "quebradillas":  "CANGREJEROS",
    "arecibo":       "CAPITANES",
    "bayamon":       "VAQUEROS",
    "caguas":        "CRIOLLOS",
    "carolina":      "GIGANTES",
    "coamo":         "OSOS",
    "guaynabo":      "METS",
    "humacao":       "INDIOS",
    "juncos":        "ATLETICOS",
    "ponce":         "LEONES",
    "san german":    "CANGREJEROS",
    "santurce":      "CANGREJEROS",
    "fajardo":       "CAPITANES",
    "toa baja":      "VAQUEROS",
}

def _norm_sofa(name):
    if not name: return None
    low = name.lower().strip()
    # Remove "de", "del", city prefixes common in PR team names
    for k, v in SOFA_NAME_MAP.items():
        if k in low: return v
    for t in BSN_TEAMS:
        if t.lower() in low: return t
    return None

# BSN tournament ID on SofaScore — confirmed from URL:
# https://www.sofascore.com/basketball/tournament/puerto-rico/bsn/17374
SOFA_BSN_TOURNAMENT = 17374
SOFA_BSN_SEASON     = 91302   # current 2026 season ID


def _sofa_session():
    """
    Create an authenticated requests.Session for SofaScore.

    Environment variables (set before running the script):
      SOFASCORE_CF_CLEARANCE  →  value of the cf_clearance cookie from Chrome DevTools
                                  (Application → Cookies → sofascore.com → cf_clearance)
      SOFASCORE_UA            →  User-Agent string from Chrome (DevTools → Network → any
                                  request → Request Headers → User-Agent). Must match the
                                  UA that was used when cf_clearance was generated.

    browser_cookie3 loads the rest of the Chrome cookies automatically.
    """
    # ── Determine User-Agent ──────────────────────────────────────────────
    # cf_clearance is bound to the UA that solved the Cloudflare challenge.
    # Use SOFASCORE_UA if set; otherwise fall back to a common Mac Chrome UA.
    ua = os.environ.get("SOFASCORE_UA", "").strip() or (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )

    s = requests.Session()
    s.headers.update({
        "User-Agent":         ua,
        "Accept":             "application/json, text/plain, */*",
        "Accept-Language":    "en-US,en;q=0.9,es;q=0.8",
        "Accept-Encoding":    "gzip, deflate, br",
        "Referer":            "https://www.sofascore.com/",
        "Origin":             "https://www.sofascore.com",
        "sec-ch-ua":          '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        "sec-ch-ua-mobile":   "?0",
        "sec-ch-ua-platform": '"macOS"',
        "sec-fetch-dest":     "empty",
        "sec-fetch-mode":     "cors",
        "sec-fetch-site":     "same-site",
        "Cache-Control":      "no-cache",
        "Pragma":             "no-cache",
    })

    # ── 1. cf_clearance env var (Cloudflare bypass) ───────────────────────
    cf_val = os.environ.get("SOFASCORE_CF_CLEARANCE", "").strip()
    if cf_val:
        # Cloudflare checks the cf_clearance cookie on the root domain
        s.cookies.set("cf_clearance", cf_val, domain=".sofascore.com")
        print(f"  🍪 SofaScore: cf_clearance desde SOFASCORE_CF_CLEARANCE ✓")

    # ── 2. browser_cookie3 — pull the rest of Chrome/Safari/Firefox cookies ─
    # cf_clearance is NOT persisted to disk by Chrome (session-only), so
    # browser_cookie3 won't find it — but we inject it above from env var.
    try:
        import browser_cookie3
        for browser_fn, bname in [
            (lambda: browser_cookie3.chrome(domain_name="sofascore.com"),  "Chrome"),
            (lambda: browser_cookie3.safari(domain_name="sofascore.com"),  "Safari"),
            (lambda: browser_cookie3.firefox(domain_name="sofascore.com"), "Firefox"),
        ]:
            try:
                cookies = browser_fn()
                clist   = list(cookies)
                sofa_c  = [c for c in clist if "sofascore" in c.domain.lower()]
                if sofa_c:
                    s.cookies.update(cookies)
                    # Don't overwrite cf_clearance if we injected it from env var
                    if cf_val:
                        s.cookies.set("cf_clearance", cf_val, domain=".sofascore.com")
                    print(f"  🍪 SofaScore: {len(sofa_c)} cookie(s) desde {bname}"
                          + (" + cf_clearance ✓" if cf_val else " (sin cf_clearance)"))
                    return s
            except Exception:
                pass
    except ImportError:
        pass

    if cf_val:
        print(f"  🍪 SofaScore: solo cf_clearance (sin cookies de Chrome)")
        return s

    print("  ⚠️  SofaScore: sin cf_clearance — obten el valor así:")
    print("      1. Abre sofascore.com en Chrome")
    print("      2. DevTools (F12) → Application → Cookies → sofascore.com")
    print("      3. Copia el valor de 'cf_clearance'")
    print("      4. export SOFASCORE_CF_CLEARANCE='<valor>'")
    print("      5. Corre bsn.py de nuevo")
    return s


def _sofa_curl_get(url, cf_clearance=None, ua=None):
    """
    Use curl subprocess as fallback when Python requests gets 403.
    Passes cf_clearance cookie and matching User-Agent if provided.
    Returns (status_code, response_text) or (None, None) on failure.
    """
    import subprocess
    if ua is None:
        ua = os.environ.get("SOFASCORE_UA", "").strip() or (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
    if cf_clearance is None:
        cf_clearance = os.environ.get("SOFASCORE_CF_CLEARANCE", "").strip()

    cmd = [
        "curl", "-s", "-L",
        "--max-time", "20",
        "-H", "Accept: application/json",
        "-H", "Accept-Language: en-US,en;q=0.9",
        "-H", "Origin: https://www.sofascore.com",
        "-H", "Referer: https://www.sofascore.com/",
        "-H", "sec-fetch-dest: empty",
        "-H", "sec-fetch-mode: cors",
        "-H", "sec-fetch-site: same-site",
        "-H", f"User-Agent: {ua}",
        "--write-out", "\n__STATUS__%{http_code}",
    ]
    if cf_clearance:
        cmd += ["-b", f"cf_clearance={cf_clearance}"]

    cmd.append(url)

    try:
        import subprocess
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=25)
        out = result.stdout
        if "__STATUS__" in out:
            body, status_str = out.rsplit("__STATUS__", 1)
            try:
                return int(status_str.strip()), body
            except ValueError:
                pass
        return None, None
    except Exception:
        return None, None


def _sofa_parse_events(events, target_dt):
    """Parse a SofaScore events list → BSN game dicts."""
    from datetime import timezone, timedelta as _td
    games = []
    for ev in events:
        home = ev.get("homeTeam", {}).get("name", "")
        away = ev.get("awayTeam", {}).get("name", "")
        t1   = _norm_sofa(away)
        t2   = _norm_sofa(home)
        if not (t1 and t2 and t1 != t2):
            continue
        ts   = ev.get("startTimestamp", 0)
        hora = ""
        if ts:
            dt_utc = datetime.fromtimestamp(ts, tz=timezone.utc)
            dt_ast = dt_utc - _td(hours=4)   # Puerto Rico = UTC-4
            if dt_ast.date() != target_dt:
                continue
            h  = dt_ast.hour; mn = dt_ast.minute
            ap = "PM" if h >= 12 else "AM"
            if h > 12: h -= 12
            if h == 0: h = 12
            hora = f"{h}:{mn:02d} {ap}"
        games.append({
            "date":      target_dt,
            "team1":     t1,
            "team2":     t2,
            "game_time": hora,
            "venue":     HOME_VENUES.get(t2, ""),
            "game":      f"{t1} vs. {t2}",
            "source":    "SofaScore",
        })
    return games


def _sofa_fetch(s, url):
    """
    Fetch a SofaScore API URL.
    Strategy (in order):
      1. api.sofascore.app  — mobile endpoint, usually no CF protection
      2. requests session   — with browser cookies + cf_clearance if set
      3. curl               — passes cf_clearance cookie explicitly
    Returns parsed JSON dict or None.
    """
    import json as _json

    # ── 1. Try api.sofascore.app (mobile, often CF-free) ─────────────────
    app_url = url.replace("api.sofascore.com", "api.sofascore.app")
    _MOBILE_UA = (
        "SofaScore/167 CFNetwork/1474 Darwin/23.0.0"
    )
    try:
        r = requests.get(
            app_url, timeout=15,
            headers={
                "User-Agent":     _MOBILE_UA,
                "Accept":         "application/json",
                "Accept-Language":"en-US,en;q=0.9",
            }
        )
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass

    # ── 2. Try requests session (browser cookies + cf_clearance) ─────────
    try:
        r = s.get(url, timeout=20)
        if r.status_code == 200:
            return r.json()
        if r.status_code != 403:
            return None
    except Exception:
        pass

    # ── 3. Fallback: curl (passes cf_clearance cookie explicitly) ─────────
    status, body = _sofa_curl_get(url)
    if status == 200 and body:
        try:
            return _json.loads(body)
        except Exception:
            pass

    # ── 3b. curl on .app domain too ───────────────────────────────────────
    if app_url != url:
        status, body = _sofa_curl_get(app_url)
        if status == 200 and body:
            try:
                return _json.loads(body)
            except Exception:
                pass

    return None


def scrape_sofascore_schedule(target_date_str):
    """
    Fetch BSN schedule from SofaScore.
    Tournament ID 17374 / Season 91302 confirmed from:
      sofascore.com/basketball/tournament/puerto-rico/bsn/17374

    Tries requests (with browser cookies) then falls back to curl for
    Cloudflare 403 bypass.
    """
    target_dt = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    s         = _sofa_session()
    games     = []

    # ── Endpoint 1: Next / Last events for BSN tournament (season 91302) ─
    endpoints = [
        f"https://api.sofascore.com/api/v1/unique-tournament/{SOFA_BSN_TOURNAMENT}"
        f"/season/{SOFA_BSN_SEASON}/events/next/0",
        f"https://api.sofascore.com/api/v1/unique-tournament/{SOFA_BSN_TOURNAMENT}"
        f"/season/{SOFA_BSN_SEASON}/events/last/0",
        f"https://api.sofascore.com/api/v1/unique-tournament/{SOFA_BSN_TOURNAMENT}"
        f"/season/latest/events/next/0",
    ]
    for url in endpoints:
        data = _sofa_fetch(s, url)
        if data is not None:
            evs = data.get("events", [])
            g   = _sofa_parse_events(evs, target_dt)
            if g:
                return _dedup_games(g)

    # ── Endpoint 2: Scheduled events for the date (all sports, filter BSN) ─
    url  = (f"https://api.sofascore.com/api/v1/sport/basketball"
             f"/scheduled-events/{target_date_str}")
    data = _sofa_fetch(s, url)
    if data is not None:
        evs = data.get("events", [])
        bsn_evs = [
            ev for ev in evs
            if (ev.get("tournament", {}).get("uniqueTournament", {}).get("id") == SOFA_BSN_TOURNAMENT
                or "bsn" in ev.get("tournament", {}).get("name", "").lower()
                or "puerto rico" in ev.get("tournament", {}).get(
                    "category", {}).get("name", "").lower())
        ]
        g = _sofa_parse_events(bsn_evs or evs, target_dt)
        games.extend(g)

    return _dedup_games(games)


def cmd_debug_schedule():
    """
    --debug-schedule: Prueba RealGM y Flashscore, muestra la respuesta cruda.
    Útil para diagnosticar por qué no se encuentran juegos.
    """
    td  = TARGET_DATE
    print(f"\n{'═'*60}")
    print(f"  DEBUG SCHEDULE — {td}")
    print(f"{'═'*60}\n")

    # ── 1. betsapi.com ────────────────────────────────────────────────────
    print("─── betsapi.com (league 4479) ───")
    bets_urls = [
        "https://betsapi.com/basketball/ls/4479/puerto-rico-superior-nacional",
        f"https://betsapi.com/basketball/ls/4479/puerto-rico-superior-nacional?date={td}",
    ]
    for url in bets_urls:
        try:
            r = requests.get(url, headers={**HEADERS, "Referer": "https://betsapi.com/"}, timeout=15)
            print(f"  {url}")
            print(f"  → Status: {r.status_code}  |  Size: {len(r.text)} chars")
            if r.status_code == 200:
                # Show any table rows
                tables = re.findall(r'<table[^>]*>(.*?)</table>', r.text, re.DOTALL|re.IGNORECASE)
                clean  = lambda s: re.sub(r'<[^>]+>', '', s).strip()
                if tables:
                    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tables[0], re.DOTALL|re.IGNORECASE)
                    print(f"  Table rows: {len(rows)}")
                    for rh in rows[:5]:
                        cells = [clean(c) for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>',
                                                                rh, re.DOTALL|re.IGNORECASE)]
                        if any(cells): print(f"    {' | '.join(cells[:6])}")
                else:
                    # Show snippet of raw text
                    snippet = re.sub(r'<[^>]+>', ' ', r.text)
                    snippet = re.sub(r'\s+', ' ', snippet).strip()[:400]
                    print(f"  Text snippet: {snippet}")
        except Exception as e:
            print(f"  ERROR: {e}")
        print()

    # scraper function
    print("─── scrape_betsapi_schedule() resultado ───")
    try:
        games = scrape_betsapi_schedule(td)
        if games:
            print(f"  ✅ {len(games)} juego(s):")
            for g in games: print(f"     {g['team1']} @ {g['team2']}  |  {g['game_time'] or 'TBD'}")
        else:
            print("  ❌ No se encontraron juegos")
    except Exception as e:
        print(f"  ERROR: {e}")

    # ── 2. RealGM — per-date URL ──────────────────────────────────────────
    print("\n─── RealGM BSN ───")
    realgm_urls = [
        f"https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/schedules/{td}",
        f"https://basketball.realgm.com/international/league/62/puerto-rican-bsn/schedules/{td}",
        f"https://basketball.realgm.com/international/league/62/puerto-rican-bsn/schedules",
    ]
    for url in realgm_urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            print(f"  {url}")
            print(f"  → Status: {r.status_code}  |  Size: {len(r.text)} chars")
            if r.status_code == 200 and len(r.text) > 500:
                # Show first table found
                tables = re.findall(r'<table[^>]*>(.*?)</table>', r.text, re.DOTALL|re.IGNORECASE)
                if tables:
                    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tables[0], re.DOTALL|re.IGNORECASE)
                    clean = lambda s: re.sub(r'<[^>]+>', '', s).strip()
                    print(f"  Table rows: {len(rows)}")
                    for row_html in rows[:6]:
                        cells = [clean(c) for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>',
                                                               row_html, re.DOTALL|re.IGNORECASE)]
                        if any(cells): print(f"    {' | '.join(cells[:6])}")
                else:
                    print(f"  (no <table> found in HTML)")
            elif r.status_code != 200:
                print(f"  Response snippet: {r.text[:200]}")
        except Exception as e:
            print(f"  ERROR: {e}")
        print()

    # Try the scraper function directly
    print("─── scrape_realgm_schedule() resultado ───")
    try:
        games = scrape_realgm_schedule(td)
        if games:
            print(f"  ✅ {len(games)} juego(s) encontrado(s):")
            for g in games:
                print(f"     {g['team1']} @ {g['team2']}  |  {g['game_time'] or 'TBD'}")
        else:
            print("  ❌ No se encontraron juegos")
    except Exception as e:
        print(f"  ERROR en scraper: {e}")

    # ── 2. Flashscore ─────────────────────────────────────────────────────
    print("\n─── Flashscore BSN ───")
    try:
        r = requests.get("https://www.flashscore.com/basketball/puerto-rico/bsn/",
                         headers={**HEADERS, "Referer": "https://www.flashscore.com/"}, timeout=15)
        print(f"  Status: {r.status_code}  |  Size: {len(r.text)} chars")
        if r.status_code == 200:
            date_s = datetime.strptime(td, "%Y-%m-%d").strftime("%d.%m.%Y")
            if td in r.text or date_s in r.text:
                print(f"  ✅ Fecha '{td}' o '{date_s}' encontrada en HTML")
            else:
                print(f"  ⚠️  Fecha no encontrada (página puede ser JS-rendered)")
    except Exception as e:
        print(f"  ERROR: {e}")

    print(f"\n{'═'*60}")
    print("  Comparte este output y lo arreglo.")
    print(f"{'═'*60}\n")


def load_scheduled_games(wb, target_date_str=None):
    """
    Obtiene juegos para la fecha dada.
    Priority: Manual → RealGM → Flashscore → Manual fallback → Excel
    """
    td = target_date_str or TARGET_DATE

    # ── 1. Juegos manuales (--add-game) — siempre van primero ────────────
    manual = _get_manual_games(td)
    if manual:
        print(f"  📋 {len(manual)} juego(s) manual(es) encontrado(s) para {td}")

    # ── 2. Intentar fuentes web ───────────────────────────────────────────
    print(f"  📅 Buscando juegos BSN para {td}...")

    # 2a. betsapi.com — league 4479
    print(f"  🔄 betsapi.com...")
    bets_games = scrape_betsapi_schedule(td)
    if bets_games:
        print(f"  ✅ {len(bets_games)} juego(s) [betsapi]")
        return _merge_games(manual, bets_games)

    # 2b. RealGM — segunda opción
    print(f"  🔄 RealGM...")
    realgm_games = scrape_realgm_schedule(td)
    if realgm_games:
        print(f"  ✅ {len(realgm_games)} juego(s) [RealGM]")
        return _merge_games(manual, realgm_games)

    # 2c. Flashscore.com — último recurso
    print(f"  🔄 Flashscore...")
    fs_games = scrape_flashscore_schedule(td)
    if fs_games:
        print(f"  ✅ {len(fs_games)} juego(s) [Flashscore]")
        return _merge_games(manual, fs_games)

    # ── 3. Juegos manuales solos (si web falló) ───────────────────────────
    if manual:
        print(f"  ✅ Usando juegos manuales.")
        return manual

    print(f"  ⚠️  No se encontraron juegos en ninguna fuente.")
    print(f"      → Agrega los juegos manualmente:")
    print(f"        python3 bsn.py --add-game SANTEROS LEONES '8:00 PM'")
    print(f"  🔄 Intentando Excel...")

    # ── 4. Fallback: Excel BSN Lines ──────────────────────────────────────
    target_dt = datetime.strptime(td, "%Y-%m-%d").date()
    games     = []
    for row in wb[BSN_LINES_SHEET].iter_rows(min_row=5, max_row=wb[BSN_LINES_SHEET].max_row,
                                              values_only=True):
        if not row[3]: continue
        try:
            row_date = row[3].date() if isinstance(row[3], datetime) \
                       else datetime.strptime(str(row[3])[:10],"%Y-%m-%d").date()
        except: continue
        if row_date != target_dt: continue
        team1 = norm_team(str(row[5])) if row[5] else None
        team2 = norm_team(str(row[6])) if row[6] else None
        if not team1 or not team2: continue
        games.append({
            "date":      row_date,
            "game":      str(row[4] or "").strip(),
            "team1":     team1,
            "team2":     team2,
            "game_time": str(row[17] if len(row)>17 and row[17] else ""),
            "venue":     HOME_VENUES.get(team2, ""),
            "source":    "Excel",
        })
    return games


def _merge_games(manual, web_games):
    """Merge manual games with web games, deduplicating by team pair."""
    if not manual:
        return web_games
    existing_keys = {tuple(sorted([g["team1"], g["team2"]])) for g in web_games}
    extra = [m for m in manual
             if tuple(sorted([m["team1"], m["team2"]])) not in existing_keys]
    return web_games + extra

# ──────────────────────────────────────────────────────
# MODELO — CÁLCULOS
# ──────────────────────────────────────────────────────

def compute_game(team1, team2, stats, injury_impact):
    """
    Calcula expected points, spread, total y win prob para un juego.
    team1 = equipo visitante, team2 = equipo local.
    """
    # Home Court Advantage — BSN: el local tiene ventaja real de cancha.
    # Investigación empírica en ligas latinoamericanas de baloncesto: ~4-5 pts.
    # Se aplica sumando al local (team2) y restando al visitante (team1).
    HCA_BSN = 4.5   # puntos de ventaja neta para el equipo local

    s1 = stats.get(team1, {"ortg":100.0,"drtg":100.0,"pace":LEAGUE_AVG_PACE})
    s2 = stats.get(team2, {"ortg":100.0,"drtg":100.0,"pace":LEAGUE_AVG_PACE})

    pace = (s1["pace"] + s2["pace"]) / 2

    raw_pts1 = (s1["ortg"] + s2["drtg"]) / 2 * pace / 100
    raw_pts2 = (s2["ortg"] + s1["drtg"]) / 2 * pace / 100

    inj1 = injury_impact.get(team1, 0.0)
    inj2 = injury_impact.get(team2, 0.0)

    # Aplicar HCA: local suma, visitante resta (split equitativo)
    pts1 = max(raw_pts1 - inj1 - (HCA_BSN / 2), 50.0)
    pts2 = max(raw_pts2 - inj2 + (HCA_BSN / 2), 50.0)

    total  = pts1 + pts2
    spread = pts1 - pts2   # positivo = team1 favorito
    wp1    = pyth_win_prob(pts1, pts2)
    wp2    = 1.0 - wp1

    ml1 = prob_to_american(wp1)
    ml2 = prob_to_american(wp2)

    # Redondear spread al .5 más cercano
    def _round_half(x):
        return round(x * 2) / 2

    fav = team1 if spread > 0 else (team2 if spread < 0 else None)
    spread_r = _round_half(abs(spread))
    if fav and spread_r >= 0.5:
        spread_line = f"{fav} -{spread_r:.1f}"
    else:
        spread_line = "PICK"

    return {
        "pts1": round(pts1, 2),
        "pts2": round(pts2, 2),
        "total": round(total, 1),
        "spread": round(spread, 1),
        "spread_line": spread_line,
        "wp1": round(wp1*100, 1),
        "wp2": round(wp2*100, 1),
        "ml1": ml1,
        "ml2": ml2,
        "inj1": round(inj1, 2),
        "inj2": round(inj2, 2),
        "pace": round(pace, 1),
    }

# ──────────────────────────────────────────────────────
# DISPLAY FUNCTIONS
# ──────────────────────────────────────────────────────

def display_lines(games_data):
    dt  = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    print(f"\n{'═'*68}")
    print(f"  LABOY PICKS — BSN   {dt.strftime('%A, %B %d %Y').upper()}")
    print(f"{'═'*68}")

    if not games_data:
        print(f"\n  No hay juegos BSN para {TARGET_DATE}.")
        print(f"  Agrega los juegos manualmente con --add-game o en el tab 'BSN Lines' del Excel.\n")
        return

    for g in games_data:
        r      = g["result"]
        t1     = g["team1"]; t2 = g["team2"]
        gtime  = g.get("game_time","")
        venue  = g.get("venue","")
        print(f"\n  {'─'*64}")
        hdr = f"  {t1:<16} @  {t2}"
        if gtime: hdr += f"   │  {gtime}"
        print(hdr)
        if venue:
            print(f"  🏟️  {venue}")
        inj_note = ""
        if r["inj1"] or r["inj2"]:
            inj_note = f"  ⚠️  Injury adj: {t1} -{r['inj1']:.1f} pts  │  {t2} -{r['inj2']:.1f} pts"
        print(f"  {'─'*64}")
        print(f"  🎯 MODEL   │  Total: {r['total']:.1f}   │  Spread: {r['spread_line']}")
        print(f"             │  {t1}: {r['pts1']:.1f} pts ({r['wp1']}%)  │  {t2}: {r['pts2']:.1f} pts ({r['wp2']}%)")
        print(f"             │  ML: {t1} {fmt_odds(r['ml1'])} / {t2} {fmt_odds(r['ml2'])}")
        if inj_note: print(inj_note)

    print(f"\n{'═'*68}\n")


# ── Parámetros probabilísticos para EV (BSN basketball) ──────────────────────
# BSN vs NBA: el mercado de BSN es menos eficiente (menos volumen, menos sharp money).
# Por eso el σ efectivo es menor — la línea no está tan bien calibrada como NBA.
# NBA referencia: σ_total ≈ 12-13 pts, σ_spread ≈ 9-10 pts.
# BSN: usamos valores ligeramente menores, reflejando que 3-4 pts de edge
#       es más significativo en una liga con líneas menos afiladas.
SIGMA_SPREAD_BSN = 10.0   # desviación estándar del margen de victoria (pts)
SIGMA_TOTAL_BSN  = 12.0   # desviación estándar del total (pts) — era 15, demasiado alto para BSN


def _normal_cdf(z):
    """Aproximación de la CDF normal estándar (Abramowitz & Stegun 26.2.17)."""
    sign = 1 if z >= 0 else -1
    z    = abs(z)
    t    = 1.0 / (1.0 + 0.2316419 * z)
    p    = t * (0.319381530 + t * (-0.356563782
               + t * (1.781477937 + t * (-1.821255978 + t * 1.330274429))))
    return (1.0 - (1.0 / math.sqrt(2 * math.pi)) * math.exp(-0.5 * z * z) * p
            ) if sign > 0 else (
             (1.0 / math.sqrt(2 * math.pi)) * math.exp(-0.5 * z * z) * p)


def _bet_win_prob(pick_type, edge_val):
    """
    Estima P(ganar la apuesta) basado en tipo de pick y tamaño del edge.
    ML: edge_val = model_prob - market_prob; P = market_prob + edge_val (calculado fuera).
    SPREAD: edge_val = pts de diferencia modelo vs mercado.
    OVER/UNDER: edge_val = pts de diferencia modelo vs mercado.
    """
    if pick_type == "SPREAD":
        return _normal_cdf(edge_val / SIGMA_SPREAD_BSN) * 100
    elif pick_type in ("OVER", "UNDER"):
        return _normal_cdf(edge_val / SIGMA_TOTAL_BSN) * 100
    return None   # ML usa win_prob directamente


def _ev_str(win_prob_pct, odds_str):
    """EV por $100 apostados, expresado como porcentaje."""
    try:
        p = max(0.0, min(1.0, win_prob_pct / 100.0))
        o = int(str(odds_str).replace("+","").strip())
        payout = (100.0 / abs(o)) if o < 0 else (o / 100.0)
        ev = p * payout - (1.0 - p)
        col = "#22c55e" if ev > 0 else "#ef4444"
        return f"{ev*100:+.1f}%", col
    except Exception:
        return "—", "#94a3b8"


MAX_MARKET_SPREAD = 5.5   # filtro: no apostar spread cuando el mercado da ≥6 pts
                          # Análisis histórico BSN: juegos con spread grande (±6+)
                          # muestran 50% win rate (-70u) → sin edge del modelo.
                          # Preferir ML o Total en esos partidos si hay señal.

def _find_value_picks(games_data, min_spread_diff=1.5, min_ml_edge=3.0, min_total_diff=3.0):
    """
    Motor principal de picks con valor real.

    Cuando hay líneas de mercado:
      • Spread edge  — modelo vs mercado difieren ≥ min_spread_diff pts
        - Mercado infla al favorito más que el modelo → apostar al UNDERDOG + puntos
        - Mercado da menos puntos al favorito que el modelo → apostar al favorito
        - Modelos y mercados discrepan en quién gana → pick fuerte en la dirección del modelo
        - FILTRO: mercado ≥6 pts → spread suprimido (sin edge histórico en BSN)
      • ML edge      — modelo win% supera implícita del mercado ≥ min_ml_edge %
      • Total edge   — diferencia modelo/mkt ≥ min_total_diff pts
        (BSN: 3.0 pts es edge real — mercado menos eficiente que NBA)

    Sin líneas de mercado:
      Lógica original — spread puro del modelo ≥ 2.0 pts → recomendar ML favorito.

    Retorna lista de dicts con todos los campos necesarios para show_picks y export HTML.
    """
    picks = []

    for g in games_data:
        r     = g["result"]
        t1    = g["team1"]
        t2    = g["team2"]
        gtime = g.get("game_time", "")

        model_spread_t1 = r["spread"]          # positivo → t1 favorito
        model_total     = r["total"]
        wp1  = float(r["wp1"]); wp2  = float(r["wp2"])
        ml1  = r["ml1"];        ml2  = r["ml2"]

        mkt = _get_market_line(t1, t2)

        # ── Sin mercado: lógica original ──────────────────────────────────────
        if not mkt:
            sp_abs = abs(model_spread_t1)
            if sp_abs >= 2.0:
                fav  = t1 if model_spread_t1 > 0 else t2
                wp   = wp1 if model_spread_t1 > 0 else wp2
                ml   = ml1 if model_spread_t1 > 0 else ml2
                picks.append({
                    "game":         f"{t1} @ {t2}",
                    "time":         gtime,
                    "pick":         f"{fav} ML",
                    "pick_type":    "ML",
                    "odds":         fmt_odds(ml),
                    "edge_val":     sp_abs,
                    "edge_str":     f"—",
                    "reason":       f"Spread modelo {sp_abs:.1f} pts (sin mercado)",
                    "wp":           f"{wp:.1f}%",
                    "model_spread": r["spread_line"],
                    "mkt_spread":   "—",
                    "model_total":  f"{model_total:.1f}",
                    "mkt_total":    "—",
                    "has_market":   False,
                    "alt_picks":    [],
                })
            continue

        # ── Con mercado: buscar discrepancias ─────────────────────────────────
        model_fav = t1 if model_spread_t1 >= 0 else t2
        model_dog = t2 if model_spread_t1 >= 0 else t1
        fav_abs   = abs(model_spread_t1)
        wp_fav    = wp1 if model_fav == t1 else wp2
        wp_dog    = wp2 if model_fav == t1 else wp1
        ml_fav    = ml1 if model_fav == t1 else ml2
        ml_dog    = ml2 if model_fav == t1 else ml1

        candidates = []

        # ── 1. SPREAD ─────────────────────────────────────────────────────────
        mkt_sp_fav      = mkt.get("spread_fav", "").upper().strip()
        mkt_sp_line     = mkt.get("spread_line", "")   # e.g., "-4.5"
        mkt_sp_odds     = mkt.get("spread_odds", "-110")
        mkt_sp_dog_odds = mkt.get("spread_dog_odds", "-110")

        if mkt_sp_fav and mkt_sp_line:
            try:
                mkt_sp_abs = abs(float(mkt_sp_line))

                # ── Filtro spread grande: sin edge en BSN cuando mkt ≥6 pts ──────
                # Históricamente 50% W rate en spreads grandes → no apostar spread.
                # Sí podemos seguir evaluando ML y Total para ese juego.
                if mkt_sp_abs > MAX_MARKET_SPREAD:
                    raise ValueError(f"spread {mkt_sp_abs} > MAX_MARKET_SPREAD — omitido")

                if mkt_sp_fav == model_fav.upper():
                    market_for_fav = mkt_sp_abs
                else:
                    market_for_fav = -mkt_sp_abs

                spread_diff = fav_abs - market_for_fav

                if spread_diff >= min_spread_diff:
                    if mkt_sp_fav == model_fav.upper():
                        # Modelo y mercado coinciden en el favorito, pero modelo lo da por más
                        pick_str  = f"{model_fav} -{mkt_sp_abs} ({mkt_sp_odds})"
                        p_odds    = mkt_sp_odds
                        reason    = (f"Modelo {model_fav} -{fav_abs:.1f} pts, mkt solo pide "
                                     f"-{mkt_sp_abs} → {model_fav} cubre")
                    else:
                        # Modelo dice que model_fav gana, pero mkt lo pone como underdog
                        # → tomar model_fav + puntos, a las odds del underdog del mercado
                        pick_str  = f"{model_fav} +{mkt_sp_abs} ({mkt_sp_dog_odds})"
                        p_odds    = mkt_sp_dog_odds   # FIX: era mkt_sp_odds (odds del fav de mkt)
                        reason    = (f"Modelo gana {model_fav} por {fav_abs:.1f} pts, "
                                     f"mkt da +{mkt_sp_abs} gratis → {model_fav} cubierto")
                    candidates.append({
                        "pick": pick_str, "pick_type": "SPREAD",
                        "pick_odds":   p_odds,
                        "edge_val":    spread_diff,
                        "edge_str":    f"+{spread_diff:.1f} pts",
                        "reason":      reason,
                        "modelo_str":  r["spread_line"],
                        "mercado_str": f"{mkt_sp_fav} {mkt_sp_line} ({mkt_sp_odds})",
                        "win_prob":    _bet_win_prob("SPREAD", spread_diff),
                    })

                elif spread_diff <= -min_spread_diff:
                    dog_odds = mkt_sp_dog_odds
                    pick_str = f"{model_dog} +{mkt_sp_abs} ({dog_odds})"
                    reason   = (f"Mkt da {mkt_sp_abs:.1f} pts al {model_fav}, "
                                f"modelo solo {fav_abs:.1f} → {model_dog} cubre")
                    candidates.append({
                        "pick": pick_str, "pick_type": "SPREAD",
                        "pick_odds":   dog_odds,
                        "edge_val":    abs(spread_diff),
                        "edge_str":    f"+{abs(spread_diff):.1f} pts",
                        "reason":      reason,
                        "modelo_str":  r["spread_line"],
                        "mercado_str": f"{mkt_sp_fav} {mkt_sp_line} ({mkt_sp_odds})",
                        "win_prob":    _bet_win_prob("SPREAD", abs(spread_diff)),
                    })

            except Exception:
                pass

        # ── 2. ML ─────────────────────────────────────────────────────────────
        mkt_ml_fav = mkt.get("ml1","") if model_fav == t1 else mkt.get("ml2","")
        mkt_ml_dog = mkt.get("ml2","") if model_fav == t1 else mkt.get("ml1","")

        for team, team_wp, mkt_ml in [
            (model_fav, wp_fav, mkt_ml_fav),
            (model_dog, wp_dog, mkt_ml_dog),
        ]:
            if not mkt_ml: continue
            prob_mkt = _american_to_prob(mkt_ml)
            if prob_mkt is None: continue
            ml_edge = team_wp - prob_mkt
            if ml_edge >= min_ml_edge:
                candidates.append({
                    "pick":        f"{team} ML ({mkt_ml})",
                    "pick_type":   "ML",
                    "pick_odds":   mkt_ml,
                    "edge_val":    ml_edge,
                    "edge_str":    f"+{ml_edge:.1f}%",
                    "reason":      (f"Win% modelo {team_wp:.1f}% vs "
                                    f"implícita mercado {prob_mkt:.1f}%"),
                    "modelo_str":  fmt_odds(prob_to_american(team_wp / 100.0)),
                    "mercado_str": fmt_odds(int(mkt_ml)) if str(mkt_ml).lstrip("+-").isdigit() else str(mkt_ml),
                    "win_prob":    team_wp,
                })

        # ── 3. TOTAL ──────────────────────────────────────────────────────────
        mkt_tot   = mkt.get("total","")
        mkt_ov_o  = mkt.get("over_odds","-110")
        mkt_un_o  = mkt.get("under_odds","-110")

        if mkt_tot:
            try:
                mkt_tv = float(mkt_tot)
                tdiff  = model_total - mkt_tv
                if tdiff >= min_total_diff:
                    candidates.append({
                        "pick":        f"OVER {mkt_tot} ({mkt_ov_o})",
                        "pick_type":   "OVER",
                        "pick_odds":   mkt_ov_o,
                        "edge_val":    tdiff,
                        "edge_str":    f"{tdiff:+.1f} pts",
                        "reason":      f"Modelo {model_total:.1f} pts > mkt {mkt_tot}",
                        "modelo_str":  f"{model_total:.1f}",
                        "mercado_str": f"{mkt_tot}",
                        "win_prob":    _bet_win_prob("OVER", tdiff),
                    })
                elif tdiff <= -min_total_diff:
                    candidates.append({
                        "pick":        f"UNDER {mkt_tot} ({mkt_un_o})",
                        "pick_type":   "UNDER",
                        "pick_odds":   mkt_un_o,
                        "edge_val":    abs(tdiff),
                        "edge_str":    f"+{abs(tdiff):.1f} pts",   # siempre positivo: edge a favor del Under
                        "reason":      f"Modelo {model_total:.1f} pts — mkt {mkt_tot} → Under por {abs(tdiff):.1f}",
                        "modelo_str":  f"{model_total:.1f}",
                        "mercado_str": f"{mkt_tot}",
                        "win_prob":    _bet_win_prob("UNDER", abs(tdiff)),
                    })
            except Exception:
                pass

        # ── Mejor candidato ───────────────────────────────────────────────────
        if not candidates:
            continue

        # Ordenar por win_prob (mismas unidades para todos los tipos de pick).
        # edge_val tiene unidades diferentes: puntos para SPREAD/TOTAL, % para ML.
        # Comparar directamente causaba que ML (18.2%) siempre ganara vs spread (7.0 pts)
        # aunque el spread tuviera mucha mayor probabilidad de ganar (75% vs 60%).
        best = max(candidates, key=lambda x: x.get("win_prob") or 0)
        _full_thresh = min(min_ml_edge, min_spread_diff)
        # alt_picks_full: candidatos con valor real (umbral completo) → tarjeta propia
        # alt_picks: candidatos de valor parcial (70%) → solo texto "Alt:"
        alt_full = [c for c in candidates
                    if c is not best and c["edge_val"] >= _full_thresh]
        alt      = [c["pick"] for c in candidates
                    if c is not best and c["edge_val"] >= _full_thresh * 0.7
                    and c not in alt_full]

        mkt_sp_str = (f"{mkt_sp_fav} {mkt_sp_line} ({mkt_sp_odds})"
                      if mkt_sp_fav and mkt_sp_line else "—")
        total_str  = (f"O {mkt_tot} ({mkt_ov_o}) / U {mkt_tot} ({mkt_un_o})"
                      if mkt_tot else "—")

        picks.append({
            "game":         f"{t1} @ {t2}",
            "time":         gtime,
            "pick":         best["pick"],
            "pick_type":    best["pick_type"],
            "pick_odds":    best.get("pick_odds",""),
            "odds":         "",
            "edge_val":     best["edge_val"],
            "edge_str":     best["edge_str"],
            "reason":       best.get("reason",""),
            "modelo_str":   best.get("modelo_str","—"),
            "mercado_str":  best.get("mercado_str","—"),
            "win_prob":     best.get("win_prob"),
            "wp":           f"{wp_fav:.1f}%",
            "model_spread": r["spread_line"],
            "mkt_spread":   mkt_sp_str,
            "model_total":  f"{model_total:.1f}",
            "mkt_total":    total_str,
            "has_market":   True,
            "alt_picks":      alt,
            "alt_picks_full": alt_full,
        })

    return picks


def show_picks(games_data, min_edge=5.0):
    """
    Muestra picks del modelo con valor real.
    Usa _find_value_picks() para encontrar discrepancias modelo vs mercado.
    """
    dt         = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    has_market = any(_get_market_line(g["team1"], g["team2"]) for g in games_data)

    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS BSN — PICKS CON VALOR   {dt.strftime('%a %b %d').upper()}")
    if has_market:
        print(f"  (Buscando discrepancias modelo vs mercado — spread / ML / totales)")
    else:
        print(f"  (Sin líneas de mercado — ejecuta --set-lines para análisis completo)")
    print(f"{'═'*72}")

    picks = _find_value_picks(games_data)

    if not picks:
        if has_market:
            print("\n  ✅  Sin discrepancias claras — el mercado está bien alineado con el modelo.")
            print("  (Ningún pick supera los umbrales: spread ≥1.5 pts, ML ≥3%, total ≥3 pts)\n")
        else:
            print("\n  No hay picks claros (todos los juegos están muy parejos).\n")
            print("  💡 Tip: ejecuta --set-lines para ingresar las odds del libro y ver edge real.\n")
    else:
        # Separar picks con y sin mercado
        mkt_picks  = [p for p in picks if p["has_market"]]
        solo_picks = [p for p in picks if not p["has_market"]]

        if mkt_picks:
            mkt_picks.sort(key=lambda x: _parse_time_sort(x["time"]))

            # ── Picks primarios ──────────────────────────────────────────────
            rows = []
            for p in mkt_picks:
                rows.append([
                    p["game"],
                    p["time"] or "—",
                    p["pick"],
                    p["edge_str"],
                    f"Mod {p['model_spread']}  |  Mkt {p['mkt_spread']}",
                    p["mkt_total"],
                ])
            print(f"\n  {len(mkt_picks)} pick(s) con valor real:\n")
            print(tab(rows, ["Juego","Hora","✅ PICK","Edge","Spread Mod vs Mkt","Total Mkt"],
                      fmt="rounded_outline"))
            for p in mkt_picks:
                print(f"  ↳ {p['game']}: {p['reason']}")
            print()

            # ── Picks adicionales con valor (alt_picks_full) ─────────────────
            # Spread y total NO son mutuamente excluyentes — se muestran ambos
            # cuando el pick secundario también supera el umbral mínimo.
            alt_full_rows = []
            for p in mkt_picks:
                for ac in p.get("alt_picks_full", []):
                    wp_str = (f"{ac['win_prob']:.1f}%"
                              if ac.get("win_prob") else "—")
                    alt_full_rows.append([
                        p["game"],
                        p["time"] or "—",
                        ac["pick"],
                        ac["edge_str"],
                        ac.get("reason", ""),
                        wp_str,
                    ])
            if alt_full_rows:
                print(f"  📌 También con valor ({len(alt_full_rows)} pick adicional):\n")
                print(tab(alt_full_rows,
                          ["Juego","Hora","Pick adicional","Edge","Razón","Win%"],
                          fmt="rounded_outline"))
                print()

        if solo_picks:
            solo_picks.sort(key=lambda x: _parse_time_sort(x["time"]))
            rows2 = []
            for p in solo_picks:
                rows2.append([
                    p["game"],
                    p["time"] or "—",
                    p["pick"],
                    p["odds"],
                    p["wp"],
                    p["model_spread"],
                ])
            print(f"\n  {len(solo_picks)} pick(s) sin líneas de mercado:\n")
            print(tab(rows2, ["Juego","Hora","Pick","ML Modelo","Win%","Spread Mod"],
                      fmt="rounded_outline"))
            print(f"\n  💡 Tip: --set-lines para agregar odds y encontrar valor real.\n")

    if not has_market:
        print(f"  📝 Para análisis completo:  python3 bsn.py --set-lines\n")

    print(f"{'═'*72}\n")

    # Auto-guardar picks del modelo en historial
    if picks:
        _model_picks_save_today_bsn(picks)

    return picks


# ──────────────────────────────────────────────────────
# MODEL PICKS HISTORY — auto-save + auto-grade (BSN)
# ──────────────────────────────────────────────────────

def _load_model_picks():
    if os.path.exists(MODEL_PICKS_FILE):
        with open(MODEL_PICKS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def _save_model_picks(picks):
    with open(MODEL_PICKS_FILE, "w", encoding="utf-8") as f:
        json.dump(picks, f, indent=2, ensure_ascii=False)

def _model_picks_save_today_bsn(picks):
    """
    Persiste los picks del modelo para TARGET_DATE en MODEL_PICKS_FILE.
    picks: lista de dicts de _find_value_picks().
    """
    history = [p for p in _load_model_picks() if p.get("date") != TARGET_DATE]
    for p in picks:
        history.append({
            "date":    TARGET_DATE,
            "game":    p.get("game", "—"),
            "pick":    p.get("pick", "—"),
            "odds":    p.get("odds", "—"),
            "modelo":  p.get("wp", "—"),
            "edge":    p.get("edge_str", "—"),
            "reason":  p.get("reason", "—"),
            "result":  None,
            "actual":  None,
        })
    _save_model_picks(history)
    print(f"  💾 {len(picks)} picks del modelo guardados en historial.")


def _fetch_bsn_scores(date_str):
    """
    Intenta obtener scores BSN desde bsnpr.com o realgm.
    Retorna lista de {team1, team2, score1, score2, status} o [].
    BSN no tiene API pública confiable, así que pedimos al usuario.
    """
    try:
        # Intentar RealGM BSN scoreboard
        import urllib.request
        url = f"https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/scores/{date_str}"
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            html = resp.read().decode("utf-8", errors="ignore")

        scores = []
        # Buscar patrones de score en el HTML
        import re as _re
        # RealGM usa tabla con equipos y scores
        # Buscamos patrones: TEAM1 XXX - YYY TEAM2
        pattern = _re.compile(
            r'<td[^>]*>([A-Za-z\s]+?)</td>\s*<td[^>]*>(\d+)</td>\s*<td[^>]*>(\d+)</td>\s*<td[^>]*>([A-Za-z\s]+?)</td>',
            _re.DOTALL
        )
        for m in pattern.finditer(html):
            t1, s1, s2, t2 = m.group(1).strip(), int(m.group(2)), int(m.group(3)), m.group(4).strip()
            team1 = norm_team(t1); team2 = norm_team(t2)
            if team1 and team2 and (s1 > 0 or s2 > 0):
                scores.append({"team1":team1,"team2":team2,"score1":s1,"score2":s2,"status":"Final"})
        return scores
    except Exception:
        return []


def _grade_bsn_pick(pick_entry, scores, manual_scores=None):
    """
    Determina W/L/P para un pick BSN dado los scores reales.
    manual_scores: dict { "LEONES vs OSOS": (95, 88) } para entrada manual.
    """
    pick_str = pick_entry["pick"].upper().strip()
    game_str = pick_entry["game"].upper().strip()

    # Buscar score — primero en scores automáticos, luego en manuales
    game_score = None
    all_scores = list(scores)
    if manual_scores:
        for key, (s1, s2) in manual_scores.items():
            k = key.upper()
            parts = k.replace(" VS.", " VS ").replace(" @ ", " VS ").split(" VS ")
            if len(parts) == 2:
                all_scores.append({"team1": norm_team(parts[0].strip()) or parts[0].strip(),
                                   "team2": norm_team(parts[1].strip()) or parts[1].strip(),
                                   "score1": s1, "score2": s2, "status": "Final"})

    for s in all_scores:
        t1 = s["team1"]; t2 = s["team2"]
        if t1 in game_str or t2 in game_str:
            game_score = s; break

    if not game_score:
        return None, "juego no encontrado"

    s1, s2   = game_score["score1"], game_score["score2"]
    total    = s1 + s2
    t1, t2   = game_score["team1"], game_score["team2"]
    actual   = f"{t1} {s1} – {s2} {t2}"

    import re as _re

    # ── Totals ─────────────────────────────────────────
    if "OVER" in pick_str or pick_str.startswith("O "):
        m = _re.search(r"([\d.]+)", pick_str)
        line = float(m.group(1)) if m else 0
        if total > line: return "W", actual
        if total < line: return "L", actual
        return "P", actual

    if "UNDER" in pick_str or pick_str.startswith("U "):
        m = _re.search(r"([\d.]+)", pick_str)
        line = float(m.group(1)) if m else 0
        if total < line: return "W", actual
        if total > line: return "L", actual
        return "P", actual

    # ── Spread ─────────────────────────────────────────
    sp_m = _re.search(r"([+-])([\d.]+)\s*$", pick_str)
    if sp_m:
        spread_val = float(sp_m.group(2)) if sp_m.group(1) == "+" else -float(sp_m.group(2))
        # Identificar pick_team
        pick_team = None
        for t in BSN_TEAMS:
            if t in pick_str:
                pick_team = t; break
        # s1=team1(visitante), s2=team2(local)
        if pick_team and pick_team == t2:
            margin = s2 - s1   # local
        else:
            margin = s1 - s2   # visitante (default)
        cover = margin + spread_val
        if cover > 0: return "W", actual
        if cover < 0: return "L", actual
        return "P", actual

    # ── Moneyline ──────────────────────────────────────
    for t in BSN_TEAMS:
        if t in pick_str and "ML" in pick_str:
            if t == t1:  return ("W" if s1 > s2 else "L" if s1 < s2 else "P"), actual
            if t == t2:  return ("W" if s2 > s1 else "L" if s2 < s1 else "P"), actual

    return None, f"pick no reconocido: {pick_str}"


def cmd_grade_picks_bsn():
    """
    --grade-picks [DATE]
    Parsea picks desde 'Laboy BSN Picks {DATE}.html', pide scores manualmente,
    actualiza el HTML en-place con colores W/L/P + score, genera tarjeta de resumen.
    """
    try:
        gi    = sys.argv.index("--grade-picks")
        gdate = sys.argv[gi+1] if gi+1 < len(sys.argv) and not sys.argv[gi+1].startswith("-") else TARGET_DATE
    except (ValueError, IndexError):
        gdate = TARGET_DATE

    # Buscar archivo HTML local: Laboy BSN Picks {DATE}-{token}.html
    import glob as _glob
    html_filename = f"Laboy BSN Picks {gdate}-{_url_token(gdate)}.html"
    html_path = os.path.join(SCRIPT_DIR, html_filename)

    # Fallback: buscar por patrón glob por si el archivo tiene token diferente
    if not os.path.exists(html_path):
        matches = _glob.glob(os.path.join(SCRIPT_DIR, f"Laboy BSN Picks {gdate}*.html"))
        if matches:
            html_path = matches[0]
            html_filename = os.path.basename(html_path)

    if not os.path.exists(html_path):
        print(f"\n  ❌ Archivo no encontrado: {html_filename}")
        print(f"     Corre --export-html primero para generarlo.\n")
        return

    print(f"\n  📄 Leyendo picks desde: {html_filename}")

    from bs4 import BeautifulSoup
    with open(html_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    pick_cards = soup.find_all("div", class_="pick-card")
    print(f"  🃏 {len(pick_cards)} cards encontrados.\n")

    if not pick_cards:
        print("  ℹ️  No se encontraron pick-cards en el HTML.\n")
        return

    # Extraer picks desde HTML
    picks_raw = []
    for card in pick_cards:
        game_label_div = card.find("div", class_="game-label")
        pick_label_div = card.find("div", class_="pick-label")
        if not game_label_div or not pick_label_div:
            continue
        # game_label: solo texto del div (sin ícono)
        game_label = game_label_div.get_text(strip=True)
        # pick_label: solo texto directo del div, SIN el odds-badge span hijo
        odds_badge = card.find("span", class_="odds-badge")
        odds_str   = odds_badge.get_text(strip=True) if odds_badge else "—"
        # NavigableString: nodos de texto puro (name=None), excluye tags hijo
        pick_label = "".join(
            str(node) for node in pick_label_div.children
            if getattr(node, "name", None) is None   # NavigableString, no Tag
        ).strip()
        # Fallback si algo falla: quitar odds_str del final
        if not pick_label:
            raw = pick_label_div.get_text(strip=True)
            pick_label = raw.replace(odds_str, "").strip() if odds_str != "—" else raw
        picks_raw.append({
            "card": card,
            "game": game_label,
            "pick": pick_label,
            "odds": odds_str,
            "pick_label_div": pick_label_div,
        })

    # Pedir scores — UNA sola vez por juego único
    print(f"  📝 Ingresa el score de cada juego (formato: 95-88, Visit-Local):\n")
    scores_map = {}
    seen_games = []
    for p in picks_raw:
        if p["game"] not in seen_games:
            seen_games.append(p["game"])
    for game_key in seen_games:
        val = input(f"  {game_key} — score (Enter para omitir): ").strip()
        if val and "-" in val:
            try:
                parts = val.split("-")
                s1, s2 = int(parts[0].strip()), int(parts[1].strip())
                scores_map[game_key] = (s1, s2)
            except ValueError:
                print(f"    ⚠️  Formato inválido, omitido.")
    print()

    picks_with_results = []
    rows = []

    for p in picks_raw:
        card = p["card"]
        game_label = p["game"]
        pick_label = p["pick"]
        odds_str   = p["odds"]
        pick_label_div = p["pick_label_div"]

        try:
            odds_v = int(odds_str.replace("+",""))
        except:
            odds_v = -110

        score_tuple = scores_map.get(game_label)
        result = "P"
        score_str = "—"
        color = "#94a3b8"

        if score_tuple:
            s1, s2 = score_tuple  # s1=visitante, s2=local
            total = s1 + s2
            score_str = f"{s1}-{s2}"
            pick_upper = pick_label.upper()

            # Detectar qué tipo de pick (BSN: totales, spread, ML)
            if "OVER" in pick_upper or re.match(r"^O\s*[\d.]+", pick_upper):
                # Extraer número de la línea (buscar cualquier número en el pick)
                m = re.search(r"([\d.]+)", pick_upper)
                line = float(m.group(1)) if m else 0
                if total > line:   result, color = "W", "#22c55e"
                elif total < line: result, color = "L", "#ef4444"
                else:              result, color = "P", "#94a3b8"

            elif "UNDER" in pick_upper or re.match(r"^U\s*[\d.]+", pick_upper):
                # Extraer número de la línea (buscar cualquier número en el pick)
                m = re.search(r"([\d.]+)", pick_upper)
                line = float(m.group(1)) if m else 0
                if total < line:   result, color = "W", "#22c55e"
                elif total > line: result, color = "L", "#ef4444"
                else:              result, color = "P", "#94a3b8"

            elif re.search(r"[+-][\d.]+\s*$", pick_upper):
                # Spread pick — e.g. "PIRATAS +8.5" o "CAPITANES -3.5"
                sp_m = re.search(r"([+-])([\d.]+)\s*$", pick_upper)
                spread_pts = float(sp_m.group(2)) if sp_m else 0
                spread_sign = sp_m.group(1) if sp_m else "+"
                spread_val  = spread_pts if spread_sign == "+" else -spread_pts

                game_teams = [t.strip() for t in re.split(r" vs\.? | @ | - ", game_label, flags=re.IGNORECASE)]
                visit_team = game_teams[0] if game_teams else ""
                local_team = game_teams[1] if len(game_teams) > 1 else ""

                pick_team = None
                for t in BSN_TEAMS:
                    if t.upper() in pick_upper:
                        pick_team = t; break
                if not pick_team:
                    pick_team = visit_team

                # s1=visitante, s2=local — obtener diferencia desde perspectiva del pick_team
                if pick_team and local_team and pick_team.upper() == local_team.upper():
                    margin = s2 - s1  # local: positivo si ganó
                else:
                    margin = s1 - s2  # visitante (default): positivo si ganó

                cover_margin = margin + spread_val
                if cover_margin > 0:   result, color = "W", "#22c55e"
                elif cover_margin < 0: result, color = "L", "#ef4444"
                else:                  result, color = "P", "#94a3b8"

            else:
                # Moneyline — detectar equipo en el pick
                game_teams = [t.strip() for t in re.split(r" vs\.? | @ | - ", game_label, flags=re.IGNORECASE)]
                visit_team = game_teams[0] if game_teams else ""
                local_team = game_teams[1] if len(game_teams) > 1 else ""

                pick_team = None
                for t in BSN_TEAMS:
                    if t.upper() in pick_upper:
                        pick_team = t; break
                if not pick_team:
                    pick_team = visit_team

                # Determinar ganador: s1=visit, s2=local
                if pick_team and local_team and pick_team.upper() == local_team.upper():
                    pick_won = s2 > s1
                else:
                    pick_won = s1 > s2  # default visitante

                if s1 == s2:          result, color = "P", "#94a3b8"
                elif pick_won:        result, color = "W", "#22c55e"
                else:                 result, color = "L", "#ef4444"

        # Actualizar card en HTML
        if result == "W":
            badge_html = '<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✅ WIN</span>'
            card_bg = "background:linear-gradient(135deg,#0d1f14 0%,#222222 60%)"
        elif result == "L":
            badge_html = '<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">❌ LOSS</span>'
            card_bg = "background:linear-gradient(135deg,#1f0d0d 0%,#222222 60%)"
        else:
            badge_html = '<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
            card_bg = "background:linear-gradient(135deg,#14161a 0%,#222222 60%)"

        badge_div = soup.new_tag("div", style="display:flex;justify-content:flex-end;margin-bottom:10px")
        badge_div.append(BeautifulSoup(badge_html, "html.parser"))

        score_div = soup.new_tag("div", style="font-size:0.75rem;color:#94a3b8;margin-top:4px")
        score_div.string = f"Score: {score_str}"

        # Garantizar que el color de la línea refleje el resultado final
        if result == "W":   color = "#22c55e"
        elif result == "L": color = "#ef4444"
        elif result == "P": color = "#94a3b8"
        card["style"] = f"border-left:4px solid {color};{card_bg}"
        card.insert(0, badge_div)
        pick_label_div.insert_after(score_div)

        picks_with_results.append({
            "game": game_label, "pick": pick_label, "odds": odds_str,
            "modelo": "—", "edge": "—", "result": result,
            "score": score_str, "color": color,
        })

        res_sym = {"W":"✅","L":"❌","P":"—"}.get(result,"⏳")
        rows.append([pick_label, odds_str, res_sym + result, score_str])

    # Guardar HTML actualizado
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(str(soup))

    print(tab(rows, ["Pick","Odds","Resultado","Score"], fmt="simple"))
    print()

    w  = sum(1 for p in picks_with_results if p["result"]=="W")
    l  = sum(1 for p in picks_with_results if p["result"]=="L")
    pu = sum(1 for p in picks_with_results if p["result"]=="P")
    tg = w + l + pu
    wp = f"{w/tg*100:.0f}%" if tg else "—"
    print(f"  📊 {gdate}: {w}W / {l}L / {pu}P  →  Win% {wp}\n")

    # Exportar tarjeta de resumen diario
    card_path = export_daily_picks_card_bsn(gdate, picks_with_results)
    if card_path:
        print(f"  📄 Card: {os.path.basename(card_path)}")

    # ── Publicar en GitHub Pages si se pidió ─────────────────────────────────
    if PUBLISH_MODE:
        to_publish = []
        if os.path.isfile(html_path):  to_publish.append(html_path)
        if card_path and os.path.isfile(card_path): to_publish.append(card_path)
        if to_publish:
            cmd_publish(to_publish)
        else:
            print("  ⚠️  No hay HTMLs para publicar.\n")

    print()


def export_model_calibration_html_bsn():
    """Genera 'Laboy Model Record - BSN.html' con historial completo + calibración."""
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    history = _load_model_picks()
    graded  = [p for p in history if p.get("result") in ("W","L","P")]

    total  = len(graded)
    wins   = sum(1 for p in graded if p["result"]=="W")
    losses = sum(1 for p in graded if p["result"]=="L")
    pushes = sum(1 for p in graded if p["result"]=="P")
    win_pct = wins / (wins + losses) if (wins + losses) else 0

    def _pnl(p):
        try:
            o_str = str(p.get("odds","")).replace("+","")
            o     = float(o_str) if o_str.lstrip("-").isdigit() else -110
            r     = p["result"]
            if r == "W": return (100/abs(o) if o < 0 else o/100)
            if r == "L": return -1.0
            return 0.0
        except: return 0.0
    roi_units = sum(_pnl(p) for p in graded)

    # Edge buckets (usando spread en puntos para BSN)
    buckets = {"0-2 pts":[], "2-4 pts":[], "4-6 pts":[], "6+ pts":[]}
    for p in graded:
        try:
            e = abs(float(str(p.get("edge","0")).replace(" pts","").replace("+","")))
            if e < 2:   buckets["0-2 pts"].append(p)
            elif e < 4: buckets["2-4 pts"].append(p)
            elif e < 6: buckets["4-6 pts"].append(p)
            else:        buckets["6+ pts"].append(p)
        except: pass

    def _bucket_row(label, items):
        if not items: return ""
        w  = sum(1 for x in items if x["result"]=="W")
        l  = sum(1 for x in items if x["result"]=="L")
        wp = f"{w/(w+l)*100:.0f}%" if (w+l) else "—"
        col = "#22c55e" if (w+l) and w/(w+l) > 0.5 else "#ef4444"
        return f"""<tr>
          <td style="padding:8px 12px;color:var(--muted)">{esc(label)}</td>
          <td style="padding:8px 12px;text-align:center">{len(items)}</td>
          <td style="padding:8px 12px;text-align:center">{w}W / {l}L</td>
          <td style="padding:8px 12px;text-align:center;color:{col};font-weight:700">{wp}</td>
        </tr>"""

    bucket_rows = "".join(_bucket_row(k, v) for k, v in buckets.items())

    pick_rows = ""
    for p in sorted(history, key=lambda x: x.get("date",""), reverse=True):
        res = p.get("result")
        if res == "W":   res_html = '<span style="color:#22c55e;font-weight:700">✅ W</span>'
        elif res == "L": res_html = '<span style="color:#ef4444;font-weight:700">❌ L</span>'
        elif res == "P": res_html = '<span style="color:#94a3b8;font-weight:700">— P</span>'
        else:            res_html = '<span style="color:#f07820">⏳</span>'
        actual = esc(p.get("actual") or "—")
        pick_rows += f"""<tr>
          <td style="padding:7px 10px;color:var(--muted);font-size:0.8rem">{esc(p.get('date',''))}</td>
          <td style="padding:7px 10px;font-size:0.85rem">{esc(p.get('game',''))}</td>
          <td style="padding:7px 10px;font-weight:600">{esc(p.get('pick',''))}</td>
          <td style="padding:7px 10px;text-align:center;color:var(--accent)">{esc(str(p.get('odds','')))}
          <td style="padding:7px 10px;text-align:center">{esc(str(p.get('edge','')))}</td>
          <td style="padding:7px 10px;text-align:center">{res_html}</td>
          <td style="padding:7px 10px;color:var(--muted);font-size:0.8rem">{actual}</td>
        </tr>"""

    win_col  = "#22c55e" if win_pct > 0.5 else "#ef4444"
    roi_col  = "#22c55e" if roi_units >= 0 else "#ef4444"
    roi_sign = "+" if roi_units >= 0 else ""

    body = f"""
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:24px">
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">PICKS</div>
        <div style="font-size:1.8rem;font-weight:900">{total}</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">RECORD</div>
        <div style="font-size:1.3rem;font-weight:900">{wins}W / {losses}L / {pushes}P</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">WIN %</div>
        <div style="font-size:1.8rem;font-weight:900;color:{win_col}">{win_pct*100:.0f}%</div>
      </div>
      <div style="background:#1a1a1a;border-radius:12px;padding:16px;text-align:center">
        <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">ROI (u)</div>
        <div style="font-size:1.8rem;font-weight:900;color:{roi_col}">{roi_sign}{roi_units:.2f}u</div>
      </div>
    </div>

    <div class="section-title" style="margin-top:0">Calibración por Spread del Modelo</div>
    <div style="background:#1a1a1a;border-radius:12px;overflow:hidden;margin-bottom:24px">
      <table style="width:100%;border-collapse:collapse">
        <thead><tr style="border-bottom:1px solid #333">
          <th style="padding:10px 12px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">SPREAD</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">PICKS</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">W/L</th>
          <th style="padding:10px 12px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">WIN%</th>
        </tr></thead>
        <tbody>{bucket_rows}</tbody>
      </table>
    </div>

    <div class="section-title">Histórico de Picks</div>
    <div style="background:#1a1a1a;border-radius:12px;overflow:hidden;overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:0.85rem">
        <thead><tr style="border-bottom:1px solid #333">
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">FECHA</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">JUEGO</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">PICK</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">ML</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">SPREAD</th>
          <th style="padding:10px;text-align:center;color:var(--muted);font-size:0.75rem;letter-spacing:1px">RES</th>
          <th style="padding:10px;text-align:left;color:var(--muted);font-size:0.75rem;letter-spacing:1px">SCORE</th>
        </tr></thead>
        <tbody>{pick_rows}</tbody>
      </table>
    </div>"""

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")
    html = _bsn_html_wrap("Laboy Model Record — BSN", "BSN", dstr, yr, body)

    fpath = os.path.join(SCRIPT_DIR, "Laboy Model Record - BSN.html")
    with open(fpath, "w", encoding="utf-8") as f:
        f.write(html)
    return fpath


def export_daily_picks_card_bsn(date_str, picks_data):
    """
    Genera 'Laboy BSN Model Card {DATE}.html' — tarjeta diaria de picks BSN con resultados.
    picks_data: lista de dicts con {game, pick, odds, modelo, edge, result, score, color}
    Retorna path del HTML generado.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    if not picks_data:
        return None

    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
    except:
        dt = datetime.now()

    dstr = dt.strftime("%A, %B %d").upper()
    yr   = dt.strftime("%Y")

    w   = sum(1 for p in picks_data if p["result"]=="W")
    l   = sum(1 for p in picks_data if p["result"]=="L")
    pu  = sum(1 for p in picks_data if p["result"]=="P")
    tot = w + l + pu
    win_pct = f"{w/tot*100:.0f}%" if tot else "—"

    picks_html = ""
    for p in picks_data:
        result = p["result"]
        color  = p.get("color", "#94a3b8")

        if result == "W":
            result_badge = '<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✅ WIN</span>'
            card_bg = "background:linear-gradient(135deg,#0d1f14 0%,#222222 60%)"
        elif result == "L":
            result_badge = '<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">❌ LOSS</span>'
            card_bg = "background:linear-gradient(135deg,#1f0d0d 0%,#222222 60%)"
        else:
            result_badge = '<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
            card_bg = "background:linear-gradient(135deg,#14161a 0%,#222222 60%)"

        picks_html += f"""
        <div class="pick-card" style="border-left:4px solid {color};{card_bg}">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
            <div style="font-size:0.75rem;color:var(--muted)">{esc(p['game'][:36])}</div>
            {result_badge}
          </div>
          <div style="font-size:0.95rem;font-weight:700;margin-bottom:4px">{esc(p['pick'])}</div>
          <div style="font-size:0.75rem;color:var(--muted)">
            Odds {esc(p['odds'])} · Score: {esc(p['score'])}
          </div>
        </div>
        """

    body = f"""
    <div class="section-title">PICKS DEL MODELO — BSN</div>
    {picks_html}

    <div style="background:#1a1a1a;border-radius:12px;padding:16px;margin-top:24px;text-align:center">
      <div style="font-size:0.7rem;color:var(--muted);letter-spacing:1px;margin-bottom:6px">RESUMEN</div>
      <div style="font-size:1.3rem;font-weight:900">{w}W · {l}L · {pu}P  →  Win% {win_pct}</div>
    </div>
    """

    html = _bsn_html_wrap(f"Laboy BSN Model Card {date_str}", "BSN", dstr, yr, body)

    _tok = _url_token(date_str)
    html_file = f"Laboy BSN Model Card {date_str}-{_tok}.html"
    html_path = os.path.join(SCRIPT_DIR, html_file)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    return html_path


def export_record_card_bsn(date_str=None):
    """
    Genera 'Laboy BSN Record Card {DATE}.html' con los picks logueados para una fecha
    o all-time. Incluye balance acumulado por pick.
    Retorna path del HTML generado.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    log = _load_log()
    if not log:
        print("  ℹ️  No hay picks en el log. Usa --log para agregar.\n")
        return None

    if date_str:
        entries = [e for e in log if e.get("date") == date_str]
    else:
        entries = log

    label = date_str or "All-Time"

    try:
        if date_str:
            dt   = datetime.strptime(date_str, "%Y-%m-%d")
            dstr = dt.strftime("%A, %B %d").upper()
            yr   = dt.strftime("%Y")
        else:
            dstr = "ALL-TIME"
            yr   = datetime.now().strftime("%Y")
    except:
        dstr = "ALL-TIME"
        yr   = datetime.now().strftime("%Y")

    def _bsn_fmt_pick(p):
        """'O 183.5' → 'Over 183.5', 'U 176.5' → 'Under 176.5'"""
        p = str(p).strip()
        if re.match(r'^O\s+\d', p): return "Over "  + p[1:].strip()
        if re.match(r'^U\s+\d', p): return "Under " + p[1:].strip()
        return p

    def _bsn_pick_logo(pick_str, game_str, size=52):
        """Logo para el pick: over_under si es total, si no el equipo del pick."""
        ps = str(pick_str).strip()
        # Detecta: "O 190.5", "U 190.5", "OVER 190.5", "UNDER 190.5"
        if re.match(r'^(OVER|UNDER|[OU])\s+\d', ps, re.IGNORECASE):
            return bsn_logo_html("over_under", size)
        # Normalizar pick (sin acentos) para comparaciones robustas
        ps_norm = _strip_accents(ps).upper()
        # Buscar el equipo del juego que aparezca en el pick
        # (normaliza ambos lados para manejar ATLÉTICOS vs ATLETICOS)
        teams = re.split(r'\s+vs\.?\s+', game_str, flags=re.IGNORECASE)
        for t in teams:
            t = t.strip()
            if t and _strip_accents(t).upper() in ps_norm:
                return bsn_logo_html(t, size)
        # Fallback: primer token del pick (antes de número o tipo de apuesta)
        tokens = ps.split()
        pick_tokens = []
        for tok in tokens:
            if re.match(r'^[+\-]?\d', tok) or tok.upper() in ('ML','RL','SPR','SPREAD','O','U'):
                break
            pick_tokens.append(tok)
        team_guess = " ".join(pick_tokens) if pick_tokens else (tokens[0] if tokens else "")
        return bsn_logo_html(team_guess, size) if team_guess else ""

    running_balance = 0.0
    picks_html = ""
    w, l, p = 0, 0, 0

    for entry in entries:
        stake    = _entry_stake(entry)
        pnl      = entry.get("pnl")
        result   = entry.get("result") or "⏳"
        raw_pick = entry.get("pick", "")
        game_str = entry.get("game", "")
        odds_v   = entry.get("odds", 0)

        if result == "W":
            running_balance += pnl if pnl is not None else stake
            w += 1
        elif result == "L":
            running_balance -= stake
            l += 1
        elif result == "P":
            p += 1

        _bpnl = pnl
        if _bpnl is not None:
            bal_fmt = f"+${_bpnl:.2f}" if _bpnl >= 0 else f"-${abs(_bpnl):.2f}"
        elif result == "W":
            bal_fmt = f"+${stake:.2f}"
        elif result == "L":
            bal_fmt = f"-${stake:.2f}"
        else:
            bal_fmt = "—"
        odds_fmt     = (f"+{odds_v}" if odds_v > 0 else str(odds_v)) if odds_v else "—"
        logo_h       = _bsn_pick_logo(raw_pick, game_str, 52)
        pick_display = esc(_bsn_fmt_pick(raw_pick))
        game_display = esc(game_str)
        _rc_cls  = {"W":"win","L":"loss","P":"push"}.get(result,"pending")
        _rc_wm   = {"W":"W","L":"L","P":"—"}.get(result,"")
        _rc_bt   = {"W":"WIN","L":"LOSS","P":"PUSH"}.get(result,"PENDING")
        _book_b  = entry.get("book","")
        _meta_b  = " · ".join(x for x in [_book_b, f"Stake: ${stake:.2f}"] if x)

        picks_html += f"""
<div class="rc-pick {_rc_cls}">
  <div class="rc-row">
    {logo_h}
    <div class="rc-main">
      <div class="rc-pick-name">{pick_display}<span class="rc-odds">{esc(odds_fmt)}</span></div>
      <div class="rc-game">{game_display}</div>
      <div class="rc-meta">{esc(_meta_b)}</div>
    </div>
    <div class="rc-result-col">
      <span class="rc-badge {_rc_cls}">{_rc_bt}</span>
      <div class="rc-pnl {_rc_cls}">{esc(bal_fmt)}</div>
    </div>
  </div>
</div>"""

    total = w + l + p
    total_pnl   = sum(e.get("pnl", 0) for e in entries if e.get("result") in ("W","L","P"))
    total_wager = sum(_entry_stake(e) for e in entries if e.get("result") in ("W","L","P"))
    roi = (total_pnl / total_wager * 100) if total_wager > 0 else 0
    win_pct = f"{w/total*100:.0f}%" if total else "—"

    # Tile de Ganancia = P&L neto del período mostrado
    pnl_fmt = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    bal_fmt = pnl_fmt   # alias para el tile de abajo

    win_col = "#22c55e" if w >= l else "#ef4444"
    pnl_col = "#22c55e" if total_pnl >= 0 else "#ef4444"
    bal_col = pnl_col
    roi_col = "#22c55e" if roi >= 0 else "#ef4444"

    _bpnl_fmt = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    _DIAS_EN  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    _MESES_EN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    if date_str:
        try:
            from datetime import date as _date
            _dobj = _date.fromisoformat(date_str)
            _rc_date_lbl = f"{_DIAS_EN[_dobj.weekday()]}, {_MESES_EN[_dobj.month-1]} {_dobj.day}"
        except Exception:
            _rc_date_lbl = date_str
    else:
        _rc_date_lbl = "All-Time"
    _roi_str = f"{roi:+.1f}%"
    _RC_CSS = """<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',system-ui,sans-serif;background:#080c12;color:#e2e8f0}
.rc-title{text-align:center;padding:12px 0 24px}
.rc-sport-lbl{font-size:1.6rem;font-weight:800;letter-spacing:.03em;color:#fff;margin-bottom:5px}
.rc-date-full{font-size:.78rem;color:#64748b;letter-spacing:.08em;text-transform:uppercase}
.rc-pick{border-radius:14px;padding:14px 16px;margin-bottom:10px;position:relative;
  overflow:hidden;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07)}
.rc-pick.win{border-left:4px solid #22c55e;background:linear-gradient(135deg,rgba(34,197,94,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.loss{border-left:4px solid #ef4444;background:linear-gradient(135deg,rgba(239,68,68,.08) 0%,rgba(255,255,255,.03) 60%)}
.rc-pick.push{border-left:4px solid #94a3b8;background:rgba(255,255,255,.04)}
.rc-pick.pending{border-left:4px solid #f59e0b;background:rgba(255,255,255,.04)}
.rc-row{display:flex;align-items:center;gap:12px}
.rc-logo{flex-shrink:0}
.rc-main{flex:1;min-width:0}
.rc-pick-name{font-size:.95rem;font-weight:600;color:#f1f5f9;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-odds{font-size:.78rem;color:#64748b;margin-left:6px;font-weight:400}
.rc-game{font-size:.78rem;color:#64748b;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.rc-meta{font-size:.72rem;color:#475569;margin-top:4px;letter-spacing:.02em}
.rc-result-col{display:flex;flex-direction:column;align-items:flex-end;gap:5px;flex-shrink:0}
.rc-badge{font-size:.68rem;font-weight:700;letter-spacing:.1em;padding:3px 9px;border-radius:20px}
.rc-badge.win{background:rgba(34,197,94,.18);color:#22c55e}
.rc-badge.loss{background:rgba(239,68,68,.18);color:#ef4444}
.rc-badge.push{background:rgba(148,163,184,.12);color:#94a3b8}
.rc-badge.pending{background:rgba(245,158,11,.12);color:#f59e0b}
.rc-pnl{font-size:.9rem;font-weight:700}
.rc-pnl.win{color:#22c55e}
.rc-pnl.loss{color:#ef4444}
.rc-pnl.push{color:#94a3b8}
.rc-summary{margin-top:18px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);
  border-radius:14px;padding:18px}
.rc-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;text-align:center}
.rc-stat-lbl{font-size:.65rem;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:5px}
.rc-stat-val{font-size:1.15rem;font-weight:700;line-height:1}
</style>"""
    body = f"""{_RC_CSS}
<div class="rc-title"><div class="rc-sport-lbl">🏀 BSN</div><div class="rc-date-full">{_rc_date_lbl}</div></div>
{picks_html}
<div class="rc-summary">
  <div class="rc-stats">
    <div class="rc-stat">
      <div class="rc-stat-lbl">Record</div>
      <div class="rc-stat-val" style="color:{win_col}">{w}-{l}-{p}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">Win Rate</div>
      <div class="rc-stat-val" style="color:{win_col}">{win_pct}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">Profit/Loss</div>
      <div class="rc-stat-val" style="color:{pnl_col}">{_bpnl_fmt}</div>
    </div>
    <div class="rc-stat">
      <div class="rc-stat-lbl">ROI</div>
      <div class="rc-stat-val" style="color:{roi_col}">{_roi_str}</div>
    </div>
  </div>
</div>
"""

    html = _bsn_html_wrap(f"Laboy BSN Record Card {label}", "BSN", dstr, yr, body)

    _tok = _url_token(label)
    html_file = f"Laboy BSN Record Card {label}-{_tok}.html"
    html_path = os.path.join(SCRIPT_DIR, html_file)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    return html_path


def export_record_card_jpg_bsn(date_str=None):
    """
    Genera el record card como SVG → JPG via ImageMagick.
    Replica exactamente el diseño de export_season_card_bsn().
    """
    import shutil, subprocess, tempfile

    log = _load_log()
    if not log:
        print("  ℹ️  No hay picks en el log.\n")
        return None

    entries = [e for e in log if e.get("date") == date_str] if date_str else list(log)
    entries = sorted(entries, key=lambda e: e.get("date",""))
    if not entries:
        print(f"  ℹ️  No hay picks para {date_str}.\n")
        return None

    label   = date_str or "All-Time"
    season  = "2025-26"
    now_str = datetime.now().strftime("%B %d, %Y")

    # ── Stats ─────────────────────────────────────────────────────
    settled = [e for e in entries if e.get("result") in ("W","L","P")]
    pending = [e for e in entries if e.get("result") not in ("W","L","P")]
    w  = sum(1 for e in settled if e.get("result") == "W")
    l  = sum(1 for e in settled if e.get("result") == "L")
    pu = sum(1 for e in settled if e.get("result") == "P")
    graded      = len(settled)
    total_wager = sum(_entry_stake(e) for e in settled)
    total_pnl   = sum(e.get("pnl", 0) or 0 for e in settled)
    roi         = (total_pnl / total_wager * 100) if total_wager > 0 else 0
    win_pct_s   = f"{w/graded*100:.0f}%" if graded else "—"
    pnl_str     = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    roi_str     = f"{roi:+.1f}%"
    rec_str     = f"{w}-{l}" + (f"-{pu}" if pu else "")
    win_col     = "#22c55e" if w >= l else "#ef4444"
    pnl_col     = "#22c55e" if total_pnl >= 0 else "#ef4444"

    # ── Pick-row SVG helper ────────────────────────────────────────
    def _fmt_date(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%-m/%-d")
        except Exception:
            return d

    def _fmt_pick_bsn(p):
        p = str(p).strip()
        if re.match(r'^O\s+\d', p): return "Over " + p[1:].strip()
        if re.match(r'^U\s+\d', p): return "Under " + p[1:].strip()
        return p

    def _esc(s):
        return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

    # ── SVG dimensions ────────────────────────────────────────────
    CARD_W   = 680        # card width
    OUTER    = 20         # outer padding
    CPAD_H   = 12         # horizontal card inner padding
    CPAD_V   = 10         # vertical card inner padding
    ROW_H    = 50         # pick row height
    ROW_GAP  = 4          # gap between rows
    COL_GAP  = 8          # gap between 2 columns
    COL_W    = (CARD_W - CPAD_H*2 - COL_GAP) // 2

    n_graded = len([e for e in entries if e.get("result")])
    n_rows   = (len(entries) + 1) // 2

    HDR_H    = 72
    SLBL_H   = 30
    PICKS_H  = n_rows * (ROW_H + ROW_GAP) - ROW_GAP + CPAD_V
    DIV_H    = 16
    STAT_H   = 88
    FTR_H    = 32
    INNER_H  = HDR_H + SLBL_H + PICKS_H + DIV_H + STAT_H + FTR_H
    TOTAL_H  = OUTER*2 + INNER_H + 16  # extra bottom breathing room

    svg_parts = []
    svg_parts.append(f'''<svg xmlns="http://www.w3.org/2000/svg" width="{CARD_W + OUTER*2}" height="{TOTAL_H}">
<defs>
  <linearGradient id="hdrgrad" x1="0%" y1="0%" x2="100%" y2="100%">
    <stop offset="0%" stop-color="#0a0a0a"/>
    <stop offset="100%" stop-color="#111111"/>
  </linearGradient>
  <linearGradient id="wingrad" x1="0%" y1="0%" x2="100%" y2="0%">
    <stop offset="0%" stop-color="#091610"/>
    <stop offset="100%" stop-color="#0a0a0a"/>
  </linearGradient>
  <linearGradient id="lossgrad" x1="0%" y1="0%" x2="100%" y2="0%">
    <stop offset="0%" stop-color="#160909"/>
    <stop offset="100%" stop-color="#0a0a0a"/>
  </linearGradient>
  <clipPath id="card-clip">
    <rect x="{OUTER}" y="{OUTER}" width="{CARD_W}" height="{INNER_H}" rx="18"/>
  </clipPath>
</defs>

<!-- Outer background -->
<rect width="{CARD_W + OUTER*2}" height="{TOTAL_H}" fill="#000000"/>

<!-- Card container -->
<rect x="{OUTER}" y="{OUTER}" width="{CARD_W}" height="{INNER_H}" rx="18" fill="#111111" stroke="#222222" stroke-width="1"/>

<!-- Header gradient bg -->
<rect x="{OUTER}" y="{OUTER}" width="{CARD_W}" height="{HDR_H}" rx="0" fill="url(#hdrgrad)" clip-path="url(#card-clip)"/>
<rect x="{OUTER}" y="{OUTER}" width="{CARD_W}" height="{HDR_H}" fill="url(#hdrgrad)" clip-path="url(#card-clip)"/>
''')

    # ── HEADER ────────────────────────────────────────────────────
    cx = OUTER  # card x
    cy = OUTER  # card y

    # Logo image (BSN logo or orange square fallback)
    logo_b64 = _bsn_logo_b64()
    lx, ly = cx + 18, cy + (HDR_H - 38) // 2
    if logo_b64:
        svg_parts.append(f'<image x="{lx}" y="{ly}" width="38" height="38" href="{logo_b64}" style="filter:drop-shadow(0 0 4px #f0782040)"/>')
    else:
        svg_parts.append(f'<rect x="{lx}" y="{ly}" width="38" height="38" rx="8" fill="#f07820"/>')
        svg_parts.append(f'<text x="{lx+19}" y="{ly+25}" text-anchor="middle" font-family="DejaVu Sans" font-weight="bold" font-size="13" fill="white">LP</text>')

    # Title — logo + "Laboy Picks" only
    tx = lx + 52
    ty = cy + HDR_H//2 + 7
    svg_parts.append(f'<text x="{tx}" y="{ty}" font-family="DejaVu Sans" font-weight="bold" font-size="22" fill="#f1f5f9">Laboy Picks</text>')

    # Header bottom border
    svg_parts.append(f'<line x1="{cx}" y1="{cy+HDR_H}" x2="{cx+CARD_W}" y2="{cy+HDR_H}" stroke="#222222" stroke-width="1"/>')

    # ── SECTION LABEL ─────────────────────────────────────────────
    ly2 = cy + HDR_H + SLBL_H - 8
    svg_parts.append(f'<text x="{cx+CPAD_H}" y="{ly2}" font-family="DejaVu Sans" font-weight="bold" font-size="9" fill="#3d4a5c" letter-spacing="2">PICKS EJECUTADOS · {len(entries)} PICKS</text>')

    # ── PICK ROWS 2-COL ───────────────────────────────────────────
    gy = cy + HDR_H + SLBL_H  # grid top

    for i, entry in enumerate(entries):
        result   = (entry.get("result") or "").upper()
        game_raw = entry.get("game", "")
        pick_raw = _fmt_pick_bsn(entry.get("pick", ""))
        date_lbl = _fmt_date(entry.get("date",""))

        col_i = i % 2
        row_i = i // 2
        rx2 = cx + CPAD_H + col_i * (COL_W + COL_GAP)
        ry  = gy + row_i * (ROW_H + ROW_GAP)

        if result == "W":
            bar_c = "#22c55e"; badge_c = "#22c55e"; badge_bg = "#22c55e20"
            grad  = "url(#wingrad)";  badge_txt = "✓"
        elif result == "L":
            bar_c = "#ef4444"; badge_c = "#ef4444"; badge_bg = "#ef444420"
            grad  = "url(#lossgrad)"; badge_txt = "✗"
        elif result == "P":
            bar_c = "#94a3b8"; badge_c = "#94a3b8"; badge_bg = "#94a3b820"
            grad  = "#0d0d0d";        badge_txt = "P"
        else:
            bar_c = "#f07820"; badge_c = "#f07820"; badge_bg = "#f0782020"
            grad  = "#0d0d0d";        badge_txt = "⏳"

        # Row background
        svg_parts.append(f'<rect x="{rx2}" y="{ry}" width="{COL_W}" height="{ROW_H}" rx="7" fill="{grad}"/>')
        # Left border bar
        svg_parts.append(f'<rect x="{rx2}" y="{ry}" width="3" height="{ROW_H}" rx="2" fill="{bar_c}"/>')

        # Date
        svg_parts.append(f'<text x="{rx2+10}" y="{ry+16}" font-family="DejaVu Sans" font-weight="bold" font-size="10" fill="{bar_c}">{_esc(date_lbl)}</text>')

        # Game
        gmax = 26
        game_disp = _esc(game_raw[:gmax] if len(game_raw)>gmax else game_raw)
        svg_parts.append(f'<text x="{rx2+10}" y="{ry+30}" font-family="DejaVu Sans" font-size="9" fill="#3d4a5c">{game_disp}</text>')

        # Pick name
        pmax = 22
        pick_disp = _esc(pick_raw[:pmax] if len(pick_raw)>pmax else pick_raw)
        svg_parts.append(f'<text x="{rx2+10}" y="{ry+44}" font-family="DejaVu Sans" font-weight="bold" font-size="14" fill="#e2e8f0">{pick_disp}</text>')

        # Result circle (right)
        bx2 = rx2 + COL_W - 20
        by2 = ry + ROW_H // 2
        svg_parts.append(f'<circle cx="{bx2}" cy="{by2}" r="11" fill="{badge_bg}"/>')
        svg_parts.append(f'<text x="{bx2}" y="{by2+5}" text-anchor="middle" font-family="DejaVu Sans" font-weight="bold" font-size="12" fill="{badge_c}">{badge_txt}</text>')

    # ── DIVIDER ───────────────────────────────────────────────────
    div_y = gy + n_rows*(ROW_H+ROW_GAP) + 4
    svg_parts.append(f'<line x1="{cx+CPAD_H}" y1="{div_y}" x2="{cx+CARD_W-CPAD_H}" y2="{div_y}" stroke="#1a1f2e" stroke-width="1"/>')

    # ── STATS 3-COL ───────────────────────────────────────────────
    stat_y  = div_y + DIV_H
    STAT_GAP2 = 8
    STAT_W  = (CARD_W - CPAD_H*2 - STAT_GAP2*2) // 3
    stat_data = [
        (rec_str,   "RECORD",  "W · L",           win_col),
        (win_pct_s, "WIN %",   f"{w} de {graded}", win_col),
        (roi_str,   "ROI",     "temporada",        pnl_col),
    ]
    for si, (val, lbl, sub, col) in enumerate(stat_data):
        sx2 = cx + CPAD_H + si*(STAT_W+STAT_GAP2)
        svg_parts.append(f'<rect x="{sx2}" y="{stat_y}" width="{STAT_W}" height="{STAT_H-8}" rx="10" fill="#0d0d0d"/>')
        svg_parts.append(f'<text x="{sx2+STAT_W//2}" y="{stat_y+16}" text-anchor="middle" font-family="DejaVu Sans" font-weight="bold" font-size="9" fill="#3d4a5c" letter-spacing="1">{_esc(lbl)}</text>')
        # value — auto-size for long strings
        fsize = 26 if len(val) <= 6 else (22 if len(val) <= 8 else 18)
        svg_parts.append(f'<text x="{sx2+STAT_W//2}" y="{stat_y+46}" text-anchor="middle" font-family="DejaVu Sans" font-weight="bold" font-size="{fsize}" fill="{col}">{_esc(val)}</text>')
        svg_parts.append(f'<text x="{sx2+STAT_W//2}" y="{stat_y+STAT_H-16}" text-anchor="middle" font-family="DejaVu Sans" font-size="9" fill="#3d4a5c">{_esc(sub)}</text>')

    # ── FOOTER ────────────────────────────────────────────────────
    ftr_y = stat_y + STAT_H + 2
    svg_parts.append(f'<line x1="{cx+CPAD_H}" y1="{ftr_y}" x2="{cx+CARD_W-CPAD_H}" y2="{ftr_y}" stroke="#1a1f2e" stroke-width="1"/>')
    ftr_txt = f"LABOYWEBSITE · BSN {season} · {now_str.upper()}"
    svg_parts.append(f'<text x="{cx+CARD_W//2}" y="{ftr_y+20}" text-anchor="middle" font-family="DejaVu Sans" font-weight="bold" font-size="9" fill="#1a1f2e" letter-spacing="1">{_esc(ftr_txt)}</text>')

    svg_parts.append('</svg>')
    svg_str = "\n".join(svg_parts)

    # ── Write SVG & convert to JPG (multi-backend) ───────────────
    fname   = f"Laboy BSN Record Card {label}.jpg"
    fpath   = os.path.join(SCRIPT_DIR, fname)
    svg_tmp = os.path.join(SCRIPT_DIR, f"_rc_{label}.svg")

    with open(svg_tmp, "w", encoding="utf-8") as f:
        f.write(svg_str)

    converted = False

    # ── Backend 1: ImageMagick convert ────────────────────────────
    convert_bin = shutil.which("convert")
    if convert_bin and not converted:
        try:
            r = subprocess.run(
                [convert_bin, "-density", "150", "-quality", "96", svg_tmp, fpath],
                capture_output=True, text=True, timeout=30
            )
            if r.returncode == 0:
                converted = True
        except Exception:
            pass

    # ── Backend 2: rsvg-convert (brew install librsvg) ────────────
    rsvg_bin = shutil.which("rsvg-convert")
    if rsvg_bin and not converted:
        png_tmp = svg_tmp.replace(".svg", ".png")
        try:
            r = subprocess.run(
                [rsvg_bin, "-w", "720", "-o", png_tmp, svg_tmp],
                capture_output=True, text=True, timeout=30
            )
            if r.returncode == 0 and os.path.exists(png_tmp):
                # PNG → JPG via Pillow
                try:
                    from PIL import Image as _PILImage
                    _PILImage.open(png_tmp).convert("RGB").save(fpath, "JPEG", quality=96)
                    converted = True
                except Exception:
                    pass
                try: os.remove(png_tmp)
                except Exception: pass
        except Exception:
            pass

    # ── Backend 3: cairosvg Python package ────────────────────────
    if not converted:
        try:
            import cairosvg
            png_bytes = cairosvg.svg2png(bytestring=svg_str.encode(), scale=2)
            from PIL import Image as _PILImage
            import io as _io
            _PILImage.open(_io.BytesIO(png_bytes)).convert("RGB").save(fpath, "JPEG", quality=96)
            converted = True
        except ImportError:
            pass
        except Exception:
            pass

    # ── Cleanup SVG temp ──────────────────────────────────────────
    try: os.remove(svg_tmp)
    except Exception: pass

    if not converted:
        print("  ⚠️  No se pudo generar el JPG. Instala una de estas opciones:")
        print("       brew install imagemagick       ← recomendado (Mac)")
        print("       brew install librsvg           ← alternativa")
        print("       pip install cairosvg           ← alternativa Python")
        return None

    print(f"  🖼️  Record Card JPG: {fname}")
    return fpath


def show_stats(stats, injury_impact):
    dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    # Detectar si algún equipo tiene GP registrado
    has_gp = any(s.get("gp") is not None for s in stats.values())

    print(f"\n{'═'*68}")
    print(f"  LABOY PICKS — BSN TEAM STATS")
    if has_gp:
        print(f"  Blend dinámico por equipo  │  GP/(GP+{GP_REGRESSION}), pleno≥{BSN_FULL_CRED_GP}GP  │  Pythagorean exp={PYTH_EXP}")
    else:
        print(f"  Blend: default 40% 2026 + 60% 2025   │   Pythagorean exp={PYTH_EXP}")
    print(f"{'═'*68}\n")
    rows = []
    for team in sorted(stats.keys()):
        s   = stats[team]
        inj = injury_impact.get(team, 0.0)
        net = round(s["ortg"] - s["drtg"], 1)
        gp  = s.get("gp")
        bl  = s.get("blend", 0.40)
        blend_str = f"{gp}gp/{bl*100:.0f}%" if gp is not None else "default"
        rows.append([team, f"{s['ortg']:.1f}", f"{s['drtg']:.1f}",
                     f"{s['pace']:.1f}", f"{net:+.1f}",
                     f"-{inj:.2f}" if inj else "—",
                     blend_str])
    print(tab(rows, ["Team","OffRtg","DefRtg","Pace","Net±","InjAdj","Blend"], fmt="rounded_outline"))
    print()


def _print_ir(entries, injury_impact):
    """Muestra el injury report actual en tabla."""
    print(f"\n{'═'*68}")
    print(f"  LABOY PICKS — BSN INJURY REPORT")
    print(f"  Rate: 1=Out(70%)  2=Doubtful(75%)  3=Limited(80%)")
    print(f"{'═'*68}\n")
    if not entries:
        print("  IR vacío — no hay jugadores reportados.\n")
    else:
        rows = []
        for i, e in enumerate(sorted(entries, key=lambda x: (x["team"], x["player"]))):
            rate_lbl = {1:"OUT", 2:"DOUBTFUL", 3:"LIMITED"}.get(e["rate"], "?")
            rows.append([str(i), e["team"], e["player"], rate_lbl,
                         f"{e['ppg']:.1f}", f"{e['usg']*100:.1f}%", f"{e['impact']:.3f}"])
        print(tab(rows, ["#","Team","Player","Rate","PPG","USG%","Impact"], fmt="rounded_outline"))
        print()
        print("  Impacto total por equipo:")
        impact_rows = [(t, f"{v:.3f}") for t, v in sorted(injury_impact.items()) if v > 0]
        if impact_rows:
            print(tab(impact_rows, ["Team","Total Impact"], fmt="rounded_outline"))
        else:
            print("  (Ninguno)")
    print()


def show_ir(entries, injury_impact):
    """Alias para compatibilidad — solo muestra, sin interacción."""
    _print_ir(entries, injury_impact)


def cmd_ir(wb):
    """
    --ir  →  Muestra IR, luego ofrece agregar o remover jugadores interactivamente.
    """
    def _reload():
        entries = load_ir_entries(wb)
        impact  = load_injury_impact(wb)
        return entries, impact

    entries, impact = _reload()
    _print_ir(entries, impact)

    while True:
        print("  ¿Qué deseas hacer?")
        print("    [a] Agregar jugador al IR")
        print("    [r] Remover jugador del IR")
        print("    [q] Salir\n")
        accion = input("  Opción: ").strip().lower()

        # ── AGREGAR ─────────────────────────────────────────────────────
        if accion == "a":
            print()

            # Equipo
            team_raw = input("  Equipo (ej: GIGANTES): ").strip().upper()
            team     = norm_team(team_raw) or team_raw
            if not team:
                print("  ❌ Equipo inválido.\n"); continue

            # Jugador
            player = input("  Jugador (apellido o nombre completo): ").strip().upper()
            if not player:
                print("  ❌ Nombre vacío.\n"); continue

            # Rate
            print("  Rate:  1 = Out (70%)  |  2 = Doubtful (75%)  |  3 = Limited (80%)")
            try:
                rate = int(input("  Rate [1/2/3]: ").strip())
                assert rate in (1, 2, 3)
            except (ValueError, AssertionError):
                print("  ❌ Rate inválido — usa 1, 2 o 3.\n"); continue

            # PPG
            try:
                ppg = float(input("  PPG (puntos por partido): ").strip())
                assert ppg >= 0
            except (ValueError, AssertionError):
                print("  ❌ PPG inválido.\n"); continue

            # USG%
            try:
                usg_raw = float(input("  USG% (ej: 18.5 para 18.5%): ").strip())
                assert 0 < usg_raw <= 100
                usg = usg_raw / 100
            except (ValueError, AssertionError):
                print("  ❌ USG% inválido.\n"); continue

            # Calcular impacto
            rf     = RATE_FACTOR[rate]
            impact_val = round(ppg * usg * rf, 4)
            rate_lbl   = {1:"OUT", 2:"DOUBTFUL", 3:"LIMITED"}[rate]
            print(f"\n  📊 Impact = {ppg:.1f} PPG × {usg*100:.1f}% USG × {rf} ({rate_lbl})")
            print(f"         = {impact_val:.3f} pts afectados\n")

            # Escribir en Excel IR - BSN
            ws_ir = wb[IR_SHEET]
            found = False
            for row in ws_ir.iter_rows(min_row=2, max_row=ws_ir.max_row):
                rt = norm_team(str(row[1].value or ""))
                rp = str(row[2].value or "").strip().upper()
                if rt == team and (player in rp or rp in player):
                    row[3].value = rate
                    row[4].value = round(ppg, 2)
                    row[5].value = round(usg, 4)
                    row[6].value = impact_val
                    found = True
                    print(f"  ✏️  Actualizado: {rp} ({team})")
                    break

            if not found:
                next_row = ws_ir.max_row + 1
                ws_ir.cell(next_row, 2).value = team
                ws_ir.cell(next_row, 3).value = player
                ws_ir.cell(next_row, 4).value = rate
                ws_ir.cell(next_row, 5).value = round(ppg, 2)
                ws_ir.cell(next_row, 6).value = round(usg, 4)
                ws_ir.cell(next_row, 7).value = impact_val
                print(f"  ➕ Agregado: {player} ({team})")

            entries_upd = load_ir_entries(wb)
            _recompute_injury_impact(wb, entries_upd)
            _preserve_advanced_blend(wb)
            wb.save(EXCEL_FILE)
            print(f"  💾 Excel guardado.\n")

            entries, impact = _reload()
            _print_ir(entries, impact)

        # ── REMOVER ─────────────────────────────────────────────────────
        elif accion == "r":
            if not entries:
                print("  IR vacío, nada que remover.\n"); continue

            print()
            sorted_entries = sorted(entries, key=lambda x: (x["team"], x["player"]))
            for i, e in enumerate(sorted_entries):
                rate_lbl = {1:"OUT", 2:"DOUBTFUL", 3:"LIMITED"}.get(e["rate"], "?")
                print(f"  [{i}] {e['team']} — {e['player']} ({rate_lbl})")
            print()

            try:
                idx = int(input("  Número a remover: ").strip())
                assert 0 <= idx < len(sorted_entries)
            except (ValueError, AssertionError):
                print("  ❌ Número inválido.\n"); continue

            target = sorted_entries[idx]
            ws_ir  = wb[IR_SHEET]
            rows_to_delete = []
            for row in ws_ir.iter_rows(min_row=2, max_row=ws_ir.max_row):
                rt = norm_team(str(row[1].value or ""))
                rp = str(row[2].value or "").strip().upper()
                if rt == target["team"] and target["player"] in rp:
                    rows_to_delete.append(row[0].row)

            for r in sorted(rows_to_delete, reverse=True):
                ws_ir.delete_rows(r)

            entries_upd = load_ir_entries(wb)
            _recompute_injury_impact(wb, entries_upd)
            _preserve_advanced_blend(wb)
            wb.save(EXCEL_FILE)
            print(f"  🗑️  {target['player']} ({target['team']}) removido del IR.")
            print(f"  💾 Excel guardado.\n")

            entries, impact = _reload()
            _print_ir(entries, impact)

        # ── SALIR ────────────────────────────────────────────────────────
        elif accion in ("q", ""):
            print()
            break
        else:
            print("  Opción inválida — usa a, r o q.\n")


def show_schedule(wb, days_ahead=14):
    today   = date.today()
    cutoff  = today + timedelta(days=days_ahead)
    games   = {}  # key=(date_str, team1, team2) → (date_str, matchup, time_str, venue_s)

    # ── 1. Juegos desde Excel (BSN Lines sheet) ───────────────────────────
    try:
        ws = wb[BSN_LINES_SHEET]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if not row[3]: continue
            try:
                row_date = row[3].date() if isinstance(row[3], datetime) \
                           else datetime.strptime(str(row[3])[:10],"%Y-%m-%d").date()
            except: continue
            if row_date < today or row_date > cutoff: continue
            t1 = norm_team(str(row[5])) if row[5] else "?"
            t2 = norm_team(str(row[6])) if row[6] else "?"
            # Hora en columna H (índice 7) si existe
            hora = str(row[7]).strip() if len(row) > 7 and row[7] else "—"
            venue_s = HOME_VENUES.get(t2, "—").split(",")[0]
            key = (str(row_date), t1, t2)
            games[key] = (str(row_date), f"{t1} @ {t2}", hora, venue_s)
    except Exception:
        pass

    # ── 2. Juegos manuales (manual_games.json) ───────────────────────────
    manual = _load_manual_games()
    for entry in manual:
        try:
            edate = entry.get("date","")
            row_date = datetime.strptime(edate[:10], "%Y-%m-%d").date()
        except: continue
        if row_date < today or row_date > cutoff: continue
        t1 = entry.get("team1") or entry.get("visit") or "?"
        t2 = entry.get("team2") or entry.get("local") or "?"
        hora = entry.get("time","—")
        venue_s = HOME_VENUES.get(t2, "—").split(",")[0]
        key = (edate[:10], t1, t2)
        if key not in games:  # no duplicar
            games[key] = (edate[:10], f"{t1} @ {t2}", hora, venue_s)

    if not games:
        print(f"\n  No hay juegos BSN en los próximos {days_ahead} días.")
        print(f"  Tip: Agrega juegos con: python3 bsn.py --add-game VISIT LOCAL 'HH:MM PM'\n")
        return

    print(f"\n  {'─'*62}")
    print(f"  PRÓXIMOS {days_ahead} DÍAS — BSN SCHEDULE  ({len(games)} juego(s))")
    print(f"  {'─'*62}")
    for key in sorted(games.keys()):
        dt_s, matchup, hora, venue_s = games[key]
        d = datetime.strptime(dt_s, "%Y-%m-%d")
        print(f"  {d.strftime('%a %b %d'):12}  {matchup:<30}  {hora:<10}  {venue_s}")
    print()

# ──────────────────────────────────────────────────────
# REALGM SCRAPING
# ──────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}

def _fetch_html_curl(url, timeout=20):
    """Fallback: usa curl del sistema para evadir bloqueos de requests/urllib."""
    import subprocess
    try:
        result = subprocess.run(
            [
                "curl", "-s", "-L",
                "--max-time", str(timeout),
                "-H", "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "-H", "Accept-Language: en-US,en;q=0.9,es;q=0.8",
                "-H", "Sec-Fetch-Dest: document",
                "-H", "Sec-Fetch-Mode: navigate",
                "-H", "Sec-Fetch-Site: none",
                "--compressed",
                "--write-out", "\n__HTTP_STATUS__%{http_code}",
                url,
            ],
            capture_output=True, text=True, timeout=timeout + 5
        )
        out = result.stdout
        if "__HTTP_STATUS__" in out:
            body, status_str = out.rsplit("__HTTP_STATUS__", 1)
            try:
                status = int(status_str.strip())
                if status == 200 and len(body) > 200:
                    return body
            except ValueError:
                pass
        return None
    except Exception:
        return None


def _fetch_html(url, timeout=20, silent=False):
    """Fetch URL con requests; si da 403 reintenta con curl."""
    try:
        session = requests.Session()
        session.headers.update(HEADERS)
        r = session.get(url, timeout=timeout, allow_redirects=True)
        if r.status_code == 200:
            return r.text
        if r.status_code == 403:
            # Retry con curl (user-agent y stack diferente)
            html = _fetch_html_curl(url, timeout)
            if html:
                return html
        r.raise_for_status()
        return r.text
    except requests.exceptions.RequestException as e:
        # Último intento con curl
        html = _fetch_html_curl(url, timeout)
        if html:
            return html
        if not silent:
            print(f"  ⚠️  Error al acceder {url}: {e}")
        return None

def _parse_table(html, id_contains=None):
    """
    Parsea tablas HTML básicamente sin BeautifulSoup.
    Retorna lista de dicts {col_header: value}.
    """
    if not html: return []
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL|re.IGNORECASE)
    headers = []
    result  = []
    for row in rows:
        cells_h = re.findall(r'<th[^>]*>(.*?)</th>', row, re.DOTALL|re.IGNORECASE)
        cells_d = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL|re.IGNORECASE)
        clean   = lambda s: re.sub(r'<[^>]+>','',s).strip()
        if cells_h and not headers:
            headers = [clean(c) for c in cells_h]
        elif cells_d and headers:
            vals = [clean(c) for c in cells_d]
            if len(vals) >= len(headers):
                result.append(dict(zip(headers, vals[:len(headers)])))
    return result

def scrape_player_ppg(player_keyword, team_keyword=None):
    """
    Scrape PPG de RealGM BSN stats page.
    Retorna (player_full_name, team_short, ppg) o None.
    """
    url  = "https://basketball.realgm.com/international/league/62/Puerto-Rican-BSN/stats"
    html = _fetch_html(url)
    if not html: return None

    rows  = _parse_table(html)
    pk    = player_keyword.upper()
    tk    = team_keyword.upper() if team_keyword else None

    best  = None
    for row in rows:
        # Columnas típicas: Player, Team, GP, MIN, PTS, FGM, ...
        name = row.get("Player","").upper()
        team = row.get("Team","").upper()
        pts  = row.get("PTS","") or row.get("PPG","") or row.get("Points","")

        # Normalizar nombre equipo
        team_n = REALGM_NAME_MAP.get(team.lower(), team)

        if pk not in name: continue
        if tk and tk not in name and tk not in team_n: continue

        try:
            ppg_val = float(pts)
            # Preferir coincidencia exacta
            if best is None or (tk and tk in team_n):
                best = (name, team_n, ppg_val)
        except: pass

    return best

def scrape_player_usg(player_keyword, team_keyword=None):
    """
    Scrape USG% de RealGM BSN advanced stats page.
    Retorna (player_full_name, usg_float) o None.
    """
    url  = ("https://basketball.realgm.com/international/league/62/"
            "Puerto-Rican-BSN/stats/2026/Advanced_Stats/Qualified/All/"
            "points/All/desc/1/Regular_Season")
    html = _fetch_html(url)
    if not html: return None

    rows = _parse_table(html)
    pk   = player_keyword.upper()
    tk   = team_keyword.upper() if team_keyword else None

    best = None
    for row in rows:
        name = row.get("Player","").upper()
        team = row.get("Team","").upper()
        usg  = row.get("USG%","") or row.get("USG","") or row.get("Usage%","") or row.get("Usage","")

        if pk not in name: continue
        if tk and tk not in name and tk not in REALGM_NAME_MAP.get(team.lower(), team): continue
        try:
            usg_val = float(usg.replace("%","")) / 100
            best    = (name, usg_val)
        except: pass

    return best

def scrape_team_advanced_stats():
    """
    Scrape stats avanzadas 2026 por equipo desde RealGM.
    Retorna dict {team_short: {ortg, drtg, pace}} o {}.
    """
    url  = ("https://basketball.realgm.com/international/league/62/"
            "Puerto-Rican-BSN/team-stats/2026/Advanced_Stats/Team_Totals")
    html = _fetch_html(url)
    if not html: return {}

    rows   = _parse_table(html)
    result = {}
    for row in rows:
        team_full = row.get("Team","") or row.get("Team Name","")
        if not team_full: continue
        team = REALGM_NAME_MAP.get(team_full.lower().strip())
        if not team: continue
        ortg = row.get("ORtg","") or row.get("OffRtg","") or row.get("Off Rtg","")
        drtg = row.get("DRtg","") or row.get("DefRtg","") or row.get("Def Rtg","")
        pace = row.get("Pace","")
        try:
            result[team] = {
                "ortg_2026": float(ortg) if ortg else None,
                "drtg_2026": float(drtg) if drtg else None,
                "pace_2026": float(pace) if pace else None,
            }
        except: pass
    return result

# ──────────────────────────────────────────────────────
# PICK TRACKER (local)
# ──────────────────────────────────────────────────────

def _load_log():
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return []

def _save_log(log):
    with open(LOG_FILE,"w",encoding="utf-8") as f: json.dump(log, f, indent=2, ensure_ascii=False)

def _entry_stake(e):
    """Retorna el monto apostado. Soporta 'stake' (nuevo, $) y 'units' (legacy)."""
    return e.get("stake") if e.get("stake") is not None else e.get("units", 1)

def _fmt_stake(e):
    """Formatea apuesta como '$15.00' (nuevo) o '1u' (legacy)."""
    if e.get("stake") is not None:
        return f"${e['stake']:.2f}"
    return f"{e.get('units', 1)}u"

def _fmt_pnl(e):
    """Formatea P&L como '+$11.54' / '-$15.00' (nuevo) o '+0.77u' (legacy)."""
    pnl = e.get("pnl")
    if pnl is None: return "—"
    if e.get("stake") is not None:
        return (f"+${pnl:.2f}" if pnl >= 0 else f"-${abs(pnl):.2f}")
    return f"{pnl:+.2f}u"

def _fmt_odds(o):
    try:
        v = int(o)
        return f"+{v}" if v > 0 else str(v)
    except: return str(o)

def _parse_odds_input(s):
    """
    Convierte string de odds a entero. Maneja: '-110', '+105', '110', '-1.5' etc.
    Fix: strip el signo ANTES de int() para evitar doble negación.
    """
    s = s.strip()
    negative = s.startswith("-")
    digits = re.sub(r"[^0-9]", "", s)
    if not digits:
        raise ValueError(f"No se pudo parsear odds: '{s}'")
    val = int(digits)
    return -val if negative else val

def _lookup_log_odds(game_str, pick_str, book="BetMGM"):
    """
    BSN no tiene API de odds de mercado, retorna None siempre.
    El usuario debe ingresar los odds manualmente.
    """
    return None

def html_to_jpg(html_path, width=800, scale=2):
    """
    Convierte un HTML file a JPG usando playwright (Chromium headless).
    Retorna el path del JPG o None si playwright no está disponible.

    Instalación (una sola vez):
      pip install playwright --break-system-packages
      playwright install chromium
    """
    jpg_path = html_path.replace(".html", ".jpg")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  💡 Para generar JPG: pip install playwright --break-system-packages && playwright install chromium")
        return None
    try:
        # Leer HTML como texto — evita el bug de '#' en filename con file:// URLs
        with open(html_path, "r", encoding="utf-8") as _f:
            html_content = _f.read()

        with sync_playwright() as pw:
            browser = pw.chromium.launch(args=["--no-sandbox","--disable-dev-shm-usage"])
            # device_scale_factor=2 → doble resolución (Retina), texto e imágenes nítidos
            page    = browser.new_page(
                viewport={"width": width, "height": 900},
                device_scale_factor=scale
            )
            page.set_content(html_content, wait_until="domcontentloaded")
            # Espera a que carguen imágenes remotas (logos de equipos)
            try:
                page.wait_for_load_state("networkidle", timeout=6000)
            except Exception:
                pass
            # Ajusta la altura al contenido real
            height = page.evaluate("document.body.scrollHeight")
            page.set_viewport_size({"width": width, "height": max(height, 400)})
            # Screenshot full-page → PNG en memoria → JPG guardado
            png_bytes = page.screenshot(full_page=True)
            browser.close()

        from PIL import Image
        import io
        img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
        img.save(jpg_path, "JPEG", quality=95, optimize=True)
        return jpg_path
    except Exception as e:
        print(f"  ⚠️  html_to_jpg falló: {e}")
        return None

def _fetch_today_games_light(date_str):
    """
    Fetches BSN games for date_str without needing the Excel workbook.
    Tries manual → betsapi → realgm → flashscore in order.
    Returns list of dicts with {team1, team2, game_time}.
    Prints minimal status — designed to be called inline during --log / --log-parlay.
    """
    games = _get_manual_games(date_str)
    if games:
        return games
    print(f"  📡 Buscando juegos BSN para {date_str}...", end="", flush=True)
    for scraper, name in [
        (scrape_betsapi_schedule,   "betsapi"),
        (scrape_realgm_schedule,    "realgm"),
        (scrape_flashscore_schedule,"flashscore"),
    ]:
        try:
            g = scraper(date_str)
            if g:
                print(f" ✅ {len(g)} juego(s) [{name}]")
                return g
        except Exception:
            pass
    print(" ⚠️  sin datos web")
    return []


def _select_game_interactively(date_str, prompt_label="Juego"):
    """
    Muestra los juegos BSN del día y deja seleccionar por número.
    Devuelve string "TEAM1 vs. TEAM2" o None si el usuario escribe manualmente.
    """
    games = _fetch_today_games_light(date_str)
    if games:
        print(f"\n  📋 Juegos BSN  —  {date_str}")
        for i, g in enumerate(games, 1):
            t1  = g.get("team1","?")
            t2  = g.get("team2","?")
            tme = g.get("game_time","")
            tme_s = f"  {tme}" if tme else ""
            print(f"   {i}.  {t1} vs. {t2}{tme_s}")
        print(f"   M.  Ingresar manualmente\n")
        raw = input(f"  {prompt_label} [1-{len(games)} / M]: ").strip().upper()
        if raw not in ("M", ""):
            try:
                idx = int(raw) - 1
                if 0 <= idx < len(games):
                    g = games[idx]
                    return f"{g['team1']} vs. {g['team2']}"
            except ValueError:
                pass
        # Fall through to manual entry
        manual = input("  Juego (ej: LEONES vs. CRIOLLOS): ").strip().upper()
        return manual if manual else None
    else:
        # No schedule data — ask directly
        manual = input(f"  {prompt_label} (ej: LEONES vs. CRIOLLOS): ").strip().upper()
        return manual if manual else None


def cmd_log_parlay():
    """
    --log-parlay
    Registra un parlay (2+ legs) en el log con resultado ya conocido.
    Se guarda como una entrada especial con type='parlay'.
    """
    print(f"\n{'═'*60}")
    print(f"  LABOY BSN — REGISTRAR PARLAY")
    print(f"{'═'*60}\n")
    try:
        date_s_raw = input("  Fecha (YYYY-MM-DD / Enter = hoy): ").strip()
        date_s     = date_s_raw if date_s_raw else TARGET_DATE
        datetime.strptime(date_s, "%Y-%m-%d")

        n_legs_s = input("  ¿Cuántos legs? (ej: 2): ").strip()
        n_legs   = int(n_legs_s)
        legs     = []
        for i in range(n_legs):
            print(f"\n  ── Leg {i+1} ──────────────────────────────────────")
            game_raw = _select_game_interactively(date_s, prompt_label=f"Leg {i+1}")
            if not game_raw:
                print("  ❌ Juego requerido."); return
            pick_raw = input(f"  Pick  (ej: CRIOLLOS ML / O 195.5 / CRIOLLOS -4.5): ").strip().upper()
            legs.append({"game": game_raw, "pick": pick_raw})

        print()
        odds_s = input("  Odds del parlay (ej: +265): ").strip()
        odds_v = _parse_odds_input(odds_s)
        stake_s = input("  Apostado (ej: 10): ").strip()
        stake   = float(re.sub(r"[^\d.]", "", stake_s.split()[0]))
        book_raw = input("  Sportsbook [BetMGM]: ").strip()
        book     = book_raw if book_raw else "BetMGM"

        result_s = input("  Resultado (W / L / Enter = PENDING): ").strip().upper()
        pnl      = None
        if result_s == "W":
            if odds_v > 0:
                auto_pnl = round(stake * (odds_v / 100), 2)
            else:
                auto_pnl = round(stake * (100 / abs(odds_v)), 2)
            pnl_s = input(f"  P&L (Enter = +${auto_pnl:.2f}): ").strip()
            pnl   = float(pnl_s) if pnl_s else auto_pnl
        elif result_s == "L":
            pnl = -stake
        elif result_s == "":
            result_s = None   # PENDING — se gradeará después
        else:
            print("  ⚠️  Solo W, L, o Enter para PENDING."); return

    except (ValueError, EOFError) as _e:
        print(f"\n  ❌ Entrada inválida ({_e}).\n"); return

    log   = _load_log()
    entry = {
        "id":     len(log),
        "date":   date_s,
        "type":   "parlay",
        "legs":   legs,
        "odds":   odds_v,
        "stake":  stake,
        "book":   book,
        "result": result_s,
        "pnl":    pnl,
        # campos compat con resto del log
        "game":   " + ".join(f"{l['game']}" for l in legs),
        "pick":   "PARLAY",
        "analysis": "",
        "sport":  "BSN",
    }
    log.append(entry)
    _save_log(log)
    status = result_s if result_s else "PENDING"
    pnl_fmt = (f"+${pnl:.2f}" if pnl is not None and pnl >= 0
               else f"-${abs(pnl):.2f}" if pnl is not None else "—")
    print(f"\n  ✅ Parlay #{entry['id']} guardado: {date_s} │ {n_legs} legs │ {_fmt_odds(odds_v)} │ {status} │ {pnl_fmt}")
    print(f"  🃏 Usa: python3 bsn.py --export-log {entry['id']}  para generar la pick card PNG.\n")


def cmd_log_special():
    """
    --log-special
    Registra una jugada especial de mismo juego con 2 condiciones (ej: Gigantes ML + Under 200.5).
    Se guarda como type='special' y se muestra como una card con ambas condiciones.

    Ejemplo de jugada: Gigantes ML & Under 200.5 (una sola apuesta, dos condiciones)
    """
    print(f"\n{'═'*60}")
    print(f"  LABOY BSN — REGISTRAR JUGADA ESPECIAL")
    print(f"  (Moneyline + Total  /  Spread + Total  /  cualquier combo)")
    print(f"{'═'*60}\n")
    try:
        date_s_raw = input("  Fecha (YYYY-MM-DD / Enter = hoy): ").strip()
        date_s     = date_s_raw if date_s_raw else TARGET_DATE
        datetime.strptime(date_s, "%Y-%m-%d")

        game_raw = _select_game_interactively(date_s, prompt_label="Juego")
        if not game_raw:
            print("  ❌ Juego requerido."); return

        def _norm_game(g):
            for sep in [" VS. ", " vs. ", " VS ", " vs ", " @ ", " - "]:
                if sep in g:
                    p = g.split(sep, 1)
                    return f"{p[0].strip()} vs. {p[1].strip()}"
            return g
        game = _norm_game(game_raw)

        print(f"\n  Juego: {game}")
        print(f"  Ingresa las dos condiciones de la jugada:\n")
        cond1 = input("  Condición 1 (ej: GIGANTES ML): ").strip().upper()
        cond2 = input("  Condición 2 (ej: U 200.5 / UNDER 200.5): ").strip().upper()
        if not cond1 or not cond2:
            print("  ❌ Ambas condiciones son requeridas."); return

        print()
        odds_s  = input("  Odds de la jugada especial (ej: -135 / +110): ").strip()
        odds_v  = _parse_odds_input(odds_s)
        stake_s = input("  Apostado (ej: 15): ").strip()
        stake   = float(re.sub(r"[^\d.]", "", stake_s.split()[0]))
        book_raw = input("  Sportsbook [BetMGM]: ").strip()
        book     = book_raw if book_raw else "BetMGM"
        analysis = input("  Análisis (opcional / Enter para omitir):\n  > ").strip()

        result_s = input("\n  Resultado (W / L / P / Enter = PENDING): ").strip().upper()
        pnl = None
        if result_s == "W":
            auto_pnl = (round(stake * (odds_v / 100), 2) if odds_v > 0
                        else round(stake * (100 / abs(odds_v)), 2))
            pnl_s = input(f"  P&L (Enter = +${auto_pnl:.2f}): ").strip()
            pnl   = float(pnl_s) if pnl_s else auto_pnl
        elif result_s == "L":
            pnl = -stake
        elif result_s == "P":
            pnl = 0.0
        elif result_s == "":
            result_s = None
        else:
            print("  ⚠️  Solo W, L, P o Enter para PENDING."); return

    except (ValueError, EOFError) as _e:
        print(f"\n  ❌ Entrada inválida ({_e}).\n"); return

    log = _load_log()
    entry = {
        "id":         len(log),
        "date":       date_s,
        "type":       "special",                      # ← jugada especial
        "game":       game,
        "conditions": [cond1, cond2],                # dos condiciones del mismo juego
        "pick":       f"{cond1} & {cond2}",          # representación compacta
        "odds":       odds_v,
        "stake":      stake,
        "book":       book,
        "result":     result_s,
        "pnl":        pnl,
        "analysis":   analysis or "",
        "sport":      "BSN",
    }
    log.append(entry)
    _save_log(log)

    status  = result_s if result_s else "PENDING"
    pnl_fmt = (f"+${pnl:.2f}" if pnl is not None and pnl >= 0
               else f"-${abs(pnl):.2f}" if pnl is not None else "—")
    print(f"\n  ✅ Especial #{entry['id']} guardado:")
    print(f"     {game}  │  {cond1} & {cond2}")
    print(f"     {_fmt_odds(odds_v)}  │  ${stake:.2f}  │  {status}  │  {pnl_fmt}")
    print(f"  🃏 Usa: python3 bsn.py --export-log {entry['id']}  para generar la pick card.\n")


def export_season_card_bsn():
    """
    Genera 'Laboy BSN Season Card.html/.jpg' con todos los picks del log
    en formato 2 columnas + resumen de temporada al fondo.
    Soporta entradas normales y parlays (type='parlay').
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    log = _load_log()
    if not log:
        print("  ⚠️  No hay picks en el log.\n"); return None

    # Ordenar por fecha, excluir pendientes del card
    log = sorted(log, key=lambda e: e.get("date", ""))
    log_graded  = [e for e in log if e.get("result")]
    log_pending = [e for e in log if not e.get("result")]

    # Logo Laboy
    logo_src = _bsn_logo_b64()
    logo_html = (f'<img src="{logo_src}" alt="Laboy" width="38" height="38" '
                 f'style="object-fit:contain;filter:drop-shadow(0 0 6px #f0782040)">'
                 if logo_src else
                 '<span style="font-size:1.5rem">🏀</span>')

    # Calcular resumen (solo picks calificados)
    w = l = pu = 0
    total_pnl = 0.0
    total_wager = 0.0
    for e in log_graded:
        res = e.get("result")
        stake = float(e.get("stake", 0))
        pnl   = e.get("pnl") or 0
        if res == "W":   w += 1; total_pnl += pnl; total_wager += stake
        elif res == "L": l += 1; total_pnl += pnl; total_wager += stake
        elif res == "P": pu += 1; total_wager += stake

    graded = w + l + pu
    roi = (total_pnl / total_wager * 100) if total_wager > 0 else 0
    win_pct = f"{w/graded*100:.0f}%" if graded else "—"
    pnl_str = f"+${total_pnl:.2f}" if total_pnl >= 0 else f"-${abs(total_pnl):.2f}"
    roi_str = f"{roi:+.1f}%"
    win_col = "#22c55e" if w >= l else "#ef4444"
    pnl_col = "#22c55e" if total_pnl >= 0 else "#ef4444"
    pending_n = len(log_pending)

    # ── Generar filas de picks ──────────────────────────────────────────
    def _fmt_date(d):
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            return dt.strftime("%-m/%-d")
        except Exception:
            return d

    def _fmt_pick_bsn(p):
        p = str(p).strip()
        if re.match(r'^O\s+\d', p): return "Over " + p[1:].strip()
        if re.match(r'^U\s+\d', p): return "Under " + p[1:].strip()
        return p

    rows_html = ""
    for e in log_graded:
        res      = e.get("result") or ""
        ptype    = e.get("type", "pick")
        date_fmt = _fmt_date(e.get("date", ""))

        if res == "W":
            bcls = "win"; badge = "✓"; badge_cls = "w"
        elif res == "L":
            bcls = "loss"; badge = "✗"; badge_cls = "l"
        elif res == "P":
            bcls = "push"; badge = "P"; badge_cls = "push"
        else:
            bcls = "pending"; badge = "⏳"; badge_cls = "pend"

        if ptype == "parlay":
            legs    = e.get("legs", [])
            legs_html = ""
            for i, lg in enumerate(legs, 1):
                legs_html += f"""
                <div class="parlay-leg">
                  <div class="leg-num">L{i}</div>
                  <div class="leg-info">
                    <div class="leg-game">{esc(lg.get('game',''))}</div>
                    <div class="leg-pick">{esc(_fmt_pick_bsn(lg.get('pick','')))}</div>
                  </div>
                </div>"""
            odds_fmt = _fmt_odds(e.get("odds", 0))
            rows_html += f"""
            <div class="parlay-row full-width {bcls}">
              <div class="parlay-top">
                <div class="parlay-label">🔗 Parlay · {len(legs)} Legs</div>
                <div class="parlay-right">
                  <div class="pick-date" style="margin:0">{date_fmt}</div>
                  <div class="parlay-odds">{esc(odds_fmt)}</div>
                  <div class="result-badge {badge_cls}">{badge}</div>
                </div>
              </div>
              <div class="parlay-legs">{legs_html}</div>
            </div>"""
        elif ptype == "special":
            conditions = e.get("conditions", e.get("pick", "").split(" & "))
            odds_fmt   = _fmt_odds(e.get("odds", 0))
            game_disp  = e.get("game", "")
            cond_html  = ""
            icons      = ["🏀", "📊"]
            for ci, cond in enumerate(conditions):
                cond_html += f"""
                <div class="parlay-leg">
                  <div class="leg-num">{icons[ci] if ci < len(icons) else "·"}</div>
                  <div class="leg-info">
                    <div class="leg-game">{esc(game_disp)}</div>
                    <div class="leg-pick">{esc(_fmt_pick_bsn(cond))}</div>
                  </div>
                </div>"""
            rows_html += f"""
            <div class="parlay-row full-width {bcls}">
              <div class="parlay-top">
                <div class="parlay-label" style="color:#a78bfa">⚡ Especial · Mismo Juego</div>
                <div class="parlay-right">
                  <div class="pick-date" style="margin:0">{date_fmt}</div>
                  <div class="parlay-odds">{esc(odds_fmt)}</div>
                  <div class="result-badge {badge_cls}">{badge}</div>
                </div>
              </div>
              <div class="parlay-legs">{cond_html}</div>
            </div>"""
        else:
            raw_game = e.get("game", "")
            raw_pick = _fmt_pick_bsn(e.get("pick", ""))
            rows_html += f"""
            <div class="pick-row {bcls}">
              <div class="pick-date">{date_fmt}</div>
              <div class="pick-info">
                <div class="pick-game">{esc(raw_game)}</div>
                <div class="pick-name">{esc(raw_pick)}</div>
              </div>
              <div class="result-badge {badge_cls}">{badge}</div>
            </div>"""

    pending_note = (f'<div style="font-size:0.58rem;color:#f07820;font-weight:700;'
                    f'text-align:center;padding:4px 0">'
                    f'⏳ {pending_n} pick(s) pendiente(s) — excluido(s) del resumen</div>'
                    if pending_n else "")

    now_str = datetime.now().strftime("%B %d, %Y")
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#080a0f;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
        display:flex;justify-content:center;padding:28px 16px}}
  .card{{width:640px;background:#0f1117;border-radius:20px;overflow:hidden;border:1px solid #1a1f2e}}

  /* Header */
  .header{{background:linear-gradient(135deg,#0a0e18 0%,#0f1420 100%);
           padding:18px 22px 14px;border-bottom:1px solid #1a1f2e;
           display:flex;justify-content:space-between;align-items:center}}
  .header-left{{display:flex;align-items:center;gap:12px}}
  .header-title h1{{font-size:1.1rem;font-weight:900;color:#f1f5f9;letter-spacing:-0.3px}}
  .header-title h1 span{{color:#f07820}}
  .header-title p{{font-size:0.6rem;color:#3d4a5c;font-weight:700;
                   letter-spacing:1.5px;text-transform:uppercase;margin-top:3px}}
  .season-badge{{background:#f0782012;border:1px solid #f0782028;color:#f07820;
                 font-size:0.68rem;font-weight:800;padding:4px 11px;
                 border-radius:20px;letter-spacing:0.5px;white-space:nowrap}}

  /* Section label */
  .section-label{{font-size:0.58rem;font-weight:800;letter-spacing:2px;
                  text-transform:uppercase;color:#3d4a5c;padding:12px 18px 7px}}

  /* Two-column grid */
  .picks-grid{{display:grid;grid-template-columns:1fr 1fr;gap:4px;padding:0 12px 10px}}
  .full-width{{grid-column:1/-1}}

  /* Pick row */
  .pick-row{{display:flex;align-items:center;gap:8px;background:#090b10;
             border-radius:8px;padding:7px 9px;border-left:3px solid transparent}}
  .pick-row.win    {{border-left-color:#22c55e}}
  .pick-row.loss   {{border-left-color:#ef4444}}
  .pick-row.push   {{border-left-color:#94a3b8}}
  .pick-row.pending{{border-left-color:#f07820}}

  .pick-date{{font-size:0.58rem;font-weight:800;color:#3d4a5c;
              min-width:28px;white-space:nowrap;flex-shrink:0}}
  .pick-info{{flex:1;min-width:0}}
  .pick-game{{font-size:0.58rem;color:#3d4a5c;font-weight:600;
              white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
  .pick-name{{font-size:0.75rem;font-weight:800;color:#e2e8f0;
              white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
  .result-badge{{font-size:0.7rem;font-weight:900;width:20px;height:20px;
                 border-radius:50%;display:flex;align-items:center;
                 justify-content:center;flex-shrink:0}}
  .result-badge.w    {{background:#22c55e20;color:#22c55e}}
  .result-badge.l    {{background:#ef444420;color:#ef4444}}
  .result-badge.push {{background:#94a3b820;color:#94a3b8;font-size:0.5rem}}
  .result-badge.pend {{background:#f0782020;color:#f07820;font-size:0.48rem}}

  /* Parlay */
  .parlay-row{{background:#091610;border-radius:8px;border-left:3px solid #22c55e;
               padding:8px 10px;display:flex;flex-direction:column;gap:5px}}
  .parlay-row.loss{{background:#160909;border-left-color:#ef4444}}
  .parlay-top{{display:flex;justify-content:space-between;align-items:center}}
  .parlay-label{{font-size:0.58rem;font-weight:800;letter-spacing:1.5px;
                 color:#22c55e;text-transform:uppercase}}
  .parlay-row.loss .parlay-label{{color:#ef4444}}
  .parlay-right{{display:flex;align-items:center;gap:7px}}
  .parlay-odds{{font-size:0.63rem;font-weight:700;color:#22c55e;
                background:#22c55e15;padding:2px 7px;border-radius:6px}}
  .parlay-row.loss .parlay-odds{{color:#ef4444;background:#ef444415}}
  .parlay-legs{{display:flex;gap:6px}}
  .parlay-leg{{flex:1;background:#0a0d10;border-radius:6px;
               padding:5px 8px;display:flex;align-items:center;gap:6px}}
  .leg-num{{font-size:0.52rem;font-weight:800;color:#22c55e;
            background:#22c55e15;padding:1px 5px;border-radius:4px;flex-shrink:0}}
  .parlay-row.loss .leg-num{{color:#ef4444;background:#ef444415}}
  .leg-game{{font-size:0.56rem;color:#3d4a5c}}
  .leg-pick{{font-size:0.72rem;font-weight:800;color:#e2e8f0}}

  /* Divider */
  .divider{{height:1px;background:#1a1f2e;margin:8px 12px}}

  /* Record summary */
  .record-wrap{{padding:10px 12px 18px}}
  .record-title{{font-size:0.58rem;font-weight:800;letter-spacing:2px;color:#3d4a5c;
                 text-transform:uppercase;margin-bottom:9px;padding-left:2px}}
  .record-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:6px}}
  .stat-box{{background:#090b10;border-radius:10px;padding:10px 8px;text-align:center}}
  .stat-lbl{{font-size:0.52rem;font-weight:700;letter-spacing:1px;color:#3d4a5c;
             text-transform:uppercase;margin-bottom:4px}}
  .stat-val{{font-size:1.15rem;font-weight:900}}
  .stat-sub{{font-size:0.52rem;color:#3d4a5c;margin-top:2px}}
  .green{{color:#22c55e}} .red{{color:#ef4444}} .orange{{color:#f07820}}

  .footer{{text-align:center;padding:9px;font-size:0.55rem;font-weight:700;
           letter-spacing:1px;color:#1a1f2e;text-transform:uppercase;
           border-top:1px solid #1a1f2e}}
</style>
</head>
<body>
<div class="card">

  <div class="header">
    <div class="header-left">
      {logo_html}
      <div class="header-title">
        <h1>Laboy <span>BSN</span> Picks</h1>
        <p>Temporada 2025–26 · Historial completo</p>
      </div>
    </div>
    <div class="season-badge">2025–26</div>
  </div>

  <div class="section-label">📋 Picks Ejecutados · {len(log_graded)} picks</div>
  <div class="picks-grid">
    {rows_html}
  </div>

  {pending_note}
  <div class="divider"></div>

  <div class="record-wrap">
    <div class="record-title">📊 Resumen de Temporada</div>
    <div class="record-grid">
      <div class="stat-box">
        <div class="stat-lbl">Record</div>
        <div class="stat-val {'green' if w >= l else 'red'}">{w}-{l}{f'-{pu}' if pu else ''}</div>
        <div class="stat-sub">W · L{' · P' if pu else ''}</div>
      </div>
      <div class="stat-box">
        <div class="stat-lbl">Win %</div>
        <div class="stat-val {'green' if w >= l else 'red'}">{win_pct}</div>
        <div class="stat-sub">{w} de {graded}</div>
      </div>
      <div class="stat-box">
        <div class="stat-lbl">P&amp;L</div>
        <div class="stat-val {'green' if total_pnl >= 0 else 'red'}">{pnl_str}</div>
        <div class="stat-sub">neto</div>
      </div>
      <div class="stat-box">
        <div class="stat-lbl">ROI</div>
        <div class="stat-val {'green' if roi >= 0 else 'red'}">{roi_str}</div>
        <div class="stat-sub">temporada</div>
      </div>
    </div>
  </div>

  <div class="footer">laboywebsite · bsn 2025–26 · {now_str}</div>
</div>
</body>
</html>"""

    fname     = "Laboy BSN Season Card.html"
    html_path = os.path.join(SCRIPT_DIR, fname)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  📄 Season Card HTML: {fname}")

    jpg_path = html_to_jpg(html_path, width=680, scale=2)
    if jpg_path:
        print(f"  🖼️  Season Card JPG: {os.path.basename(jpg_path)}")
    return html_path, jpg_path


def cmd_serve(port=5001):
    """
    --serve [PORT]
    Levanta un servidor web local con formulario para loguear picks desde el celular.
    Expón con ngrok para acceso remoto:  ngrok http 5001
    """
    import socket
    import threading
    from http.server import BaseHTTPRequestHandler, HTTPServer
    from urllib.parse import parse_qs, urlparse, unquote_plus

    actual_port = port
    for i, a in enumerate(sys.argv):
        if a == "--serve" and i+1 < len(sys.argv):
            try: actual_port = int(sys.argv[i+1])
            except ValueError: pass

    # ── Local IP ──────────────────────────────────────────────────
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = "127.0.0.1"

    local_url = f"http://{local_ip}:{actual_port}"

    teams = sorted(BSN_TEAMS)
    team_opts = "\n".join(f'<option value="{t}">{t.title()}</option>' for t in teams)

    def _load(): return _load_log()
    def _save(lg): _save_log(lg)

    def _html_form(msg="", msg_type=""):
        log  = _load()
        rec  = [e for e in log if e.get("result")]
        pend = [e for e in log if not e.get("result")]
        w    = sum(1 for e in rec if e.get("result")=="W")
        l    = sum(1 for e in rec if e.get("result")=="L")
        total_pnl = sum(e.get("pnl",0) or 0 for e in rec)
        pnl_s = (f"+${total_pnl:.2f}" if total_pnl>=0 else f"-${abs(total_pnl):.2f}")
        today = date.today().strftime("%Y-%m-%d")

        # last 5 picks
        recent = list(reversed(log[-5:])) if log else []
        rows_html = ""
        for e in recent:
            res = e.get("result","⏳")
            c = {"W":"#22c55e","L":"#ef4444","P":"#94a3b8"}.get(res,"#f07820")
            rows_html += f'''<tr>
              <td style="color:#64748b;font-size:0.75rem">{e.get("date","")}</td>
              <td style="font-weight:600;font-size:0.85rem">{e.get("pick","")}</td>
              <td style="color:#94a3b8;font-size:0.75rem">{e.get("game","")[:20]}</td>
              <td style="color:{c};font-weight:700;text-align:center">{res}</td>
            </tr>'''

        alert = ""
        if msg:
            bg = "#14532d" if msg_type=="ok" else "#7f1d1d"
            alert = f'<div style="background:{bg};color:#fff;padding:14px 18px;border-radius:10px;margin-bottom:18px;font-weight:600">{msg}</div>'

        return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Laboy BSN · Log Pick</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#000;color:#f1f5f9;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
       min-height:100vh;padding:0 0 40px}}
  .header{{background:#111;border-bottom:2px solid #f07820;padding:14px 20px;
           display:flex;align-items:center;gap:12px}}
  .header h1{{font-size:1.1rem;font-weight:800;color:#f1f5f9}}
  .header h1 span{{color:#f07820}}
  .badge{{background:#f0782018;border:1px solid #f07820;color:#f07820;
          font-size:0.65rem;font-weight:700;padding:3px 9px;border-radius:20px}}
  .stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;padding:16px}}
  .stat{{background:#111;border-radius:10px;padding:12px;text-align:center;border:1px solid #1e293b}}
  .stat-val{{font-size:1.4rem;font-weight:900}}
  .stat-lbl{{font-size:0.6rem;color:#64748b;letter-spacing:1px;text-transform:uppercase;margin-top:2px}}
  .section{{padding:0 16px;margin-top:8px}}
  .card{{background:#111;border-radius:14px;padding:20px;border:1px solid #1e293b}}
  label{{display:block;font-size:0.72rem;color:#64748b;font-weight:700;
         letter-spacing:1px;text-transform:uppercase;margin-bottom:6px}}
  input,select,textarea{{width:100%;background:#0a0a0a;color:#f1f5f9;border:1px solid #1e293b;
    border-radius:8px;padding:12px 14px;font-size:1rem;margin-bottom:14px;
    -webkit-appearance:none;appearance:none;font-family:inherit}}
  input:focus,select:focus,textarea:focus{{outline:none;border-color:#f07820}}
  textarea{{height:80px;resize:none}}
  .row{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
  .btn{{width:100%;background:#f07820;color:#fff;border:none;border-radius:10px;
        padding:16px;font-size:1.05rem;font-weight:800;cursor:pointer;margin-top:4px;
        letter-spacing:0.5px}}
  .btn:active{{background:#d96a10}}
  .recent{{margin-top:20px}}
  .recent h3{{font-size:0.7rem;color:#64748b;letter-spacing:2px;text-transform:uppercase;
             margin-bottom:10px;padding-bottom:6px;border-bottom:1px solid #1e293b}}
  table{{width:100%;border-collapse:collapse}}
  td{{padding:8px 4px;border-bottom:1px solid #0f172a}}
  .divider{{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:10px;
            margin-bottom:16px;color:#1e293b;font-size:0.75rem;color:#334155}}
  .divider::before,.divider::after{{content:"";height:1px;background:#1e293b}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>Laboy <span>BSN</span> Picks</h1>
  </div>
  <div class="badge">LOG PICK</div>
</div>

<div class="stats">
  <div class="stat">
    <div class="stat-val" style="color:{'#22c55e' if w>=l else '#ef4444'}">{w}-{l}</div>
    <div class="stat-lbl">Record</div>
  </div>
  <div class="stat">
    <div class="stat-val" style="color:{'#22c55e' if w>=l else '#ef4444'}">{f'{w/(w+l)*100:.0f}%' if (w+l) else '—'}</div>
    <div class="stat-lbl">Win %</div>
  </div>
  <div class="stat">
    <div class="stat-val" style="color:{'#22c55e' if total_pnl>=0 else '#ef4444'}">{pnl_s}</div>
    <div class="stat-lbl">P&amp;L</div>
  </div>
</div>

<div class="section">
  {alert}
  <div class="card">
    <form method="POST" action="/log">
      <label>Fecha</label>
      <input type="date" name="date" value="{today}" required>

      <label>Juego (Equipo A vs. Equipo B)</label>
      <div class="row">
        <select name="team1" required>
          <option value="">Equipo A</option>
          {team_opts}
        </select>
        <select name="team2" required>
          <option value="">Equipo B</option>
          {team_opts}
        </select>
      </div>

      <label>Pick</label>
      <input type="text" name="pick" placeholder="Ej: LEONES ML / O 155.5 / LEONES -3.5" required autocomplete="off" autocapitalize="characters">

      <div class="row">
        <div>
          <label>Odds</label>
          <input type="text" name="odds" placeholder="-110 / +150" required autocomplete="off">
        </div>
        <div>
          <label>Apuesta ($)</label>
          <input type="number" name="stake" placeholder="15" step="0.01" min="1" required>
        </div>
      </div>

      <label>Sportsbook</label>
      <select name="book">
        <option value="BetMGM">BetMGM</option>
        <option value="DraftKings">DraftKings</option>
        <option value="FanDuel">FanDuel</option>
        <option value="Caesars">Caesars</option>
        <option value="Bet365">Bet365</option>
        <option value="PointsBet">PointsBet</option>
        <option value="Otro">Otro</option>
      </select>

      <label>Análisis (opcional)</label>
      <textarea name="analysis" placeholder="Razón del pick, tendencias, matchup..."></textarea>

      <button type="submit" class="btn">✅ Loguear Pick</button>
    </form>

    <div class="recent">
      <h3>Últimos picks ({len(pend)} pendiente{'s' if len(pend)!=1 else ''})</h3>
      <table>
        <tbody>{rows_html if rows_html else '<tr><td colspan="4" style="color:#334155;text-align:center;padding:16px">Sin picks aún</td></tr>'}</tbody>
      </table>
    </div>
  </div>
</div>
</body>
</html>"""

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, fmt, *args): pass  # silencia logs del servidor

        def do_GET(self):
            if self.path == "/favicon.ico":
                self.send_response(204); self.end_headers(); return
            body = _html_form().encode()
            self.send_response(200)
            self.send_header("Content-Type","text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_POST(self):
            if self.path != "/log":
                self.send_response(404); self.end_headers(); return
            length = int(self.headers.get("Content-Length", 0))
            raw    = self.rfile.read(length).decode("utf-8")
            data   = parse_qs(raw, keep_blank_values=True)
            def g(k): return unquote_plus(data.get(k,[""])[0]).strip()

            t1   = g("team1").upper()
            t2   = g("team2").upper()
            pick = g("pick").upper()
            odds_s = g("odds")
            stake_s = g("stake")
            book = g("book") or "BetMGM"
            analysis = g("analysis")
            d_val = g("date") or date.today().strftime("%Y-%m-%d")

            err = ""
            if not t1 or not t2: err = "⚠️ Selecciona ambos equipos."
            elif t1 == t2:        err = "⚠️ Los equipos no pueden ser iguales."
            elif not pick:        err = "⚠️ El pick es requerido."
            elif not odds_s:      err = "⚠️ Los odds son requeridos."
            elif not stake_s:     err = "⚠️ La apuesta es requerida."

            if err:
                body = _html_form(err, "err").encode()
                self.send_response(200)
                self.send_header("Content-Type","text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body); return

            try:
                odds_v  = _parse_odds_input(odds_s)
                stake_v = float(re.sub(r"[^\d.]","",stake_s.split()[0]))
                game    = f"{t1} vs. {t2}"
                log     = _load()
                entry   = {
                    "id":       len(log),
                    "date":     d_val,
                    "game":     game,
                    "pick":     pick,
                    "odds":     odds_v,
                    "stake":    stake_v,
                    "book":     book,
                    "result":   None,
                    "pnl":      None,
                    "analysis": analysis,
                }
                log.append(entry)
                _save(log)
                odds_fmt = _fmt_odds(odds_v)
                msg = f"✅ Pick #{entry['id']} logueado — {game} | {pick} {odds_fmt} | ${stake_v:.2f}"
                body = _html_form(msg, "ok").encode()
                self.send_response(200)
                self.send_header("Content-Type","text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                print(f"  ✅ Pick logueado: {game} | {pick} | {odds_fmt} | ${stake_v:.2f}")
            except Exception as ex:
                body = _html_form(f"⚠️ Error: {ex}", "err").encode()
                self.send_response(200)
                self.send_header("Content-Type","text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

    server = HTTPServer(("0.0.0.0", actual_port), Handler)

    print(f"\n{'═'*62}")
    print(f"  🌐  LABOY BSN · SERVIDOR DE LOG DE PICKS")
    print(f"{'═'*62}")
    print(f"\n  📱  Red local:   {local_url}")
    print(f"  💻  Localhost:   http://127.0.0.1:{actual_port}")
    print(f"\n  Para acceso desde CUALQUIER lugar (fuera de casa):")
    print(f"  1. Abre otra terminal y corre:  ngrok http {actual_port}")
    print(f"  2. Copia la URL https://xxxx.ngrok.io que te da")
    print(f"  3. Ábrela en el celular")
    print(f"\n  Ctrl+C para detener el servidor\n")
    print(f"{'═'*62}\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\n  🛑  Servidor detenido.\n")
        server.server_close()


def cmd_log_pick():
    print(f"\n{'═'*60}")
    print(f"  LABOY PICKS — REGISTRAR JUGADA")
    print(f"{'═'*60}\n")
    try:
        game_raw = _select_game_interactively(TARGET_DATE, prompt_label="Juego")
        if not game_raw:
            print("  ❌ Juego requerido."); return
        # Normalizar separador al guardar
        def _norm_game(g):
            for sep in [" VS. ", " vs. ", " VS ", " vs ", " @ ", " - "]:
                if sep in g:
                    p = g.split(sep, 1)
                    return f"{p[0].strip()} vs. {p[1].strip()}"
            for t in sorted(BSN_TEAMS, key=len, reverse=True):
                if g.startswith(t + " "):
                    rest = g[len(t):].strip()
                    if rest: return f"{t} vs. {rest}"
            return g
        game = _norm_game(game_raw)
        pick     = input("  Pick     (ej: LEONES ML / O 155.5 / LEONES -3.5): ").strip().upper()
        book_raw = input("  Sportsbook [BetMGM]: ").strip()
        book     = book_raw if book_raw else "BetMGM"
        # BSN no tiene API de odds — usuario ingresa manualmente
        odds_s = input("  Odds (ej: +150 o -110): ").strip()
        odds_v = _parse_odds_input(odds_s)
        stake_s  = input("  Apostado (ej: 15 para $15.00): ").strip()
        stake_clean = re.sub(r"[^\d.]", "", stake_s.split()[0])
        stake    = float(stake_clean)
        analysis = input("  Análisis (opcional — razón del pick, Enter para omitir):\n  > ").strip()
    except (ValueError, EOFError) as _e:
        print(f"\n  ❌ Entrada inválida ({_e}). Intenta de nuevo.\n"); return
    log   = _load_log()
    entry = {"id":len(log),"date":TARGET_DATE,"game":game,"pick":pick,
             "odds":odds_v,"stake":stake,"book":book,"result":None,"pnl":None,
             "analysis": analysis or ""}
    log.append(entry); _save_log(log)
    # Mostrar payout potencial
    if odds_v > 0:
        pot = round(stake * (odds_v / 100), 2)
    else:
        pot = round(stake * (100 / abs(odds_v)), 2)
    print(f"\n  ✅ Pick #{entry['id']}: {game} │ {pick} │ {_fmt_odds(odds_v)} │ ${stake:.2f} → potencial +${pot:.2f}")
    # Auto-generate HTML card
    try:
        html_path = export_log_pick_html(entry)
        if html_path:
            print(f"     → {os.path.basename(html_path)}")
    except Exception as e:
        print(f"     ⚠️  HTML export falló: {e}")
    print()

def cmd_export_log():
    """
    --export-log [IDX]
    Re-exporta un pick logueado como HTML.
    Si no se da IDX, exporta el último pick logueado.
    Ejemplo: python3 bsn.py --export-log 3
    """
    log = _load_log()
    if not log:
        print("  ❌ No hay picks logueados. Usa --log primero.\n"); return
    try:
        ei  = sys.argv.index("--export-log")
        idx = int(sys.argv[ei+1]) if ei+1 < len(sys.argv) and sys.argv[ei+1].lstrip('-').isdigit() else len(log)-1
    except (ValueError, IndexError):
        idx = len(log) - 1
    if not (0 <= idx < len(log)):
        print(f"  ❌ Índice {idx} inválido. Hay {len(log)} picks (0–{len(log)-1}).\n"); return
    entry = log[idx]
    print(f"\n  🌐 Exportando Pick #{idx}: {entry['game']} │ {entry['pick']} │ {_fmt_odds(entry['odds'])}")
    html_path = export_log_pick_html(entry)
    if html_path:
        print(f"  ✅ Guardado: {os.path.basename(html_path)}\n")
        if PUBLISH_MODE:
            cmd_publish([html_path])
    else:
        print("  ❌ Error al generar HTML.\n")

def cmd_log_retro():
    """
    --log-retro
    Agrega picks retroactivos al log (con fecha, resultado y P&L ya conocidos).
    Útil para registrar picks del pasado que no fueron logueados en su momento.
    """
    print(f"\n{'═'*60}")
    print(f"  LABOY BSN — PICK RETROACTIVO")
    print(f"{'═'*60}\n")
    print("  (Escribe 'listo' en Fecha para terminar)\n")

    log = _load_log()

    def _norm_game(g):
        for sep in [" VS. ", " vs. ", " VS ", " vs ", " @ ", " - "]:
            if sep in g:
                p = g.split(sep, 1)
                return f"{p[0].strip()} vs. {p[1].strip()}"
        return g

    count = 0
    while True:
        try:
            print(f"  {'─'*40}")
            date_s = input("  Fecha (YYYY-MM-DD o 'listo'): ").strip()
            if date_s.lower() in ("listo", "done", "q", "exit", ""):
                break
            # Validar fecha
            try:
                datetime.strptime(date_s, "%Y-%m-%d")
            except ValueError:
                print("  ⚠️  Formato inválido. Usa YYYY-MM-DD.\n"); continue

            game_raw = input("  Juego    (ej: LEONES vs. OSOS): ").strip().upper()
            game     = _norm_game(game_raw)
            pick     = input("  Pick     (ej: LEONES ML / O 155.5): ").strip().upper()
            book_raw = input("  Sportsbook [BetMGM]: ").strip()
            book     = book_raw if book_raw else "BetMGM"
            odds_s   = input("  Odds (ej: -110 / +135): ").strip()
            odds_v   = _parse_odds_input(odds_s)
            stake_s  = input("  Apostado (ej: 15): ").strip()
            stake    = float(re.sub(r"[^\d.]", "", stake_s.split()[0]))

            result_s = input("  Resultado (W / L / P): ").strip().upper()
            if result_s not in ("W", "L", "P"):
                print("  ⚠️  Resultado debe ser W, L o P.\n"); continue

            # Calcular P&L automático si no se ingresa
            if result_s == "W":
                if odds_v > 0:
                    auto_pnl = round(stake * (odds_v / 100), 2)
                else:
                    auto_pnl = round(stake * (100 / abs(odds_v)), 2)
                pnl_s = input(f"  P&L (Enter = +${auto_pnl:.2f}): ").strip()
                pnl   = float(pnl_s) if pnl_s else auto_pnl
            elif result_s == "L":
                pnl_s = input(f"  P&L (Enter = -${stake:.2f}): ").strip()
                pnl   = float(pnl_s) if pnl_s else -stake
            else:  # Push
                pnl   = 0.0

            analysis = input("  Análisis (opcional, Enter para omitir):\n  > ").strip()

        except (ValueError, EOFError) as _e:
            print(f"\n  ❌ Entrada inválida ({_e}).\n"); continue

        entry = {
            "id":       len(log),
            "date":     date_s,
            "game":     game,
            "pick":     pick,
            "odds":     odds_v,
            "stake":    stake,
            "book":     book,
            "result":   result_s,
            "pnl":      pnl,
            "analysis": analysis or "",
        }
        log.append(entry)
        _save_log(log)
        pnl_fmt = f"+${pnl:.2f}" if pnl >= 0 else f"-${abs(pnl):.2f}"
        print(f"\n  ✅ Pick #{entry['id']} guardado: {date_s} │ {game} │ {pick} │ {result_s} │ {pnl_fmt}\n")
        count += 1

    if count:
        print(f"\n  📋 {count} pick(s) retroactivo(s) agregados al log.\n")
    else:
        print(f"\n  (Sin cambios)\n")


def cmd_grade_pick():
    """
    Uso: python3 bsn.py --grade IDX W|L|P
    IDX  = número del pick (sale en --record, empieza en 0)
    W    = Win (ganaste)
    L    = Loss (perdiste)
    P    = Push (devuelven dinero)

    Ejemplo: python3 bsn.py --grade 0 W
    """
    try:
        gi  = sys.argv.index("--grade")
        idx = int(sys.argv[gi+1])
        res = sys.argv[gi+2].upper()
        assert res in ("W","L","P")
    except (ValueError, IndexError, AssertionError):
        print("  ❌ Uso: python3 bsn.py --grade IDX W|L|P")
        print("     Ejemplo: python3 bsn.py --grade 0 W")
        print("\n  IDX  = número del pick (ver python3 bsn.py --record)")
        print("  W    = Win   L = Loss   P = Push")
        return
    log = _load_log()
    if not (0 <= idx < len(log)):
        print(f"  ❌ Índice {idx} no válido. Hay {len(log)} picks (0–{len(log)-1}).\n"
              f"     Corre: python3 bsn.py --record  para ver los índices.")
        return
    e = log[idx]; e["result"] = res
    sv = _entry_stake(e)   # stake en $ o units (legacy)
    if res == "W":
        e["pnl"] = round(sv * (e["odds"] / 100) if e["odds"] > 0
                         else sv * (100 / abs(e["odds"])), 2)
    elif res == "L": e["pnl"] = round(-sv, 2)
    else:            e["pnl"] = 0.0
    _save_log(log)
    emoji = {"W":"✅","L":"❌","P":"🔄"}[res]
    print(f"\n  {emoji} Pick #{idx} → {res} | {e['game']} {e['pick']} {_fmt_odds(e['odds'])} | P&L: {_fmt_pnl(e)}")
    # Regenerar HTML pick card con badge WIN/LOSS actualizado
    try:
        html_path = export_log_pick_html(e)
        if html_path:
            print(f"  🖼️  Pick card actualizado: {os.path.basename(html_path)}")
    except Exception as _ge:
        print(f"  ⚠️  No se pudo regenerar HTML: {_ge}")
    print()

def cmd_remove_pick():
    """
    Elimina uno o varios picks del log por su índice (ID).
    Uso: python3 bsn.py --remove 2
         python3 bsn.py --remove 2 4 7

    Los IDs restantes se renumeran automáticamente para mantener consistencia.
    """
    # Colectar todos los índices que siguen a --remove
    try:
        ri = sys.argv.index("--remove")
        idxs = []
        for a in sys.argv[ri+1:]:
            if a.startswith("--"): break
            idxs.append(int(a))
    except (ValueError, IndexError):
        print("  ❌ Uso: python3 bsn.py --remove IDX [IDX2 ...]")
        print("     Ejemplo: python3 bsn.py --remove 3")
        return

    if not idxs:
        print("  ❌ Especifica al menos un índice. Usa --record para ver los IDs.\n")
        return

    log = _load_log()
    if not log:
        print("  ❌ No hay picks en el log.\n"); return

    # Validar todos los índices primero
    invalid = [i for i in idxs if not (0 <= i < len(log))]
    if invalid:
        print(f"  ❌ Índice(s) inválido(s): {invalid}. El log tiene {len(log)} picks (0–{len(log)-1}).\n")
        return

    # Mostrar qué se va a borrar y pedir confirmación
    print(f"\n{'═'*52}")
    print(f"  ⚠️  ELIMINAR {len(idxs)} PICK(S)")
    print(f"{'═'*52}")
    for i in sorted(idxs):
        e = log[i]
        print(f"  #{i} — {e['date']} | {e['game']} | {e['pick']} {_fmt_odds(e['odds'])}")
    print()
    confirm = input("  ¿Confirmar eliminación? (s/N): ").strip().lower()
    if confirm not in ("s","si","sí","y","yes"):
        print("  ↩️  Cancelado.\n"); return

    # Eliminar y renumerar
    to_remove = set(idxs)
    new_log   = [e for i, e in enumerate(log) if i not in to_remove]
    for new_id, e in enumerate(new_log):
        e["id"] = new_id
    _save_log(new_log)

    print(f"\n  ✅ {len(idxs)} pick(s) eliminado(s). Quedan {len(new_log)} en el log.\n")

def cmd_edit_pick():
    """
    Edita un campo de un pick existente.
    Uso: python3 bsn.py --edit IDX campo valor

    Campos editables: date, game, pick, odds, stake, book, result, pnl, analysis
    Ejemplos:
      python3 bsn.py --edit 33 date 2026-04-19
      python3 bsn.py --edit 33 game "OSOS vs. VAQUEROS"
      python3 bsn.py --edit 33 pick "OVER 169.5"
      python3 bsn.py --edit 33 odds -110
      python3 bsn.py --edit 33 stake 20
    """
    EDITABLE = {"date","game","pick","odds","stake","book","result","pnl","analysis"}

    try:
        ei    = sys.argv.index("--edit")
        idx   = int(sys.argv[ei + 1])
        field = sys.argv[ei + 2].lower()
        value = sys.argv[ei + 3]
    except (IndexError, ValueError):
        print("  ❌ Uso: python3 bsn.py --edit IDX campo valor")
        print("  Ejemplo: python3 bsn.py --edit 33 date 2026-04-19")
        print(f"  Campos disponibles: {', '.join(sorted(EDITABLE))}")
        return

    if field not in EDITABLE:
        print(f"  ❌ Campo '{field}' no válido.")
        print(f"  Campos disponibles: {', '.join(sorted(EDITABLE))}")
        return

    log = _load_log()
    if not (0 <= idx < len(log)):
        print(f"  ❌ Pick #{idx} no existe. El log tiene {len(log)} picks (0–{len(log)-1}).")
        return

    entry = log[idx]
    old_val = entry.get(field)

    # Coerce numeric fields
    if field in ("odds",):
        try:   value = int(value)
        except ValueError: value = float(value)
    elif field in ("stake", "pnl"):
        try:   value = float(value)
        except ValueError: pass

    entry[field] = value
    _save_log(log)

    print(f"\n  ✅ Pick #{idx} actualizado:")
    print(f"     {field}: {repr(old_val)}  →  {repr(value)}")
    print(f"     Juego: {entry['game']} | {entry['pick']} | {_fmt_odds(entry['odds'])}\n")


def cmd_record():
    """
    --record              → últimos 30 picks (más recientes primero)
    --record all          → todos los picks
    --record 2026-04-18   → solo picks de esa fecha
    """
    log = _load_log()
    print(f"\n{'═'*80}")
    print(f"  LABOY PICKS — REGISTRO")
    print(f"{'═'*80}")
    if not log:
        print("\n  No hay jugadas. Usa: python3 bsn.py --log\n"); return

    # ── Parse argument after --record ────────────────────────
    date_filter = None
    show_all    = False
    try:
        ri = sys.argv.index("--record")
        arg = sys.argv[ri + 1] if ri + 1 < len(sys.argv) and not sys.argv[ri + 1].startswith("--") else None
        if arg:
            if arg.lower() == "all":
                show_all = True
            elif re.match(r"^\d{4}-\d{2}-\d{2}$", arg):
                date_filter = arg
    except (ValueError, IndexError):
        pass

    # ── Build running balance over ALL picks first ────────────
    running_balance = 0.0
    balance_by_id   = {}
    for e in log:
        res = e.get("result") or "—"
        pnl = e.get("pnl")
        stake = _entry_stake(e)
        if res == "W":
            running_balance += pnl if pnl is not None else stake
        elif res == "L":
            running_balance -= stake
        balance_by_id[e["id"]] = running_balance

    # ── Filter ───────────────────────────────────────────────
    if date_filter:
        display_log = [e for e in log if e.get("date","") == date_filter]
        filter_label = f"  📅 Filtrado: {date_filter}  ({len(display_log)} picks)"
    elif show_all:
        display_log  = log
        filter_label = f"  📋 Mostrando todos los picks ({len(log)})"
    else:
        display_log  = log[-30:]       # last 30, already newest at bottom
        filter_label = f"  📋 Últimos {len(display_log)} picks  (usa --record all para ver todos)"

    if not display_log:
        print(f"\n  Sin picks para {date_filter}.\n"); return

    # ── Newest first ─────────────────────────────────────────
    display_log = list(reversed(display_log))

    rows = []
    for e in display_log:
        res     = e.get("result") or "—"
        bal_str = (f"+${balance_by_id[e['id']]:.2f}" if balance_by_id[e['id']] >= 0
                   else f"-${abs(balance_by_id[e['id']]):.2f}")
        rows.append([e["id"], e["date"], e["game"][:22], e["pick"][:14],
                     _fmt_odds(e["odds"]), _fmt_stake(e), res, _fmt_pnl(e), bal_str])

    print(f"\n{filter_label}")
    print("\n" + tab(rows, ["#","Fecha","Juego","Pick","Odds","Apostado","Res","P&L","Ganancia"]))

    # ── Stats always over full log ────────────────────────────
    graded = [e for e in log if e.get("result") in ("W","L","P")]
    wins   = [e for e in graded if e["result"]=="W"]
    pnl_t  = sum(e["pnl"] for e in graded if e.get("pnl") is not None)
    wag    = sum(_entry_stake(e) for e in graded)
    roi    = (pnl_t / wag * 100) if wag > 0 else 0
    use_dollars = any(e.get("stake") is not None for e in graded)
    pnl_str = f"+${pnl_t:.2f}" if (use_dollars and pnl_t >= 0) else (f"-${abs(pnl_t):.2f}" if use_dollars else f"{pnl_t:+.2f}u")
    wag_str = f"${wag:.2f}" if use_dollars else f"{wag}u"
    bal_final = (f"+${running_balance:.2f}" if running_balance >= 0
                 else f"-${abs(running_balance):.2f}")

    print(f"\n  📊 Récord total: {len(wins)}-{len([e for e in graded if e['result']=='L'])}-"
          f"{len([e for e in graded if e['result']=='P'])}  "
          f"Pending: {len(log)-len(graded)}")
    if graded:
        print(f"     Win%: {len(wins)/len(graded)*100:.1f}%  │  P&L: {pnl_str}  │  Jugado: {wag_str}  │  ROI: {roi:+.1f}%")
    print(f"  💰 Ganancia actual: {bal_final}")

    print(f"\n  Tips:")
    print(f"    python3 bsn.py --record all               ← ver todos los picks")
    print(f"    python3 bsn.py --record 2026-04-18        ← picks de una fecha")
    print(f"    python3 bsn.py --grade N W|L|P            ← califica pick #N")
    print(f"    python3 bsn.py --edit N campo valor       ← edita un campo del pick")
    print(f"    python3 bsn.py --remove N                  ← elimina pick #N del log")
    print(f"    python3 bsn.py --export-record [DATE]      ← exporta tarjeta HTML+JPG\n")

def cmd_feedback():
    log    = _load_log()
    graded = [e for e in log if e.get("result") in ("W","L")]
    print(f"\n{'═'*72}")
    print(f"  LABOY PICKS — ANÁLISIS DE RENDIMIENTO")
    print(f"{'═'*72}")
    if len(graded) < 3:
        print(f"\n  Necesitas ≥3 picks calificados. Tienes {len(graded)}.\n"); return
    use_dollars = any(e.get("stake") is not None for e in graded)
    def _pnl_s(v): return (f"+${v:.2f}" if v >= 0 else f"-${abs(v):.2f}") if use_dollars else f"{v:+.2f}u"

    by_type = {}
    for e in graded:
        p = e.get("pick","").upper()
        bt = "ML" if "ML" in p else "Over" if (p.startswith("O ") or "OVER" in p) \
             else "Under" if (p.startswith("U ") or "UNDER" in p) \
             else "Spread" if any(x in p for x in ["-1.5","+1.5","SPREAD","-3","+3"]) else "Other"
        s = by_type.setdefault(bt, {"W":0,"L":0,"pnl":0.0})
        s[e["result"]] += 1; s["pnl"] += e.get("pnl", 0)
    type_rows = []
    for bt, s in sorted(by_type.items()):
        tot = s["W"] + s["L"]
        type_rows.append([bt, f"{s['W']}-{s['L']}", f"{s['W']/tot*100:.0f}%", _pnl_s(s["pnl"])])
    print(f"\n  Por tipo:\n" + tab(type_rows, ["Tipo","W-L","Win%","P&L"], fmt="simple"))

    def bucket(o):
        if o >= 200:  return "+200 y más"
        if o >= 101:  return "+101 a +199"
        if o >= -120: return "EVEN a -120"
        if o >= -150: return "-121 a -150"
        return "-151 y menos"
    by_o = {}
    for e in graded:
        b = bucket(e.get("odds", 0))
        s = by_o.setdefault(b, {"W":0,"L":0,"pnl":0.0})
        s[e["result"]] += 1; s["pnl"] += e.get("pnl", 0)
    odds_rows = []
    for b in ["+200 y más","+101 a +199","EVEN a -120","-121 a -150","-151 y menos"]:
        if b in by_o:
            s = by_o[b]; tot = s["W"] + s["L"]
            odds_rows.append([b, f"{s['W']}-{s['L']}", f"{s['W']/tot*100:.0f}%", _pnl_s(s["pnl"])])
    if odds_rows:
        print(f"\n  Por rango odds:\n" + tab(odds_rows, ["Odds Range","W-L","Win%","P&L"], fmt="simple"))
    wp  = sum(1 for e in graded if e["result"]=="W") / len(graded) * 100
    pnl = sum(e.get("pnl", 0) for e in graded)
    print(f"\n  💡 Win rate: {wp:.0f}%  P&L: {_pnl_s(pnl)}")
    if by_type:
        best  = max(by_type.items(), key=lambda x: x[1]["pnl"])
        worst = min(by_type.items(), key=lambda x: x[1]["pnl"])
        print(f"  ✅ Mejor tipo: {best[0]} ({_pnl_s(best[1]['pnl'])})")
        if worst[0] != best[0]:
            print(f"  ⚠️  Área a mejorar: {worst[0]} ({_pnl_s(worst[1]['pnl'])})")

    # ── Análisis AI (si hay ANTHROPIC_API_KEY) ────────────────────
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if api_key:
        print(f"\n  🤖 Analizando patrones con IA...\n")
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)

            # Construir resumen de picks para el prompt
            summary_lines = []
            for e in graded[-30:]:  # Últimos 30 picks
                summary_lines.append(
                    f"  {e.get('date','')} | {e.get('game','')} | {e.get('pick','')} | "
                    f"Odds:{e.get('odds','')} | {e.get('result','')} | P&L:{e.get('pnl','')}"
                )
            summary = "\n".join(summary_lines)

            losses_only = [e for e in graded if e["result"] == "L"]
            loss_summary = "\n".join(
                f"  {e.get('date','')} | {e.get('game','')} | {e.get('pick','')} | Odds:{e.get('odds','')}"
                for e in losses_only[-15:]
            )

            prompt = f"""Eres un analista de apuestas deportivas especializado en baloncesto de Puerto Rico (BSN).
Analiza el siguiente historial de picks y proporciona retroalimentación en español.

HISTORIAL RECIENTE ({len(graded)} picks calificados):
{summary}

PÉRDIDAS RECIENTES:
{loss_summary}

ESTADÍSTICAS GENERALES:
- Win rate: {wp:.0f}%
- P&L total: {_pnl_s(pnl)}
- Mejor tipo de pick: {best[0] if by_type else 'N/A'}
- Área a mejorar: {worst[0] if by_type and worst[0] != best[0] else 'N/A'}

Por favor analiza:
1. ¿Hay patrones en las pérdidas? (tipos de picks, rangos de odds, equipos específicos, horarios)
2. ¿Qué estrategias están funcionando mejor?
3. ¿Hay equipos o matchups donde el modelo parece tener más o menos ventaja?
4. Recomendaciones concretas para mejorar el rendimiento en BSN

Sé específico, usa los datos, y responde en español claro y conciso (máximo 300 palabras)."""

            message = client.messages.create(
                model="claude-opus-4-6",
                max_tokens=1024,
                messages=[{"role": "user", "content": prompt}]
            )
            ai_text = message.content[0].text
            print(f"  {'─'*60}")
            print(f"  🧠 ANÁLISIS IA:")
            print(f"  {'─'*60}")
            for line in ai_text.split("\n"):
                print(f"  {line}")
            print()
        except ImportError:
            print("  💡 Instala anthropic: pip install anthropic --break-system-packages")
        except Exception as e:
            print(f"  ⚠️  Error AI: {e}\n")
    else:
        print(f"\n  💡 Para análisis AI, configura: export ANTHROPIC_API_KEY='tu_key'")

    print()


# ──────────────────────────────────────────────────────
# PARLAY PICK CARD — HTML + PNG
# ──────────────────────────────────────────────────────

def export_log_parlay_html(entry):
    """
    Genera 'Laboy BSN Parlay YYYY-MM-DD #N.html' con diseño parlay:
    - Header de 2 legs con game/pick de cada uno
    - Odds combinadas + Stake + badge PENDING/WIN/LOSS
    Retorna path o None.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    try:
        pick_date  = entry.get("date", TARGET_DATE)
        pick_id    = entry.get("id", 0)
        legs       = entry.get("legs", [])
        odds_v     = entry.get("odds", 0)
        stake      = entry.get("stake", 0)
        book       = entry.get("book", "BetMGM")
        result     = entry.get("result")
        pnl        = entry.get("pnl")
        n_legs     = len(legs)

        odds_fmt   = _fmt_odds(odds_v)
        stake_disp = f"${stake:.0f}" if stake == int(stake) else f"${stake:.2f}"
        dt         = datetime.strptime(pick_date, "%Y-%m-%d")
        dstr       = dt.strftime("%A, %B %d · %Y").upper()
        yr         = dt.strftime("%Y")

        # Result badge
        if result == "W":
            result_color = "#22c55e"
            result_html  = '<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✓ WIN</span>'
            card_bg      = "background:linear-gradient(135deg,#0d1f14 0%,#161e2e 60%)"
            border_col   = "#22c55e"
        elif result == "L":
            result_color = "#ef4444"
            result_html  = '<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✗ LOSS</span>'
            card_bg      = "background:linear-gradient(135deg,#1f0d0d 0%,#161e2e 60%)"
            border_col   = "#ef4444"
        else:
            result_color = "#f07820"
            result_html  = '<span style="background:#f0782022;color:#f07820;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">⏳ PENDING</span>'
            card_bg      = ""
            border_col   = "#f07820"

        pnl_html = ""
        if pnl is not None:
            pnl_col  = "#22c55e" if pnl >= 0 else "#ef4444"
            pnl_disp = f"+${pnl:.2f}" if pnl >= 0 else f"-${abs(pnl):.2f}"
            pnl_html = f'<div class="stat"><div class="stat-label">P&L</div><div class="stat-val" style="color:{pnl_col}">{esc(pnl_disp)}</div></div>'

        # Build leg HTML
        legs_html = ""
        for i, leg in enumerate(legs):
            lg_game = leg.get("game", "").upper()
            lg_pick = leg.get("pick", "").upper()
            # Color from team
            lg_col  = "#f07820"
            for t, c in _BSN_COLORS.items():
                if t in lg_pick or t in lg_game:
                    lg_col = f"rgb({c[0]},{c[1]},{c[2]})"
                    break
            connector = ""
            if i > 0:
                connector = '<div style="text-align:center;font-size:1.2rem;color:#f07820;margin:6px 0">+</div>'
            legs_html += f"""
            {connector}
            <div style="background:#0a0e1a;border-radius:10px;padding:14px 16px;border-left:3px solid {lg_col}">
              <div style="font-size:0.6rem;color:#3d4a5c;letter-spacing:1px;font-family:monospace;margin-bottom:4px">LEG {i+1} &nbsp;·&nbsp; {esc(lg_game)}</div>
              <div style="font-size:1.2rem;font-weight:900;color:#e2e8f0;letter-spacing:0.5px">{esc(lg_pick)}</div>
            </div>"""

        card_html = f"""
        <div style="border-left:4px solid {border_col};border-radius:14px;padding:20px;margin-bottom:14px;{card_bg};background-color:#111827;background-origin:border-box">
          <!-- top row -->
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
            <div style="font-size:0.65rem;color:#3d4a5c;font-family:monospace">
              🔗 PARLAY &nbsp;·&nbsp; {n_legs} LEGS &nbsp;·&nbsp; PICK #{pick_id}
            </div>
            {result_html}
          </div>

          <!-- legs -->
          <div style="margin-bottom:16px">{legs_html}</div>

          <!-- stats -->
          <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:14px">
            <div style="background:#0a0e1a;border-radius:8px;padding:10px 12px;text-align:center">
              <div style="font-size:0.6rem;color:#3d4a5c;font-family:monospace;margin-bottom:4px">{esc(book)}</div>
              <div style="font-size:1rem;font-weight:700;color:#f07820">{esc(odds_fmt)}</div>
            </div>
            <div style="background:#0a0e1a;border-radius:8px;padding:10px 12px;text-align:center">
              <div style="font-size:0.6rem;color:#3d4a5c;font-family:monospace;margin-bottom:4px">APOSTADO</div>
              <div style="font-size:1rem;font-weight:700;color:#e2e8f0">{esc(stake_disp)}</div>
            </div>
            <div style="background:#0a0e1a;border-radius:8px;padding:10px 12px;text-align:center">
              <div style="font-size:0.6rem;color:#3d4a5c;font-family:monospace;margin-bottom:4px">P&L</div>
              <div style="font-size:1rem;font-weight:700;{'color:#22c55e' if pnl is not None and pnl >= 0 else 'color:#ef4444' if pnl is not None else 'color:#3d4a5c'}">
                {(f'+${pnl:.2f}' if pnl >= 0 else f'-${abs(pnl):.2f}') if pnl is not None else '—'}</div>
            </div>
          </div>
        </div>"""

        logo_b64 = _bsn_logo_b64()
        logo_img  = (f'<img src="{logo_b64}" alt="Laboy" style="height:32px;opacity:0.85">'
                     if logo_b64 else '<span style="font-weight:900;color:#f07820;font-size:1.1rem">LABOY</span>')
        picks_url = ""  # no hay picks HTML en parlay stand-alone

        html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Laboy BSN Parlay {pick_date} #{pick_id}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&family=IBM+Plex+Mono:wght@400;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
<style>
  :root{{--bg:#060b14;--card:#111827;--muted:#3d4a5c;--ice:#e4eeff;--orange:#f07820}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--ice);font-family:'Inter',sans-serif;min-height:100vh;
       display:flex;flex-direction:column;align-items:center;justify-content:center;padding:20px}}
  .wrap{{max-width:480px;width:100%}}
  .top-bar{{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px}}
  .sport-badge{{font-size:0.6rem;font-weight:700;letter-spacing:2px;color:var(--orange);
                font-family:'IBM Plex Mono',monospace;border:1px solid var(--orange);
                border-radius:4px;padding:3px 8px}}
  .date-str{{font-size:0.6rem;color:var(--muted);font-family:'IBM Plex Mono',monospace}}
  .section-title{{font-size:0.65rem;letter-spacing:2px;color:var(--muted);
                  font-family:'IBM Plex Mono',monospace;margin-bottom:14px}}
  .footer{{margin-top:24px;text-align:center;font-size:0.6rem;color:var(--muted);font-family:'IBM Plex Mono',monospace}}
</style>
</head><body>
<div class="wrap">
  <div class="top-bar">
    {logo_img}
    <div style="text-align:right">
      <div class="sport-badge">BSN · PARLAY</div>
      <div class="date-str" style="margin-top:4px">{esc(dstr)}</div>
    </div>
  </div>
  {card_html}
  <div class="footer">dubclub.win &nbsp;·&nbsp; Laboy Picks &nbsp;·&nbsp; BSN</div>
</div>
</body></html>"""

        fname = f"Laboy BSN Parlay {pick_date} #{pick_id}.html"
        fpath = os.path.join(SCRIPT_DIR, fname)
        if os.path.exists(fpath) and not FORCE_EXPORT:
            print(f"  🔒 BSN Parlay HTML #{pick_id} ya existe — protegido de sobreescritura.")
            print(f"     Usa --force-export para regenerar.")
            return fpath
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"  🔗 Parlay HTML: {fname}")
        # Also generate PNG pick card
        png_path = export_parlay_png(entry)
        return fpath
    except Exception as e:
        print(f"  ❌ export_log_parlay_html error: {e}")
        import traceback; traceback.print_exc()
        return None


def export_parlay_png(entry):
    """
    Genera PNG 1080×1920 con pick card de parlay.
    Sigue EXACTAMENTE el mismo estilo que _draw_bsn_picks_page:
    cada leg = un card (panel, barra naranja, game pequeño, pick grande, badge odds).
    Los cards se encadenan con un conector '+' entre ellos.
    """
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export")
        return None
    try:
        W, H  = 1080, 1920
        PAD   = 54
        GAP   = 20

        legs      = entry.get("legs", [])
        odds_v    = entry.get("odds", 0)
        stake     = entry.get("stake", 0)
        book      = entry.get("book", "BetMGM")
        result    = entry.get("result")
        pick_date = entry.get("date", TARGET_DATE)
        pick_id   = entry.get("id", 0)
        n_legs    = len(legs)
        odds_s    = f"{odds_v:+d}" if odds_v else "—"
        stake_s   = f"${stake:.0f}" if stake == int(stake) else f"${stake:.2f}"

        dt   = datetime.strptime(pick_date, "%Y-%m-%d")
        dstr = dt.strftime("%A %B %d · %Y").upper()

        # Result colors (no emoji — font may not support it)
        if result == "W":
            acc_col = _GREEN_B;  acc_bg = _GREEN_BGB; badge_t = "WIN"
        elif result == "L":
            acc_col = _RED_B;    acc_bg = (40, 8, 8);  badge_t = "LOSS"
        else:
            acc_col = _AMBER_B;  acc_bg = (40, 22, 4); badge_t = "PENDING"

        img = Image.new("RGB", (W, H), _BG_B)
        d   = ImageDraw.Draw(img)
        _png_dot_bg_b(d, W, H)

        d.rectangle([(0, 0),(W, 8)],    fill=_AMBER_B)
        d.rectangle([(0, H-8),(W, H)],  fill=_AMBER_B)

        # ── Fonts (identical to _draw_bsn_picks_page) ─────────────────────
        F_HERO  = _fnt_bsn("BigShoulders-Bold.ttf", 96)
        F_SUB   = _fnt_bsn("IBMPlexMono-Bold.ttf",  22)
        F_PICK  = _fnt_bsn("BigShoulders-Bold.ttf", 68)   # big pick text
        F_MONOB = _fnt_bsn("IBMPlexMono-Bold.ttf",  30)   # odds badge
        F_LBL   = _fnt_bsn("GeistMono-Regular.ttf", 22)   # game + units
        F_MICRO = _fnt_bsn("GeistMono-Regular.ttf", 18)   # small labels
        F_CONN  = _fnt_bsn("BigShoulders-Bold.ttf", 48)   # "+" connector

        # Corner marks
        M, L = 40, 28
        for (x1,y1),(x2,y2) in [
            [(M,M),(M+L,M)],     [(M,M),(M,M+L)],
            [(W-M-L,M),(W-M,M)], [(W-M,M),(W-M,M+L)],
            [(M,H-M),(M+L,H-M)], [(M,H-M-L),(M,H-M)],
            [(W-M-L,H-M),(W-M,H-M)], [(W-M,H-M-L),(W-M,H-M)],
        ]:
            d.line([(x1,y1),(x2,y2)], fill=_AMBER_MB, width=2)

        # ── Header ────────────────────────────────────────────────────────
        y = 40
        _cx_b(d, "LABOY PICKS", F_HERO, W, y, _ICE_B)
        y += 112
        d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
        y += 16
        _cx_b(d, f"BSN  ·  {dstr}", F_MICRO, W, y, _MUTED_B)
        y += 40
        _cx_b(d, f"PARLAY  ·  {n_legs} LEGS", F_SUB, W, y, _AMBER_B)
        y += 44
        d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
        y += 28

        # ── One card per leg (same structure as _draw_bsn_picks_page) ─────
        CARD_H  = 260   # identical to single-pick cards
        CONN_H  = 56    # height of the "+" connector zone between cards

        for i, leg in enumerate(legs):
            lg_game = leg.get("game", "").upper()
            lg_pick = leg.get("pick", "").upper()

            # Card background — same as single picks
            d.rounded_rectangle([(PAD, y),(W-PAD, y+CARD_H)],
                                 radius=14, fill=_PANEL_B)
            d.rounded_rectangle([(PAD+1, y+1),(W-PAD-1, y+CARD_H-1)],
                                 radius=13, outline=(22,28,50), width=1)
            # Left accent bar — amber for parlay
            d.rounded_rectangle([(PAD, y),(PAD+4, y+CARD_H)],
                                 radius=2, fill=_AMBER_B)

            # Small LEG label — top-left inside card
            d.text((PAD+16, y+14), f"LEG {i+1}", font=F_MICRO, fill=_AMBER_B)

            # Game — centered, dimmed (same as single picks)
            _cx_b(d, lg_game, F_LBL, W, y+20, _ICE_DB)

            # Pick — big, centered (same as single picks)
            _cx_b(d, lg_pick, F_PICK, W, y+58, _ICE_B)

            # No per-leg odds badge — parlay odds shown once at the bottom
            # Show stake on first leg instead
            if i == 0:
                _cx_b(d, f"APOSTADO: {stake_s}", F_LBL, W, y+196, _MUTED_B)

            # Pick ID bottom-left only on last leg
            if i == n_legs - 1:
                d.text((PAD+16, y+CARD_H-28), f"Pick #{pick_id}",
                       font=F_MICRO, fill=_MUTED_B)

            y += CARD_H

            # ── Connector between legs ─────────────────────────────────
            if i < n_legs - 1:
                # Thin rules on each side of the "+"
                mid_y = y + CONN_H // 2
                d.line([(PAD+16, mid_y),(W//2-36, mid_y)], fill=_RULE_B, width=1)
                d.line([(W//2+36, mid_y),(W-PAD-16, mid_y)], fill=_RULE_B, width=1)
                # Circle with "+"
                r = 26
                cx = W // 2
                d.ellipse([(cx-r, mid_y-r),(cx+r, mid_y+r)],
                           fill=_PANEL_B, outline=_AMBER_B, width=2)
                pw, ph = _tw_b(d, "+", F_CONN)
                d.text((cx - pw//2, mid_y - ph//2 - 2), "+", font=F_CONN, fill=_AMBER_B)
                y += CONN_H + GAP
            else:
                y += GAP

        # ── Parlay odds badge (centered, matching single-pick style) ──────
        y += 10
        is_pos = not odds_s.startswith("-")
        ow, oh = _tw_b(d, odds_s, F_MONOB)
        ox = (W - ow - 32) // 2
        d.rounded_rectangle([(ox, y),(ox+ow+32, y+oh+16)], radius=10,
                             fill=acc_bg if result else (_GREEN_BGB if is_pos else (18,20,38)))
        d.rounded_rectangle([(ox, y),(ox+ow+32, y+oh+16)], radius=10,
                             outline=(acc_col if result else (_GREEN_B if is_pos else (40,50,80))),
                             width=1)
        d.text((ox+16, y+8), odds_s, font=F_MONOB,
               fill=acc_col if result else (_GREEN_B if is_pos else _ICE_B))
        y += oh + 36

        # Result badge (text, no emoji)
        badge_col = acc_col if result else _MUTED_B
        _cx_b(d, badge_t, F_LBL, W, y, badge_col)
        y += 40

        # Book label
        _cx_b(d, book.upper(), F_MICRO, W, y, _MUTED_B)

        # ── Footer ────────────────────────────────────────────────────────
        yf = H - 60
        d.line([(PAD, yf),(W-PAD, yf)], fill=_RULE_B, width=1)
        _cx_b(d, "dubclub.win  ·  Laboy Picks  ·  BSN", F_MICRO, W, yf+14, _MUTED_B)

        img   = _png_vignette_b(img, W, H)
        fname = f"Laboy BSN Parlay {pick_date} #{pick_id}.png"
        fpath = os.path.join(SCRIPT_DIR, fname)
        img.save(fpath, "PNG", dpi=(300, 300))
        print(f"  🖼️  Parlay PNG: {fname}")
        return fpath
    except Exception as e:
        print(f"  ❌ export_parlay_png error: {e}")
        import traceback; traceback.print_exc()
        return None


def export_log_pick_html(entry):
    """
    Genera 'Laboy Pick YYYY-MM-DD #N.html' con:
      Si es parlay → delega a export_log_parlay_html()
      Card 1 — pick card (mismo diseño que picks/lines HTML)
      Card 2 — análisis (si hay texto)
    Retorna el path o None si falla.
    """
    # Parlays get their own renderer
    if entry.get("type") == "parlay":
        return export_log_parlay_html(entry)

    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    try:
        pick_date  = entry.get("date", TARGET_DATE)
        pick_id    = entry.get("id", 0)
        game       = entry.get("game", "")
        pick       = entry.get("pick", "")
        odds_v     = entry.get("odds", 0)
        analysis   = entry.get("analysis", "")
        result     = entry.get("result")
        pnl        = entry.get("pnl")
        book       = entry.get("book", "BetMGM")   # sportsbook donde se jugó
        stake_disp = _fmt_stake(entry)    # "$15.00" o "1u" (legacy)

        # Normalizar game string → "TEAM1 vs. TEAM2"
        def _fmt_game(g):
            """Convierte 'LEONES OSOS', 'LEONES @ OSOS', 'LEONES VS OSOS' → 'LEONES vs. OSOS'"""
            g = g.strip()
            for sep in [" VS. ", " vs. ", " VS ", " vs ", " @ ", " - "]:
                if sep in g:
                    parts = g.split(sep, 1)
                    return f"{parts[0].strip()} vs. {parts[1].strip()}"
            # Sin separador reconocido — intentar dividir por equipos BSN conocidos
            for t in sorted(BSN_TEAMS, key=len, reverse=True):
                if g.startswith(t + " "):
                    rest = g[len(t):].strip()
                    if rest:
                        return f"{t} vs. {rest}"
            return g

        game_disp = _fmt_game(game)

        # Detectar equipo para logo y color
        pick_upper = pick.upper()
        is_total   = any(kw in pick_upper for kw in ("OVER","UNDER")) or bool(re.match(r'^[OU][\s]?[\d.]', pick_upper))
        team_name  = None
        if not is_total:
            for t in BSN_TEAMS:
                if t in pick_upper:
                    team_name = t; break
        if not team_name and not is_total:
            parts = game_disp.split(" vs. ")
            team_name = parts[0].strip() if parts else None

        if is_total:
            logo_html = bsn_logo_html("over_under", 64)
            color     = "#f97316"
        elif team_name:
            logo_html = bsn_logo_html(team_name, 64)
            color     = _bsn_team_color(team_name)
        else:
            logo_html = ""
            color     = "#f07820"

        odds_fmt = _fmt_odds(odds_v)
        dt        = datetime.strptime(pick_date, "%Y-%m-%d")
        dstr      = dt.strftime("%A, %B %d · %Y").upper()
        yr        = dt.strftime("%Y")

        # Result color + badge
        if result == "W":
            result_color = "#22c55e"
            result_html  = f'<span style="background:#22c55e22;color:#22c55e;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✓ WIN</span>'
        elif result == "L":
            result_color = "#ef4444"
            result_html  = f'<span style="background:#ef444422;color:#ef4444;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">✗ LOSS</span>'
        elif result == "P":
            result_color = "#94a3b8"
            result_html  = f'<span style="background:#94a3b822;color:#94a3b8;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">— PUSH</span>'
        else:
            result_color = color   # color equipo mientras pendiente
            result_html  = f'<span style="background:#f0782022;color:#f07820;border-radius:6px;padding:3px 10px;font-size:0.85rem;font-weight:700">⏳ PENDING</span>'

        card_border = result_color

        # Tint de fondo AI-style según resultado
        if result == "W":
            card_bg_style = "background:linear-gradient(140deg,#061410 0%,#07080d 60%)"
            _glow_col = "#22c55e"
        elif result == "L":
            card_bg_style = "background:linear-gradient(140deg,#14060a 0%,#07080d 60%)"
            _glow_col = "#ef4444"
        elif result == "P":
            card_bg_style = "background:linear-gradient(140deg,#090a10 0%,#07080d 60%)"
            _glow_col = "#94a3b8"
        else:
            card_bg_style = "background:linear-gradient(140deg,#090910 0%,#07080d 100%)"
            _glow_col = color

        pnl_html = ""
        if pnl is not None:
            pnl_col  = "#22c55e" if pnl >= 0 else "#ef4444"
            pnl_disp = _fmt_pnl(entry)
            pnl_html = f'<div class="stat"><div class="stat-label">P&L</div><div class="stat-val" style="color:{pnl_col}">{esc(pnl_disp)}</div></div>'

        # Card 1 — pick (AI style)
        card1 = f"""
        <div class="pick-card" style="border-left:4px solid {card_border};{card_bg_style};
             border-top:1px solid {_glow_col}22;
             box-shadow:0 0 0 1px {_glow_col}10,0 0 28px {_glow_col}0a,0 6px 32px rgba(0,0,0,.85)">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
            <div style="font-size:0.62rem;color:#4a6272;font-family:monospace;letter-spacing:0.08em">
              <i class="fa-solid fa-ticket fa-icon"></i>PICK&nbsp;#{pick_id}&nbsp;&nbsp;·&nbsp;&nbsp;{esc(pick_date)}
            </div>
            {result_html}
          </div>
          <div class="teams-row">
            {logo_html}
            <div>
              <div class="game-label"><i class="fa-solid fa-basketball fa-icon"></i>{esc(game_disp)}</div>
              <div class="pick-label">{esc(pick)} <span class="odds-badge">{esc(odds_fmt)}</span></div>
            </div>
          </div>
          <div class="stats-grid">
            <div class="stat"><div class="stat-label">{esc(book)}</div><div class="stat-val" style="color:#f07820">{esc(odds_fmt)}</div></div>
            <div class="stat"><div class="stat-label">Apostado</div><div class="stat-val">{esc(stake_disp)}</div></div>
            <div class="stat"><div class="stat-label">Fecha</div><div class="stat-val" style="font-size:0.75rem">{esc(pick_date)}</div></div>
            {pnl_html if pnl_html else '<div class="stat"><div class="stat-label">Resultado</div><div class="stat-val">—</div></div>'}
          </div>
        </div>"""

        # Card 2 — análisis (AI dark)
        card2 = ""
        if analysis and analysis.strip():
            analysis_html = esc(analysis).replace("\n", "<br>")
            card2 = f"""
        <div class="pick-card" style="border-left:4px solid {card_border};
             background:linear-gradient(150deg,#070812 0%,#06060a 100%);
             box-shadow:0 0 0 1px rgba(0,220,255,.05),0 4px 24px rgba(0,0,0,.8)">
          <div class="section-title" style="margin-top:0;margin-bottom:12px">
            <i class="fa-solid fa-magnifying-glass-chart fa-icon"></i>Análisis
          </div>
          <div style="font-size:0.88rem;line-height:1.85;color:#c8d4e0;
                      letter-spacing:0.015em;white-space:pre-wrap">{analysis_html}</div>
        </div>"""

        body = card1 + card2

        html  = _bsn_html_wrap(f"Laboy Pick #{pick_id} · {dstr}", "BSN", dstr, yr, body)
        fname = f"Laboy Pick {pick_date} #{pick_id}.html"
        fpath = os.path.join(SCRIPT_DIR, fname)
        if os.path.exists(fpath) and not FORCE_EXPORT:
            print(f"  🔒 BSN Pick HTML #{pick_id} ya existe — protegido de sobreescritura.")
            print(f"     → {fname}")
            print(f"     Usa --force-export para regenerar.")
            return fpath
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(html)

        # Auto-generar JPG del mismo HTML
        jpg_path = html_to_jpg(fpath)
        if jpg_path:
            print(f"  🖼️  JPG: {os.path.basename(jpg_path)}")

        return fpath

    except Exception as e:
        print(f"     ⚠️  HTML export error: {e}")
        return None

def _publish_update_index(repo):
    """
    1. Mantiene index.html COMPLETAMENTE VACÍO — nadie puede browsear el directorio.
    2. Regenera dashboard-{DASHBOARD_TOKEN}.html con listado de todos los archivos
       publicados — solo tú tienes esa URL secreta.
    """
    import glob as _glob

    # ── 1. .nojekyll — desactiva Jekyll para que GitHub Pages sirva manifest.json ─
    with open(os.path.join(repo, ".nojekyll"), "w") as f:
        f.write("")

    # ── 2. index.html — completamente vacío ──────────────────────────────────
    blank_html = (
        "<!DOCTYPE html>\n"
        '<html lang="es"><head><meta charset="UTF-8">\n'
        "<title>Laboy Picks BSN</title>\n"
        "<style>*{box-sizing:border-box;margin:0;padding:0}"
        "html,body{height:100%;background:#0a0a0a}</style>\n"
        "</head><body></body></html>"
    )
    with open(os.path.join(repo, "index.html"), "w", encoding="utf-8") as f:
        f.write(blank_html)

    # ── 2. Dashboard secreto ──────────────────────────────────────────────────
    all_html = sorted(
        _glob.glob(os.path.join(repo, "Laboy BSN *.html")),
        key=os.path.getmtime,
        reverse=True,
    )

    def _fname_icon(fname):
        if "Lines" in fname:       return "📊"
        if "Model Card" in fname:  return "🏆"
        if "Record Card" in fname: return "📈"
        return "🎯"

    def _fname_label(fname):
        base = os.path.basename(fname)
        name = base[:-5]  # strip .html
        # Strip token suffix (-xxxxxxx at end, 8 chars)
        if len(name) > 8 and name[-8] == "-":
            name = name[:-8]
        name = name.replace("Laboy BSN ", "", 1)
        return f"{_fname_icon(base)} {name}"

    rows_html = ""
    for hp in all_html:
        base    = os.path.basename(hp)
        encoded = base.replace(" ", "%20")
        url     = f"{GITHUB_PAGES_URL}/{encoded}"
        label   = _fname_label(base)
        mtime   = datetime.fromtimestamp(os.path.getmtime(hp)).strftime("%b %d · %H:%M")
        rows_html += (
            f'<a href="{url}" target="_blank" class="item">'
            f'  <span class="label">{label}</span>'
            f'  <span class="ts">{mtime}</span>'
            f'</a>\n'
        )

    if not rows_html:
        rows_html = '<p class="empty">No hay archivos publicados aún.</p>'

    dash_html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>🏀 Laboy Picks — BSN Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a0a;color:#e2e8f0;font-family:'Inter',sans-serif;
        min-height:100vh;padding:24px 16px 48px}}
  .header{{text-align:center;margin-bottom:32px;padding-top:24px}}
  .header h1{{font-size:1.6rem;font-weight:900;letter-spacing:-0.5px}}
  .header h1 span{{color:#f07820}}
  .header p{{font-size:0.8rem;color:#64748b;margin-top:6px}}
  .list{{max-width:520px;margin:0 auto;display:flex;flex-direction:column;gap:10px}}
  .item{{display:flex;justify-content:space-between;align-items:center;
         background:#141414;border:1px solid #1e293b;border-radius:12px;
         padding:14px 18px;text-decoration:none;color:#e2e8f0;
         transition:border-color .2s,background .2s}}
  .item:hover{{background:#1a1a1a;border-color:#f07820}}
  .label{{font-size:0.95rem;font-weight:600}}
  .ts{{font-size:0.75rem;color:#64748b;white-space:nowrap;margin-left:12px}}
  .empty{{text-align:center;color:#64748b;padding:40px 0}}
  .footer{{text-align:center;margin-top:40px;font-size:0.7rem;color:#334155}}
</style>
</head>
<body>
  <div class="header">
    <h1>🏀 Laboy <span>BSN</span></h1>
    <p>Dashboard privado · {datetime.now().strftime("%B %d, %Y · %H:%M")}</p>
  </div>
  <div class="list">
{rows_html}  </div>
  <div class="footer">Solo tú tienes esta URL. No la compartas.</div>
</body>
</html>"""

    dash_path = os.path.join(repo, f"dashboard-{DASHBOARD_TOKEN}.html")
    with open(dash_path, "w", encoding="utf-8") as f:
        f.write(dash_html)

    # ── manifest.json — para el dashboard principal laboy-picks ────────────
    manifest = {"sport": "BSN", "base_url": GITHUB_PAGES_URL, "files": []}
    for hp in all_html[:20]:
        base = os.path.basename(hp)
        enc  = base.replace(" ", "%20")
        ftype = "picks" if "Picks" in base else ("lines" if "Lines" in base else "record")
        fsubtype = ("model_record" if ("Model Card" in base or "Model Record" in base) else "my_record") if ftype == "record" else ""
        date_m = re.search(r"(\d{4}-\d{2}-\d{2})", base)
        manifest["files"].append({
            "name": base, "url": f"{GITHUB_PAGES_URL}/{enc}",
            "type": ftype,
            "subtype": fsubtype,
            "date": date_m.group(1) if date_m else "",
        })
    # Imágenes de picks personales
    all_imgs = sorted(
        _glob.glob(os.path.join(repo, "Laboy Pick *.jpg")) +
        _glob.glob(os.path.join(repo, "Laboy Pick *.png")) +
        _glob.glob(os.path.join(repo, "Laboy BSN MyPicks*.png")) +
        _glob.glob(os.path.join(repo, "Laboy BSN Parlay*.png")),
        key=os.path.getmtime, reverse=True,
    )
    # Load BSN picks log for game names
    _bsn_pick_game = {}
    try:
        with open(LOG_FILE, encoding="utf-8") as _lf:
            _bsn_log_data = json.load(_lf)
        _bsn_entries = _bsn_log_data if isinstance(_bsn_log_data, list) else _bsn_log_data.get("picks", [])
        for _e in _bsn_entries:
            _bsn_pick_game[int(_e["id"])] = _e.get("game", "")
    except Exception:
        pass
    for ip in all_imgs[:30]:
        base   = os.path.basename(ip)
        enc    = base.replace(" ", "%20").replace("#", "%23")
        date_m = re.search(r"(\d{4}-\d{2}-\d{2})", base)
        img_date = date_m.group(1) if date_m else ""
        _real_today_bsn = date.today().strftime("%Y-%m-%d")
        img_subtype = "today" if img_date == _real_today_bsn else ("archive" if img_date else "")
        # Extract pick id (#N) from filename and look up game
        id_m   = re.search(r"#(\d+)", base)
        img_game = _bsn_pick_game.get(int(id_m.group(1)), "") if id_m else ""
        manifest["files"].append({
            "name": base, "url": f"{GITHUB_PAGES_URL}/{enc}",
            "type": "mypick_img", "subtype": img_subtype,
            "date": img_date, "game": img_game,
        })

    with open(os.path.join(repo, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)


def cmd_publish(html_paths):
    """
    Copia los HTMLs generados al repo local de GitHub Pages y hace push.
    Uso: python3 bsn.py --export-html --publish
         python3 bsn.py --grade-picks DATE --publish
         python3 bsn.py --export-record --publish

    Requiere:
      - Repo clonado localmente en GITHUB_PAGES_REPO (o env BSN_GITHUB_REPO)
        git clone https://github.com/laboywebsite-lgtm/bsn-picks ~/repos/bsn-picks
      - git configurado con acceso push al repo
    """
    import shutil, subprocess, glob as _glob

    repo      = GITHUB_PAGES_REPO
    clone_url = "https://github.com/laboywebsite-lgtm/bsn-picks"

    if not os.path.isdir(repo):
        print(f"\n  📥 Repo no encontrado localmente. Clonando...")
        print(f"     {clone_url}  →  {repo}")
        parent = os.path.dirname(repo)
        os.makedirs(parent, exist_ok=True)
        result = subprocess.run(["git", "clone", clone_url, repo],
                                capture_output=True, text=True)
        if result.returncode != 0:
            print(f"\n  ❌ Error al clonar el repo:")
            print(f"     {result.stderr.strip()}")
            print(f"\n  Asegúrate de tener acceso push a:")
            print(f"     {clone_url}")
            return
        print(f"  ✅ Repo clonado exitosamente.\n")

    copied = []
    for hp in (html_paths or []):
        if hp and os.path.isfile(hp):
            dest = os.path.join(repo, os.path.basename(hp))
            shutil.copy2(hp, dest)
            copied.append(os.path.basename(hp))

    # Copiar imágenes de picks personales (JPG/PNG)
    for img_pat in ["Laboy Pick *.jpg", "Laboy Pick *.png",
                    "Laboy BSN MyPicks*.png", "Laboy BSN Parlay*.png"]:
        for img in _glob.glob(os.path.join(SCRIPT_DIR, img_pat)):
            dest = os.path.join(repo, os.path.basename(img))
            shutil.copy2(img, dest)

    if not copied:
        print("\n  ⚠️  No hay HTMLs para publicar. Corre con --export-html primero.")
        return

    # ── Regenerar index.html (vacío) + dashboard secreto ───────────────────
    _publish_update_index(repo)

    # ── git add / commit / pull --rebase / push ──────────────────────────────
    def _git(args):
        r = subprocess.run(["git", "-C", repo] + args,
                           capture_output=True, text=True)
        return r.returncode, r.stdout.strip(), r.stderr.strip()

    _git(["add", "--all"])
    msg = f"🏀 BSN {TARGET_DATE} — {', '.join(copied)}"
    code, out, err = _git(["commit", "-m", msg])
    if code != 0 and "nothing to commit" in (out + err):
        print("\n  ℹ️  Sin cambios nuevos en el repo (archivos idénticos).")
    elif code != 0:
        print(f"\n  ❌ git commit falló: {err or out}")
        return

    # Pull antes de push para evitar "rejected — fetch first"
    # Saltarse si el repo remoto está vacío (primer push)
    code_ls, out_ls, _ = _git(["ls-remote", "--heads", "origin"])
    repo_vacio = (out_ls.strip() == "")

    if repo_vacio:
        print("  🆕 Repo remoto vacío — primer push, saltando pull...")
        # Asegurarse de que la rama local se llame 'main'
        _git(["checkout", "-B", "main"])
    else:
        print("  🔄 git pull --rebase...")
        code, out, err = _git(["pull", "--rebase"])
        if code != 0:
            print(f"\n  ❌ git pull falló: {err or out}")
            print(f"     Resuelve conflictos manualmente y vuelve a intentar.")
            return

    code, out, err = _git(["push", "-u", "origin", "main"] if repo_vacio else ["push"])
    if code != 0:
        print(f"\n  ❌ git push falló: {err or out}")
        print(f"     Verifica que tienes acceso SSH/HTTPS configurado.")
        return

    print(f"\n  ✅ Publicado en GitHub Pages!")
    for fname in copied:
        encoded = fname.replace(" ", "%20")
        print(f"  🌐 {GITHUB_PAGES_URL}/{encoded}")
    print(f"\n  📱 Dashboard privado:")
    print(f"     {GITHUB_PAGES_URL}/dashboard-{DASHBOARD_TOKEN}.html")
    print(f"     (Guarda este enlace en Safari / iPhone)")
    print()

# ──────────────────────────────────────────────────────
# INJURY MANAGEMENT
# ──────────────────────────────────────────────────────

def _recompute_injury_impact(wb, ir_entries):
    """
    Recalcula INJURY IMPACT - BSN.
    - Suma impactos por equipo desde ir_entries (solo BSN teams)
    - Escribe col G del IR como valor plano (no depende de la fórmula Excel)
    - Escribe el total en INJURY IMPACT sheet col C
    """
    # ── Recalcular impacts en IR sheet como valores planos ────────────────
    ws_ir  = wb[IR_SHEET]
    totals = {}
    for row in ws_ir.iter_rows(min_row=2, max_row=ws_ir.max_row):
        raw_team = row[1].value
        rate_v   = row[3].value
        ppg_v    = row[4].value
        usg_v    = row[5].value
        if not raw_team: continue
        team = norm_team(str(raw_team))
        try:
            rate   = int(rate_v) if rate_v else 0
            ppg    = float(ppg_v) if ppg_v else 0.0
            usg    = float(usg_v) if usg_v else 0.0
            rf     = RATE_FACTOR.get(rate, 0.0)
            impact = round(ppg * usg * rf, 4)
            row[6].value = impact   # col G — siempre plain number
            if team and team in BSN_TEAMS:
                totals[team] = totals.get(team, 0.0) + impact
        except Exception:
            row[6].value = 0.0

    # ── Actualizar INJURY IMPACT sheet ───────────────────────────────────
    ws_ii = wb[INJ_IMPACT_SHEET]
    for row in ws_ii.iter_rows(min_row=4, max_row=ws_ii.max_row):
        raw_team = row[1].value
        if not raw_team: continue
        team = norm_team(str(raw_team))
        if team:
            row[2].value = round(totals.get(team, 0.0), 4)   # col C


def _preserve_advanced_blend(wb):
    """
    Escribe (o restaura) las fórmulas de blend 80/20 en cols D-F del BSN Advanced sheet.
    D = OffRtg blend (80/20), E = DefRtg blend (80/20), F = PACE 2026 (100%).
    Usa VLOOKUP por col H (nombre completo 2026) para mapear a la tabla 2025 (cols M:P).
    """
    try:
        ws = wb[BSN_ADV_SHEET]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            if not row[2].value:
                continue   # col C = team short name
            r = row[0].row
            row[3].value = f"=0.8*I{r}+0.2*VLOOKUP(H{r},$M$5:$P$16,2,FALSE)"  # D = OffRtg
            row[4].value = f"=0.8*J{r}+0.2*VLOOKUP(H{r},$M$5:$P$16,3,FALSE)"  # E = DefRtg
            row[5].value = f"=K{r}"  # F = PACE (100% 2026, sin blend)
    except Exception:
        pass

def cmd_add_injury(wb):
    """
    --add-injury TEAM PLAYER RATE [PPG] [USG_PCT]
    Agrega o actualiza jugador en IR - BSN y recalcula injury impact.
    Si PPG y USG_PCT se pasan como args, omite el scrape de RealGM.
    """
    try:
        fi     = sys.argv.index("--add-injury")
        team_a = sys.argv[fi+1].upper().strip()
        player = sys.argv[fi+2].upper().strip()
        rate   = int(sys.argv[fi+3])
        assert rate in (1,2,3)
    except (ValueError, IndexError, AssertionError):
        print("  ❌ Uso: python3 bsn.py --add-injury TEAM PLAYER RATE [PPG] [USG_PCT]")
        print("     Ej: python3 bsn.py --add-injury GIGANTES WATERS 1 18.5 22.3")
        print("     RATE: 1=Out  2=Doubtful  3=Limited")
        return

    # Optional direct PPG / USG args — skip RealGM scrape when provided
    _ppg_direct = None
    _usg_direct = None
    try:
        if len(sys.argv) > fi+4:
            _ppg_direct = float(sys.argv[fi+4])
        if len(sys.argv) > fi+5:
            _usg_direct = float(sys.argv[fi+5]) / 100.0  # convert pct → ratio
    except (ValueError, IndexError):
        pass

    team = norm_team(team_a) or team_a
    rate_lbl = {1:"OUT",2:"DOUBTFUL",3:"LIMITED"}[rate]
    rf       = RATE_FACTOR[rate]

    print(f"\n  🏥 Injury: {team} — {player} — Rate {rate} ({rate_lbl})")

    # ── PPG ─────────────────────────────────────────────
    if _ppg_direct is not None:
        ppg = _ppg_direct
        print(f"  ✅ PPG (manual): {ppg:.1f}")
    else:
        print(f"  Buscando PPG en RealGM...")
        ppg_result = scrape_player_ppg(player, team)
        if ppg_result:
            _, _, ppg = ppg_result
            print(f"  ✅ PPG encontrado: {ppg:.1f}")
        else:
            print(f"  ⚠️  No se encontró PPG en RealGM para '{player}' / '{team}'")
            try:
                ppg = float(input("     Ingresa PPG manualmente: ").strip())
            except (ValueError, EOFError):
                print("  ❌ Operación cancelada."); return

    # ── USG% ────────────────────────────────────────────
    if _usg_direct is not None:
        usg = _usg_direct
        print(f"  ✅ USG% (manual): {usg*100:.1f}%")
    else:
        print(f"  Buscando USG% en RealGM...")
        usg_result = scrape_player_usg(player, team)
        if usg_result:
            _, usg = usg_result
            print(f"  ✅ USG% encontrado: {usg*100:.1f}%")
        else:
            print(f"  ⚠️  No se encontró USG% en RealGM para '{player}'")
            try:
                usg_pct = float(input("     Ingresa USG% manualmente (ej: 17.6): ").strip())
                usg     = usg_pct / 100
            except (ValueError, EOFError):
                print("  ❌ Operación cancelada."); return

    impact = round(ppg * usg * rf, 4)
    print(f"\n  📊 Impact = {ppg:.1f} × {usg*100:.1f}% × {rf} = {impact:.3f} pts")

    # Actualizar IR - BSN
    ws_ir   = wb[IR_SHEET]
    entries = load_ir_entries(wb)

    # Buscar si ya existe (mismo equipo + jugador)
    found   = False
    for row in ws_ir.iter_rows(min_row=2, max_row=ws_ir.max_row):
        rt = norm_team(str(row[1].value or ""))
        rp = str(row[2].value or "").strip().upper()
        if rt == team and (player in rp or rp in player):
            row[3].value = rate
            row[4].value = round(ppg, 2)
            row[5].value = round(usg, 4)
            row[6].value = impact
            found = True
            print(f"  ✏️  Actualizado entry existente: {rp}")
            break

    if not found:
        # Agregar nueva fila
        next_row = ws_ir.max_row + 1
        ws_ir.cell(next_row, 2).value = team
        ws_ir.cell(next_row, 3).value = player
        ws_ir.cell(next_row, 4).value = rate
        ws_ir.cell(next_row, 5).value = round(ppg, 2)
        ws_ir.cell(next_row, 6).value = round(usg, 4)
        ws_ir.cell(next_row, 7).value = impact
        print(f"  ➕ Nueva entrada agregada.")

    # Recalcular impacts y preservar blend antes de guardar
    entries_updated = load_ir_entries(wb)
    _recompute_injury_impact(wb, entries_updated)
    _preserve_advanced_blend(wb)   # ← evita pérdida de caché de fórmulas Excel

    wb.save(EXCEL_FILE)
    print(f"\n  💾 Excel actualizado: {os.path.basename(EXCEL_FILE)}")
    print(f"  📋 Injury impact {team}: {impact:.3f} pts afectados\n")


def cmd_remove_injury(wb):
    """--remove-injury TEAM PLAYER"""
    try:
        fi     = sys.argv.index("--remove-injury")
        team_a = sys.argv[fi+1].upper().strip()
        player = sys.argv[fi+2].upper().strip()
    except (ValueError, IndexError):
        print("  ❌ Uso: python3 bsn.py --remove-injury TEAM PLAYER")
        print("     Ej: python3 bsn.py --remove-injury GIGANTES WATERS")
        return

    team = norm_team(team_a) or team_a
    ws_ir = wb[IR_SHEET]
    rows_to_delete = []
    for row in ws_ir.iter_rows(min_row=2, max_row=ws_ir.max_row):
        rt = norm_team(str(row[1].value or ""))
        rp = str(row[2].value or "").strip().upper()
        if rt == team and (player in rp or rp in player):
            rows_to_delete.append(row[0].row)

    if not rows_to_delete:
        print(f"  ❌ No se encontró '{player}' / '{team}' en el IR.\n")
        return

    for r in sorted(rows_to_delete, reverse=True):
        ws_ir.delete_rows(r)

    entries_updated = load_ir_entries(wb)
    _recompute_injury_impact(wb, entries_updated)
    _preserve_advanced_blend(wb)   # ← evita pérdida de caché de fórmulas Excel

    wb.save(EXCEL_FILE)
    print(f"  ✅ {player} ({team}) eliminado del IR y Excel actualizado.\n")


# ──────────────────────────────────────────────────────
# REFRESH 2026 STATS FROM REALGM
# ──────────────────────────────────────────────────────

def cmd_refresh(wb):
    """Actualiza stats 2026 en BSN - Advanced desde RealGM y recalcula blend."""
    print(f"\n  🌐 Scraping RealGM — BSN Advanced Stats 2026...")
    new_stats = scrape_team_advanced_stats()

    if not new_stats:
        print("  ❌ No se pudo obtener data de RealGM.")
        print("     Verifica tu conexión o que la URL sea accesible.")
        return

    ws     = wb[BSN_ADV_SHEET]
    BLEND  = 0.80   # 80% 2026 + 20% 2025

    updated = 0
    # Leer 2025 data (col M=team, N=ortg, O=drtg, P=pace)  para blend
    stats_2025 = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        t25 = row[12]  # col M
        if not t25: continue
        team25 = norm_team(str(t25))
        if team25:
            try:
                stats_2025[team25] = {
                    "ortg": float(row[13]) if row[13] else None,   # col N
                    "drtg": float(row[14]) if row[14] else None,   # col O
                    "pace": float(row[15]) if row[15] and str(row[15]).strip() not in ("","  ") else None,  # col P
                }
            except: pass

    # Actualizar 2026 y blend en hoja
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        raw_blend = row[2].value   # col C = team blend
        if not raw_blend: continue
        team = norm_team(str(raw_blend))
        if not team: continue

        s26  = new_stats.get(team)
        s25  = stats_2025.get(team)
        if not s26: continue

        # Actualizar cols H-K (2026: team, ortg, drtg, pace)
        row[7].value  = team             # col H
        if s26["ortg_2026"]: row[8].value  = s26["ortg_2026"]   # col I
        if s26["drtg_2026"]: row[9].value  = s26["drtg_2026"]   # col J
        if s26["pace_2026"]: row[10].value = s26["pace_2026"]   # col K

        # Escribir fórmulas blend en cols D-F (preserva lógica Excel)
        r = row[0].row
        row[3].value = f"=0.8*I{r}+0.2*VLOOKUP(H{r},$M$5:$P$16,2,FALSE)"  # col D = OffRtg
        row[4].value = f"=0.8*J{r}+0.2*VLOOKUP(H{r},$M$5:$P$16,3,FALSE)"  # col E = DefRtg
        row[5].value = f"=K{r}"  # col F = PACE (100% 2026, sin blend)

        updated += 1

    wb.save(EXCEL_FILE)
    print(f"  ✅ {updated} equipos actualizados con stats 2026 de RealGM.")
    print(f"  💾 Guardado: {os.path.basename(EXCEL_FILE)}\n")


# ──────────────────────────────────────────────────────
# PNG / PDF EXPORTS  (Pillow + img2pdf)
# ──────────────────────────────────────────────────────

try:
    from PIL import Image, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

_FONTS_DIR = os.path.join(SCRIPT_DIR, "..", ".claude", "skills", "canvas-design", "canvas-fonts")

def _fnt_bsn(name, size):
    if not HAS_PIL: return None
    try:    return ImageFont.truetype(os.path.join(_FONTS_DIR, name), size)
    except:
        try:    return ImageFont.load_default(size=size)
        except: return ImageFont.load_default()

# Paleta (identical to mlb.py)
_BG_B      = (6,  8, 15)
_PANEL_B   = (12, 15, 24)
_AMBER_B   = (240, 95,  8)
_AMBER_MB  = (140, 52,  4)
_ICE_B     = (228, 238, 255)
_ICE_DB    = (120, 135, 168)
_MUTED_B   = (58,  68, 98)
_GREEN_B   = (52,  211, 108)
_GREEN_BGB = (8,   40,  22)
_RED_B     = (239, 68,  68)
_RULE_B    = (16,  20,  36)

# BSN team colors (by their normalized names)
_BSN_COLORS = {
    "ATLETICOS": (0, 85, 164), "CANGREJEROS": (200, 16, 46),
    "CAPITANES":  (0, 56, 168), "CRIOLLOS":   (139, 0,  0),
    "GIGANTES":   (0, 128, 0),  "INDIOS":     (255, 140, 0),
    "LEONES":     (255, 215, 0),"MARATONISTAS":(70,130,180),
    "METS":       (0,  45, 114),"PIRATAS":    (39, 37, 31),
    "SANTEROS":   (148, 0, 211),"VAQUEROS":   (192, 160, 0),
    "BRUJOS":     (75,  0, 130),"TROTAMUNDOS":(0, 100, 160),
}

def _tw_b(d, text, fnt):
    bb = d.textbbox((0,0), text, font=fnt)
    return bb[2]-bb[0], bb[3]-bb[1]

def _cx_b(d, text, fnt, W, y, col):
    bb = d.textbbox((0,0), text, font=fnt)
    d.text(((W-(bb[2]-bb[0]))//2, y), text, font=fnt, fill=col)

def _png_dot_bg_b(d, W, H):
    for gy in range(0, H, 54):
        for gx in range(0, W, 54):
            d.ellipse([(gx-1,gy-1),(gx+1,gy+1)], fill=(14,18,32))

def _png_vignette_b(img, W, H):
    vgn = Image.new("RGBA",(W,H),(0,0,0,0))
    vd  = ImageDraw.Draw(vgn)
    for s in range(0,90,6):
        vd.rectangle([(s,s),(W-s,H-s)], outline=(0,0,0,int(s*1.6)), width=6)
    return Image.alpha_composite(img.convert("RGBA"), vgn).convert("RGB")


def _draw_bsn_lines_page(rows_page, dt, dstr, page_num, total_pages):
    """Dibuja una página de BSN model lines."""
    PAD   = 48
    W     = 1080
    RH    = 116
    GAP   = 10
    HDR_H = 196
    FTR_H = 66
    n     = max(1, len(rows_page))
    H     = HDR_H + n*(RH+GAP) + FTR_H

    img = Image.new("RGB", (W, H), _BG_B)
    d   = ImageDraw.Draw(img)
    _png_dot_bg_b(d, W, H)

    F_HERO  = _fnt_bsn("BigShoulders-Bold.ttf",   84)
    F_SUB   = _fnt_bsn("IBMPlexMono-Bold.ttf",    22)
    F_DATE  = _fnt_bsn("GeistMono-Regular.ttf",   18)
    F_TEAM  = _fnt_bsn("BigShoulders-Bold.ttf",   32)
    F_NUM   = _fnt_bsn("BigShoulders-Bold.ttf",   52)
    F_LBL   = _fnt_bsn("GeistMono-Regular.ttf",   17)
    F_SMALL = _fnt_bsn("GeistMono-Regular.ttf",   15)
    F_MONOB = _fnt_bsn("IBMPlexMono-Bold.ttf",    20)

    # ── Header ──
    y = 16
    _cx_b(d, "LABOY PICKS", F_HERO, W, y, _ICE_B)
    y += 108
    _cx_b(d, "BSN — MODEL LINES", F_SUB, W, y, _AMBER_B)
    y += 34
    page_s = f"Pág. {page_num}/{total_pages}   ·   " if total_pages > 1 else ""
    _cx_b(d, f"{page_s}{dstr}", F_DATE, W, y, _MUTED_B)
    y += 38
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
    y += 2

    # ── Game rows ──
    for g in rows_page:
        r      = g["result"]
        t1     = g.get("team1","?").upper()
        t2     = g.get("team2","?").upper()
        gtime  = g.get("game_time","")
        spread = r.get("spread_line","—")
        total  = r.get("total", 0)
        wp1    = r.get("wp1", "—")
        wp2    = r.get("wp2", "—")

        d.rounded_rectangle([(PAD, y+8),(W-PAD, y+RH-4)], radius=10, fill=_PANEL_B)
        d.rounded_rectangle([(PAD+1,y+9),(W-PAD-1,y+RH-5)], radius=9,
                            outline=(22,28,50), width=1)

        # Amber left accent strip
        c1  = _BSN_COLORS.get(t1, _AMBER_MB)
        d.rounded_rectangle([(PAD, y+8),(PAD+4, y+RH-4)], radius=2, fill=c1)

        # Time chip
        if gtime:
            d.rounded_rectangle([(PAD+14, y+16),(PAD+90, y+34)], radius=6, fill=(16,20,38))
            d.text((PAD+20, y+18), gtime, font=F_SMALL, fill=_MUTED_B)

        # Teams
        SPLIT = W // 2
        d.text((PAD+14, y+38), t1, font=F_TEAM, fill=_ICE_B)
        d.text((PAD+14, y+72), f"@ {t2}", font=F_LBL, fill=_ICE_DB)

        # Total (big number, right of split)
        tot_s   = f"{total:.1f}"
        tot_col = _GREEN_B if total >= 200 else _ICE_B
        d.text((SPLIT+10, y+24), "TOTAL", font=F_SMALL, fill=_MUTED_B)
        d.text((SPLIT+10, y+38), tot_s, font=F_NUM, fill=tot_col)

        # Spread + WP
        d.text((SPLIT+120, y+24), "SPREAD", font=F_SMALL, fill=_MUTED_B)
        d.text((SPLIT+120, y+40), str(spread), font=F_MONOB, fill=_ICE_DB)
        d.text((SPLIT+120, y+64), f"{wp1}%  /  {wp2}%", font=F_SMALL, fill=_MUTED_B)

        y += RH + GAP

    # ── Footer ──
    y += 4
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
    y += 12
    _cx_b(d, "dubclub.win  ·  Laboy Picks Data Model  ·  BSN", F_SMALL, W, y, _MUTED_B)

    return img


def _draw_bsn_picks_page(picks_page, dt, dstr, page_num, total_pages):
    """Dibuja una página de BSN logged picks (1080×1920)."""
    W, H = 1080, 1920
    PAD  = 54

    img = Image.new("RGB", (W, H), _BG_B)
    d   = ImageDraw.Draw(img)
    _png_dot_bg_b(d, W, H)

    d.rectangle([(0,0),(W,8)], fill=_AMBER_B)
    d.rectangle([(0,H-8),(W,H)], fill=_AMBER_B)

    F_HERO  = _fnt_bsn("BigShoulders-Bold.ttf", 96)
    F_SUB   = _fnt_bsn("IBMPlexMono-Bold.ttf",  22)
    F_DATE  = _fnt_bsn("GeistMono-Regular.ttf", 18)
    F_TEAM  = _fnt_bsn("BigShoulders-Bold.ttf", 52)
    F_PICK  = _fnt_bsn("BigShoulders-Bold.ttf", 68)
    F_MONOB = _fnt_bsn("IBMPlexMono-Bold.ttf",  30)
    F_LBL   = _fnt_bsn("GeistMono-Regular.ttf", 22)
    F_MICRO = _fnt_bsn("GeistMono-Regular.ttf", 18)

    # Corner marks
    M, L = 40, 28
    for (x1,y1),(x2,y2) in [
        [(M,M),(M+L,M)],[(M,M),(M,M+L)],
        [(W-M-L,M),(W-M,M)],[(W-M,M),(W-M,M+L)],
        [(M,H-M),(M+L,H-M)],[(M,H-M-L),(M,H-M)],
        [(W-M-L,H-M),(W-M,H-M)],[(W-M,H-M-L),(W-M,H-M)],
    ]:
        d.line([(x1,y1),(x2,y2)], fill=_AMBER_MB, width=2)

    y = 40
    _cx_b(d, "LABOY PICKS", F_HERO, W, y, _ICE_B)
    y += 112
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
    y += 16
    _cx_b(d, f"BSN  ·  {dstr}", F_MICRO, W, y, _MUTED_B)
    y += 40

    page_note = f"  ({page_num}/{total_pages})" if total_pages > 1 else ""
    _cx_b(d, f"EV+ PICKS{page_note}", F_SUB, W, y, _AMBER_B)
    y += 44
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
    y += 28

    CARD_H = 260
    GAP    = 20

    for entry in picks_page:
        game   = entry.get("game","").upper()
        pick   = entry.get("pick","").upper()
        odds_v = entry.get("odds", 0)
        odds_s = f"{odds_v:+d}" if odds_v else "—"
        units  = entry.get("units", 1)
        idx    = entry.get("id", "?")
        etime  = entry.get("time","")

        # Card background
        d.rounded_rectangle([(PAD, y),(W-PAD, y+CARD_H)], radius=14, fill=_PANEL_B)
        d.rounded_rectangle([(PAD+1,y+1),(W-PAD-1,y+CARD_H-1)], radius=13,
                            outline=(22,28,50), width=1)
        d.rounded_rectangle([(PAD,y),(PAD+4,y+CARD_H)], radius=2, fill=_AMBER_B)

        # Game header + time chip
        if etime:
            d.rounded_rectangle([(PAD+16,y+16),(PAD+100,y+38)], radius=6, fill=(16,20,38))
            d.text((PAD+22,y+18), etime, font=F_MICRO, fill=_MUTED_B)
        _cx_b(d, game, F_LBL, W, y+20, _ICE_DB)

        # Pick text
        _cx_b(d, pick, F_PICK, W, y+58, _ICE_B)

        # Odds badge
        is_pos = not odds_s.startswith("-")
        ow, oh = _tw_b(d, odds_s, F_MONOB)
        ox = (W - ow - 32) // 2
        d.rounded_rectangle([(ox,y+136),(ox+ow+32,y+136+oh+16)], radius=10,
                            fill=_GREEN_BGB if is_pos else (18,20,38))
        d.rounded_rectangle([(ox,y+136),(ox+ow+32,y+136+oh+16)], radius=10,
                            outline=(_GREEN_B if is_pos else (40,50,80)), width=1)
        d.text((ox+16,y+144), odds_s, font=F_MONOB,
               fill=_GREEN_B if is_pos else _ICE_B)

        # Units + Pick ID
        _cx_b(d, f"UNIDADES: {units}u", F_LBL, W, y+196, _MUTED_B)
        d.text((PAD+16, y+CARD_H-28), f"Pick #{idx}", font=F_MICRO, fill=_MUTED_B)

        y += CARD_H + GAP

    # Footer
    y = H - 60
    d.line([(PAD,y),(W-PAD,y)], fill=_RULE_B, width=1)
    y += 14
    _cx_b(d, "dubclub.win  ·  Laboy Picks  ·  BSN", F_MICRO, W, y, _MUTED_B)

    return img


def export_bsn_lines_png(games_data, max_per_page=6):
    """PNG(s) con model lines BSN. Retorna lista de paths."""
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export"); return []

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A %B %d · %Y").upper()
    pages = [games_data[i:i+max_per_page] for i in range(0, max(1,len(games_data)), max_per_page)]
    if not games_data: pages = [[]]
    total = len(pages)
    paths = []

    for pi, page in enumerate(pages, 1):
        fname = (f"Laboy BSN Lines {TARGET_DATE}.png" if total==1
                 else f"Laboy BSN Lines {TARGET_DATE} p{pi}.png")
        fpath = os.path.join(SCRIPT_DIR, fname)
        img   = _draw_bsn_lines_page(page, dt, dstr, pi, total)
        img   = _png_vignette_b(img, *img.size)
        img.save(fpath, "PNG", dpi=(300,300))
        print(f"  📊 BSN Lines PNG ({pi}/{total}): {fname}")
        paths.append(fpath)

    return paths


def export_bsn_log_picks_png(date_str=None):
    """PNG(s) 1080×1920 con picks logueados para TARGET_DATE. Retorna lista de paths."""
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages  para PNG export"); return []

    # BSN usa el mismo log que MLB si están en el mismo workspace, o uno propio
    bsn_log = os.path.join(SCRIPT_DIR, "laboy_bsn_picks_log.json")
    mlb_log = os.path.join(SCRIPT_DIR, "..", "MLB", "laboy_picks_log.json")
    log_path = bsn_log if os.path.exists(bsn_log) else (
               mlb_log if os.path.exists(mlb_log) else bsn_log)

    td  = date_str or TARGET_DATE
    log = []
    if os.path.exists(log_path):
        try:
            with open(log_path) as f: log = json.load(f)
        except Exception: log = []

    today_picks = [e for e in log if e.get("date","") == td
                   and e.get("sport","").upper() in ("BSN","")]

    if not today_picks:
        print(f"\n  ℹ️  No hay picks BSN logueados para {td}.")
        print(f"  Usa: python3 bsn.py --log  para registrar un pick.")
        return []

    today_picks.sort(key=lambda e: (_parse_time_sort(e.get("time","")), e.get("id",0)))

    dt   = datetime.strptime(td, "%Y-%m-%d")
    dstr = dt.strftime("%A %B %d · %Y").upper()

    # Parlays get their own dedicated PNG card — extract them separately
    parlay_picks  = [e for e in today_picks if e.get("type") == "parlay"]
    regular_picks = [e for e in today_picks if e.get("type") != "parlay"]

    paths = []

    # ── Parlay pick cards (one PNG each) ──────────────
    for e in parlay_picks:
        pp = export_parlay_png(e)
        if pp:
            paths.append(pp)

    # ── Regular picks pages ───────────────────────────
    if regular_picks:
        MAX_PER = 4
        pages   = [regular_picks[i:i+MAX_PER] for i in range(0, len(regular_picks), MAX_PER)]
        total   = len(pages)

        for pi, page in enumerate(pages, 1):
            fname = (f"Laboy BSN MyPicks {td}.png" if total==1
                     else f"Laboy BSN MyPicks {td} p{pi}.png")
            fpath = os.path.join(SCRIPT_DIR, fname)
            img   = _draw_bsn_picks_page(page, dt, dstr, pi, total)
            img   = _png_vignette_b(img, *img.size)
            img.save(fpath, "PNG", dpi=(300,300))
            print(f"  🖼️  BSN My Picks PNG ({pi}/{total}): {fname}")
            paths.append(fpath)

    return paths


def pngs_to_pdf_bsn(png_paths, out_path):
    """Convierte lista de PNGs a PDF con img2pdf. Retorna path o None."""
    try:
        import img2pdf as _img2pdf
    except ImportError:
        import subprocess as _sp, importlib as _il
        _sp.run([sys.executable, "-m", "pip", "install", "img2pdf", "-q"], check=True)
        _il.invalidate_caches()
        import img2pdf as _img2pdf

    existing = [p for p in png_paths if os.path.exists(p)]
    if not existing:
        print("  ⚠️  pngs_to_pdf: ningún PNG encontrado."); return None

    with open(out_path, "wb") as f:
        f.write(_img2pdf.convert(existing))
    print(f"  📄 PDF generado: {os.path.basename(out_path)}")
    return out_path


def crop_story_to_post_bsn(story_path, out_path=None):
    """Recorta PNG 1080×1920 a 1080×1080 centrado. Retorna path o None."""
    if not HAS_PIL:
        print("  ⚠️  pip install Pillow --break-system-packages"); return None

    img = Image.open(story_path).convert("RGB")
    w, h = img.size
    if w == h:
        out_path = out_path or story_path.replace(".png","_post.png")
        img.save(out_path,"PNG",dpi=(300,300)); return out_path

    size = min(w, h, 1080)
    left = (w - size) // 2
    top  = max(0, (h - size) // 3)
    cropped = img.crop((left, top, left+size, top+size))
    out_path = out_path or story_path.replace(".png","_post.png")
    cropped.save(out_path,"PNG",dpi=(300,300))
    print(f"  🟥 Post 1080×1080: {os.path.basename(out_path)}")
    return out_path


# ──────────────────────────────────────────────────────
# HTML EXPORT — mismo diseño que mlb.py
# ──────────────────────────────────────────────────────

def _bsn_logo_b64():
    """
    Carga laboy_logo.png desde SCRIPT_DIR (BSN folder), remueve el fondo
    negro y retorna un data-URI base64 listo para <img src="...">.
    Retorna None si no existe o hay error.
    """
    logo_path = os.path.join(SCRIPT_DIR, "laboy_logo.png")
    # Intenta también en la carpeta MLB hermana
    if not os.path.exists(logo_path):
        logo_path = os.path.join(SCRIPT_DIR, "..", "MLB", "laboy_logo.png")
    if not os.path.exists(logo_path):
        return None
    try:
        from PIL import Image
        import io, base64
        img = Image.open(logo_path).convert("RGBA")
        data = img.load()
        w, h = img.size
        for y in range(h):
            for x in range(w):
                r, g, b, a = data[x, y]
                if r < 35 and g < 35 and b < 35:
                    data[x, y] = (0, 0, 0, 0)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/png;base64,{b64}"
    except Exception:
        return None


def _bsn_team_color(team):
    """Retorna el hex de color del equipo BSN, o el accent naranja por defecto."""
    return BSN_TEAM_COLORS.get(team, "#f07820")


def _strip_accents(s):
    """Normaliza acentos: 'ATLÉTICOS' → 'ATLETICOS'"""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s))
        if unicodedata.category(c) != "Mn"
    )

def bsn_logo_html(team, size=52):
    """
    Retorna HTML de logo para un equipo BSN.
    Orden de búsqueda:
      1. SCRIPT_DIR/{team_ascii.lower()}.png       → raíz del script (over_under.png, etc.)
      2. SCRIPT_DIR/logos/{team_ascii.lower()}.png → carpeta logos
      3. SVG inline con inicial del equipo y color del equipo (fallback elegante)
    Normaliza acentos automáticamente para la búsqueda de archivo.
    """
    team_lower      = team.lower()
    team_ascii_lower = _strip_accents(team_lower)   # 'atléticos' → 'atleticos'
    logo_dir        = os.path.join(SCRIPT_DIR, "logos")
    # Build candidate filenames: try accent-stripped first, then original
    candidates = []
    for slug in dict.fromkeys([team_ascii_lower, team_lower]):   # deduplicated
        candidates += [
            os.path.join(SCRIPT_DIR, f"{slug}.png"),
            os.path.join(logo_dir,   f"{slug}.png"),
        ]
    for check_path in candidates:
        if os.path.exists(check_path):
            try:
                import base64
                with open(check_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                return (f'<img src="data:image/png;base64,{b64}" alt="{team}" '
                        f'width="{size}" height="{size}" style="object-fit:contain">')
            except Exception:
                pass
    # SVG badge fallback
    color   = _bsn_team_color(team)
    initial = team[0]
    half    = size // 2
    fs      = max(10, size // 2)
    ty      = half + fs // 3
    return (f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" '
            f'xmlns="http://www.w3.org/2000/svg">'
            f'<circle cx="{half}" cy="{half}" r="{half-1}" fill="{color}" opacity="0.9"/>'
            f'<text x="{half}" y="{ty}" text-anchor="middle" fill="white" '
            f'font-family="-apple-system,BlinkMacSystemFont,sans-serif" '
            f'font-size="{fs}" font-weight="900">{initial}</text>'
            f'</svg>')


def _bsn_html_css():
    return """
  :root{--bg:#050508;--card:#0d0d10;--accent:#f07820;--green:#22c55e;--red:#ef4444;
        --text:#e8eef4;--muted:#4a6272;--border:#151520;--cyan:#00dcff;}
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);
    background-image:linear-gradient(rgba(0,220,255,.018) 1px,transparent 1px),
                     linear-gradient(90deg,rgba(0,220,255,.018) 1px,transparent 1px);
    background-size:32px 32px;
    color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:0 0 40px}
  .section{max-width:700px;margin:0 auto;padding:32px 16px 0}
  .section-title{font-size:0.75rem;font-weight:800;letter-spacing:3px;text-transform:uppercase;
    background:linear-gradient(90deg,#00dcff,#7a8fa0);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;
    margin:28px 0 12px;padding-bottom:6px;border-bottom:1px solid var(--border)}
  .pick-card{background:linear-gradient(160deg,#0d0d10 0%,#0a0a0d 100%);border-radius:14px;
    padding:18px;margin-bottom:14px;border-left:4px solid var(--accent);
    border-top:1px solid rgba(0,220,255,.08);
    box-shadow:0 0 0 1px rgba(0,220,255,.04),0 4px 32px rgba(0,0,0,.8),
               inset 0 1px 0 rgba(255,255,255,.03)}
  .pick-time{font-size:0.7rem;color:var(--muted);margin-bottom:8px;font-family:monospace}
  .teams-row{display:flex;align-items:center;gap:12px;margin-bottom:14px}
  .game-label{font-size:0.8rem;color:var(--muted);margin-bottom:4px}
  .pick-label{font-size:1.3rem;font-weight:800}
  @keyframes bsn-odds-glow{0%,100%{box-shadow:0 0 4px rgba(240,120,32,.3)}50%{box-shadow:0 0 10px rgba(240,120,32,.6)}}
  .odds-badge{background:#f0782022;color:var(--accent);border-radius:6px;padding:2px 8px;font-size:1rem;
    animation:bsn-odds-glow 2.5s ease-in-out infinite}
  .stats-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}
  .stat{background:rgba(0,220,255,.025);border:1px solid rgba(0,220,255,.08);
    border-radius:8px;padding:8px;text-align:center}
  .stat-label{font-size:0.6rem;text-transform:uppercase;letter-spacing:1px;color:#4a6272;margin-bottom:3px}
  .stat-val{font-size:0.88rem;font-weight:700;
    font-family:'JetBrains Mono','SF Mono','Fira Code','Courier New',monospace;letter-spacing:-.5px}
  .no-picks{color:var(--muted);text-align:center;padding:24px}
  .po-notes{margin-top:10px;padding:7px 10px;background:rgba(0,220,255,.03);
    border:1px solid rgba(0,220,255,.06);border-radius:8px;
    font-size:0.68rem;color:#7a8fa0;letter-spacing:0.5px;line-height:1.6}
  .po-note{color:#c8a84b}
  .line-card{background:linear-gradient(160deg,#0d0d10 0%,#0a0a0d 100%);border-radius:12px;
    padding:14px 16px;margin-bottom:10px;
    border:1px solid rgba(0,220,255,.07);
    box-shadow:0 0 0 1px rgba(0,220,255,.03),0 4px 24px rgba(0,0,0,.7)}
  .line-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px}
  .team-logo{display:flex;align-items:center;gap:6px;font-size:0.8rem;font-weight:700}
  .team-logo svg,.team-logo img{width:36px;height:36px;flex-shrink:0}
  .line-time{font-size:0.75rem;color:var(--muted);font-family:monospace}
  .line-stats{display:flex;gap:12px;flex-wrap:wrap;font-size:0.8rem;margin-bottom:6px}
  .venue-row{font-size:0.72rem;color:#3a4f5c;margin-top:4px}
  .inj-row{font-size:0.72rem;color:var(--red);margin-top:3px}
  .footer{text-align:center;padding:32px 16px 0;color:#1a2530;font-size:0.8rem;
    border-top:1px solid rgba(0,220,255,.06);margin-top:20px;padding-top:16px}
  .footer a{color:rgba(0,220,255,.35);text-decoration:none}
  ::-webkit-scrollbar{width:5px}
  ::-webkit-scrollbar-track{background:#050508}
  ::-webkit-scrollbar-thumb{background:linear-gradient(180deg,#00dcff40,#f0782040);border-radius:3px}"""


def _bsn_html_wrap(title, header_sub, dstr, yr, body_html):
    """Envuelve body_html en el shell HTML completo — AI style (BSN)."""
    logo_src = _bsn_logo_b64()
    logo_html = (f'<img class="dbg-logo" src="{logo_src}" alt="Laboy Picks">'
                 if logo_src else '<span class="dbg-wordmark">LABOY</span>')
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
<style>{_bsn_html_css()}
  /* ── AI Header ── */
  .dbg-header{{background:linear-gradient(180deg,#000 0%,#06060a 100%);
    padding:22px 24px 18px;text-align:center;border-bottom:1px solid rgba(0,220,255,.13);
    position:relative;overflow:hidden}}
  @keyframes bsn-scan{{0%{{transform:translateY(-120%)}}100%{{transform:translateY(1200%)}}}}
  .dbg-header::before{{content:'';position:absolute;top:0;left:0;right:0;height:60px;
    background:linear-gradient(180deg,transparent,rgba(0,220,255,.09),transparent);
    animation:bsn-scan 4s linear infinite;pointer-events:none}}
  .dbg-header::after{{content:'';position:absolute;bottom:-1px;left:10%;right:10%;height:1px;
    background:linear-gradient(90deg,transparent,rgba(0,220,255,.5),rgba(240,120,32,.5),transparent)}}
  .dbg-logo{{height:80px;width:auto;display:block;margin:0 auto 10px;
    filter:drop-shadow(0 0 12px rgba(240,120,32,.45))}}
  .dbg-wordmark{{font-size:2rem;font-weight:900;letter-spacing:6px;
    background:linear-gradient(90deg,#f07820,#00dcff);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;
    display:block;margin-bottom:8px}}
  @keyframes bsn-gradient{{0%{{background-position:0% center}}100%{{background-position:200% center}}}}
  .dbg-title{{font-size:0.6rem;font-weight:800;letter-spacing:4px;text-transform:uppercase;
    margin-bottom:4px;background:linear-gradient(90deg,#00dcff,#7c3aed,#00dcff);
    background-size:200% auto;-webkit-background-clip:text;-webkit-text-fill-color:transparent;
    animation:bsn-gradient 4s linear infinite}}
  .dbg-date{{color:#475569;font-size:0.65rem;letter-spacing:2.5px;text-transform:uppercase;margin-top:2px}}
  @keyframes bsn-pulse{{0%,100%{{opacity:1;transform:scale(1)}}50%{{opacity:.4;transform:scale(.7)}}}}
  .dbg-badge{{display:inline-flex;align-items:center;gap:5px;
    background:rgba(0,220,255,.07);border:1px solid rgba(0,220,255,.2);
    border-radius:20px;padding:2px 10px;margin-top:8px;
    font-size:0.58rem;font-weight:700;letter-spacing:2px;color:rgba(0,220,255,.6);
    text-transform:uppercase}}
  .dbg-badge-dot{{width:5px;height:5px;border-radius:50%;background:#00dcff;
    box-shadow:0 0 6px #00dcff;animation:bsn-pulse 1.8s ease-in-out infinite}}
  .fa-icon{{font-size:0.85em;opacity:0.75;margin-right:4px}}
  .section-title .fa-icon{{font-size:0.9em;opacity:1;margin-right:6px}}
</style>
</head>
<body>
<div class="dbg-header">
  {logo_html}
  <div class="dbg-title">&#9632;&nbsp;Model Report&nbsp;&#9632;</div>
  <div class="dbg-date">BSN &nbsp;·&nbsp; {dstr}</div>
  <div><span class="dbg-badge"><span class="dbg-badge-dot"></span>AI Analysis Engine · Active</span></div>
</div>
<div class="section">
{body_html}
</div>
<div class="footer">
  <p>Data Model by <a href="https://instagram.com/laboypicks">Laboy Picks</a> &nbsp;·&nbsp; {yr}</p>
</div>
</body>
</html>"""


def export_bsn_picks_html(games_data):
    """
    Genera 'Laboy BSN Picks YYYY-MM-DD.html' con picks de valor real.
    Usa _find_value_picks() — misma lógica que el terminal.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")

    # Obtener picks con valor real (misma lógica que show_picks)
    picks = _find_value_picks(games_data)
    picks.sort(key=lambda x: _parse_time_sort(x.get("time", "")))

    if not picks:
        body = '<div class="no-picks">No hay picks con valor claro hoy — mercado bien alineado con el modelo.</div>'
    else:
        body = '<div class="section-title"><i class="fa-solid fa-bullseye fa-icon"></i>Picks con Valor — BSN</div>\n'
        for p in picks:
            ptype    = p.get("pick_type","")
            _raw_time = p["time"] or ""
            time_str  = (esc(_raw_time) + " AST") if _raw_time and "AST" not in _raw_time else (esc(_raw_time) if _raw_time else "—")

            # Logo: OVER/UNDER usan over_under.png; otros usan el equipo del pick
            if ptype in ("OVER","UNDER"):
                logo_h = bsn_logo_html("over_under", 52)
                color  = {"OVER":"#f97316","UNDER":"#a78bfa"}.get(ptype,"#f07820")
            else:
                pick_team = p["pick"].split()[0]
                logo_h    = bsn_logo_html(pick_team, 52)
                color     = _bsn_team_color(pick_team)

            # Type badge → etiqueta del tipo de pick
            _badge_map = {
                "ML":     ("Moneyline", "#1e3a5f", "#93c5fd"),
                "SPREAD": ("Spread",    "#1a2e1a", "#4ade80"),
                "OVER":   ("Totals",    "#2d1a3d", "#c084fc"),
                "UNDER":  ("Totals",    "#2d1a3d", "#c084fc"),
            }
            _bl, _bbg, _bfg = _badge_map.get(ptype, ("⏳ PENDING", "#f0782022", "#f07820"))
            type_badge = (f'<span style="background:{_bbg};color:{_bfg};border-radius:6px;'
                          f'padding:3px 10px;font-size:0.75rem;font-weight:700">{_bl}</span>')

            # ── Stats: Modelo / Mercado / Edge / EV ─────────────────────────
            modelo_str  = esc(p.get("modelo_str","—"))
            mercado_str = esc(p.get("mercado_str","—"))
            edge_str    = esc(p.get("edge_str","—"))
            edge_col    = "#22c55e" if not p.get("edge_str","").startswith("-") else "#ef4444"

            win_prob  = p.get("win_prob")
            pick_odds = p.get("pick_odds","")
            if win_prob is not None and pick_odds:
                ev_str, ev_col = _ev_str(win_prob, pick_odds)
            else:
                ev_str, ev_col = "—", "#94a3b8"

            # Pick display: quitar odds del string (están en odds-badge)
            pick_display = esc(re.sub(r'\s*\([^)]+\)\s*$', '', p['pick']).strip())

            # Alt picks de valor parcial — solo texto pequeño
            alt_html = ""
            if p.get("alt_picks"):
                alts = " · ".join(esc(a) for a in p["alt_picks"])
                alt_html = (f'<div style="font-size:0.78rem;color:#94a3b8;font-weight:500;'
                            f'margin-top:8px;padding-top:8px;border-top:1px solid #2a2a2a">'
                            f'Alt: {alts}</div>')

            # Helper para generar tarjeta de alt con valor completo
            def _alt_card(a, game_str, lg_html, t_str):
                a_pick_odds = a.get("pick_odds","")
                a_wp = a.get("win_prob")
                if a_wp is not None and a_pick_odds:
                    a_ev_str, a_ev_col = _ev_str(a_wp, a_pick_odds)
                else:
                    a_ev_str, a_ev_col = "—", "#94a3b8"
                a_edge_col = "#22c55e" if "+" in a.get("edge_str","") else "#f97316"
                a_pick_d   = esc(re.sub(r'\s*\([^)]+\)\s*$', '', a["pick"]).strip())
                # Logo propio del alt pick — no hereda el del pick primario
                a_ptype = a.get("pick_type", "")
                if a_ptype in ("OVER", "UNDER"):
                    lg_html = bsn_logo_html("over_under", 52)
                else:
                    a_first = a["pick"].split()[0] if a.get("pick") else ""
                    lg_html = bsn_logo_html(a_first, 52) if a_first else lg_html
                _a_badge_map = {
                    "ML":     ("Moneyline", "#1e3a5f", "#93c5fd"),
                    "SPREAD": ("Spread",    "#1a2e1a", "#4ade80"),
                    "OVER":   ("Totals",    "#2d1a3d", "#c084fc"),
                    "UNDER":  ("Totals",    "#2d1a3d", "#c084fc"),
                }
                _a_bl, _a_bbg, _a_bfg = _a_badge_map.get(a.get("pick_type",""), ("ALT", "#1e3a5f", "#93c5fd"))
                a_tb = (f'<span style="background:{_a_bbg};color:{_a_bfg};font-size:0.65rem;'
                        f'font-weight:700;padding:3px 8px;border-radius:10px">'
                        f'{_a_bl}</span>')
                return f"""
            <div class="pick-card" style="border-left:4px solid #f97316;opacity:0.92">
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
                <div class="pick-time">{t_str}</div>
                {a_tb}
              </div>
              <div class="teams-row">
                {lg_html}
                <div class="pick-main">
                  <div class="game-label">{esc(game_str)}</div>
                  <div class="pick-label" style="font-size:1.1rem">{a_pick_d} <span class="odds-badge">{esc(str(a_pick_odds))}</span></div>
                </div>
              </div>
              <div class="stats-grid">
                <div class="stat"><div class="stat-label">Modelo</div><div class="stat-val">{esc(a.get('modelo_str','—'))}</div></div>
                <div class="stat"><div class="stat-label">Mercado</div><div class="stat-val" style="color:#94a3b8;font-size:0.8rem">{esc(a.get('mercado_str','—'))}</div></div>
                <div class="stat"><div class="stat-label">Edge</div><div class="stat-val" style="color:{a_edge_col}">{esc(a.get('edge_str','—'))}</div></div>
                <div class="stat"><div class="stat-label">EV</div><div class="stat-val" style="color:{a_ev_col}">{esc(a_ev_str)}</div></div>
              </div>
            </div>"""

            body += f"""
            <div class="pick-card" style="border-left:4px solid {color}">
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
                <div class="pick-time">{time_str}</div>
                {type_badge}
              </div>
              <div class="teams-row">
                {logo_h}
                <div class="pick-main">
                  <div class="game-label">{esc(p['game'])}</div>
                  <div class="pick-label" style="font-size:1.2rem">{pick_display} <span class="odds-badge">{esc(str(pick_odds))}</span></div>
                </div>
              </div>
              <div class="stats-grid">
                <div class="stat">
                  <div class="stat-label">Modelo</div>
                  <div class="stat-val">{modelo_str}</div>
                </div>
                <div class="stat">
                  <div class="stat-label">Mercado</div>
                  <div class="stat-val" style="color:#94a3b8;font-size:0.8rem">{mercado_str}</div>
                </div>
                <div class="stat">
                  <div class="stat-label">Edge</div>
                  <div class="stat-val" style="color:{edge_col}">{edge_str}</div>
                </div>
                <div class="stat">
                  <div class="stat-label">EV</div>
                  <div class="stat-val" style="color:{ev_col}">{esc(ev_str)}</div>
                </div>
              </div>
              {alt_html}
            </div>"""

            # Render full-value alt picks as separate cards
            for _alt in p.get("alt_picks_full", []):
                body += _alt_card(_alt, p["game"], logo_h, time_str)

    html  = _bsn_html_wrap(f"Laboy BSN Picks · {dstr}", "BSN", dstr, yr, body)
    fname = f"Laboy BSN Picks {TARGET_DATE}-{_url_token(TARGET_DATE)}.html"
    fpath = os.path.join(SCRIPT_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  🎯 BSN Picks HTML: {fname}")
    return fpath


def export_bsn_lines_html(games_data):
    """
    Genera 'Laboy BSN Lines YYYY-MM-DD.html' con model lines de todos los juegos.
    """
    def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    dt   = datetime.strptime(TARGET_DATE, "%Y-%m-%d")
    dstr = dt.strftime("%A, %B %d · %Y").upper()
    yr   = dt.strftime("%Y")

    sorted_games = sorted(games_data, key=lambda g: _parse_time_sort(g.get("game_time","")))

    body = '<div class="section-title"><i class="fa-solid fa-chart-simple fa-icon"></i>Data Model Lines — Todos los Juegos</div>\n'
    for g in sorted_games:
        r      = g["result"]
        t1     = g["team1"]; t2 = g["team2"]
        gtime  = g.get("game_time","")
        venue  = g.get("venue","")
        venue_short = venue.split(",")[0] if venue else ""

        logo1 = bsn_logo_html(t1, 36)
        logo2 = bsn_logo_html(t2, 36)
        time_str = ((esc(gtime) + " AST") if gtime and "AST" not in gtime else (esc(gtime) if gtime else "—"))

        fav_color = _bsn_team_color(t1 if r["spread"] > 0 else t2)

        inj_html = ""
        if r["inj1"] or r["inj2"]:
            parts = []
            if r["inj1"]: parts.append(f'{esc(t1)} -{r["inj1"]:.1f}')
            if r["inj2"]: parts.append(f'{esc(t2)} -{r["inj2"]:.1f}')
            inj_html = f'<div class="inj-row"><i class="fa-solid fa-triangle-exclamation fa-icon"></i>Injury: {" · ".join(parts)} pts</div>'

        body += f"""
        <div class="line-card" style="border-left:3px solid {fav_color}">
          <div class="line-header">
            <div class="team-logo">{logo1}<span>{esc(t1)}</span></div>
            <div class="line-time">{time_str}</div>
            <div class="team-logo">{logo2}<span>{esc(t2)}</span></div>
          </div>
          <div class="line-stats">
            <span><i class="fa-solid fa-basketball fa-icon"></i>Total: {r['total']:.1f}</span>
            <span><i class="fa-solid fa-arrows-left-right fa-icon"></i>{esc(r['spread_line'])}</span>
            <span><i class="fa-solid fa-percent fa-icon"></i>{esc(t1)} {r['wp1']}% / {esc(t2)} {r['wp2']}%</span>
          </div>
          <div class="line-stats" style="margin-top:2px">
            <span><i class="fa-solid fa-coins fa-icon"></i>ML: {esc(t1)} {fmt_odds(r['ml1'])} / {esc(t2)} {fmt_odds(r['ml2'])}</span>
            <span><i class="fa-solid fa-bolt fa-icon"></i>Pace: {r['pace']:.1f}</span>
          </div>
          {f'<div class="venue-row"><i class="fa-solid fa-location-dot fa-icon"></i>{esc(venue_short)}</div>' if venue_short else ""}
          {inj_html}
        </div>"""

    html  = _bsn_html_wrap(f"Laboy BSN Lines · {dstr}", "BSN", dstr, yr, body)
    fname = f"Laboy BSN Lines {TARGET_DATE}-{_url_token(TARGET_DATE)}.html"
    fpath = os.path.join(SCRIPT_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  📊 BSN Lines HTML: {fname}")
    return fpath


def export_bsn_html(games_data):
    """Genera ambos HTMLs (picks + lines). Retorna lista [picks_path, lines_path]."""
    p = export_bsn_picks_html(games_data)
    l = export_bsn_lines_html(games_data)
    return [hp for hp in [p, l] if hp]


# ──────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n{'═'*52}")
    print(f"  LABOY PICKS — BSN")

    if   LOG_MODE:                    mode = "LOG PICK"
    elif GRADE_MODE:                  mode = "GRADE PICK"
    elif RECORD_MODE:                 mode = "RECORD"
    elif FEEDBACK_MODE:               mode = "FEEDBACK"
    elif EXPORT_LOG_MODE:             mode = "EXPORT LOG"
    elif REMOVE_MODE:                 mode = "REMOVE PICK(S)"
    elif SET_LINES_MODE:              mode = "SET MARKET LINES"
    elif CLEAR_LINES_MODE:            mode = "CLEAR MARKET LINES"
    elif GRADE_PICKS_MODE:            mode = "GRADE PICKS"
    elif LINES_MODE or PICKS_MODE:    mode = "SHOW LINES / PICKS"
    elif STATS_MODE:                   mode = "TEAM STATS"
    elif IR_MODE:                      mode = "INJURY REPORT"
    elif ADD_INJURY_MODE:              mode = "ADD INJURY"
    elif REMOVE_INJURY_MODE:           mode = "REMOVE INJURY"
    elif ADD_GAME_MODE:                mode = "ADD GAME"
    elif REMOVE_GAME_MODE:             mode = "REMOVE GAME"
    elif LIST_GAMES_MODE:              mode = "LIST GAMES"
    elif DEBUG_SCHEDULE:               mode = "DEBUG SCHEDULE"
    elif REFRESH_MODE:                 mode = "REFRESH 2026 STATS"
    elif SCHEDULE_MODE:                mode = "SCHEDULE"
    elif EXPORT_LINES:                 mode = "EXPORT LINES PNG"
    elif EXPORT_LINES_PDF:             mode = "EXPORT LINES PDF"
    elif EXPORT_PICKS or EXPORT_STORY: mode = "EXPORT PICKS PNG"
    elif EXPORT_PICKS_PDF:             mode = "EXPORT PICKS PDF"
    elif EXPORT_POST:                  mode = "EXPORT POST PNG"
    elif EXPORT_HTML_MODE:             mode = "EXPORT HTML"
    else:                              mode = f"DAILY — {TARGET_DATE}"
    print(f"  {mode}")
    print(f"{'═'*52}\n")

    # ── Pick tracker (immediate exit modes) ────────────────────
    if GP_MODE:          cmd_set_gp();            sys.exit(0)
    if GRADE_PICKS_MODE:   cmd_grade_picks_bsn();  sys.exit(0)
    if SET_LINES_MODE:     cmd_set_lines_bsn();    sys.exit(0)
    if CLEAR_LINES_MODE:   cmd_clear_lines();      sys.exit(0)
    if SERVE_MODE:         cmd_serve();            sys.exit(0)
    if LOG_MODE:           cmd_log_pick();         sys.exit(0)
    if LOG_RETRO_MODE:     cmd_log_retro();        sys.exit(0)
    if LOG_PARLAY_MODE:    cmd_log_parlay();       sys.exit(0)
    if LOG_SPECIAL_MODE:   cmd_log_special();      sys.exit(0)
    if SEASON_CARD_MODE:
        result = export_season_card_bsn()
        if result and PUBLISH_MODE:
            _, jpg = result
            if jpg: cmd_publish([jpg])
        sys.exit(0)
    if EXPORT_LOG_MODE:    cmd_export_log();       sys.exit(0)
    if GRADE_MODE:         cmd_grade_pick();       sys.exit(0)
    if REMOVE_MODE:        cmd_remove_pick();      sys.exit(0)
    if EDIT_MODE:          cmd_edit_pick();        sys.exit(0)
    if RECORD_MODE:        cmd_record();           sys.exit(0)
    if FEEDBACK_MODE:      cmd_feedback();         sys.exit(0)
    if EXPORT_RECORD_MODE:
        try:
            ri = sys.argv.index("--export-record")
            rd = sys.argv[ri+1] if ri+1 < len(sys.argv) and not sys.argv[ri+1].startswith("-") else None
        except (ValueError, IndexError):
            rd = None
        html_path = export_record_card_bsn(rd)
        jpg_path  = None
        if html_path:
            print(f"  📄 Record Card HTML: {os.path.basename(html_path)}")
            jpg_path = html_to_jpg(html_path, width=680, scale=2)
            if jpg_path:
                print(f"  🖼️  Record Card JPG:  {os.path.basename(jpg_path)}")
        if PUBLISH_MODE:
            to_publish = [p for p in [html_path, jpg_path] if p]
            if to_publish:
                cmd_publish(to_publish)
        sys.exit(0)

    # ── PNG / PDF exports (no Excel needed for picks) ─
    if EXPORT_PICKS or EXPORT_STORY:
        paths = export_bsn_log_picks_png()
        if EXPORT_STORY:
            print(f"\n  📱 Story (1080×1920): {len(paths)} archivo(s) generado(s)")
        sys.exit(0)

    if EXPORT_PICKS_PDF:
        picks_paths = export_bsn_log_picks_png()
        if picks_paths:
            pdf_out = os.path.join(SCRIPT_DIR, f"Laboy BSN MyPicks {TARGET_DATE}.pdf")
            pngs_to_pdf_bsn(picks_paths, pdf_out)
        sys.exit(0)

    if EXPORT_POST:
        story_paths = export_bsn_log_picks_png()
        for sp in story_paths:
            crop_story_to_post_bsn(sp, sp.replace(".png","_post.png"))
        sys.exit(0)

    # Lines/PDF exports need games data — loaded below
    # (EXPORT_LINES and EXPORT_LINES_PDF handled after load_scheduled_games)

    # ── Injury management (write mode) ────────────────
    if ADD_INJURY_MODE:
        wb = load_workbook(EXCEL_FILE, keep_links=False)
        cmd_add_injury(wb)
        sys.exit(0)

    if REMOVE_INJURY_MODE:
        wb = load_workbook(EXCEL_FILE, keep_links=False)
        cmd_remove_injury(wb)
        sys.exit(0)

    if ADD_GAME_MODE:
        cmd_add_game()
        sys.exit(0)

    if REMOVE_GAME_MODE:
        cmd_remove_game()
        sys.exit(0)

    if EDIT_GAME_MODE:
        cmd_edit_game()
        sys.exit(0)

    if LIST_GAMES_MODE:
        cmd_list_games()
        sys.exit(0)

    if DEBUG_SCHEDULE:
        cmd_debug_schedule()
        sys.exit(0)

    if REFRESH_MODE:
        wb = load_workbook(EXCEL_FILE, keep_links=False)
        cmd_refresh(wb)
        sys.exit(0)

    # ── Read-only modes ────────────────────────────────
    wb = load_workbook(EXCEL_FILE, data_only=True, keep_links=False)

    if SCHEDULE_MODE:
        show_schedule(wb)
        sys.exit(0)

    stats         = load_bsn_advanced(wb)
    injury_impact = load_injury_impact(wb)

    if STATS_MODE:
        show_stats(stats, injury_impact)
        sys.exit(0)

    if IR_MODE:
        cmd_ir(wb)
        sys.exit(0)

    # ── Lines / Picks / Daily ──────────────────────────
    raw_games = load_scheduled_games(wb, TARGET_DATE)

    if not raw_games:
        print(f"  No hay juegos BSN para {TARGET_DATE}.")
        print(f"  Tip: Agrega juegos en el tab 'BSN Lines' del Excel.")
        print(f"       Columnas: D=Fecha  F=Team1  G=Team2\n")
        sys.exit(0)

    # Calcular proyecciones para cada juego
    games_data = []
    for g in raw_games:
        result = compute_game(g["team1"], g["team2"], stats, injury_impact)
        games_data.append({**g, "result": result})

    # Ordenar por hora (si hay hora disponible)
    games_data.sort(key=lambda x: _parse_time_sort(x.get("game_time","")))

    if EXPORT_LINES:
        export_bsn_lines_png(games_data)
        sys.exit(0)

    if EXPORT_LINES_PDF:
        png_paths = export_bsn_lines_png(games_data)
        if png_paths:
            pdf_out = os.path.join(SCRIPT_DIR, f"Laboy BSN Lines {TARGET_DATE}.pdf")
            pngs_to_pdf_bsn(png_paths, pdf_out)
        sys.exit(0)

    if EXPORT_HTML_MODE:
        html_paths = export_bsn_html(games_data)   # lista [picks_path, lines_path]
        if PUBLISH_MODE and html_paths:
            cmd_publish(html_paths)
        sys.exit(0)

    display_lines(games_data)

    if PICKS_MODE:
        show_picks(games_data)
    elif not LINES_MODE:
        # Default: muestra picks también
        show_picks(games_data)
